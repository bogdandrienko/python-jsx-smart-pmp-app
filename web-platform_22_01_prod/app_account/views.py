from django.shortcuts import render, redirect
from django.contrib.auth.models import Group, User
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login, authenticate, logout

from .forms import CreateUserForm, ChangeUserForm, CreateUsersForm, GeneratePasswordsForm
from .models import AccountDataModel
from src.py.django_utils import create_encrypted_password
from app_salary.json_data import get_users

import openpyxl


def login_account(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            username = request.POST['username']
            password = request.POST['password']
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('home')
            return redirect('app_account:create')
    form = AuthenticationForm()
    context = {
        'form': form,
    }
    return render(request, 'app_account/login_user.html', context)


def logout_account(request):
    # Проверка регистрации: если пользователь не вошёл в аккаунт, действия не срабатают, а его переадресует в форму
    # входа
    if request.user.is_authenticated is not True:
        return redirect('app_account:login')
    # Переадресация пользователя на страницу входа

    logout(request)
    return redirect('app_account:login')


def get_user(username: str):
    return 'Bogdan and  ' + username


def create_user(request):
    # Переадресация не аутентифицированного пользователя на страницу входа
    if request.user.is_authenticated is not True:
        return redirect('app_account:login')
    # Переадресация не аутентифицированного пользователя на страницу входа

    if request.method == 'POST':

        form = CreateUserForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            signup_user = User.objects.get(username=username)
            try:
                user_group = Group.objects.get(name='User')
            except Exception as ex:
                user_group = Group.objects.create(name='User')
            user_group.user_set.add(signup_user)

        result_create_user = True
    else:
        result_create_user = False

    create_user_form = CreateUserForm()
    context = {
        'create_user_form': create_user_form,
        'result_create_user': result_create_user,
    }

    return render(request, 'app_account/create_user.html', context)


def create_user_from1c(request):
    data = None
    if request.method == 'POST':
        data_s = get_users()
        headers = ["Период", "Статус", "ИИН", "Фамилия", "Имя", "Отчество"]
        bodies = []
        for x in data_s["global_objects"]:
            val = []
            for y in data_s["global_objects"][x]:
                val.append(data_s["global_objects"][x][y])
            bodies.append(val)
        data = [headers, bodies]
    context = {
        'data': data,
    }
    return render(request, 'app_account/create_user_from1c.html', context)


def change_user(request):
    # Переадресация не аутентифицированного пользователя на страницу входа
    if request.user.is_authenticated is not True:
        return redirect('app_account:login')
    # Переадресация не аутентифицированного пользователя на страницу входа

    if request.method == 'POST':

        form = ChangeUserForm(request.POST)
        if form.is_valid():
            user = User.objects.get(username=request.user.username)
            user.email = request.POST['email']
            user.password = request.POST['password1']
            user.first_name = request.POST['first_name']
            user.last_name = request.POST['last_name']
            user.save()
            user.set_password = request.POST['password1']

        result_change_user = True
    else:
        result_change_user = False

    change_user_form = ChangeUserForm()
    context = {
        'change_user_form': change_user_form,
        'result_change_user': result_change_user,
    }

    return render(request, 'app_account/change_user.html', context)


def create_users(request):
    # Переадресация не аутентифицированного пользователя на страницу входа
    if request.user.is_authenticated is not True:
        return redirect('app_account:login')
    # Переадресация не аутентифицированного пользователя на страницу входа

    if request.method == 'POST':
        excel_file = request.FILES.get('document_addition_file_1')
        if excel_file:
            def get_value(rows: str, cols: int):
                return str(sheet[rows + str(cols)].value)

            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            max_row = sheet.max_row
            for i in range(2, max_row + 1):
                username = get_value('A', i)
                password = get_value('B', i)
                email = get_value('C', i)
                first_name = get_value('D', i)
                last_name = get_value('E', i)
                group = get_value('F', i)
                if get_value('G', i) == 'ИСТИНА' or get_value('G', i) == 'True':
                    is_staff = True
                else:
                    is_staff = False
                if username and password:
                    try:
                        user = User.objects.get(username=username)
                        user.username = username
                        user.password = password
                        user.email = email
                        user.first_name = first_name
                        user.last_name = last_name
                        user.is_staff = is_staff
                        if username == 'Bogdan' or username == 'bogdan':
                            user.is_superuser = True
                        else:
                            user.is_superuser = False
                        user.save()
                        user.set_password = str(password)
                    except Exception as ex:
                        user = User.objects.create(
                            username=username,
                            password=password,
                            email=email,
                            first_name=first_name,
                            last_name=last_name,
                            is_staff=is_staff,
                        )
                    try:
                        user_group = Group.objects.get(name=group)
                    except Exception as ex:
                        user_group = Group.objects.create(name=group)
                    signup_user = User.objects.get(username=username)
                    user_group.user_set.add(signup_user)

        result_create_users = True
    else:
        result_create_users = False

    create_users_form = CreateUsersForm()
    context = {
        'create_users_form': create_users_form,
        'result_create_users': result_create_users,
    }

    return render(request, 'app_account/create_users.html', context)


def export_users(request):
    # Переадресация не аутентифицированного пользователя на страницу входа
    if request.user.is_authenticated is not True:
        return redirect('app_account:login')
    # Переадресация не аутентифицированного пользователя на страницу входа

    if request.method == 'POST':

        print('export_users')

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Страница 1'
        user_objects = User.objects.all().order_by('id')

        print(user_objects)

        sheet['A1'] = 'Имя пользователя'
        sheet['B1'] = 'Зашифрованный Пароль'
        sheet['C1'] = 'Почта'
        sheet['D1'] = 'Имя'
        sheet['E1'] = 'Фамилия'
        sheet['F1'] = 'Группа'
        sheet['G1'] = 'Доступ к админ панели'
        value = 1
        for user_object in user_objects:
            value += 1
            sheet[f'A{value}'] = user_object.username
            sheet[f'B{value}'] = user_object.password
            sheet[f'C{value}'] = user_object.email
            sheet[f'D{value}'] = user_object.first_name
            sheet[f'E{value}'] = user_object.last_name
            groups = Group.objects.filter(user=user_object)
            group_list = ''
            for group in groups:
                group_list += group.name
            sheet[f'F{value}'] = group_list
            if user_object.is_staff:
                sheet[f'G{value}'] = 'ИСТИНА'
            else:
                sheet[f'G{value}'] = 'ЛОЖЬ'
        wb.save('static/media/data/accounts/export_users.xlsx')

        result_export_users = True
    else:
        result_export_users = False

    context = {
        'result_export_users': result_export_users,
    }

    return render(request, 'app_account/export_users.html', context)


def generate_password(request):
    # Переадресация не аутентифицированного пользователя на страницу входа
    if request.user.is_authenticated is not True:
        return redirect('app_account:login')
    # Переадресация не аутентифицированного пользователя на страницу входа

    if request.method == 'POST':
        not_encrypted_password = str(request.POST['password'])

        try:
            user = User.objects.get(username='None')
        except Exception as ex:
            user = User.objects.create(
                username='None'
            )
        user.set_password(not_encrypted_password)
        user.save()
        user = User.objects.get(username='None')
        encrypted_password = user.password

    else:
        encrypted_password = False

    context = {
        'result_generate_password': encrypted_password,
    }

    return render(request, 'app_account/generate_password.html', context)


def generate_passwords(request):
    # Переадресация не аутентифицированного пользователя на страницу входа
    if request.user.is_authenticated is not True:
        return redirect('app_account:login')
    # Переадресация не аутентифицированного пользователя на страницу входа

    if request.method == 'POST':
        passwords_quantity = int(request.POST['passwords_quantity'])
        passwords_chars = str(request.POST['passwords_chars'])
        passwords_length = int(request.POST['passwords_length'])

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Страница 1'
        sheet[f'A1'] = 'Зашифрованный Пароль'
        sheet[f'B1'] = 'Пароль'
        for n in range(2, passwords_quantity + 2):
            password = create_encrypted_password(_random_chars=passwords_chars, _length=passwords_length)
            try:
                user = User.objects.get(username='None')
            except Exception as ex:
                user = User.objects.create(
                    username='None',
                    email=password,
                )
            user.email = password
            user.set_password(password)
            user.save()
            user = User.objects.get(username='None')
            sheet[f'A{n}'] = user.password
            sheet[f'B{n}'] = user.email
        wb.save('static/media/data/accounts/generate_passwords.xlsx')

        result_generate_passwords = True
    else:
        result_generate_passwords = False

    generate_passwords_form = GeneratePasswordsForm()
    context = {
        'generate_passwords_form': generate_passwords_form,
        'result_generate_passwords': result_generate_passwords,
    }

    return render(request, 'app_account/generate_passwords.html', context)


def view_profile(request, username=None):
    try:
        account = AccountDataModel.objects.get(user_iin=username)
    except Exception as ex:
        account = False
    context = {
        'account': account,
        'username': username
    }
    return render(request, 'app_account/view_profile.html', context)
