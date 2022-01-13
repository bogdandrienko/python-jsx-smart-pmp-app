import datetime
import math
import os
import random
import time
import cv2
import httplib2
import numpy
import openpyxl
from openpyxl.utils import get_column_letter
from fastkml import kml
from concurrent.futures import ThreadPoolExecutor

from django.conf import settings
from django.contrib.auth.models import User
from django.contrib.staticfiles import finders
from django.core.handlers.wsgi import WSGIRequest
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils import timezone

from app_admin.models import UserModel, GroupModel, LoggingModel, ActionModel, ComputerVisionModuleModel, \
    ComputerVisionComponentModel
from app_admin.utils import ExcelClass, DirPathFolderPathClass, DateTimeUtils


class DjangoClass:
    class AuthorizationClass:
        @staticmethod
        def try_to_access(request, access: str):
            DjangoClass.LoggingClass.logging_actions(request=request)
            if str(request.META.get("REMOTE_ADDR")) == '192.168.1.202':
                return 'local'
            if access == 'only_logging':
                return False
            if request.user.is_authenticated:
                try:
                    user = User.objects.get(username=request.user.username)
                    user_model = UserModel.objects.get_or_create(user_foreign_key_field=user)[0]
                    if user.is_superuser:
                        return False
                    if user_model.activity_boolean_field is False:
                        return 'account_logout'
                    else:
                        if user_model.email_field and user_model.secret_question_char_field and \
                                user_model.secret_answer_char_field:
                            try:
                                action_model = ActionModel.objects.get(name_slug_field=access)
                                if action_model:
                                    groups = GroupModel.objects.filter(
                                        user_many_to_many_field=user_model,
                                        action_many_to_many_field=action_model,
                                    )
                                    if groups:
                                        return False
                                    else:
                                        return 'home'
                            except Exception as error:
                                return 'home'
                        else:
                            return 'account_change_password'
                except Exception as error:
                    return 'home'
            else:
                return 'account_login'

    class LoggingClass:
        @staticmethod
        def logging_errors(request, error):
            # for k, v in request.META.items():
            #     print(f'\n{k}: {v}')
            username = request.user.username
            ip = request.META.get("REMOTE_ADDR")
            request_path = request.path
            request_method = request.method
            datetime_now = datetime.datetime.now()
            LoggingModel.objects.create(
                username_slug_field=username,
                ip_genericipaddress_field=ip,
                request_path_slug_field=request_path,
                request_method_slug_field=request_method,
                error_text_field=f'error: {error}'
            )
            text = [username, ip, request_path, request_method, datetime_now, error]
            string = ''
            for val in text:
                string = string + f', {val}'
            with open('static/media/admin/logging/logging_errors.txt', 'a') as log:
                log.write(f'\n{string[2:]}\n')

        @staticmethod
        def logging_errors_local(error, function_error):
            datetime_now = datetime.datetime.now()
            LoggingModel.objects.create(username='', ip='', request_path=function_error,
                                        request_method='', error=error)
            text = [function_error, datetime_now, error]
            string = ''
            for val in text:
                string = string + f', {val}'
            with open('static/media/admin/logging/logging_errors.txt', 'a') as log:
                log.write(f'\n{string[2:]}\n')

        @staticmethod
        def logging_actions(request):
            # for k, v in request.META.items():
            #     print(f'\n{k}: {v}')
            username = request.user.username
            ip = request.META.get("REMOTE_ADDR")
            request_path = request.path
            request_method = request.method
            datetime_now = datetime.datetime.now()
            LoggingModel.objects.create(
                username_slug_field=username,
                ip_genericipaddress_field=ip,
                request_path_slug_field=request_path,
                request_method_slug_field=request_method,
                error_text_field='successful'
            )
            text = [username, ip, request_path, request_method, datetime_now]
            string = ''
            for val in text:
                string = string + f', {val}'
            with open('static/media/admin/logging/logging_actions.txt', 'a') as log:
                log.write(f'\n{string[2:]}\n')

    class AccountClass:
        @staticmethod
        def create_django_encrypt_password(password: str):
            try:
                user = User.objects.get_or_create(username='sha256')[0]
                user.set_password(password)
                encrypt_password = user.password
                UserModel.objects.get(user_foreign_key_field=user).delete()
                user.delete()
                return encrypt_password
            except Exception as error:
                DjangoClass.LoggingClass.logging_errors_local(error=error, function_error='get_sha256_password')
                return False

        @staticmethod
        def create_password_from_chars(chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890',
                                       length=8):
            password = ''
            for i in range(1, length + 1):
                password += random.choice(chars)
            return password

    class RequestClass:
        @staticmethod
        def get_value(request: WSGIRequest, key: str, none_is_error=False, strip=True):
            if none_is_error:
                # If key not have is Exception Error
                value = request.POST[key]
            else:
                # If key not have value is None
                value = request.POST.get(key)
            if strip and value:
                value.strip()
            return value

        @staticmethod
        def get_check(request: WSGIRequest, key: str, none_is_error=False):
            if none_is_error:
                # If key not have is Exception Error
                value = request.POST[key]
            else:
                # If key not have value is None
                value = request.POST.get(key)
            if value is None:
                return False
            else:
                return True

        @staticmethod
        def get_file(request: WSGIRequest, key: str, none_is_error=False):
            if none_is_error:
                # If key not have is Exception Error
                file = request.FILES[key]
            else:
                # If key not have value is None
                file = request.FILES.get(key)
            return file


class PaginationClass:
    @staticmethod
    def paginate(request, objects, num_page):
        # Пагинатор: постраничный вывод объектов
        paginator = Paginator(objects, num_page)
        pages = request.GET.get('page')
        try:
            page = paginator.page(pages)
        except PageNotAnInteger:
            page = paginator.page(1)
        except EmptyPage:
            page = paginator.page(paginator.num_pages)
        return page


class ComputerVisionClass:
    class SoloCamSettingsComputerVisionClassExample:
        """
        Настройки одной камеры для модуля: грохота 16 операции, 10 отметка
        """

        def __init__(self, ip: str, correct: float):
            # Ip address
            self.ip = ip
            # Коэффициент корректировки
            self.correct = round(correct, 3)

    class EventLoopClass:
        @staticmethod
        def loop_modules_global(tick_delay=0.5):
            """
            Цикл прохода по всем модулям-циклам
            """
            try:
                def global_tick():
                    # Начальное время
                    start_time = time.time()
                    # Получение словаря с названиями модулей из константы модели: "ComputerVisionModuleModel"
                    dict_modules = {str(x[0]).strip(): str(x[1]).strip() for x in
                                    ComputerVisionModuleModel.get_all_modules()}
                    # Менеджер контекста для многопоточное прохода по всем модулям
                    with ThreadPoolExecutor() as executor:
                        # Получение всех активных модулей
                        modules = ComputerVisionModuleModel.objects.filter(play_boolean_field=True)
                        # Получение всех активных модулей
                        if modules:
                            delay = tick_delay
                        else:
                            delay = tick_delay * 5
                        # Цикл прохода по всем активным модулям
                        for module in modules:
                            # Проверка, что цикл модуля отработал в последний раз не позже задержки цикла
                            if module.datetime_field and (module.datetime_field + datetime.timedelta(hours=6)) > \
                                    DateTimeUtils.get_difference_datetime(seconds=-module.delay_float_field):
                                continue
                            # Получение имени модуля из константы в модели ComputerVisionModuleModel
                            name_module = dict_modules[module.path_slug_field]
                            if name_module == 'Грохота, 16 операция, 10 отметка':
                                executor.submit(ComputerVisionClass.EventLoopClass.Grohota16OperationClass.
                                                loop_module_grohota_16_operation, module=module)
                            elif name_module == 'Грохота, 26 операция, 10 отметка':
                                executor.submit(ComputerVisionClass.EventLoopClass.Grohota16OperationClass.
                                                loop_module_grohota_16_operation, module=module)
                            elif name_module == 'Грохота, 36 операция, 10 отметка':
                                executor.submit(ComputerVisionClass.EventLoopClass.Grohota16OperationClass.
                                                loop_module_grohota_16_operation, module=module)
                    time.sleep(delay)
                    print(f'\n#######################################'
                          f'\n##### {round(time.time() - start_time, 2)} seconds for global_tick ####'
                          f'\n#######################################\n')

                while True:
                    global_tick()
            except Exception as error:
                print(f'\nloop_modules_global | error : {error}\n')
                time.sleep(tick_delay * 10)
                ComputerVisionClass.EventLoopClass.loop_modules_global()

        class Grohota16OperationClass:
            @staticmethod
            def loop_module_grohota_16_operation(module):
                """
                Цикл прохода по компонентам-функциям внутри модуля: Грохота, 16 операция, 10 отметка
                """
                try:
                    # Начальное время
                    start_time = time.time()
                    # Данные из базы с настройками по каждому активному компоненту
                    components = ComputerVisionComponentModel.objects.filter(play_boolean_field=True)
                    with ThreadPoolExecutor() as executor:
                        # Цикл для прохода по настройкам и запуску компонентов
                        for component in components:
                            executor.submit(ComputerVisionClass.EventLoopClass.Grohota16OperationClass.
                                            component_grohota_16_operation, component=component)
                    module = ComputerVisionModuleModel.objects.get(path_slug_field=module.path_slug_field)
                    module.duration_float_field = round(time.time() - start_time, 2)
                    module.datetime_field = timezone.now()
                    module.save()
                except Exception as error:
                    error_text_field = f'loop_module_grohota_16_operation | {timezone.now()} | error : {error}'
                    module = ComputerVisionModuleModel.objects.get(path_slug_field=module.path_slug_field)
                    module.error_text_field = error_text_field
                    module.save()
                    print(f'\n{error_text_field}\n')

            @staticmethod
            def component_grohota_16_operation(component):
                """
                Компонент-функция расчёта % схода на грохотах 16 операции, 10 отметка
                """
                try:
                    # Создание папки с кешем для библиотеки
                    temp_path = DirPathFolderPathClass.create_folder_in_this_dir(
                        folder_name='static/media/data/computer_vision/temp'
                    )
                    # Создание экземпляра объекта библиотеки, установка папки с кешем для библиотеки и таймаута
                    h = httplib2.Http(cache=temp_path, timeout=3)
                    # Установка логина от камеры
                    login = 'admin'
                    # Установка пароля от камеры
                    password = 'q1234567'
                    # Добавление логина и пароля к авторизации
                    h.add_credentials(login, password)
                    # Заполнение api-пути для получения изображения по сети от камеры
                    sources = f'http://{str(component.genericipaddress_field)}:80' \
                              f'/ISAPI/Streaming/channels/101/picture?snapShotImageType=JPEG'
                    # Установка заголовка для запроса
                    headers = {
                        'user-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:72.0) Gecko/20100101 Firefox/72.0',
                        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'
                    }
                    # Получение данных с api камеры
                    response, content = h.request(uri=sources, method="GET", headers=headers)
                    # Чтение изображения-маски
                    mask = cv2.imread(component.mask_char_field, 0)
                    # Превращение массива байтов в массив пикселей и чтение массива в объект-изображение cv2
                    image = cv2.imdecode(numpy.frombuffer(content, numpy.uint8), cv2.IMREAD_COLOR)
                    # Наложение на изображение маски с тёмным режимом: закрашивание участков в чёрный
                    bitwise_and = cv2.bitwise_and(image, image, mask=mask)
                    # Перевод изображения BGR(RGB - для не cv2) в формат HSV
                    cvtcolor = cv2.cvtColor(bitwise_and, cv2.COLOR_BGR2HSV)
                    # Заполнение массива изображения по диапазону с выбранной чувствительностью
                    inrange = cv2.inRange(
                        cvtcolor,
                        numpy.array([0, 0, 255 - 120], dtype=numpy.uint8),
                        numpy.array([255, 120, 255], dtype=numpy.uint8)
                    )
                    # Подсчёт белых пикселей после заполнения к белым пикселям на маске и умножение на коррекцию
                    value = numpy.sum(inrange > 0) / numpy.sum(mask > 0) * 100 * float(1)
                    # Запись результата в базу данных
                    LoggingModel.objects.create(
                        username_slug_field='computer vision',
                        ip_genericipaddress_field=component.genericipaddress_field,
                        request_path_slug_field='component_grohota_16_operation',
                        request_method_slug_field='POST',
                        error_text_field=f'value: {round(value, 2)} %'
                    )
                    # value = value/100
                    #
                    # device_row = component.alias_char_field
                    # if value > 30:
                    #     alarm_row = 1
                    # else:
                    #     alarm_row = 0
                    # datetime_row = datetime.datetime.now().strftime('%Y-%m-%dT%H:%M:%S')
                    # rows = ['device_row', 'value_row', 'alarm_row', 'datetime_row']
                    # values = [device_row, round(value, 2), alarm_row, datetime_row]
                    #
                    # connection = SQLClass.pyodbc_connect(
                    #     ip='192.168.15.122',
                    #     server='WINCC',
                    #     port='49279',
                    #     database='KM_Fabrika',
                    #     username='computer_vision',
                    #     password='vision12345678'
                    # )
                    # cursor = connection.cursor()
                    # cursor.fast_executemany = True
                    # __rows = ''
                    # for x in rows:
                    #     __rows = f"{__rows}{str(x)}, "
                    # query = f"UPDATE {'grohot16_now_table'} SET {rows[1]} = '{values[1]}',{rows[2]} = '{values[2]}' ,
                    # {rows[3]} = '{values[3]}' " \
                    #         f"WHERE {rows[0]} = '{values[0]}'"
                    # cursor.execute(query)
                    # connection.commit()

                    # Вывод результата в консоль
                    print(f'{component.genericipaddress_field} | {component.alias_char_field} : {round(value, 2)} %')
                except Exception as error:
                    print(
                        f'\ncomponent_grohota_16_operation| {component.genericipaddress_field} | '
                        f'{component.alias_char_field} | error : {error}\n')


class SalaryClass:
    @staticmethod
    def create_arr_table(title: str, footer: str, json_obj, exclude: list):
        headers = []

        json_obj = dict(json_obj).copy()

        for x in json_obj["Fields"]:
            headers.append(json_obj["Fields"][x])
        del json_obj["Fields"]
        bodies = [["", title]]

        if exclude:
            hours = 0
            days = 0
            sum_ = 0
            for x in json_obj:
                val = [x]
                i = 0
                for y in json_obj[x]:
                    i += 1
                    if i == exclude[0]:
                        hours += json_obj[x][y]
                        continue
                    if i == exclude[1]:
                        days += json_obj[x][y]
                        continue
                    if i == len(json_obj[x]):
                        sum_ += json_obj[x][y]
                    val.append(json_obj[x][y])
                bodies.append(val)
            footers = ["", footer, "", hours, days, round(sum_, 2)]
        else:
            sum_ = 0
            for x in json_obj:
                val = [x]
                i = 0
                for y in json_obj[x]:
                    i += 1
                    if i == len(json_obj[x]):
                        sum_ += json_obj[x][y]
                    val.append(json_obj[x][y])
                bodies.append(val)
            footers = ["", footer, "", round(sum_, 2)]

        return [headers, bodies, footers]

    @staticmethod
    def create_arr_from_json(json_obj, parent_key: str):
        headers = []
        for x in json_obj[parent_key]["Fields"]:
            headers.append(json_obj[parent_key]["Fields"][x])
        del json_obj[parent_key]["Fields"]
        bodies = []
        for x in json_obj[parent_key]:
            val = [x]
            for y in json_obj[parent_key][x]:
                val.append(json_obj[parent_key][x][y])
            bodies.append(val)
        return [parent_key, headers, bodies]


class Xhtml2pdfClass:
    @staticmethod
    def link_callback(uri):
        """
        Convert HTML URIs to absolute system paths so xhtml2pdf can access those
        resources
        """
        result = finders.find(uri)
        s_url = ''
        m_url = ''
        if result:
            if not isinstance(result, (list, tuple)):
                result = [result]
            result = list(os.path.realpath(path) for path in result)
            path = result[0]
        else:
            s_url = settings.STATIC_URL  # Typically /static/
            s_root = settings.STATIC_ROOT  # Typically /home/userX/project_static/
            m_url = settings.MEDIA_URL  # Typically /media/
            m_root = settings.MEDIA_ROOT  # Typically /home/userX/project_static/media/

            if uri.startswith(m_url):
                path = os.path.join(m_root, uri.replace(m_url, ""))
            elif uri.startswith(s_url):
                path = os.path.join(s_root, uri.replace(s_url, ""))
            else:
                return uri
        if not os.path.isfile(path):
            raise Exception(
                'media URI must start with %s or %s' % (s_url, m_url)
            )
        return path


class GeoClass:
    @staticmethod
    def get_hypotenuse(x1: float, y1: float, x2: float, y2: float):
        """"
        Принимает: "пару" точек - их широту и долготу.
        Возвращает: корень из суммы квадратов разностей широты и долготы двух пар точек, ака гипотенузу.
        """
        return math.sqrt(((x2 - x1) ** 2) + ((y2 - y1) ** 2))

    @staticmethod
    def get_haversine(latitude_1: float, longitude_1: float, latitude_2: float, longitude_2: float):
        """"
        Принимает: "пару" точек - их широту и долготу.
        Возвращает: значение расстояния по гипотенузе двух точек, в метрах, ака формула гаверсинуса.
        """
        delta_latitude = latitude_2 * math.pi / 180 - latitude_1 * math.pi / 180
        delta_longitude = longitude_2 * math.pi / 180 - longitude_1 * math.pi / 180
        a = math.sin(delta_latitude / 2) * math.sin(delta_latitude / 2) + math.cos(latitude_1 * math.pi / 180) * \
            math.cos(latitude_2 * math.pi / 180) * math.sin(delta_longitude / 2) * math.sin(delta_longitude / 2)
        return round(math.atan2(math.sqrt(a), math.sqrt(1 - a)) * 6378.137 * 2 * 1000)

    @staticmethod
    def find_near_point(point_arr: list, point_latitude: float, point_longitude: float):
        """"
        Принимает: массив точек в которых надо искать, где первый элемент это широта, а второй долгоа. Также данные
        точки ближайшие координаты которой надо найти.
        Возвращает: данные точки из массива точек, которая соответствует ближайшему значению целевой точки.
        """
        result = None
        for coord in point_arr:
            if coord[0] <= point_latitude and coord[1] <= point_longitude:
                result = coord
        if result is None:
            result = point_arr[0]
        return result

    @staticmethod
    def get_vector_arr(point_arr):
        """
        Принимает: массив "точек" - первый элемент это имя точки, второй и третий это широта и долгота, а четвёртый
        связи.
        Возвращает: массив "векторов" - первый элемент это имя вектора, второй это расстояние через формулу
        "гаверсинуса".
        """
        # Points = [PointName, latitude, longitude, PointLinks]
        # point1 = ["1", 52.14303, 61.22812, "2"]
        # point2 = ["2", 52.1431, 61.22829, "1|3"]
        # Vectors = [VectorName, length(meters)]
        # vector1 = ["1|2", 12]
        # vector2 = ["2|3", 14]

        vector_arr = []
        for point in point_arr:
            lat1 = point[0]
            lon1 = point[1]
            vector1 = point[2]
            link_arr = point[3].split("|")
            for link in link_arr:
                for point_1 in point_arr:
                    if point_1[2] == link:
                        lat2 = point_1[0]
                        lon2 = point_1[1]
                        vector2 = link
                        length = GeoClass.get_haversine(lat1, lon1, lat2, lon2)
                        vector_arr.append([f"{vector1}|{vector2}", length])
        return vector_arr

    @staticmethod
    def create_cube_object(point: list):
        latitude = point[0]
        longitude = point[1]
        first = [latitude - 0.000002 * 1.62, longitude - 0.000002, 0]
        second = [latitude - 0.000002 * 1.62, longitude + 0.000002, 0]
        third = [latitude + 0.000002 * 1.62, longitude + 0.000002, 0]
        fourth = [latitude + 0.000002 * 1.62, longitude - 0.000002, 0]
        string_object = ''
        for iteration in [first, second, third, fourth, first]:
            num = 1
            for i in iteration:
                if num == 3:
                    string_object += f"{i} "
                else:
                    string_object += f"{i},"
                num += 1
        text_d = '' \
            # f"""<Placemark>
        # 	<name>object</name>
        # 	<Polygon>
        # 		<outerBoundaryIs>
        # 			<LinearRing>
        # 				<coordinates>
        # 					{string_object}
        # 				</coordinates>
        # 			</LinearRing>
        # 		</outerBoundaryIs>
        # 	</Polygon>
        # </Placemark>"""
        return text_d

    @staticmethod
    def create_point_object(point: list):
        latitude = point[0]
        longitude = point[1]
        id_s = point[2]
        # first = [latitude - 0.000002 * 1.62, longitude - 0.000002, 0]
        # second = [latitude - 0.000002 * 1.62, longitude + 0.000002, 0]
        # third = [latitude + 0.000002 * 1.62, longitude + 0.000002, 0]
        # fourth = [latitude + 0.000002 * 1.62, longitude - 0.000002, 0]
        # string_object = ''
        # for iteration in [first, second, third, fourth, first]:
        #     num = 1
        #     for i in iteration:
        #         if num == 3:
        #             string_object += f"{i} "
        #         else:
        #             string_object += f"{i},"
        #         num += 1
        text_d = f"""<Placemark>
                <name>Точка: {id_s}</name>
                <Point>
                    <coordinates>{latitude},{longitude},0</coordinates>
                </Point>
            </Placemark>"""
        return text_d

    @staticmethod
    def generate_xlsx():
        # connection = pg.connect(
        #     host="192.168.1.6",
        #     database="navSections",
        #     port="5432",
        #     user="postgres",
        #     password="nF2ArtXK"
        # )
        # # postgresql_select_query = f"SELECT device, navtime, ROUND(CAST(latitude AS numeric),
        # {request.POST['request_value']}), ROUND(CAST(longitude AS numeric), {request.POST['request_value']}) " \
        # #                           "FROM public.navdata_202108 " \
        # #                           f"WHERE device BETWEEN {request.POST['request_between_first']} AND
        # {request.POST['request_between_last']} AND timezone('UTC', to_timestamp(navtime)) >
        # (CURRENT_TIMESTAMP - INTERVAL '{request.POST['request_hours']} hours') AND flags != 64 " \
        # #                           "ORDER BY device, navtime DESC;"
        # postgresql_select_query = f"SELECT device, navtime, ROUND(CAST(latitude AS numeric),
        # {request.POST['request_value']}), ROUND(CAST(longitude AS numeric), {request.POST['request_value']}) " \
        #                           "FROM public.navdata_202108 " \
        #                           f"WHERE device BETWEEN {request.POST['request_between_first']} AND
        #                           {request.POST['request_between_last']} AND timezone('UTC', to_timestamp(navtime)) >
        #                           (CURRENT_TIMESTAMP - INTERVAL '{request.POST['request_minutes']} minutes') AND
        #                           flags != 64 " \
        #                           "ORDER BY device, navtime DESC;"
        # cursor = connection.cursor()
        # cursor.execute(postgresql_select_query)
        # mobile_records = cursor.fetchall()
        mobile_records = []
        cols = ["устройство", "дата и время", "широта", "долгота", "высота", "скорость", "ds", "направление",
                "флаги ошибок"]
        all_arr = []
        for rows in mobile_records:
            arr = []
            for value in rows:
                id_s = rows.index(value)
                if id_s == 1:
                    arr.append(datetime.datetime.fromtimestamp(int(value - 21600)).strftime('%Y-%m-%d %H:%M:%S'))
                else:
                    arr.append(value)
            all_arr.append(arr)
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Страница 1'
        for n in cols:
            sheet[f'{get_column_letter(cols.index(n) + 1)}1'] = n
        for n in all_arr:
            for i in n:
                if i:
                    sheet[f'{get_column_letter(n.index(i) + 1)}{all_arr.index(n) + 2}'] = i
                else:
                    sheet[f'{get_column_letter(n.index(i) + 1)}{all_arr.index(n) + 2}'] = '0.0'
        wb.save('static/media/data/geo.xlsx')
        postgresql_select_query = None
        return [cols, all_arr, postgresql_select_query]

    @staticmethod
    def generate_kml():
        # Чтение Excel
        file_xlsx = 'static/media/data/geo_1.xlsx'
        workbook = openpyxl.load_workbook(file_xlsx)
        sheet = workbook.active

        # Чтение из Excel
        final = 0
        for num in range(2, 10000000):
            if ExcelClass.get_sheet_value("A", num, sheet=sheet) is None or '':
                final = num
                break
        array = []
        for num in range(2, final):
            array.append(
                [ExcelClass.get_sheet_value("D", num, sheet=sheet),
                 ExcelClass.get_sheet_value("C", num, sheet=sheet),
                 ExcelClass.get_sheet_value("A", num, sheet=sheet)]
            )

        # Генерация объекта и субъекта
        point_obj = [61.22330, 52.14113, 'БелАЗ']
        point_sub = [61.22891, 52.14334, 'Экскаватор']
        text_x = GeoClass.create_point_object(point_obj)
        text_y = GeoClass.create_point_object(point_sub)

        # Цена рёбер
        count = 0
        for current in array:
            index = array.index(current)
            if index != 0:
                previous = [array[index - 1][0], array[index - 1][1]]
            else:
                previous = [current[0], current[1]]

            count += GeoClass.get_hypotenuse(current[0], current[1], previous[0], previous[1])
        print(round(count, 5))

        # Генерация линий по цветам
        device_arr = []
        previous_device = 0
        text_b = ''
        colors = ['FFFFFFFF', 'FF0000FF', 'FFFF0000', 'FF00FF00', 'FF00FFFF', 'FF0F0F0F', 'FFF0F0F0']
        colors_alias = [': Чё', ': Кр', ': Си', ': Зе', ': Жё', ': Доп1', ': Доп2']
        current_color = colors[0]
        for current in array:
            # index = array.index(current)
            if previous_device is not current[2]:
                device_arr.append(str(current[2]) + colors_alias[len(device_arr) + 1])
                previous_device = current[2]
                current_color = colors[colors.index(current_color) + 1]
            # if index != 0:
            #     previous = [array[index - 1][0], array[index - 1][1]]
            # else:
            #     previous = [current[0], current[1]]
            # text_b += f"""<Placemark>
            #       <Style>
            #         <LineStyle>
            #           <color>{current_color}</color>
            #         </LineStyle>
            #       </Style>
            #       <LineString>
            #         <coordinates>{previous[0]},{previous[1]},0 {current[0]},{current[1]},0 </coordinates>
            #       </LineString>
            #   </Placemark>
            # """
            #
            # # Генерация точек
            # text_b += GeoClass.create_cube_object([current[0], current[1], index])

        # Генерация имени документа из цветов
        dev = ''
        for x in device_arr:
            dev += f"{x} | "

        # Начало kml
        text_a = f"""<?xml version="1.0" encoding="utf-8"?>
          <kml xmlns="http://earth.google.com/kml/2.2">
            <Document>
              <name>{dev}</name>
                """

        # Конец kml
        text_c = R"""</Document>
        </kml>"""

        # Запись в kml
        with open("static/media/data/geo.kml", "w", encoding="utf-8") as file:
            file.write(text_a + text_b + text_x + text_y + text_c)

    @staticmethod
    def generate_way(object_, subject_, point_arr):
        # Генерация общей карты
        text_map = GeoClass.generate_map(point_arr, 'FF0000FF')
        # text_map = ''

        # Генерация объекта и субъекта
        point_obj = [object_[0], object_[1], "ЭКГ"]
        point_sub = [subject_[0], subject_[1], "БелАЗ"]
        text_obj = GeoClass.create_point_object(point_obj)
        text_sub = GeoClass.create_point_object(point_sub)

        path = GeoClass.generate_path(object_, subject_, point_arr)
        # print(path[0])
        # Генерация карты пути
        text_path = GeoClass.generate_map(path[0], 'FFFF0000')

        # Генерация имени документа из цветов
        dev = ''
        for x in [object_[0], subject_[0]]:
            dev += f"{x} | "
        dev += f"{path[1]} meters"
        print(f"{path[1]} meters")

        # Начало kml
        text_title = f"""<?xml version="1.0" encoding="utf-8"?>
          <kml xmlns="http://earth.google.com/kml/2.2">
            <Document>
              <name>{dev}</name>
                """

        # Конец kml
        text_footer = R"""</Document>
        </kml>"""

        # Запись в kml
        try:
            os.remove("static/media/data/geo_1.kml")
        except Exception as error:
            DjangoClass.LoggingClass.logging_errors_local(error=error, function_error='generate_way')
        with open("static/media/data/geo_5.kml", "w", encoding="utf-8") as file:
            file.write(text_title + text_map + text_path + text_obj + text_sub + text_footer)

    @staticmethod
    def generate_path(object_, subject_, point_arr):
        """"
        Генерация пути синего цвета из всех валидных точек и расчёт расстояния.
        """
        # Путь маршрут
        path = []
        # Расстояние пути
        path_distantion = 0

        # Откуда начинаем путь
        from_point = object_
        print(f"from_point: {from_point}")
        # from_point: [61.232109, 52.146845, '38', '37|39']

        # Где завершаем путь
        to_point = subject_
        print(f"to_point: {to_point}")
        # to_point: [61.229219, 52.143999, '12', '11|13']

        # Генерация маршрута
        # От исходной точки идём к её линиям связи, выбираем самую короткую к финальной точке, "наступаем" туда, цикл.
        start_point = from_point
        final_point = to_point

        path.append(from_point)
        while True:
            print('\n*****************')
            print('while', start_point[2], final_point[2])
            links = start_point[3].split("|")
            print(f"links: {links}")
            # links: ['38', '40']

            near_points = []
            for point in point_arr:
                for link in links:
                    if point[2] == link:
                        try:
                            near_points.index(point)
                        except Exception as error:
                            DjangoClass.LoggingClass.logging_errors_local(error=error, function_error='generate_path')
                            near_points.append(point)
            print(f"near_points: {near_points}")
            # near_points: [[61.232054, 52.146927, '38', '37|39'], [61.232384, 52.147085, '40', '39|41']]

            destinations = []
            for point in near_points:
                destinations.append(GeoClass.get_haversine(point[0], point[1], final_point[0], final_point[1]))
            print(f"destinations: {destinations}")
            # destinations: [[61.232054, 52.146927, '38', '37|39'], [61.232384, 52.147085, '40', '39|41']]

            min_point = min(destinations)
            print(f"min_point: {min_point}")
            # min_point: 367

            if min_point == 0:
                next_point = final_point
                print(f"next_point: {next_point}")
                # next_point: [61.232101, 52.146821, '45', '44|46']

                dist = GeoClass.get_haversine(start_point[0], start_point[1], final_point[0], final_point[1])
                print(f"dist: {dist}")
                # dist: 22

                path_distantion += dist
                path.append(next_point)
                break

            next_point = near_points[destinations.index(min_point)]
            print(f"next_point: {next_point}")
            # next_point: [61.232101, 52.146821, '45', '44|46']

            dist = GeoClass.get_haversine(start_point[0], start_point[1], next_point[0], next_point[1])
            print(f"dist: {dist}")
            # dist: 22

            path_distantion += dist
            start_point = next_point
            path.append(next_point)
            print('*****************\n')
        return [path, path_distantion]

    @staticmethod
    def generate_path_old(object_, subject_, point_arr):
        """"
        Генерация пути синего цвета из всех валидных точек.
        """
        # Генерация маршрута
        # От исходной точки идём к её линиям связи, выбираем самую короткую к финальной точке, "наступаем" туда, цикл.
        path = []
        path_dist = 0
        current_point = subject_
        # subject_: [61.232230610495556, 52.14697472947569, '21', '20|22']
        previous_point = subject_
        for loop in range(len(point_arr)):
            # links: ['20', '22']
            links = current_point[3].split("|")
            near_points = []
            for point in point_arr:
                for link in links:
                    if point[2] == link:
                        # print(point[2])
                        try:
                            near_points.index(point)
                        except Exception as error:
                            DjangoClass.LoggingClass.logging_errors_local(
                                error=error, function_error='generate_path_old'
                            )
                            near_points.append(point)

            # Первые две линии не брать, а к последней по индексу добавлять ещё одну связь.
            min_dist = 9999
            # print(near_points)
            for point in near_points:
                # object_: [61.2293618582404, 52.143978995225346, '4', '3|5']
                dist = GeoClass.get_haversine(point[0], point[1], object_[0], object_[1])
                # print(f"dist: {dist}")
                if min_dist > dist:
                    min_dist = dist
                    current_point = point
                    path.append(point)
                    distantion = GeoClass.get_haversine(point[0], point[1], previous_point[0], previous_point[1])
                    print(f"prev: {previous_point}   |   curr: {point}     |       deist: {distantion}m")
                    path_dist += distantion
                previous_point = point
        print(path_dist)
        path_dist = 0
        for x in path:
            path_dist += GeoClass.get_haversine(x[0], x[1], object_[0], object_[1])
            print(GeoClass.get_haversine(x[0], x[1], object_[0], object_[1]))
        return [path, path_dist]

    @staticmethod
    def generate_map(point_arr, color):
        """"
        Генерация карты выбранного цвета из всех валидных точек.
        """
        text = ''
        for current in point_arr:
            index = point_arr.index(current)
            if index != 0:
                previous = [point_arr[index - 1][0], point_arr[index - 1][1]]
            else:
                previous = [current[0], current[1]]
            text += f"""<Placemark>
                  <Style>
                    <LineStyle>
                      <color>{color}</color>
                    </LineStyle>
                  </Style>
                  <LineString>
                    <coordinates>{previous[0]},{previous[1]},0 {current[0]},{current[1]},0 </coordinates>
                  </LineString>
              </Placemark>
            """
            # Генерация точек
            text += GeoClass.create_cube_object([current[0], current[1], index])
        return text

    @staticmethod
    def read_kml(val: list):
        """"
        Чтение из kml-файла
        """
        with open("static/media/data/geo.kml", 'rt', encoding="utf-8") as file:
            data = file.read()
        k = kml.KML()
        k.from_string(data)
        features = list(k.features())
        k2 = list(features[0].features())
        arr = []
        for feat in k2:
            string = str(feat.geometry).split('(')[1].split('0.0')[0].split(' ')
            arr.append([float(string[0]), float(string[1])])
        if val is None:
            val = [61.2200083333333, 52.147525]
        val2 = 0
        val3 = 0
        for loop1 in arr:
            # Мы должны найти к какой из точек он ближе(разница двух элементов массива)
            if val[0] > loop1[0]:
                for loop2 in arr:
                    if val[1] > loop2[1]:
                        val2 = loop1[0]
                        val3 = loop1[1]
                        break
        print([val2, val3])
        return [val2, val3]

    @staticmethod
    def create_style():
        # f"""
        # <gx:CascadingStyle kml:id="__managed_style_25130D559F1CA685BFB3">
        # 	<Style>
        # 		<IconStyle>
        # 			<scale>1.2</scale>
        # 			<Icon>
        # 				<href>https://earth.google.com/earth/rpc/cc/icon?color=1976d2&amp;id=2000&amp;scale=4</href>
        # 			</Icon>
        # 			<hotSpot x="64" y="128" xunits="pixels" yunits="insetPixels"/>
        # 		</IconStyle>
        # 		<LabelStyle>
        # 		</LabelStyle>
        # 		<LineStyle>
        # 			<color>ff2dc0fb</color>
        # 			<width>6</width>
        # 		</LineStyle>
        # 		<PolyStyle>
        # 			<color>40ffffff</color>
        # 		</PolyStyle>
        # 		<BalloonStyle>
        # 			<displayMode>hide</displayMode>
        # 		</BalloonStyle>
        # 	</Style>
        # </gx:CascadingStyle>
        # <gx:CascadingStyle kml:id="__managed_style_1A4EFD26461CA685BFB3">
        # 	<Style>
        # 		<IconStyle>
        # 			<Icon>
        # 				<href>https://earth.google.com/earth/rpc/cc/icon?color=1976d2&amp;id=2000&amp;scale=4</href>
        # 			</Icon>
        # 			<hotSpot x="64" y="128" xunits="pixels" yunits="insetPixels"/>
        # 		</IconStyle>
        # 		<LabelStyle>
        # 		</LabelStyle>
        # 		<LineStyle>
        # 			<color>ff2dc0fb</color>
        # 			<width>4</width>
        # 		</LineStyle>
        # 		<PolyStyle>
        # 			<color>40ffffff</color>
        # 		</PolyStyle>
        # 		<BalloonStyle>
        # 			<displayMode>hide</displayMode>
        # 		</BalloonStyle>
        # 	</Style>
        # </gx:CascadingStyle>
        # <StyleMap id="__managed_style_047C2286A81CA685BFB3">
        # 	<Pair>
        # 		<key>normal</key>
        # 		<styleUrl>#__managed_style_1A4EFD26461CA685BFB3</styleUrl>
        # 	</Pair>
        # 	<Pair>
        # 		<key>highlight</key>
        # 		<styleUrl>#__managed_style_25130D559F1CA685BFB3</styleUrl>
        # 	</Pair>
        # </StyleMap>
        # <Placemark id="0D045F86381CA685BFB2">
        # 	<name>Самосвал</name>
        # 	<LookAt>
        # 		<longitude>61.2344061029136</longitude>
        # 		<latitude>52.17019183209385</latitude>
        # 		<altitude>282.7747547496757</altitude>
        # 		<heading>0</heading>
        # 		<tilt>0</tilt>
        # 		<gx:fovy>35</gx:fovy>
        # 		<range>1198.571236050484</range>
        # 		<altitudeMode>absolute</altitudeMode>
        # 	</LookAt>
        # 	<styleUrl>#__managed_style_047C2286A81CA685BFB3</styleUrl>
        # 	<Point>
        # 		<coordinates>61.23500224897153,52.17263169824412,281.7092496784567</coordinates>
        # 	</Point>
        # </Placemark>
        # <Placemark id="0B4EA6F59B1CA68601E0">
        # 	<name>Экскаватор</name>
        # 	<LookAt>
        # 		<longitude>61.23624067458115</longitude>
        # 		<latitude>52.17416232356366</latitude>
        # 		<altitude>277.5968564918906</altitude>
        # 		<heading>-0.5372217869872089</heading>
        # 		<tilt>53.57834275643886</tilt>
        # 		<gx:fovy>35</gx:fovy>
        # 		<range>2536.120178802812</range>
        # 		<altitudeMode>absolute</altitudeMode>
        # 	</LookAt>
        # 	<styleUrl>#__managed_style_047C2286A81CA685BFB3</styleUrl>
        # 	<Point>
        # 		<coordinates>61.23654046107902,52.16710625511239,297.4562999141254</coordinates>
        # 	</Point>
        # </Placemark>"""
        pass


class UtilsClass:
    @staticmethod
    def create_encrypted_password(_random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890',
                                  _length=8):
        password = ''
        for i in range(1, _length + 1):
            password += random.choice(_random_chars)
        return password

    @staticmethod
    def decrypt_text_with_hash(massivsimvolov: str, massivkhesha: str):
        rasshifrovat_tekst = ''
        pozitsiyasimvolakhesha = 0
        dlinakhesha = len(massivkhesha)
        propusk = False
        for num in massivsimvolov:
            if propusk:
                propusk = False
                continue
            nomersimvola = ord(str(num))
            if pozitsiyasimvolakhesha >= dlinakhesha - 1:
                pozitsiyasimvolakhesha = 0
            pozitsiyasimvolakhesha = pozitsiyasimvolakhesha + 1
            simvolkhesha = ord(str(massivkhesha[pozitsiyasimvolakhesha]))
            kod_zashifrovannyy_simvol = nomersimvola - simvolkhesha
            # print(f"nomersimvola:{chr(nomersimvola)}:{nomersimvola}|simvolkhesha:{chr(simvolkhesha)}:{simvolkhesha}")
            zashifrovannyy_simvol = chr(kod_zashifrovannyy_simvol)
            rasshifrovat_tekst = rasshifrovat_tekst + zashifrovannyy_simvol
            if round(simvolkhesha / 2, 0) == simvolkhesha / 2:
                propusk = True
        return rasshifrovat_tekst
