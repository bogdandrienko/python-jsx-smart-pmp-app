import asyncio
import concurrent
import datetime
import json
import multiprocessing
import os
import smtplib
import sys
import threading
import time
import tkinter
import tkinter as tk
import tkinter.ttk as ttk
from concurrent.futures import ThreadPoolExecutor
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor
from multiprocessing import freeze_support
from threading import Thread
from time import sleep
from typing import Union

import PySide6.QtGui as QtGui
import PySide6.QtWidgets as QtWidgets
import aiohttp
import chardet
import cv2
import httplib2
import numpy
import openpyxl
import pandas
import pyodbc
import pyttsx3
import requests
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter


########################################################################################################################


class ExcelClass:
    @staticmethod
    def workbook_create():
        workbook = openpyxl.Workbook()
        return workbook

    @staticmethod
    def workbook_load(excel_file: str):
        workbook = openpyxl.load_workbook(excel_file)
        return workbook

    @staticmethod
    def workbook_activate(workbook):
        sheet = workbook.active
        return sheet

    @staticmethod
    def workbook_save(workbook, excel_file: str):
        try:
            os.remove(excel_file)
        except Exception as error:
            pass
        try:
            workbook.save(excel_file)
        except Exception as error:
            print(f'\n ! Please, close the excel_file! \n: {excel_file} | {error}')

    @staticmethod
    def workbook_close(workbook):
        openpyxl.Workbook.close(workbook)

    @staticmethod
    def set_sheet_title(sheet, page_name='page 1'):
        sheet.title = page_name

    @staticmethod
    def get_sheet_value(col: Union[str, int], row: int, sheet):
        if isinstance(col, int):
            col = ExcelClass.get_column_letter(col)
        value = str(sheet[str(col).upper() + str(row)].value)
        if value == 'None' or value is None:
            value = ''
        else:
            value = str(value)
        return value

    @staticmethod
    def set_sheet_value(col: Union[str, int], row: int, value: str, sheet):
        if isinstance(col, int):
            col = ExcelClass.get_column_letter(col)
        if value == 'None' or value is None:
            value = ''
        sheet[str(col) + str(row)] = str(value)

    @staticmethod
    def get_column_letter(num: int):
        return get_column_letter(num)

    @staticmethod
    def get_max_num_rows(sheet):
        return int(sheet.max_row)

    class Example:
        @staticmethod
        def example_read_from_excel_file_col_int():
            excel_file = 'export.xlsx'
            workbook = ExcelClass.workbook_load(excel_file=excel_file)
            sheet = ExcelClass.workbook_activate(workbook=workbook)
            max_num_rows = ExcelClass.get_max_num_rows(sheet=sheet)
            max_num_cols = 10
            global_list = []
            for row in range(1, max_num_rows + 1):
                local_list = []
                for col in range(1, max_num_cols + 1):
                    value = ExcelClass.get_sheet_value(col=col, row=row, sheet=sheet)
                    local_list.append(value)
                global_list.append(local_list)
            for row in global_list:
                print(row)
            ExcelClass.workbook_close(workbook=workbook)

        @staticmethod
        def example_read_from_excel_file_col_char():
            excel_file = 'export.xlsx'
            workbook = ExcelClass.workbook_load(excel_file=excel_file)
            sheet = ExcelClass.workbook_activate(workbook=workbook)
            max_num_rows = ExcelClass.get_max_num_rows(sheet=sheet)
            char_cols = 'ACDF'
            global_list = []
            for row in range(1, max_num_rows + 1):
                local_list = []
                for col in char_cols:
                    value = ExcelClass.get_sheet_value(col=col, row=row, sheet=sheet)
                    local_list.append(value)
                global_list.append(local_list)
            for row in global_list:
                print(row)
            ExcelClass.workbook_close(workbook=workbook)

        @staticmethod
        def example_write_to_excel_file():
            global_list = [
                ['title_1', 'title_2', 'title_3'],
                ['body_1_1', 'body_1_2', 'body_1_3'],
                ['body_2_1', 'body_2_2', 'body_2_3'],
                ['body_3_1', 'body_3_3', 'body_3_3'],
            ]

            excel_file = 'import.xlsx'
            workbook = ExcelClass.workbook_create()
            sheet = ExcelClass.workbook_activate(workbook=workbook)
            for row in global_list:
                for value in row:
                    ExcelClass.set_sheet_value(
                        col=row.index(value) + 1,
                        row=global_list.index(row) + 1,
                        value=value,
                        sheet=sheet
                    )
            ExcelClass.workbook_save(workbook=workbook, excel_file=excel_file)


class DateTimeUtils:
    @staticmethod
    def get_current_datetime():
        return f"{time.strftime('%Y-%m-%d %H:%M:%S')}"

    @staticmethod
    def get_current_date():
        return f"{time.strftime('%Y-%m-%d')}"

    @staticmethod
    def get_current_time():
        return f"{time.strftime('%H:%M:%S')}"

    @staticmethod
    def get_difference_datetime(days=0.0, hours=0.0, minutes=0.0, seconds=0.0):
        value = datetime.datetime.now() + datetime.timedelta(days=days, hours=hours, minutes=minutes, seconds=seconds)
        return value.replace(tzinfo=datetime.timezone.utc)

    class Example:
        @staticmethod
        def example_get_datetime():
            print(DateTimeUtils.get_current_datetime())

        @staticmethod
        def example_get_difference_datetime():
            print(DateTimeUtils.get_difference_datetime(hours=-2))


class FileReadWriteClass:
    @staticmethod
    def write_to_file(file_name, text_file, mode='a'):
        with open(file_name, mode=mode) as file:
            file.write(text_file)

    @staticmethod
    def read_from_file_lines(file_name, mode='r'):
        with open(file_name, mode=mode) as file:
            return file.readlines()

    @staticmethod
    def read_from_file(file_name, mode='r'):
        with open(file_name, mode=mode) as file:
            return file.read()

    @staticmethod
    def example_write_to_file_add():
        FileReadWriteClass.write_to_file(file_name='file.txt', text_file='text\n', mode='a')

    @staticmethod
    def example_write_to_file_rewrite():
        FileReadWriteClass.write_to_file(file_name='file.txt', text_file='text\n', mode='w')

    @staticmethod
    def example_read_from_file_lines():
        print(FileReadWriteClass.read_from_file_lines(file_name='file.txt', mode='r'))

    @staticmethod
    def example_read_from_file():
        print(FileReadWriteClass.read_from_file(file_name='file.txt', mode='r'))


class LoggingClass:
    @staticmethod
    def logging_to_local_file(logging_text, error_func, file_name='logging.txt', type_write='a'):
        try:
            with open(file_name, type_write) as file:
                file.write(f'\n{DateTimeUtils.get_current_datetime()} | error_func: {error_func} | {logging_text}\n')
        except Exception as error:
            print(f'\n{time.strftime("%x %X")} | error_func: {error_func} | {error}\n')

    class Example:
        @staticmethod
        def example_logging():
            LoggingClass.logging_to_local_file(
                logging_text='example',
                error_func='LoggingClass.Example.example_logging',
                file_name='logging.txt',
                type_write='a'
            )


class DirPathFolderPathClass:
    @staticmethod
    def create_folder_in_this_dir(folder_name='new_folder', current_path=os.path.dirname(os.path.abspath('__file__'))):
        full_path = current_path + f'\\{folder_name}'
        try:
            os.makedirs(full_path)
        except Exception as error:
            # print(f'directory already yet | {error}')
            pass
        finally:
            return full_path

    @staticmethod
    def get_all_files_in_path(path=os.path.dirname(os.path.abspath('__file__'))):
        files_list = []
        for root, dirs, files in os.walk(path, topdown=True):
            for name in files:
                files_list.append(f"{os.path.join(root, name)}")
        return files_list

    @staticmethod
    def get_all_dirs_in_path(path=os.path.dirname(os.path.abspath('__file__'))):
        directories_list = []
        for root, dirs, files in os.walk(path, topdown=True):
            for name in dirs:
                directories_list.append(f"{os.path.join(root, name)}")
        return directories_list

    class Example:
        @staticmethod
        def example_create_folder_in_this_folder():
            path = DirPathFolderPathClass.create_folder_in_this_dir(folder_name='new_folder')
            print(path)

        @staticmethod
        def example_create_folder_in_folder_in_this_folder():
            path = DirPathFolderPathClass.create_folder_in_this_dir(folder_name='new_folder\\new')
            print(path)

        @staticmethod
        def example_create_folder_in_external_folder():
            path = DirPathFolderPathClass.create_folder_in_this_dir(folder_name='new_folder\\new', current_path='C:\\')
            print(path)

        @staticmethod
        def example_get_all_files_in_path():
            files_list = DirPathFolderPathClass.get_all_files_in_path(
                path=r'C:\Project\Github_Projects\python-jsx-smart-pmp-app\web-platform_12_21_dev\app_admin'
            )
            for file in files_list:
                print(file)

        @staticmethod
        def example_get_all_folders_in_path():
            path_list = DirPathFolderPathClass.get_all_dirs_in_path(
                path=r'C:\Project\Github_Projects\python-jsx-smart-pmp-app\web-platform_12_21_dev\app_admin'
            )
            for path in path_list:
                print(path)


class SyncAsyncThreadingPoolExecutorClass:
    class Example:
        @staticmethod
        # Синхронный код
        def example_sync_compute():
            # Начальное время
            start_time = time.time()
            print('start')
            # Генерация ссылок
            page_urls = ["https://en.wikipedia.org/wiki/" + str(i) for i in range(10)]

            # Функция получения текста из ссылки
            def get_page_from_url(url):
                try:
                    # Синхронная библиотека
                    response = requests.get(url=url, timeout=3)
                    value = f'{page_urls.index(url) + 1}: {response.text[0:15]}'
                    print(value)
                    return value
                except Exception as error:
                    get_page_from_url(url=url)

            responses_list = []
            # Цикл для прохода по ссылкам
            for page_url in page_urls:
                resp = get_page_from_url(url=page_url)
                responses_list.append(resp)
            for response_from_list in responses_list:
                print(response_from_list)

            # Финальное время
            print(f"Final time: {round(time.time() - start_time, 1)}")
            print('end')

        @staticmethod
        # Acинхронный код
        def example_async_compute():
            # Начальное время
            start_time = time.time()
            print('start')
            # Генерация ссылок
            page_urls = ["https://en.wikipedia.org/wiki/" + str(i) for i in range(10)]

            # Функция получения текста из ссылки
            async def get_page_from_url(session, url):
                try:
                    # Асинхронная библиотека
                    async with session.get(url) as resp:
                        response = await resp.text()
                    value = f'{page_urls.index(url) + 1}: {response[0:15]}'
                    print(value)
                    return value
                except Exception as error:
                    await get_page_from_url(session=session, url=url)

            # Цикл для прохода по ссылкам
            async def main():
                # Асинхронная библиотека
                async with aiohttp.ClientSession() as session:
                    tasks = []
                    for page_url in page_urls:
                        tasks.append(
                            asyncio.ensure_future(
                                get_page_from_url(
                                    session=session,
                                    url=page_url
                                )
                            )
                        )

                    responses_list = await asyncio.gather(*tasks)

                    for response_from_list in responses_list:
                        print(response_from_list)

            asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
            asyncio.run(main())

            # Финальное время
            print(f"Final time: {round(time.time() - start_time, 1)}")
            print('end')

        @staticmethod
        # Многопоточный Threading код
        def example_threading_compute():
            # Начальное время
            start_time = time.time()
            print('start')
            # Генерация ссылок
            page_urls = ["https://en.wikipedia.org/wiki/" + str(i) for i in range(10)]

            # Функция получения текста из ссылки
            def get_page_from_url(page_url):
                try:
                    # Синхронная библиотека
                    response = requests.get(url=page_url, timeout=3)
                    value = f'{page_urls.index(page_url) + 1}: {response.text[0:15]}'
                    print(value)
                    return value
                except Exception as error:
                    get_page_from_url(page_url=page_url)

            # Цикл для прохода по ссылкам
            for url in page_urls:
                threading.Thread(target=get_page_from_url, args=([url])).start()

            # Финальное время
            print(f"Final time: {round(time.time() - start_time, 1)}")
            print('end')

        @staticmethod
        # Многопоточный ThreadPoolExecutor код
        def example_thread_pool_executor_compute():
            # Начальное время
            start_time = time.time()
            print('start')
            # Генерация ссылок
            page_urls = ["https://en.wikipedia.org/wiki/" + str(i) for i in range(10)]

            # Функция получения текста из ссылки
            def get_page_from_url(page_url):
                try:
                    # Синхронная библиотека
                    response = requests.get(url=page_url, timeout=3)
                    value = f'{page_urls.index(page_url) + 1}: {response.text[0:15]}'
                    print(value)
                    return value
                except Exception as error:
                    get_page_from_url(page_url=page_url)

            # Менеджер контекста для многопотока под ThreadPoolExecutor
            with ThreadPoolExecutor() as executor:
                futures = []
                # Цикл для прохода по ссылкам
                for url in page_urls:
                    futures.append(executor.submit(get_page_from_url, page_url=url))
                for future in concurrent.futures.as_completed(futures):
                    print(future.result())

            # Финальное время
            print(f"Final time: {round(time.time() - start_time, 1)}")
            print('end')

        @staticmethod
        # Функция получения текста из ссылки
        def get_page_from_url_sync(page_url):
            try:
                # Синхронная библиотека
                response = requests.get(url=page_url, timeout=3)
                value = f'{response.text[0:15]}'
                print(value)
                return value
            except Exception as error:
                SyncAsyncThreadingPoolExecutorClass.Example.get_page_from_url_sync(page_url=page_url)

        @staticmethod
        # Мультипоточный ProcessPoolExecutor код
        def example_process_pool_executor_compute():
            # Начальное время
            start_time = time.time()
            print('start')
            # Генерация ссылок
            page_urls = ["https://en.wikipedia.org/wiki/" + str(i) for i in range(10)]

            # Менеджер контекста для многопотока под ThreadPoolExecutor
            with ProcessPoolExecutor() as executor:
                futures = []
                # Цикл для прохода по ссылкам
                for url in page_urls:
                    futures.append(executor.submit(
                        SyncAsyncThreadingPoolExecutorClass.Example.get_page_from_url_sync,
                        page_url=url
                    ))
                for future in concurrent.futures.as_completed(futures):
                    print(future.result())

            # Финальное время
            print(f"Final time: {round(time.time() - start_time, 1)}")
            print('end')


class EncryptingClass:
    @staticmethod
    def encrypt_text(text: str, hash_chars: str):
        chars = 'abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890' + ' _-:' + hash_chars
        chars = set(chars)
        forward_ord_list = []
        for char in chars:
            forward_ord_list.append(ord(char))
        forward_ord_list.sort(reverse=False)
        reverse_ord_list = forward_ord_list.copy()
        reverse_ord_list.sort(reverse=True)
        forward_dictinary = {reverse_ord_list[forward_ord_list.index(x)]: x for x in forward_ord_list}
        return ''.join([chr(forward_dictinary[ord(x)]) for x in text])

    @staticmethod
    def decrypt_text(text: str, hash_chars: str):
        chars = 'abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890' + ' _-:' + hash_chars
        chars = set(chars)
        forward_ord_list = []
        for char in chars:
            forward_ord_list.append(ord(char))
        forward_ord_list.sort(reverse=False)
        reverse_ord_list = forward_ord_list.copy()
        reverse_ord_list.sort(reverse=True)
        reverse_dictinary = {forward_ord_list[reverse_ord_list.index(x)]: x for x in reverse_ord_list}
        return ''.join([chr(reverse_dictinary[ord(x)]) for x in text])

    class Example:
        @staticmethod
        def example_encrypt_text():
            value = EncryptingClass.encrypt_text(text='12345', hash_chars='321')
            print(value)

        @staticmethod
        def example_decrypt_text():
            value = EncryptingClass.decrypt_text(text='wvuts', hash_chars='321')
            print(value)


class TypeVariablesClass:
    @staticmethod
    def example_get_type_of_variable(variable):
        return type(variable)

    @staticmethod
    def example_check_type_of_variable(variable, type_variable: str):
        if type_variable == 'bool':
            if isinstance(variable, bool):
                return True
            else:
                return False
        elif type_variable == 'int':
            if isinstance(variable, int):
                return True
            else:
                return False
        elif type_variable == 'float':
            if isinstance(variable, float):
                return True
            else:
                return False
        elif type_variable == 'str' or type_variable == 'string':
            if isinstance(variable, str):
                return True
            else:
                return False
        elif type_variable == 'list':
            if isinstance(variable, list):
                return True
            else:
                return False
        elif type_variable == 'dict' or type_variable == 'dictionary':
            if isinstance(variable, dict):
                return True
            else:
                return False
        elif type_variable == 'tuple':
            if isinstance(variable, tuple):
                return True
            else:
                return False
        elif type_variable == 'set':
            if isinstance(variable, set):
                return True
            else:
                return False
        else:
            return None

    class Example:
        @staticmethod
        def example_get_type():
            value = TypeVariablesClass.example_get_type_of_variable(variable=12.5)
            print(value)

        @staticmethod
        def example_check_type():
            value = TypeVariablesClass.example_check_type_of_variable(variable=12.5, type_variable='int')
            print(value)


class JsonClass:
    @staticmethod
    def write_json_to_file(dictionary: dict, file_name="json"):
        with open(f'{file_name}.json', 'w') as file:
            json.dump(dictionary, file)

    @staticmethod
    def read_json_from_file(file_name="json"):
        with open(f'{file_name}.json', 'r') as file:
            return json.load(file)

    class Example:
        @staticmethod
        def example_write_json_to_file():
            data = {
                'key_1': 'value_1',
                'key_2': 3,
                'key_3': {
                    'key_3_1': True,
                    'key_3_2': 'value_3_2',
                }
            }
            JsonClass.write_json_to_file(dictionary=data, file_name='json')

        @staticmethod
        def example_read_json_to_file():
            dictionary = JsonClass.read_json_from_file(file_name='json')
            print(dictionary)
            print(dictionary['key_1'])


class DictionaryClass:
    @staticmethod
    def get_all_keys(dictionary: dict):
        return dictionary.keys()

    @staticmethod
    def get_all_values(dictionary: dict):
        return dictionary.values()

    @staticmethod
    def get_all_sources(dictionary: dict, values: dict):
        dictionary_local = dictionary.copy()
        for key, value in values.items():
            dictionary_local[key] = value
        return dictionary_local


class EncodingClass:
    @staticmethod
    def find_encoding(text: Union[str, bytes]):
        if isinstance(text, str):
            text = text.encode()
        return chardet.detect(text)['encoding']

    @staticmethod
    def convert_encoding(text: Union[str, bytes], encoding=None, new_encoding='utf-8'):
        if encoding is None:
            encoding = EncodingClass.find_encoding(text)
        if isinstance(text, str):
            text = text.encode(encoding=encoding)
        return text.decode(encoding=new_encoding)

    class Example:
        @staticmethod
        def example_find_encoding():
            text = 'РђР»РµРєСЃР°РЅРґСЂР° РџСЂРѕРєРѕС„СЊРµРІР°'
            encoding = EncodingClass.find_encoding(text=text.encode(encoding='utf-8'))
            print(encoding)

        @staticmethod
        def example_convert_encoding():
            text = '\xd0\xa0\xd1\x92\xd0\xa0\xc2'
            text = 'РџСЂРѕРєРѕС„СЊРµРІР°'
            encoding = EncodingClass.find_encoding(text=text)
            text = EncodingClass.convert_encoding(text=text, encoding='Windows-1251', new_encoding='utf-8')
            print(text)


class ParserBeautifulSoupClass:
    @staticmethod
    def get_url():
        pass

    @staticmethod
    def parse_local_html():
        pass

    class Example:
        @staticmethod
        def example_parse_weather():
            arr_url = []
            for year in range(2011, 2012):
                for month in range(1, 2):
                    for day in range(1, 32):
                        url = f"http://www.pogodaiklimat.ru/weather.php?id=35042&b" \
                              f"day={day}&fday={day}&amonth={month}&ayear={year}&bot=2"
                        arr_url.append(url)

            arr_responces = []
            for url in arr_url:
                headers = {'user-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:72.0) Gecko/20100101 Firefox/72.0',
                           'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'}
                response = requests.get(url, headers=headers)
                response.encoding = 'utf-8'
                soup = BeautifulSoup(response.text, "html.parser")
                soup.encoded = 'utf-8'
                rows = soup.find('div', class_='archive-table-wrap')
                arr_responces.append(rows)

            all_data = []
            for request in arr_responces:
                data_list = []
                for row in request.find_all('tr'):
                    local_data = []
                    for col in row.find_all('td'):
                        local_data.append(col.text)
                    data_list.append(local_data)
                all_data.append([''])
                all_data.append(data_list)
            for data in all_data:
                print(data)

        @staticmethod
        def example_perser_exams_kiney():
            version = 3
            with open(f'{version}_data.html', 'r+', encoding='utf-8') as file:
                questions = []
                index = 1
                for lines in file:
                    c1 = f"""
                    <div id="content_test"><br><div class="workspace-header"><i>Самопроверка промежуточного рейтинга 1</i></div><p><b class="question" id="1">Сұрақ № 1 / Вопрос № 1 </b>АСУ из регулятора и объекта управления, называется:</p><p><input type="radio" name="variant" value="1" id="ans1"><label for="ans1">автоматической системой регулирования АСР</label></p><p><input type="radio" name="variant" value="2" id="ans2"><label for="ans2">дистанционным управлением</label></p><p><input type="radio" name="variant" value="3" id="ans3"><label for="ans3">автоматической системой управления АСУ</label></p><p><input type="radio" name="variant" value="4" id="ans4"><label for="ans4">автоматическим управлением</label></p><p><input type="radio" name="variant" value="5" id="ans5"><label for="ans5">комплексной автоматизацией</label></p><button class="btn btn-number" number="1" style="width:40px;margin:5px;border: 2px solid #e34761;">1</button><button class="btn btn-number" number="2" style="width:40px;margin:5px;">2</button><button class="btn btn-number" number="3" style="width:40px;margin:5px;">3</button><button class="btn btn-number" number="4" style="width:40px;margin:5px;">4</button><button class="btn btn-number" number="5" style="width:40px;margin:5px;">5</button><button class="btn btn-number" number="6" style="width:40px;margin:5px;">6</button><button class="btn btn-number" number="7" style="width:40px;margin:5px;">7</button><button class="btn btn-number" number="8" style="width:40px;margin:5px;">8</button><button class="btn btn-number" number="9" style="width:40px;margin:5px;">9</button><button class="btn btn-number" number="10" style="width:40px;margin:5px;">10</button><button class="btn btn-number" number="11" style="width:40px;margin:5px;">11</button><button class="btn btn-number" number="12" style="width:40px;margin:5px;">12</button><button class="btn btn-number" number="13" style="width:40px;margin:5px;">13</button><button class="btn btn-number" number="14" style="width:40px;margin:5px;">14</button><button class="btn btn-number" number="15" style="width:40px;margin:5px;">15</button><button class="btn btn-number" number="16" style="width:40px;margin:5px;">16</button><button class="btn btn-number" number="17" style="width:40px;margin:5px;">17</button><button class="btn btn-number" number="18" style="width:40px;margin:5px;">18</button><button class="btn btn-number" number="19" style="width:40px;margin:5px;">19</button><button class="btn btn-number" number="20" style="width:40px;margin:5px;">20</button><button class="btn btn-number" number="21" style="width:40px;margin:5px;">21</button><button class="btn btn-number" number="22" style="width:40px;margin:5px;">22</button><button class="btn btn-number" number="23" style="width:40px;margin:5px;">23</button><button class="btn btn-number" number="24" style="width:40px;margin:5px;">24</button><button class="btn btn-number" number="25" style="width:40px;margin:5px;">25</button><button class="btn btn-number" number="26" style="width:40px;margin:5px;">26</button><button class="btn btn-number" number="27" style="width:40px;margin:5px;">27</button><button class="btn btn-number" number="28" style="width:40px;margin:5px;">28</button><button class="btn btn-number" number="29" style="width:40px;margin:5px;">29</button><button class="btn btn-number" number="30" style="width:40px;margin:5px;">30</button><button class="btn btn-number" number="31" style="width:40px;margin:5px;">31</button><button class="btn btn-number" number="32" style="width:40px;margin:5px;">32</button><button class="btn btn-number" number="33" style="width:40px;margin:5px;">33</button><button class="btn btn-number" number="34" style="width:40px;margin:5px;">34</button><button class="btn btn-number" number="35" style="width:40px;margin:5px;">35</button><button class="btn btn-number" number="36" style="width:40px;margin:5px;">36</button><button class="btn btn-number" number="37" style="width:40px;margin:5px;">37</button><button class="btn btn-number" number="38" style="width:40px;margin:5px;">38</button><button class="btn btn-number" number="39" style="width:40px;margin:5px;">39</button><button class="btn btn-number" number="40" style="width:40px;margin:5px;">40</button><hr></div>
<div id="content_test"><br><div class="w
                    """
                    final_data = f"<strong>{index}</strong><br><hr><br>"
                    data = lines.split(" </b>")[1].split("<button class=")[0]
                    try:
                        new_data = data.split('<img src="')
                        if len(new_data) > 2:
                            final_data += data
                        else:
                            final_data += data.split('<img src="')[0] + """<img src="https://sdo.kineu.kz""" + \
                                          data.split('<img src="')[1]
                    except Exception as ex:
                        final_data += data
                    final_data += "<br><hr><br>"
                    reverse = False
                    for question in questions:
                        if question.split("</strong>")[1] == final_data.split("</strong>")[1]:
                            print('повторение!')
                            reverse = True
                            break
                    if reverse is False:
                        questions.append(final_data)
                        index += 1
            with open(f'{version}_new_data.html', 'w', encoding='utf-8') as file:
                title = """<!DOCTYPE html>
            <html lang="en">
            <head>
                <meta charset="UTF-8">
                <title>Title</title>
            </head>
            <body>
                """
                footer = """
                </body>
            </html>"""
                file.write(title)
                for line in questions:
                    file.write(line)
                file.write(footer)


class CycleClass:
    class Example:
        @staticmethod
        def example_cycle_for():
            for i in range(1, 10):
                print(i)

        @staticmethod
        def example_cycle_while():
            sec = 10
            while sec < 50:
                sec += 1
                print(sec)


class SendMail:
    class Example:
        @staticmethod
        def example_send_email(subject='subj', text='text'):
            host = 'smtp.yandex.ru'
            port = '465'
            login = 'eevee.cycle'
            password = '31284bogdan'
            writer = 'eevee.cycle@yandex.ru'
            recipient = 'eevee.cycle@yandex.ru'

            message = f"""From: {recipient}\nTo: {writer}\nSubject: {subject}\n\n{text}"""

            smtpobj = smtplib.SMTP_SSL(host=host, port=port)
            smtpobj.ehlo()
            smtpobj.login(user=login, password=password)
            smtpobj.sendmail(from_addr=writer, to_addrs=recipient, msg=message)
            smtpobj.quit()


class Pagination:
    class Example:
        @staticmethod
        def example():
            alpabet_list = list('abcdefghhklmnt')
            p = Pagination.Example(alpabet_list, 4)
            print(p.get_visible_items())

        def __init__(self, items=None, page_size=10):
            if items is None:
                items = []
            self.items = items
            self.page_size = page_size
            self.total_pages = 1 if not self.items else len(self.items) // self.page_size + 1
            self.current_page = 1

        def get_items(self):
            return self.items

        def get_page_size(self):
            return self.page_size

        def get_current_page(self):
            return self.current_page

        def prev_page(self):
            if self.current_page == 1:
                return self
            self.current_page -= 1
            return self

        def next_page(self):
            if self.current_page == self.total_pages:
                return self
            self.current_page += 1
            return self

        def first_page(self):
            self.current_page = 1
            return self

        def last_page(self):
            self.current_page = self.total_pages
            return self

        def go_to_page(self, page):
            if page < 1:
                page = 1
            elif page > self.total_pages:
                page = self.total_pages
            self.current_page = page
            return self

        def get_visible_items(self):
            start = (self.current_page - 1) * self.page_size
            return self.items[start:start + self.page_size]


########################################################################################################################


class VoiceClass:
    @staticmethod
    def example(self):
        # Not create class:
        self.speak()
        self.speak('Приветики!')
        # With create class:
        voice = self()
        voice.say()
        voice.say('Я уничтожу человечество!!!')

    def __init__(self, obj_id=0, volume=1.0, rate=200, voice=None, voices=None):
        self.engine = pyttsx3.init()
        self.volume = volume
        self.rate = rate
        self.voice = voice
        self.voices = voices
        self.obj_id = obj_id
        self.properties = [self.volume, self.rate, self.voice, self.voices]

    @staticmethod
    def speak(text: str = 'Inicialization successfull.'):
        pyttsx3.speak(text)

    @staticmethod
    async def async_speak(text: str = 'Inicialization successfull.'):
        await pyttsx3.speak(text)
        return text

    def say(self, text: str = 'Inicialization successfull.'):
        self.engine.say(text)
        self.engine.runAndWait()

    def get_property(self, name='volume'):
        return self.engine.getProperty(name)

    def get_properties(self):
        return [self.engine.getProperty(name) for name in self.properties]

    def set_property(self, name='volume', value=1.0):
        self.engine.setProperty(name, value)

    def set_properties(self, volume=1.0, rate=200):
        self.set_property(self.properties[0], volume)
        self.set_property(self.properties[1], rate)


class ObjectOrientedProgrammingClass:
    @staticmethod
    def example_worker():
        class Worker:
            """
            Класс, который содержит в себе работника, со значениями по строке
            """

            def __init__(self, A_1='', B_1='', C_1='', D_1='', E_1='', F_1='', G_1='', H_1='', M_1=''):
                # Подразделение
                self.A_1 = A_1
                # Цех или Служба
                self.B_1 = B_1
                # Отдел или участок
                self.C_1 = C_1
                # Фамилия
                self.D_1 = D_1
                # Имя
                self.E_1 = E_1
                # Отчество
                self.F_1 = F_1
                # Табельный №
                self.G_1 = G_1
                # Категория
                self.H_1 = H_1
                # Пол
                self.M_1 = M_1

            def print_worker(self):
                """

                """
                print(
                    f'{self.A_1}+{self.B_1}+{self.C_1}+{self.D_1}+{self.E_1}+{self.F_1}+{self.G_1}+{self.H_1}+{self.M_1}')

            def get_worker_value(self, index):
                """

                :param index:
                :return:
                """
                value_ = list(
                    (self.A_1, self.B_1, self.C_1, self.D_1, self.E_1, self.F_1, self.G_1, self.H_1, self.M_1))
                return value_[index]

            def get_worker_id(self):
                """

                :return:
                """
                return self.G_1

    @staticmethod
    def example_objects():
        class Actions:
            def __init__(self):
                pass

            @staticmethod
            def action(first_value, second_value):
                return first_value + second_value

        print(Actions.action(10, 30))

        class Circle:
            pi = 3.14

            def __init__(self, radius=1):
                self.radius = radius
                self.circle_reference = 2 * self.pi * self.radius

            def get_area(self):
                return self.pi * (self.radius ** 2)

            def get_circle_reference(self):
                return 2 * self.pi * self.radius

        circle_1 = Circle(4)
        print(circle_1.get_area())
        print(circle_1.circle_reference)
        print(circle_1.get_circle_reference())

        # Классы
        class Car:
            wheels_number = 4

            def __init__(self, name, color, year, is_crashed):
                self.name = name
                self.color = color
                self.year = year
                self.is_crashed = is_crashed

        mazda_car = Car(name="Mazda CX7", color="red", year=2017, is_crashed=True)
        print(mazda_car.name, mazda_car.is_crashed, mazda_car.wheels_number)
        bmw_car = Car(name="Mazda", color="black", year=2019, is_crashed=False)
        print(bmw_car.name, bmw_car.is_crashed, bmw_car.wheels_number)
        print(Car.wheels_number * 3)


########################################################################################################################

class ExampleClass:
    @staticmethod
    def example():
        # Сравнение данных из 1с и данных в базе данных енбека
        def extra_enbek():
            start_time = time.time()
            print('start')

            input_1 = '1c.xlsx'
            input_2 = 'enbek.xlsx'
            output_1 = 'final.xlsx'

            workbook = ExcelClass.workbook_load(excel_file=input_1)
            sheet = ExcelClass.workbook_activate(workbook=workbook)
            max_num_rows = ExcelClass.get_max_num_rows(sheet=sheet)

            global_workers_id_list = []
            for row in range(1, max_num_rows + 1):
                global_workers_id_list.append(
                    ExcelClass.get_sheet_value(col=ExcelClass.get_column_letter(11), row=row, sheet=sheet)
                )
            print(global_workers_id_list)
            ExcelClass.workbook_close(workbook=workbook)

            workbook = ExcelClass.workbook_load(excel_file=input_2)
            sheet = ExcelClass.workbook_activate(workbook=workbook)
            max_num_rows = ExcelClass.get_max_num_rows(sheet=sheet)
            global_workers_enbek_list = []
            for row in range(1, max_num_rows + 1):
                local_workers_enbek_list = []
                for col in range(1, 21 + 1):
                    local_workers_enbek_list.append(
                        ExcelClass.get_sheet_value(col=ExcelClass.get_column_letter(col), row=row, sheet=sheet)
                    )
                if local_workers_enbek_list[1] in global_workers_id_list:
                    local_workers_enbek_list.append('+')
                else:
                    local_workers_enbek_list.append('-')
                global_workers_enbek_list.append(local_workers_enbek_list)
            print(global_workers_enbek_list)
            ExcelClass.workbook_close(workbook=workbook)

            workbook = ExcelClass.workbook_load(excel_file=output_1)
            sheet = ExcelClass.workbook_activate(workbook=workbook)
            for row in global_workers_enbek_list:
                print(row)
                for value in row:
                    ExcelClass.set_sheet_value(
                        col=ExcelClass.get_column_letter(row.index(value) + 1),
                        row=global_workers_enbek_list.index(row) + 1,
                        value=value,
                        sheet=sheet
                    )
            ExcelClass.workbook_save(workbook=workbook, excel_file=output_1)

            # Финальное время
            print(f"Final time: {round(time.time() - start_time, 1)}")
            print('end')

        threading.Thread(target=extra_enbek, args=()).start()


class OpenCvClass:
    @staticmethod
    def example_computer_vision():
        # UTILITES
        class LoggingClass:
            @staticmethod
            def logging(message, file_name='log.txt', type_write='a'):
                with open(file_name, type_write) as log:
                    log.write(f'{TimeUtils.get_current_time()} : {message}\n')

        class CopyDictionary:
            @staticmethod
            def get_all_sources(source: dict, values: dict):
                value = source.copy()
                for _key, _value in values.items():
                    value[_key] = _value
                return value

        class FileSettings:
            @staticmethod
            def export_settings(data: dict):
                with open(f"{data['import_file']}.json", 'w') as file:
                    json.dump(data, file)

            @staticmethod
            def import_settings(data: dict):
                try:
                    with open(f"{data['import_file']}.json", "r") as read_file:
                        data = json.load(read_file)
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'import_settings error : {ex}')
                return data

        class TimeUtils:
            @staticmethod
            def get_current_time():
                return f"{time.strftime('%X')}"

        class SendMail:
            @staticmethod
            def sender_email(subject='subj', text='text'):
                host = 'smtp.yandex.ru'
                port = '465'
                login = 'eevee.cycle'
                password = '31284bogdan'
                writer = 'eevee.cycle@yandex.ru'
                recipient = 'eevee.cycle@yandex.ru'

                message = f"""From: {recipient}\nTo: {writer}\nSubject: {subject}\n\n{text}"""

                smtpobj = smtplib.SMTP_SSL(host=host, port=port)
                smtpobj.ehlo()
                smtpobj.login(user=login, password=password)
                smtpobj.sendmail(from_addr=writer, to_addrs=recipient, msg=message)
                smtpobj.quit()

        # SQL
        class SQLClass:
            @staticmethod
            def sql_post_data(ip: str, server: str, port: str, database: str, username: str, password: str, table: str,
                              rows: list, values: list):
                try:
                    sql = SQLClass.pyodbc_connect(ip=ip, server=server, port=port, database=database, username=username,
                                                  password=password)
                    SQLClass.execute_data_query(connection=sql, table=table, rows=rows, values=values)
                except Exception as ex:
                    print(ex)
                    with open('log.txt', 'a') as log:
                        log.write(f'\n{ex}\n')

            @staticmethod
            def sql_post_now(ip: str, server: str, port: str, database: str, username: str, password: str, table: str,
                             rows: list, values: list):
                try:
                    sql = SQLClass.pyodbc_connect(ip=ip, server=server, port=port, database=database, username=username,
                                                  password=password)
                    SQLClass.execute_now_query(connection=sql, table=table, rows=rows, values=values)
                except Exception as ex:
                    print(ex)
                    with open('log.txt', 'a') as log:
                        log.write(f'\n{ex}\n')

            @staticmethod
            def pyodbc_connect(ip: str, server: str, port: str, database: str, username: str, password: str):
                conn_str = (
                        r'DRIVER={ODBC Driver 17 for SQL Server};SERVER=tcp:' + ip + '\\' + server + ',' + port +
                        ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password + ';'
                )
                return pyodbc.connect(conn_str)

            @staticmethod
            def pd_read_sql_query(connection, query: str, database: str, table: str):
                return pandas.read_sql_query(f'{query} {database}.dbo.{table} ORDER BY id', connection)

            @staticmethod
            def execute_data_query(connection, table: str, rows: list, values: list):
                cursor = connection.cursor()
                cursor.fast_executemany = True
                _rows = ''
                for x in rows:
                    _rows = f"{_rows}{str(x)}, "
                value = f"INSERT INTO {table} (" + _rows[:-2:] + f") VALUES {tuple(values)}"
                cursor.execute(value)
                connection.commit()

            @staticmethod
            def execute_now_query(connection, table, rows: list, values: list):
                cursor = connection.cursor()
                cursor.fast_executemany = True
                __rows = ''
                for x in rows:
                    __rows = f"{__rows}{str(x)}, "
                value = f"UPDATE {table} SET {rows[1]} = '{values[1]}',{rows[2]} = '{values[2]}' ,{rows[3]} = '{values[3]}' " \
                        f"WHERE {rows[0]} = '{values[0]}'"
                cursor.execute(value)
                connection.commit()

        # OPENCV
        class AnalyzeClass:
            @staticmethod
            def start_analyze(data: dict):
                sources = AnalyzeClass.get_full_sources(source_type=data['source_type'], sources=data['ip_cam'],
                                                        masks=data['mask_cam'], protocol=data['protocol_cam_type'],
                                                        login=data['login_cam'], password=data['password_cam'],
                                                        port=data['port_cam'], sensitivity=data['sensitivity_analysis'],
                                                        correct=data['correct_coefficient'],
                                                        alias_device=data['alias_device'],
                                                        alarm_level=data['alarm_level'])
                data = CopyDictionary.get_all_sources(data, {'sources': sources})
                if data['compute_debug'] == 'sync':
                    AnalyzeClass.sync_method(data)
                elif data['compute_debug'] == 'async':
                    AnalyzeClass.async_method(data)
                elif data['compute_debug'] == 'multithread':
                    AnalyzeClass.multithread_method(data)
                elif data['compute_debug'] == 'multiprocess':
                    AnalyzeClass.multiprocess_method(data)
                elif data['compute_debug'] == 'complex':
                    AnalyzeClass.complex_method(data)

            @staticmethod
            def sync_method(data: dict):
                while True:
                    if not data['pause']():
                        cv2.destroyAllWindows()
                        break
                    else:
                        for source in data['sources']:
                            AnalyzeClass.analyze(source, data)
                    time.sleep(1 / data['speed_analysis'])

            @staticmethod
            def async_method(data: dict):
                def thread_loop(source):
                    while True:
                        if not data['pause']():
                            cv2.destroyAllWindows()
                            break
                        else:
                            asyncio.run(AnalyzeClass.async_analiz(source, data))
                        time.sleep(1 / data['speed_analysis'])

                for src in data['sources']:
                    threading.Thread(target=thread_loop, args=(src,)).start()

            @staticmethod
            async def async_analiz(source, data: dict):
                image = await AnalyzeClass.async_get_source(data['source_type'], source[0], data['login_cam'],
                                                            data['password_cam'])
                mask = source[1]
                name = source[0].split("192.168.")[1].strip().split(":")[0].strip()
                AnalyzeClass.render_final(image=image, mask=mask,
                                          sensitivity_analysis=source[2],
                                          correct_coefficient=source[3],
                                          name=name,
                                          resolution_debug=data['resolution_debug'])
                cv2.waitKey(round(1000 / data['speed_video_stream']))
                values = AnalyzeClass.result_final(image=image, mask=mask,
                                                   sensitivity_analysis=source[2],
                                                   correct_coefficient=source[3])
                AnalyzeClass.write_result(ip_sql=data['ip_sql'],
                                          server_sql=data['server_sql'],
                                          port_sql=data['port_sql'],
                                          database_sql=data['database_sql'],
                                          user_sql=data['user_sql'],
                                          password_sql=data['password_sql'],
                                          sql_now_check=data['sql_now_check'],
                                          table_now_sql=data['table_now_sql'],
                                          rows_now_sql=data['rows_now_sql'],
                                          sql_data_check=data['sql_data_check'],
                                          table_data_sql=data['table_data_sql'],
                                          rows_data_sql=data['rows_data_sql'],
                                          values=values,
                                          source=name,
                                          widget=data['widget'],
                                          widget_write=data['widget_write'],
                                          text_write=data['text_write'],
                                          sql_val=data['sql_write'],
                                          alarm_level=source[5]
                                          )

            @staticmethod
            async def async_get_source(source_type: str, sources: str, login: str, password: str):
                if source_type == 'image-http':
                    h = httplib2.Http(os.path.abspath('__file__'))
                    h.add_credentials(login, password)
                    response, content = h.request(sources)
                    image = cv2.imdecode(numpy.frombuffer(content, numpy.uint8), cv2.IMREAD_COLOR)
                    return image
                elif source_type == 'video-rtsp' or 'video-file':
                    cam_stream = cv2.VideoCapture(sources)
                    _, image = cam_stream.read()
                    cam_stream.release()
                    return image

            @staticmethod
            def multithread_method(data: dict):
                def thread_loop(source):
                    while True:
                        if not data['pause']():
                            cv2.destroyAllWindows()
                            break
                        else:
                            AnalyzeClass.analyze(source, data)
                        time.sleep(1 / data['speed_analysis'])

                for src in data['sources']:
                    threading.Thread(target=thread_loop, args=(src,)).start()

            @staticmethod
            def multiprocess_method(data: dict):
                def process_loop():
                    while True:
                        if not data['pause']():
                            cv2.destroyAllWindows()
                            break
                        else:
                            with multiprocessing.Pool(data['process_cores']) as process:
                                process.map(AnalyzeClass.multi, _data)
                        time.sleep(1 / data['speed_analysis'])

                _data = []
                for loop in data['sources']:
                    __data = data.copy()
                    del __data['widget']
                    __data['source'] = loop
                    _data.append(__data)
                threading.Thread(target=process_loop).start()

            @staticmethod
            def multi(kwargs):
                source = kwargs['source']
                kwargs['widget'] = None
                AnalyzeClass.analyze(source, kwargs)

            @staticmethod
            def complex_method(data: dict):
                def thread_loop(source):
                    while True:
                        if not data['pause']():
                            cv2.destroyAllWindows()
                            break
                        else:
                            asyncio.run(AnalyzeClass.async_complex_analiz(source, data))
                        time.sleep(1 / data['speed_analysis'])

                for src in data['sources']:
                    threading.Thread(target=thread_loop, args=(src,)).start()

            @staticmethod
            async def async_complex_analiz(source, data: dict):
                try:
                    image = await AnalyzeClass.async_get_source(data['source_type'], source[0], data['login_cam'],
                                                                data['password_cam'])
                    mask = source[1]
                    name = source[4]
                    if data['render_debug'] == 'all':
                        AnalyzeClass.render_flip(image=image, flipcode=0, name=name,
                                                 resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_cvtcolor(image=image, color_type=cv2.COLOR_RGB2GRAY, name=name,
                                                     resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_shapes(name=name, resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_origin(image=image, name=name,
                                                   resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_cropping_image(image=image, name=name,
                                                           resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_bitwise_and(image=image, mask=mask, name=name,
                                                        resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_bitwise_not_white(image=image, mask=mask, name=name,
                                                              resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_cvtcolor_to_hsv(image=image, name=name,
                                                            resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_threshold(image=image, name=name,
                                                      resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_inrange(image=image, sensitivity_analysis=source[2],
                                                    name=name,
                                                    resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_canny_edges(image=image,
                                                        sensitivity_analysis=source[2],
                                                        name=name,
                                                        resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_final(image=image, mask=mask,
                                                  sensitivity_analysis=source[2],
                                                  correct_coefficient=source[3],
                                                  name=name,
                                                  resolution_debug=data['resolution_debug'])
                    elif data['render_debug'] == 'extended':
                        AnalyzeClass.render_origin(image=image, name=name,
                                                   resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_inrange(image=image, sensitivity_analysis=source[2],
                                                    name=name,
                                                    resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_final(image=image, mask=mask,
                                                  sensitivity_analysis=source[2],
                                                  correct_coefficient=source[3],
                                                  name=name,
                                                  resolution_debug=data['resolution_debug'])
                    elif data['render_debug'] == 'final':
                        AnalyzeClass.render_final(image=image, mask=mask,
                                                  sensitivity_analysis=source[2],
                                                  correct_coefficient=source[3],
                                                  name=name,
                                                  resolution_debug=data['resolution_debug'])
                    elif data['render_debug'] == 'source':
                        AnalyzeClass.render_origin(image=image, name=name,
                                                   resolution_debug=data['resolution_debug'])
                    cv2.waitKey(round(1000 / data['speed_video_stream']))
                    values = AnalyzeClass.result_final(image=image, mask=mask,
                                                       sensitivity_analysis=source[2],
                                                       correct_coefficient=source[3])
                    AnalyzeClass.write_result(ip_sql=data['ip_sql'],
                                              server_sql=data['server_sql'],
                                              port_sql=data['port_sql'],
                                              database_sql=data['database_sql'],
                                              user_sql=data['user_sql'],
                                              password_sql=data['password_sql'],
                                              sql_now_check=data['sql_now_check'],
                                              table_now_sql=data['table_now_sql'],
                                              rows_now_sql=data['rows_now_sql'],
                                              sql_data_check=data['sql_data_check'],
                                              table_data_sql=data['table_data_sql'],
                                              rows_data_sql=data['rows_data_sql'],
                                              values=values,
                                              source=name,
                                              widget=data['widget'],
                                              widget_write=data['widget_write'],
                                              text_write=data['text_write'],
                                              sql_val=data['sql_write'],
                                              alarm_level=source[5])
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'analyze func error')
                    print(ex)
                    # SendMail.sender_email(subject='error', text='analyze func error')

            @staticmethod
            def analyze(source, data: dict):
                try:
                    image = AnalyzeClass.get_source(data['source_type'], source[0], data['login_cam'],
                                                    data['password_cam'])
                    mask = source[1]
                    name = source[0].split("192.168.")[1].strip().split(":")[0].strip()
                    if data['render_debug'] == 'all':
                        AnalyzeClass.render_origin(image=image, name=name,
                                                   resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_cropping_image(image=image, name=name,
                                                           resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_bitwise_and(image=image, mask=mask, name=name,
                                                        resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_bitwise_not_white(image=image, mask=mask, name=name,
                                                              resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_cvtcolor_to_hsv(image=image, name=name,
                                                            resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_threshold(image=image, name=name,
                                                      resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_inrange(image=image, sensitivity_analysis=source[2],
                                                    name=name,
                                                    resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_canny_edges(image=image,
                                                        sensitivity_analysis=source[2],
                                                        name=name,
                                                        resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_final(image=image, mask=mask,
                                                  sensitivity_analysis=source[2],
                                                  correct_coefficient=source[3],
                                                  name=name,
                                                  resolution_debug=data['resolution_debug'])
                    elif data['render_debug'] == 'extended':
                        AnalyzeClass.render_origin(image=image, name=name,
                                                   resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_inrange(image=image, sensitivity_analysis=source[2],
                                                    name=name,
                                                    resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_final(image=image, mask=mask,
                                                  sensitivity_analysis=source[2],
                                                  correct_coefficient=source[3],
                                                  name=name,
                                                  resolution_debug=data['resolution_debug'])
                    elif data['render_debug'] == 'final':
                        AnalyzeClass.render_final(image=image, mask=mask,
                                                  sensitivity_analysis=source[2],
                                                  correct_coefficient=source[3],
                                                  name=name,
                                                  resolution_debug=data['resolution_debug'])
                    elif data['render_debug'] == 'source':
                        AnalyzeClass.render_origin(image=image, name=name,
                                                   resolution_debug=data['resolution_debug'])
                    cv2.waitKey(round(1000 / data['speed_video_stream']))
                    values = AnalyzeClass.result_final(image=image, mask=mask,
                                                       sensitivity_analysis=source[2],
                                                       correct_coefficient=source[3])
                    AnalyzeClass.write_result(
                        ip_sql=data['ip_sql'],
                        server_sql=data['server_sql'],
                        port_sql=data['port_sql'],
                        database_sql=data['database_sql'],
                        user_sql=data['user_sql'],
                        password_sql=data['password_sql'],
                        sql_now_check=data['sql_now_check'],
                        table_now_sql=data['table_now_sql'],
                        rows_now_sql=data['rows_now_sql'],
                        sql_data_check=data['sql_data_check'],
                        table_data_sql=data['table_data_sql'],
                        rows_data_sql=data['rows_data_sql'],
                        values=values,
                        source=name,
                        widget=data['widget'],
                        widget_write=data['widget_write'],
                        text_write=data['text_write'],
                        sql_val=data['sql_write'],
                        alarm_level=source[5]
                    )
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'{TimeUtils.get_current_time()} : analyze func error')
                    print(ex)

            @staticmethod
            def get_full_sources(source_type: str, sources: list, masks: list, protocol: str, login: str, password: str,
                                 port: int, sensitivity: list, correct: list, alias_device: list, alarm_level: list):
                try:
                    _sources = []
                    for x in sources:
                        index = sources.index(x)
                        if source_type == 'image-http':
                            _sources.append([f'{protocol}://192.168.{x}:{port}/ISAPI/Streaming/channels/101/'
                                             f'picture?snapShotImageType=JPEG', cv2.imread(masks[index], 0),
                                             sensitivity[index], correct[index], alias_device[index],
                                             alarm_level[index]])
                        elif source_type == 'video-rtsp':
                            _sources.append(
                                [f'{protocol}://{login}:{password}@192.168.{x}:{port}', cv2.imread(masks[index], 0),
                                 sensitivity[index], correct[index], alias_device[index], alarm_level[index]])
                        elif source_type == 'video-file':
                            _sources.append([f'{x}', cv2.imread(masks[index], 0), sensitivity[index], correct[index],
                                             alias_device[index], alarm_level[index]])
                    return _sources
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'get_full_sources error : {ex}')

            @staticmethod
            def get_source(source_type: str, sources: str, login: str, password: str):
                try:
                    if source_type == 'image-http':
                        h = httplib2.Http("/path/to/cache-directory")
                        h.add_credentials(login, password)
                        response, content = h.request(sources)
                        image = cv2.imdecode(numpy.frombuffer(content, numpy.uint8), cv2.IMREAD_COLOR)
                        return image
                    elif source_type == 'video-rtsp' or 'video-file':
                        cam_stream = cv2.VideoCapture(sources)
                        _, image = cam_stream.read()
                        cam_stream.release()
                        return image
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'get_source func error')
                    print(ex)

            @staticmethod
            def render_origin(image, name, resolution_debug):
                AnalyzeClass.render(name=f"origin : {name}", source=image, resolution_debug=resolution_debug)

            @staticmethod
            def render_cropping_image(image, name, resolution_debug):
                cropping_image = image[250:1080, 600:1720]
                AnalyzeClass.render(name=f"cropping_image : {name}", source=cropping_image,
                                    resolution_debug=resolution_debug)

            @staticmethod
            def render_bitwise_not_white(image, mask, name, resolution_debug):
                bitwise_not = cv2.bitwise_not(image, image, mask=mask)
                AnalyzeClass.render(name=f"_bitwise_not_white : {name}", source=bitwise_not,
                                    resolution_debug=resolution_debug)

            @staticmethod
            def render_bitwise_and(image, mask, name, resolution_debug):
                bitwise_and = cv2.bitwise_and(image, image, mask=mask)
                AnalyzeClass.render(name=f"bitwise_and : {name}", source=bitwise_and, resolution_debug=resolution_debug)

            @staticmethod
            def render_threshold(image, name, resolution_debug):
                _, threshold = cv2.threshold(image, 220, 255, cv2.THRESH_BINARY_INV)
                AnalyzeClass.render(name=f"threshold : {name}", source=threshold, resolution_debug=resolution_debug)

            @staticmethod
            def render_cvtcolor_to_hsv(image, name, resolution_debug):
                cvtcolor = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
                AnalyzeClass.render(name=f"cvtcolor_to_hsv : {name}", source=cvtcolor,
                                    resolution_debug=resolution_debug)

            @staticmethod
            def render_inrange(image, sensitivity_analysis, name, resolution_debug):
                cvtcolor = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
                inrange = cv2.inRange(cvtcolor, numpy.array([0, 0, 255 - sensitivity_analysis], dtype=numpy.uint8),
                                      numpy.array([255, sensitivity_analysis, 255], dtype=numpy.uint8))
                AnalyzeClass.render(name=f"inrange : {name}", source=inrange, resolution_debug=resolution_debug)

            @staticmethod
            def render_canny_edges(image, sensitivity_analysis, name, resolution_debug):
                canny = cv2.Canny(image, sensitivity_analysis, sensitivity_analysis, apertureSize=3, L2gradient=True)
                AnalyzeClass.render(name=f"canny_edges : {name}", source=canny, resolution_debug=resolution_debug)

            @staticmethod
            def render_shapes(name, resolution_debug):
                red = (0, 0, 255)
                green = (0, 255, 0)
                blue = (255, 0, 0)
                yellow = (0, 255, 255)
                numpy.set_printoptions(threshold=0)
                img = numpy.zeros(shape=(512, 512, 3), dtype=numpy.uint8)
                cv2.line(
                    img=img,
                    pt1=(0, 0),
                    pt2=(311, 511),
                    color=blue,
                    thickness=10
                )
                cv2.rectangle(
                    img=img,
                    pt1=(30, 166),
                    pt2=(130, 266),
                    color=green,
                    thickness=3
                )
                cv2.circle(
                    img=img,
                    center=(222, 222),
                    radius=50,
                    color=(255.111, 111),
                    thickness=-1
                )
                cv2.ellipse(
                    img=img,
                    center=(333, 333),
                    axes=(50, 20),
                    angle=0,
                    startAngle=0,
                    endAngle=150,
                    color=red,
                    thickness=-1
                )
                pts = numpy.array(
                    [[10, 5], [20, 30], [70, 20], [50, 10]],
                    dtype=numpy.int32
                )
                pts = pts.reshape((-1, 1, 2,))
                cv2.polylines(
                    img=img,
                    pts=[pts],
                    isClosed=True,
                    color=yellow,
                    thickness=5
                )
                cv2.putText(
                    img=img,
                    text="SOL",
                    org=(10, 400),
                    fontFace=cv2.FONT_ITALIC,
                    fontScale=3.5,
                    color=(255, 255, 255),
                    thickness=2
                )
                AnalyzeClass.render(name=f"shapes : {name}", source=img, resolution_debug=resolution_debug)

            @staticmethod
            def render_cvtcolor(image, color_type, name, resolution_debug):
                cvtcolor = cv2.cvtColor(image, color_type)
                AnalyzeClass.render(name=f"cvtcolor : {name}", source=cvtcolor, resolution_debug=resolution_debug)

            @staticmethod
            def render_flip(image, flipcode, name, resolution_debug):
                flip = cv2.flip(image, flipcode)
                AnalyzeClass.render(name=f"flip : {name}", source=flip, resolution_debug=resolution_debug)

            @staticmethod
            def render_final(image, mask, sensitivity_analysis, correct_coefficient, name, resolution_debug):
                bitwise_and = cv2.bitwise_and(image, image, mask=mask)
                cvtcolor = cv2.cvtColor(bitwise_and, cv2.COLOR_BGR2HSV)
                inrange = cv2.inRange(cvtcolor, numpy.array([0, 0, 255 - sensitivity_analysis], dtype=numpy.uint8),
                                      numpy.array([255, sensitivity_analysis, 255], dtype=numpy.uint8))
                cv2.putText(inrange,
                            f"{numpy.sum(inrange > 0) / numpy.sum(mask > 0) * 100 * correct_coefficient:0.2f}%",
                            (int(1920 / 5), int(1080 / 2)), cv2.FONT_HERSHEY_SIMPLEX, 3, (255, 255, 255), 3)
                AnalyzeClass.render(name=f"final : {name}", source=inrange, resolution_debug=resolution_debug)

            @staticmethod
            def render(name: str, source, resolution_debug: list):
                try:
                    if source is not None:
                        img = cv2.resize(source, (resolution_debug[0], resolution_debug[1]),
                                         interpolation=cv2.INTER_AREA)
                        cv2.imshow(name, img)
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(ex)

            @staticmethod
            def result_final(image, mask, sensitivity_analysis, correct_coefficient):
                bitwise_and = cv2.bitwise_and(image, image, mask=mask)
                cvtcolor = cv2.cvtColor(bitwise_and, cv2.COLOR_BGR2HSV)
                inrange = cv2.inRange(cvtcolor, numpy.array([0, 0, 255 - sensitivity_analysis], dtype=numpy.uint8),
                                      numpy.array([255, sensitivity_analysis, 255], dtype=numpy.uint8))
                try:
                    return round(numpy.sum(inrange > 0) / numpy.sum(mask > 0) * 100 * correct_coefficient, 2)
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'result_final func error : {ex}')
                    return 0.0

            @staticmethod
            def write_result(ip_sql: str, server_sql: str, port_sql: str, database_sql: str, user_sql: str,
                             password_sql: str,
                             sql_now_check: bool, table_now_sql: str, rows_now_sql: list, sql_data_check: bool,
                             table_data_sql: str, rows_data_sql: list, source: str, values: float, widget,
                             widget_write: bool,
                             text_write: bool, sql_val: bool, alarm_level: int):
                sql_datetime = datetime.datetime.now().strftime('%Y-%m-%dT%H:%M:%S')
                boolean = 0
                if values > alarm_level:
                    boolean = 1
                _values = [source, values, boolean, sql_datetime]
                if widget_write:
                    try:
                        widget(f'{_values}')
                    except Exception as ex:
                        LoggingClass.logging(ex)
                        print(ex)
                if text_write:
                    try:
                        with open('db.txt', 'a') as db:
                            db.write(f'{_values}\n')
                    except Exception as ex:
                        LoggingClass.logging(ex)
                        print(ex)
                if sql_val:
                    try:
                        if sql_now_check:
                            SQLClass.sql_post_now(ip=ip_sql, server=server_sql, port=port_sql, database=database_sql,
                                                  username=user_sql, password=password_sql, table=table_now_sql,
                                                  rows=rows_now_sql, values=_values)
                        if sql_data_check:
                            SQLClass.sql_post_data(ip=ip_sql, server=server_sql, port=port_sql, database=database_sql,
                                                   username=user_sql, password=password_sql, table=table_data_sql,
                                                   rows=rows_data_sql, values=_values)
                    except Exception as ex:
                        LoggingClass.logging(ex)
                        print(ex)

            @staticmethod
            def make_snapshot(data: dict):
                h = httplib2.Http("/path/to/cache-directory")
                h.add_credentials(name=data['login_cam'], password=data['password_cam'])
                sources = f"{data['protocol_cam_type']}://192.168.{data['ip_cam_snapshot']}:{data['port_cam']}/" \
                          f"ISAPI/Streaming/channels/101/picture?snapShotImageType=JPEG"
                response, content = h.request(sources)
                with open(data["name_snapshot"], 'wb') as file:
                    file.write(content)
                # cv2.imwrite('1.png', img, [int(cv2.IMWRITE_JPEG_QUALITY), 95])
                # cv2.imwrite('1.png', img, [int(cv2.IMWRITE_PNG_COMPRESSION), 9])

            @staticmethod
            def play_rtsp():
                cap = cv2.VideoCapture('rtsp://admin:q1234567@192.168.15.229:554/cam/realmonitor?channel=1&subtype=0')
                # cap = cv2.VideoCapture('rtsp://admin:nehrfvths123@192.168.15.140:554')
                # cap = cv2.VideoCapture('rtsp://192.168.15.229:554/cam/realmonitor?channel=1&subtype=0&unicast=true&proto=Onvif')
                # cap = cv2.VideoCapture('rtsp://192.168.15.229:554/live')
                # cap = cv2.VideoCapture('rtsp://admin:q1234567@192.168.15.229:554/cam/realmonitor?channel=2&subtype=1')
                # cap = cv2.VideoCapture('rtsp://admin:q1234567@192.168.15.229:554/cam/realmonitor?channel=33&subtype=0')
                # sources = f"http://192.168.15.227/cgi-bin/snapshot.cgi?channel=road?loginuse=admin&loginpas=q1234567"
                # sources = f"http://192.168.15.227/cgi-bin/snapshot.cgi?loginuse=admin&loginpas=q1234567"
                # sources = f"http://192.168.15.227/cgi-bin/snapshot.cgi?chn=1&u=admin&p=q1234567"
                # sources = f"http://192.168.15.227/cgi-bin/snapshot.cgi?1"
                # sources = f"rtsp://192.168.15.227:554/cam/realmonitor?channel=1&subtype=0&unicast=true&proto=Onvif"
                # sources = f"rtsp://192.168.15.227:554/live"
                # sources = f"rtsp://admin:q1234567@192.168.15.227:554/cam/realmonitor?channel=1&subtype=1"
                # sources = f"rtsp://admin:q1234567@192.168.15.227:554/cam/realmonitor?channel=1&subtype=0"
                # sources = f"rtsp://admin:nehrfvths123@192.168.15.140:554"
                while True:
                    try:
                        ret, frame = cap.read()
                        cv2.imshow("Capturing", frame)
                        if cv2.waitKey(1) & 0xFF == ord('q'):
                            break
                    except Exception as ex:
                        print(ex)
                cv2.destroyAllWindows()
                cap.release()

        # UI
        class AppContainerClass:
            def __init__(self):
                self.app = QtWidgets.QApplication([])
                self.widget = None

            def create_ui(self, title, width, height, icon, play_f, stop_f, quit_f, snapshot_f):
                self.widget = MainWidgetClass(title, width, height, icon, play_f, stop_f, quit_f, snapshot_f)
                return self.widget

            @staticmethod
            def create_qlable(text: str, _parent, background=False):
                _widget = QtWidgets.QLabel(text)
                if background:
                    _widget.setAutoFillBackground(True)
                    _widget.setStyleSheet("QLabel { background-color: rgba(0, 0, 0, 255); color : white; }")
                _parent.addWidget(_widget)
                return _widget

            @staticmethod
            def create_qpushbutton(_parent, _connect_func, _text='set'):
                _widget = QtWidgets.QPushButton(_text)
                _parent.addWidget(_widget)
                _widget.clicked.connect(_connect_func)
                return _widget

            @staticmethod
            def create_qcheckbox(_parent, _text='check?', default=False):
                _widget = QtWidgets.QCheckBox(_text)
                _widget.setChecked(default)
                _parent.addWidget(_widget)
                return _widget

            @staticmethod
            def create_qcombobox(_parent, _text: list, default=None):
                _widget = QtWidgets.QComboBox()
                _widget.addItems([x for x in _text])
                _widget.setCurrentText(default)
                _parent.addWidget(_widget)
                return _widget

            @staticmethod
            def create_qradiobutton(_parent, _text: str, default=False):
                _widget = QtWidgets.QRadioButton(_text)
                _widget.setChecked(default)
                _parent.addWidget(_widget)
                return _widget

        class MainWidgetClass(QtWidgets.QWidget):
            def __init__(self, title="APP", width=640, height=480, icon="", play_f=None, stop_f=None, quit_f=None,
                         snapshot_f=None):
                super().__init__()

                self.play_f = play_f
                self.snapshot_f = snapshot_f
                self.resize(width, height)
                self.setWindowTitle(title)
                self.setWindowIcon(QtGui.QIcon(icon))
                self.resolution_debug = []
                self.v_layout_m = QtWidgets.QVBoxLayout(self)

                # MANAGEMENT
                self.h_layout_g_management = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_management)
                self.g_management_set = AppContainerClass.create_qlable('MANAGEMENT', self.h_layout_g_management,
                                                                        background=True)
                self.h_layout_management_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_management_1)
                self.play_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_management_1,
                                                                             self.play_btn_func,
                                                                             'play')
                self.stop_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_management_1, stop_f, 'stop')
                self.quit_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_management_1, quit_f, 'quit')
                # CAMERAS
                self.h_layout_g_cam = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_cam)
                self.g_cam_set = AppContainerClass.create_qlable('CAMERAS', self.h_layout_g_cam, background=True)
                self.h_layout_cam_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_1)
                self.protocol_cam_type = AppContainerClass.create_qlable('PROTOCOL TYPE : http', self.h_layout_cam_1)
                self.set_protocol_cam_type = AppContainerClass.create_qpushbutton(self.h_layout_cam_1,
                                                                                  self.get_protocol_cam_type_button)
                self.port_cam = AppContainerClass.create_qlable('PORT CAM : 80', self.h_layout_cam_1)
                self.set_port_cam = AppContainerClass.create_qpushbutton(self.h_layout_cam_1, self.get_port_cam_button)
                self.login_cam = AppContainerClass.create_qlable('LOGIN CAM : admin', self.h_layout_cam_1)
                self.set_login_cam = AppContainerClass.create_qpushbutton(self.h_layout_cam_1,
                                                                          self.get_login_cam_button)
                self.password_cam = AppContainerClass.create_qlable('PASSWORD CAM : q1234567', self.h_layout_cam_1)
                self.set_password_cam = AppContainerClass.create_qpushbutton(self.h_layout_cam_1,
                                                                             self.get_password_cam_button)
                self.h_layout_cam_2 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_2)
                self.alias_device = AppContainerClass.create_qlable(
                    'ALIAS DEVICE : 16/1 | 16/2 | 16/3 | 16/4 | 16/5 | 16/6 '
                    '| 16/7 | 16/8 | 16/9 | 16/10', self.h_layout_cam_2)
                self.set_alias_device = AppContainerClass.create_qpushbutton(self.h_layout_cam_2,
                                                                             self.get_alias_device_button)
                self.h_layout_cam_3 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_3)
                self.ip_cam = AppContainerClass.create_qlable(
                    'IP CAM : 15.202 | 15.206 | 15.207 | 15.208 | 15.209 | 15.210 '
                    '| 15.211 | 15.203 | 15.204 | 15.205', self.h_layout_cam_3)
                self.set_ip_cam = AppContainerClass.create_qpushbutton(self.h_layout_cam_3, self.get_ip_cam_button)
                self.h_layout_cam_4 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_4)
                self.mask_cam = AppContainerClass.create_qlable(
                    'MASK CAM : m_16_1.jpg | m_16_2.jpg | m_16_3.jpg | m_16_4.jpg '
                    '| m_16_5.jpg | m_16_6.jpg | m_16_7.jpg | m_16_8.jpg '
                    '| m_16_9.jpg | m_16_10.jpg', self.h_layout_cam_4)
                self.set_mask_cam = AppContainerClass.create_qpushbutton(self.h_layout_cam_4, self.get_mask_cam_button)
                self.h_layout_cam_5 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_5)
                self.sensitivity_analysis = AppContainerClass.create_qlable(
                    'SENSITIVITY ANALYSIS : 115 | 115 | 115 | 115 '
                    '| 115 | 115 | 115 | 115 | 115 | 115',
                    self.h_layout_cam_5)
                self.set_sensitivity_analysis = AppContainerClass.create_qpushbutton(self.h_layout_cam_5,
                                                                                     self.get_sensitivity_analysis_button)
                self.h_layout_cam_6 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_6)
                self.alarm_level = AppContainerClass.create_qlable('ALARM LEVEL : 30 | 30 | 30 | 30 '
                                                                   '| 30 | 30 | 30 | 30 | 30 | 30',
                                                                   self.h_layout_cam_6)
                self.set_alarm_level = AppContainerClass.create_qpushbutton(self.h_layout_cam_6,
                                                                            self.get_alarm_level_button)
                self.h_layout_cam_7 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_7)
                self.correct_coefficient = AppContainerClass.create_qlable(
                    'CORRECT COEFFICIENT : 1.0 | 1.0 | 1.0 | 1.0 | 1.0 '
                    '| 1.0 | 1.0 | 1.0 | 1.0 | 1.0', self.h_layout_cam_7)
                self.set_correct_coefficient = AppContainerClass.create_qpushbutton(self.h_layout_cam_7,
                                                                                    self.get_correct_coefficient_button)
                # SQL
                self.h_layout_g_sql = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_sql)
                self.g_sql_set = AppContainerClass.create_qlable('SQL', self.h_layout_g_sql, background=True)
                self.h_layout_sql_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_sql_1)
                self.sql_write = AppContainerClass.create_qcheckbox(self.h_layout_sql_1, 'CONNECT TO SQL?')
                self.ip_sql = AppContainerClass.create_qlable('IP SQL SERVER : 192.168.15.87', self.h_layout_sql_1)
                self.set_ip_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_1, self.get_ip_sql_button)
                self.server_sql = AppContainerClass.create_qlable(r'SERVER SQL : DESKTOP-SM7K050\COMPUTER_VISION',
                                                                  self.h_layout_sql_1)
                self.set_server_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_1,
                                                                           self.get_server_sql_button)
                self.port_sql = AppContainerClass.create_qlable('PORT SQL SERVER : 1433', self.h_layout_sql_1)
                self.set_port_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_1, self.get_port_sql_button)
                self.h_layout_sql_2 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_sql_2)
                self.database_sql = AppContainerClass.create_qlable('DATABASE SQL : analiz_16grohot',
                                                                    self.h_layout_sql_2)
                self.set_database_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_2,
                                                                             self.get_database_sql_button)
                self.user_sql = AppContainerClass.create_qlable('USER SQL : computer_vision', self.h_layout_sql_2)
                self.set_user_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_2, self.get_user_sql_button)
                self.h_layout_sql_3 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_sql_3)
                self.password_sql = AppContainerClass.create_qlable('PASSWORD SQL : vision12345678',
                                                                    self.h_layout_sql_2)
                self.set_password_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_2,
                                                                             self.get_password_sql_button)
                self.sql_now_check = AppContainerClass.create_qcheckbox(self.h_layout_sql_3, 'WRITE NOW SQL?')
                self.table_now_sql = AppContainerClass.create_qlable('TABLE NOW SQL : grohot16_now_table',
                                                                     self.h_layout_sql_3)
                self.set_table_now_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_3,
                                                                              self.get_table_now_sql_button)
                self.rows_now_sql = AppContainerClass.create_qlable('ROWS NOW SQL : device_row | value_row | alarm_row '
                                                                    '| datetime_row', self.h_layout_sql_3)
                self.set_rows_now_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_3,
                                                                             self.get_rows_now_sql_button)
                self.h_layout_sql_4 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_sql_4)
                self.sql_data_check = AppContainerClass.create_qcheckbox(self.h_layout_sql_4, 'WRITE DATA SQL?')
                self.table_data_sql = AppContainerClass.create_qlable('TABLE DATA SQL : grohot16_data_table',
                                                                      self.h_layout_sql_4)
                self.set_table_data_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_4,
                                                                               self.get_table_data_sql_button)
                self.rows_data_sql = AppContainerClass.create_qlable(
                    'ROWS DATA SQL : device_row | value_row | alarm_row '
                    '| datetime_row', self.h_layout_sql_4)
                self.set_rows_data_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_4,
                                                                              self.get_rows_data_sql_button)
                # DEBUG
                self.h_layout_g_debug = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_debug)
                self.g_debug_set = AppContainerClass.create_qlable('DEBUG', self.h_layout_g_debug, background=True)
                self.h_layout_debug_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_debug_1)
                self.auto_import_check = AppContainerClass.create_qcheckbox(self.h_layout_debug_1, 'AUTO IMPORT?')
                self.auto_play_check = AppContainerClass.create_qcheckbox(self.h_layout_debug_1, 'AUTO PLAY?')
                self.speed_analysis = AppContainerClass.create_qlable('SPEED ANALYSIS : 1.0', self.h_layout_debug_1)
                self.set_speed_analysis = AppContainerClass.create_qpushbutton(self.h_layout_debug_1,
                                                                               self.get_speed_analysis_button)
                self.speed_video_stream = AppContainerClass.create_qlable('SPEED VIDEO-STREAM : 1.0',
                                                                          self.h_layout_debug_1)
                self.set_speed_video_stream = AppContainerClass.create_qpushbutton(self.h_layout_debug_1,
                                                                                   self.get_speed_video_stream_button)
                self.h_layout_debug_2 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_debug_2)
                self.widget_data_value = AppContainerClass.create_qlable('0.00%', self.h_layout_debug_2)
                self.h_layout_debug_2.addStretch()
                self.widget_write = AppContainerClass.create_qcheckbox(self.h_layout_debug_2, 'WRITE TO WIDGET?')
                self.text_write = AppContainerClass.create_qcheckbox(self.h_layout_debug_2, 'WRITE TO TEXT?')
                self.source_win_type = AppContainerClass.create_qlable('SOURCE TYPE :', self.h_layout_debug_2)
                self.source_type = AppContainerClass.create_qcombobox(self.h_layout_debug_2,
                                                                      ['image-http', 'video-rtsp', 'video-file'],
                                                                      'image-http')
                self.compute_win_debug = AppContainerClass.create_qlable('COMPUTE TYPE :', self.h_layout_debug_2)
                self.compute_debug = AppContainerClass.create_qcombobox(self.h_layout_debug_2,
                                                                        ['sync', 'async', 'multithread',
                                                                         'multiprocess', 'complex'],
                                                                        'complex')
                self.process_cores = AppContainerClass.create_qlable('PROCESS CORES : 4', self.h_layout_debug_2)
                self.set_process_cores = AppContainerClass.create_qpushbutton(self.h_layout_debug_2,
                                                                              self.get_process_cores_button)
                self.h_layout_debug_3 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_debug_3)
                self.render_win_debug = AppContainerClass.create_qlable('Render windows :', self.h_layout_debug_3)
                self.render_debug = AppContainerClass.create_qcombobox(self.h_layout_debug_3,
                                                                       ['none', 'source', 'final', 'extended',
                                                                        'all'], 'none')
                self.resolution_debug_1 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '320x240',
                                                                                default=True)
                self.resolution_debug_1.toggled.connect(self.set_resolution_debug(self.resolution_debug_1, 320, 240))
                self.resolution_debug_2 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '640x480')
                self.resolution_debug_2.toggled.connect(self.set_resolution_debug(self.resolution_debug_2, 640, 480))
                self.resolution_debug_3 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '1280x720')
                self.resolution_debug_3.toggled.connect(self.set_resolution_debug(self.resolution_debug_3, 1280, 720))
                self.resolution_debug_4 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '1920x1080')
                self.resolution_debug_4.toggled.connect(self.set_resolution_debug(self.resolution_debug_4, 1920, 1080))
                self.resolution_debug_5 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '2560x1600')
                self.resolution_debug_5.toggled.connect(self.set_resolution_debug(self.resolution_debug_5, 2560, 1600))
                self.resolution_debug_6 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '3840x2160')
                self.resolution_debug_6.toggled.connect(self.set_resolution_debug(self.resolution_debug_6, 3840, 2160))
                # IMPORTS
                self.h_layout_g_imports = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_imports)
                self.g_imports_set = AppContainerClass.create_qlable('IMPORTS', self.h_layout_g_imports,
                                                                     background=True)
                self.h_layout_imports_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_imports_1)
                self.import_file = AppContainerClass.create_qlable('SETTINGS FILE NAME : settings',
                                                                   self.h_layout_imports_1)
                self.set_import_file = AppContainerClass.create_qpushbutton(self.h_layout_imports_1,
                                                                            self.get_settings_file_name)
                self.export_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_imports_1,
                                                                               self.export_settings_func, 'export')
                self.import_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_imports_1,
                                                                               self.import_settings_func, 'import')
                # SHOT
                self.h_layout_g_shot = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_shot)
                self.g_shot_set = AppContainerClass.create_qlable('SHOT', self.h_layout_g_shot, background=True)
                self.h_layout_shot_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_shot_1)
                self.ip_cam_snapshot = AppContainerClass.create_qlable('ip-cam : 15.204', self.h_layout_shot_1)
                self.set_ip_cam_snapshot = AppContainerClass.create_qpushbutton(self.h_layout_shot_1,
                                                                                self.get_ip_cam_snapshot_button)
                self.name_snapshot = AppContainerClass.create_qlable('file name : picture.jpg', self.h_layout_shot_1)
                self.set_name_snapshot = AppContainerClass.create_qpushbutton(self.h_layout_shot_1,
                                                                              self.set_name_snapshot_button)
                self.snapshot_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_shot_1,
                                                                                 self.snapshot_btn_func,
                                                                                 'snapshot')

                self.setLayout(self.v_layout_m)
                self.auto_play_func()
                self.auto_import_settings_func()

            def get_speed_analysis_button(self):
                widget = self.speed_analysis
                value, success = QtWidgets.QInputDialog.getDouble(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                  f'{widget.text().split(":")[0].strip()} value:',
                                                                  1.0, 0.01, 50.0, 2)
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_speed_video_stream_button(self):
                widget = self.speed_video_stream
                value, success = QtWidgets.QInputDialog.getDouble(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                  f'{widget.text().split(":")[0].strip()} value:',
                                                                  1.0, 0.01, 50.0, 2)
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_sensitivity_analysis_button(self):
                widget = self.sensitivity_analysis
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_alarm_level_button(self):
                widget = self.alarm_level
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_correct_coefficient_button(self):
                widget = self.correct_coefficient
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_protocol_cam_type_button(self):
                widget = self.protocol_cam_type
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_port_cam_button(self):
                widget = self.port_cam
                value, success = QtWidgets.QInputDialog.getInt(self, f'Set {widget.text().split(":")[0].strip()}',
                                                               f'{widget.text().split(":")[0].strip()} value:',
                                                               554, 1, 9999, 5)
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_login_cam_button(self):
                widget = self.login_cam
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_password_cam_button(self):
                widget = self.password_cam
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_alias_device_button(self):
                widget = self.alias_device
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_ip_cam_button(self):
                widget = self.ip_cam
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_mask_cam_button(self):
                widget = self.mask_cam
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_ip_sql_button(self):
                widget = self.ip_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_server_sql_button(self):
                widget = self.server_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_port_sql_button(self):
                widget = self.port_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_database_sql_button(self):
                widget = self.database_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_user_sql_button(self):
                widget = self.user_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_password_sql_button(self):
                widget = self.password_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_table_now_sql_button(self):
                widget = self.table_now_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_rows_now_sql_button(self):
                widget = self.rows_now_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_table_data_sql_button(self):
                widget = self.table_data_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_rows_data_sql_button(self):
                widget = self.rows_data_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_settings_file_name(self):
                widget = self.import_file
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def set_resolution_debug(self, radio, width: int, height: int):
                self.resolution_debug.append([radio, width, height])

            def get_window_resolution(self):
                for radio in self.resolution_debug:
                    try:
                        if radio[0].isChecked():
                            return [int(radio[1]), int(radio[2])]
                    except Exception as ex:
                        print(ex)

            def set_window_resolution(self, value):
                for radio in self.resolution_debug:
                    try:
                        if radio[1] == value[0]:
                            radio[0].setChecked(True)
                        else:
                            radio[0].setChecked(False)
                    except Exception as ex:
                        print(ex)

            def get_process_cores_button(self):
                widget = self.process_cores
                value, success = QtWidgets.QInputDialog.getInt(self, f'Set {widget.text().split(":")[0].strip()}',
                                                               f'{widget.text().split(":")[0].strip()} value:',
                                                               4, 1, 16, 5)
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_ip_cam_snapshot_button(self):
                widget = self.ip_cam_snapshot
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def set_name_snapshot_button(self):
                widget = self.name_snapshot
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_string_from_list(self, source: list):
                value = ''
                for x in source:
                    value = f'{value} | {x}'
                return value[3::]

            def set_data_func(self, value: str):
                try:
                    self.widget_data_value.setText(f"{value}")
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'set_data_func error : {ex}')

            def create_data_func(self):
                try:
                    data = {
                        'protocol_cam_type': str(self.protocol_cam_type.text().split(':')[1].strip()),
                        'port_cam': int(self.port_cam.text().split(':')[1].strip()),
                        'login_cam': str(self.login_cam.text().split(':')[1].strip()),
                        'password_cam': str(self.password_cam.text().split(':')[1].strip()),
                        'alias_device': list(
                            [x.strip() for x in self.alias_device.text().split(':')[1].strip().split('|')]),
                        'ip_cam': list([x.strip() for x in self.ip_cam.text().split(':')[1].strip().split('|')]),
                        'mask_cam': list([x.strip() for x in self.mask_cam.text().split(':')[1].strip().split('|')]),
                        'sensitivity_analysis': list([int(x.strip()) for x in
                                                      self.sensitivity_analysis.text().split(':')[1].strip().split(
                                                          '|')]),
                        'alarm_level': list(
                            [int(x.strip()) for x in self.alarm_level.text().split(':')[1].strip().split('|')]),
                        'correct_coefficient': list([float(x.strip()) for x in
                                                     self.correct_coefficient.text().split(':')[1].strip().split('|')]),

                        'sql_write': bool(self.sql_write.isChecked()),
                        'ip_sql': str(self.ip_sql.text().split(':')[1].strip()),
                        'server_sql': str(self.server_sql.text().split(':')[1].strip()),
                        'port_sql': str(self.port_sql.text().split(':')[1].strip()),
                        'database_sql': str(self.database_sql.text().split(':')[1].strip()),
                        'user_sql': str(self.user_sql.text().split(':')[1].strip()),
                        'password_sql': str(self.password_sql.text().split(':')[1].strip()),
                        'sql_now_check': bool(self.sql_now_check.isChecked()),
                        'table_now_sql': str(self.table_now_sql.text().split(':')[1].strip()),
                        'rows_now_sql': list(
                            [x.strip() for x in self.rows_now_sql.text().split(':')[1].strip().split('|')]),
                        'sql_data_check': bool(self.sql_data_check.isChecked()),
                        'table_data_sql': str(self.table_data_sql.text().split(':')[1].strip()),
                        'rows_data_sql': list(
                            [x.strip() for x in self.rows_data_sql.text().split(':')[1].strip().split('|')]),

                        'auto_import_check': bool(self.auto_import_check.isChecked()),
                        'auto_play_check': bool(self.auto_play_check.isChecked()),
                        'speed_analysis': float(self.speed_analysis.text().split(':')[1].strip()),
                        'speed_video_stream': float(self.speed_video_stream.text().split(':')[1].strip()),

                        'widget_write': bool(self.widget_write.isChecked()),
                        'text_write': bool(self.text_write.isChecked()),
                        'widget': self.set_data_func,
                        'source_type': str(self.source_type.currentText().strip()),
                        'compute_debug': str(self.compute_debug.currentText().strip()),
                        'process_cores': int(self.process_cores.text().split(':')[1].strip()),
                        'render_debug': str(self.render_debug.currentText().strip()),
                        'resolution_debug': list(self.get_window_resolution()),

                        'import_file': str(self.import_file.text().split(':')[1].strip()),

                        'ip_cam_snapshot': str(self.ip_cam_snapshot.text().split(":")[1].strip()),
                        'name_snapshot': str(self.name_snapshot.text().split(":")[1].strip()),
                    }
                    return data
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'create_data_func error : {ex}')

            def play_btn_func(self):
                try:
                    data = self.create_data_func()
                    self.play_f(data=data)
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'play_btn_func error : {ex}')

            def snapshot_btn_func(self):
                try:
                    data = self.create_data_func()
                    self.snapshot_f(data=data)
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'snapshot_btn_func error : {ex}')

            def export_settings_func(self):
                try:
                    data = self.create_data_func()
                    del data['widget']
                    FileSettings.export_settings(data)
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'export_settings_func error : {ex}')

            def import_settings_func(self):
                try:
                    data = self.create_data_func()
                    data = FileSettings.import_settings(data)
                    self.protocol_cam_type.setText(f'{self.protocol_cam_type.text().split(":")[0].strip()} : '
                                                   f'{str(data["protocol_cam_type"])}')
                    self.port_cam.setText(f'{self.port_cam.text().split(":")[0].strip()} : '
                                          f'{str(data["port_cam"])}')
                    self.login_cam.setText(f'{self.login_cam.text().split(":")[0].strip()} : '
                                           f'{str(data["login_cam"])}')
                    self.password_cam.setText(f'{self.password_cam.text().split(":")[0].strip()} : '
                                              f'{str(data["password_cam"])}')
                    self.alias_device.setText(f'{self.alias_device.text().split(":")[0].strip()} : '
                                              f'{self.get_string_from_list(data["alias_device"])}')
                    self.ip_cam.setText(f'{self.ip_cam.text().split(":")[0].strip()} : '
                                        f'{self.get_string_from_list(data["ip_cam"])}')
                    self.mask_cam.setText(f'{self.mask_cam.text().split(":")[0].strip()} : '
                                          f'{self.get_string_from_list(data["mask_cam"])}')
                    self.sensitivity_analysis.setText(f'{self.sensitivity_analysis.text().split(":")[0].strip()} : '
                                                      f'{self.get_string_from_list(data["sensitivity_analysis"])}')
                    self.alarm_level.setText(f'{self.alarm_level.text().split(":")[0].strip()} : '
                                             f'{self.get_string_from_list(data["alarm_level"])}')
                    self.correct_coefficient.setText(f'{self.correct_coefficient.text().split(":")[0].strip()} : '
                                                     f'{self.get_string_from_list(data["correct_coefficient"])}')
                    self.sql_write.setChecked(data["sql_write"])
                    self.ip_sql.setText(f'{self.ip_sql.text().split(":")[0].strip()} : {str(data["ip_sql"])}')
                    self.server_sql.setText(
                        f'{self.server_sql.text().split(":")[0].strip()} : {str(data["server_sql"])}')
                    self.port_sql.setText(f'{self.port_sql.text().split(":")[0].strip()} : {str(data["port_sql"])}')
                    self.database_sql.setText(f'{self.database_sql.text().split(":")[0].strip()} : '
                                              f'{str(data["database_sql"])}')
                    self.user_sql.setText(f'{self.user_sql.text().split(":")[0].strip()} : '
                                          f'{str(data["user_sql"])}')
                    self.password_sql.setText(f'{self.password_sql.text().split(":")[0].strip()} : '
                                              f'{str(data["password_sql"])}')
                    self.sql_now_check.setChecked(data["sql_now_check"])
                    self.table_now_sql.setText(f'{self.table_now_sql.text().split(":")[0].strip()} : '
                                               f'{str(data["table_now_sql"])}')
                    self.rows_now_sql.setText(f'{self.rows_now_sql.text().split(":")[0].strip()} : '
                                              f'{self.get_string_from_list(data["rows_now_sql"])}')
                    self.sql_data_check.setChecked(data["sql_data_check"])
                    self.table_data_sql.setText(f'{self.table_data_sql.text().split(":")[0].strip()} : '
                                                f'{str(data["table_data_sql"])}')
                    self.rows_data_sql.setText(f'{self.rows_data_sql.text().split(":")[0].strip()} : '
                                               f'{self.get_string_from_list(data["rows_data_sql"])}')
                    self.auto_import_check.setChecked(data["auto_import_check"])
                    self.auto_play_check.setChecked(data["auto_play_check"])
                    self.speed_analysis.setText(f'{self.speed_analysis.text().split(":")[0].strip()} : '
                                                f'{str(data["speed_analysis"])}')
                    self.speed_video_stream.setText(f'{self.speed_video_stream.text().split(":")[0].strip()} : '
                                                    f'{str(data["speed_video_stream"])}')
                    self.widget_write.setChecked(data["widget_write"])
                    self.text_write.setChecked(data["text_write"])
                    self.source_type.setCurrentText(data["source_type"])
                    self.compute_debug.setCurrentText(data["compute_debug"])
                    self.process_cores.setText(f'{self.process_cores.text().split(":")[0].strip()} : '
                                               f'{str(data["process_cores"])}')
                    self.render_debug.setCurrentText(data["render_debug"])
                    self.set_window_resolution(data["resolution_debug"])
                    self.import_file.setText(f'{self.import_file.text().split(":")[0].strip()} : '
                                             f'{str(data["import_file"])}')
                    self.ip_cam_snapshot.setText(f'{self.ip_cam_snapshot.text().split(":")[0].strip()} : '
                                                 f'{str(data["ip_cam_snapshot"])}')
                    self.name_snapshot.setText(f'{self.name_snapshot.text().split(":")[0].strip()} : '
                                               f'{str(data["name_snapshot"])}')
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'import_settings_func error : {ex}')

            def auto_import_settings_func(self):
                try:
                    data = self.create_data_func()
                    data = FileSettings.import_settings(data)
                    if data['auto_import_check']:
                        self.import_settings_func()
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'auto_import_settings_func error : {ex}')

            def auto_play_func(self):
                try:
                    data = self.create_data_func()
                    _data = FileSettings.import_settings(data)
                    _data['widget'] = data['widget']
                    if _data['auto_play_check']:
                        self.play_f(data=_data)
                        self.showMinimized()
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'auto_play_func error : {ex}')

        def play_func(data: dict):
            global play
            try:
                play = True

                def play_analyze():
                    AnalyzeClass.start_analyze(data=CopyDictionary.get_all_sources(data, {'pause': pause}))

                threading.Thread(target=play_analyze, args=()).start()
            except Exception as ex:
                print(ex)
                with open('log.txt', 'a') as log:
                    log.write(f'\n{ex}\n')

        def stop_func():
            global play
            play = False

        def pause():
            global play
            return play

        def quit_func():
            stop_func()
            sys.exit(app_container.app.exec())

        def snapshot_func(data: dict):
            threading.Thread(target=AnalyzeClass.make_snapshot, args=(data,)).start()

        # MAIN
        if __name__ == "__main__":
            freeze_support()
            play = True
            app_container = AppContainerClass()
            widget = app_container.create_ui(title="analysis", width=300, height=300, icon="icon.ico",
                                             play_f=play_func, stop_f=stop_func, quit_f=quit_func,
                                             snapshot_f=snapshot_func)
            ui_thread = threading.Thread(target=widget.show())
            sys.exit(app_container.app.exec())


class TkinterClass:
    @staticmethod
    def example():
        def play_func(data: str):
            global play
            play = False
            sleep(0.5)
            play = True

            def whiles():
                print(data)

            thread_render = Thread(target=whiles)
            thread_render.start()
            app.set_title("Завершено")

        class Application(tkinter.Frame):
            def __init__(self, root, **kw):
                super().__init__(**kw)
                self.id = 0
                self.iid = 0
                self.root = root
                self.root.title("ожидание")
                self.root.grid_rowconfigure(0, weight=1)
                self.root.grid_columnconfigure(0, weight=1)
                self.root.config(background="black")
                self.root.geometry('1280x720')
                self.master.minsize(1280, 720)
                self.master.maxsize(1280, 720)

                self.play_btn = tkinter.Button(self.root, text="Запустить", font="100", command=self.play_button)
                self.play_btn.grid(row=0, column=0, sticky=tkinter.W)
                self.stop_btn = tkinter.Button(self.root, text="Остановить", font="100", command=self.stop_button)
                self.stop_btn.grid(row=0, column=1, sticky=tkinter.W)
                self.quit_btn = tkinter.Button(self.root, text="Выход", font="100", command=self.quit_button)
                self.quit_btn.grid(row=0, column=2, sticky=tkinter.W)

                self.export_label = tkinter.Label(self.root, text="Видеофайл для анализа/ip для анализа", font="100")
                self.export_label.grid(row=1, column=0, sticky=tkinter.W)
                self.export_entry = tkinter.Entry(self.root, font="100")
                self.export_entry.grid(row=1, column=1, sticky=tkinter.W)
                self.export_entry.insert(0, 'video.mp4')

                chk_state = tk.BooleanVar()
                chk_state.set(False)
                chk = ttk.Checkbutton(self.root, text='Выбрать', var=chk_state)
                chk_state.set(False)
                chk.grid(row=2, column=1)

                self.combo = ttk.Combobox(self.root)
                self.combo['values'] = (1, 2, 3, 4, 5, "Text")
                self.combo.current(1)
                self.combo.grid(row=3, column=1, sticky=tkinter.W)

                self.text = tkinter.Text(self.root, font="100")
                self.text.grid(row=4, column=0, sticky=tkinter.W)

            def play_button(self):
                self.set_title("в процессе")
                play_func(data="старт")

            def stop_button(self):
                self.set_title("пауза")
                global play
                play = False

            def quit_button(self):
                self.set_title("выход")
                global play
                play = False
                self.quit()

            def set_title(self, title: str):
                self.root.title(title)

        if __name__ == "__main__":
            play = False
            app = Application(tk.Tk())
            thread_main = Thread(target=app.root.mainloop())
            thread_main.start()

    @staticmethod
    def example_generate_list():
        class Application(tk.Frame):
            def __init__(self, root, **kw):
                super().__init__(**kw)
                self.root = root
                self.initialize_user_interface()

            def initialize_user_interface(self):
                self.root.title("Попытка в приложение")
                self.root.grid_rowconfigure(0, weight=1)
                self.root.grid_columnconfigure(0, weight=1)
                self.root.config(background="black")
                self.root.geometry('1280x720')

                # Define the different GUI widgets
                self.SurnameNumber_label = tk.Label(self.root, text="Фамилия:", font="100")
                self.SurnameNumber_entry = tk.Entry(self.root, font="100")
                self.SurnameNumber_label.grid(row=0, column=0, sticky=tk.W)
                self.SurnameNumber_entry.grid(row=0, column=0)

                self.NameName_label = tk.Label(self.root, text="Имя:", font="100")
                self.NameName_entry = tk.Entry(self.root, font="100")
                self.NameName_label.grid(row=1, column=0, sticky=tk.W)
                self.NameName_entry.grid(row=1, column=0)

                self.PatronymicName_label = tk.Label(self.root, text="Отчество:", font="100")
                self.PatronymicName_entry = tk.Entry(self.root, font="100")
                self.PatronymicName_label.grid(row=2, column=0, sticky=tk.W)
                self.PatronymicName_entry.grid(row=2, column=0)

                self.Extra_label = tk.Label(self.root, text="Дополнительно:", font="100")
                self.Extra_entry = tk.Entry(self.root, font="100")
                self.Extra_label.grid(row=3, column=0, sticky=tk.W)
                self.Extra_entry.grid(row=3, column=0)

                self.submit_button = tk.Button(self.root, text="Добавить", font="100", command=self.insert_data)
                self.submit_button.grid(row=2, column=1, sticky=tk.W)

                self.exit_button = tk.Button(self.root, text="Выход", font="100", command=self.quit)
                self.exit_button.grid(row=0, column=1, sticky=tk.W)

                # Set the treeview
                self.tree = ttk.Treeview(self.root, columns=('№', 'Фамилия:', 'Имя:', 'Отчество:', 'Дополнительно:'))

                # Set the heading (Attribute Names)
                self.tree.heading('#0', text='№')
                self.tree.heading('#1', text='Фамилия')
                self.tree.heading('#2', text='Имя')
                self.tree.heading('#3', text='Отчество')
                self.tree.heading('#4', text='Дополнительно')

                # Specify attributes of the columns (We want to stretch it!)
                self.tree.column('#0', stretch=tk.YES)
                self.tree.column('#1', stretch=tk.YES)
                self.tree.column('#2', stretch=tk.YES)
                self.tree.column('#3', stretch=tk.YES)
                self.tree.column('#4', stretch=tk.YES)

                self.tree.grid(row=3, columnspan=4, sticky='nsew')
                self.treeview = self.tree

                self.id = 0
                self.iid = 0

            def insert_data(self):
                self.treeview.insert('', 'end', iid=self.iid, text=str(self.id + 1),
                                     values=(self.SurnameNumber_entry.get(), self.NameName_entry.get(),
                                             self.PatronymicName_entry.get(),
                                             str(self.SurnameNumber_entry.get() + self.NameName_entry.get() +
                                                 self.PatronymicName_entry.get())))
                self.iid = self.iid + 1
                self.id = self.id + 1

        app = Application(tk.Tk())
        thread_main = Thread(target=app.root.mainloop())
        thread_main.start()


class PySideClass:
    @staticmethod
    def example():
        def play_func(data):
            global play
            play = False
            sleep(0.5)
            play = True

            def whiles():
                print(data)

            thread_render = Thread(target=whiles)
            thread_render.start()
            widget.set_text_func('завершено')

        class MyWidget(QtWidgets.QWidget):
            def __init__(self, title="ожидание"):
                super().__init__()
                self.play_button = QtWidgets.QPushButton("play")
                self.temp_box = QtWidgets.QDoubleSpinBox()
                self.stop_button = QtWidgets.QPushButton("stop")
                self.quit_button = QtWidgets.QPushButton("quit")
                self.temp_box.setValue(36.6)
                self.setWindowTitle(title)

                self.ui_window = QtWidgets.QHBoxLayout(self)
                self.ui_window.addWidget(self.play_button)
                self.ui_window.addWidget(self.temp_box)
                self.ui_window.addWidget(self.stop_button)
                self.ui_window.addWidget(self.quit_button)

                self.play_button.clicked.connect(self.play_btn_func)
                self.stop_button.clicked.connect(self.stop_btn_func)
                self.quit_button.clicked.connect(self.quit_btn_func)

            def play_btn_func(self):
                self.set_text_func("в процессе")
                play_func(data=self.temp_box.value())

            def stop_btn_func(self):
                self.set_text_func("пауза")
                global play
                play = False

            def quit_btn_func(self):
                self.set_text_func("выйти")
                global play
                play = False
                global app
                sys.exit(app.exec())

            def set_text_func(self, text: str):
                self.setWindowTitle(text)

        if __name__ == "__main__":
            play = False
            app = QtWidgets.QApplication([])
            widget = MyWidget()
            widget.resize(640, 480)
            thread_main = Thread(target=widget.show())
            thread_main.start()
            sys.exit(app.exec())

    @staticmethod
    def example_cv():
        class AppContainerClass:
            def __init__(self):
                self.app = QtWidgets.QApplication([])
                self.widget = None

            def create_ui(self, title, width, height, icon, play_f, stop_f, quit_f, snapshot_f):
                self.widget = MainWidgetClass(title, width, height, icon, play_f, stop_f, quit_f, snapshot_f)
                return self.widget

            @staticmethod
            def create_qlable(text: str, _parent, background=False):
                _widget = QtWidgets.QLabel(text)
                if background:
                    _widget.setAutoFillBackground(True)
                    _widget.setStyleSheet("QLabel { background-color: rgba(0, 0, 0, 255); color : white; }")
                _parent.addWidget(_widget)
                return _widget

            @staticmethod
            def create_qpushbutton(_parent, _connect_func, _text='set'):
                _widget = QtWidgets.QPushButton(_text)
                _parent.addWidget(_widget)
                _widget.clicked.connect(_connect_func)
                return _widget

            @staticmethod
            def create_qcheckbox(_parent, _text='check?', default=False):
                _widget = QtWidgets.QCheckBox(_text)
                _widget.setChecked(default)
                _parent.addWidget(_widget)
                return _widget

            @staticmethod
            def create_qcombobox(_parent, _text: list, default=None):
                _widget = QtWidgets.QComboBox()
                _widget.addItems([x for x in _text])
                _widget.setCurrentText(default)
                _parent.addWidget(_widget)
                return _widget

            @staticmethod
            def create_qradiobutton(_parent, _text: str, default=False):
                _widget = QtWidgets.QRadioButton(_text)
                _widget.setChecked(default)
                _parent.addWidget(_widget)
                return _widget

        class MainWidgetClass(QtWidgets.QWidget):
            def __init__(self, title="APP", width=640, height=480, icon="", play_f=None, stop_f=None, quit_f=None,
                         snapshot_f=None):
                super().__init__()

                self.play_f = play_f
                self.snapshot_f = snapshot_f
                self.resize(width, height)
                self.setWindowTitle(title)
                self.setWindowIcon(QtGui.QIcon(icon))
                self.resolution_debug = []
                self.v_layout_m = QtWidgets.QVBoxLayout(self)

                # MANAGEMENT
                self.h_layout_g_management = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_management)
                self.g_management_set = AppContainerClass.create_qlable('MANAGEMENT', self.h_layout_g_management,
                                                                        background=True)
                self.h_layout_management_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_management_1)
                self.play_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_management_1,
                                                                             self.play_btn_func,
                                                                             'play')
                self.stop_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_management_1, stop_f, 'stop')
                self.quit_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_management_1, quit_f, 'quit')
                # CAMERAS
                self.h_layout_g_cam = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_cam)
                self.g_cam_set = AppContainerClass.create_qlable('CAMERAS', self.h_layout_g_cam, background=True)
                self.h_layout_cam_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_1)
                self.protocol_cam_type = AppContainerClass.create_qlable('PROTOCOL TYPE : http', self.h_layout_cam_1)
                self.set_protocol_cam_type = AppContainerClass.create_qpushbutton(self.h_layout_cam_1,
                                                                                  self.get_protocol_cam_type_button)
                self.port_cam = AppContainerClass.create_qlable('PORT CAM : 80', self.h_layout_cam_1)
                self.set_port_cam = AppContainerClass.create_qpushbutton(self.h_layout_cam_1, self.get_port_cam_button)
                self.login_cam = AppContainerClass.create_qlable('LOGIN CAM : admin', self.h_layout_cam_1)
                self.set_login_cam = AppContainerClass.create_qpushbutton(self.h_layout_cam_1,
                                                                          self.get_login_cam_button)
                self.password_cam = AppContainerClass.create_qlable('PASSWORD CAM : q1234567', self.h_layout_cam_1)
                self.set_password_cam = AppContainerClass.create_qpushbutton(self.h_layout_cam_1,
                                                                             self.get_password_cam_button)
                self.h_layout_cam_2 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_2)
                self.alias_device = AppContainerClass.create_qlable(
                    'ALIAS DEVICE : 16/1 | 16/2 | 16/3 | 16/4 | 16/5 | 16/6 '
                    '| 16/7 | 16/8 | 16/9 | 16/10', self.h_layout_cam_2)
                self.set_alias_device = AppContainerClass.create_qpushbutton(self.h_layout_cam_2,
                                                                             self.get_alias_device_button)
                self.h_layout_cam_3 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_3)
                self.ip_cam = AppContainerClass.create_qlable(
                    'IP CAM : 15.202 | 15.206 | 15.207 | 15.208 | 15.209 | 15.210 '
                    '| 15.211 | 15.203 | 15.204 | 15.205', self.h_layout_cam_3)
                self.set_ip_cam = AppContainerClass.create_qpushbutton(self.h_layout_cam_3, self.get_ip_cam_button)
                self.h_layout_cam_4 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_4)
                self.mask_cam = AppContainerClass.create_qlable(
                    'MASK CAM : m_16_1.jpg | m_16_2.jpg | m_16_3.jpg | m_16_4.jpg '
                    '| m_16_5.jpg | m_16_6.jpg | m_16_7.jpg | m_16_8.jpg '
                    '| m_16_9.jpg | m_16_10.jpg', self.h_layout_cam_4)
                self.set_mask_cam = AppContainerClass.create_qpushbutton(self.h_layout_cam_4, self.get_mask_cam_button)
                self.h_layout_cam_5 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_5)
                self.sensitivity_analysis = AppContainerClass.create_qlable(
                    'SENSITIVITY ANALYSIS : 115 | 115 | 115 | 115 '
                    '| 115 | 115 | 115 | 115 | 115 | 115',
                    self.h_layout_cam_5)
                self.set_sensitivity_analysis = AppContainerClass.create_qpushbutton(self.h_layout_cam_5,
                                                                                     self.get_sensitivity_analysis_button)
                self.h_layout_cam_6 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_6)
                self.alarm_level = AppContainerClass.create_qlable('ALARM LEVEL : 30 | 30 | 30 | 30 '
                                                                   '| 30 | 30 | 30 | 30 | 30 | 30',
                                                                   self.h_layout_cam_6)
                self.set_alarm_level = AppContainerClass.create_qpushbutton(self.h_layout_cam_6,
                                                                            self.get_alarm_level_button)
                self.h_layout_cam_7 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_7)
                self.correct_coefficient = AppContainerClass.create_qlable(
                    'CORRECT COEFFICIENT : 1.0 | 1.0 | 1.0 | 1.0 | 1.0 '
                    '| 1.0 | 1.0 | 1.0 | 1.0 | 1.0', self.h_layout_cam_7)
                self.set_correct_coefficient = AppContainerClass.create_qpushbutton(self.h_layout_cam_7,
                                                                                    self.get_correct_coefficient_button)
                # SQL
                self.h_layout_g_sql = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_sql)
                self.g_sql_set = AppContainerClass.create_qlable('SQL', self.h_layout_g_sql, background=True)
                self.h_layout_sql_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_sql_1)
                self.sql_write = AppContainerClass.create_qcheckbox(self.h_layout_sql_1, 'CONNECT TO SQL?')
                self.ip_sql = AppContainerClass.create_qlable('IP SQL SERVER : 192.168.15.87', self.h_layout_sql_1)
                self.set_ip_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_1, self.get_ip_sql_button)
                self.server_sql = AppContainerClass.create_qlable(r'SERVER SQL : DESKTOP-SM7K050\COMPUTER_VISION',
                                                                  self.h_layout_sql_1)
                self.set_server_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_1,
                                                                           self.get_server_sql_button)
                self.port_sql = AppContainerClass.create_qlable('PORT SQL SERVER : 1433', self.h_layout_sql_1)
                self.set_port_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_1, self.get_port_sql_button)
                self.h_layout_sql_2 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_sql_2)
                self.database_sql = AppContainerClass.create_qlable('DATABASE SQL : analiz_16grohot',
                                                                    self.h_layout_sql_2)
                self.set_database_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_2,
                                                                             self.get_database_sql_button)
                self.user_sql = AppContainerClass.create_qlable('USER SQL : computer_vision', self.h_layout_sql_2)
                self.set_user_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_2, self.get_user_sql_button)
                self.h_layout_sql_3 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_sql_3)
                self.password_sql = AppContainerClass.create_qlable('PASSWORD SQL : vision12345678',
                                                                    self.h_layout_sql_2)
                self.set_password_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_2,
                                                                             self.get_password_sql_button)
                self.sql_now_check = AppContainerClass.create_qcheckbox(self.h_layout_sql_3, 'WRITE NOW SQL?')
                self.table_now_sql = AppContainerClass.create_qlable('TABLE NOW SQL : grohot16_now_table',
                                                                     self.h_layout_sql_3)
                self.set_table_now_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_3,
                                                                              self.get_table_now_sql_button)
                self.rows_now_sql = AppContainerClass.create_qlable('ROWS NOW SQL : device_row | value_row | alarm_row '
                                                                    '| datetime_row', self.h_layout_sql_3)
                self.set_rows_now_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_3,
                                                                             self.get_rows_now_sql_button)
                self.h_layout_sql_4 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_sql_4)
                self.sql_data_check = AppContainerClass.create_qcheckbox(self.h_layout_sql_4, 'WRITE DATA SQL?')
                self.table_data_sql = AppContainerClass.create_qlable('TABLE DATA SQL : grohot16_data_table',
                                                                      self.h_layout_sql_4)
                self.set_table_data_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_4,
                                                                               self.get_table_data_sql_button)
                self.rows_data_sql = AppContainerClass.create_qlable(
                    'ROWS DATA SQL : device_row | value_row | alarm_row '
                    '| datetime_row', self.h_layout_sql_4)
                self.set_rows_data_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_4,
                                                                              self.get_rows_data_sql_button)
                # DEBUG
                self.h_layout_g_debug = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_debug)
                self.g_debug_set = AppContainerClass.create_qlable('DEBUG', self.h_layout_g_debug, background=True)
                self.h_layout_debug_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_debug_1)
                self.auto_import_check = AppContainerClass.create_qcheckbox(self.h_layout_debug_1, 'AUTO IMPORT?')
                self.auto_play_check = AppContainerClass.create_qcheckbox(self.h_layout_debug_1, 'AUTO PLAY?')
                self.speed_analysis = AppContainerClass.create_qlable('SPEED ANALYSIS : 1.0', self.h_layout_debug_1)
                self.set_speed_analysis = AppContainerClass.create_qpushbutton(self.h_layout_debug_1,
                                                                               self.get_speed_analysis_button)
                self.speed_video_stream = AppContainerClass.create_qlable('SPEED VIDEO-STREAM : 1.0',
                                                                          self.h_layout_debug_1)
                self.set_speed_video_stream = AppContainerClass.create_qpushbutton(self.h_layout_debug_1,
                                                                                   self.get_speed_video_stream_button)
                self.h_layout_debug_2 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_debug_2)
                self.widget_data_value = AppContainerClass.create_qlable('0.00%', self.h_layout_debug_2)
                self.h_layout_debug_2.addStretch()
                self.widget_write = AppContainerClass.create_qcheckbox(self.h_layout_debug_2, 'WRITE TO WIDGET?')
                self.text_write = AppContainerClass.create_qcheckbox(self.h_layout_debug_2, 'WRITE TO TEXT?')
                self.source_win_type = AppContainerClass.create_qlable('SOURCE TYPE :', self.h_layout_debug_2)
                self.source_type = AppContainerClass.create_qcombobox(self.h_layout_debug_2,
                                                                      ['image-http', 'video-rtsp', 'video-file'],
                                                                      'image-http')
                self.compute_win_debug = AppContainerClass.create_qlable('COMPUTE TYPE :', self.h_layout_debug_2)
                self.compute_debug = AppContainerClass.create_qcombobox(self.h_layout_debug_2,
                                                                        ['sync', 'async', 'multithread',
                                                                         'multiprocess', 'complex'],
                                                                        'complex')
                self.process_cores = AppContainerClass.create_qlable('PROCESS CORES : 4', self.h_layout_debug_2)
                self.set_process_cores = AppContainerClass.create_qpushbutton(self.h_layout_debug_2,
                                                                              self.get_process_cores_button)
                self.h_layout_debug_3 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_debug_3)
                self.render_win_debug = AppContainerClass.create_qlable('Render windows :', self.h_layout_debug_3)
                self.render_debug = AppContainerClass.create_qcombobox(self.h_layout_debug_3,
                                                                       ['none', 'source', 'final', 'extended',
                                                                        'all'], 'none')
                self.resolution_debug_1 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '320x240',
                                                                                default=True)
                self.resolution_debug_1.toggled.connect(self.set_resolution_debug(self.resolution_debug_1, 320, 240))
                self.resolution_debug_2 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '640x480')
                self.resolution_debug_2.toggled.connect(self.set_resolution_debug(self.resolution_debug_2, 640, 480))
                self.resolution_debug_3 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '1280x720')
                self.resolution_debug_3.toggled.connect(self.set_resolution_debug(self.resolution_debug_3, 1280, 720))
                self.resolution_debug_4 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '1920x1080')
                self.resolution_debug_4.toggled.connect(self.set_resolution_debug(self.resolution_debug_4, 1920, 1080))
                self.resolution_debug_5 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '2560x1600')
                self.resolution_debug_5.toggled.connect(self.set_resolution_debug(self.resolution_debug_5, 2560, 1600))
                self.resolution_debug_6 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '3840x2160')
                self.resolution_debug_6.toggled.connect(self.set_resolution_debug(self.resolution_debug_6, 3840, 2160))
                # IMPORTS
                self.h_layout_g_imports = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_imports)
                self.g_imports_set = AppContainerClass.create_qlable('IMPORTS', self.h_layout_g_imports,
                                                                     background=True)
                self.h_layout_imports_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_imports_1)
                self.import_file = AppContainerClass.create_qlable('SETTINGS FILE NAME : settings',
                                                                   self.h_layout_imports_1)
                self.set_import_file = AppContainerClass.create_qpushbutton(self.h_layout_imports_1,
                                                                            self.get_settings_file_name)
                self.export_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_imports_1,
                                                                               self.export_settings_func, 'export')
                self.import_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_imports_1,
                                                                               self.import_settings_func, 'import')
                # SHOT
                self.h_layout_g_shot = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_shot)
                self.g_shot_set = AppContainerClass.create_qlable('SHOT', self.h_layout_g_shot, background=True)
                self.h_layout_shot_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_shot_1)
                self.ip_cam_snapshot = AppContainerClass.create_qlable('ip-cam : 15.204', self.h_layout_shot_1)
                self.set_ip_cam_snapshot = AppContainerClass.create_qpushbutton(self.h_layout_shot_1,
                                                                                self.get_ip_cam_snapshot_button)
                self.name_snapshot = AppContainerClass.create_qlable('file name : picture.jpg', self.h_layout_shot_1)
                self.set_name_snapshot = AppContainerClass.create_qpushbutton(self.h_layout_shot_1,
                                                                              self.set_name_snapshot_button)
                self.snapshot_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_shot_1,
                                                                                 self.snapshot_btn_func,
                                                                                 'snapshot')

                self.setLayout(self.v_layout_m)
                self.auto_play_func()
                self.auto_import_settings_func()

            def get_speed_analysis_button(self):
                widget = self.speed_analysis
                value, success = QtWidgets.QInputDialog.getDouble(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                  f'{widget.text().split(":")[0].strip()} value:',
                                                                  1.0, 0.01, 50.0, 2)
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_speed_video_stream_button(self):
                widget = self.speed_video_stream
                value, success = QtWidgets.QInputDialog.getDouble(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                  f'{widget.text().split(":")[0].strip()} value:',
                                                                  1.0, 0.01, 50.0, 2)
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_sensitivity_analysis_button(self):
                widget = self.sensitivity_analysis
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_alarm_level_button(self):
                widget = self.alarm_level
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_correct_coefficient_button(self):
                widget = self.correct_coefficient
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_protocol_cam_type_button(self):
                widget = self.protocol_cam_type
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_port_cam_button(self):
                widget = self.port_cam
                value, success = QtWidgets.QInputDialog.getInt(self, f'Set {widget.text().split(":")[0].strip()}',
                                                               f'{widget.text().split(":")[0].strip()} value:',
                                                               554, 1, 9999, 5)
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_login_cam_button(self):
                widget = self.login_cam
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_password_cam_button(self):
                widget = self.password_cam
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_alias_device_button(self):
                widget = self.alias_device
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_ip_cam_button(self):
                widget = self.ip_cam
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_mask_cam_button(self):
                widget = self.mask_cam
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_ip_sql_button(self):
                widget = self.ip_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_server_sql_button(self):
                widget = self.server_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_port_sql_button(self):
                widget = self.port_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_database_sql_button(self):
                widget = self.database_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_user_sql_button(self):
                widget = self.user_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_password_sql_button(self):
                widget = self.password_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_table_now_sql_button(self):
                widget = self.table_now_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_rows_now_sql_button(self):
                widget = self.rows_now_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_table_data_sql_button(self):
                widget = self.table_data_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_rows_data_sql_button(self):
                widget = self.rows_data_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_settings_file_name(self):
                widget = self.import_file
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def set_resolution_debug(self, radio, width: int, height: int):
                self.resolution_debug.append([radio, width, height])

            def get_window_resolution(self):
                for radio in self.resolution_debug:
                    try:
                        if radio[0].isChecked():
                            return [int(radio[1]), int(radio[2])]
                    except Exception as ex:
                        print(ex)

            def set_window_resolution(self, value):
                for radio in self.resolution_debug:
                    try:
                        if radio[1] == value[0]:
                            radio[0].setChecked(True)
                        else:
                            radio[0].setChecked(False)
                    except Exception as ex:
                        print(ex)

            def get_process_cores_button(self):
                widget = self.process_cores
                value, success = QtWidgets.QInputDialog.getInt(self, f'Set {widget.text().split(":")[0].strip()}',
                                                               f'{widget.text().split(":")[0].strip()} value:',
                                                               4, 1, 16, 5)
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_ip_cam_snapshot_button(self):
                widget = self.ip_cam_snapshot
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def set_name_snapshot_button(self):
                widget = self.name_snapshot
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_string_from_list(self, source: list):
                value = ''
                for x in source:
                    value = f'{value} | {x}'
                return value[3::]

            def set_data_func(self, value: str):
                try:
                    self.widget_data_value.setText(f"{value}")
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'set_data_func error : {ex}')

            def create_data_func(self):
                try:
                    data = {
                        'protocol_cam_type': str(self.protocol_cam_type.text().split(':')[1].strip()),
                        'port_cam': int(self.port_cam.text().split(':')[1].strip()),
                        'login_cam': str(self.login_cam.text().split(':')[1].strip()),
                        'password_cam': str(self.password_cam.text().split(':')[1].strip()),
                        'alias_device': list(
                            [x.strip() for x in self.alias_device.text().split(':')[1].strip().split('|')]),
                        'ip_cam': list([x.strip() for x in self.ip_cam.text().split(':')[1].strip().split('|')]),
                        'mask_cam': list([x.strip() for x in self.mask_cam.text().split(':')[1].strip().split('|')]),
                        'sensitivity_analysis': list([int(x.strip()) for x in
                                                      self.sensitivity_analysis.text().split(':')[1].strip().split(
                                                          '|')]),
                        'alarm_level': list(
                            [int(x.strip()) for x in self.alarm_level.text().split(':')[1].strip().split('|')]),
                        'correct_coefficient': list([float(x.strip()) for x in
                                                     self.correct_coefficient.text().split(':')[1].strip().split('|')]),

                        'sql_write': bool(self.sql_write.isChecked()),
                        'ip_sql': str(self.ip_sql.text().split(':')[1].strip()),
                        'server_sql': str(self.server_sql.text().split(':')[1].strip()),
                        'port_sql': str(self.port_sql.text().split(':')[1].strip()),
                        'database_sql': str(self.database_sql.text().split(':')[1].strip()),
                        'user_sql': str(self.user_sql.text().split(':')[1].strip()),
                        'password_sql': str(self.password_sql.text().split(':')[1].strip()),
                        'sql_now_check': bool(self.sql_now_check.isChecked()),
                        'table_now_sql': str(self.table_now_sql.text().split(':')[1].strip()),
                        'rows_now_sql': list(
                            [x.strip() for x in self.rows_now_sql.text().split(':')[1].strip().split('|')]),
                        'sql_data_check': bool(self.sql_data_check.isChecked()),
                        'table_data_sql': str(self.table_data_sql.text().split(':')[1].strip()),
                        'rows_data_sql': list(
                            [x.strip() for x in self.rows_data_sql.text().split(':')[1].strip().split('|')]),

                        'auto_import_check': bool(self.auto_import_check.isChecked()),
                        'auto_play_check': bool(self.auto_play_check.isChecked()),
                        'speed_analysis': float(self.speed_analysis.text().split(':')[1].strip()),
                        'speed_video_stream': float(self.speed_video_stream.text().split(':')[1].strip()),

                        'widget_write': bool(self.widget_write.isChecked()),
                        'text_write': bool(self.text_write.isChecked()),
                        'widget': self.set_data_func,
                        'source_type': str(self.source_type.currentText().strip()),
                        'compute_debug': str(self.compute_debug.currentText().strip()),
                        'process_cores': int(self.process_cores.text().split(':')[1].strip()),
                        'render_debug': str(self.render_debug.currentText().strip()),
                        'resolution_debug': list(self.get_window_resolution()),

                        'import_file': str(self.import_file.text().split(':')[1].strip()),

                        'ip_cam_snapshot': str(self.ip_cam_snapshot.text().split(":")[1].strip()),
                        'name_snapshot': str(self.name_snapshot.text().split(":")[1].strip()),
                    }
                    return data
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'create_data_func error : {ex}')

            def play_btn_func(self):
                try:
                    data = self.create_data_func()
                    self.play_f(data=data)
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'play_btn_func error : {ex}')

            def snapshot_btn_func(self):
                try:
                    data = self.create_data_func()
                    self.snapshot_f(data=data)
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'snapshot_btn_func error : {ex}')

            def export_settings_func(self):
                try:
                    data = self.create_data_func()
                    del data['widget']
                    JsonClass.write_json_to_file(dictionary=data)
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'export_settings_func error : {ex}')

            def import_settings_func(self):
                try:
                    data = self.create_data_func()
                    data = JsonClass.read_json_from_file(data)
                    self.protocol_cam_type.setText(f'{self.protocol_cam_type.text().split(":")[0].strip()} : '
                                                   f'{str(data["protocol_cam_type"])}')
                    self.port_cam.setText(f'{self.port_cam.text().split(":")[0].strip()} : '
                                          f'{str(data["port_cam"])}')
                    self.login_cam.setText(f'{self.login_cam.text().split(":")[0].strip()} : '
                                           f'{str(data["login_cam"])}')
                    self.password_cam.setText(f'{self.password_cam.text().split(":")[0].strip()} : '
                                              f'{str(data["password_cam"])}')
                    self.alias_device.setText(f'{self.alias_device.text().split(":")[0].strip()} : '
                                              f'{self.get_string_from_list(data["alias_device"])}')
                    self.ip_cam.setText(f'{self.ip_cam.text().split(":")[0].strip()} : '
                                        f'{self.get_string_from_list(data["ip_cam"])}')
                    self.mask_cam.setText(f'{self.mask_cam.text().split(":")[0].strip()} : '
                                          f'{self.get_string_from_list(data["mask_cam"])}')
                    self.sensitivity_analysis.setText(f'{self.sensitivity_analysis.text().split(":")[0].strip()} : '
                                                      f'{self.get_string_from_list(data["sensitivity_analysis"])}')
                    self.alarm_level.setText(f'{self.alarm_level.text().split(":")[0].strip()} : '
                                             f'{self.get_string_from_list(data["alarm_level"])}')
                    self.correct_coefficient.setText(f'{self.correct_coefficient.text().split(":")[0].strip()} : '
                                                     f'{self.get_string_from_list(data["correct_coefficient"])}')
                    self.sql_write.setChecked(data["sql_write"])
                    self.ip_sql.setText(f'{self.ip_sql.text().split(":")[0].strip()} : {str(data["ip_sql"])}')
                    self.server_sql.setText(
                        f'{self.server_sql.text().split(":")[0].strip()} : {str(data["server_sql"])}')
                    self.port_sql.setText(f'{self.port_sql.text().split(":")[0].strip()} : {str(data["port_sql"])}')
                    self.database_sql.setText(f'{self.database_sql.text().split(":")[0].strip()} : '
                                              f'{str(data["database_sql"])}')
                    self.user_sql.setText(f'{self.user_sql.text().split(":")[0].strip()} : '
                                          f'{str(data["user_sql"])}')
                    self.password_sql.setText(f'{self.password_sql.text().split(":")[0].strip()} : '
                                              f'{str(data["password_sql"])}')
                    self.sql_now_check.setChecked(data["sql_now_check"])
                    self.table_now_sql.setText(f'{self.table_now_sql.text().split(":")[0].strip()} : '
                                               f'{str(data["table_now_sql"])}')
                    self.rows_now_sql.setText(f'{self.rows_now_sql.text().split(":")[0].strip()} : '
                                              f'{self.get_string_from_list(data["rows_now_sql"])}')
                    self.sql_data_check.setChecked(data["sql_data_check"])
                    self.table_data_sql.setText(f'{self.table_data_sql.text().split(":")[0].strip()} : '
                                                f'{str(data["table_data_sql"])}')
                    self.rows_data_sql.setText(f'{self.rows_data_sql.text().split(":")[0].strip()} : '
                                               f'{self.get_string_from_list(data["rows_data_sql"])}')
                    self.auto_import_check.setChecked(data["auto_import_check"])
                    self.auto_play_check.setChecked(data["auto_play_check"])
                    self.speed_analysis.setText(f'{self.speed_analysis.text().split(":")[0].strip()} : '
                                                f'{str(data["speed_analysis"])}')
                    self.speed_video_stream.setText(f'{self.speed_video_stream.text().split(":")[0].strip()} : '
                                                    f'{str(data["speed_video_stream"])}')
                    self.widget_write.setChecked(data["widget_write"])
                    self.text_write.setChecked(data["text_write"])
                    self.source_type.setCurrentText(data["source_type"])
                    self.compute_debug.setCurrentText(data["compute_debug"])
                    self.process_cores.setText(f'{self.process_cores.text().split(":")[0].strip()} : '
                                               f'{str(data["process_cores"])}')
                    self.render_debug.setCurrentText(data["render_debug"])
                    self.set_window_resolution(data["resolution_debug"])
                    self.import_file.setText(f'{self.import_file.text().split(":")[0].strip()} : '
                                             f'{str(data["import_file"])}')
                    self.ip_cam_snapshot.setText(f'{self.ip_cam_snapshot.text().split(":")[0].strip()} : '
                                                 f'{str(data["ip_cam_snapshot"])}')
                    self.name_snapshot.setText(f'{self.name_snapshot.text().split(":")[0].strip()} : '
                                               f'{str(data["name_snapshot"])}')
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'import_settings_func error : {ex}')

            def auto_import_settings_func(self):
                try:
                    data = self.create_data_func()
                    data = JsonClass.read_json_from_file(data)
                    if data['auto_import_check']:
                        self.import_settings_func()
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'auto_import_settings_func error : {ex}')

            def auto_play_func(self):
                try:
                    data = self.create_data_func()
                    _data = JsonClass.read_json_from_file(data)
                    _data['widget'] = data['widget']
                    if _data['auto_play_check']:
                        self.play_f(data=_data)
                        self.showMinimized()
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'auto_play_func error : {ex}')


class SQLClass:

    @staticmethod
    def example_cv():
        class SQLClass:
            @staticmethod
            def sql_post_data(ip: str, server: str, port: str, database: str, username: str, password: str, table: str,
                              rows: list, values: list):
                try:
                    sql = SQLClass.pyodbc_connect(ip=ip, server=server, port=port, database=database, username=username,
                                                  password=password)
                    SQLClass.execute_data_query(connection=sql, table=table, rows=rows, values=values)
                except Exception as ex:
                    print(ex)
                    with open('log.txt', 'a') as log:
                        log.write(f'\n{ex}\n')

            @staticmethod
            def sql_post_now(ip: str, server: str, port: str, database: str, username: str, password: str, table: str,
                             rows: list, values: list):
                try:
                    sql = SQLClass.pyodbc_connect(ip=ip, server=server, port=port, database=database, username=username,
                                                  password=password)
                    SQLClass.execute_now_query(connection=sql, table=table, rows=rows, values=values)
                except Exception as ex:
                    print(ex)
                    with open('log.txt', 'a') as log:
                        log.write(f'\n{ex}\n')

            @staticmethod
            def pyodbc_connect(ip: str, server: str, port: str, database: str, username: str, password: str):
                conn_str = (
                        r'DRIVER={ODBC Driver 17 for SQL Server};SERVER=tcp:' + ip + '\\' + server + ',' + port +
                        ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password + ';'
                )
                return pyodbc.connect(conn_str)

            @staticmethod
            def pd_read_sql_query(connection, query: str, database: str, table: str):
                return pandas.read_sql_query(f'{query} {database}.dbo.{table} ORDER BY id', connection)

            @staticmethod
            def execute_data_query(connection, table: str, rows: list, values: list):
                cursor = connection.cursor()
                cursor.fast_executemany = True
                _rows = ''
                for x in rows:
                    _rows = f"{_rows}{str(x)}, "
                value = f"INSERT INTO {table} (" + _rows[:-2:] + f") VALUES {tuple(values)}"
                cursor.execute(value)
                connection.commit()

            @staticmethod
            def execute_now_query(connection, table, rows: list, values: list):
                cursor = connection.cursor()
                cursor.fast_executemany = True
                __rows = ''
                for x in rows:
                    __rows = f"{__rows}{str(x)}, "
                value = f"UPDATE {table} SET {rows[1]} = '{values[1]}',{rows[2]} = '{values[2]}' ,{rows[3]} = '{values[3]}' " \
                        f"WHERE {rows[0]} = '{values[0]}'"
                cursor.execute(value)
                connection.commit()

    @staticmethod
    def sql_post_data(ip: str, server: str, port: str, database: str, username: str, password: str, table: str,
                      rows: list, values: list):
        try:
            sql = SQLClass.pyodbc_connect(ip=ip, server=server, port=port, database=database, username=username,
                                          password=password)
            SQLClass.execute_data_query(connection=sql, table=table, rows=rows, values=values)
        except Exception as ex:
            print(ex)
            with open('log.txt', 'a') as log:
                log.write(f'\n{ex}\n')

    @staticmethod
    def sql_post_now(ip: str, server: str, port: str, database: str, username: str, password: str, table: str,
                     rows: list, values: list):
        try:
            sql = SQLClass.pyodbc_connect(ip=ip, server=server, port=port, database=database, username=username,
                                          password=password)
            SQLClass.execute_now_query(connection=sql, table=table, rows=rows, values=values)
        except Exception as ex:
            print(ex)
            with open('log.txt', 'a') as log:
                log.write(f'\n{ex}\n')

    @staticmethod
    def pyodbc_connect(ip: str, server: str, port: str, database: str, username: str, password: str):
        conn_str = (
                r'DRIVER={ODBC Driver 17 for SQL Server};SERVER=tcp:' + ip + '\\' + server + ',' + port +
                ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password + ';'
        )
        return pyodbc.connect(conn_str)

    @staticmethod
    def pd_read_sql_query(connection, query: str, database: str, table: str):
        return pandas.read_sql_query(f'{query} {database}.dbo.{table} ORDER BY id', connection)

    @staticmethod
    def execute_data_query(connection, table: str, rows: list, values: list):
        cursor = connection.cursor()
        cursor.fast_executemany = True
        _rows = ''
        for x in rows:
            _rows = f"{_rows}{str(x)}, "
        value = f"INSERT INTO {table} (" + _rows[:-2:] + f") VALUES {tuple(values)}"
        cursor.execute(value)
        connection.commit()

    @staticmethod
    def execute_now_query(connection, table, rows: list, values: list):
        cursor = connection.cursor()
        cursor.fast_executemany = True
        __rows = ''
        for x in rows:
            __rows = f"{__rows}{str(x)}, "
        value = f"UPDATE {table} SET {rows[1]} = '{values[1]}',{rows[2]} = '{values[2]}' ,{rows[3]} = '{values[3]}' " \
                f"WHERE {rows[0]} = '{values[0]}'"
        cursor.execute(value)
        connection.commit()

    # class SQLclass:
    #     def __init__(self, server, database, username, password, table):
    #         self.server = server
    #         self.database = database
    #         self.username = username
    #         self.password = password
    #         self.table = table
    #         self.cursor = self.pyodbc_connect(server=server, database=database, username=username,
    #                                           password=password).cursor()
    #
    #     @staticmethod
    #     def pyodbc_connect(server, database, username, password):
    #         return pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' + server + ';DATABASE=' +
    #                               database + ';UID=' + username + ';PWD=' + password +
    #                               ';Trusted_Connection=yes;')
    #
    #     @staticmethod
    #     def pd_read_sql_query(query: str, database: str, table: str, connection):
    #         return pd.read_sql_query(f'{query} {database}.dbo.{table} ORDER BY id', connection)
    #
    #     @staticmethod
    #     def execute_query(connection, table, rows: list, values: list):
    #         cursor = connection.cursor()
    #         cursor.fast_executemany = True
    #         __rows = ''
    #         for x in rows:
    #             __rows = f"{__rows}{str(x)}, "
    #         value = f"INSERT INTO {table} (" + __rows[:-2:] + f") VALUES {tuple(values)}"
    #         # value = f"INSERT INTO {table} (" + __rows[:-2:] + f") VALUES {tuple(values)}"
    #         # value = f"INSERT INTO {table} (id_row, device_row, percent_row, time_row, data_row, extra_row)
    #         # VALUES {tuple(values)}"
    #         cursor.execute(value)
    #         connection.commit()

    # # Server variables
    # _server = 'WIN-P4E9N6ORCNP\\ANALIZ_SQLSERVER'
    # _database = 'ruda_db'
    # _username = 'ruda_user'
    # _password = 'ruda_user'
    # _table = 'ruda_table'
    #
    # _date = f'{str(datetime.datetime.now()).split(" ")[0]}'
    # _time = f'{str(datetime.datetime.now()).split(" ")[1].split(".")[0]}'
    #
    # _rows = ['id_row', 'device_row', 'percent_row', 'time_row', 'data_row', 'extra_row']
    # _values = ['id_row', 'device_row', 'percent_row', f'{str(datetime.datetime.now()).split(" ")[1].split(".")[0]}',
    #            f'{str(datetime.datetime.now()).split(" ")[0]}', 'extra_row']

    # Read SQL data with Class
    # sql = SQLclass.pyodbc_connect(server=_server, database=_database, username=_username,
    #                               password=_password, table=_table)
    # data = SQLclass.pd_read_sql_query(query='SELECT * FROM', database=_database, table=_table, connection=sql)
    # print(data)
    # print(type(data))
    # for row in data:
    #     print(row)

    # # Read SQL data native
    # connection_native = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' + _server + ';DATABASE=' +
    #                                    _database + ';UID=' + _username + ';PWD=' + _password +
    #                                    ';Trusted_Connection=yes;')
    # sql_query_read = pd.read_sql_query(f'SELECT * FROM {_database}.dbo.{_table}', connection_native)
    # print(sql_query_read)
    # print(type(sql_query_read))
    # for row in sql_query_read:
    #     print(row)
    # print(type(row))

    # Write SQL data with Class
    # sql = SQLclass.pyodbc_connect(server=_server, database=_database, username=_username, password=_password)
    # SQLclass.execute_query(connection=sql, table=_table, rows=_rows, values=_values)

    # # Write SQL data native
    # connection_native = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' + _server + ';DATABASE=' +
    #                                    _database + ';UID=' + _username + ';PWD=' + _password +
    #                                    ';Trusted_Connection=yes;')
    # cursor = connection_native.cursor()
    # cursor.fast_executemany = True
    # count = cursor.execute(f"""INSERT INTO {_table} (id_row, device_row, percent_row, time_row, data_row, extra_row)
    # VALUES (id_row, device_row, percent_row, '{_time}', '{_date}', extra_row)""").rowcount
    # connection_native.commit()

    def example(self):
        # sql_select_query = f"SELECT * " \
        #                    f"FROM dbtable " \
        #                    f"WHERE CAST(temperature AS FLOAT) >= {temp} AND personid = '6176' OR personid = '25314' OR personid = '931777' OR personid = '5863' " \
        #                    f"ORDER BY date1 DESC, date2 DESC;"
        # sql_select_query = f"SELECT * " \
        #                    f"FROM dbtable " \
        #                    f"WHERE CAST(temperature AS FLOAT) < 37.0 AND date1 BETWEEN '2021-07-25' AND '2021-08-25' " \
        #                    f"ORDER BY date1 DESC, date2 DESC;"
        connect_db = self.pyodbc_connect()
        cursor = connect_db.cursor()
        cursor.fast_executemany = True
        # cursor.execute(sql_select_query)
        data = cursor.fetchall()


########################################################################################################################
if __name__ == '__main__':
    # SyncAsyncThreadingPoolExecutorClass.Example.example_sync_compute()
    # SyncAsyncThreadingPoolExecutorClass.Example.example_async_compute()
    # SyncAsyncThreadingPoolExecutorClass.Example.example_threading_compute()
    # SyncAsyncThreadingPoolExecutorClass.Example.example_thread_pool_executor_compute()
    # SyncAsyncThreadingPoolExecutorClass.Example.example_process_pool_executor_compute()

    pass
