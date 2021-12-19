# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook('BI 2000 в лицо.xlsx')
print(wb)
sheet = wb.get_sheet_by_name('Личные данные')
print(sheet)
sheet.title