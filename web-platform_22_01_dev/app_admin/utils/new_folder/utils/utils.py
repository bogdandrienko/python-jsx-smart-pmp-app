
# импортируем библиотеки
import openpyxl
from openpyxl.utils import get_column_letter

print('begin')

# Читаем данные с файла больницы
bol_var = 'mineraly.xlsx'
workbook = openpyxl.load_workbook(bol_var)
sheet = workbook.active
max_row = sheet.max_row
global_list_matrix_bol = []
for row in range(1, max_row + 1):
    local_list_matrix = []
    for col in range(1, 8 + 1):
        value = sheet[str(get_column_letter(col)) + str(row)].value
        if value == None or value == '' or value == 'None' or value == 'none':
            value = ' '
        local_list_matrix.append(value)



print('complete')