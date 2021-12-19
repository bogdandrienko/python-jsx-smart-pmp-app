# импортируем библиотеки
import time

import openpyxl
from openpyxl.utils import get_column_letter

from app_admin.utils import ExcelClass

bool_object = True


# print(type(bool_object))


class CustomPersonWorker:
    """
    Класс для обучения, будет содержать данные по одному работнику
    """

    def __init__(self, name, surname, patronumyc, position, subdivion, age=20):
        self.age = age
        self.name = name
        self.surname = surname
        self.patronumyc = patronumyc
        self.position = position
        self.subdivion = subdivion

    def happy_birthday(self):
        self.age = self.age + 2

    def say_your_name(self):
        print('I am a ' + str(self.age))

    @staticmethod
    def say_your_age(text):
        time.sleep(0.1)
        print(text)


# CustomPersonWorker.say_hello(text='Привет')

# user_object = CustomPersonWorker(age=29, name='Иван', surname='Иванов', position='Раб')
# user_object2 = CustomPersonWorker(age=25, name='Григорий', surname='Иванов', position='Раб')
#
# print(user_object.name)
# print(user_object2.name)
#
# user_object.say_your_name()
# user_object.happy_birthday()
# user_object.say_your_name()

print('begin')

# Читаем данные с файла больницы
import_file = 'import.xlsx'
# Загружаем эксель файл в оперативу
workbook = openpyxl.load_workbook(import_file)
# Активируем рабочую страницу
sheet = workbook.active
# Берём со страницы максимум количества строк
max_row = sheet.max_row
# Создаём пустой массив для работников
global_list_matrix_bol = []
for row in range(1, max_row + 1):
    # Создаём объект-работника
    user_object = CustomPersonWorker(
        age=20,
        name=sheet[str(get_column_letter(4)) + str(row)].value,
        surname=sheet[str(get_column_letter(5)) + str(row)].value,
        patronumyc=sheet[str(get_column_letter(6)) + str(row)].value,
        position=sheet[str(get_column_letter(8)) + str(row)].value,
        subdivion=sheet[str(get_column_letter(1)) + str(row)].value,
    )
    # Добавляем объект-работника в общий массив
    global_list_matrix_bol.append(user_object)
# Закрытие эксель-файла
workbook.close()


    # local_list_matrix = []
    # for col in range(1, 8 + 1):
    #     value = sheet[str(get_column_letter(col)) + str(row)].value
    #     if value is None or value == '' or value == 'None' or value == 'none':
    #         value = ' '
    #     local_list_matrix.append(value)
    # global_list_matrix_bol.append(local_list_matrix)

print(global_list_matrix_bol)

# Вывод на экран массива с данными объекта
for user in global_list_matrix_bol:
    print(user.name + ' ' + user.surname + ' : ' + user.subdivion)
    if user.subdivion == 'Энергоуправление':
        print(user.name + ' ' + user.surname + ' : Я работник энергоуправления!')

# записываем результат в файл
final = 'final.xlsx'
# workbook = openpyxl.load_workbook(final)
workbook = ExcelClass.workbook_load(excel_file=final)
# sheet = workbook.active
sheet = ExcelClass.workbook_activate(workbook=workbook)
index = 0
for user in global_list_matrix_bol:
    index = index + 1

    sheet['A' + str(index)] = user.name
    sheet['B' + str(index)] = user.surname
    sheet['C' + str(index)] = user.patronumyc
    sheet['D' + str(index)] = user.subdivion


    # for man in user:
    #     local_index = user.index(man) + 1
    #     local_char = get_column_letter(local_index)
    #     address = local_char + str(index)
    #     sheet[address] = str(man)
# workbook.save(final)
# workbook.close()
ExcelClass.workbook_save(excel_file=final, workbook=workbook)
ExcelClass.workbook_close(workbook=workbook)

print('complete')
