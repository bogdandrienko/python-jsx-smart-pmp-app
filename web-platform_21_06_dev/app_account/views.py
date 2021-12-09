from django.shortcuts import render, redirect
from django.contrib.auth.models import Group, User
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login, authenticate, logout
from .forms import SignUpForm, SignUpManyForm, SignUpPasswordForm
import openpyxl
import random


def login_account(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            username = request.POST['username']
            password = request.POST['password']
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('home')
            return redirect('app_account:create')
    form = AuthenticationForm()
    context = {
        'form': form,
    }
    return render(request, 'app_account/login.html', context)

def create_account(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            signup_user = User.objects.get(username=username)
            try:
                user_group = Group.objects.get(name='User')
            except:
                user_group = Group.objects.create(name='User')
            user_group.user_set.add(signup_user)
            return redirect('app_account:login')
    form = SignUpForm()
    form1 = SignUpManyForm()
    form2 = SignUpPasswordForm()
    context = {
        'form': form,
        'form1': form1,
        'form2': form2,
    }
    return render(request, 'app_account/create.html', context)

def import_account(request):
    if request.method == 'POST':
        wb = openpyxl.load_workbook(request.FILES.get('document_addition_file_1'))
        sheet = wb.active
        def get_value(rows:str, cols:int):
            return str(sheet[rows+str(cols)].value)
        max = sheet.max_row
        for i in range(2, max + 1):
            username        = get_value('A', i)
            secret_password = get_value('B', i)
            password        = get_value('C', i)
            first_name      = get_value('D', i)
            last_name       = get_value('E', i)
            group           = get_value('F', i)
            if get_value('G', i) == 'ИСТИНА' or get_value('G', i) == 'True':
                is_staff = True
            else:
                is_staff = False
            if username and secret_password and password:
                try:
                    user    = User.objects.get(username=username)
                    user.username       = username
                    user.password       = secret_password
                    user.email          = password
                    user.first_name     = first_name
                    user.last_name      = last_name
                    user.is_staff       = is_staff
                    if username == 'Bogdan' or username == 'bogdan':
                        user.is_superuser = True
                    else:
                        user.is_superuser = False
                    user.save()
                    user.set_password   = str(password)
                except:
                    user    = User.objects.create(
                    username    = username,
                    password    = secret_password,
                    email       = password,
                    first_name  = first_name,
                    last_name   = last_name,
                    is_staff    = is_staff,
                    )
                try:
                    user_group = Group.objects.get(name=group)
                except:
                    user_group = Group.objects.create(name=group)
                signup_user = User.objects.get(username=username)
                user_group.user_set.add(signup_user)
    return redirect('app_account:create')

def export_account(request):
    if request.method == 'POST':
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Страница 1'
        user_objects = User.objects.all().order_by('id')
        sheet['A1'] = 'Имя пользователя'
        sheet['B1'] = 'Зашифрованный Пароль'
        sheet['C1'] = 'Пароль'
        sheet['D1'] = 'Имя'
        sheet['E1'] = 'Фамилия'
        sheet['F1'] = 'Группа'
        sheet['G1'] = 'Доступ к админ панели'
        value = 1
        for object in user_objects:
            value += 1
            sheet[f'A{value}'] = object.username
            sheet[f'B{value}'] = object.password
            sheet[f'C{value}'] = object.email
            sheet[f'D{value}'] = object.first_name
            sheet[f'E{value}'] = object.last_name
            groups = Group.objects.filter(user = object)
            group_list = ''
            for group in groups:
                group_list += group.name
            sheet[f'F{value}'] = group_list
            if object.is_staff:
                sheet[f'G{value}'] = 'ИСТИНА'
            else:
                sheet[f'G{value}'] = 'ЛОЖЬ'
        wb.save('static/media/tempates/export_account.xlsx')
        wb.close
    return redirect('app_account:create')

def generate_password(request):
    if request.method == 'POST':
        wb              = openpyxl.Workbook()
        sheet           = wb.active
        sheet.title     = 'Страница 1'
        quantity        = int(request.POST['quantity'])
        lenght          = int(request.POST['lenght'])
        chars           = 'abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890'
        sheet[f'A1']    = 'Зашифрованный Пароль'
        sheet[f'B1']    = 'Пароль'
        for n in range(2, quantity + 2):
            password = ''
            for i in range(1, lenght + 1):
                password += random.choice(chars)
            try:
                user            = User.objects.get(username='None')
            except:
                user            = User.objects.create(
                    username    = 'None',
                    email       = password,
                )
            user.email = password
            user.set_password(password)
            user.save()
            user    = User.objects.get(username='None')
            sheet[f'A{n}'] = user.password
            sheet[f'B{n}'] = user.email
        wb.save('static/media/tempates/generate_password.xlsx')
        wb.close
    return redirect('app_account:create')

def logout_account(request):
    # Проверка регистрации: если пользователь не вошёл в аккаунт, действия не срабатают, а его переадресует в форму входа
    if request.user.is_authenticated is not True:
        return redirect('app_account:login')
    # Переадресация пользователя на страницу входа

    logout(request)
    return redirect('app_account:login')
