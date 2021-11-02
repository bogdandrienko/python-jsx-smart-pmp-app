import openpyxl
from openpyxl.utils import get_column_letter
import requests
from bs4 import BeautifulSoup


def pandas(url):
    print(url)
    headers = {'user-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:72.0) Gecko/20100101 Firefox/72.0',
               'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'}
    # url = 'http://www.pogodaiklimat.ru/weather.php?id=35042&bday=1&fday=1&amonth=10&ayear=2021&bot=2'
    response = requests.get(url, headers=headers)
    response.encoding = 'utf-8'
    soup = BeautifulSoup(response.text, "html.parser")
    soup.encoded = 'utf-8'
    rows = soup.find('div', class_='archive-table-wrap')
    return rows


def set_sheet_value(column, row, value_, sheet):
    try:
        if value_:
            sheet[str(column) + str(row)] = value_
        else:
            sheet[str(column) + str(row)] = ''
    except:
        sheet[str(column) + str(row)] = ''


print('begin')
arr_url = []
for year in range(2011, 2012):
    for month in range(1, 2):
        for day in range(1, 32):
            url = f"http://www.pogodaiklimat.ru/weather.php?id=35042&bday={day}&fday={day}&amonth={month}&ayear={year}&bot=2"
            # print(f"day={day} month={month} year={year}")
            arr_url.append(url)
# print(arr_url)

arr_requests = []
for url in arr_url:
    arr_requests.append(pandas(url))
# print(arr_requests)

all_data = []
for request in arr_requests:
    data_list = []
    for row in request.find_all('tr'):
        local_data = []
        # print(row)
        # print('\n *************** \n')
        for col in row.find_all('td'):
            local_data.append(col.text)
            # print(col.text)
            # print('\n *************** \n')
        data_list.append(local_data)
    # print(data_list)
    all_data.append([''])
    all_data.append(data_list)
# print(all_data)

_import_file = 'data_1.xlsx'
workbook = openpyxl.load_workbook(_import_file)
sheet = workbook.active

row_index = 0
for data in all_data:
    # print(data)
    for row in data:
        # print(row)
        row_index += 1
        col_index = 0
        for cell in row:
            col_index += 1
            sheet[get_column_letter(col_index) + str(row_index)] = str(cell)

workbook.save(_import_file)
workbook.close()
print('complete')
