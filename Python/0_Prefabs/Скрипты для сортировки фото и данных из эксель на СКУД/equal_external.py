import openpyxl
from openpyxl.utils import get_column_letter


def get_sheet_value(column, row, _sheet):
    return str(_sheet[str(column) + str(row)].value)


def set_sheet_value(column, row, value):
    sheet[f'{column}{row}'] = str(value)


file_1c = 'Export_1.xlsx'
workers_id_1c = []
id_1c = 3

workbook = openpyxl.load_workbook(file_1c)
sheet = workbook.active
for num in range(1, 350):
    stringes = get_sheet_value(get_column_letter(int(id_1c)), num, _sheet=sheet).split(' ')
    try:
        string = f"{stringes[0]} {stringes[1]}"
        workers_id_1c.append(string)
    except:
        workers_id_1c.append(f"None None")
workbook.close()

# print(workers_id_1c)

file_skud = 'import.xlsx'
workers_id_skud = []
id_skud = 2

workbook = openpyxl.load_workbook(file_skud)
sheet = workbook.active
for num in range(1, 350):
    string = f"{get_sheet_value(get_column_letter(int(id_skud)), num, _sheet=sheet)} " \
             f"{get_sheet_value(get_column_letter(int(id_skud)-1), num, _sheet=sheet)}"
    workers_id_skud.append(string)
workbook.close()

# print(workers_id_skud)
workbook = openpyxl.load_workbook(file_skud)
sheet = workbook.active

for x in workers_id_skud:
    for y in workers_id_1c:
        print(workers_id_skud.index(x))
        if x == y and x is not None:
            set_sheet_value(column=get_column_letter(int(18)), row=workers_id_skud.index(x)+1, value='МСС')
workbook.save('Import_1.xlsx')
