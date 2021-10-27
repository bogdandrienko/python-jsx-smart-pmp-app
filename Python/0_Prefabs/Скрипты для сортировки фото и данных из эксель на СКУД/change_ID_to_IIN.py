import os
import tkinter
import openpyxl
from openpyxl.utils import get_column_letter
from shutil import move
from fnmatch import fnmatch


def create_dir(dir_name='new folder', current_path=os.path.dirname(os.path.abspath('__file__'))):
    f"""
Create directory with {dir_name} - name in {current_path} - path. And return this full path.
    :param current_path: 
    :param dir_name: 
    :return: Full path with created directory
    """
    full_path = current_path + f'\\{dir_name}'
    try:
        os.makedirs(full_path)
    except:
        print('directory already yet')
    finally:
        return full_path


import_folder = create_dir('Исходные фото')
equal_folder = create_dir('Изменённые названия')
not_equal_folder = create_dir('Нет в базе')
correct_folder = create_dir('Корректировка')
error_folder = create_dir('Ошибка')


def click_button(export_file='', col_entry='754'):
    """

    :param export_file:
    :param col_entry:
    """
    app.root.config(background="red")
    print('start')

    pattern = '*.jpg'
    formatting = '.jpg'
    cols = [get_column_letter(int(x)) for x in col_entry]
    workers_id_all = []
    workers_full_name_all = []

    workbook = openpyxl.load_workbook(export_file)
    sheet = workbook.active
    for num in range(1, 5000):
        worker_list = []
        for x in cols:
            value = get_sheet_value(x, num, sheet=sheet)
            if value == 'None' or value is None or value == '':
                value = ''
            worker_list.append(value)
        workers_id_all.append(worker_list)
    workbook.close()
    # print(workers_id_all)

    for path, subdirs, files in os.walk(import_folder):
        for name in files:
            if fnmatch(name, pattern):
                try:
                    name_start = name.split('_')[0].strip()
                    name_end = name.split('.')[1].strip()
                    id_ = name.split('.')[0].strip().split('_')[1].strip()
                    for num in workers_id_all:
                        if id_ == num[0]:
                            Actions.action(f'{import_folder}\\{name}',
                                           f'{equal_folder}\\{name_start+"_"+num[1]+"."+name_end}')
                        else:
                            pass
                except Exception as ex:
                    print(name, ex)

    print('complete')
    app.root.config(background="green")


def get_sheet_value(column, row, sheet):
    return str(sheet[str(column) + str(row)].value)


class Actions:
    def __init__(self, obj_id):
        self.obj_id = obj_id

    @staticmethod
    def action(src, dst):
        move(src, dst)

    @staticmethod
    def equal(file):
        Actions.action(f'{import_folder}\\{file}', f'{equal_folder}\\{file}')

    @staticmethod
    def not_equal(file):
        Actions.action(f'{import_folder}\\{file}', f'{not_equal_folder}\\{file}')

    @staticmethod
    def correct(file):
        Actions.action(f'{import_folder}\\{file}', f'{correct_folder}\\{file}')

    @staticmethod
    def error(file):
        Actions.action(f'{import_folder}\\{file}', f'{error_folder}\\{file}')


class Application(tkinter.Frame):
    """

    """

    def __init__(self, root, **kw):
        super().__init__(**kw)
        self.root = root

        # Configure the root object for the Application
        self.root.title("Смена табельных на фото на ИИН")
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        self.root.config(background="black")
        self.root.geometry('640x480')
        self.master.minsize(640, 480)
        self.master.maxsize(640, 480)
        self.id = 0
        self.iid = 0

        self.submit_button = tkinter.Button(self.root, text="Запустить", font="100", command=self.insert_data)
        self.submit_button.grid(row=0, column=0, sticky=tkinter.W)

        self.exit_button = tkinter.Button(self.root, text="Выход", font="100", command=self.quit)
        self.exit_button.grid(row=0, column=1, sticky=tkinter.W)

        # Define the different GUI widgets
        self.export_label = tkinter.Label(self.root, text="Файл excel для сравнения", font="100")
        self.export_label.grid(row=1, column=0, sticky=tkinter.W)

        self.export_entry = tkinter.Entry(self.root, font="100")
        self.export_entry.grid(row=1, column=1, sticky=tkinter.W)
        self.export_entry.insert(0, 'export_test.xlsx')

        self.col_label = tkinter.Label(self.root, text="Введите цифру столбца с 'ИИН' 'Таб' 'Имя' 'Фамилия' -(9754)",
                                       font="100")
        self.col_label.grid(row=2, column=0, sticky=tkinter.W)

        self.col_entry = tkinter.Entry(self.root, font="100")
        self.col_entry.grid(row=2, column=1, sticky=tkinter.W)
        self.col_entry.insert(0, '34')

    def insert_data(self):
        """

        """
        click_button(self.export_entry.get(), self.col_entry.get())


if __name__ == "__main__":
    app = Application(tkinter.Tk())
    app.root.mainloop()
