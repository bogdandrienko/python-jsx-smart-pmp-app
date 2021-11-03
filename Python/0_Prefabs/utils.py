# import datetime
# from datetime import datetime
# from datetime import datetime as dt
import os
import time
import chardet
import openpyxl
import pyttsx3
import pyodbc
import pandas
from typing import Union
from openpyxl.utils import get_column_letter
from functools import wraps


########################################################################################################################
# # Импорт данного файла в проект
# from utilities import type_variable
#
# # Инициализация переменной
# var = 11.0
#
# # Вывод значения и типа переменной на экран
# print_variable_and_type_variable(var)
#
# # Конвертирование и присвоение нового значения переменной в несуществующий формат
# var = convert_variable(var, 'привет')
#
# # Конвертирование и присвоение нового значения переменной
# var = convert_variable(var, 'int')
#
# # Вывод значения и типа переменной на экран
# print_variable_and_type_variable(var)
########################################################################################################################


class MeasureTimeClass:
    @staticmethod
    def measure_time(func):
        @wraps(func)
        def wrap(*args, **kwargs):
            start = time.perf_counter()
            result = func(*args, **kwargs)
            elapsed = time.perf_counter() - start
            print(f'Executed {func} in {elapsed:0.4f} seconds')
            return result

        return wrap

    @staticmethod
    def async_measure_time(func):
        @wraps(func)
        async def wrap(*args, **kwargs):
            start = time.perf_counter()
            result = await func(*args, **kwargs)
            elapsed = time.perf_counter() - start
            print(f'Executed {func} in {elapsed:0.4f} seconds')
            return result

        return wrap


class EncodingClass:
    @staticmethod
    def example(self):
        a = 'РђР»РµРєСЃР°РЅРґСЂР° РџСЂРѕРєРѕС„СЊРµРІР°'
        b = a.encode()
        c = self.find_encoding(b)
        print(c)
        d = self.convert_encoding(a)
        print(d)

        # value = ''
        # try:
        #     val = value.encode('1251').decode('utf-8')
        # except Exception as ex_1:
        #     val = str(value).split(" ")
        #     string = "N" + val[0][2:] + " " + "N" + val[1][2:]
        #     try:
        #         sheet[f'{get_column_letter(value_index)}{row_index + 1}'] = string.encode('1251').decode(
        #             'utf-8')
        #     except Exception as ex_2:
        #         sheet[f'{get_column_letter(value_index)}{row_index + 1}'] = value
        #         print(ex_2)
        #     print(ex_1)

        # encoding = chardet.detect(content)['encoding']
        # print('\n ***************** \n')
        # print(encoding)
        # print('\n ***************** \n')
        # print(content[:100])
        # print('\n ***************** \n')
        # print(content.decode()[:100])
        # print('\n ***************** \n')
        # print(content.decode(encoding='UTF-8', errors='strict')[:100])
        # print('\n ***************** \n')
        # print(content.decode(encoding='UTF-8-SIG', errors='strict')[:100])
        # print('\n ***************** \n')
        # # print(content.decode(encoding='1251', errors='strict'))
        # # print('\n ***************** \n')

        # print('\n ************ \n')
        # for num in str(MassivKhesha)[1:10]:
        #     print(ord(num))
        # print('\n ************ \n')
        # MassivKhesha_1 = str(sha256(KodShifra.encode('utf-8')).hexdigest())
        # print(blake2b(b'XWew151299Ioo@').hexdigest())
        # print(hashlib.sha256("XWew151299Ioo@".encode()).digest().decode())
        # for num in str(hashlib.sha256("XWew151299Ioo@".encode()).digest().decode())[:5]:
        #     print(ord(num))
        # print(MassivKhesha_1)
        # for num in MassivKhesha_1:
        #     print(ord(num))

    @staticmethod
    def convert_encoding(data: str, new_coding='utf-8'):
        encoding = chardet.detect(data)['encoding']
        _data = data.encode(encoding)
        __data = _data.decode(encoding, data)
        ___data = __data.encode(new_coding)
        return ___data

    @staticmethod
    def find_encoding(data):
        encoding = chardet.detect(data)['encoding']
        return encoding


class VoiceClass:
    @staticmethod
    def example(self):
        # Not create class:
        self.speak()
        self.speak('Приветики!')
        # With create class:
        voice = self()
        voice.say()
        voice.say('Я уничтожу человечество!!!')

    def __init__(self, obj_id=0, volume=1.0, rate=200, voice=None, voices=None):
        self.engine = pyttsx3.init()
        self.volume = volume
        self.rate = rate
        self.voice = voice
        self.voices = voices
        self.obj_id = obj_id
        self.properties = [self.volume, self.rate, self.voice, self.voices]

    @staticmethod
    def speak(text: str = 'Inicialization successfull.'):
        pyttsx3.speak(text)

    @staticmethod
    async def async_speak(text: str = 'Inicialization successfull.'):
        await pyttsx3.speak(text)
        return text

    def say(self, text: str = 'Inicialization successfull.'):
        self.engine.say(text)
        self.engine.runAndWait()

    def get_property(self, name='volume'):
        return self.engine.getProperty(name)

    def get_properties(self):
        return [self.engine.getProperty(name) for name in self.properties]

    def set_property(self, name='volume', value=1.0):
        self.engine.setProperty(name, value)

    def set_properties(self, volume=1.0, rate=200):
        self.set_property(self.properties[0], volume)
        self.set_property(self.properties[1], rate)


class TypeVariablesClass:
    @staticmethod
    def check_type_of_variable(variable):
        return type(variable)

    @staticmethod
    def convert_variable(variable, new_type):
        if new_type == 'bool':
            return bool(variable)
        if new_type == 'float':
            return float(variable)
        if new_type == 'int':
            return int(variable)
        if new_type == 'str':
            return str(variable)
        if new_type == 'list':
            return list(variable)
        if new_type == 'dict':
            return dict(variable)
        if new_type == 'tuple':
            return tuple(variable)
        if new_type == 'set':
            return set(variable)
        else:
            print('Error: "' + str(new_type) + '" - not correct Parameter' +
                  ' / Parameter = new_type / Function = convert_variable')
        return variable


class Pagination:
    @staticmethod
    def example(self):
        alpabet_list = list('abcdefghhklmnt')
        p = Pagination(alpabet_list, 4)
        print(p.get_visible_items())

    def __init__(self, items=None, page_size=10):
        if items is None:
            items = []
        self.items = items
        self.page_size = page_size
        self.total_pages = 1 if not self.items else len(self.items) // self.page_size + 1
        self.current_page = 1

    def get_items(self):
        return self.items

    def get_page_size(self):
        return self.page_size

    def get_current_page(self):
        return self.current_page

    def prev_page(self):
        if self.current_page == 1:
            return self
        self.current_page -= 1
        return self

    def next_page(self):
        if self.current_page == self.total_pages:
            return self
        self.current_page += 1
        return self

    def first_page(self):
        self.current_page = 1
        return self

    def last_page(self):
        self.current_page = self.total_pages
        return self

    def go_to_page(self, page):
        if page < 1:
            page = 1
        elif page > self.total_pages:
            page = self.total_pages
        self.current_page = page
        return self

    def get_visible_items(self):
        start = (self.current_page - 1) * self.page_size
        return self.items[start:start + self.page_size]


class SQLClass:
    @staticmethod
    def example(self):
        # sql_select_query = f"SELECT * " \
        #                    f"FROM dbtable " \
        #                    f"WHERE CAST(temperature AS FLOAT) >= {temp} AND personid = '6176' OR personid = '25314' OR personid = '931777' OR personid = '5863' " \
        #                    f"ORDER BY date1 DESC, date2 DESC;"
        sql_select_query = f"SELECT * " \
                           f"FROM dbtable " \
                           f"WHERE CAST(temperature AS FLOAT) < 37.0 AND date1 BETWEEN '2021-07-25' AND '2021-08-25' " \
                           f"ORDER BY date1 DESC, date2 DESC;"
        connect_db = self.pyodbc_connect()
        cursor = connect_db.cursor()
        cursor.fast_executemany = True
        cursor.execute(sql_select_query)
        data = cursor.fetchall()

    @staticmethod
    def sql_post_data(ip: str, server: str, port: str, database: str, username: str, password: str, table: str,
                      rows: list, values: list):
        try:
            sql = SQLClass.pyodbc_connect(ip=ip, server=server, port=port, database=database, username=username,
                                          password=password)
            SQLClass.execute_data_query(connection=sql, table=table, rows=rows, values=values)
        except Exception as ex:
            print(ex)
            with open('log.txt', 'a') as log:
                log.write(f'\n{ex}\n')

    @staticmethod
    def sql_post_now(ip: str, server: str, port: str, database: str, username: str, password: str, table: str,
                     rows: list, values: list):
        try:
            sql = SQLClass.pyodbc_connect(ip=ip, server=server, port=port, database=database, username=username,
                                          password=password)
            SQLClass.execute_now_query(connection=sql, table=table, rows=rows, values=values)
        except Exception as ex:
            print(ex)
            with open('log.txt', 'a') as log:
                log.write(f'\n{ex}\n')

    @staticmethod
    def pyodbc_connect(ip: str, server: str, port: str, database: str, username: str, password: str):
        conn_str = (
                r'DRIVER={ODBC Driver 17 for SQL Server};SERVER=tcp:' + ip + '\\' + server + ',' + port +
                ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password + ';'
        )
        return pyodbc.connect(conn_str)

    @staticmethod
    def pd_read_sql_query(connection, query: str, database: str, table: str):
        return pandas.read_sql_query(f'{query} {database}.dbo.{table} ORDER BY id', connection)

    @staticmethod
    def execute_data_query(connection, table: str, rows: list, values: list):
        cursor = connection.cursor()
        cursor.fast_executemany = True
        _rows = ''
        for x in rows:
            _rows = f"{_rows}{str(x)}, "
        value = f"INSERT INTO {table} (" + _rows[:-2:] + f") VALUES {tuple(values)}"
        cursor.execute(value)
        connection.commit()

    @staticmethod
    def execute_now_query(connection, table, rows: list, values: list):
        cursor = connection.cursor()
        cursor.fast_executemany = True
        __rows = ''
        for x in rows:
            __rows = f"{__rows}{str(x)}, "
        value = f"UPDATE {table} SET {rows[1]} = '{values[1]}',{rows[2]} = '{values[2]}' ,{rows[3]} = '{values[3]}' " \
                f"WHERE {rows[0]} = '{values[0]}'"
        cursor.execute(value)
        connection.commit()

    # class SQLclass:
    #     def __init__(self, server, database, username, password, table):
    #         self.server = server
    #         self.database = database
    #         self.username = username
    #         self.password = password
    #         self.table = table
    #         self.cursor = self.pyodbc_connect(server=server, database=database, username=username,
    #                                           password=password).cursor()
    #
    #     @staticmethod
    #     def pyodbc_connect(server, database, username, password):
    #         return pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' + server + ';DATABASE=' +
    #                               database + ';UID=' + username + ';PWD=' + password +
    #                               ';Trusted_Connection=yes;')
    #
    #     @staticmethod
    #     def pd_read_sql_query(query: str, database: str, table: str, connection):
    #         return pd.read_sql_query(f'{query} {database}.dbo.{table} ORDER BY id', connection)
    #
    #     @staticmethod
    #     def execute_query(connection, table, rows: list, values: list):
    #         cursor = connection.cursor()
    #         cursor.fast_executemany = True
    #         __rows = ''
    #         for x in rows:
    #             __rows = f"{__rows}{str(x)}, "
    #         value = f"INSERT INTO {table} (" + __rows[:-2:] + f") VALUES {tuple(values)}"
    #         # value = f"INSERT INTO {table} (" + __rows[:-2:] + f") VALUES {tuple(values)}"
    #         # value = f"INSERT INTO {table} (id_row, device_row, percent_row, time_row, data_row, extra_row)
    #         # VALUES {tuple(values)}"
    #         cursor.execute(value)
    #         connection.commit()

    # # Server variables
    # _server = 'WIN-P4E9N6ORCNP\\ANALIZ_SQLSERVER'
    # _database = 'ruda_db'
    # _username = 'ruda_user'
    # _password = 'ruda_user'
    # _table = 'ruda_table'
    #
    # _date = f'{str(datetime.datetime.now()).split(" ")[0]}'
    # _time = f'{str(datetime.datetime.now()).split(" ")[1].split(".")[0]}'
    #
    # _rows = ['id_row', 'device_row', 'percent_row', 'time_row', 'data_row', 'extra_row']
    # _values = ['id_row', 'device_row', 'percent_row', f'{str(datetime.datetime.now()).split(" ")[1].split(".")[0]}',
    #            f'{str(datetime.datetime.now()).split(" ")[0]}', 'extra_row']

    # Read SQL data with Class
    # sql = SQLclass.pyodbc_connect(server=_server, database=_database, username=_username,
    #                               password=_password, table=_table)
    # data = SQLclass.pd_read_sql_query(query='SELECT * FROM', database=_database, table=_table, connection=sql)
    # print(data)
    # print(type(data))
    # for row in data:
    #     print(row)

    # # Read SQL data native
    # connection_native = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' + _server + ';DATABASE=' +
    #                                    _database + ';UID=' + _username + ';PWD=' + _password +
    #                                    ';Trusted_Connection=yes;')
    # sql_query_read = pd.read_sql_query(f'SELECT * FROM {_database}.dbo.{_table}', connection_native)
    # print(sql_query_read)
    # print(type(sql_query_read))
    # for row in sql_query_read:
    #     print(row)
    # print(type(row))

    # Write SQL data with Class
    # sql = SQLclass.pyodbc_connect(server=_server, database=_database, username=_username, password=_password)
    # SQLclass.execute_query(connection=sql, table=_table, rows=_rows, values=_values)

    # # Write SQL data native
    # connection_native = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' + _server + ';DATABASE=' +
    #                                    _database + ';UID=' + _username + ';PWD=' + _password +
    #                                    ';Trusted_Connection=yes;')
    # cursor = connection_native.cursor()
    # cursor.fast_executemany = True
    # count = cursor.execute(f"""INSERT INTO {_table} (id_row, device_row, percent_row, time_row, data_row, extra_row)
    # VALUES (id_row, device_row, percent_row, '{_time}', '{_date}', extra_row)""").rowcount
    # connection_native.commit()


class ExcelClass:
    @staticmethod
    def example(data: list):

        # Read
        file_name = 'data.xlsx'
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        global_arr = []
        for num in range(1, 5000):
            local_list = []
            for x in 'ABCDEFGHM':
                value = ExcelClass.get_sheet_value(x, num, sheet=sheet)
                if value == 'None' or value is None or value == '':
                    value = ''
                local_list.append(value)
            global_arr.append(local_list)
        workbook.close()

        # Write
        file_name = 'data.xlsx'
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        row_index = 0
        for row in data:
            row_index += 1
            col_index = 0
            for cell in row:
                col_index += 1
                ExcelClass.set_sheet_value(col=col_index, row=row_index, value=cell, sheet=sheet)
        workbook.close()

    @staticmethod
    def get_sheet_value(col: Union[str, int], row: int, sheet=None):
        if isinstance(col, int):
            col = ExcelClass.get_column_letter(col)
        return str(sheet[str(col) + str(row)].value)

    @staticmethod
    def set_sheet_value(col: Union[str, int], row: int, value: str, sheet=None):
        if isinstance(col, int):
            col = ExcelClass.get_column_letter(col)
        if value:
            sheet[str(col) + str(row)] = value
        else:
            sheet[str(col) + str(row)] = ''

    @staticmethod
    def get_column_letter(num: int):
        return get_column_letter(num)


class PathClass:
    export_file = 'file.txt'
    relative_path = os.path.dirname(os.path.abspath('__file__')) + '\\'
    export_path = relative_path + export_file


class ClassWorker:
    """
    Класс, который содержит в себе работника, со значениями по строке
    """

    def __init__(self, A_1='', B_1='', C_1='', D_1='', E_1='', F_1='', G_1='', H_1='', M_1=''):
        # Подразделение
        self.A_1 = A_1
        # Цех или Служба
        self.B_1 = B_1
        # Отдел или участок
        self.C_1 = C_1
        # Фамилия
        self.D_1 = D_1
        # Имя
        self.E_1 = E_1
        # Отчество
        self.F_1 = F_1
        # Табельный №
        self.G_1 = G_1
        # Категория
        self.H_1 = H_1
        # Пол
        self.M_1 = M_1

    def print_worker(self):
        """

        """
        print(f'{self.A_1}+{self.B_1}+{self.C_1}+{self.D_1}+{self.E_1}+{self.F_1}+{self.G_1}+{self.H_1}+{self.M_1}')

    def get_worker_value(self, index):
        """

        :param index:
        :return:
        """
        value_ = list((self.A_1, self.B_1, self.C_1, self.D_1, self.E_1, self.F_1, self.G_1, self.H_1, self.M_1))
        return value_[index]

    def get_worker_id(self):
        """

        :return:
        """
        return self.G_1


def input_integer_value(description):
    """

    :param description:
    :return:
    """
    while True:
        try:
            value_ = round(int(input(f'{description}')))
            if value_ > 0:
                break
        except:
            print('Ошибка, введите ещё раз.')
    return value_


if __name__ == '__main__':
    pass
