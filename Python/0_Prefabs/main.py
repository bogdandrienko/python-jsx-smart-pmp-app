import openpyxl
from openpyxl.utils import get_column_letter


def get_sheet_value(_column, _row, _sheet):
    """"
    Принимает: индексы колонки и строки для извлечения данных, а также лист откуда извлекать.
    Возвращает: значение, находящееся по индексам на нужном листе.
    """
    try:
        return _sheet[str(_column) + str(_row)].value
    except Exception as ex:
        print(ex)
        return ''


def set_sheet_value(_column, _row, _sheet, _value):
    """"
    Принимает: индекс колонки и строку для записи данных, а также лист откуда куда записывать и значение для записи.
    """
    try:
        int(_column)
        _column = get_column_letter(_column)
    except ValueError:
        pass
    try:
        sheet[f'{_column}{_row}'] = str(_value)
    except Exception as _ex:
        print(_ex)


file_xlsx = 'excel.xlsx'
workbook = openpyxl.load_workbook(file_xlsx)
sheet = workbook.active
old_arr = []
for row in range(3, 209+1):
    arr_local = []
    for column in "ABCJ":
        cell_vall = get_sheet_value(_column=column, _row=row, _sheet=sheet)
        arr_local.append(cell_vall)
    old_arr.append(arr_local)
workbook.close()
print(old_arr)

file_xlsx = 'new_excel.xlsx'
workbook = openpyxl.load_workbook(file_xlsx)
sheet = workbook.active
for row_1 in old_arr:
    for col_1 in row_1:
        set_sheet_value(_row=old_arr.index(row_1) + 2, _column=row_1.index(col_1) + 1, _sheet=sheet, _value=col_1)
workbook.save('new_excel.xlsx')

file_xlsx = 'new_excel.xlsx'
workbook = openpyxl.load_workbook(file_xlsx)
sheet = workbook.active
clear_arr = []
index = 0
for row in range(2, 2000):
    arr_local = []
    for column in "ABCD":
        cell_vall = get_sheet_value(_column=column, _row=row, _sheet=sheet)
        arr_local.append(cell_vall)
    if arr_local[0]:
        index += 1
        arr_local.append(index)
        clear_arr.append(arr_local)
workbook.close()
print(clear_arr)

file_xlsx = 'clear_excel.xlsx'
workbook = openpyxl.load_workbook(file_xlsx)
sheet = workbook.active
for row_1 in clear_arr:
    for col_1 in row_1:
        set_sheet_value(_row=clear_arr.index(row_1) + 2, _column=row_1.index(col_1) + 1, _sheet=sheet, _value=col_1)
workbook.save('clear_excel.xlsx')

# arr = ['A', 'C', 'D']
# for val in "BCD":
#     try:
#         arr.index(val)
#         print(f'Этот элемент: {val} уже есть, мы его не добавляем')
#         pass
#     except Exception as ex:
#         print(f'Этого элемента: {val} нет, мы его добавляем')
#         arr.append(val)
# print(arr)

# temp_1 = 'Управление предприятия (ЦК)'
# temp_2 = 'Отдел закупок'
#
# for row in arr:
#     # print(f"строка: {row}")
#     row_1 = row[0]
#     row_2 = row[1]
#     row_3 = row[2]
#     if row_1 == temp_1 and row_2 == temp_2:
#         for cell in row:
#             print(f"ячейка: {cell}")

# arr = []
# for row in range(2, 10):
#     arr_local = []
#     for column in range(1, 6):
#         cell_vall = get_sheet_value(_column=get_column_letter(column), _row=row, _sheet=sheet)
#         print(cell_vall)
#         arr_local.append(cell_vall)
#     arr.append(arr_local)
# # print(arr)
