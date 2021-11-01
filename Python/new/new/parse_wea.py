import openpyxl
from openpyxl.utils import get_column_letter
import requests
from bs4 import BeautifulSoup
import csv


def pandas(url):
    print(url)
    headers = {'user-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:72.0) Gecko/20100101 Firefox/72.0',
               'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'}
    # url = 'http://www.pogodaiklimat.ru/weather.php?id=35042&bday=1&fday=1&amonth=10&ayear=2021&bot=2'
    response = requests.get(url, headers=headers)
    response.encoding = 'utf-8'
    soup = BeautifulSoup(response.text, "html.parser")
    soup.encoded = 'utf-8'
    rows = soup.find('div', class_='archive-table-wrap')
    if rows is None or rows == '':
        rows = 'ТАКОГО ДНЯ НЕ СУЩЕСТВУЕТ'
    return rows


def set_sheet_value(column, row, value_, sheet):
    try:
        if value_:
            sheet[str(column) + str(row)] = value_
        else:
            sheet[str(column) + str(row)] = ''
    except:
        sheet[str(column) + str(row)] = ''


print('begin')
arr_url = []
for year in range(2011, 2022):
    for month in range(1, 13):
        for day in range(1, 32):
            url = f"http://www.pogodaiklimat.ru/weather.php?id=35042&bday={day}&fday={day}&amonth={month}&ayear={year}&bot=2"
            print(f"day={day} month={month} year={year}")
            arr_url.append([f"{url}", f"{day}.{month}.{year}"])
# print(arr_url[:100])

arr_requests = []
for url in arr_url[:3]:
    data = pandas(url[0])
    arr_requests.append([url[1], data])
# print(arr_requests)

_import_file = 'data_1.xlsx'
workbook = openpyxl.load_workbook(_import_file)
sheet = workbook.active
# row_index = 0
# for req in arr_requests:
#     row_index += 1
#     for one_req in req:
#         if req.index(one_req) == 0:
#             sheet['A' + str(row_index)] = str(one_req)
#         if req.index(one_req) == 1:
#             sheet['B' + str(row_index)] = str(one_req)


final_arr = []
for one_request in arr_requests:
    data = one_request[0]
    req = one_request[1].find_all('tr')
    # req - ТУТ ЛЕЖИТ ОДИН ОТВЕТ ОТ КАЖДОЙ ДАТЫ
    # print(req)
    temp_arr = []
    sum_temp = []
    for one_req in req[1:]:
        # one_req - ТУТ ЛЕЖИТ ОДНА СТРОКА ОТ КАЖДОЙ ТАБЛИЦЫ
        temp_temp = []
        try:
            temp = str(one_req).split('class="temp')[1].split('</nobr></td>')[0].split('nobr>')[1]
            # temp - ТУТ ЛЕЖИТ ТЕМПЕРАТУРА ДЛЯ КАЖДОЙ СТРОКИ
            temp_temp.append(float(temp))
            sum_temp.append(float(temp))
            print(f"{temp}")
            temp_arr.append([temp])
        except Exception as ex:
            print(ex)
            temp = -1000
    sum_ = 0.0
    for tem in sum_temp:
        print(f"temperature: {tem}")
        sum_ += tem
    sum_ = sum_ / len(sum_temp)

    print(f"summ temperature: {round(sum_, 1)}")

    #     try:
    #         pressure = str(one_req).split('class="temp_33">')[1].split('</td>')[0]
    #         print(f"{pressure} \n\n\n\n\n")
    #     except:
    #         pressure = -1000
    #     temp_arr.append([temp, pressure])
    str_temp = ''
    for t in sum_temp:
        str_temp += f", {t}"

    arr = [data, str_temp[2:], temp_arr[0][0]]
    final_arr.append(arr)


row_index = 0
for arr in final_arr:
    row_index += 1
    col_index = 0
    for one_req in arr:
        col_index += 1
        sheet[str(get_column_letter(col_index)) + str(row_index)] = one_req

        # print(f"{one_req} \n\n\n\n\n")
    # row_index = 0
    # col_index = 0
    # for row in req.find_all('td'):
    # for row in req.find_all('td'):
    #     print(row)

        # row_index += 1
        # col_index = 0
        # for col in row.find_all('td'):
        #     col_index += 1
        #     # print(f"COL: \n {col.text} \n")
        #     if col_index == 4 and row_index != 1:
        #         value = str(col).split('</td>')[0].split('>')[1].strip()
        #     elif col_index == 5 and row_index != 1:
        #         # try:
        #         #     value = str(col.split('</')[0])
        #         # except:
        #         #     value = str(col)
        #         value = str(col).split('<')[1].split('>')[1].strip()
        #         # print(str(col))
        #         print(value)
        #     else:
        #         value = str(col).split('</td>')[0].split('>')[1].strip()
        #
        #     # print(f"COL: \n {value} \n")
        #     try:
        #         sheet[str(get_column_letter(col_index)) + str(row_index)] = value
        #     except Exception as ex:
        #         # print(f"ERROR: {ex}")
        #         pass
workbook.save(_import_file)
workbook.close()
print('complete')

stroing = '-23.7, -21.2, -18.7, -16.7, -17.7, -19.4, -23.0'

vals = stroing.split(', ')
print(vals)

# url = 'http://www.pogodaiklimat.ru/weather.php?id=35042'
# response.encoding = 'utf-8'
# file = open('data.csv', 'w')
# writer = csv.writer(file)
# soup = BeautifulSoup(response.text, "html.parser")
# soup.encoded = 'utf-8'
# rows = soup.find('div', class_='archive-table-wrap').find_all('tr')
# for row in rows:
#     columns = row.find_all('td')
#     print(columns)
#     data_list = [columns[0].text, columns[1].text, columns[2].text, columns[3].text, columns[4].text, columns[5].text,
#                  columns[6].text, columns[7].text, columns[8].text, columns[9].text, columns[10].text, columns[11].text,
#                  columns[12].text, columns[13].text, columns[14].text, columns[15].text, columns[16].text]
#     writer.writerow(data_list)
# file.close()
