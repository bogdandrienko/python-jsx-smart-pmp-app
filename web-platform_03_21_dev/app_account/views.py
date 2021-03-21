from django.shortcuts import render, redirect
from django.contrib.auth.models import Group, User
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login, authenticate, logout, get_user_model
from .forms import SignUpForm, SignUpManyForm
from django.core.mail import BadHeaderError
import openpyxl


def login_account(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            username = request.POST['username']
            password = request.POST['password']
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('home')
            else:
                return redirect('app_account:create')
    else:
        form = AuthenticationForm()

    context = {
        'form': form,
    }
    return render(request, 'app_account/login.html', context)

def create_account(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            signup_user = User.objects.get(username=username)
            try:
                user_group = Group.objects.get(name='User')
            except:
                user_group = Group.objects.create(name='User')
            user_group.user_set.add(signup_user)
    else:
        form = SignUpForm()
    
    context = {
        'form': form,
    }
    return render(request, 'app_account/create.html', context)

def create_many_account(request):
    if request.method == 'POST':
        try:
            # читаем excel-файлы из запроса
            wb1 = openpyxl.load_workbook(request.FILES.get('document_addition_file_1'))
            # получаем активные листы
            sheet = wb1.active
            # функция получения значения по строке и колонне
            def get_value(rows:str, cols:int):
                value = str(sheet[rows+str(cols)].value)
                return value
            # Выбираем количество строк для цикла
            max = sheet.max_row
            # max_column = sheet.max_column
            # создаём цикл для записи в модель
            for i in range(2, max + 1):
                username    = get_value('A', i)
                password    = get_value('B', i)
                email       = get_value('C', i)
                first_name  = get_value('D', i)
                last_name   = get_value('E', i)
                group       = get_value('F', i)
                if get_value('G', i) == 'да' or get_value('G', i) == 'Да':
                    is_staff = True
                else:
                    is_staff = False
                if username and password and first_name and last_name:
                    try:
                        UserModel   = get_user_model()
                        user        = UserModel.objects.create_user(
                            username    = username,
                            password    = password,
                            email       = email,
                            first_name  = first_name,
                            last_name   = last_name,
                            is_staff    = is_staff,
                        )
                        try:
                            user_group = Group.objects.get(name=group)
                        except:
                            user_group = Group.objects.create(name=group)
                        signup_user = User.objects.get(username=user.username)
                        user_group.user_set.add(signup_user)
                    except BadHeaderError:
                        return redirect('app_account:create_many')
        except:
            return redirect('app_account:create_many')
    form = SignUpManyForm()
    context = {
        'form': form
    }
    return render(request, 'app_account/create_many.html', context)

def logout_account(request):
    # Проверка регистрации: если пользователь не вошёл в аккаунт, действия не срабатают, а его переадресует в форму входа
    if request.user.is_authenticated is not True:
        return redirect('app_account:login')
    # Переадресация пользователя на страницу входа

    logout(request)
    return redirect('app_account:login')
