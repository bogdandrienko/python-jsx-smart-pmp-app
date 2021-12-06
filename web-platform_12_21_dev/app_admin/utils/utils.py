import asyncio
import concurrent
import datetime
import json
import aiohttp
import multiprocessing
import os
import smtplib
import sys
import threading
import time

import chardet
import colorama
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor
from functools import wraps

from typing import Union

import cv2
import httplib2
import numpy
import numpy as np
import openpyxl
import pyttsx3
import requests
import psycopg2 as pg
from bs4 import BeautifulSoup
from gmplot import gmplot
from matplotlib import pyplot as plt, cm
from matplotlib.ticker import LinearLocator
from mpl_toolkits.mplot3d import axes3d
from openpyxl.utils import get_column_letter
from django.utils import timezone
from app_admin.models import ComputerVisionModuleModel


class ComputerVisionClass:
    class SoloCamSettingsComputerVisionClassExample:
        """
        Настройки одной камеры для модуля: грохота 16 операции, 10 отметка
        """

        def __init__(self, ip: str, correct: float):
            # Ip address
            self.ip = ip
            # Коэффициент корректировки
            self.correct = round(correct, 3)

    class EventLoopClass:
        @staticmethod
        def loop_modules_global(tick_delay=1.0, module_re_start_delay=5.0):
            """
            Цикл прохода по всем модулям-циклам
            """
            # try:
            if True:
                print('global tick')
                dict_modules = {str(x[0]).strip(): str(x[1]).strip() for x in
                                ComputerVisionModuleModel.get_all_modules()}
                for module in ComputerVisionModuleModel.objects.filter(play_boolean_field=True):
                    if module.play_boolean_field:
                        # Проверка, что цикл модуля отработал в последний раз не позже "tick_delay" секунд
                        print(module.datetime_field + datetime.timedelta(hours=6))
                        print(DateTimeUtils.get_difference_datetime(seconds=module_re_start_delay))
                        if module.datetime_field and (module.datetime_field + datetime.timedelta(hours=6)) > \
                                DateTimeUtils.get_difference_datetime(seconds=-module_re_start_delay):
                            play_module = False
                        else:
                            play_module = True
                        print(play_module)
                    else:
                        continue
                    if play_module:
                        # Получение имени модуля из константы в модели ComputerVisionModuleModel
                        name_module = dict_modules[module.path_slug_field]
                        if name_module == 'Грохота, 16 операция, 26 отметка':
                            threading.Thread(
                                target=ComputerVisionClass.EventLoopClass.
                                    Grohota16OperationClass.loop_module_grohota_16_operation(module),
                                args=([])
                            ).start()
                        elif name_module == 'Грохота, 26 операция, ?? отметка':
                            # threading.Thread(
                            #     target=ComputerVisionClass.EventLoopClass.
                            #         Grohota26OperationClass.loop_module_grohota_26_operation(),
                            #     args=([])
                            # ).start()
                            pass
                    else:
                        continue
                time.sleep(tick_delay)
                ComputerVisionClass.EventLoopClass.loop_modules_global()
            # except Exception as error:
            #     print(error)
            #     time.sleep(1)
            #     ComputerVisionClass.EventLoopClass.loop_modules_global()

        class Grohota16OperationClass:
            @staticmethod
            def loop_module_grohota_16_operation(module):
                """
                Цикл прохода по компонентам-функциям внутри модуля: Грохота, 16 операция, 10 отметка
                """
                try:
                    # Начальное время
                    start_time = time.time()
                    # Данные из базы с настройками по каждому компоненту
                    settings_components = [
                        {
                            'ip': '192.168.15.202',
                            'correct': 1.2
                        },
                        {
                            'ip': '192.168.15.203',
                            'correct': 1.3
                        },
                        {
                            'ip': '192.168.15.204',
                            'correct': 1.4
                        },
                        {
                            'ip': '192.168.15.205',
                            'correct': 1.5
                        },
                        {
                            'ip': '192.168.15.206',
                            'correct': 1.6
                        },
                        {
                            'ip': '192.168.15.207',
                            'correct': 1.7
                        },
                        {
                            'ip': '192.168.15.208',
                            'correct': 1.8
                        },
                        {
                            'ip': '192.168.15.209',
                            'correct': 1.9
                        },
                        {
                            'ip': '192.168.15.210',
                            'correct': 2.0
                        },
                        {
                            'ip': '192.168.15.211',
                            'correct': 2.1
                        },
                    ]
                    error_text_field = None
                    with ThreadPoolExecutor() as executor:
                        try:
                            futures = []
                            # Цикл для прохода по настройкам и запуску компонентов
                            for settings in settings_components:
                                futures.append(
                                    executor.submit(
                                        ComputerVisionClass.EventLoopClass.
                                            Grohota16OperationClass.component_grohota_16_operation,
                                        settings=settings
                                    )
                                )
                            for future in concurrent.futures.as_completed(futures):
                                value = f"{float(future.result()):0.3f}%"
                                print(value)
                        except Exception as error:
                            error_text_field = f'error : {error}'
                            pass

                    module.duration_float_field = round(time.time() - start_time, 4)
                    module.datetime_field = timezone.now()
                    if error_text_field:
                        module.error_text_field = error_text_field
                    module.save()
                    time.sleep(1.0)
                    ComputerVisionClass.EventLoopClass.\
                        Grohota16OperationClass.loop_module_grohota_16_operation(module=module)
                except Exception as error:
                    print(error)

            @staticmethod
            def component_grohota_16_operation(settings: dict):
                """
                Компонент-функция расчёта % схода на грохотах 16 операции, 10 отметка
                """
                try:
                    sources = f'http://{str(settings["ip"])}:80' \
                              f'/ISAPI/Streaming/channels/101/picture?snapShotImageType=JPEG'
                    login = 'admin'
                    password = 'q1234567'
                    temp_path = DirPathFolderPathClass.create_folder_in_this_dir(
                        folder_name='static/media/data/computer_vision/temp'
                    )
                    h = httplib2.Http(temp_path)
                    h.add_credentials(login, password)
                    response, content = h.request(sources)
                    image = cv2.imdecode(numpy.frombuffer(content, numpy.uint8), cv2.IMREAD_COLOR)
                    image = cv2.resize(image, (750, 500), interpolation=cv2.INTER_AREA)
                    mask = cv2.imread('static/media/data/computer_vision/mask/m_16_8.jpg', 0)
                    mask = cv2.resize(mask, (750, 500), interpolation=cv2.INTER_AREA)
                    bitwise_and = cv2.bitwise_and(image, image, mask=mask)
                    cvtcolor = cv2.cvtColor(bitwise_and, cv2.COLOR_BGR2HSV)
                    inrange = cv2.inRange(cvtcolor, numpy.array([0, 0, 255 - 120], dtype=numpy.uint8),
                                          numpy.array([255, 120, 255], dtype=numpy.uint8))
                    value = numpy.sum(inrange > 0) / numpy.sum(mask > 0) * 100 * float(settings['correct'])
                    return round(value, 3)
                except Exception as error:
                    return None

        class Grohota26OperationClass:
            @staticmethod
            def loop_module_grohota_26_operation():
                """
                Цикл прохода по компонентам-функциям внутри модуля: Грохота, 26 операция, ?? отметка
                """
                pass

            @staticmethod
            def component_grohota_26_operation():
                """
                Компонент-функция расчёта % схода на грохотах 26 операции, ?? отметка
                """
                return None

    @staticmethod
    def example_analyse():

        def analyse_image_open_cv(ip=203):
            # Начальное время
            # start_time = time.time()
            # print('start')

            sources = f'http://192.168.15.{ip}:80/ISAPI/Streaming/channels/101/picture?snapShotImageType=JPEG'
            login = 'admin'
            password = 'q1234567'
            h = httplib2.Http(os.path.abspath('__file__'))
            h.add_credentials(login, password)
            response, content = h.request(sources)
            image = cv2.imdecode(numpy.frombuffer(content, numpy.uint8), cv2.IMREAD_COLOR)
            image = cv2.resize(image, (750, 500), interpolation=cv2.INTER_AREA)
            mask = cv2.imread('static/media/data/computer_vision/mask/m_16_8.jpg', 0)
            mask = cv2.resize(mask, (750, 500), interpolation=cv2.INTER_AREA)
            bitwise_and = cv2.bitwise_and(image, image, mask=mask)
            cvtcolor = cv2.cvtColor(bitwise_and, cv2.COLOR_BGR2HSV)
            inrange = cv2.inRange(cvtcolor, numpy.array([0, 0, 255 - 120], dtype=numpy.uint8),
                                  numpy.array([255, 120, 255], dtype=numpy.uint8))
            value = f"{numpy.sum(inrange > 0) / numpy.sum(mask > 0) * 100 * 1.0:0.2f}%"

            # Финальное время
            # print(f"Final time: {round(time.time() - start_time, 1)}")
            # print('end')

            # print(value)
            return value

        def analyse():

            list_ip = []
            for num in range(202, 211 + 1):
                list_ip.append(num)
            print(list_ip)

            index = 0
            while index < 100:
                index += 1
                # Начальное время
                start_time = time.time()
                print('start')
                # Менеджер контекста для многопотока под ThreadPoolExecutor
                with ThreadPoolExecutor() as executor:
                    futures = []
                    # Цикл для прохода по ссылкам
                    for ip in list_ip:
                        futures.append(executor.submit(analyse_image_open_cv, ip=ip))
                    for future in concurrent.futures.as_completed(futures):
                        print(future.result())

                # Финальное время
                print(f"Final time: {round(time.time() - start_time, 1)}")
                print('end')
                time.sleep(1)

        threading.Thread(target=analyse, args=([])).start()


class DateTimeUtils:
    @staticmethod
    def get_current_datetime():
        return f"{time.strftime('%Y-%m-%d %H:%M:%S')}"

    @staticmethod
    def get_current_date():
        return f"{time.strftime('%Y-%m-%d')}"

    @staticmethod
    def get_current_time():
        return f"{time.strftime('%H:%M:%S')}"

    @staticmethod
    def get_difference_datetime(days=0.0, hours=0.0, minutes=0.0, seconds=0.0):
        value = datetime.datetime.now() + datetime.timedelta(days=days, hours=hours, minutes=minutes, seconds=seconds)
        return value.replace(tzinfo=datetime.timezone.utc)

    class Example:
        @staticmethod
        def example_get_datetime():
            print(DateTimeUtils.get_current_datetime())

        @staticmethod
        def example_get_difference_datetime():
            print(DateTimeUtils.get_difference_datetime(hours=-2))


class FileReadWriteClass:
    @staticmethod
    def write_to_file(file_name, text_file, mode='a'):
        with open(file_name, mode=mode) as file:
            file.write(text_file)

    @staticmethod
    def read_from_file_lines(file_name, mode='r'):
        with open(file_name, mode=mode) as file:
            return file.readlines()

    @staticmethod
    def read_from_file(file_name, mode='r'):
        with open(file_name, mode=mode) as file:
            return file.read()

    @staticmethod
    def example_write_to_file_add():
        FileReadWriteClass.write_to_file(file_name='file.txt', text_file='text\n', mode='a')

    @staticmethod
    def example_write_to_file_rewrite():
        FileReadWriteClass.write_to_file(file_name='file.txt', text_file='text\n', mode='w')

    @staticmethod
    def example_read_from_file_lines():
        print(FileReadWriteClass.read_from_file_lines(file_name='file.txt', mode='r'))

    @staticmethod
    def example_read_from_file():
        print(FileReadWriteClass.read_from_file(file_name='file.txt', mode='r'))


class LoggingClass:
    @staticmethod
    def logging_to_local_file(logging_text, error_func, file_name='logging.txt', type_write='a'):
        try:
            with open(file_name, type_write) as file:
                file.write(f'\n{DateTimeUtils.get_current_datetime()} | error_func: {error_func} | {logging_text}\n')
        except Exception as error:
            print(f'\n{time.strftime("%x %X")} | error_func: {error_func} | {error}\n')

    class Example:
        @staticmethod
        def example_logging():
            LoggingClass.logging_to_local_file(
                logging_text='example',
                error_func='LoggingClass.Example.example_logging',
                file_name='logging.txt',
                type_write='a'
            )


class ExcelClass:
    @staticmethod
    def workbook_create():
        workbook = openpyxl.Workbook()
        return workbook

    @staticmethod
    def workbook_load(excel_file: str):
        workbook = openpyxl.load_workbook(excel_file)
        return workbook

    @staticmethod
    def workbook_activate(workbook):
        sheet = workbook.active
        return sheet

    @staticmethod
    def workbook_save(workbook, excel_file: str):
        try:
            os.remove(excel_file)
        except Exception as error:
            pass
        try:
            workbook.save(excel_file)
        except Exception as error:
            print(f'Please, close the excel_file: {excel_file} | {error}')

    @staticmethod
    def workbook_close(workbook):
        openpyxl.Workbook.close(workbook)

    @staticmethod
    def set_sheet_title(sheet, page_name='page 1'):
        sheet.title = page_name

    @staticmethod
    def get_sheet_value(col: Union[str, int], row: int, sheet):
        if isinstance(col, int):
            col = ExcelClass.get_column_letter(col)
        value = str(sheet[str(col).upper() + str(row)].value)
        if value == 'None' or value is None:
            value = ''
        else:
            value = str(value)
        return value

    @staticmethod
    def set_sheet_value(col: Union[str, int], row: int, value: str, sheet):
        if isinstance(col, int):
            col = ExcelClass.get_column_letter(col)
        if value == 'None' or value is None:
            value = ''
        sheet[str(col) + str(row)] = str(value)

    @staticmethod
    def get_column_letter(num: int):
        return get_column_letter(num)

    @staticmethod
    def get_max_num_rows(sheet):
        return int(sheet.max_row)

    class Example:
        @staticmethod
        def example_read_from_excel_file_col_int():
            excel_file = 'export.xlsx'
            workbook = ExcelClass.workbook_load(excel_file=excel_file)
            sheet = ExcelClass.workbook_activate(workbook=workbook)
            max_num_rows = ExcelClass.get_max_num_rows(sheet=sheet)
            max_num_cols = 10
            global_list = []
            for row in range(1, max_num_rows + 1):
                local_list = []
                for col in range(1, max_num_cols + 1):
                    value = ExcelClass.get_sheet_value(col=col, row=row, sheet=sheet)
                    local_list.append(value)
                global_list.append(local_list)
            for row in global_list:
                print(row)
            ExcelClass.workbook_close(workbook=workbook)

        @staticmethod
        def example_read_from_excel_file_col_char():
            excel_file = 'export.xlsx'
            workbook = ExcelClass.workbook_load(excel_file=excel_file)
            sheet = ExcelClass.workbook_activate(workbook=workbook)
            max_num_rows = ExcelClass.get_max_num_rows(sheet=sheet)
            char_cols = 'ACDF'
            global_list = []
            for row in range(1, max_num_rows + 1):
                local_list = []
                for col in char_cols:
                    value = ExcelClass.get_sheet_value(col=col, row=row, sheet=sheet)
                    local_list.append(value)
                global_list.append(local_list)
            for row in global_list:
                print(row)
            ExcelClass.workbook_close(workbook=workbook)

        @staticmethod
        def example_write_to_excel_file():
            global_list = [
                ['title_1', 'title_2', 'title_3'],
                ['body_1_1', 'body_1_2', 'body_1_3'],
                ['body_2_1', 'body_2_2', 'body_2_3'],
                ['body_3_1', 'body_3_3', 'body_3_3'],
            ]

            excel_file = 'import.xlsx'
            workbook = ExcelClass.workbook_create()
            sheet = ExcelClass.workbook_activate(workbook=workbook)
            for row in global_list:
                for value in row:
                    ExcelClass.set_sheet_value(
                        col=row.index(value) + 1,
                        row=global_list.index(row) + 1,
                        value=value,
                        sheet=sheet
                    )
            ExcelClass.workbook_save(workbook=workbook, excel_file=excel_file)


class DirPathFolderPathClass:
    @staticmethod
    def create_folder_in_this_dir(folder_name='new_folder', current_path=os.path.dirname(os.path.abspath('__file__'))):
        full_path = current_path + f'\\{folder_name}'
        try:
            os.makedirs(full_path)
        except Exception as error:
            # print(f'directory already yet | {error}')
            pass
        finally:
            return full_path

    @staticmethod
    def get_all_files_in_path(path=os.path.dirname(os.path.abspath('__file__'))):
        files_list = []
        for root, dirs, files in os.walk(path, topdown=True):
            for name in files:
                files_list.append(f"{os.path.join(root, name)}")
        return files_list

    @staticmethod
    def get_all_dirs_in_path(path=os.path.dirname(os.path.abspath('__file__'))):
        directories_list = []
        for root, dirs, files in os.walk(path, topdown=True):
            for name in dirs:
                directories_list.append(f"{os.path.join(root, name)}")
        return directories_list

    class Example:
        @staticmethod
        def example_create_folder_in_this_folder():
            path = DirPathFolderPathClass.create_folder_in_this_dir(folder_name='new_folder')
            print(path)

        @staticmethod
        def example_create_folder_in_folder_in_this_folder():
            path = DirPathFolderPathClass.create_folder_in_this_dir(folder_name='new_folder\\new')
            print(path)

        @staticmethod
        def example_create_folder_in_external_folder():
            path = DirPathFolderPathClass.create_folder_in_this_dir(folder_name='new_folder\\new', current_path='C:\\')
            print(path)

        @staticmethod
        def example_get_all_files_in_path():
            files_list = DirPathFolderPathClass.get_all_files_in_path(
                path=r'C:\Project\Github_Projects\python-jsx-smart-pmp-app\web-platform_12_21_dev\app_admin'
            )
            for file in files_list:
                print(file)

        @staticmethod
        def example_get_all_folders_in_path():
            path_list = DirPathFolderPathClass.get_all_dirs_in_path(
                path=r'C:\Project\Github_Projects\python-jsx-smart-pmp-app\web-platform_12_21_dev\app_admin'
            )
            for path in path_list:
                print(path)


class SyncAsyncThreadingPoolExecutorClass:
    class Example:
        @staticmethod
        # Синхронный код
        def example_sync_compute():
            # Начальное время
            start_time = time.time()
            print('start')
            # Генерация ссылок
            page_urls = ["https://en.wikipedia.org/wiki/" + str(i) for i in range(10)]

            # Функция получения текста из ссылки
            def get_page_from_url(url):
                try:
                    # Синхронная библиотека
                    response = requests.get(url=url, timeout=3)
                    value = f'{page_urls.index(url) + 1}: {response.text[0:15]}'
                    print(value)
                    return value
                except Exception as error:
                    get_page_from_url(url=url)

            responses_list = []
            # Цикл для прохода по ссылкам
            for page_url in page_urls:
                resp = get_page_from_url(url=page_url)
                responses_list.append(resp)
            for response_from_list in responses_list:
                print(response_from_list)

            # Финальное время
            print(f"Final time: {round(time.time() - start_time, 1)}")
            print('end')

        @staticmethod
        # Acинхронный код
        def example_async_compute():
            # Начальное время
            start_time = time.time()
            print('start')
            # Генерация ссылок
            page_urls = ["https://en.wikipedia.org/wiki/" + str(i) for i in range(10)]

            # Функция получения текста из ссылки
            async def get_page_from_url(session, url):
                try:
                    # Асинхронная библиотека
                    async with session.get(url) as resp:
                        response = await resp.text()
                    value = f'{page_urls.index(url) + 1}: {response[0:15]}'
                    print(value)
                    return value
                except Exception as error:
                    await get_page_from_url(session=session, url=url)

            # Цикл для прохода по ссылкам
            async def main():
                # Асинхронная библиотека
                async with aiohttp.ClientSession() as session:
                    tasks = []
                    for page_url in page_urls:
                        tasks.append(
                            asyncio.ensure_future(
                                get_page_from_url(
                                    session=session,
                                    url=page_url
                                )
                            )
                        )

                    responses_list = await asyncio.gather(*tasks)

                    for response_from_list in responses_list:
                        print(response_from_list)

            asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
            asyncio.run(main())

            # Финальное время
            print(f"Final time: {round(time.time() - start_time, 1)}")
            print('end')

        @staticmethod
        # Многопоточный Threading код
        def example_threading_compute():
            # Начальное время
            start_time = time.time()
            print('start')
            # Генерация ссылок
            page_urls = ["https://en.wikipedia.org/wiki/" + str(i) for i in range(10)]

            # Функция получения текста из ссылки
            def get_page_from_url(page_url):
                try:
                    # Синхронная библиотека
                    response = requests.get(url=page_url, timeout=3)
                    value = f'{page_urls.index(page_url) + 1}: {response.text[0:15]}'
                    print(value)
                    return value
                except Exception as error:
                    get_page_from_url(page_url=page_url)

            # Цикл для прохода по ссылкам
            for url in page_urls:
                threading.Thread(target=get_page_from_url, args=([url])).start()

            # Финальное время
            print(f"Final time: {round(time.time() - start_time, 1)}")
            print('end')

        @staticmethod
        # Многопоточный ThreadPoolExecutor код
        def example_thread_pool_executor_compute():
            # Начальное время
            start_time = time.time()
            print('start')
            # Генерация ссылок
            page_urls = ["https://en.wikipedia.org/wiki/" + str(i) for i in range(10)]

            # Функция получения текста из ссылки
            def get_page_from_url(page_url):
                try:
                    # Синхронная библиотека
                    response = requests.get(url=page_url, timeout=3)
                    value = f'{page_urls.index(page_url) + 1}: {response.text[0:15]}'
                    print(value)
                    return value
                except Exception as error:
                    get_page_from_url(page_url=page_url)

            # Менеджер контекста для многопотока под ThreadPoolExecutor
            with ThreadPoolExecutor() as executor:
                futures = []
                # Цикл для прохода по ссылкам
                for url in page_urls:
                    futures.append(executor.submit(get_page_from_url, page_url=url))
                for future in concurrent.futures.as_completed(futures):
                    print(future.result())

            # Финальное время
            print(f"Final time: {round(time.time() - start_time, 1)}")
            print('end')

        @staticmethod
        # Функция получения текста из ссылки
        def get_page_from_url_sync(page_url):
            try:
                # Синхронная библиотека
                response = requests.get(url=page_url, timeout=3)
                value = f'{response.text[0:15]}'
                print(value)
                return value
            except Exception as error:
                SyncAsyncThreadingPoolExecutorClass.Example.get_page_from_url_sync(page_url=page_url)

        @staticmethod
        # Мультипоточный ProcessPoolExecutor код
        def example_process_pool_executor_compute():
            # Начальное время
            start_time = time.time()
            print('start')
            # Генерация ссылок
            page_urls = ["https://en.wikipedia.org/wiki/" + str(i) for i in range(10)]

            # Менеджер контекста для многопотока под ThreadPoolExecutor
            with ProcessPoolExecutor() as executor:
                futures = []
                # Цикл для прохода по ссылкам
                for url in page_urls:
                    futures.append(executor.submit(
                        SyncAsyncThreadingPoolExecutorClass.Example.get_page_from_url_sync,
                        page_url=url
                    ))
                for future in concurrent.futures.as_completed(futures):
                    print(future.result())

            # Финальное время
            print(f"Final time: {round(time.time() - start_time, 1)}")
            print('end')


class EncryptingClass:
    @staticmethod
    def encrypt_text(text: str, hash_chars: str):
        chars = 'abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890' + ' _-:' + hash_chars
        chars = set(chars)
        forward_ord_list = []
        for char in chars:
            forward_ord_list.append(ord(char))
        forward_ord_list.sort(reverse=False)
        reverse_ord_list = forward_ord_list.copy()
        reverse_ord_list.sort(reverse=True)
        forward_dictinary = {reverse_ord_list[forward_ord_list.index(x)]: x for x in forward_ord_list}
        return ''.join([chr(forward_dictinary[ord(x)]) for x in text])

    @staticmethod
    def decrypt_text(text: str, hash_chars: str):
        chars = 'abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890' + ' _-:' + hash_chars
        chars = set(chars)
        forward_ord_list = []
        for char in chars:
            forward_ord_list.append(ord(char))
        forward_ord_list.sort(reverse=False)
        reverse_ord_list = forward_ord_list.copy()
        reverse_ord_list.sort(reverse=True)
        reverse_dictinary = {forward_ord_list[reverse_ord_list.index(x)]: x for x in reverse_ord_list}
        return ''.join([chr(reverse_dictinary[ord(x)]) for x in text])

    class Example:
        @staticmethod
        def example_encrypt_text():
            value = EncryptingClass.encrypt_text(text='12345', hash_chars='321')
            print(value)

        @staticmethod
        def example_decrypt_text():
            value = EncryptingClass.decrypt_text(text='wvuts', hash_chars='321')
            print(value)


class TypeVariablesClass:
    @staticmethod
    def example_get_type_of_variable(variable):
        return type(variable)

    @staticmethod
    def example_check_type_of_variable(variable, type_variable: str):
        if type_variable == 'bool':
            if isinstance(variable, bool):
                return True
            else:
                return False
        elif type_variable == 'int':
            if isinstance(variable, int):
                return True
            else:
                return False
        elif type_variable == 'float':
            if isinstance(variable, float):
                return True
            else:
                return False
        elif type_variable == 'str' or type_variable == 'string':
            if isinstance(variable, str):
                return True
            else:
                return False
        elif type_variable == 'list':
            if isinstance(variable, list):
                return True
            else:
                return False
        elif type_variable == 'dict' or type_variable == 'dictionary':
            if isinstance(variable, dict):
                return True
            else:
                return False
        elif type_variable == 'tuple':
            if isinstance(variable, tuple):
                return True
            else:
                return False
        elif type_variable == 'set':
            if isinstance(variable, set):
                return True
            else:
                return False
        else:
            return None

    class Example:
        @staticmethod
        def example_get_type():
            value = TypeVariablesClass.example_get_type_of_variable(variable=12.5)
            print(value)

        @staticmethod
        def example_check_type():
            value = TypeVariablesClass.example_check_type_of_variable(variable=12.5, type_variable='int')
            print(value)


class JsonClass:
    @staticmethod
    def write_json_to_file(dictionary: dict, file_name="json"):
        with open(f'{file_name}.json', 'w') as file:
            json.dump(dictionary, file)

    @staticmethod
    def read_json_from_file(file_name="json"):
        with open(f'{file_name}.json', 'r') as file:
            return json.load(file)

    class Example:
        @staticmethod
        def example_write_json_to_file():
            data = {
                'key_1': 'value_1',
                'key_2': 3,
                'key_3': {
                    'key_3_1': True,
                    'key_3_2': 'value_3_2',
                }
            }
            JsonClass.write_json_to_file(dictionary=data, file_name='json')

        @staticmethod
        def example_read_json_to_file():
            dictionary = JsonClass.read_json_from_file(file_name='json')
            print(dictionary)
            print(dictionary['key_1'])


class DictionaryClass:
    @staticmethod
    def get_all_keys(dictionary: dict):
        return dictionary.keys()

    @staticmethod
    def get_all_values(dictionary: dict):
        return dictionary.values()

    @staticmethod
    def get_all_sources(dictionary: dict, values: dict):
        dictionary_local = dictionary.copy()
        for key, value in values.items():
            dictionary_local[key] = value
        return dictionary_local


class EncodingClass:
    @staticmethod
    def find_encoding(text: Union[str, bytes]):
        if isinstance(text, str):
            text = text.encode()
        return chardet.detect(text)['encoding']

    @staticmethod
    def convert_encoding(text: Union[str, bytes], encoding=None, new_encoding='utf-8'):
        if encoding is None:
            encoding = EncodingClass.find_encoding(text)
        if isinstance(text, str):
            text = text.encode(encoding=encoding)
        return text.decode(encoding=new_encoding)

    class Example:
        @staticmethod
        def example_find_encoding():
            text = 'РђР»РµРєСЃР°РЅРґСЂР° РџСЂРѕРєРѕС„СЊРµРІР°'
            encoding = EncodingClass.find_encoding(text=text.encode(encoding='utf-8'))
            print(encoding)

        @staticmethod
        def example_convert_encoding():
            text = '\xd0\xa0\xd1\x92\xd0\xa0\xc2'
            text = 'РџСЂРѕРєРѕС„СЊРµРІР°'
            encoding = EncodingClass.find_encoding(text=text)
            text = EncodingClass.convert_encoding(text=text, encoding='Windows-1251', new_encoding='utf-8')
            print(text)


class ParserBeautifulSoupClass:
    @staticmethod
    def get_url():
        pass

    @staticmethod
    def parse_local_html():
        pass

    class Example:
        @staticmethod
        def example_parse_weather():
            arr_url = []
            for year in range(2011, 2012):
                for month in range(1, 2):
                    for day in range(1, 32):
                        url = f"http://www.pogodaiklimat.ru/weather.php?id=35042&b" \
                              f"day={day}&fday={day}&amonth={month}&ayear={year}&bot=2"
                        arr_url.append(url)

            arr_responces = []
            for url in arr_url:
                headers = {'user-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:72.0) Gecko/20100101 Firefox/72.0',
                           'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'}
                response = requests.get(url, headers=headers)
                response.encoding = 'utf-8'
                soup = BeautifulSoup(response.text, "html.parser")
                soup.encoded = 'utf-8'
                rows = soup.find('div', class_='archive-table-wrap')
                arr_responces.append(rows)

            all_data = []
            for request in arr_responces:
                data_list = []
                for row in request.find_all('tr'):
                    local_data = []
                    for col in row.find_all('td'):
                        local_data.append(col.text)
                    data_list.append(local_data)
                all_data.append([''])
                all_data.append(data_list)
            for data in all_data:
                print(data)

        @staticmethod
        def example_perser_exams_kiney():
            version = 3
            with open(f'{version}_data.html', 'r+', encoding='utf-8') as file:
                questions = []
                index = 1
                for lines in file:
                    c1 = f"""
                    <div id="content_test"><br><div class="workspace-header"><i>Самопроверка промежуточного рейтинга 1</i></div><p><b class="question" id="1">Сұрақ № 1 / Вопрос № 1 </b>АСУ из регулятора и объекта управления, называется:</p><p><input type="radio" name="variant" value="1" id="ans1"><label for="ans1">автоматической системой регулирования АСР</label></p><p><input type="radio" name="variant" value="2" id="ans2"><label for="ans2">дистанционным управлением</label></p><p><input type="radio" name="variant" value="3" id="ans3"><label for="ans3">автоматической системой управления АСУ</label></p><p><input type="radio" name="variant" value="4" id="ans4"><label for="ans4">автоматическим управлением</label></p><p><input type="radio" name="variant" value="5" id="ans5"><label for="ans5">комплексной автоматизацией</label></p><button class="btn btn-number" number="1" style="width:40px;margin:5px;border: 2px solid #e34761;">1</button><button class="btn btn-number" number="2" style="width:40px;margin:5px;">2</button><button class="btn btn-number" number="3" style="width:40px;margin:5px;">3</button><button class="btn btn-number" number="4" style="width:40px;margin:5px;">4</button><button class="btn btn-number" number="5" style="width:40px;margin:5px;">5</button><button class="btn btn-number" number="6" style="width:40px;margin:5px;">6</button><button class="btn btn-number" number="7" style="width:40px;margin:5px;">7</button><button class="btn btn-number" number="8" style="width:40px;margin:5px;">8</button><button class="btn btn-number" number="9" style="width:40px;margin:5px;">9</button><button class="btn btn-number" number="10" style="width:40px;margin:5px;">10</button><button class="btn btn-number" number="11" style="width:40px;margin:5px;">11</button><button class="btn btn-number" number="12" style="width:40px;margin:5px;">12</button><button class="btn btn-number" number="13" style="width:40px;margin:5px;">13</button><button class="btn btn-number" number="14" style="width:40px;margin:5px;">14</button><button class="btn btn-number" number="15" style="width:40px;margin:5px;">15</button><button class="btn btn-number" number="16" style="width:40px;margin:5px;">16</button><button class="btn btn-number" number="17" style="width:40px;margin:5px;">17</button><button class="btn btn-number" number="18" style="width:40px;margin:5px;">18</button><button class="btn btn-number" number="19" style="width:40px;margin:5px;">19</button><button class="btn btn-number" number="20" style="width:40px;margin:5px;">20</button><button class="btn btn-number" number="21" style="width:40px;margin:5px;">21</button><button class="btn btn-number" number="22" style="width:40px;margin:5px;">22</button><button class="btn btn-number" number="23" style="width:40px;margin:5px;">23</button><button class="btn btn-number" number="24" style="width:40px;margin:5px;">24</button><button class="btn btn-number" number="25" style="width:40px;margin:5px;">25</button><button class="btn btn-number" number="26" style="width:40px;margin:5px;">26</button><button class="btn btn-number" number="27" style="width:40px;margin:5px;">27</button><button class="btn btn-number" number="28" style="width:40px;margin:5px;">28</button><button class="btn btn-number" number="29" style="width:40px;margin:5px;">29</button><button class="btn btn-number" number="30" style="width:40px;margin:5px;">30</button><button class="btn btn-number" number="31" style="width:40px;margin:5px;">31</button><button class="btn btn-number" number="32" style="width:40px;margin:5px;">32</button><button class="btn btn-number" number="33" style="width:40px;margin:5px;">33</button><button class="btn btn-number" number="34" style="width:40px;margin:5px;">34</button><button class="btn btn-number" number="35" style="width:40px;margin:5px;">35</button><button class="btn btn-number" number="36" style="width:40px;margin:5px;">36</button><button class="btn btn-number" number="37" style="width:40px;margin:5px;">37</button><button class="btn btn-number" number="38" style="width:40px;margin:5px;">38</button><button class="btn btn-number" number="39" style="width:40px;margin:5px;">39</button><button class="btn btn-number" number="40" style="width:40px;margin:5px;">40</button><hr></div>
<div id="content_test"><br><div class="w
                    """
                    final_data = f"<strong>{index}</strong><br><hr><br>"
                    data = lines.split(" </b>")[1].split("<button class=")[0]
                    try:
                        new_data = data.split('<img src="')
                        if len(new_data) > 2:
                            final_data += data
                        else:
                            final_data += data.split('<img src="')[0] + """<img src="https://sdo.kineu.kz""" + \
                                          data.split('<img src="')[1]
                    except Exception as ex:
                        final_data += data
                    final_data += "<br><hr><br>"
                    reverse = False
                    for question in questions:
                        if question.split("</strong>")[1] == final_data.split("</strong>")[1]:
                            print('повторение!')
                            reverse = True
                            break
                    if reverse is False:
                        questions.append(final_data)
                        index += 1
            with open(f'{version}_new_data.html', 'w', encoding='utf-8') as file:
                title = """<!DOCTYPE html>
            <html lang="en">
            <head>
                <meta charset="UTF-8">
                <title>Title</title>
            </head>
            <body>
                """
                footer = """
                </body>
            </html>"""
                file.write(title)
                for line in questions:
                    file.write(line)
                file.write(footer)


class CycleClass:
    class Example:
        @staticmethod
        def example_cycle_for():
            for i in range(1, 10):
                print(i)

        @staticmethod
        def example_cycle_while():
            sec = 10
            while sec < 50:
                sec += 1
                print(sec)


class SendMail:
    class Example:
        @staticmethod
        def example_send_email(subject='subj', text='text'):
            host = 'smtp.yandex.ru'
            port = '465'
            login = 'eevee.cycle'
            password = '31284bogdan'
            writer = 'eevee.cycle@yandex.ru'
            recipient = 'eevee.cycle@yandex.ru'

            message = f"""From: {recipient}\nTo: {writer}\nSubject: {subject}\n\n{text}"""

            smtpobj = smtplib.SMTP_SSL(host=host, port=port)
            smtpobj.ehlo()
            smtpobj.login(user=login, password=password)
            smtpobj.sendmail(from_addr=writer, to_addrs=recipient, msg=message)
            smtpobj.quit()


class Pagination:
    class Example:
        @staticmethod
        def example():
            alpabet_list = list('abcdefghhklmnt')
            p = Pagination.Example(alpabet_list, 4)
            print(p.get_visible_items())

        def __init__(self, items=None, page_size=10):
            if items is None:
                items = []
            self.items = items
            self.page_size = page_size
            self.total_pages = 1 if not self.items else len(self.items) // self.page_size + 1
            self.current_page = 1

        def get_items(self):
            return self.items

        def get_page_size(self):
            return self.page_size

        def get_current_page(self):
            return self.current_page

        def prev_page(self):
            if self.current_page == 1:
                return self
            self.current_page -= 1
            return self

        def next_page(self):
            if self.current_page == self.total_pages:
                return self
            self.current_page += 1
            return self

        def first_page(self):
            self.current_page = 1
            return self

        def last_page(self):
            self.current_page = self.total_pages
            return self

        def go_to_page(self, page):
            if page < 1:
                page = 1
            elif page > self.total_pages:
                page = self.total_pages
            self.current_page = page
            return self

        def get_visible_items(self):
            start = (self.current_page - 1) * self.page_size
            return self.items[start:start + self.page_size]


########################################################################################################################


class VoiceClass:
    @staticmethod
    def example(self):
        # Not create class:
        self.speak()
        self.speak('Приветики!')
        # With create class:
        voice = self()
        voice.say()
        voice.say('Я уничтожу человечество!!!')

    def __init__(self, obj_id=0, volume=1.0, rate=200, voice=None, voices=None):
        self.engine = pyttsx3.init()
        self.volume = volume
        self.rate = rate
        self.voice = voice
        self.voices = voices
        self.obj_id = obj_id
        self.properties = [self.volume, self.rate, self.voice, self.voices]

    @staticmethod
    def speak(text: str = 'Inicialization successfull.'):
        pyttsx3.speak(text)

    @staticmethod
    async def async_speak(text: str = 'Inicialization successfull.'):
        await pyttsx3.speak(text)
        return text

    def say(self, text: str = 'Inicialization successfull.'):
        self.engine.say(text)
        self.engine.runAndWait()

    def get_property(self, name='volume'):
        return self.engine.getProperty(name)

    def get_properties(self):
        return [self.engine.getProperty(name) for name in self.properties]

    def set_property(self, name='volume', value=1.0):
        self.engine.setProperty(name, value)

    def set_properties(self, volume=1.0, rate=200):
        self.set_property(self.properties[0], volume)
        self.set_property(self.properties[1], rate)


class ObjectOrientedProgrammingClass:
    @staticmethod
    def example_worker():
        class Worker:
            """
            Класс, который содержит в себе работника, со значениями по строке
            """

            def __init__(self, A_1='', B_1='', C_1='', D_1='', E_1='', F_1='', G_1='', H_1='', M_1=''):
                # Подразделение
                self.A_1 = A_1
                # Цех или Служба
                self.B_1 = B_1
                # Отдел или участок
                self.C_1 = C_1
                # Фамилия
                self.D_1 = D_1
                # Имя
                self.E_1 = E_1
                # Отчество
                self.F_1 = F_1
                # Табельный №
                self.G_1 = G_1
                # Категория
                self.H_1 = H_1
                # Пол
                self.M_1 = M_1

            def print_worker(self):
                """

                """
                print(
                    f'{self.A_1}+{self.B_1}+{self.C_1}+{self.D_1}+{self.E_1}+{self.F_1}+{self.G_1}+{self.H_1}+{self.M_1}')

            def get_worker_value(self, index):
                """

                :param index:
                :return:
                """
                value_ = list(
                    (self.A_1, self.B_1, self.C_1, self.D_1, self.E_1, self.F_1, self.G_1, self.H_1, self.M_1))
                return value_[index]

            def get_worker_id(self):
                """

                :return:
                """
                return self.G_1

    @staticmethod
    def example_objects():
        class Actions:
            def __init__(self):
                pass

            @staticmethod
            def action(first_value, second_value):
                return first_value + second_value

        print(Actions.action(10, 30))

        class Circle:
            pi = 3.14

            def __init__(self, radius=1):
                self.radius = radius
                self.circle_reference = 2 * self.pi * self.radius

            def get_area(self):
                return self.pi * (self.radius ** 2)

            def get_circle_reference(self):
                return 2 * self.pi * self.radius

        circle_1 = Circle(4)
        print(circle_1.get_area())
        print(circle_1.circle_reference)
        print(circle_1.get_circle_reference())

        # Классы
        class Car:
            wheels_number = 4

            def __init__(self, name, color, year, is_crashed):
                self.name = name
                self.color = color
                self.year = year
                self.is_crashed = is_crashed

        mazda_car = Car(name="Mazda CX7", color="red", year=2017, is_crashed=True)
        print(mazda_car.name, mazda_car.is_crashed, mazda_car.wheels_number)
        bmw_car = Car(name="Mazda", color="black", year=2019, is_crashed=False)
        print(bmw_car.name, bmw_car.is_crashed, bmw_car.wheels_number)
        print(Car.wheels_number * 3)


########################################################################################################################
if __name__ == '__main__':
    # SyncAsyncThreadingPoolExecutorClass.Example.example_sync_compute()
    # SyncAsyncThreadingPoolExecutorClass.Example.example_async_compute()
    # SyncAsyncThreadingPoolExecutorClass.Example.example_threading_compute()
    # SyncAsyncThreadingPoolExecutorClass.Example.example_thread_pool_executor_compute()
    # SyncAsyncThreadingPoolExecutorClass.Example.example_process_pool_executor_compute()

    pass
