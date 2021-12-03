import os
import shutil
import sys
import time
from fnmatch import fnmatch
from multiprocessing import freeze_support
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor
from shutil import move, copy
import threading
import concurrent.futures
import asyncio
from time import sleep

import PySide6.QtWidgets as QtWidgets
import PySide6.QtGui as QtGui
import cv2
import openpyxl
import requests
from openpyxl.utils import get_column_letter
from skimage import io

from app_admin.utils.utils import ExcelClass, DirFolderPathClass


# Синхронная функция
def get_url(page_url, timeout=2):
    try:
        response = requests.get(url=page_url, timeout=timeout)
        page_status = "unknown"
        if response.status_code == 200:
            page_status = "exists"
        elif response.status_code == 404:
            page_status = "does not exist"

        return page_url + " - " + page_status
    except Exception as error:
        return page_url + " - " + f'error: {error}'


# UI
class AppContainerClass:
    def __init__(self):
        self.app = QtWidgets.QApplication([])
        self.widget = None

    def create_ui(self, title, width, height, icon):
        self.widget = MainWidgetClass(title, width, height, icon)
        return self.widget

    @staticmethod
    def create_qlable(text: str, _parent, background=False):
        _widget = QtWidgets.QLabel(text)
        if background:
            _widget.setAutoFillBackground(True)
            _widget.setStyleSheet("QLabel { background-color: rgba(0, 0, 0, 255); color : white; }")
        _parent.addWidget(_widget)
        return _widget

    @staticmethod
    def create_qpushbutton(_parent, _connect_func, _text='set'):
        _widget = QtWidgets.QPushButton(_text)
        _parent.addWidget(_widget)
        _widget.clicked.connect(_connect_func)
        return _widget

    @staticmethod
    def create_qcheckbox(_parent, _text='check?', default=False):
        _widget = QtWidgets.QCheckBox(_text)
        _widget.setChecked(default)
        _parent.addWidget(_widget)
        return _widget

    @staticmethod
    def create_qcombobox(_parent, _text: list, default=None):
        _widget = QtWidgets.QComboBox()
        _widget.addItems([x for x in _text])
        _widget.setCurrentText(default)
        _parent.addWidget(_widget)
        return _widget

    @staticmethod
    def create_qradiobutton(_parent, _text: str, default=False):
        _widget = QtWidgets.QRadioButton(_text)
        _widget.setChecked(default)
        _parent.addWidget(_widget)
        return _widget


class MainWidgetClass(QtWidgets.QWidget):
    def __init__(self, title="приложение", width=640, height=480, icon="icon.ico"):
        super().__init__()

        self.pause = False

        self.resize(width, height)
        self.setWindowTitle(title)
        self.setWindowIcon(QtGui.QIcon(icon))
        self.resolution_debug = []
        self.v_layout_m = QtWidgets.QVBoxLayout(self)

        # MANAGEMENT
        self.h_layout_g_management = QtWidgets.QHBoxLayout()
        self.v_layout_m.addLayout(self.h_layout_g_management)
        self.g_management_set = AppContainerClass.create_qlable(
            'УПРАВЛЕНИЕ',
            self.h_layout_g_management,
            background=True
        )
        self.h_layout_management_1 = QtWidgets.QHBoxLayout()
        self.v_layout_m.addLayout(self.h_layout_management_1)
        self.play_QPushButton = AppContainerClass.create_qpushbutton(
            self.h_layout_management_1,
            self.play_btn_func,
            'play'
        )
        self.stop_QPushButton = AppContainerClass.create_qpushbutton(
            self.h_layout_management_1,
            self.stop_btn_func,
            'stop'
        )
        self.quit_QPushButton = AppContainerClass.create_qpushbutton(
            self.h_layout_management_1,
            self.quit_btn_func,
            'quit'
        )

    def quit_btn_func(self):
        try:
            print('quit')
            self.pause = True
            sys.exit(app_container.app.exec())
        except Exception as error:
            print(error)

    def stop_btn_func(self):
        try:
            print('pause')
            self.pause = True
        except Exception as error:
            print(error)

    def play_btn_func(self):
        try:
            print('play')
            self.pause = False

            # Находит и обрезает лица, добавляет края + сжимает фото
            def find_face(max_workers=3, remove_old_file=False):
                start_time = time.time()
                print('start')
                print(f'max_workers={max_workers}')

                # Создание папок для исходящих и входящих изображений
                input_folder = DirFolderPathClass.create_folder_in_this_dir('input')
                output_folder = DirFolderPathClass.create_folder_in_this_dir('output\\completed')
                error_folder = DirFolderPathClass.create_folder_in_this_dir('output\\error')

                # Выбор формата изображений для обработки
                pattern = '*.jpg'

                # Функция для обрезки фото по именам
                def loop_crop_img(file_name):
                    try:
                        if self.pause is False:
                            if fnmatch(file_name, pattern):
                                # Чтение изображения в память
                                src_img = io.imread(input_folder + '\\' + file_name)

                                # Множитель
                                multiply = round(len(src_img) / 9248 * 0.5, 2)
                                print(multiply)

                                # Коррекция краёв изображения
                                correct_field = int(1000 * multiply)

                                # Размеры исходного изображения

                                # Загрузка алгоритма классификатора для поиска лиц
                                haar_face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_alt.xml')

                                # Значение, добавляемое по краям изображения
                                # width_addiction = int(image_width * 1.25 * multiply)
                                # height_addiction = int(image_height * 1.25 * multiply)

                                # Качество сжатия изображения
                                quality = 100

                                # Получение значения от исходного числа и % от него
                                def get_percent(value, side):
                                    return side * value // 100

                                # Установка обрезки краёв и обрезка
                                # crop_top = (image_height - (8 * multiply)) // 2
                                # crop_down = (image_height - (8 * multiply)) // 2
                                # crop_left = (image_width - (6 * multiply)) // 2
                                # crop_right = (image_width - (6 * multiply)) // 2
                                # src_img = src_img[
                                #           int(crop_top):int(image_height) - int(crop_down),
                                #           int(crop_left):int(image_width) - int(crop_right)
                                #           ]

                                # Предобработка изображения и поиск лиц алгоритмом
                                gray_img = cv2.cvtColor(src_img, cv2.COLOR_BGR2GRAY)
                                detect_faces = haar_face_cascade.detectMultiScale(gray_img)

                                # Очистка некорректно найдённых лиц на изображении
                                correct_faces = []
                                for (x, y, w, h) in detect_faces:
                                    if x > correct_field and y > correct_field and w > correct_field and \
                                            h > correct_field:
                                        correct_faces.append([x, y, w, h])

                                # Формирование финального результата
                                output = None
                                for (x, y, w, h) in correct_faces:
                                    print(f'y={y} | h={h} | x={x} | w={w}')
                                    height_1 = int(y - (h * 0.5))
                                    if height_1 < 0:
                                        height_1 = 0
                                    height_2 = int(y + h + (h * 0.5))
                                    if height_2 < 0:
                                        height_2 = 0
                                    width_1 = int(x - (w * 0.5 * 0.4))
                                    if width_1 < 0:
                                        width_1 = 0
                                    width_2 = int(x + w + (w * 0.5 * 0.4))
                                    if width_2 < 0:
                                        width_2 = 0
                                    print(f'height_1={height_1} | height_2={height_2} | '
                                          f'width_1={width_1} |  width_2={width_2}')
                                    if w > correct_field or h > correct_field:
                                        if abs(height_2 - height_1) > correct_field and \
                                                abs(width_2 - width_1) > correct_field:
                                            output = src_img[int(height_1):int(height_2), int(width_1):int(width_2)]

                                # Удаление исходного фото
                                try:
                                    if remove_old_file:
                                        os.remove(input_folder + '\\' + file_name)
                                except Exception as error:
                                    pass

                                # Сохранение результата
                                io.imsave(output_folder + '\\' + file_name, output, quality=quality)

                                # Вывод имени обработанного фото
                                return f'{file_name}\n'
                            # Если изображение не того формата, копирование его в выбранную папку
                            else:
                                try:
                                    copy(f'{input_folder}\\{file_name}', f'{error_folder}\\{file_name}')
                                    # Вывод на экран сообщения и имени файла
                                    return f"ошибка формата изображения: {file_name}\n"
                                except Exception as error:
                                    copy(f'{input_folder}\\{file_name}', f'{error_folder}\\{file_name}')
                                    # Вывод на экран сообщения и имени файла
                                    return f"ошибка: {file_name}\n"
                    except Exception as error:
                        copy(f'{input_folder}\\{file_name}', f'{error_folder}\\{file_name}')
                        # Удаление исходного фото
                        try:
                            if remove_old_file:
                                os.remove(input_folder + '\\' + file_name)
                        except Exception as _:
                            pass
                        return f'{file_name}: {error}\n'

                # Запуск многопоточности
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    for path, subdirs, files in os.walk(input_folder):
                        for file in files:
                            if self.pause is False:
                                futures = executor.submit(loop_crop_img, file_name=file)
                                print(futures.result())

                # Финальное время
                print(f"Final time: {round(time.time() - start_time, 1)}")
                print('end')

            # Сравнивает массив фото с excel-файлом выгруженным из системы 1С
            def equal_foto(max_workers=6, remove_old_file=False):
                start_time = time.time()
                print('start')
                print(f'max_workers={max_workers}')

                # Выбор excel-файла для загрузки данных из 1с
                export_file_from_1c = 'export.xlsx'
                # Выбор столбцов с данными в excel-файле
                export_cols_from_1c = '549'

                # Создание папок для исходящих и входящих изображений
                input_folder = DirFolderPathClass.create_folder_in_this_dir('input')
                equal_folder = DirFolderPathClass.create_folder_in_this_dir('output\\Есть в базе')
                not_equal_folder = DirFolderPathClass.create_folder_in_this_dir('output\\Нет в базе')
                error_format_folder = DirFolderPathClass.create_folder_in_this_dir('output\\Неверный формат фото')
                error_id_folder = DirFolderPathClass.create_folder_in_this_dir('output\\Нет идентификатора')
                error_folder = DirFolderPathClass.create_folder_in_this_dir('output\\Ошибка')

                # Выбор формата изображений для обработки
                pattern = '*.jpg'

                # Загрузка в память excel-файла
                workbook = ExcelClass.workbook_load(file=export_file_from_1c)
                sheet = ExcelClass.workbook_activate(workbook=workbook)
                max_num_rows = ExcelClass.get_max_num_rows(sheet=sheet)

                # Создание и наполнение матрицы с данными по людям из 1с
                global_workers_list = []
                # Создание и наполнение массива с идентификаторами с КМ
                global_workers_id_list = []
                for num in range(1, max_num_rows + 1):
                    if self.pause is False:
                        # Создание и наполнение массива с данными по одного человеку
                        local_workers_list = []
                        for x in export_cols_from_1c:
                            local_workers_list.append(
                                str(ExcelClass.get_sheet_value(col=int(x), row=num, sheet=sheet)).strip()
                            )
                        if local_workers_list[2]:
                            global_workers_list.append(local_workers_list)
                            global_workers_id_list.append(local_workers_list[2])
                ExcelClass.workbook_close(workbook)
                # Вывод на экран матрицы с данными и её длины
                print(global_workers_list)
                print(len(global_workers_list))

                # Функция для сравнения фото по именам и данным из excel-файла
                def loop_equal_foto(file_name: str):
                    message = ''
                    if self.pause is False:
                        # Проверка на формат изображения
                        if fnmatch(file_name, pattern):
                            try:
                                # Получение полного имени файла
                                full_foto_name = file_name.split('.')[0].strip()
                                name_foto = file_name.split('+')[1].strip()
                                try:
                                    # Получение идентификатора из имени файла
                                    id_foto = full_foto_name.split('_')[1].strip()
                                    # Проверка, есть ли идентификатор в массиве с идентификаторами с 1с
                                    try:
                                        index = global_workers_id_list.index(id_foto)
                                        worker = global_workers_list[index]
                                        # Копирование файла, если его имя полностью совпадает с матрицей
                                        if full_foto_name == f'{worker[0]}+{worker[1]}_{worker[2]}':
                                            copy(f'{input_folder}\\{file_name}', f'{equal_folder}\\{file_name}')
                                            # Вывод на экран сообщения и имени файла
                                            message = f"соответветствует: {file_name}\n"
                                        # Копирование файла и замена имени, если его идентификатор совпадает
                                        else:
                                            copy(f'{input_folder}\\{file_name}',
                                                 f'{equal_folder}\\{worker[0]}+{worker[1]}_{worker[2]}.jpg')
                                            # Вывод на экран сообщения и имени файла
                                            message = f"соответветствует с корректировкой: {file_name}\n"
                                    # Копирование файла, если его нет в массиве с идентификаторами с 1с
                                    except Exception as error:
                                        copy(f'{input_folder}\\{file_name}', f'{not_equal_folder}\\{file_name}')
                                        # Вывод на экран сообщения и имени файла
                                        message = f"не соответветствует: {file_name}\n"
                                # Копирование файла, если у него нет идентификатора
                                except Exception as error:
                                    copy(f'{input_folder}\\{file_name}', f'{error_id_folder}\\{file_name}')
                                    # Вывод на экран сообщения и имени файла
                                    message = f"нет идентификатора: {file_name}\n"
                            except Exception as error:
                                copy(f'{input_folder}\\{file_name}', f'{error_folder}\\{file_name}')
                                # Вывод на экран сообщения и имени файла
                                message = f"ошибка: {file_name}\n"
                        # Если изображение не того формата, копирование его в выбранную папку
                        else:
                            try:
                                copy(f'{input_folder}\\{file_name}', f'{error_format_folder}\\{file_name}')
                                # Вывод на экран сообщения и имени файла
                                message = f"ошибка формата изображения: {file_name}\n"
                            except Exception as error:
                                copy(f'{input_folder}\\{file_name}', f'{error_folder}\\{file_name}')
                                # Вывод на экран сообщения и имени файла
                                message = f"ошибка: {file_name}\n"
                        # Удаление исходного файла
                        try:
                            if remove_old_file:
                                os.remove(f'{input_folder}\\{file_name}')
                        except Exception as error:
                            pass

                        return message

                # Запуск многопоточности
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    for path, subdirs, files in os.walk(input_folder):
                        for file in files:
                            if self.pause is False:
                                futures = executor.submit(loop_equal_foto, file_name=file)
                                print(futures.result())

                # Финальное время
                print(f"Final time: {round(time.time() - start_time, 1)}")
                print('end')

            # Удаление идентификатора с названия файла
            def remove_id_from_jpg(max_workers=1, remove_old_file=False):
                start_time = time.time()
                print('start')
                print(f'max_workers={max_workers}')

                # Создание папок для исходящих и входящих изображений
                input_folder = DirFolderPathClass.create_folder_in_this_dir('input')
                equal_folder = DirFolderPathClass.create_folder_in_this_dir('output\\completed')
                error_format_folder = DirFolderPathClass.create_folder_in_this_dir('output\\format_error')
                error_folder = DirFolderPathClass.create_folder_in_this_dir('output\\error')

                # Выбор формата изображений для обработки
                pattern = '*.jpg'

                # Функция для удаления идентификатора с фото
                def loop_remove_id(file_name: str):
                    message = ''
                    try:
                        if self.pause is False:
                            if fnmatch(file_name, pattern):
                                name = file_name.split('.')[0].strip().split('_')[0].strip() + pattern[1:]
                                copy(f'{input_folder}\\{file_name}', f'{equal_folder}\\{name}')
                                message = f"{name}\n"
                            # Если изображение не того формата, копирование его в выбранную папку
                            else:
                                copy(f'{input_folder}\\{file_name}', f'{error_format_folder}\\{file_name}')
                                # Вывод на экран сообщения и имени файла
                                message = f"ошибка формата изображения: {file_name}\n"
                    except Exception as error:
                        copy(f'{input_folder}\\{file_name}', f'{error_folder}\\{file_name}')
                        message = f'{file_name}: {error}\n'

                    # Удаление исходного файла
                    try:
                        if remove_old_file:
                            os.remove(f'{input_folder}\\{file_name}')
                    except Exception as error:
                        pass

                    return message

                # Запуск многопоточности
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    for path, subdirs, files in os.walk(input_folder):
                        for file in files:
                            if self.pause is False:
                                futures = executor.submit(loop_remove_id, file_name=file)
                                print(futures.result())

                # Финальное время
                print(f"Final time: {round(time.time() - start_time, 1)}")
                print('end')

            # Удаление идентификатора с названия файла
            def add_id_to_file(max_workers=1, remove_old_file=False):
                start_time = time.time()
                print('start')
                print(f'max_workers={max_workers}')

                # Создание папок для исходящих и входящих изображений
                input_folder = DirFolderPathClass.create_folder_in_this_dir('input')
                equal_folder = DirFolderPathClass.create_folder_in_this_dir('output\\completed')
                error_format_folder = DirFolderPathClass.create_folder_in_this_dir('output\\format_error')
                error_folder = DirFolderPathClass.create_folder_in_this_dir('output\\error')

                # Выбор формата изображений для обработки
                pattern = '*.jpg'

                # Начальный идентификатор, с которого нужно начинать нумерацию
                start_index = 1

                # Функция для удаления идентификатора с фото
                def loop_remove_id(file_name: str, start_index: int):
                    message = ''
                    try:
                        if self.pause is False:
                            if fnmatch(file_name, pattern):
                                try:
                                    name = file_name.split('.')[0].strip().split('_')[0].strip() + pattern[1:]
                                    copy(f'{input_folder}\\{file_name}', f'{error_folder}\\{name}')
                                    message = f"{name}\n"
                                except Exception as error:
                                    name = file_name.split('.')[0].strip() + f'_{start_index}{pattern[1:]}'
                                    copy(f'{input_folder}\\{file_name}', f'{error_folder}\\{name}')
                                    message = f"{name}\n"

                            # Если изображение не того формата, копирование его в выбранную папку
                            else:
                                copy(f'{input_folder}\\{file_name}', f'{error_format_folder}\\{file_name}')
                                # Вывод на экран сообщения и имени файла
                                message = f"ошибка формата изображения: {file_name}\n"
                    except Exception as error:
                        copy(f'{input_folder}\\{file_name}', f'{error_folder}\\{file_name}')
                        message = f'{file_name}: {error}\n'

                    # Удаление исходного файла
                    try:
                        if remove_old_file:
                            os.remove(f'{input_folder}\\{file_name}')
                    except Exception as error:
                        pass

                    return message

                # Запуск многопоточности
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    for path, subdirs, files in os.walk(input_folder):
                        for file in files:
                            if self.pause is False:
                                start_index += 1
                                futures = executor.submit(loop_remove_id, file_name=file, start_index=start_index)
                                print(futures.result())

                # Финальное время
                print(f"Final time: {round(time.time() - start_time, 1)}")
                print('end')

            # Меняет русскую букву "Р" на английскую в имени файла или наоборот
            def change_p_to_eng_to_rus(max_workers=12, to_eng=True, remove_old_file=False):
                start_time = time.time()
                print('start')
                print(f'max_workers={max_workers}')

                # Создание папок для исходящих и входящих изображений
                input_folder = DirFolderPathClass.create_folder_in_this_dir('input')
                output_folder = DirFolderPathClass.create_folder_in_this_dir('output\\completed')
                error_format_folder = DirFolderPathClass.create_folder_in_this_dir('output\\Неверный формат фото')
                error_folder = DirFolderPathClass.create_folder_in_this_dir('output\\Ошибка')
                pattern = '*.jpg'

                # Функция для замены русской Р на английскую
                def loop_rename_file(file_name):
                    message = ''
                    try:
                        # Проверка на формат изображения
                        if fnmatch(file_name, pattern):
                            full_name = file_name.strip().split('.')[0].strip()
                            first_name = full_name.split('+')[0].strip()
                            second_name = full_name.split('+')[1].strip()
                            new_name = f'{first_name}+{second_name}.jpg'
                            if to_eng:
                                eng = 1056
                                rus = 80
                            else:
                                eng = 80
                                rus = 1056
                            if ord(first_name[0:1:]) == eng or ord(second_name[0:1:]) == eng:
                                if ord(first_name[0:1:]) == eng:
                                    first_name = chr(rus) + first_name[1::]
                                if ord(second_name[0:1:]) == eng:
                                    second_name = chr(rus) + second_name[1::]
                                copy(f'{input_folder}\\{file_name}', f'{output_folder}\\{new_name}')
                                message = f'{new_name}\n'
                            else:
                                copy(f'{input_folder}\\{file_name}', f'{output_folder}\\{new_name}')
                                message = 'pass\n'
                        else:
                            copy(f'{input_folder}\\{file_name}', f'{error_format_folder}\\{file_name}')
                            message = 'error: неверный формат файла\n'
                    except Exception as error:
                        copy(f'{input_folder}\\{file_name}', f'{error_folder}\\{file_name}')
                        message = f"{file_name} : error : {error}\n"

                    # Удаление исходного файла
                    try:
                        if remove_old_file:
                            os.remove(f'{input_folder}\\{file_name}')
                    except Exception as error:
                        pass

                    return message

                # Запуск многопоточности
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    for path, subdirs, files in os.walk(input_folder):
                        for file in files:
                            if self.pause is False:
                                futures = executor.submit(loop_rename_file, file_name=file)
                                print(futures.result())

                # Финальное время
                print(f"Final time: {round(time.time() - start_time, 1)}")
                print('end')

            # Заполнение данных экспортированных из системы, данными из 1с
            def filling_extra_data_from_1c_to_import_hikvision(max_workers=3):
                start_time = time.time()
                print('start')

                input_1 = '1c.xlsx'
                input_2 = 'export.xlsx'
                output_1 = 'import.xlsx'

                workbook = ExcelClass.workbook_load(file=input_1)
                sheet = ExcelClass.workbook_activate(workbook=workbook)
                max_num_rows = ExcelClass.get_max_num_rows(sheet=sheet)

                global_workers_list = []
                global_workers_id_list = []
                for row_export in range(1, max_num_rows + 1):
                    local_workers_list = []
                    for col in range(1, 10 + 1):
                        local_workers_list.append(
                            ExcelClass.get_sheet_value(
                                col=ExcelClass.get_column_letter(col),
                                row=row_export,
                                sheet=sheet
                            )
                        )
                    if local_workers_list[8]:
                        global_workers_list.append(local_workers_list)
                        global_workers_id_list.append(local_workers_list[8])
                print(global_workers_list)
                print(global_workers_id_list)
                ExcelClass.workbook_close(workbook=workbook)

                workbook = ExcelClass.workbook_load(file=input_2)
                sheet = ExcelClass.workbook_activate(workbook=workbook)
                max_num_rows = ExcelClass.get_max_num_rows(sheet=sheet)
                global_workers_export_list = []
                for row_export in range(1, max_num_rows + 1):
                    local_workers_export_list = []
                    for col in range(1, 27 + 1):
                        local_workers_export_list.append(
                            ExcelClass.get_sheet_value(
                                col=ExcelClass.get_column_letter(col),
                                row=row_export,
                                sheet=sheet
                            )
                        )
                    global_workers_export_list.append(local_workers_export_list)
                print(global_workers_export_list)
                ExcelClass.workbook_close(workbook=workbook)

                workbook = ExcelClass.workbook_load(file=output_1)
                sheet = ExcelClass.workbook_activate(workbook=workbook)
                for row_export in global_workers_export_list:
                    print(f'{row_export[0]} {row_export[1]} {row_export[2]}')
                    for value in row_export:
                        try:
                            index = global_workers_id_list.index(row_export[2])
                            worker = global_workers_list[index]
                            col_index = row_export.index(value) + 1
                            if col_index == 1:
                                value = worker[4]
                            elif col_index == 2:
                                value = worker[3]
                            elif col_index == 18:
                                value = worker[0]
                            elif col_index == 19:
                                value = worker[1]
                            elif col_index == 20:
                                value = worker[2]
                            elif col_index == 21:
                                value = worker[5]
                            elif col_index == 22:
                                value = worker[9]
                            ExcelClass.set_sheet_value(
                                col=ExcelClass.get_column_letter(col_index),
                                row=global_workers_export_list.index(row_export) + 1,
                                value=value,
                                sheet=sheet
                            )
                        except Exception as error:
                            ExcelClass.set_sheet_value(
                                col=ExcelClass.get_column_letter(row_export.index(value) + 1),
                                row=global_workers_export_list.index(row_export) + 1,
                                value=value,
                                sheet=sheet
                            )
                ExcelClass.workbook_save(workbook=workbook, filename=output_1)

                # Финальное время
                print(f"Final time: {round(time.time() - start_time, 1)}")
                print('end')

            # Сравнение данных экспортированных из системы и данных из 1с, на выходе файл с фильтром
            # по работникам не в базе
            def equal_system(max_workers=1):
                start_time = time.time()
                print('start')
                print(f'max_workers={max_workers}')

                input_1 = 'export.xlsx'
                input_2 = '1c.xlsx'
                output_1 = 'import.xlsx'

                workbook = ExcelClass.workbook_load(file=input_1)
                sheet = ExcelClass.workbook_activate(workbook=workbook)
                max_num_rows = ExcelClass.get_max_num_rows(sheet=sheet)

                global_workers_id_list = []
                for row_export in range(1, max_num_rows + 1):
                    value = ExcelClass.get_sheet_value(col=ExcelClass.get_column_letter(3), row=row_export, sheet=sheet)
                    if value:
                        global_workers_id_list.append(value)
                print(global_workers_id_list)
                ExcelClass.workbook_close(workbook=workbook)

                workbook = ExcelClass.workbook_load(file=input_2)
                sheet = ExcelClass.workbook_activate(workbook=workbook)
                max_num_rows = ExcelClass.get_max_num_rows(sheet=sheet)

                global_workers_list = []
                for row_export in range(1, max_num_rows + 1):
                    local_workers_list = []
                    for col in range(1, 10 + 1):
                        local_workers_list.append(
                            ExcelClass.get_sheet_value(
                                col=ExcelClass.get_column_letter(col),
                                row=row_export,
                                sheet=sheet
                            )
                        )
                    if local_workers_list[8]:
                        if local_workers_list[8] in global_workers_id_list:
                            local_workers_list.append('+')
                        else:
                            local_workers_list.append('-')
                        global_workers_list.append(local_workers_list)
                print(global_workers_list)
                print(global_workers_id_list)
                ExcelClass.workbook_close(workbook=workbook)

                workbook = ExcelClass.workbook_load(file=output_1)
                sheet = ExcelClass.workbook_activate(workbook=workbook)
                for row_export in global_workers_list:
                    print(f'{row_export[3]} {row_export[4]} {row_export[8]}')
                    for value in row_export:
                        ExcelClass.set_sheet_value(
                            col=ExcelClass.get_column_letter(row_export.index(value) + 1),
                            row=global_workers_list.index(row_export) + 1,
                            value=value,
                            sheet=sheet
                        )
                ExcelClass.workbook_save(workbook=workbook, filename=output_1)

                # Финальное время
                print(f"Final time: {round(time.time() - start_time, 1)}")
                print('end')

                # def get_sheet_value(column, row, _sheet):
                #     return str(_sheet[str(column) + str(row)].value)
                #
                # def set_sheet_value(column, row, value):
                #     sheet[f'{column}{row}'] = str(value)
                #
                # file_1c = 'export.xlsx'
                # workers_id_1c = []
                # id_1c = 9
                #
                # workbook = openpyxl.load_workbook(file_1c)
                # sheet = workbook.active
                # for num in range(1, 3000):
                #     workers_id_1c.append(get_sheet_value(get_column_letter(int(id_1c)), num, _sheet=sheet))
                # workbook.close()
                #
                # file_skud = 'import.xlsx'
                # workers_id_skud = []
                # id_skud = 3
                #
                # workbook = openpyxl.load_workbook(file_skud)
                # sheet = workbook.active
                # for num in range(1, 3000):
                #     workers_id_skud.append(get_sheet_value(get_column_letter(int(id_skud)), num, _sheet=sheet))
                # workbook.close()
                #
                # workbook = openpyxl.load_workbook(file_1c)
                # sheet = workbook.active
                #
                # for x in workers_id_1c:
                #     for y in workers_id_skud:
                #         if x == y and x is not None:
                #             set_sheet_value(column=get_column_letter(int(8)), row=workers_id_1c.index(x) + 1, value='+')
                # workbook.save('export.xlsx')

            # Переименование файлов в папках, включая все подпапки
            def rename_file_in_folders(max_workers=1):
                start_time = time.time()
                print('start')
                print(f'max_workers={max_workers}')

                # Создание папок для исходящих и входящих изображений
                input_folder = DirFolderPathClass.create_folder_in_this_dir('input')

                directories_ = []
                for root, dirs, files in os.walk(input_folder, topdown=True):
                    for name in dirs:
                        directories_.append(f"{os.path.join(root, name)}")

                # Финальное время
                print(f"Final time: {round(time.time() - start_time, 1)}")
                print('end')

            # Сравнение данных из 1с и данных в базе данных енбека
            def extra_enbek():
                start_time = time.time()
                print('start')

                input_1 = '1c.xlsx'
                input_2 = 'enbek.xlsx'
                output_1 = 'final.xlsx'

                workbook = ExcelClass.workbook_load(file=input_1)
                sheet = ExcelClass.workbook_activate(workbook=workbook)
                max_num_rows = ExcelClass.get_max_num_rows(sheet=sheet)

                global_workers_id_list = []
                for row in range(1, max_num_rows + 1):
                    global_workers_id_list.append(
                        ExcelClass.get_sheet_value(col=ExcelClass.get_column_letter(11), row=row, sheet=sheet)
                    )
                print(global_workers_id_list)
                ExcelClass.workbook_close(workbook=workbook)

                workbook = ExcelClass.workbook_load(file=input_2)
                sheet = ExcelClass.workbook_activate(workbook=workbook)
                max_num_rows = ExcelClass.get_max_num_rows(sheet=sheet)
                global_workers_enbek_list = []
                for row in range(1, max_num_rows + 1):
                    local_workers_enbek_list = []
                    for col in range(1, 21 + 1):
                        local_workers_enbek_list.append(
                            ExcelClass.get_sheet_value(col=ExcelClass.get_column_letter(col), row=row, sheet=sheet)
                        )
                    if local_workers_enbek_list[1] in global_workers_id_list:
                        local_workers_enbek_list.append('+')
                    else:
                        local_workers_enbek_list.append('-')
                    global_workers_enbek_list.append(local_workers_enbek_list)
                print(global_workers_enbek_list)
                ExcelClass.workbook_close(workbook=workbook)

                workbook = ExcelClass.workbook_load(file=output_1)
                sheet = ExcelClass.workbook_activate(workbook=workbook)
                for row in global_workers_enbek_list:
                    print(row)
                    for value in row:
                        ExcelClass.set_sheet_value(
                            col=ExcelClass.get_column_letter(row.index(value) + 1),
                            row=global_workers_enbek_list.index(row) + 1,
                            value=value,
                            sheet=sheet
                        )
                ExcelClass.workbook_save(workbook=workbook, filename=output_1)

                # Финальное время
                print(f"Final time: {round(time.time() - start_time, 1)}")
                print('end')

            # threading.Thread(target=extra_enbek, args=()).start()

            # threading.Thread(target=find_face, args=([3])).start()
            # threading.Thread(target=equal_foto, args=([6])).start()
            # threading.Thread(target=remove_id_from_jpg, args=([6])).start()
            # threading.Thread(target=change_p_to_eng_to_rus, args=([12, True])).start()
            # threading.Thread(target=filling_extra_data_from_1c_to_import_hikvision, args=([1])).start()
            # threading.Thread(target=equal_system, args=([1])).start()

        except Exception as error:
            print(error)


# 1 Перевести в единый формат .jpg фото из папок: готовые в импорт и уже добавлены. Сменить латинскую P на кириллицу.
# 2 Объединить все фото в одной папке, заменив перед этим старые свежими.
# 3 Обрезать лица алгоритмом: поиграться с отношением, разрешением и сжатием.
# 4 Найти в базе фото совпадения с КМ, скорректировать ошибки.
# 5 Найти в базе фото совпадения с ТОО, скорректировать ошибки.
# 6 Сменить кириллическую P на латинскую.
# 7 Импортировать в базу данных hikvision.
# 8 Переделать некорректные названия фото и снова импортировать.
# 9 Выгрузить из системы список работников, добавить поля в базу hikvision и загрузить дополненный список из 1с.

# MAIN
if __name__ == "__main__":
    freeze_support()
    app_container = AppContainerClass()
    widget = app_container.create_ui(
        title="приложение",
        width=640,
        height=480,
        icon="icon.ico"
    )
    ui_thread = threading.Thread(target=widget.show())
    sys.exit(app_container.app.exec())
