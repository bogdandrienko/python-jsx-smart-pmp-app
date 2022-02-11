import base64
import datetime
import hashlib
import json
import os
import random
import bs4
import httplib2
import requests

from django.contrib import admin
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User, Group, update_last_login
from django.core.mail import send_mail
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.template.loader import get_template
from django.urls import reverse
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill
from rest_framework import status
from rest_framework import viewsets, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from xhtml2pdf import pisa

from backend import models as backend_models, settings as backend_settings, serializers as backend_serializers, \
    forms as backend_forms, service as backend_service, utils as backend_utils


def index(request):
    """
    React app index page
    """

    # logging
    backend_service.DjangoClass.LoggingClass.logging_actions(request=request)

    # if str(request.META.get("REMOTE_ADDR")) == '192.168.1.202':
    #     redirect('http://192.168.1.68:8000/home/')

    context = {}
    return render(request, 'index.html', context)


def admin_(request):
    """
    Admin page
    """

    # logging
    backend_service.DjangoClass.LoggingClass.logging_actions(request=request)

    # if str(request.META.get("REMOTE_ADDR")) == '192.168.1.202':
    #     redirect('http://192.168.1.68:8000/admin/')

    return render(request, admin.site.urls)


@api_view(http_method_names=['GET'])
def routes(request):
    """
    All django-rest-framework routes
    """

    # logging
    backend_service.DjangoClass.LoggingClass.logging_actions(request=request)

    print(f'\n\nroutes:\n\n')
    _routes = [
        {
            'Endpoints': '$BASEPATH$/router/',
            'methods': 'GET, HEAD, OPTIONS',
            'descriptions': 'default router for rest_framework ViewSet',
            'helps': '''$BASEPATH$/router/''',
            'code': '''from rest_framework import routers\n
            router = routers.DefaultRouter()\n
            router.register(prefix=r'users', viewset=views.UserViewSet, basename='users')'''
        },
        {
            'name': '''JWT token''',
            'endpoint': '''$BASEPATH$/tokens/''',
            'method': '''...''',
            'body': '''...''',
            'descriptions': '''All endpoints for "JWT token" with django-rest_framework api''',
            'helps': '''$BASEPATH$/tokens/''',
            'code': '''...'''
        },
        {
            'name': '''get token''',
            'endpoint': '''$BASEPATH$/token/''',
            'method': '''...''',
            'body': '''...''',
            'descriptions': '''All actions for "note" with django-rest_framework api''',
            'helps': '''...''',
            'code': '''...'''
        },
        {
            'name': '''refresh token''',
            'endpoint': '''$BASEPATH$/token/refresh/''',
            'method': '''...''',
            'body': '''...''',
            'descriptions': '''refresh token with django-rest_framework api''',
            'helps': '''-''',
            'code': '''...'''
        },
        {
            'name': '''verify token''',
            'endpoint': '''$BASEPATH$/token/verify/''',
            'method': '''...''',
            'body': '''...''',
            'descriptions': '''verify token with django-rest_framework api''',
            'helps': '''-''',
            'code': '''...'''
        },
        {
            'name': '''note_api''',
            'endpoint': '''$BASEPATH$/note_api/ && $BASEPATH$/note_api/<id>/''',
            'method': '''...''',
            'body': '''...''',
            'descriptions': '''All actions for "note" with django-rest_framework api''',
            'helps': '''$BASEPATH$/note_api/-1/''',
            'code': '''...'''
        },
    ]
    return Response(_routes)


@api_view(http_method_names=['GET', 'POST', 'PUT', 'DELETE'])
def api_login_user(request):
    """
    Api django-rest-framework login
    """

    # logging
    backend_service.DjangoClass.LoggingClass.logging_actions(request=request)

    print('\n\n\n')
    print('get_tokens_for_user:')
    print('\n')
    request_method, request_action_type, request_user, request_body = \
        backend_service.DjangoClass.DRFClass.request_utils(request=request)
    print(f"datetime: {backend_utils.DateTimeUtils.get_current_datetime()}", f"method: {request_method}",
          f"action_type: {request_action_type}", f"user: {request_user}", f"body: {request_body}")

    try:
        if request_method == 'GET':
            return Response({"error": "This method not allowed for endpoint."})
        elif request_method == 'POST':
            if request_action_type == "LOGIN":
                #  { "body": {"Action-type": "USER_LOGIN", "username": "Bogdan", "password": "31284bogdan"} }
                try:
                    username = request.data["body"]["username"]
                    password = request.data["body"]["password"]

                    ip = request.META.get("REMOTE_ADDR")
                    request_path = request.path
                    request_method = request.method
                    now = (datetime.datetime.now()).strftime('%Y-%m-%d %H:%M')
                    access_count = 0
                    for dat in backend_models.LoggingModel.objects.filter(
                            username_slug_field=username,
                            ip_genericipaddress_field=ip,
                            request_path_slug_field=request_path,
                            request_method_slug_field=request_method,
                            error_text_field=f'action: LOGIN'
                    ):
                        if (dat.datetime_field + datetime.timedelta(hours=6, minutes=59)).strftime('%Y-%m-%d %H:%M') \
                                >= now:
                            access_count += 1

                    print(f"access_count: {access_count}")

                    if access_count < 10:
                        backend_models.LoggingModel.objects.create(
                            username_slug_field=username,
                            ip_genericipaddress_field=ip,
                            request_path_slug_field=request_path,
                            request_method_slug_field=request_method,
                            error_text_field=f'action: LOGIN'
                        )

                        is_authenticated = authenticate(username=username, password=password)
                        if is_authenticated is not None:
                            user = User.objects.get(username=username)
                            user_model = backend_models.UserModel.objects.get(user_foreign_key_field=user)
                            update_last_login(sender=None, user=user)
                            refresh = RefreshToken.for_user(user=user)
                            response = {"response": {
                                "refresh": str(refresh),
                                "access": str(refresh.access_token),
                                "token": str(refresh.access_token),
                                "username": str(user.username),
                                "name": str(f'{user_model.last_name_char_field} {user_model.first_name_char_field}'),
                            }}
                        else:
                            response = {'error': 'Внимание, данные не совпадают!'}
                    else:
                        response = {'error': 'Внимание, попыток входа можно совершать не более 10 в час!'}
                    print(f"response: {response}")
                    return Response(response)
                except Exception as error:
                    print(error)
                    backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                    return Response({"error": "This action has error."})
            else:
                return Response({"error": "This action not allowed for this method."})
        elif request_method == 'PUT':
            return Response({"error": "This method not allowed for endpoint."})
        elif request_method == 'DELETE':
            return Response({"error": "This method not allowed for endpoint."})
        else:
            return Response({"error": "This method not allowed for endpoint."})
    except Exception as error:
        print(error)
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        return Response({"error": "This endpoint has error."})


@api_view(http_method_names=['GET', 'POST', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def api_user_profile(request):
    """
    Api django-rest-framework profile
    """

    # logging
    backend_service.DjangoClass.LoggingClass.logging_actions(request=request)

    print('\n\n\n')
    print('api_user_profile:')
    print('\n')
    request_method, request_action_type, request_user, request_body = \
        backend_service.DjangoClass.DRFClass.request_utils(request=request)
    print(f"datetime: {backend_utils.DateTimeUtils.get_current_datetime()}", f"method: {request_method}",
          f"action_type: {request_action_type}", f"user: {request_user}", f"body: {request_body}")

    try:
        if request_method == 'GET':
            return Response({"error": "This method not allowed for endpoint."})
        elif request_method == 'POST':
            if request_action_type == "PROFILE":
                try:
                    serializer = backend_serializers.UserSerializer(request_user, many=False)
                    response = {"response": serializer.data}
                    print(f"response: {response}")
                    return Response(response)
                except Exception as error:
                    print(error)
                    backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                    return Response({"error": "This action has error."})
            else:
                return Response({"error": "This action not allowed for this method."})
        elif request_method == 'PUT':
            return Response({"error": "This method not allowed for endpoint."})
        elif request_method == 'DELETE':
            return Response({"error": "This method not allowed for endpoint."})
        else:
            return Response({"error": "This method not allowed for endpoint."})
    except Exception as error:
        print(error)
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        return Response({"error": "This endpoint has error."})


@api_view(http_method_names=['GET', 'POST', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def api_user_change_profile(request):
    """
    Api django-rest-framework change profile
    """

    # logging
    backend_service.DjangoClass.LoggingClass.logging_actions(request=request)

    print('\n\n\n')
    print('api_user_change_profile:')
    print('\n')
    request_method, request_action_type, request_user, request_body = \
        backend_service.DjangoClass.DRFClass.request_utils(request=request)
    print(f"datetime: {backend_utils.DateTimeUtils.get_current_datetime()}", f"method: {request_method}",
          f"action_type: {request_action_type}", f"user: {request_user}", f"body: {request_body}")

    try:
        if request_method == 'GET':
            return Response({"error": "This method not allowed for endpoint."})
        elif request_method == 'POST':
            if request_action_type == "CHANGE":
                try:
                    user_model = backend_models.UserModel.objects.get(user_foreign_key_field=request_user)
                    password = str(request_body["password"]).strip()
                    password2 = str(request_body["password2"]).strip()
                    if len(password) > 1 and password == password2 and password != user_model.password_slug_field:
                        request_user.set_password(password)
                        request_user.save()
                        user_model.password_slug_field = password
                        user_model.temp_password_boolean_field = False
                        try:
                            if request_body["email"] and request_body["email"] != user_model.email_field:
                                user_model.email_field = request_body["email"]
                        except Exception as error:
                            pass
                        try:
                            if request_body["secretQuestion"] and request_body["secretQuestion"] != \
                                    user_model.secret_question_char_field:
                                user_model.secret_question_char_field = request_body["secretQuestion"]
                        except Exception as error:
                            pass
                        try:
                            if request_body["secretAnswer"] and request_body["secretAnswer"] != \
                                    user_model.secret_answer_char_field:
                                user_model.secret_answer_char_field = request_body["secretAnswer"]
                        except Exception as error:
                            pass
                        user_model.save()
                        response = {"response": "Изменение пароля успешно проведено!"}
                    else:
                        response = {"error": "Пароли пустые, не совпадают или идентичны предыдущему"}
                    print(f"response: {response}")
                    return Response(response)
                except Exception as error:
                    print(error)
                    backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                    return Response({"error": "This action has error."})
            else:
                return Response({"error": "This action not allowed for this method."})
        elif request_method == 'PUT':
            return Response({"error": "This method not allowed for endpoint."})
        elif request_method == 'DELETE':
            return Response({"error": "This method not allowed for endpoint."})
        else:
            return Response({"error": "This method not allowed for endpoint."})
    except Exception as error:
        print(error)
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        return Response({"error": "This endpoint has error."})


@api_view(http_method_names=['GET', 'POST', 'PUT', 'DELETE'])
def api_user_recover_password(request):
    """
    Api django-rest-framework recover password
    """

    # logging
    backend_service.DjangoClass.LoggingClass.logging_actions(request=request)

    print('\n\n\n')
    print('api_user_recover_password:')
    print('\n')
    request_method, request_action_type, request_user, request_body = \
        backend_service.DjangoClass.DRFClass.request_utils(request=request)
    print(f"datetime: {backend_utils.DateTimeUtils.get_current_datetime()}", f"method: {request_method}",
          f"action_type: {request_action_type}", f"user: {request_user}", f"body: {request_body}")

    try:
        if request_method == 'GET':
            return Response({"error": "This method not allowed for endpoint."})
        elif request_method == 'POST':
            if request_action_type == "FIND_USER":
                try:
                    username = request_body["username"]

                    user = User.objects.get(username=username)
                    user_model = backend_models.UserModel.objects.get(user_foreign_key_field=user)
                    if user_model.temp_password_boolean_field == True:
                        return Response({"error": {"Пользователь ещё ни разу не менял пароль!"}})
                    response = {"response": {
                        "username": str(user.username),
                        "secret_question_char_field": str(user_model.secret_question_char_field),
                        "email_field": str(user_model.email_field),
                        "success": False
                    }}
                    print(f"response: {response}")
                    return Response(response)
                except Exception as error:
                    print(error)
                    backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                    return Response({"error": "This action has error."})
            if request_action_type == "CHECK_ANSWER":
                try:
                    username = request_body["username"]
                    secret_answer_char_field = request_body["secretAnswer"]

                    user = User.objects.get(username=username)
                    user_model = backend_models.UserModel.objects.get(user_foreign_key_field=user)

                    if str(secret_answer_char_field).strip().lower() == \
                            str(user_model.secret_answer_char_field).strip().lower():
                        response = {"response": {
                            "username": user.username,
                            "success": True
                        }}
                    else:
                        response = {"error": "Ответ не верный!"}
                    print(f"response: {response}")
                    return Response(response)
                except Exception as error:
                    print(error)
                    backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                    return Response({"error": "This action has error."})
            if request_action_type == "SEND_EMAIL_PASSWORD":
                try:
                    username = request_body["username"]

                    user = User.objects.get(username=username)
                    user_model = backend_models.UserModel.objects.get(user_foreign_key_field=user)
                    password = str(user_model.password_slug_field)
                    email_ = str(user_model.email_field)

                    ip = request.META.get("REMOTE_ADDR")
                    request_path = request.path
                    request_method = request.method
                    now = (datetime.datetime.now()).strftime('%Y-%m-%d %H:%M')
                    access_count = 0
                    for dat in backend_models.LoggingModel.objects.filter(
                            username_slug_field=username,
                            ip_genericipaddress_field=ip,
                            request_path_slug_field=request_path,
                            request_method_slug_field=request_method,
                            error_text_field=f'action: SEND_EMAIL_PASSWORD'
                    ):
                        if (dat.datetime_field + datetime.timedelta(hours=6, minutes=1)).strftime('%Y-%m-%d %H:%M') \
                                >= now:
                            access_count += 1

                    if access_count < 1:
                        backend_models.LoggingModel.objects.create(
                            username_slug_field=username,
                            ip_genericipaddress_field=ip,
                            request_path_slug_field=request_path,
                            request_method_slug_field=request_method,
                            error_text_field=f'action: SEND_EMAIL_PASSWORD'
                        )

                        text = f"{datetime.datetime.now().strftime('%Y-%m-%dT%H%M')}_{password[-1]}" \
                               f"{str(user_model.user_foreign_key_field)}{password}"
                        encrypt_text = backend_utils.EncryptingClass.encrypt_text(
                            text,
                            '31284'
                        )
                        # print(f"send pass to email: {encrypt_text}")

                        subject = 'Восстановление пароля от веб платформы'
                        message_s = f'{user_model.first_name_char_field} {user_model.last_name_char_field}, ' \
                                    f'перейдите по ссылке: http://web.km.kz:88/recover_password => ' \
                                    f'восстановление пароля, введите Ваш идентификатор и затем в окне почты ' \
                                    f'введите код (без кавычек): "{encrypt_text}". Внимание, этот код действует ' \
                                    f'в течении часа с момента отправки!'
                        from_email = 'web@km.kz'
                        to_email = email_
                        if subject and message_s and to_email:
                            send_mail(subject, message_s, from_email, [to_email, ''], fail_silently=False)

                        response = {"response": {
                            "username": str(user.username),
                            "secret_question_char_field": str(user_model.secret_question_char_field),
                            "email_field": str(user_model.email_field),
                            "success": False
                        }}
                    else:
                        response = {"error": f"Внимание, отправлять письмо можно не чаще раза в 3 минуты!"}
                    print(f"response: {response}")
                    return Response(response)
                except Exception as error:
                    print(error)
                    backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                    return Response({"error": "This action has error."})
            if request_action_type == "CHECK_EMAIL_PASSWORD":
                try:
                    username = request_body["username"]
                    recover_password = request_body["recoverPassword"]

                    user = User.objects.get(username=username)
                    user_model = backend_models.UserModel.objects.get(user_foreign_key_field=user)
                    password = str(user_model.password_slug_field)

                    decrypt_text = backend_utils.EncryptingClass.decrypt_text(recover_password, '31284')
                    text = f"{datetime.datetime.now().strftime('%Y-%m-%dT%H%M')}_{password[-1]}" \
                           f"{str(user_model.user_foreign_key_field)}{password}"

                    time1 = int(decrypt_text.split('_')[0].split('T')[1])
                    time2 = int(text.split('_')[0].split('T')[1])

                    # vxvvyxvyxpTvvssMwoqxpxwuswwqowvutsrqpM:H1Mw

                    if time1 - time2 > -60:
                        if str(decrypt_text.split('_')[1]).strip() == str(text.split('_')[1]).strip():
                            response = {"response": {
                                "username": user.username,
                                "success": True
                            }}
                        else:
                            response = {"error": f"Код не верный!"}
                    else:
                        response = {"error": f"Код не верный!"}
                    print(f"response: {response}")
                    return Response(response)
                except Exception as error:
                    print(error)
                    backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                    return Response({"error": "This action has error."})
            if request_action_type == "CHANGE_PASSWORD":
                try:
                    username = request_body["username"]
                    password = str(request_body["password"]).strip()
                    password2 = str(request_body["password2"]).strip()

                    user = User.objects.get(username=username)
                    user_model = backend_models.UserModel.objects.get(user_foreign_key_field=user)

                    if password == password2 and password != str(user_model.password_slug_field).strip():
                        user.set_password(password)
                        user.save()
                        user_model.password_slug_field = password
                        user_model.temp_password_boolean_field = False
                        user_model.save()
                        response = {"response": {
                            "username": user.username,
                            "success": False
                        }}
                    else:
                        response = {"error": f"Пароли не совпадают или старый пароль идентичный!"}
                    print(f"response: {response}")
                    return Response(response)
                except Exception as error:
                    print(error)
                    backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                    return Response({"error": "This action has error."})
            else:
                return Response({"error": "This action not allowed for this method."})
        elif request_method == 'PUT':
            return Response({"error": "This method not allowed for endpoint."})
        elif request_method == 'DELETE':
            return Response({"error": "This method not allowed for endpoint."})
        else:
            return Response({"error": "This method not allowed for endpoint."})
    except Exception as error:
        print(error)
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        return Response({"error": "This endpoint has error."})


@api_view(http_method_names=['GET', 'POST', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated, IsAdminUser])
def api_user_all(request):
    """
    Api django-rest-framework user all
    """

    # logging
    backend_service.DjangoClass.LoggingClass.logging_actions(request=request)

    print('\n\n\n')
    print('api_user_all:')
    print('\n')
    request_method, request_action_type, request_user, request_body = \
        backend_service.DjangoClass.DRFClass.request_utils(request=request)
    print(f"datetime: {backend_utils.DateTimeUtils.get_current_datetime()} | ", f"method: {request_method} | ",
          f"action_type: {request_action_type} | ", f"user: {request_user} | ", f"body: {request_body} | ")

    try:
        if request_method == 'GET':
            return Response({"error": "This method not allowed for endpoint."})
        elif request_method == 'POST':
            if request_action_type == "ALL":
                try:
                    users = User.objects.all()
                    serializer = backend_serializers.UserSerializer(users, many=True)
                    response = {"response": serializer.data}
                    print(f"response: {response}")
                    return Response(response)
                except Exception as error:
                    print(error)
                    backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                    return Response({"error": "This action has error."})
            else:
                return Response({"error": "This action not allowed for this method."})
        elif request_method == 'PUT':
            return Response({"error": "This method not allowed for endpoint."})
        elif request_method == 'DELETE':
            return Response({"error": "This method not allowed for endpoint."})
        else:
            return Response({"error": "This method not allowed for endpoint."})
    except Exception as error:
        print(error)
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        return Response({"error": "This endpoint has error."})


@api_view(http_method_names=['GET', 'POST', 'PUT', 'DELETE'])
def api_user_temp_all(request):
    """
    Api django-rest-framework user temp_all
    """

    # logging
    backend_service.DjangoClass.LoggingClass.logging_actions(request=request)

    print('\n\n\n')
    print('api_user_temp_all:')
    print('\n')
    request_method, request_action_type, request_user, request_body = \
        backend_service.DjangoClass.DRFClass.request_utils(request=request)
    print(f"datetime: {backend_utils.DateTimeUtils.get_current_datetime()} | ", f"method: {request_method} | ",
          f"action_type: {request_action_type} | ", f"user: {request_user} | ", f"body: {request_body} | ")

    user_models = backend_models.UserModel.objects.filter(temp_password_boolean_field=True)

    objects = []
    for user_model in user_models:
        if not user_model.user_foreign_key_field.is_superuser:
            objects.append(
                {
                    f"{str(user_model.user_foreign_key_field)}":
                        base64.b64encode(str(f"12{user_model.password_slug_field}345").encode()).decode()
                }
            )

    # for obj in objects:
    #     for key, value in obj.items():
    #         print(f"{key}: {str(base64.b64decode(value).decode())[2: -3]}")

    return Response({"response": objects})


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = backend_serializers.UserSerializer
    permission_classes = [permissions.AllowAny]


class GroupViewSet(viewsets.ModelViewSet):
    queryset = Group.objects.all()
    serializer_class = backend_serializers.GroupSerializer
    permission_classes = [permissions.AllowAny]


class ChatViewSet(viewsets.ModelViewSet):
    queryset = backend_models.ChatModel.objects.all()
    serializer_class = backend_serializers.ChatModelSerializer
    permission_classes = [permissions.AllowAny]


@api_view(http_method_names=['GET', 'POST', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def api_salary(request):
    """
    Api django-rest-framework salary
    """

    # logging
    backend_service.DjangoClass.LoggingClass.logging_actions(request=request)

    print('\n\n\n')
    print('api_salary:')
    print('\n')
    request_method, request_action_type, request_user, request_body = \
        backend_service.DjangoClass.DRFClass.request_utils(request=request)
    print(f"datetime: {backend_utils.DateTimeUtils.get_current_datetime()} | ", f"method: {request_method} | ",
          f"action_type: {request_action_type} | ", f"user: {request_user} | ", f"body: {request_body} | ")

    try:
        if request_method == 'GET':
            return Response({"error": "This method not allowed for endpoint."})
        elif request_method == 'POST':
            if request_action_type == "SALARY":
                try:
                    # Get json response from 1c
                    ####################################################################################################
                    key = backend_service.UtilsClass.create_encrypted_password(
                        _random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890',
                        _length=10
                    )
                    hash_key_obj = hashlib.sha256()
                    hash_key_obj.update(key.encode('utf-8'))
                    key_hash = str(hash_key_obj.hexdigest().strip().upper())
                    key_hash_base64 = base64.b64encode(str(key_hash).encode()).decode()
                    iin = request_user.username
                    if str(iin).lower() == 'bogdan':
                        iin = 970801351179
                    iin_base64 = base64.b64encode(str(iin).encode()).decode()
                    date_base64 = base64.b64encode(f'{request_body["dateTime"]}'.encode()).decode()
                    url = f'http://192.168.1.10/KM_1C/hs/zp/rl/{iin_base64}_{key_hash_base64}/{date_base64}'
                    relative_path = os.path.dirname(os.path.abspath('__file__')) + '\\'
                    h = httplib2.Http(relative_path + "\\static\\media\\data\\temp\\get_salary_data")
                    _login = 'Web_adm_1c'
                    password = '159159qqww!'
                    h.add_credentials(_login, password)
                    response, content = h.request(url)
                    data = backend_service.UtilsClass.decrypt_text_with_hash(content.decode()[1:], key_hash)
                    error_word_list = ['ошибка', 'error', 'failed']
                    if data.find('send') == 0:
                        return Response({"error": f"{data.split('send')[1].strip()}"})
                    for error_word in error_word_list:
                        if data.find(error_word.lower()) >= 0:
                            return Response({"error": f"Неизвестная ошибка."})

                    json_data = json.loads(data)
                    try:
                        json_data["global_objects"]["3.Доходы в натуральной форме"]
                    except Exception as error:
                        json_data["global_objects"]["3.Доходы в натуральной форме"] = {
                            "Fields": {
                                "1": "Вид",
                                "2": "Период",
                                "3": "Сумма"
                            },
                        }
                    ####################################################################################################

                    # Get local test json response from 1c
                    ####################################################################################################
                    # Временное чтение файла для отладки без доступа к 1С
                    # with open("static/media/data/json_data.json", "r", encoding="utf-8") as file:
                    #     json_data = json.loads(json.load(file))
                    # time.sleep(1)
                    print(json_data)
                    ####################################################################################################

                    # Return pretty integer and float value
                    ####################################################################################################
                    def return_float_value(_value):
                        if isinstance(_value, int) or isinstance(_value, float):
                            if len(f'{_value:.2f}') < 10:
                                _value = f'{_value:.2f}'[:-6] + ' ' + f'{_value:.2f}'[-6:]
                            else:
                                _value = f'{_value:.2f}'[:-9] + ' ' + f'{_value:.2f}'[-9:-6] + ' ' + f'{_value:.2f}'[
                                                                                                     -6:]
                        return _value
                    ####################################################################################################

                    # Create 'Ends' and pretty integer and float value
                    ####################################################################################################
                    def create_ends(table: str, extracols=False):
                        if extracols:
                            _days = 0
                            _hours = 0
                            _summ = 0
                            for __key in json_data['global_objects'][table].keys():
                                if __key != 'Fields':
                                    try:
                                        _days += json_data['global_objects'][table][f'{__key}']['ВсегоДни']
                                        _hours += json_data['global_objects'][table][f'{__key}']['ВсегоЧасы']
                                        _summ_local = json_data['global_objects'][table][f'{__key}']['Сумма']
                                        json_data['global_objects'][table][f'{__key}']['Сумма'] = return_float_value(
                                            _summ_local)
                                        _summ += _summ_local
                                    except Exception as _error:
                                        print(_error)
                            json_data['global_objects'][table]['Ends'] = {
                                'Вид': 'Итого', 'Период': '', 'Дни': _days, 'Часы': _hours,
                                'ВсегоДни': 0, 'ВсегоЧасы': 0, 'Сумма': return_float_value(_summ)
                            }
                        else:
                            _summ = 0
                            for __key in json_data['global_objects'][table].keys():
                                if __key != 'Fields':
                                    try:
                                        _summ_local = json_data['global_objects'][table][f'{__key}']['Сумма']
                                        json_data['global_objects'][table][f'{__key}']['Сумма'] = return_float_value(
                                            _summ_local)
                                        _summ += _summ_local
                                    except Exception as _error:
                                        print(_error)
                            json_data['global_objects'][table]['Ends'] = {
                                'Вид': 'Итого', 'Период': '', 'Сумма': return_float_value(_summ)
                            }

                    create_ends(table='1.Начислено', extracols=True)
                    create_ends(table='2.Удержано', extracols=False)
                    create_ends(table='3.Доходы в натуральной форме', extracols=False)
                    create_ends(table='4.Выплачено', extracols=False)
                    create_ends(table='5.Налоговые вычеты', extracols=False)
                    ####################################################################################################

                    # pretty integer and float value in headers
                    ####################################################################################################
                    for _key in json_data.keys():
                        if _key != 'global_objects':
                            json_data[_key] = return_float_value(json_data[_key])
                    ####################################################################################################
                    try:
                        # create excel
                        ################################################################################################
                        key = backend_service.UtilsClass.create_encrypted_password(
                            _random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890',
                            _length=24
                        )
                        date = backend_utils.DateTimeUtils.get_current_date()
                        path = 'static/media/salary/'
                        path_excel_file = f"{path}salary_{key}_{date}.xlsx"
                        workbook = backend_utils.ExcelClass.workbook_create()
                        sheet = backend_utils.ExcelClass.workbook_activate(workbook)

                        # Delete old files
                        #######################################################
                        for root, dirs, files in os.walk(path, topdown=True):
                            for file in files:
                                try:
                                    date_file = str(file).strip().split('.')[0].strip().split('_')[-1]
                                    if date != date_file:
                                        os.remove(f'{path}{file}')
                                except Exception as error:
                                    print(error)
                        #######################################################

                        # Create 'TitleComponent'
                        #######################################################
                        backend_utils.ExcelClass.set_sheet_value(
                            col=1,
                            row=1,
                            value='РАСЧЕТНЫЙ ЛИСТ',
                            sheet=sheet
                        )
                        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
                        #######################################################

                        # Create 'Headers'
                        #######################################################
                        header_arr = []
                        for key, value in json_data.items():
                            if key != 'global_objects' and key != 'Долг за организацией на начало месяца' \
                                    and key != 'Долг за организацией на конец месяца':
                                header_arr.append(f'{key}: {value}')

                        header_len_devide = len(header_arr) - 8
                        if header_len_devide % 2 != 0:
                            header_len_devide += 1

                        row_i_1 = 1 + 1
                        for header in header_arr[0:4]:
                            col_i = 1
                            backend_utils.ExcelClass.set_sheet_value(
                                col=col_i,
                                row=row_i_1,
                                value=header,
                                sheet=sheet
                            )
                            sheet.merge_cells(start_row=row_i_1, start_column=col_i, end_row=row_i_1,
                                              end_column=col_i + 4)
                            row_i_1 += 1

                        row_i_2 = 1 + 1
                        for header in header_arr[4:8]:
                            col_i = 6
                            backend_utils.ExcelClass.set_sheet_value(
                                col=col_i,
                                row=row_i_2,
                                value=header,
                                sheet=sheet
                            )
                            sheet.merge_cells(start_row=row_i_2, start_column=col_i, end_row=row_i_2,
                                              end_column=col_i + 2)
                            row_i_2 += 1

                        row_i_3 = row_i_1
                        for header in header_arr[8:8 + header_len_devide // 2]:
                            col_i = 1
                            backend_utils.ExcelClass.set_sheet_value(
                                col=col_i,
                                row=row_i_3,
                                value=header,
                                sheet=sheet
                            )
                            sheet.merge_cells(start_row=row_i_3, start_column=col_i, end_row=row_i_3,
                                              end_column=col_i + 4)
                            row_i_3 += 1

                        row_i_4 = row_i_2
                        for header in header_arr[8 + header_len_devide // 2:]:
                            col_i = 6
                            backend_utils.ExcelClass.set_sheet_value(
                                col=col_i,
                                row=row_i_4,
                                value=header,
                                sheet=sheet
                            )
                            sheet.merge_cells(start_row=row_i_4, start_column=col_i, end_row=row_i_4,
                                              end_column=col_i + 2)
                            row_i_4 += 1
                        header_low_row = row_i_4

                        #######################################################

                        # Create 'Bodyes'
                        #######################################################
                        def create_bodyes(table: str, extracols=False):
                            bodyes_arr = [[table]]
                            for __key, _value in json_data['global_objects'][table].items():
                                local_bodyes_arr = []
                                for ___key, __value in json_data['global_objects'][table][__key].items():
                                    if extracols:
                                        if ___key != '6' and ___key != '7' and \
                                                ___key != 'ВсегоДни' and ___key != 'ВсегоЧасы':
                                            local_bodyes_arr.append(__value)
                                    else:
                                        local_bodyes_arr.append(__value)
                                bodyes_arr.append(local_bodyes_arr)
                            return bodyes_arr

                        bodyes_arr_1 = create_bodyes('1.Начислено', extracols=True)
                        bodyes_arr_2 = create_bodyes('2.Удержано', extracols=False)
                        bodyes_arr_3 = create_bodyes('3.Доходы в натуральной форме', extracols=False)
                        bodyes_arr_4 = create_bodyes('4.Выплачено', extracols=False)
                        bodyes_arr_5 = create_bodyes('5.Налоговые вычеты', extracols=False)

                        bold_arr = []

                        body_low_row_1 = header_low_row
                        bold_arr.append(body_low_row_1)
                        sheet.merge_cells(start_row=body_low_row_1, start_column=1, end_row=body_low_row_1,
                                          end_column=1 + 4)
                        for body in bodyes_arr_1:
                            col_i = 1
                            for value in body:
                                if isinstance(value, int) and value == 0:
                                    value = ''
                                backend_utils.ExcelClass.set_sheet_value(
                                    col=col_i,
                                    row=body_low_row_1,
                                    value=value,
                                    sheet=sheet
                                )
                                col_i += 1
                            body_low_row_1 += 1

                        body_low_row_2 = header_low_row
                        bold_arr.append(body_low_row_2)
                        sheet.merge_cells(start_row=body_low_row_2, start_column=6, end_row=body_low_row_2,
                                          end_column=6 + 2)
                        for body in bodyes_arr_2:
                            col_i = 6
                            for value in body:
                                backend_utils.ExcelClass.set_sheet_value(
                                    col=col_i,
                                    row=body_low_row_2,
                                    value=value,
                                    sheet=sheet
                                )
                                col_i += 1
                            body_low_row_2 += 1

                        if body_low_row_1 >= body_low_row_2:
                            body_low_row_3 = body_low_row_1
                            body_low_row_4 = body_low_row_1
                        else:
                            body_low_row_3 = body_low_row_2
                            body_low_row_4 = body_low_row_2

                        bold_arr.append(body_low_row_3)
                        sheet.merge_cells(start_row=body_low_row_3, start_column=1, end_row=body_low_row_3,
                                          end_column=1 + 4)
                        for body in bodyes_arr_3:
                            col_i = 1
                            for value in body:
                                backend_utils.ExcelClass.set_sheet_value(
                                    col=col_i,
                                    row=body_low_row_3,
                                    value=value,
                                    sheet=sheet
                                )
                                col_i += 1
                            body_low_row_3 += 1

                        bold_arr.append(body_low_row_4)
                        sheet.merge_cells(start_row=body_low_row_4, start_column=6, end_row=body_low_row_4,
                                          end_column=6 + 2)
                        for body in bodyes_arr_4:
                            col_i = 6
                            for value in body:
                                backend_utils.ExcelClass.set_sheet_value(
                                    col=col_i,
                                    row=body_low_row_4,
                                    value=value,
                                    sheet=sheet
                                )
                                col_i += 1
                            body_low_row_4 += 1

                        if body_low_row_3 >= body_low_row_4:
                            body_low_row_5 = body_low_row_3
                            body_low_row_6 = body_low_row_3
                        else:
                            body_low_row_5 = body_low_row_4
                            body_low_row_6 = body_low_row_4

                        bold_arr.append(body_low_row_5)
                        sheet.merge_cells(start_row=body_low_row_5, start_column=1, end_row=body_low_row_5,
                                          end_column=1 + 4)
                        for body in bodyes_arr_5:
                            col_i = 1
                            for value in body:
                                backend_utils.ExcelClass.set_sheet_value(
                                    col=col_i,
                                    row=body_low_row_5,
                                    value=value,
                                    sheet=sheet
                                )
                                col_i += 1
                            body_low_row_5 += 1

                        lowest = [
                            f'Долг за организацией на начало месяца: '
                            f'{json_data["Долг за организацией на начало месяца"]}',
                            f'Долг за организацией на конец месяца: '
                            f'{json_data["Долг за организацией на конец месяца"]}',
                        ]

                        bold_arr.append(body_low_row_6)
                        for body in ['.', 'Вид', *lowest]:
                            col_i = 6
                            backend_utils.ExcelClass.set_sheet_value(
                                col=col_i,
                                row=body_low_row_6,
                                value=body,
                                sheet=sheet
                            )
                            sheet.merge_cells(start_row=body_low_row_6, start_column=col_i, end_row=body_low_row_6,
                                              end_column=col_i + 2)
                            body_low_row_6 += 1

                        if body_low_row_6 >= body_low_row_5:
                            body_low_row = body_low_row_6
                        else:
                            body_low_row = body_low_row_5

                        if body_low_row_1 >= body_low_row_2:
                            body_color_2 = body_low_row_1
                        else:
                            body_color_2 = body_low_row_2
                        if body_low_row_3 >= body_low_row_4:
                            body_color_3 = body_low_row_3
                        else:
                            body_color_3 = body_low_row_4
                        #######################################################

                        # Set fonts
                        #######################################################
                        font_headers = Font(name='Arial', size=8, bold=False)
                        for row in range(1, header_low_row):
                            for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                cell = sheet[f'{col}{row}']
                                cell.font = font_headers

                        font_bodyes = Font(name='Arial', size=7, bold=False)
                        for row in range(header_low_row, body_low_row + 1):
                            for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                cell = sheet[f'{col}{row}']
                                cell.font = font_bodyes

                        font_tables = Font(name='Arial', size=8, bold=True)
                        for row in [header_low_row, body_color_2, body_color_3]:
                            for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                cell = sheet[f'{col}{row}']
                                cell.font = font_tables
                        #######################################################

                        # Set aligments
                        #######################################################
                        # wrap_text = Alignment(wrap_text=True)
                        # shrink_to_fit = Alignment(shrink_to_fit=True)
                        aligment_center = Alignment(horizontal='center', vertical='center', wrap_text=True,
                                                    shrink_to_fit=True)
                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                            cell = sheet[f'{col}{1}']
                            cell.alignment = aligment_center
                        aligment_left = Alignment(horizontal='left', vertical='center', wrap_text=True,
                                                  shrink_to_fit=True)
                        for row in range(2, header_low_row):
                            for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                cell = sheet[f'{col}{row}']
                                cell.alignment = aligment_left

                        aligment_right = Alignment(horizontal='right', vertical='center', wrap_text=True,
                                                   shrink_to_fit=True)
                        for row in range(header_low_row, body_low_row + 1):
                            for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                cell = sheet[f'{col}{row}']
                                if col == 'A' or col == 'F':
                                    cell.alignment = aligment_left
                                elif col == 'E' or col == 'H':
                                    cell.alignment = aligment_right
                                else:
                                    cell.alignment = aligment_center
                        #######################################################

                        # Set borders
                        #######################################################
                        side_medium = Side(border_style="thin", color="FF808080")
                        # side_think = Side(border_style="thin", color="FF808080")
                        border_horizontal_middle = Border(
                            top=side_medium,
                            # left=side_medium,
                            # right=side_medium,
                            # bottom=side_medium
                        )
                        border_vertical_middle = Border(
                            # top=side_think,
                            left=side_medium,
                            # right=side_think,
                            # bottom=side_think
                        )
                        border_vertical_light = Border(
                            # top=side_think,
                            left=side_medium,
                            # right=side_think,
                            # bottom=side_think
                        )
                        for row in range(header_low_row, body_low_row):
                            for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                if col == 'G' and row > body_low_row_4 or col == 'H' and row > body_low_row_4:
                                    pass
                                else:
                                    cell = sheet[f'{col}{row}']
                                    cell.border = border_vertical_light
                            cell = sheet[f'{backend_utils.ExcelClass.get_column_letter(1)}{row}']
                            cell.border = border_vertical_middle
                            cell = sheet[f'{backend_utils.ExcelClass.get_column_letter(6)}{row}']
                            cell.border = border_vertical_middle
                            cell = sheet[f'{backend_utils.ExcelClass.get_column_letter(9)}{row}']
                            cell.border = border_vertical_middle
                        side_think = Side(border_style="dotted", color="FF808080")
                        # {'mediumDashDotDot', 'thin', 'dashed', 'mediumDashed', 'dotted', 'double', 'thick',
                        # 'medium', 'dashDot','dashDotDot', 'hair', 'mediumDashDot', 'slantDashDot'}
                        border_think = Border(
                            top=side_think,
                            left=side_think,
                            right=side_think,
                            bottom=side_think
                        )
                        for row in range(1, header_low_row):
                            for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                cell = sheet[f'{col}{row}']
                                cell.border = border_think
                        side_medium = Side(border_style="thin", color="FF808080")
                        border_medium = Border(
                            top=side_medium,
                            left=side_medium,
                            right=side_medium,
                            bottom=side_medium
                        )
                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                            for row in [header_low_row + 1, body_color_2 + 1, body_color_3 + 1]:
                                cell = sheet[f'{col}{row - 1}']
                                cell.border = border_medium
                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                            cell = sheet[f'{col}{body_low_row}']
                            cell.border = border_horizontal_middle
                        #######################################################

                        # Colored styles
                        #######################################################
                        color_green = PatternFill(fgColor="E6E6FF", fill_type="solid")
                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                            for row in [header_low_row + 1, body_color_2 + 1, body_color_3 + 1]:
                                cell = sheet[f'{col}{row}']
                                cell.fill = color_green
                                cell = sheet[f'{col}{row - 1}']
                                cell.border = border_medium
                        color_yellow = PatternFill(fgColor="d0ffd8", fill_type="solid")
                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 5 + 1)]:
                            cell = sheet[f'{col}{body_low_row_1 - 1}']
                            cell.fill = color_yellow

                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(6, 8 + 1)]:
                            cell = sheet[f'{col}{body_low_row_2 - 1}']
                            cell.fill = color_yellow

                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 3 + 1)]:
                            cell = sheet[f'{col}{body_low_row_3 - 1}']
                            cell.fill = color_yellow
                            cell.fill = color_yellow

                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(6, 8 + 1)]:
                            cell = sheet[f'{col}{body_low_row_4 - 1}']
                            cell.fill = color_yellow

                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 3 + 1)]:
                            cell = sheet[f'{col}{body_low_row_5 - 1}']
                            cell.fill = color_yellow
                            cell.fill = color_yellow

                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(6, 8 + 1)]:
                            cell = sheet[f'{col}{body_low_row_6 - 1}']
                            cell.fill = color_yellow

                        #######################################################

                        # Height and width styles
                        #######################################################
                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                            width = 1
                            for row in range(1, body_low_row + 1):
                                cell = sheet[f'{col}{row}']
                                value = len(str(cell.value))
                                if value > width:
                                    width = value
                            if col == 'A' or col == 'F':
                                width = width / 2
                            sheet.column_dimensions[col].height = 1
                            sheet.column_dimensions[col].width = round((width * 0.95), 3)
                        #######################################################

                        # Set global page and book settings
                        #######################################################
                        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
                        sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
                        sheet.page_margins.left = 0.05
                        sheet.page_margins.right = 0.05
                        sheet.page_margins.header = 0.1
                        sheet.page_margins.bottom = 0.2
                        sheet.page_margins.footer = 0.2
                        sheet.page_margins.top = 0.1
                        sheet.print_options.horizontalCentered = True
                        # sheet.print_options.verticalCentered = True
                        sheet.page_setup.fitToHeight = 1
                        sheet.page_setup.scale = 80
                        sheet.page_setup.fitToHeight = 1
                        sheet.page_setup.fitToWidth = 1
                        sheet.protection.password = key + '_1'
                        sheet.protection.sheet = True
                        sheet.protection.enable()
                        #######################################################
                        backend_utils.ExcelClass.workbook_save(workbook=workbook, excel_file=path_excel_file)
                        json_data["excel_path"] = path_excel_file
                    except Exception as error:
                        print(error)
                        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                        json_data["excel_path"] = ''
                    response = {"response": json_data}
                    # print(f"response: {response}")
                    return Response(response)
                except Exception as error:
                    print(error)
                    backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                    return Response({"error": "This action has error."})
            else:
                return Response({"error": "This action not allowed for this method."})
        elif request_method == 'PUT':
            return Response({"error": "This method not allowed for endpoint."})
        elif request_method == 'DELETE':
            return Response({"error": "This method not allowed for endpoint."})
        else:
            return Response({"error": "This method not allowed for endpoint."})
    except Exception as error:
        print(error)
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        return Response({"error": "This endpoint has error."})


########################################################################################################################

# @api_view(http_method_names=['GET'])
# @permission_classes([IsAuthenticated])
# def get_user_profile(request):
#     print(f'\n\nget_user_profile:\n\n')
#     user = request.user
#     serializer = backend_serializers.UserSerializer(user, many=False)
#     return Response(serializer.data)
#
#
# @api_view(http_method_names=['POST'])
# @permission_classes([IsAuthenticated])
# def change_user_profile(request):
#     print(f'\n\nchange_user_profile:\n\n')
#     print('data: ', request.data)
#     user = User.objects.get(username=request.user.username)
#     print('user: ', user)
#     user_model = backend_models.UserModel.objects.get(user_foreign_key_field=user)
#     try:
#         password = str(request.data["password"]).strip()
#         password2 = str(request.data["password2"]).strip()
#         if password and password2 and password != user_model.password_slug_field:
#             user_model.password_slug_field = password
#             user.set_password(password)
#             print(f'password changed to "{password}"')
#             # print(make_password(request.data["password"]))
#             user.save()
#     except Exception as error:
#         print(error)
#     try:
#         if request.data["email"] and request.data["email"] != user_model.email_field:
#             user_model.email_field = request.data["email"]
#         if request.data["secretQuestion"] and request.data["secretQuestion"] != user_model.secret_question_char_field:
#             user_model.secret_question_char_field = request.data["secretQuestion"]
#         if request.data["secretAnswer"] and request.data["secretAnswer"] != user_model.secret_answer_char_field:
#             user_model.secret_answer_char_field = request.data["secretAnswer"]
#     except Exception as error:
#         print(error)
#     user_model.save()
#     return Response({'error': False})
#
#
# @api_view(http_method_names=['POST'])
# def recover_user_password(request):
#     print(f'\n\nrecover_user_password:\n\n')
#     response = {"error": "error"}
#     if request.method == "POST":
#         try:
#             username = request.data["username"]
#         except Exception as error:
#             username = ''
#
#         try:
#             secret_answer_char_field = request.data["secret_answer_char_field"]
#         except Exception as error:
#             secret_answer_char_field = ''
#
#         try:
#             password = str(request.data["password"]).strip().lower()
#             password2 = str(request.data["password2"]).strip().lower()
#         except Exception as error:
#             password = ''
#             password2 = ''
#
#         if username and not secret_answer_char_field and not password:
#             try:
#                 user = User.objects.get(username=username)
#                 user_model = backend_models.UserModel.objects.get(user_foreign_key_field=user)
#                 secret_question_char_field = user_model.secret_question_char_field
#                 response = {
#                     "username": user.username,
#                     "secret_question_char_field": secret_question_char_field,
#                     "success": False
#                 }
#             except Exception as error:
#                 print(error)
#         elif username and secret_answer_char_field and not password:
#             try:
#                 user = User.objects.get(username=username)
#                 user_model = backend_models.UserModel.objects.get(user_foreign_key_field=user)
#                 if str(secret_answer_char_field).strip().lower() == \
#                         str(user_model.secret_answer_char_field).strip().lower():
#                     response = {
#                         "username": user.username,
#                         "secret_question_char_field": None,
#                         "success": True
#                     }
#                 else:
#                     response = {
#                         "username": user.username,
#                         "secret_question_char_field": None,
#                         "success": False
#                     }
#             except Exception as error:
#                 print(error)
#         elif username and not secret_answer_char_field and password:
#             try:
#                 user = User.objects.get(username=username)
#                 user_model = backend_models.UserModel.objects.get(user_foreign_key_field=user)
#                 if password == password2 and str(user_model.password_slug_field).strip().lower():
#                     user.set_password(password)
#                     user.save()
#                     user_model.password_slug_field = password
#                     user_model.save()
#                     response = {
#                         "username": user.username,
#                         "secret_question_char_field": None,
#                         "success": False
#                     }
#                 else:
#                     response = {
#                         "username": user.username,
#                         "secret_question_char_field": None,
#                         "success": False
#                     }
#             except Exception as error:
#                 print(error)
#
#     print("response: ", response)
#     return Response(response)
#
#
# @api_view(http_method_names=['GET'])
# def recover_user_password_email(request):
#     print(f'\n\nrecover_user_password:\n\n')
#     response = {"error": "error"}
#     if request.method == "POST":
#         try:
#             username = request.data["username"]
#             user = User.objects.get(username=username)
#             user_model = backend_models.UserModel.objects.get(user_foreign_key_field=user)
#         except Exception as error:
#             username = None
#             user = None
#             user_model = None
#
#         try:
#             password = user_model.password_slug_field
#             email_ = user_model.email_field
#             if password and email_:
#                 subject = 'Восстановление пароля от веб платформы'
#                 encrypt_message = backend_utils.EncryptingClass.encrypt_text(
#                     text=f'_{password}_{datetime.datetime.now().strftime("%Y-%m-%dT%H:%M")}_',
#                     hash_chars=f'_{password}_{datetime.datetime.now().strftime("%Y-%m-%dT%H:%M")}_'
#                 )
#                 message_s = f'{user_model.first_name_char_field} {user_model.last_name_char_field}, ' \
#                             f'перейдите по ссылке: http://192.168.1.68:80/account_recover_password/ , ' \
#                             f'введите иин и затем в окне почты введите код (без кавычек): "{encrypt_message}"'
#                 # from_email = 'eevee.cycle@yandex.ru'
#                 from_email = 'web.km.kz'
#                 to_email = email_
#                 if subject and message and to_email:
#                     send_mail(subject, message_s, from_email, [to_email, ''], fail_silently=False)
#                     response = 2
#                 else:
#                     response = -2
#             else:
#                 response = -2
#         except Exception as error:
#             backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
#
#         try:
#             encrypt_text = backend_service.DjangoClass.RequestClass.get_value(request, "recover")
#             decrypt_text = backend_utils.EncryptingClass.decrypt_text(
#                 text=encrypt_text,
#                 hash_chars=f'_{user_model.password_slug_field}_'
#                            f'{datetime.datetime.now().strftime("%Y-%m-%dT%H:%M")}_'
#             )
#             if decrypt_text.split('_')[2] >= \
#                     (datetime.datetime.now() - datetime.timedelta(hours=1)).strftime('%Y-%m-%d %H:%M') and \
#                     decrypt_text.split('_')[1] == user_model.password_slug_field:
#                 login(request, user)
#                 user_model.secret_question_char_field = ''
#                 user_model.secret_answer_char_field = ''
#                 user_model.save()
#                 response = 2
#             else:
#                 response = -2
#         except Exception as error:
#             backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
#
#     print("response: ", response)
#     return Response(response)
#
#
# @api_view(http_method_names=['GET'])
# def get_users(request):
#     print(f'\n\nget_users:\n\n')
#     user = User.objects.all()
#     serializer = backend_serializers.UserSerializer(user, many=True)
#     return Response(serializer.data)


# def get_tokens_for_user(user):
#     refresh = RefreshToken.for_user(user)
#
#     return {
#         'refresh': str(refresh),
#         'access': str(refresh.access_token),
#         'username': str(refresh.username),
#     }


# class MyTokenObtainPairSerializer(TokenObtainPairSerializer):
#     def validate(self, attrs):
#         data = super().validate(attrs)
#
#         if settings.SIMPLE_JWT["UPDATE_LAST_LOGIN"]:
#             update_last_login(None, self.user)
#
#         serializer = backend_serializers.UserSerializerWithToken(self.user)
#         for k, v in serializer.data.items():
#             data[f'{k}'] = v
#
#         try:
#             user = User.objects.get(username=data["username"])
#             user_model = backend_models.UserModel.objects.get(user_foreign_key_field=user)
#             data['name'] = f'{user_model.last_name_char_field} {user_model.first_name_char_field}'
#         except Exception as error:
#             data['name'] = f'Данные не заполнены'
#
#         # try:
#         #     user = User.objects.get(username=data["username"])
#         #     user_model = backend_models.UserModel.objects.get(user_foreign_key_field=user)
#         #     user_model_serializer = backend_serializers.UserModelSerializer(instance=user_model, many=False)
#         #     data['user_model'] = user_model_serializer.data
#         # except Exception as error:
#         #     data['user_model'] = {'error': f'{error}'}
#         #
#         # try:
#         #     user = User.objects.get(username=data["username"])
#         #     user_model = backend_models.UserModel.objects.get(user_foreign_key_field=user)
#         #     group_model = backend_models.GroupModel.objects.filter(user_many_to_many_field=user_model)
#         #     group_model_serializer = backend_serializers.GroupModelSerializer(instance=group_model, many=True)
#         #     data['group_model'] = group_model_serializer.data
#         # except Exception as error:
#         #     data['group_model'] = {'error': f'{error}'}
#
#         return data
#
#     @classmethod
#     def get_token(cls, user):
#         token = super().get_token(user)
#         token['username'] = user.username
#
#         return token


# class MyTokenObtainPairSerializer(TokenObtainPairSerializer):
#     @classmethod
#     def get_token(cls, user):
#         token = super().get_token(user)
#
#         # Add custom claims
#         token['username'] = user.username
#         # ...
#
#         return token


# class MyTokenObtainPairView(TokenObtainPairView):
#     serializer_class = MyTokenObtainPairSerializer


@api_view(http_method_names=['GET', 'POST', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def note_api(request, pk=None):
    # for key, value in request.META.items():
    #     print(f'{key}: {value}')
    # print(request)
    # print(request.user)
    # print(request.user.username)
    if request.method == 'GET':
        if pk == '-1':
            _routes = [
                {
                    'Endpoint': '$BASEPATH$/note_api/',
                    'method': 'GET',
                    'body': None,
                    'descriptions': 'Returns an array of notes objects',
                    'code': '''notes = NoteComponent.objects.all().order_by('-updated')
                    serializer = NoteSerializer(notes, many=True)
                    return Response(serializer.data)'''
                },
                {
                    'Endpoint': '$BASEPATH$/note_api/<id>/',
                    'method': 'GET',
                    'body': None,
                    'descriptions': 'Returns a single note object',
                    'code': '''note = NoteComponent.objects.get(id=pk)
                    serializer = NoteSerializer(note, many=False)
                    return Response(serializer.data)'''
                },
                {
                    'Endpoint': '$BASEPATH$/note_api/',
                    'method': 'POST',
                    'body': {'body': ""},
                    'descriptions': 'Creates new note with data sent in request',
                    'code': '''data = request.data
                    note = NoteComponent.objects.create(body=data['body'])
                    serializer = NoteSerializer(note, many=False)
                    return Response(serializer.data)'''
                },
                {
                    'Endpoint': '$BASEPATH$/note_api/<id>/',
                    'method': 'PUT',
                    'body': None,
                    'descriptions': 'Updates note with data sent in request',
                    'code': '''data = request.data
                    note = NoteComponent.objects.get(id=pk)
                    equal = data['body'] != note.body
                    serializer = NoteSerializer(instance=note, data=data)
                    if serializer.is_valid() and equal:
                        serializer.save()
                    return Response(serializer.data)'''
                },
                {
                    'Endpoint': '$BASEPATH$/note_api/<id>/',
                    'method': 'DELETE',
                    'body': None,
                    'descriptions': 'Deletes note',
                    'code': '''note = NoteComponent.objects.get(id=pk)
                    note.delete()
                    return Response(f'delete note №{pk} successfull')'''
                },
            ]
            return Response(_routes)
        elif pk:
            note = backend_models.NoteModel.objects.get(id=pk)
            serializer = backend_serializers.NoteSerializer(note, many=False)
            return Response(serializer.data)
        else:
            notes = backend_models.NoteModel.objects.all().order_by('-updated')
            serializer = backend_serializers.NoteSerializer(notes, many=True)
            return Response(serializer.data)
    elif request.method == 'POST':
        data = request.data
        note = backend_models.NoteModel.objects.create(body=data['body'],
                                                       user=User.objects.get(username=request.user.username))
        serializer = backend_serializers.NoteSerializer(note, many=False)
        return Response(serializer.data)
    elif request.method == 'PUT':
        data = request.data
        note = backend_models.NoteModel.objects.get(id=pk)
        equal = data['body'] != note.body
        serializer = backend_serializers.NoteSerializer(instance=note, data=data["body"])
        if serializer.is_valid() and equal:
            serializer.save()
        return Response(serializer.data)
    elif request.method == 'DELETE':
        note = backend_models.NoteModel.objects.get(id=pk)
        note.delete()
        return Response(f'delete note №{pk} successfull')


@api_view(http_method_names=['GET', 'POST', 'PUT', 'DELETE'])
def _products(request, pk=None):
    # for key, value in request.META.items():
    #     print(f'{key}: {value}')
    # print(request)
    # print(request.user)
    # print(request.user.username)
    if request.method == 'GET':
        if pk == '-1':
            _routes = [
                {
                    'Endpoint': '$BASEPATH$/note_api/',
                    'method': 'GET',
                    'body': None,
                    'descriptions': 'Returns an array of notes objects',
                    'code': '''notes = NoteComponent.objects.all().order_by('-updated')
                    serializer = NoteSerializer(notes, many=True)
                    return Response(serializer.data)'''
                },
                {
                    'Endpoint': '$BASEPATH$/note_api/<id>/',
                    'method': 'GET',
                    'body': None,
                    'descriptions': 'Returns a single note object',
                    'code': '''note = NoteComponent.objects.get(id=pk)
                    serializer = NoteSerializer(note, many=False)
                    return Response(serializer.data)'''
                },
                {
                    'Endpoint': '$BASEPATH$/note_api/',
                    'method': 'POST',
                    'body': {'body': ""},
                    'descriptions': 'Creates new note with data sent in request',
                    'code': '''data = request.data
                    note = NoteComponent.objects.create(body=data['body'])
                    serializer = NoteSerializer(note, many=False)
                    return Response(serializer.data)'''
                },
                {
                    'Endpoint': '$BASEPATH$/note_api/<id>/',
                    'method': 'PUT',
                    'body': None,
                    'descriptions': 'Updates note with data sent in request',
                    'code': '''data = request.data
                    note = NoteComponent.objects.get(id=pk)
                    equal = data['body'] != note.body
                    serializer = NoteSerializer(instance=note, data=data)
                    if serializer.is_valid() and equal:
                        serializer.save()
                    return Response(serializer.data)'''
                },
                {
                    'Endpoint': '$BASEPATH$/note_api/<id>/',
                    'method': 'DELETE',
                    'body': None,
                    'descriptions': 'Deletes note',
                    'code': '''note = NoteComponent.objects.get(id=pk)
                    note.delete()
                    return Response(f'delete note №{pk} successfull')'''
                },
            ]
            return Response(_routes)
        elif pk:
            note = backend_models.NoteModel.objects.get(id=pk)
            serializer = backend_serializers.NoteSerializer(note, many=False)
            return Response(serializer.data)
        else:
            notes = backend_models.NoteModel.objects.all().order_by('-updated')
            serializer = backend_serializers.NoteSerializer(notes, many=True)
            return Response(serializer.data)
    elif request.method == 'POST':
        data = request.data
        note = backend_models.NoteModel.objects.create(body=data['body'])
        serializer = backend_serializers.NoteSerializer(note, many=False)
        return Response(serializer.data)
    elif request.method == 'PUT':
        data = request.data
        note = backend_models.NoteModel.objects.get(id=pk)
        equal = data['body'] != note.body
        serializer = backend_serializers.NoteSerializer(instance=note, data=data)
        if serializer.is_valid() and equal:
            serializer.save()
        return Response(serializer.data)
    elif request.method == 'DELETE':
        note = backend_models.NoteModel.objects.get(id=pk)
        note.delete()
        return Response(f'delete note №{pk} successfull')


@api_view(http_method_names=['GET', 'POST', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def _note_api(request, pk=None):
    # for key, value in request.META.items():
    #     print(f'{key}: {value}')
    # print(request)
    # print(request.user)
    # print(request.user.username)
    if request.method == 'GET':
        if pk == '-1':
            _routes = [
                {
                    'Endpoint': '$BASEPATH$/note_api/',
                    'method': 'GET',
                    'body': None,
                    'descriptions': 'Returns an array of notes objects',
                    'code': '''notes = NoteComponent.objects.all().order_by('-updated')
                    serializer = backend_serializers.NoteSerializer(notes, many=True)
                    return Response(serializer.data)'''
                },
                {
                    'Endpoint': '$BASEPATH$/note_api/<id>/',
                    'method': 'GET',
                    'body': None,
                    'descriptions': 'Returns a single note object',
                    'code': '''note = NoteComponent.objects.get(id=pk)
                    serializer = backend_serializers.(note, many=False)
                    return Response(serializer.data)'''
                },
                {
                    'Endpoint': '$BASEPATH$/note_api/',
                    'method': 'POST',
                    'body': {'body': ""},
                    'descriptions': 'Creates new note with data sent in request',
                    'code': '''data = request.data
                    note = NoteComponent.objects.create(body=data['body'])
                    serializer = backend_serializers.NoteSerializer(note, many=False)
                    return Response(serializer.data)'''
                },
                {
                    'Endpoint': '$BASEPATH$/note_api/<id>/',
                    'method': 'PUT',
                    'body': None,
                    'descriptions': 'Updates note with data sent in request',
                    'code': '''data = request.data
                    note = NoteComponent.objects.get(id=pk)
                    equal = data['body'] != note.body
                    serializer = backend_serializers.NoteSerializer(instance=note, data=data)
                    if serializer.is_valid() and equal:
                        serializer.save()
                    return Response(serializer.data)'''
                },
                {
                    'Endpoint': '$BASEPATH$/note_api/<id>/',
                    'method': 'DELETE',
                    'body': None,
                    'descriptions': 'Deletes note',
                    'code': '''note = NoteComponent.objects.get(id=pk)
                    note.delete()
                    return Response(f'delete note №{pk} successfull')'''
                },
            ]
            return Response(_routes)
        elif pk:
            note = backend_models.NoteModel.objects.get(id=pk)
            serializer = backend_serializers.NoteSerializer(note, many=False)
            return Response(serializer.data)
        else:
            notes = backend_models.NoteModel.objects.all().order_by('-updated')
            serializer = backend_serializers.NoteSerializer(notes, many=True)
            return Response(serializer.data)
    elif request.method == 'POST':
        data = request.data
        note = backend_models.NoteModel.objects.create(body=data['body'])
        serializer = backend_serializers.NoteSerializer(note, many=False)
        return Response(serializer.data)
    elif request.method == 'PUT':
        data = request.data
        note = backend_models.NoteModel.objects.get(id=pk)
        equal = data['body'] != note.body
        serializer = backend_serializers.NoteSerializer(instance=note, data=data)
        if serializer.is_valid() and equal:
            serializer.save()
        return Response(serializer.data)
    elif request.method == 'DELETE':
        note = backend_models.NoteModel.objects.get(id=pk)
        note.delete()
        return Response(f'delete note №{pk} successfull')


@api_view(http_method_names=['GET'])
@permission_classes([IsAuthenticated])
def get_products(request):
    query = request.query_params.get('keyword')
    if query is None:
        query = ''

    products = backend_models.ProductModel.objects.filter(
        name__icontains=query).order_by('-createdAt')

    page = request.query_params.get('page')
    paginator = Paginator(products, 5)

    try:
        products = paginator.page(page)
    except PageNotAnInteger:
        products = paginator.page(1)
    except EmptyPage:
        products = paginator.page(paginator.num_pages)

    if page is None:
        page = 1

    print(products)

    page = int(page)
    print('Page:', page)

    products = backend_models.ProductModel.objects.all().order_by('-createdAt')
    serializer = backend_serializers.ProductSerializer(products, many=True)
    return Response({'productsTest': serializer.data})


@api_view(http_method_names=['POST'])
def register_user(request):
    try:
        data = request.data['body']
        user = User.objects.create(
            username=data['username'],
            password=make_password(data['password'])
        )
        serializer = backend_serializers.UserSerializerWithToken(user, many=False)
        return Response(f'user: [ {user.username} | {serializer.data} ] successfull created')
    except Exception as error:
        return Response(f'user not created | error: {error}', status.HTTP_400_BAD_REQUEST)


@api_view(http_method_names=['GET'])
def get_product(request, pk):
    product = backend_models.ProductModel.objects.get(_id=pk)
    serializer = backend_serializers.ProductSerializer(product, many=False)
    return Response(serializer.data)


########################################################################################################################


def example(request):
    """
    Страница с примерами разных frontend элементов
    """

    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        response = 0
        data = [
            ['Заголовок_1', 'Заголовок_2', 'Заголовок_3'],
            [
                ['Тело_1_1', 'Тело_1_2'],
                ['Тело_2_1', 'Тело_2_2'],
                ['Тело_3_1', 'Тело_3_2'],
            ]
        ]
        if request.method == 'POST':
            response = 1
        context = {
            'response': response,
            'data': data,
            'form_1': backend_forms.ExamplesModelForm,
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': -1,
            'data': None,
            'form_1': None,
        }

    return render(request, 'backend/example.html', context)


def examples_forms(request):
    """
    Страница с примерами разных frontend форм
    """

    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        response = 0
        data = None
        if request.method == 'POST':
            response = 1
            data = [
                ['Заголовок_1', 'Заголовок_2', 'Заголовок_3'],
                [
                    ['Тело_1_1', 'Тело_1_2'],
                    ['Тело_2_1', 'Тело_2_2'],
                    ['Тело_3_1', 'Тело_3_2'],
                ]
            ]
        context = {
            'response': response,
            'data': data,
            'form_1': backend_forms.ExamplesModelForm,
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': -1,
            'data': None,
            'form_1': backend_forms.ExamplesModelForm,
        }

    return render(request, 'backend/examples.html', context)


def local(request):
    """
    Перенаправляет пользователей внутренней сети (192.168.1.202) на локальный адрес - ускорение работы
    """

    # logging
    backend_service.DjangoClass.LoggingClass.logging_actions(request=request)

    return redirect('http://192.168.1.68:8000/')


def logging(request):
    """
    Страница показа логов системы
    """

    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        response = 0
        data = None
        if request.method == 'POST':
            try:
                start = backend_service.DjangoClass.RequestClass.get_check(request, 'checkbox_start')
                start_datetime = datetime.datetime.strptime(
                    backend_service.DjangoClass.RequestClass.get_value(request, 'datetime_start', strip=False),
                    '%Y-%m-%dT%H:%M'
                ).replace(tzinfo=datetime.timezone.utc)
                end = backend_service.DjangoClass.RequestClass.get_check(request, 'checkbox_end')
                end_datetime = datetime.datetime.strptime(
                    backend_service.DjangoClass.RequestClass.get_value(request, 'datetime_end', strip=False),
                    '%Y-%m-%dT%H:%M'
                ).replace(tzinfo=datetime.timezone.utc)
                logs = backend_models.LoggingModel.objects.all()
                if backend_service.DjangoClass.RequestClass.get_value(request, 'username_slug_field'):
                    logs = logs.filter(username_slug_field=backend_service.DjangoClass.RequestClass.get_value(
                        request, 'username_slug_field')
                    )
                if backend_service.DjangoClass.RequestClass.get_value(request, 'ip_genericipaddress_field'):
                    logs = logs.filter(ip_genericipaddress_field=backend_service.DjangoClass.RequestClass.get_value(
                        request, 'ip_genericipaddress_field')
                    )
                if backend_service.DjangoClass.RequestClass.get_value(request, 'request_path_slug_field'):
                    logs = logs.filter(request_path_slug_field=backend_service.DjangoClass.RequestClass.get_value(
                        request, 'request_path_slug_field')
                    )
                if backend_service.DjangoClass.RequestClass.get_value(request, 'request_method_slug_field'):
                    logs = logs.filter(request_method_slug_field=backend_service.DjangoClass.RequestClass.get_value(
                        request, 'request_method_slug_field')
                    )
                if backend_service.DjangoClass.RequestClass.get_value(request, 'error_text_field'):
                    logs = logs.filter(error_text_field=backend_service.DjangoClass.RequestClass.get_value(
                        request, 'error_text_field')
                    )
                titles = ['username_slug_field', 'ip_genericipaddress_field', 'request_path_slug_field',
                          'request_method_slug_field', 'error_text_field', 'datetime_field']
                body = []
                for log in logs:
                    if start and end and \
                            start_datetime <= (log.datetime_field + datetime.timedelta(hours=6)) <= end_datetime:
                        body.append(
                            [log.username_slug_field,
                             log.ip_genericipaddress_field,
                             log.request_path_slug_field,
                             log.request_method_slug_field,
                             log.error_text_field,
                             log.datetime_field]
                        )
                    elif start and end is False and \
                            start_datetime <= (log.datetime_field + datetime.timedelta(hours=6)):
                        body.append(
                            [log.username_slug_field,
                             log.ip_genericipaddress_field,
                             log.request_path_slug_field,
                             log.request_method_slug_field,
                             log.error_text_field,
                             log.datetime_field]
                        )
                    elif end and start is False and \
                            (log.datetime_field + datetime.timedelta(hours=6)) <= end_datetime:
                        body.append(
                            [log.username_slug_field,
                             log.ip_genericipaddress_field,
                             log.request_path_slug_field,
                             log.request_method_slug_field,
                             log.error_text_field,
                             log.datetime_field]
                        )
                    elif start is False and end is False:
                        body.append(
                            [log.username_slug_field,
                             log.ip_genericipaddress_field,
                             log.request_path_slug_field,
                             log.request_method_slug_field,
                             log.error_text_field,
                             log.datetime_field]
                        )
                data = [titles, body]
                workbook = backend_utils.ExcelClass.workbook_create()
                sheet = backend_utils.ExcelClass.workbook_activate(workbook)
                for title in titles:
                    backend_utils.ExcelClass.set_sheet_value(
                        col=titles.index(title) + 1,
                        row=1,
                        value=title,
                        sheet=sheet
                    )
                for row in body:
                    for value in row:
                        backend_utils.ExcelClass.set_sheet_value(
                            col=row.index(value) + 1,
                            row=body.index(row) + 2,
                            value=value,
                            sheet=sheet
                        )
                backend_utils.ExcelClass.workbook_save(workbook=workbook,
                                                       excel_file='static/media/data/logging/logging.xlsx')
                response = 1
            except Exception as error:
                backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                response = -1
        context = {
            'response': response,
            'data': data,
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': -1,
            'data': None,
        }

    return render(request, 'account/account_logging.html', context)


def home(request):
    """
    Домашняя страница
    """

    # logging
    backend_service.DjangoClass.LoggingClass.logging_actions(request=request)

    return render(request, 'backend/home.html')


def create_modules(request):
    """
    Страница для создания действий, групп и модулей веб-платформы
    """

    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        response = 0
        if request.method == 'POST':
            response = 1
            actions = backend_service.DjangoClass.RequestClass.get_file(request, "excel_file_actions")
            groups = backend_service.DjangoClass.RequestClass.get_file(request, "excel_file_groups")
            _modules = backend_service.DjangoClass.RequestClass.get_file(request, "excel_file_modules")
            if actions:
                workbook = backend_utils.ExcelClass.workbook_load(excel_file=actions)
                sheet = backend_utils.ExcelClass.workbook_activate(workbook=workbook)
                dictionary = {x[1]: x[0] for x in backend_models.ActionModel.LIST_DB_VIEW_CHOICES}
                for row in range(2, backend_utils.ExcelClass.get_max_num_rows(sheet) + 1):
                    try:
                        action = backend_models.ActionModel.objects.get_or_create(
                            name_slug_field=backend_utils.ExcelClass.get_sheet_value('C', row, sheet)
                        )[0]
                        action.type_slug_field = dictionary[
                            str(backend_utils.ExcelClass.get_sheet_value('A', row, sheet)).strip()]
                        action.name_char_field = backend_utils.ExcelClass.get_sheet_value('B', row, sheet)
                        action.save()
                    except Exception as error:
                        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                backend_utils.ExcelClass.workbook_close(workbook=workbook)
            if groups:
                workbook = backend_utils.ExcelClass.workbook_load(excel_file=groups)
                sheet = backend_utils.ExcelClass.workbook_activate(workbook=workbook)
                for row in range(2, backend_utils.ExcelClass.get_max_num_rows(sheet) + 1):
                    try:
                        name_char_field = backend_utils.ExcelClass.get_sheet_value('A', row, sheet)
                        name_slug_field = backend_utils.ExcelClass.get_sheet_value('B', row, sheet)
                        action_many_to_many_field = backend_utils.ExcelClass.get_sheet_value('D', row, sheet)
                        user_many_to_many_field = backend_utils.ExcelClass.get_sheet_value('E', row, sheet)
                        position_float_field = backend_utils.ExcelClass.get_sheet_value('C', row, sheet)
                        if len(name_slug_field.strip()) <= 0:
                            continue
                        try:
                            position_float_field = int(position_float_field)
                        except Exception as error:
                            print(error)
                            position_float_field = 1
                        group = Group.objects.get_or_create(name=name_slug_field.strip().lower())[0]
                        group_model = backend_models.GroupModel.objects.get_or_create(group_foreign_key_field=group)[0]
                        if group_model.name_char_field != name_char_field:
                            group_model.name_char_field = name_char_field
                        if group_model.name_slug_field != name_slug_field:
                            group_model.name_slug_field = name_slug_field
                        if group_model.position_float_field != position_float_field:
                            group_model.position_float_field = position_float_field
                        actions = [str(x).strip() for x in action_many_to_many_field.strip().split(',')]
                        for action in actions:
                            try:
                                group_model.action_many_to_many_field.add(
                                    backend_models.ActionModel.objects.get(name_slug_field=action)
                                )
                            except Exception as error:
                                print(error)
                        users = [str(x).strip() for x in user_many_to_many_field.strip().split(',')]
                        for user in users:
                            try:
                                group_model.user_many_to_many_field.add(
                                    backend_models.UserModel.objects.get(
                                        user_foreign_key_field=User.objects.get(username=user))
                                )
                            except Exception as error:
                                print(error)
                        group_model.save()
                    except Exception as error:
                        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                backend_utils.ExcelClass.workbook_close(workbook=workbook)
            if _modules:
                workbook = backend_utils.ExcelClass.workbook_load(excel_file=_modules)
                sheet = backend_utils.ExcelClass.workbook_activate(workbook=workbook)
                dictionary = {x[1]: x[0] for x in backend_models.ModulesModel.LIST_DB_VIEW_CHOICES}
                for row in range(2, backend_utils.ExcelClass.get_max_num_rows(sheet) + 1):
                    try:
                        module = backend_models.ModulesModel.objects.get_or_create(
                            next_path_slug_field=backend_utils.ExcelClass.get_sheet_value('E', row, sheet)
                        )[0]
                        module.type_slug_field = dictionary[
                            str(backend_utils.ExcelClass.get_sheet_value('A', row, sheet)).strip()]
                        module.name_char_field = backend_utils.ExcelClass.get_sheet_value('B', row, sheet)
                        module.previous_path_slug_field = backend_utils.ExcelClass.get_sheet_value('C', row, sheet)
                        module.current_path_slug_field = backend_utils.ExcelClass.get_sheet_value('D', row, sheet)
                        module.position_float_field = backend_utils.ExcelClass.get_sheet_value('F', row, sheet)
                        module.text_field = backend_utils.ExcelClass.get_sheet_value('G', row, sheet)
                        module.save()
                    except Exception as error:
                        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                backend_utils.ExcelClass.workbook_close(workbook=workbook)

        context = {
            'response': response,
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': -1,
            '_modules': None,
        }

    return render(request, 'backend/create_modules.html', context)


def modules(request, url_slug='module_modules'):
    """
    Страница отображающая модули, секции или компоненты веб-платформы
    """

    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        _modules = backend_models.ModulesModel.objects.filter(current_path_slug_field=url_slug)
        response = 0
        context = {
            'response': response,
            '_modules': _modules,
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': -1,
            '_modules': None,
        }

    return render(request, 'backend/modules.html', context)


def account_login(request):
    """
    Страница логина пользователей
    """

    # logging
    backend_service.DjangoClass.LoggingClass.logging_actions(request=request)

    try:
        response = 0
        access_count = None
        if request.method == 'POST':
            try:
                now = (datetime.datetime.now() - datetime.timedelta(hours=1)).strftime('%Y-%m-%d %H:%M')
                access_count = 0
                for dat in backend_models.LoggingModel.objects.filter(
                        username_slug_field=request.user.username,
                        ip_genericipaddress_field=request.META.get("REMOTE_ADDR"),
                        request_path_slug_field='/account_login/',
                        request_method_slug_field='POST',
                        error_text_field='successful'
                ):
                    if (dat.datetime_field + datetime.timedelta(hours=6)).strftime('%Y-%m-%d %H:%M') >= now:
                        access_count += 1
                user = authenticate(
                    username=backend_service.DjangoClass.RequestClass.get_value(request, "username"),
                    password=backend_service.DjangoClass.RequestClass.get_value(request, "password")
                )
                if user and access_count <= 10:
                    login(request, user)
                    response = 1
                else:
                    response = -1
            except Exception as error:
                print(error)
                response = -1
        context = {
            'response': response,
            'access_count': access_count,
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': -1,
            'access_count': None,
        }

    return render(request, 'account/account_login.html', context)


def account_logout(request):
    """
    Ссылка на выход из аккаунта и перенаправление не страницу входа
    """

    # logging
    backend_service.DjangoClass.LoggingClass.logging_actions(request=request)

    try:
        logout(request)
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)

    return redirect('account_login')


def account_change_password(request):
    """
    Страница смены пароля пользователей
    """

    # logging
    backend_service.DjangoClass.LoggingClass.logging_actions(request=request)

    try:
        response = 0
        user = User.objects.get(username=request.user.username)
        user_model = backend_models.UserModel.objects.get_or_create(user_foreign_key_field=user)[0]
        if request.method == 'POST':
            try:
                # User password data
                password_1 = backend_service.DjangoClass.RequestClass.get_value(request, "password_1")
                password_2 = backend_service.DjangoClass.RequestClass.get_value(request, "password_2")
                email = backend_service.DjangoClass.RequestClass.get_value(request, "email")
                secret_question = backend_service.DjangoClass.RequestClass.get_value(request, "secret_question")
                secret_answer = backend_service.DjangoClass.RequestClass.get_value(request, "secret_answer")
                if password_1 == password_2 and len(password_1) >= 8 and password_1 != user_model.password_slug_field:
                    user_model.password_slug_field = password_1
                    user.password = password_1
                    user.set_password(password_1)
                # Third data account
                if email and email != user_model.email_field:
                    user_model.email_field = email
                if secret_question and secret_question != user_model.secret_question_char_field:
                    user_model.secret_question_char_field = secret_question
                if secret_answer and secret_answer != user_model.secret_answer_char_field:
                    user_model.secret_answer_char_field = secret_answer
                user.save()
                user_model.save()
                response = 1
            except Exception as error:
                backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                response = -1
        context = {
            'response': response,
            'user': user,
            'user_model': user_model,
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': -1,
            'user': None,
            'user_model': None,
        }

    return render(request, 'account/account_change_password.html', context)


def account_change_profile(request):
    """
    Страница изменения профиля пользователя
    """
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request,
                                                                        access='superuser')
    if page:
        return redirect(page)

    try:
        response = 0
        user = User.objects.get(username=request.user.username)
        user_model = backend_models.UserModel.objects.get_or_create(user_foreign_key_field=user)[0]
        if request.method == 'POST':
            try:
                education = backend_service.DjangoClass.RequestClass.get_value(request, "education")
                achievements = backend_service.DjangoClass.RequestClass.get_value(request, "achievements")
                biography = backend_service.DjangoClass.RequestClass.get_value(request, "biography")
                hobbies = backend_service.DjangoClass.RequestClass.get_value(request, "hobbies")
                image_avatar = backend_service.DjangoClass.RequestClass.get_file(request, "image_avatar")
                # Second data account
                if education and education != user_model.education_text_field:
                    user_model.education_text_field = education
                if achievements and achievements != user_model.achievements_text_field:
                    user_model.achievements_text_field = achievements
                if biography and biography != user_model.biography_text_field:
                    user_model.biography_text_field = biography
                if hobbies and hobbies != user_model.hobbies_text_field:
                    user_model.hobbies_text_field = hobbies
                if image_avatar and image_avatar != user_model.image_field:
                    user_model.image_field = image_avatar
                user.save()
                user_model.save()
                response = 1
            except Exception as error:
                backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                response = -1
        context = {
            'response': response,
            'user': user,
            'user_model': user_model,
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': -1,
            'user': None,
            'user_model': None,
        }

    return render(request, 'account/account_change_profile.html', context)


def account_recover_password(request, type_slug='iin'):
    """
    Страница восстановления пароля пользователей
    """

    # logging
    backend_service.DjangoClass.LoggingClass.logging_actions(request=request)

    try:
        response = 0
        data = None
        user = None
        user_model = None
        access_count = None
        if request.method == 'POST':
            try:
                user = User.objects.get(
                    username=backend_service.DjangoClass.RequestClass.get_value(request, "username"))
                user_model = backend_models.UserModel.objects.get_or_create(user_foreign_key_field=user)[0]
            except Exception as error:
                print(error)
            if type_slug.lower() == 'iin':
                try:
                    if user:
                        response = 1
                    now = (datetime.datetime.now() - datetime.timedelta(hours=1)).strftime('%Y-%m-%d %H:%M')
                    access_count = 0
                    for dat in backend_models.LoggingModel.objects.filter(
                            username_slug_field=request.user.username,
                            ip_genericipaddress_field=request.META.get("REMOTE_ADDR"),
                            request_path_slug_field='/account_login/',
                            request_method_slug_field='POST',
                            error_text_field='successful'
                    ):
                        if (dat.datetime_field + datetime.timedelta(hours=6)).strftime('%Y-%m-%d %H:%M') >= now:
                            access_count += 1
                    if access_count > 10:
                        response = -1
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
            elif type_slug.lower() == 'secret_answer':
                try:
                    secret_answer = backend_service.DjangoClass.RequestClass.get_value(request, "secret_answer")
                    password_1 = backend_service.DjangoClass.RequestClass.get_value(request, "password_1")
                    password_2 = backend_service.DjangoClass.RequestClass.get_value(request, "password_2")
                    if user.profile.secret_answer.lower() == secret_answer.lower() and password_1 == password_2:
                        user.profile.password = password_1
                        user.set_password(password_1)
                        user.save()
                        response = 2
                    else:
                        response = -2
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
            elif type_slug.lower() == 'email':
                try:
                    password = user_model.password_slug_field
                    email_ = user_model.email_field
                    email__ = backend_service.DjangoClass.RequestClass.get_value(request, "email")
                    if password and email_ and email_ == email__:
                        subject = 'Восстановление пароля от веб платформы'
                        encrypt_message = backend_utils.EncryptingClass.encrypt_text(
                            text=f'_{password}_{datetime.datetime.now().strftime("%Y-%m-%dT%H:%M")}_',
                            hash_chars=f'_{password}_{datetime.datetime.now().strftime("%Y-%m-%dT%H:%M")}_'
                        )
                        message_s = f'{user_model.first_name_char_field} {user_model.last_name_char_field}, ' \
                                    f'перейдите по ссылке: http://192.168.1.68:80/account_recover_password/ , ' \
                                    f'введите иин и затем в окне почты введите код (без кавычек): "{encrypt_message}"'
                        # from_email = 'eevee.cycle@yandex.ru'
                        from_email = 'webapp@km.kz'
                        to_email = email_
                        if subject and message_s and to_email:
                            send_mail(subject, message_s, from_email, [to_email, ''], fail_silently=False)
                            response = 2
                        else:
                            response = -2
                    else:
                        response = -2
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
            elif type_slug.lower() == 'recover':
                try:
                    encrypt_text = backend_service.DjangoClass.RequestClass.get_value(request, "recover")
                    decrypt_text = backend_utils.EncryptingClass.decrypt_text(
                        text=encrypt_text,
                        hash_chars=f'_{user_model.password_slug_field}_'
                                   f'{datetime.datetime.now().strftime("%Y-%m-%dT%H:%M")}_'
                    )
                    if decrypt_text.split('_')[2] >= \
                            (datetime.datetime.now() - datetime.timedelta(hours=1)).strftime('%Y-%m-%d %H:%M') and \
                            decrypt_text.split('_')[1] == user_model.password_slug_field:
                        login(request, user)
                        user_model.secret_question_char_field = ''
                        user_model.secret_answer_char_field = ''
                        user_model.save()
                        response = 2
                    else:
                        response = -2
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': response,
            'data': data,
            'access_count': access_count,
            'user': user,
            'user_model': user_model,
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': -1,
            'data': None,
            'access_count': None,
            'user': None,
            'user_model': None,
        }

    return render(request, 'account/account_recover_password.html', context)


def account_profile(request, user_id=0):
    """
    Страница профиля пользователя
    """

    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        if user_id <= 0:
            user = request.user
        else:
            user = User.objects.get(id=user_id)
        user_model = backend_models.UserModel.objects.get_or_create(user_foreign_key_field=user)[0]
        response = 1
        context = {
            'response': response,
            'user': user,
            'user_model': user_model,
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': -1,
            'user': None,
            'user_model': None,
        }

    return render(request, 'account/account_profile.html', context)


def account_notification(request, type_slug='All'):
    """
    Страница уведомлений пользователя
    """

    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        user_model = backend_models.UserModel.objects.get_or_create(
            user_foreign_key_field=User.objects.get(id=request.user.id)
        )[0]
        if type_slug.lower() != 'all':
            notifications = backend_models.NotificationModel.objects.filter(
                user_foreign_key_field=user_model,
                type_slug_field=type_slug,
            ).order_by('status_boolean_field', '-created_datetime_field')
        else:
            notifications = backend_models.NotificationModel.objects.filter(
                user_foreign_key_field=user_model,
            ).order_by('status_boolean_field', '-created_datetime_field')
        types = backend_models.NotificationModel.get_all_types()
        try:
            page = backend_service.PaginationClass.paginate(request=request, objects=notifications, num_page=100)
            response = 0
        except Exception as error:
            backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
            response = -1

        context = {
            'response': response,
            'page': page,
            'types': types,
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': -1,
            'page': None,
            'types': None,
        }

    return render(request, 'account/account_notifications.html', context)


def account_create_notification(request):
    """
    Страница создания уведомления пользователя
    """

    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request,
                                                                        access='superuser')
    if page:
        return redirect(page)

    # try:
    if True:
        if request.method == 'POST':
            type_slug_field = backend_service.DjangoClass.RequestClass.get_value(request, "type_slug_field")
            name_char_field = backend_service.DjangoClass.RequestClass.get_value(request, "name_char_field")
            text_field = backend_service.DjangoClass.RequestClass.get_value(request, "text_field")
            superusers = User.objects.filter(is_superuser=True)
            for superuser in superusers:
                user_model = backend_models.UserModel.objects.get_or_create(user_foreign_key_field=superuser)[0]
                backend_models.NotificationModel.objects.create(
                    user_foreign_key_field=user_model,
                    type_slug_field=type_slug_field,
                    name_char_field=name_char_field,
                    text_field=text_field
                )
    # except Exception as error:
    #     backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)

    return redirect('home')


def account_delete_or_change_notification(request, notification_id: int):
    """
    Страница создания уведомления пользователя
    """

    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request,
                                                                        access='superuser')
    if page:
        return redirect(page)

    # try:
    if True:
        if request.method == 'POST':

            type_action = backend_service.DjangoClass.RequestClass.get_value(request, "type_action")
            notification = backend_models.NotificationModel.objects.get(id=notification_id)
            if type_action == 'change':
                _status = backend_service.DjangoClass.RequestClass.get_value(request, "hidden")
                if _status == 'true':
                    _status = True
                elif _status == 'false':
                    _status = False
                notification.status_boolean_field = _status
                notification.decision_datetime_field = backend_utils.DateTimeUtils.get_current_datetime()
                notification.save()
            elif type_action == 'delete':
                notification.delete()
    # except Exception as error:
    #     backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)

    return redirect('account_notification')


def account_create_or_change_accounts(request):
    """
    Страница создания или изменения пользователей
    """

    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request,
                                                                        access='superuser')
    if page:
        return redirect(page)

    try:
        response = 0
        if request.method == 'POST':
            excel_file = request.FILES.get('excel_file')
            if excel_file:
                try:
                    workbook = backend_utils.ExcelClass.workbook_load(excel_file=excel_file)
                    sheet = backend_utils.ExcelClass.workbook_activate(workbook=workbook)
                    for row in range(2, backend_utils.ExcelClass.get_max_num_rows(sheet) + 1):
                        try:
                            # authorization data
                            username = backend_utils.ExcelClass.get_sheet_value('J', row, sheet)
                            password = backend_utils.ExcelClass.get_sheet_value('K', row, sheet)
                            # technical data
                            is_active = backend_utils.ExcelClass.get_sheet_value('L', row, sheet)
                            if is_active.lower() == 'true':
                                is_active = True
                            else:
                                is_active = False
                            is_staff = backend_utils.ExcelClass.get_sheet_value('M', row, sheet)
                            if is_staff.lower() == 'true':
                                is_staff = True
                            else:
                                is_staff = False
                            is_superuser = backend_utils.ExcelClass.get_sheet_value('N', row, sheet)
                            if is_superuser.lower() == 'true':
                                is_superuser = True
                            else:
                                is_superuser = False
                            groups = backend_utils.ExcelClass.get_sheet_value('O', row, sheet)
                            email_ = backend_utils.ExcelClass.get_sheet_value('P', row, sheet)
                            secret_question = backend_utils.ExcelClass.get_sheet_value('Q', row, sheet)
                            secret_answer = backend_utils.ExcelClass.get_sheet_value('R', row, sheet)
                            # first data
                            last_name = backend_utils.ExcelClass.get_sheet_value('D', row, sheet)
                            first_name = backend_utils.ExcelClass.get_sheet_value('E', row, sheet)
                            patronymic = backend_utils.ExcelClass.get_sheet_value('F', row, sheet)
                            # second data
                            personnel_number = backend_utils.ExcelClass.get_sheet_value('G', row, sheet)
                            subdivision = backend_utils.ExcelClass.get_sheet_value('A', row, sheet)
                            workshop_service = backend_utils.ExcelClass.get_sheet_value('B', row, sheet)
                            department_site = backend_utils.ExcelClass.get_sheet_value('C', row, sheet)
                            position = backend_utils.ExcelClass.get_sheet_value('H', row, sheet)
                            category = backend_utils.ExcelClass.get_sheet_value('I', row, sheet)
                            # utils
                            force_change_account = backend_service.DjangoClass.RequestClass.get_check(
                                request, 'force_change_account'
                            )
                            force_change_password = backend_service.DjangoClass.RequestClass.get_check(
                                request, 'force_change_password'
                            )
                            force_clear_groups = backend_service.DjangoClass.RequestClass.get_check(
                                request, 'force_clear_groups'
                            )
                            if len(username) < 1:
                                continue
                            # Пользователь уже существует: изменение
                            try:
                                user = User.objects.get(username=username)
                                # Возврат, если пользователь обладает правами суперпользователя или его нельзя изменять
                                if user.is_superuser or force_change_account is False:
                                    continue
                                new_account = False
                            # Пользователь не существует: создание
                            except Exception as error:
                                backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                                user = User.objects.create(
                                    # authorization data
                                    username=username,
                                    password=backend_service.DjangoClass.AccountClass.create_django_encrypt_password(
                                        password),
                                    # technical data
                                    is_active=True,
                                    is_staff=is_staff,
                                    is_superuser=is_superuser,
                                )
                                new_account = True
                            user_model = backend_models.UserModel.objects.get_or_create(user_foreign_key_field=user)[0]
                            # authorization data
                            if new_account:
                                user_model.password_slug_field = password
                            else:
                                if force_change_password:
                                    user.password = backend_service.DjangoClass. \
                                        AccountClass.create_django_encrypt_password(password)
                                    user.save()
                                    user_model.password_slug_field = password
                            # technical data
                            user_model.activity_boolean_field = is_active
                            user_model.email_field = email_
                            user_model.secret_question_char_field = secret_question
                            user_model.secret_answer_char_field = secret_answer
                            # first data
                            user_model.last_name_char_field = last_name
                            user_model.first_name_char_field = first_name
                            user_model.patronymic_char_field = patronymic
                            # second data
                            user_model.personnel_number_slug_field = personnel_number
                            user_model.subdivision_char_field = subdivision
                            user_model.workshop_service_char_field = workshop_service
                            user_model.department_site_char_field = department_site
                            user_model.position_char_field = position
                            user_model.category_char_field = category
                            # save
                            user_model.save()
                            # update data
                            user_model = backend_models.UserModel.objects.get_or_create(
                                user_foreign_key_field=User.objects.get(username=username)
                            )[0]
                            if force_clear_groups:
                                for group in backend_models.GroupModel.objects.filter(
                                        user_many_to_many_field=user_model):
                                    group.user_many_to_many_field.remove(user_model)
                            response = 1
                            groups = [group.strip() for group in str(groups).lower().strip().split(',')]
                            for group in groups:
                                if len(group) > 0:
                                    group_object = Group.objects.get_or_create(name=group)[0]
                                    try:
                                        group_model = backend_models.GroupModel.objects.get(
                                            group_foreign_key_field=group_object
                                        )
                                    except Exception as error:
                                        backend_service.DjangoClass.LoggingClass.logging_errors(request=request,
                                                                                                error=error)
                                        group_model = backend_models.GroupModel.objects.create(
                                            group_foreign_key_field=group_object,
                                            name_char_field=group,
                                            name_slug_field=group,
                                        )
                                    group_model.user_many_to_many_field.add(user_model)
                        except Exception as error:
                            backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                            response = -1
                    backend_utils.ExcelClass.workbook_close(workbook=workbook)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                    response = -1
        context = {
            'response': response,
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': -1,
        }

    return render(request, 'account/account_create_accounts.html', context)


def account_export_accounts(request):
    """
    Страница экспорта пользователей
    """

    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request,
                                                                        access='superuser')
    if page:
        return redirect(page)

    try:
        response = 0
        data = None
        if request.method == 'POST':
            try:
                workbook = backend_utils.ExcelClass.workbook_create()
                sheet = backend_utils.ExcelClass.workbook_activate(workbook)
                user_objects = User.objects.all().order_by('-id')
                titles = ['Подразделение', 'Цех/Служба', 'Отдел/Участок', 'Фамилия', 'Имя', 'Отчество',
                          'Табельный номер', 'Должность', 'Категория работника', 'Имя пользователя',
                          'Пароль аккаунта', 'Активность аккаунта', 'Доступ к панели управления',
                          'Права суперпользователя', 'Группы доступа', 'Электронная почта', 'Секретный вопрос',
                          'Секретный ответ']
                for title in titles:
                    backend_utils.ExcelClass.set_sheet_value(
                        col=titles.index(title) + 1,
                        row=1,
                        value=title,
                        sheet=sheet
                    )
                _index = 1
                body = []
                response = 1
                for user_object in user_objects:
                    try:
                        if User.objects.get(username=user_object.username).is_superuser:
                            continue
                        _index += 1
                        sub_body = []
                        # authorization data
                        username = user_object.username
                        user_model = backend_models.UserModel.objects.get_or_create(user_foreign_key_field=user_object)[
                            0]
                        backend_utils.ExcelClass.set_sheet_value('J', _index, username, sheet)

                        password = user_model.password_slug_field
                        backend_utils.ExcelClass.set_sheet_value('K', _index, password, sheet)

                        # technical data
                        is_active = user_model.activity_boolean_field
                        if is_active:
                            is_active = 'true'
                        else:
                            is_active = 'false'
                        backend_utils.ExcelClass.set_sheet_value('L', _index, is_active, sheet)

                        is_staff = user_object.is_staff
                        if is_staff:
                            is_staff = 'true'
                        else:
                            is_staff = 'false'
                        backend_utils.ExcelClass.set_sheet_value('M', _index, is_staff, sheet)

                        is_superuser = user_object.is_superuser
                        if is_superuser:
                            is_superuser = 'true'
                        else:
                            is_superuser = 'false'
                        backend_utils.ExcelClass.set_sheet_value('N', _index, is_superuser, sheet)

                        group_string = ''
                        groups = backend_models.GroupModel.objects.filter(user_many_to_many_field=user_model)
                        if groups:
                            for group in groups:
                                group_string += f", {group}"
                            groups = group_string[2:]
                        else:
                            groups = ''
                        backend_utils.ExcelClass.set_sheet_value('P', _index, groups, sheet)

                        _email = user_model.email_field
                        backend_utils.ExcelClass.set_sheet_value('O', _index, _email, sheet)

                        secret_question = user_model.secret_question_char_field
                        backend_utils.ExcelClass.set_sheet_value('Q', _index, secret_question, sheet)

                        secret_answer = user_model.secret_answer_char_field
                        backend_utils.ExcelClass.set_sheet_value('R', _index, secret_answer, sheet)

                        # first data
                        last_name = user_model.last_name_char_field
                        backend_utils.ExcelClass.set_sheet_value('D', _index, last_name, sheet)

                        first_name = user_model.first_name_char_field
                        backend_utils.ExcelClass.set_sheet_value('E', _index, first_name, sheet)

                        patronymic = user_model.patronymic_char_field
                        backend_utils.ExcelClass.set_sheet_value('F', _index, patronymic, sheet)
                        # second data

                        personnel_number = user_model.personnel_number_slug_field
                        backend_utils.ExcelClass.set_sheet_value('G', _index, personnel_number, sheet)

                        subdivision = user_model.subdivision_char_field
                        backend_utils.ExcelClass.set_sheet_value('A', _index, subdivision, sheet)

                        workshop_service = user_model.workshop_service_char_field
                        backend_utils.ExcelClass.set_sheet_value('B', _index, workshop_service, sheet)

                        department_site = user_model.department_site_char_field
                        backend_utils.ExcelClass.set_sheet_value('C', _index, department_site, sheet)

                        position = user_model.position_char_field
                        backend_utils.ExcelClass.set_sheet_value('H', _index, position, sheet)

                        category = user_model.category_char_field
                        backend_utils.ExcelClass.set_sheet_value('I', _index, category, sheet)

                        sub_body.append(subdivision)
                        sub_body.append(workshop_service)
                        sub_body.append(department_site)
                        sub_body.append(last_name)
                        sub_body.append(first_name)
                        sub_body.append(patronymic)
                        sub_body.append(personnel_number)
                        sub_body.append(position)
                        sub_body.append(category)
                        sub_body.append(username)
                        sub_body.append(password)
                        sub_body.append(is_active)
                        sub_body.append(is_staff)
                        sub_body.append(is_superuser)
                        sub_body.append(groups)
                        sub_body.append(_email)
                        sub_body.append(secret_question)
                        sub_body.append(secret_answer)
                        body.append(sub_body)
                    except Exception as error:
                        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                        response = -1
                    data = [titles, body]
                backend_utils.ExcelClass.workbook_save(workbook=workbook,
                                                       excel_file='static/media/admin/account/export_users.xlsx')
            except Exception as error:
                backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                response = -1
        context = {
            'response': response,
            'data': data,
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': -1,
            'data': None,
        }

    return render(request, 'account/account_export_accounts.html', context)


def django_account_generate_passwords(request):
    """
    Страница генерации паролей для аккаунтов пользователей
    """

    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request,
                                                                        access='superuser')
    if page:
        return redirect(page)

    try:
        response = 0
        data = None
        if request.method == 'POST':
            try:
                passwords_chars = backend_service.DjangoClass.RequestClass.get_value(request, "passwords_chars")
                passwords_quantity = backend_service.DjangoClass.RequestClass.get_value(request, "passwords_quantity")
                passwords_length = backend_service.DjangoClass.RequestClass.get_value(request, "passwords_length")
                workbook = backend_utils.ExcelClass.workbook_create()
                sheet = backend_utils.ExcelClass.workbook_activate(workbook=workbook)
                titles = ['Пароль']
                for title in titles:
                    backend_utils.ExcelClass.set_sheet_value(
                        col=titles.index(title) + 1,
                        row=1,
                        value=title,
                        sheet=sheet
                    )
                body = []
                for n in range(2, int(passwords_quantity) + 2):
                    password = backend_service.DjangoClass.AccountClass.create_password_from_chars(
                        chars=passwords_chars,
                        length=int(passwords_length)
                    )
                    sheet[f'A{n}'] = f'temp_{password}'
                    body.append([f'temp_{password}'])
                backend_utils.ExcelClass.workbook_save(
                    workbook=workbook, excel_file='static/media/admin/account/generate_passwords.xlsx'
                )
                backend_utils.ExcelClass.workbook_close(workbook=workbook)
                data = [titles, body]
                response = 1
            except Exception as error:
                backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                response = -1
        context = {
            'response': response,
            'data': data,
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': -1,
            'data': None,
        }

    return render(request, 'account/django_account_generate_passwords.html', context)


def account_update_accounts_1c(request):
    """
    Страница обновления аккаунтов пользователей из системы 1С
    """

    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request,
                                                                        access='superuser')
    if page:
        return redirect(page)

    # try:
    if True:
        response = 0
        data = None
        if request.method == 'POST':
            # try:
            if True:
                key = backend_service.UtilsClass.create_encrypted_password(
                    _random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890',
                    _length=10
                )
                hash_key_obj = hashlib.sha256()
                hash_key_obj.update(key.encode('utf-8'))
                key_hash = str(hash_key_obj.hexdigest().strip().upper())
                key_hash_base64 = base64.b64encode(str(key_hash).encode()).decode()
                date = datetime.datetime.now().strftime("%Y%m%d")
                date_base64 = base64.b64encode(str(date).encode()).decode()
                url = f'http://192.168.1.10/KM_1C/hs/iden/change/{date_base64}_{key_hash_base64}'
                relative_path = os.path.dirname(os.path.abspath('__file__')) + '\\'
                h = httplib2.Http(relative_path + "\\static\\media\\data\\temp\\get_users")
                _login = 'Web_adm_1c'
                password = '159159qqww!'
                h.add_credentials(_login, password)
                try:
                    response_, content = h.request(url)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                    content = None
                json_data = None
                success_web_read = False
                if content:
                    success = True
                    error_word_list = ['Ошибка', 'ошибка',
                                       'Error', 'error', 'Failed', 'failed']
                    for error_word in error_word_list:
                        if str(content.decode()).find(error_word) >= 0:
                            success = False
                    if success:
                        json_data = json.loads(
                            backend_service.UtilsClass.decrypt_text_with_hash(content.decode()[1:], key_hash))
                        with open("static/media/admin/account/accounts.json", "w", encoding="utf-8") as file:
                            json.dump(backend_service.UtilsClass.decrypt_text_with_hash(content.decode()[1:], key_hash),
                                      file)
                        success_web_read = True
                if success_web_read is False:
                    print('read temp file')
                    with open("static/media/admin/account/accounts_temp.json", "r", encoding="utf-8") as file:
                        json_data = json.load(file)
                # Генерация объектов для создания аккаунтов
                titles_1c = ['Период', 'Статус', 'ИИН', 'Фамилия', 'Имя', 'Отчество', 'ТабельныйНомер', 'Подразделение',
                             'Цех_Служба', 'Отдел_Участок', 'Должность', 'Категория']
                for user in json_data["global_objects"]:
                    try:
                        # authorization data
                        username = json_data["global_objects"][user]["ИИН"]
                        password = backend_service.DjangoClass.AccountClass.create_password_from_chars(length=8)
                        # technical data
                        if json_data["global_objects"][user]["Статус"] == 'created' or \
                                json_data["global_objects"][user]["Статус"] == 'changed':
                            is_active = True
                        else:
                            is_active = False
                        # first data
                        last_name = json_data["global_objects"][user]["Фамилия"]
                        first_name = json_data["global_objects"][user]["Имя"]
                        patronymic = json_data["global_objects"][user]["Отчество"]
                        # second data
                        personnel_number = json_data["global_objects"][user]["ТабельныйНомер"]
                        subdivision = json_data["global_objects"][user]["Подразделение"]
                        workshop_service = json_data["global_objects"][user]["Цех_Служба"]
                        department_site = json_data["global_objects"][user]["Отдел_Участок"]
                        position = json_data["global_objects"][user]["Должность"]
                        category = json_data["global_objects"][user]["Категория"]
                        # Пользователь уже существует: изменение
                        try:
                            user = User.objects.get(username=username)
                            new_account = False
                        # Пользователь не существует: создание
                        except Exception as error:
                            backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                            user = User.objects.create(
                                # authorization data
                                username=username,
                                password=backend_service.DjangoClass.AccountClass.create_django_encrypt_password(
                                    password=password),
                                # technical data
                                is_active=True,
                                is_staff=False,
                                is_superuser=False,
                            )
                            new_account = True
                        user_model = backend_models.UserModel.objects.get_or_create(user_foreign_key_field=user)[0]
                        # authorization data
                        if new_account:
                            user_model.password_slug_field = password
                        # technical data
                        user_model.activity_boolean_field = is_active
                        if new_account:
                            user_model.email_field = ''
                            user_model.secret_question_char_field = ''
                            user_model.secret_answer_char_field = ''
                        # first data
                        user_model.last_name_char_field = last_name
                        user_model.first_name_char_field = first_name
                        user_model.patronymic_char_field = patronymic
                        # second data
                        user_model.personnel_number_slug_field = personnel_number
                        user_model.subdivision_char_field = subdivision
                        user_model.workshop_service_char_field = workshop_service
                        user_model.department_site_char_field = department_site
                        user_model.position_char_field = position
                        user_model.category_char_field = category
                        # save
                        user_model.save()
                        # update data
                        user_model = backend_models.UserModel.objects.get_or_create(
                            user_foreign_key_field=User.objects.get(username=username)
                        )[0]
                        groups = ['user']
                        for group in groups:
                            if len(group) > 0:
                                group_object = Group.objects.get_or_create(name=group)[0]
                                try:
                                    group_model = backend_models.GroupModel.objects.get(
                                        group_foreign_key_field=group_object
                                    )
                                except Exception as error:
                                    backend_service.DjangoClass.LoggingClass.logging_errors(request=request,
                                                                                            error=error)
                                    group_model = backend_models.GroupModel.objects.create(
                                        group_foreign_key_field=group_object,
                                        name_char_field=group,
                                        name_slug_field=group,
                                    )
                                group_model.user_many_to_many_field.add(user_model)
                        response = 1
                    except Exception as error:
                        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                        response = -1

                # Генерация ответа для отрисовки в таблицу на странице
                titles = ['Период', 'Статус', 'ИИН', 'Фамилия', 'Имя', 'Отчество', 'Табельный', 'Подразделение',
                          'Цех/Служба', 'Отдел/Участок', 'Должность', 'Категория']
                bodies = []
                for user in json_data["global_objects"]:
                    user_object = []
                    for key in titles_1c:
                        value = str(json_data["global_objects"][user][key]).strip()
                        user_object.append(value)
                    bodies.append(user_object)
                data = [titles, bodies]
            # except Exception as error:
            #     backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': response,
            'data': data,
        }
    # except Exception as error:
    #     backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
    #     context = {
    #         'response': -1,
    #         'data': None,
    #     }

    return render(request, 'account/account_update_accounts_1c.html', context)


def account_change_groups(request):
    """
    Страница создания пользователей
    """

    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    # try:
    if True:
        response = 0
        if request.method == 'POST':
            type_select = backend_service.DjangoClass.RequestClass.get_value(request, 'type_select')

            group_name_char_field = backend_service.DjangoClass.RequestClass.get_value(request, 'group_name_char_field')
            group_name_slug_field = backend_service.DjangoClass.RequestClass.get_value(request, 'group_name_slug_field')
            action_name_char_field = backend_service.DjangoClass.RequestClass.get_value(request,
                                                                                        'action_name_char_field')
            action_name_slug_field = backend_service.DjangoClass.RequestClass.get_value(request,
                                                                                        'action_name_slug_field')

            group = backend_service.DjangoClass.RequestClass.get_value(request, 'group')
            action = backend_service.DjangoClass.RequestClass.get_value(request, 'action')
            username = backend_service.DjangoClass.RequestClass.get_value(request, 'username')
            if group_name_char_field and group_name_slug_field and action_name_char_field and action_name_slug_field:
                action_model = backend_models.ActionModel.objects.get_or_create(
                    name_char_field=action_name_char_field,
                    name_slug_field=action_name_slug_field,
                )[0]
                group = Group.objects.get_or_create(name=group_name_slug_field)[0]
                group_model = backend_models.GroupModel.objects.get_or_create(
                    group_foreign_key_field=group,
                    name_slug_field=group_name_slug_field,
                )[0]
                if type_select == 'add':
                    group_model.action_many_to_many_field.add(action_model)
                elif type_select == 'remove':
                    group_model.action_many_to_many_field.remove(action_model)
                group_model.save()
            elif action:
                group = Group.objects.get_or_create(name=group)[0]
                group_model = backend_models.GroupModel.objects.get_or_create(group_foreign_key_field=group)[0]
                action_model = backend_models.ActionModel.objects.get_or_create(name_slug_field=action)[0]
                if type_select == 'add':
                    group_model.action_many_to_many_field.add(action_model)
                elif type_select == 'remove':
                    group_model.action_many_to_many_field.remove(action_model)
                group_model.save()
            elif username:
                group = Group.objects.get_or_create(name=group)[0]
                group_model = backend_models.GroupModel.objects.get_or_create(group_foreign_key_field=group)[0]
                user_model = backend_models.UserModel.objects.get_or_create(
                    user_foreign_key_field=User.objects.get_or_create(username=username)[0]
                )[0]
                if type_select == 'add':
                    group_model.user_many_to_many_field.add(user_model)
                elif type_select == 'remove':
                    group_model.user_many_to_many_field.remove(user_model)
                group_model.save()
            response = 1
        actions = backend_models.ActionModel.objects.all()
        groups = backend_models.GroupModel.objects.all()
        users = backend_models.UserModel.objects.all()
        context = {
            'response': response,
            'actions': actions,
            'groups': groups,
            'users': users,
        }
    # except Exception as error:
    #     backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
    #     context = {
    #         'response': -1,
    #         'actions': None,
    #         'groups': None,
    #         'users': None,
    #     }

    return render(request, 'account/account_change_groups.html', context)


def idea_create(request):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        response = 0
        category = backend_models.IdeaModel.get_all_category()
        if request.method == 'POST':
            author_foreign_key_field = backend_models.UserModel.objects.get(user_foreign_key_field=request.user)
            name_char_field = backend_service.DjangoClass.RequestClass.get_value(request, "name_char_field")
            category_slug_field = backend_service.DjangoClass.RequestClass.get_value(request, "category_slug_field")
            short_description_char_field = backend_service.DjangoClass.RequestClass.get_value(
                request, "short_description_char_field"
            )
            full_description_text_field = backend_service.DjangoClass.RequestClass.get_value(
                request, "full_description_text_field"
            )
            avatar_image_field = backend_service.DjangoClass.RequestClass.get_file(request, "avatar_image_field")
            addiction_file_field = backend_service.DjangoClass.RequestClass.get_file(request, "addiction_file_field")
            backend_models.IdeaModel.objects.create(
                author_foreign_key_field=author_foreign_key_field,
                name_char_field=name_char_field,
                category_slug_field=category_slug_field,
                short_description_char_field=short_description_char_field,
                full_description_text_field=full_description_text_field,
                avatar_image_field=avatar_image_field,
                addiction_file_field=addiction_file_field,
                visibility_boolean_field=False,
            )
            response = 1
        context = {
            'response': response,
            'category': category,
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': -1,
            'category': None,
        }

    return render(request, 'idea/idea_create.html', context)


def idea_change(request, idea_int):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        response = 0
        idea = backend_models.IdeaModel.objects.get(id=idea_int)
        users = backend_models.UserModel.objects.all()
        categoryes = backend_models.IdeaModel.get_all_category()
        if request.method == 'POST':
            author_foreign_key_field_id = backend_service.DjangoClass.RequestClass.get_value(
                request, "author_foreign_key_field_id"
            )
            author_foreign_key_field = backend_models.UserModel.objects.get(id=author_foreign_key_field_id)
            name_char_field = backend_service.DjangoClass.RequestClass.get_value(
                request, "name_char_field"
            )
            category_slug_field = backend_service.DjangoClass.RequestClass.get_value(
                request, "category_slug_field"
            )
            short_description_char_field = backend_service.DjangoClass.RequestClass.get_value(
                request, "short_description_char_field"
            )
            full_description_text_field = backend_service.DjangoClass.RequestClass.get_value(
                request, "full_description_text_field"
            )
            avatar_image_field = backend_service.DjangoClass.RequestClass.get_file(
                request, "avatar_image_field"
            )
            addiction_file_field = backend_service.DjangoClass.RequestClass.get_file(
                request, "addiction_file_field"
            )

            if author_foreign_key_field and author_foreign_key_field != idea.author_foreign_key_field:
                idea.author_foreign_key_field = author_foreign_key_field
            if name_char_field and name_char_field != idea.name_char_field:
                idea.name_char_field = name_char_field
            if category_slug_field and category_slug_field != idea.category_slug_field:
                idea.category_slug_field = category_slug_field
            if short_description_char_field and short_description_char_field != idea.short_description_char_field:
                idea.short_description_char_field = short_description_char_field
            if full_description_text_field and full_description_text_field != idea.full_description_text_field:
                idea.full_description_text_field = full_description_text_field
            if avatar_image_field and avatar_image_field != idea.avatar_image_field:
                idea.avatar_image_field = avatar_image_field
            if addiction_file_field and addiction_file_field != idea.addiction_file_field:
                idea.addiction_file_field = addiction_file_field

            idea.save()
            response = 1
        context = {
            'response': response,
            'idea': idea,
            'users': users,
            'categoryes': categoryes,
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': -1,
            'idea': None,
            'users': None,
            'categoryes': None,
        }

    return render(request, 'idea/idea_change.html', context)


def idea_list(request, category_slug='All'):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        categoryes = backend_models.IdeaModel.get_all_category()
        num_page = 5
        if category_slug == 'idea_change_visibility':
            ideas = backend_models.IdeaModel.objects.filter(visibility_boolean_field=False)
        elif category_slug.lower() != 'all':
            ideas = backend_models.IdeaModel.objects.filter(category_slug_field=category_slug,
                                                            visibility_boolean_field=True)
        else:
            ideas = backend_models.IdeaModel.objects.filter(visibility_boolean_field=True)
        if request.method == 'POST':
            search_char_field = backend_service.DjangoClass.RequestClass.get_value(request, "search_char_field")
            if search_char_field:
                ideas = ideas.filter(name_char_field__icontains=search_char_field)
            num_page = 100
        page = backend_service.PaginationClass.paginate(request=request, objects=ideas, num_page=num_page)
        response = 0
        context = {
            'response': response,
            'page': page,
            'categoryes': categoryes,
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': -1,
            'page': None,
            'categoryes': None
        }

    return render(request, 'idea/idea_list.html', context)


def idea_change_visibility(request, idea_int):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request,
                                                                        access='superuser')
    if page:
        return redirect(page)

    try:
        if request.method == 'POST':
            _status = backend_service.DjangoClass.RequestClass.get_value(request, "hidden")
            if _status == 'true':
                _status = True
            elif _status == 'false':
                _status = False
            data = backend_models.IdeaModel.objects.get(id=idea_int)
            data.visibility_boolean_field = _status

            data.save()
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)

    return redirect(reverse('idea_list', args=()))


def idea_view(request, idea_int):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        idea = backend_models.IdeaModel.objects.get(id=idea_int)
        comments = backend_models.IdeaCommentModel.objects.filter(idea_foreign_key_field=idea)
        try:
            page = backend_service.PaginationClass.paginate(request=request, objects=comments, num_page=5)
            response = 0
        except Exception as error:
            response = -1
            backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': response,
            'idea': idea,
            'page': page,
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': -1,
            'idea': None,
            'page': None,
        }

    return render(request, 'idea/idea_view.html', context)


def idea_like(request, idea_int):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        idea = backend_models.IdeaModel.objects.get(id=idea_int)
        author = backend_models.UserModel.objects.get(user_foreign_key_field=request.user)
        if request.POST['status'] == 'like':
            try:
                backend_models.IdeaRatingModel.objects.get(
                    author_foreign_key_field=author,
                    idea_foreign_key_field=idea,
                    status_boolean_field=True
                ).delete()
            except Exception as error:
                print(error)
                backend_models.IdeaRatingModel.objects.create(
                    author_foreign_key_field=author,
                    idea_foreign_key_field=idea,
                    status_boolean_field=True
                )
            try:
                backend_models.IdeaRatingModel.objects.get(
                    author_foreign_key_field=author,
                    idea_foreign_key_field=idea,
                    status_boolean_field=False
                ).delete()
            except Exception as error:
                print(error)
        else:
            try:
                backend_models.IdeaRatingModel.objects.get(
                    author_foreign_key_field=author,
                    idea_foreign_key_field=idea,
                    status_boolean_field=False
                ).delete()
            except Exception as error:
                print(error)
                backend_models.IdeaRatingModel.objects.create(
                    author_foreign_key_field=author,
                    idea_foreign_key_field=idea,
                    status_boolean_field=False
                )
                backend_models.IdeaCommentModel.objects.create(
                    author_foreign_key_field=backend_models.UserModel.objects.get(user_foreign_key_field=request.user),
                    idea_foreign_key_field=backend_models.IdeaModel.objects.get(id=idea_int),
                    text_field=request.POST['text_field']
                )
            try:
                backend_models.IdeaRatingModel.objects.get(
                    author_foreign_key_field=author,
                    idea_foreign_key_field=idea,
                    status_boolean_field=True
                ).delete()
            except Exception as error:
                print(error)
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)

    return redirect(reverse('idea_view', args=(idea_int,)))


def idea_comment(request, idea_int):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        if request.method == 'POST':
            backend_models.IdeaCommentModel.objects.create(
                author_foreign_key_field=backend_models.UserModel.objects.get(user_foreign_key_field=request.user),
                idea_foreign_key_field=backend_models.IdeaModel.objects.get(id=idea_int),
                text_field=backend_service.DjangoClass.RequestClass.get_value(request, "text_field")
            )
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)

    return redirect(reverse('idea_view', args=(idea_int,)))


def idea_rating(request):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        idea = backend_models.IdeaModel.objects.order_by('-id')
        authors = []
        for query in idea:
            authors.append(query.author_foreign_key_field)
        authors_dict = {}
        for author in authors:
            authors_dict[author] = authors.count(author)
        user_counts = []
        for author in authors_dict:
            ideas = backend_models.IdeaModel.objects.filter(author_foreign_key_field=author)
            total_rating = 0
            for idea in ideas:
                total_rating += idea.get_total_rating()
            user_counts.append(
                {'author': author, 'count': ideas.count(), 'rating': total_rating}
            )
        sorted_by_rating = True
        if request.method == 'POST':
            if request.POST['sorted'] == 'idea':
                sorted_by_rating = True
            if request.POST['sorted'] == 'count':
                sorted_by_rating = False
        if sorted_by_rating:
            page = sorted(user_counts, key=lambda k: k['rating'], reverse=True)
        else:
            page = sorted(user_counts, key=lambda k: k['count'], reverse=True)
        page = backend_service.PaginationClass.paginate(request=request, objects=page, num_page=5)
        response = 0
        context = {
            'response': response,
            'page': page,
            'sorted': sorted_by_rating
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': -1,
            'page': None,
            'sorted': None
        }

    return render(request, 'idea/idea_rating.html', context)


def salary_(request):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        date = [int(x) for x in str(backend_utils.DateTimeUtils.get_current_date()).split('-')]
        year = date[0]
        month = date[1] - 1
        day = date[2]
        if day < 10:
            month -= 1
        if month == 0:
            month = 12
            year -= 1
        elif month < 0:
            month = 11
            year -= 1
        months = []
        months_list = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 'Июль', 'Август', 'Сентябрь', 'Октябрь',
                       'Ноябрь', 'Декабрь']
        for month_list in months_list:
            _index = months_list.index(month_list) + 1
            if _index == month:
                months.append([_index, month_list, True])
            else:
                months.append([_index, month_list, False])
        years = []
        years_list = [2021, 2022, 2023, 2024, 2025]
        for year_list in years_list:
            _index = years_list.index(year) + 1
            if _index == month:
                years.append([year_list, True])
            else:
                years.append([year_list, False])
        data = None
        response = 0
        if request.method == 'POST':
            key = backend_service.UtilsClass.create_encrypted_password(
                _random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890',
                _length=10
            )
            hash_key_obj = hashlib.sha256()
            hash_key_obj.update(key.encode('utf-8'))
            key_hash = str(hash_key_obj.hexdigest().strip().upper())
            key_hash_base64 = base64.b64encode(str(key_hash).encode()).decode()
            iin = request.user.username
            if str(request.user.username).lower() == 'bogdan':
                iin = 970801351179
            iin_base64 = base64.b64encode(str(iin).encode()).decode()
            month = backend_service.DjangoClass.RequestClass.get_value(request, "month")
            if int(month) < 10:
                month = f'0{month}'
            year = backend_service.DjangoClass.RequestClass.get_value(request, "year")
            date_base64 = base64.b64encode(f'{year}{month}'.encode()).decode()
            # url = f'http://192.168.1.158/Tanya_perenos/hs/zp/rl/{iin_base64}_{key_hash_base64}/{date_base64}'
            url = f'http://192.168.1.10/KM_1C/hs/zp/rl/{iin_base64}_{key_hash_base64}/{date_base64}'
            relative_path = os.path.dirname(os.path.abspath('__file__')) + '\\'
            h = httplib2.Http(
                relative_path + "\\static\\media\\data\\temp\\get_salary_data")
            _login = 'Web_adm_1c'
            password = '159159qqww!'
            h.add_credentials(_login, password)
            response, content = h.request(url)
            if content:
                _message = backend_service.UtilsClass.decrypt_text_with_hash(content.decode()[1:], key_hash)
                success = True
                error_word_list = ['Ошибка', 'ошибка', 'Error', 'error', 'Failed', 'failed']
                for error_word in error_word_list:
                    if _message.find(error_word) >= 0:
                        success = False
                if _message.find('send') == 0:
                    data = _message.split('send')[1].strip()
                    success = False
            else:
                success = False
            if success:
                json_data = json.loads(
                    backend_service.UtilsClass.decrypt_text_with_hash(content.decode()[1:], key_hash))
                with open("static/media/data/zarplata.json", "w", encoding="utf-8") as file:
                    encode_data = json.dumps(json_data, ensure_ascii=False)
                    json.dump(encode_data, file, ensure_ascii=False)
            if True:
                # Временное чтение файла для отладки без доступа к 1С
                with open("static/media/data/zarplata_temp.json", "r", encoding="utf-8") as file:
                    json_data = json.load(file)

                try:
                    json_data["global_objects"]["3.Доходы в натуральной форме"]
                except Exception as error:
                    print(error)
                    json_data["global_objects"]["3.Доходы в натуральной форме"] = {
                        "Fields": {
                            "1": "Вид",
                            "2": "Период",
                            "3": "Сумма"
                        },
                    }
                print(json_data)
                new_data = dict(json_data).copy()
                del (new_data["global_objects"])
                new_arr = []
                for key, value in new_data.items():
                    if key == 'Долг за организацией на начало месяца' or key == 'Долг за организацией на конец месяца':
                        continue
                    new_arr.append([key, value])
                data = {
                    "Table_0_1": new_arr[:len(new_arr) // 2],
                    "Table_0_2": new_arr[len(new_arr) // 2:],
                    "Table_1": backend_service.SalaryClass.create_arr_table(
                        title="1.Начислено",
                        footer="Всего начислено",
                        json_obj=json_data["global_objects"]["1.Начислено"],
                        exclude=[5, 6]
                    ),
                    "Table_2": backend_service.SalaryClass.create_arr_table(
                        title="2.Удержано",
                        footer="Всего удержано",
                        json_obj=json_data["global_objects"]["2.Удержано"],
                        exclude=[]
                    ),
                    "Table_3": backend_service.SalaryClass.create_arr_table(
                        title="3.Доходы в натуральной форме",
                        footer="Всего натуральных доходов",
                        json_obj=json_data["global_objects"]["3.Доходы в натуральной форме"],
                        exclude=[]
                    ),
                    "Table_4": backend_service.SalaryClass.create_arr_table(
                        title="4.Выплачено",
                        footer="Всего выплат",
                        json_obj=json_data["global_objects"]["4.Выплачено"],
                        exclude=[]
                    ),
                    "Table_5": backend_service.SalaryClass.create_arr_table(
                        title="5.Налоговые вычеты",
                        footer="Всего вычеты",
                        json_obj=json_data["global_objects"]["5.Налоговые вычеты"],
                        exclude=[]
                    ),
                    "Down": {
                        "first": [
                            "Долг за организацией на начало месяца",
                            json_data["Долг за организацией на начало месяца"]
                        ],
                        "last": ["Долг за организацией на конец месяца",
                                 json_data["Долг за организацией на конец месяца"]],
                    },
                    "Final": [
                        ["Период", json_data["Период"]],
                        ["Долг за организацией на конец месяца", json_data["Долг за организацией на конец месяца"]],
                    ],
                }
                response = 1
        context = {
            'response': response,
            'months': months,
            'years': years,
            'data': data,
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'response': -1,
            'months': None,
            'years': None,
            'data': None,
        }
    return render(request, 'salary/salary.html', context)


def salary_pdf_(request):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    # try:
    if True:
        template_path = 'salary/salary_pdf.html'
        key = backend_service.UtilsClass.create_encrypted_password(
            _random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890', _length=10
        )
        print('\n ***************** \n')
        print(f"key: {key}")
        hash_key_obj = hashlib.sha256()
        hash_key_obj.update(key.encode('utf-8'))
        key_hash = str(hash_key_obj.hexdigest().strip().upper())
        print('\n ***************** \n')
        print(f"key_hash: {key_hash}")
        key_hash_base64 = base64.b64encode(str(key_hash).encode()).decode()
        print('\n ***************** \n')
        print(f"key_hash_base64: {key_hash_base64}")

        iin = request.user.username
        if str(request.user.username).lower() == 'bogdan':
            iin = 970801351179
        print('\n ***************** \n')
        print(f"iin: {iin}")
        iin_base64 = base64.b64encode(str(iin).encode()).decode()
        print('\n ***************** \n')
        print(f"iin_base64: {iin_base64}")
        print('\n ***************** \n')

        month = 10
        if int(month) < 10:
            month = f'0{month}'
        year = 2021
        date = f'{year}{month}'
        print(f"date: {date}")
        date_base64 = base64.b64encode(str(date).encode()).decode()
        print('\n ***************** \n')
        print(f"date_base64: {date_base64}")

        # url = f'http://192.168.1.158/Tanya_perenos/hs/zp/rl/{iin_base64}_{key_hash_base64}/{date_base64}'
        url = f'http://192.168.1.10/KM_1C/hs/zp/rl/{iin_base64}_{key_hash_base64}/{date_base64}'
        print('\n ***************** \n')
        print(f"url: {url}")

        relative_path = os.path.dirname(os.path.abspath('__file__')) + '\\'
        h = httplib2.Http(
            relative_path + "\\static\\media\\data\\temp\\get_salary_data")
        _login = 'Web_adm_1c'
        password = '159159qqww!'
        h.add_credentials(_login, password)
        try:
            response, content = h.request(url)
        except Exception as error:
            backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
            content = None
        success_web_read = False
        if content:

            print('\n ***************** \n')
            print(f"content: {content}")

            print('\n ***************** \n')
            print(f"content_utf: {content.decode()}")

            content_decrypt = backend_service.UtilsClass.decrypt_text_with_hash(
                content.decode(encoding='UTF-8', errors='strict')[1:], key_hash
            )
            print('\n ***************** \n')
            print(f"content_decrypt: {content_decrypt}")

            success = True
            error_word_list = ['Ошибка', 'ошибка',
                               'Error', 'error', 'Failed', 'failed']
            for error_word in error_word_list:
                if str(content.decode()).find(error_word) >= 0:
                    success = False
            if success:
                try:
                    json_data = json.loads(
                        backend_service.UtilsClass.decrypt_text_with_hash(content.decode()[1:], key_hash))
                    with open("static/media/data/zarplata.json", "w", encoding="utf-8") as file:
                        encode_data = json.dumps(json_data, ensure_ascii=False)
                        json.dump(encode_data, file, ensure_ascii=False)
                    success_web_read = True
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        if success_web_read is False:
            print('read temp file')
            with open("static/media/data/zarplata_temp.json", "r", encoding="utf-8") as file:
                json_data = json.load(file)

        print('\n ***************** \n')
        print(f"json_data: {json_data}")
        print('\n ***************** \n')

        try:
            json_data["global_objects"]["3.Доходы в натуральной форме"]
        except Exception as error:
            backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
            json_data["global_objects"]["3.Доходы в натуральной форме"] = {
                "Fields": {
                    "1": "Вид",
                    "2": "Период",
                    "3": "Дни",
                    "4": "Часы",
                    "5": "Сумма",
                    "6": "ВсегоДни",
                    "7": "ВсегоЧасы"
                },
            }

        data = {
            "Table_1": backend_service.SalaryClass.create_arr_table(
                title="1.Начислено", footer="Всего начислено", json_obj=json_data["global_objects"]["1.Начислено"],
                exclude=[5, 6]
            ),
            "Table_2": backend_service.SalaryClass.create_arr_table(
                title="2.Удержано", footer="Всего удержано", json_obj=json_data["global_objects"]["2.Удержано"],
                exclude=[]
            ),
            "Table_3": backend_service.SalaryClass.create_arr_table(
                title="3.Доходы в натуральной форме", footer="Всего натуральных доходов",
                json_obj=json_data["global_objects"]["3.Доходы в натуральной форме"], exclude=[
                ]
            ),
            "Table_4": backend_service.SalaryClass.create_arr_table(
                title="4.Выплачено", footer="Всего выплат", json_obj=json_data["global_objects"]["4.Выплачено"],
                exclude=[]
            ),
            "Table_5": backend_service.SalaryClass.create_arr_table(
                title="5.Налоговые вычеты", footer="Всего вычеты",
                json_obj=json_data["global_objects"]["5.Налоговые вычеты"],
                exclude=[]
            ),
            "Down": {
                "first": ["Долг за организацией на начало месяца", json_data["Долг за организацией на начало месяца"]],
                "last": ["Долг за организацией на конец месяца", json_data["Долг за организацией на конец месяца"]],
            },
        }
        context = {
            'data': data,
            'STATIC_ROOT': backend_settings.STATIC_ROOT,
        }
        # Create a Django response object, and specify content_type as pdf
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="report.pdf"'
        # find the template and render it.
        template = get_template(template_path)
        html = template.render(context)
        # create a pdf
        pisa_status = pisa.CreatePDF(html, dest=response, encoding='utf-8',
                                     link_callback=backend_service.Xhtml2pdfClass.link_callback)
        # template = render_to_string(template_path, context)
        # pdf = pisa.pisaDocument(io.BytesIO(template.encode('UTF-8')), response,
        #                         encoding='utf-8',
        #                         link_callback=link_callback)
        # pdf = pisa.pisaDocument(io.StringIO(html), response, encoding='UTF-8')
        # if error then show some funny view
        if pisa_status.err:
            return HttpResponse('We had some errors <pre>' + html + '</pre>')
    # except Exception as error:
    #     backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
    #     response = None
    return response


def career(request):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        data = None
        response = 0
        if request.method == 'POST':
            headers = {
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:45.0) Gecko/20100101 Firefox/45.0'
            }
            vacancies_urls = []
            url = 'https://www.km-open.online/property'
            r = requests.get(url, headers=headers)
            soup = bs4.BeautifulSoup(r.content.decode("utf-8"), features="lxml")
            list_objs = soup.find_all('div', {"class": "collection-item w-dyn-item"})
            for list_obj in list_objs:
                vacancies_urls.append(url.split('/property')[0] + str(list_obj).split('href="')[1].split('"')[0])
            vacancies_title = []
            for list_obj in list_objs:
                vacancies_title.append(str(list_obj).split('class="heading-12">')[1].split('</h5>')[0])
            vacancies_data = []
            for title in vacancies_title:
                vacancies_data.append([title, vacancies_urls[vacancies_title.index(title)]])
            data = [["Вакансия"], vacancies_data]
            response = 1
        context = {
            'data': data,
            'response': response
        }
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {
            'data': None,
            'response': None,
        }

    return render(request, 'hr/career.html', context)


def video_study(request):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        context = {}
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
        context = {}

    return render(request, 'news/video_study.html', context)


def chat(request):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        if request.method == 'POST':
            text = backend_service.DjangoClass.RequestClass.get_value(request, "text")
            try:
                user = User.objects.get(username=request.user.username)
                user_model = backend_models.UserModel.objects.get_or_create(user_foreign_key_field=user)[0]
                usr = f'{user_model.last_name_char_field} {user_model.first_name_char_field}'
                if len(usr.strip()) <= 1:
                    usr = 'Не заполнено'
            except Exception as error:
                print(error)
                usr = 'скрыто'

            backend_models.ChatModel.objects.create(
                author_char_field=usr,
                text_field=text
            )
        context = {}
        return render(request, 'chat/chat.html', context)
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)


def chat_react(request):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        if request.method == 'POST':
            pass
        context = {}
        return render(request, 'index.html', context)
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)


def react(request):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        if request.method == 'POST':
            pass
        context = {}
        return render(request, 'chat/chat_react.html', context)
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)


def geo(request):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    try:
        data = None
        form = backend_forms.GeoForm()
        if request.method == 'POST':
            print('begin')

            # data = generate_xlsx(request)
            # print('generate_xlsx successfully')

            # generate_kml()
            # print('generate_kml successfully')

            # Points = [PointName, latitude, longitude, PointLinks]

            # point1 = [61.22812, 52.14303, "1", "2"]
            # point2 = [61.22829, 52.1431, "2", "1|3"]
            # point3 = [61.22862, 52.14323, "3", "2|4"]
            # point4 = [61.22878, 52.14329, "4", "3|5"]
            # point5 = [61.22892201, 52.14332617, "5", "4|6"]
            # point_arr = [point1, point2, point3, point4, point5]

            # Получение значений из формы
            count_points = int(request.POST['count_points'])
            correct_rad = int(request.POST['correct_rad'])
            rounded_val = int(request.POST['rounded_val'])

            points_arr = []
            val = 0
            for num in range(1, count_points):
                x = 61.22812
                y = 52.14303
                val += random.random() / 10000 * 2
                var = [round(x + val, rounded_val), round(y + val - random.random() / 10000 * correct_rad, rounded_val),
                       str(num), str(f"{num - 1}|{num + 1}")]
                points_arr.append(var)

            # Near Point
            subject_ = backend_service.GeoClass.find_near_point(points_arr, 61.27, 52.147)
            print(subject_)
            object_ = backend_service.GeoClass.find_near_point(points_arr, 61.24, 52.144)
            print(object_)

            # Vectors = [VectorName, length(meters)]
            # vector_arr = backend_service.GeoClass.get_vector_arr(points_arr)
            # print(vector_arr)

            # print(points_arr)

            # New KML Object
            backend_service.GeoClass.generate_way(object_, subject_, points_arr)

            print('end')
        context = {
            'data': data,
            'form': form,
        }
        return render(request, 'extra/geo.html', context)
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)


def analyse(request):
    """
    Машинное зрение
    """
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    # response = requests.get(url='http://127.0.0.1/drf/users/2/', timeout=3)
    # print(response.text)

    login_ = 'Bogdan'
    password = '31284bogdan'
    h = httplib2.Http(backend_utils.DirPathFolderPathClass.create_folder_in_this_dir(folder_name='static/media/temp'))
    h.add_credentials(login_, password)
    response, content = h.request('http://127.0.0.1/drf/ideas/')
    print(response)
    print(content.decode())

    # try:
    #     with ThreadPoolExecutor() as executor:
    #         executor.submit(ComputerVisionClass.EventLoopClass.loop_modules_global, tick_delay=0.1)
    # except Exception as error:
    #     backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
    #     print(f'\nanalyse | error : {error}\n')
    #     with ThreadPoolExecutor() as executor:
    #         executor.submit(ComputerVisionClass.EventLoopClass.loop_modules_global, tick_delay=0.2)

    return redirect(to='home')


def passages_thermometry(request):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    data = None
    if request.method == 'POST':
        date_start = backend_service.DjangoClass.RequestClass.get_value(request, 'date_start').split('T')[0]
        date_end = backend_service.DjangoClass.RequestClass.get_value(request, 'date_end').split('T')[0]
        check = backend_service.DjangoClass.RequestClass.get_check(request, 'check')
        personid = backend_service.DjangoClass.RequestClass.get_value(request, 'personid')
        connect_db = backend_utils.SQLClass.pyodbc_connect(ip="192.168.15.87", server="DESKTOP-SM7K050", port="1434",
                                                           database="thirdpartydb", username="sa",
                                                           password="skud12345678")
        cursor = connect_db.cursor()
        cursor.fast_executemany = True
        try:
            if check:
                sql_select_query = f"SELECT * " \
                                   f"FROM dbtable " \
                                   f"WHERE date1 BETWEEN '{date_start}' AND '{date_end}' AND personid = '{personid}' " \
                                   f"AND CAST(temperature AS FLOAT) >= 37.0 " \
                                   f"ORDER BY date1 DESC, date2 DESC;"
            else:
                sql_select_query = f"SELECT * " \
                                   f"FROM dbtable " \
                                   f"WHERE date1 BETWEEN '{date_start}' AND '{date_end}' " \
                                   f"AND CAST(temperature AS FLOAT) >= 37.0 " \
                                   f"ORDER BY date1 DESC, date2 DESC;"
            cursor.execute(sql_select_query)
            data = cursor.fetchall()
            bodies = []
            for row in data:
                local_bodies = []
                value_index = 0
                for val in row:
                    if value_index == 4:
                        try:
                            val = val.encode('1251').decode('utf-8')
                        except Exception as error:
                            backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                            try:
                                value = str(val).split(" ")
                                try:
                                    name = value[0].encode('1251').decode('utf-8')
                                except Exception as error:
                                    backend_service.DjangoClass.LoggingClass.logging_errors(request=request,
                                                                                            error=error)
                                    name = "И" + \
                                           value[0][2:].encode('1251').decode('utf-8')
                                try:
                                    surname = value[1].encode(
                                        '1251').decode('utf-8')
                                except Exception as error:
                                    backend_service.DjangoClass.LoggingClass.logging_errors(request=request,
                                                                                            error=error)
                                    surname = "И" + \
                                              value[1][2:].encode('1251').decode('utf-8')
                                string = name + " " + surname
                                val = string
                            except Exception as error:
                                backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                    value_index += 1
                    local_bodies.append(val)
                bodies.append(local_bodies)
            headers = ["табельный", "доступ", "дата", "время", "данные", "точка", "номер карты", "температура",
                       "маска", "алкотест"]
            data = [headers, bodies]
        except Exception as error:
            backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
    context = {
        'data': data,
    }
    return render(request, 'skud/passages_thermometry.html', context)


def passages_select(request):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    data = None
    if request.method == 'POST':
        personid = str(request.POST['personid'])
        connect_db = backend_utils.SQLClass.pyodbc_connect(ip="192.168.15.87", server="DESKTOP-SM7K050", port="1434",
                                                           database="thirdpartydb", username="sa",
                                                           password="skud12345678")
        cursor = connect_db.cursor()
        cursor.fast_executemany = True
        try:
            check = request.POST['check']
            date = str(request.POST['date']).split('T')[0]
            time = str(request.POST['date']).split('T')[1]
            sql_select_query = f"SELECT * " \
                               f"FROM dbtable " \
                               f"WHERE date1 = '{date}' AND date2 BETWEEN '{time}:00' AND '{time}:59' " \
                               f"AND personid = '{personid}' " \
                               f"ORDER BY date1 DESC, date2 DESC;"
        except Exception as error:
            backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
            sql_select_query = f"SELECT * " \
                               f"FROM dbtable " \
                               f"WHERE date1 BETWEEN '2021-07-30' AND '2023-12-31' AND personid = '{personid}' " \
                               f"ORDER BY date1 DESC, date2 DESC;"
        cursor.execute(sql_select_query)
        print(sql_select_query)
        print(data)
        data = cursor.fetchall()
        bodies = []
        for row in data:
            local_bodies = []
            value_index = 0
            for val in row:
                if value_index == 4:
                    try:
                        val = val.encode('1251').decode('utf-8')
                    except Exception as error:
                        backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                        try:
                            value = str(val).split(" ")
                            try:
                                name = value[0].encode('1251').decode('utf-8')
                            except Exception as error:
                                backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                                name = "И" + \
                                       value[0][2:].encode('1251').decode('utf-8')
                            try:
                                surname = value[1].encode(
                                    '1251').decode('utf-8')
                            except Exception as error:
                                backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                                surname = "И" + \
                                          value[1][2:].encode('1251').decode('utf-8')
                            string = name + " " + surname
                            val = string
                        except Exception as error:
                            backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                value_index += 1
                local_bodies.append(val)
            bodies.append(local_bodies)
        headers = ["табельный", "доступ", "дата", "время", "данные",
                   "точка", "номер карты", "температура", "маска"]
        data = [headers, bodies]
    context = {
        'data': data,
    }
    return render(request, 'skud/passages_select.html', context)


def passages_update(request):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    data = None
    if request.method == 'POST':
        personid_old = request.POST['personid_old']
        date_old = str(request.POST['datetime_old']).split('T')[0]
        time_old = str(request.POST['datetime_old']).split('T')[1]
        date_new = str(request.POST['datetime_new']).split('T')[0]
        time_new = str(request.POST['datetime_new']).split('T')[1] + ':00'
        accessdateandtime_new = date_new + 'T' + time_new
        connect_db = backend_utils.SQLClass.pyodbc_connect(ip="192.168.15.87", server="DESKTOP-SM7K050", port="1434",
                                                           database="thirdpartydb", username="sa",
                                                           password="skud12345678")
        cursor = connect_db.cursor()
        cursor.fast_executemany = True
        value = f"UPDATE dbtable SET accessdateandtime = '{accessdateandtime_new}', date1 = '{date_new}', " \
                f"date2 = '{time_new}', alko = '0' " \
                f"WHERE date1 = '{date_old}' AND date2 BETWEEN '{time_old}:00' AND '{time_old}:59' " \
                f"AND personid = '{personid_old}' "
        cursor.execute(value)
        connect_db.commit()
    context = {
        'data': data,
    }
    return render(request, 'skud/passages_update.html', context)


def passages_insert(request):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    data = None
    if request.method == 'POST':
        personid = request.POST['personid']
        date = str(request.POST['datetime']).split('T')[0]
        time = str(request.POST['datetime']).split('T')[1] + ':00'
        accessdateandtime = date + 'T' + time
        devicename = str(request.POST['devicename'])
        cardno = str(request.POST['cardno'])
        temperature = str(request.POST['temperature'])
        if temperature == '0':
            temperature = ''
        mask = str(request.POST['mask'])
        try:
            connect_db = backend_utils.SQLClass.pyodbc_connect(
                ip="192.168.15.87", server="DESKTOP-SM7K050", port="1434", database="thirdpartydb", username="sa",
                password="skud12345678"
            )
            cursor = connect_db.cursor()
            cursor.fast_executemany = True
            sql_select_query = f'SELECT TOP (1) personname ' \
                               f'FROM dbtable ' \
                               f'WHERE personid = \'{personid}\' ' \
                               f'ORDER BY date1 DESC, date2 DESC;'
            cursor.execute(sql_select_query)
            personname_all = cursor.fetchall()
            personname = personname_all[0][0]
        except Exception as error:
            backend_service.DjangoClass.LoggingClass.logging_errors(request=request, error=error)
            personname = 'None'
        connection = backend_utils.SQLClass.pyodbc_connect(ip="192.168.15.87", server="DESKTOP-SM7K050", port="1434",
                                                           database="thirdpartydb", username="sa",
                                                           password="skud12345678")
        cursor = connection.cursor()
        cursor.fast_executemany = True
        rows = ['personid', 'accessdateandtime', 'date1', 'date2', 'personname', 'devicename', 'cardno',
                'temperature', 'mask', 'alko']
        values = [personid, accessdateandtime, date, time,
                  personname, devicename, cardno, temperature, mask, 0]
        _rows = ''
        for x in rows:
            _rows = f"{_rows}{str(x)}, "
        value = f"INSERT INTO dbtable (" + \
                _rows[:-2:] + f") VALUES {tuple(values)}"
        cursor.execute(value)
        connection.commit()
    context = {
        'data': data,
    }
    return render(request, 'skud/passages_insert.html', context)


def passages_delete(request):
    # access and logging
    page = backend_service.DjangoClass.AuthorizationClass.try_to_access(request=request, access='superuser')
    if page:
        return redirect(page)

    data = None
    if request.method == 'POST':
        personid = str(request.POST['personid'])
        date = str(request.POST['datetime']).split('T')[0]
        time = str(request.POST['datetime']).split('T')[1]
        connect_db = backend_utils.SQLClass.pyodbc_connect(ip="192.168.15.87", server="DESKTOP-SM7K050", port="1434",
                                                           database="thirdpartydb", username="sa",
                                                           password="skud12345678")
        cursor = connect_db.cursor()
        cursor.fast_executemany = True
        value = f"DELETE FROM dbtable " \
                f"WHERE date1 = '{date}' AND date2 BETWEEN '{time}:00' AND '{time}:59' AND personid = '{personid}' "
        cursor.execute(value)
        connect_db.commit()
    context = {
        'data': data,
    }
    return render(request, 'skud/passages_delete.html', context)
