# TODO download modules ################################################################################################
import base64
import datetime
import hashlib
import json
import os
from concurrent.futures import ThreadPoolExecutor
from typing import Union
import httplib2
import openpyxl
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill
from openpyxl.utils import get_column_letter
# TODO django modules ##################################################################################################
from django.contrib.auth import authenticate
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User, Group, update_last_login
from django.core.mail import send_mail
from django.shortcuts import render
from django.utils import timezone
# TODO drf modules #####################################################################################################
from rest_framework import viewsets, permissions
from rest_framework.authentication import BasicAuthentication
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
# TODO custom modules ##################################################################################################
from backend import models as backend_models, serializers as backend_serializers, service as backend_service, \
    settings as backend_settings


# TODO index ###########################################################################################################
@permission_classes([AllowAny])
def index(request):
    """
    index React Single Page Applications
    """

    try:
        # TODO Request #################################################################################################
        req_inst = backend_service.DjangoClass.TemplateClass.request(request=request)
        # TODO Methods #################################################################################################
        if req_inst.method == "GET":
            # TODO Actions #############################################################################################
            if req_inst.action_type == "":
                try:
                    response = {"response": "Данные успешно получены!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return render(request, 'index.html', response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return render(request, "backend/404.html")
            else:
                return Response({"error": "This action not allowed for this method."})
        else:
            return Response({"error": "This method not allowed for this endpoint."})
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
        return render(request, "backend/404.html")


# TODO routes ##########################################################################################################
@api_view(http_method_names=["GET", "POST", "PUT", "DELETE"])
@authentication_classes([BasicAuthentication])
def api_auth_routes(request):
    """
    routes django-rest-framework
    """

    try:
        # TODO Request #################################################################################################
        req_inst = backend_service.DjangoClass.TemplateClass.request(request=request)
        # TODO Methods #################################################################################################
        if req_inst.method == "GET":
            # TODO Actions #############################################################################################
            if req_inst.action_type == "":
                try:
                    response = {"response": [
                        {
                            'Endpoints': '$BASEPATH$/router/',
                            'methods': 'GET, HEAD, OPTIONS',
                            'descriptions': 'default router for rest_framework ViewSet',
                            'helps': '''$BASEPATH$/router/''',
                            'code': '''from rest_framework import routers\n
                            router = routers.DefaultRouter()\n
                            router.register(prefix=r'users', viewset=views.UserViewSet, basename='users')'''
                        },
                        {
                            'name': '''JWT token''',
                            'endpoint': '''$BASEPATH$/tokens/''',
                            'method': '''...''',
                            'body': '''...''',
                            'descriptions': '''All endpoints for "JWT token" with django-rest_framework api''',
                            'helps': '''$BASEPATH$/tokens/''',
                            'code': '''...'''
                        },
                        {
                            'name': '''get token''',
                            'endpoint': '''$BASEPATH$/token/''',
                            'method': '''...''',
                            'body': '''...''',
                            'descriptions': '''All actions for "note" with django-rest_framework api''',
                            'helps': '''...''',
                            'code': '''...'''
                        },
                        {
                            'name': '''refresh token''',
                            'endpoint': '''$BASEPATH$/token/refresh/''',
                            'method': '''...''',
                            'body': '''...''',
                            'descriptions': '''refresh token with django-rest_framework api''',
                            'helps': '''-''',
                            'code': '''...'''
                        },
                        {
                            'name': '''verify token''',
                            'endpoint': '''$BASEPATH$/token/verify/''',
                            'method': '''...''',
                            'body': '''...''',
                            'descriptions': '''verify token with django-rest_framework api''',
                            'helps': '''-''',
                            'code': '''...'''
                        },
                        {
                            'name': '''note_api''',
                            'endpoint': '''$BASEPATH$/note_api/ && $BASEPATH$/note_api/<id>/''',
                            'method': '''...''',
                            'body': '''...''',
                            'descriptions': '''All actions for "note" with django-rest_framework api''',
                            'helps': '''$BASEPATH$/note_api/-1/''',
                            'code': '''...'''
                        },
                    ]}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            else:
                return Response({"error": "This action not allowed for this method."})
        else:
            return Response({"error": "This method not allowed for this endpoint."})
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
        return render(request, "backend/404.html")


# TODO view sets #######################################################################################################
class ExamplesModelViewSet(viewsets.ModelViewSet):
    queryset = backend_models.ExamplesModel.objects.all()
    serializer_class = backend_serializers.ExamplesModelSerializer
    permission_classes = [permissions.AllowAny]


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = backend_serializers.UserSerializer
    permission_classes = [permissions.AllowAny]


class GroupViewSet(viewsets.ModelViewSet):
    queryset = Group.objects.all()
    serializer_class = backend_serializers.GroupSerializer
    permission_classes = [permissions.AllowAny]


# TODO base views ######################################################################################################
@api_view(http_method_names=["GET", "POST", "PUT", "DELETE"])
@permission_classes([AllowAny])
def api_any_user(request):
    """
    any user django-rest-framework
    """

    try:
        # TODO Request #################################################################################################
        req_inst = backend_service.DjangoClass.TemplateClass.request(request=request)
        # TODO Methods #################################################################################################
        if req_inst.method == "POST":
            # TODO Actions #############################################################################################
            if req_inst.action_type == "USER_LOGIN":
                try:
                    # TODO get_value ###################################################################################
                    username = req_inst.get_value("username")
                    password = req_inst.get_value("password")
                    # TODO actions #####################################################################################
                    access_count = 0
                    for log in backend_models.LoggingModel.objects.filter(
                            username_slug_field=username,
                            ip_genericipaddress_field=req_inst.ip,
                            request_path_slug_field=req_inst.path,
                            request_method_slug_field=f"{req_inst.method} | {req_inst.action_type}",
                            error_text_field="-"
                    ):
                        if (log.created_datetime_field +
                            datetime.timedelta(hours=6, minutes=59)).strftime('%Y-%m-%d %H:%M') >= \
                                (datetime.datetime.now()).strftime('%Y-%m-%d %H:%M'):
                            access_count += 1
                    if access_count < 20:
                        backend_models.LoggingModel.objects.create(
                            username_slug_field=username,
                            ip_genericipaddress_field=req_inst.ip,
                            request_path_slug_field=req_inst.path,
                            request_method_slug_field=f"{req_inst.method} | {req_inst.action_type}",
                            error_text_field="-"
                        )
                        is_authenticated = authenticate(username=username, password=password)
                        if is_authenticated is not None:
                            user = User.objects.get(username=username)
                            user_model = backend_models.UserModel.objects.get(user_foreign_key_field=user)
                            if user_model.activity_boolean_field is False:
                                return Response({"error": "Внимание, Ваш аккаунт заблокирован!"})
                            update_last_login(sender=None, user=user)
                            refresh = RefreshToken.for_user(user=user)
                            response = {"response": {
                                "refresh": str(refresh),
                                "access": str(refresh.access_token),
                                "token": str(refresh.access_token),
                                "username": str(user.username),
                                "name": str(f'{user_model.last_name_char_field} {user_model.first_name_char_field}'),
                            }}
                        else:
                            response = {"error": "Внимание, данные не совпадают!"}
                    else:
                        response = {"error": "Внимание, попыток входа можно совершать не более 20 в час!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "FIND_USER":
                try:
                    # TODO get_value ###################################################################################
                    username = req_inst.get_value("username")
                    # TODO actions #####################################################################################
                    try:
                        user_model = backend_models.UserModel.objects.get(
                            user_foreign_key_field=User.objects.get(username=username)
                        )
                        if user_model.temp_password_boolean_field:
                            response = {"error": {"Пользователь ещё ни разу не менял пароль!"}}
                        else:
                            response = {"response": {
                                "username": str(user_model.user_foreign_key_field.username),
                                "secretQuestion": str(user_model.secret_question_char_field),
                                "email": str(user_model.email_field),
                                "success": False
                            }}
                    except Exception as error:
                        response = {"error": "Пользователя не существует или произошла ошибка!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "CHECK_ANSWER":
                try:
                    # TODO get_value ###################################################################################
                    username = req_inst.get_value("username")
                    secret_answer_char_field = req_inst.get_value("secretAnswer")
                    # TODO actions #####################################################################################
                    user_model = backend_models.UserModel.objects.get(
                        user_foreign_key_field=User.objects.get(username=username)
                    )
                    if str(secret_answer_char_field).strip().lower() == \
                            str(user_model.secret_answer_char_field).strip().lower():
                        response = {"response": {
                            "username": user_model.user_foreign_key_field.username,
                            "success": True
                        }}
                    else:
                        response = {"error": "Секретный ответ не совпадает!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "SEND_EMAIL_PASSWORD":
                try:
                    # TODO get_value ###################################################################################
                    username = req_inst.get_value("username")
                    # TODO actions #####################################################################################
                    user_model = backend_models.UserModel.objects.get(
                        user_foreign_key_field=User.objects.get(username=username)
                    )
                    access_count = 0
                    for log in backend_models.LoggingModel.objects.filter(
                            username_slug_field=username,
                            ip_genericipaddress_field=req_inst.ip,
                            request_path_slug_field=req_inst.path,
                            request_method_slug_field=f"{req_inst.method} | {req_inst.action_type}",
                            error_text_field="-",
                    ):
                        if (log.created_datetime_field +
                            datetime.timedelta(hours=6, minutes=3)).strftime('%Y-%m-%d %H:%M') \
                                >= (datetime.datetime.now()).strftime('%Y-%m-%d %H:%M'):
                            access_count += 1
                    if access_count < 1:
                        backend_models.LoggingModel.objects.create(
                            username_slug_field=username,
                            ip_genericipaddress_field=req_inst.ip,
                            request_path_slug_field=req_inst.path,
                            request_method_slug_field=f"{req_inst.method} | {req_inst.action_type}",
                            error_text_field="-",
                        )
                        text = f"{datetime.datetime.now().strftime('%Y-%m-%dT%H:%M')}_" \
                               f"{user_model.password_char_field[-1]} {str(user_model.user_foreign_key_field)}" \
                               f"{user_model.password_char_field}"
                        encrypt_text = backend_service.EncryptingClass.encrypt_text(
                            text,
                            '31284'
                        )
                        subject = 'Восстановление пароля от веб платформы'
                        message_s = f'{user_model.first_name_char_field} {user_model.last_name_char_field}, ' \
                                    f'перейдите по ссылке: https://web.km.kz/ => войти => ' \
                                    f'восстановить доступ к аккаунту => введите Ваш ИИН и затем в окне восстановления' \
                                    f' через почту введите этот код восстановления (без кавычек): "{encrypt_text}". ' \
                                    f'Внимание! Этот код действует в течении часа с момента отправки!'
                        if subject and message_s and user_model.email_field:
                            send_mail(
                                subject,
                                message_s,
                                "kostanayminerals@web.km.kz",
                                [user_model.email_field, ''],
                                fail_silently=False
                            )
                        response = {
                            "response":
                                {"username": str(user_model.user_foreign_key_field.username),
                                 "secret_question_char_field": str(user_model.secret_question_char_field),
                                 "email_field": str(user_model.email_field),
                                 "success": False}
                        }
                    else:
                        response = {"error": "Внимание, отправлять письмо можно не чаще раза в 3 минуты!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "CHECK_EMAIL_PASSWORD":
                try:
                    # TODO get_value ###################################################################################
                    username = req_inst.get_value("username")
                    recover_password = req_inst.get_value("recoverPassword")
                    # TODO actions #####################################################################################
                    user_model = backend_models.UserModel.objects.get(
                        user_foreign_key_field=User.objects.get(username=username)
                    )
                    decrypt_text = backend_service.EncryptingClass.decrypt_text(
                        recover_password, user_model.password_char_field
                    )
                    text = f"{datetime.datetime.now().strftime('%Y-%m-%dT%H:%M')}_" \
                           f"{user_model.password_char_field[-1]}" \
                           f" {str(user_model.user_foreign_key_field)}{user_model.password_char_field} "
                    time1 = int(
                        decrypt_text.split('_')[0].split('T')[1].split(":")[0] +
                        decrypt_text.split('_')[0].split('T')[1].split(":")[1]
                    )
                    time2 = int(
                        text.split('_')[0].split('T')[1].split(":")[0] +
                        text.split('_')[0].split('T')[1].split(":")[1]
                    )
                    if time1 - time2 > -60:
                        if str(decrypt_text.split('_')[1]).strip() == str(text.split('_')[1]).strip():
                            response = {"response": {
                                "username": user_model.user_foreign_key_field.username,
                                "success": True
                            }}
                        else:
                            response = {"error": "Код не верный!"}
                    else:
                        response = {"error": "Код не верный или просрочен!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "CHANGE_PASSWORD":
                try:
                    # TODO get_value ###################################################################################
                    username = req_inst.get_value("username")
                    password = req_inst.get_value("password")
                    password2 = req_inst.get_value("password2")
                    # TODO actions #####################################################################################
                    user = User.objects.get(username=username)
                    user_model = backend_models.UserModel.objects.get(user_foreign_key_field=user)
                    if password == password2 and password != str(user_model.password_char_field).strip():
                        user.set_password(password)
                        user.save()
                        user_model.password_char_field = password
                        user_model.temp_password_boolean_field = False
                        user_model.save()
                        response = {"response": {
                            "username": user.username,
                            "success": False
                        }}
                    else:
                        response = {"error": "Пароли не совпадают или старый пароль идентичный!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            else:
                return Response({"error": "This action not allowed for this method."})
        else:
            return Response({"error": "This method not allowed for this endpoint."})
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
        return render(request, "backend/404.html")


@api_view(http_method_names=["GET", "POST", "PUT", "DELETE"])
@permission_classes([IsAuthenticated])
def api_auth_user(request):
    """
    auth user django-rest-framework
    """

    try:
        # TODO Request #################################################################################################
        req_inst = backend_service.DjangoClass.TemplateClass.request(request=request)
        # TODO Methods #################################################################################################
        if req_inst.method == "POST":
            # TODO Actions #############################################################################################
            if req_inst.action_type == "USER_DETAIL":
                try:
                    # TODO actions #####################################################################################
                    response = {"response": backend_serializers.UserSerializer(req_inst.user, many=False).data}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "CHANGE":
                try:
                    # TODO get_value ###################################################################################
                    password = req_inst.get_value("password")
                    password2 = req_inst.get_value("password2")
                    secret_question = req_inst.get_value("secretQuestion")
                    secret_answer = req_inst.get_value("secretAnswer")
                    email = req_inst.get_value("email")
                    # TODO actions #####################################################################################
                    user = User.objects.get(id=req_inst.user.id)
                    user_model = backend_models.UserModel.objects.get(user_foreign_key_field=user)
                    if len(password) < 1:
                        response = {"error": "Пароль пустой!"}
                    elif password != password2:
                        response = {"error": "Пароли не совпадают!"}
                    elif password == user_model.password_char_field:
                        response = {"error": "Пароль такой же как и предыдущий!"}
                    else:
                        user.set_password(password)
                        user.save()
                        user_model.password_char_field = password
                        user_model.temp_password_boolean_field = False
                        user_model.save()
                        if secret_question and secret_question != user_model.secret_question_char_field:
                            user_model.secret_question_char_field = secret_question
                            user_model.save()
                        if secret_answer and secret_answer != user_model.secret_answer_char_field:
                            user_model.secret_answer_char_field = secret_answer
                            user_model.save()
                        if email and email != user_model.email_field:
                            user_model.email_field = email
                            user_model.save()
                        response = {"response": "Изменение успешно проведено."}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "CHANGE_PASSWORD":
                try:
                    # TODO get_value ###################################################################################
                    username = req_inst.get_value("username")
                    password = req_inst.get_value("password")
                    password2 = req_inst.get_value("password2")
                    # TODO actions #####################################################################################
                    user = User.objects.get(username=username)
                    user_model = backend_models.UserModel.objects.get(user_foreign_key_field=user)
                    if password == password2 and password != str(user_model.password_char_field).strip():
                        user.set_password(password)
                        user.save()
                        user_model.password_char_field = password
                        user_model.temp_password_boolean_field = False
                        user_model.save()
                        response = {"response": {
                            "username": user.username,
                            "success": False
                        }}
                    else:
                        response = {"error": f"Пароли не совпадают или старый пароль идентичный!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "USER_LIST_ALL":
                try:
                    # TODO get_value ###################################################################################
                    user_models = backend_models.UserModel.objects.filter(
                        activity_boolean_field=True
                    ).order_by("last_name_char_field")
                    # TODO actions #####################################################################################
                    users = []
                    for user_model in user_models:
                        if user_model.user_foreign_key_field.is_superuser:
                            continue
                        users.append(f"{user_model.last_name_char_field} {user_model.first_name_char_field} "
                                     f"{user_model.personnel_number_slug_field} ")
                    response = {"response": users}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "NOTIFICATION_CREATE":
                try:
                    # TODO get_value ###################################################################################
                    name = req_inst.get_value("name")
                    place = req_inst.get_value("place")
                    # TODO actions #####################################################################################
                    if place == "банк идей":
                        model = backend_models.GroupModel.objects.get(name_slug_field="moderator_idea")
                    else:
                        model = backend_models.GroupModel.objects.get(name_slug_field="moderator")
                    description = req_inst.get_value("description")
                    backend_models.NotificationModel.objects.create(
                        author_foreign_key_field=req_inst.user_model,
                        model_foreign_key_field=model,
                        name_char_field=name,
                        place_char_field=place,
                        description_text_field=description,
                    )
                    response = {"response": "Успешно отправлено!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "NOTIFICATION_DELETE":
                try:
                    # TODO get_value ###################################################################################
                    _id = req_inst.get_value("id")
                    # TODO actions #####################################################################################
                    obj = backend_models.NotificationModel.objects.get(id=_id)
                    obj.visibility_boolean_field = False
                    obj.hide_datetime_field = datetime.datetime.now()
                    obj.save()
                    response = {"response": "Успешно удалено!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "NOTIFICATION_LIST":
                try:
                    # TODO actions #####################################################################################
                    objects = backend_models.NotificationModel.objects.filter(
                        model_foreign_key_field=None,
                        target_foreign_key_field=req_inst.user_model,
                        visibility_boolean_field=True,
                    )
                    serializer = backend_serializers.NotificationModelSerializer(objects, many=True).data
                    group_models = backend_models.GroupModel.objects.filter(user_many_to_many_field=req_inst.user_model)
                    objects = backend_models.NotificationModel.objects.filter(
                        model_foreign_key_field__in=group_models,
                        target_foreign_key_field=None,
                        visibility_boolean_field=True,
                    ).order_by("-hide_datetime_field")
                    serializer1 = backend_serializers.NotificationModelSerializer(objects, many=True).data
                    response = {"response": serializer + serializer1}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            else:
                return Response({"error": "This action not allowed for this method."})
        else:
            return Response({"error": "This method not allowed for this endpoint."})
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
        return render(request, "backend/404.html")


# TODO admin views #####################################################################################################
@api_view(http_method_names=["GET", "POST", "PUT", "DELETE"])
@permission_classes([IsAuthenticated])
def api_auth_admin(request):
    """
    auth admin django-rest-framework
    """

    try:
        # TODO Request #################################################################################################
        req_inst = backend_service.DjangoClass.TemplateClass.request(request=request)
        # TODO Methods #################################################################################################
        if req_inst.method == "POST":
            # TODO Actions #############################################################################################
            if req_inst.action_type == "CHECK_USER":
                try:
                    # TODO get_value ###################################################################################
                    username = req_inst.get_value("username")
                    # TODO actions #####################################################################################
                    user = User.objects.get(username=username)
                    response = {"response": {
                        "username": str(user.username),
                        "success": True
                    }}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "CHANGE_USER_PASSWORD":
                try:
                    # TODO get_value ###################################################################################
                    username = req_inst.get_value("username")
                    password = req_inst.get_value("password")
                    password2 = req_inst.get_value("password2")
                    # TODO actions #####################################################################################
                    user = User.objects.get(username=username)
                    user_model = backend_models.UserModel.objects.get(user_foreign_key_field=user)
                    if password == password2 and password != str(user_model.password_char_field).strip():
                        user.set_password(password)
                        user.save()
                        user_model.password_char_field = password
                        user_model.temp_password_boolean_field = True
                        user_model.secret_question_char_field = ""
                        user_model.secret_answer_char_field = ""
                        user_model.email_field = ""
                        user_model.save()
                        response = {"response": {
                            "username": str(user.username),
                            "success": False
                        }}
                    else:
                        response = {"error": f"Пароли не совпадают или старый пароль идентичный!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "CREATE_OR_CHANGE_USERS":
                try:
                    # TODO get_value ###################################################################################
                    additional_excel = req_inst.get_value("additionalExcel")
                    change_user = req_inst.get_value("changeUser")
                    change_user_password = req_inst.get_value("changeUserPassword")
                    clear_user_groups = req_inst.get_value("clearUserGroups")
                    # TODO actions #####################################################################################
                    if additional_excel:
                        def get_value(_col: Union[str, int], _row: Union[str, int], _sheet):
                            if isinstance(_col, int):
                                _col = get_column_letter(_col)
                            if isinstance(_row, str):
                                _row = str(_row)
                            _value = str(_sheet[str(_col).upper() + str(_row)].value).strip()
                            if _value.lower() == "none":
                                return ""
                            elif _value.lower() == "true":
                                return True
                            elif _value.lower() == "false":
                                return False
                            else:
                                return _value

                        workbook = openpyxl.load_workbook(additional_excel)
                        sheet = workbook.active
                        max_rows = sheet.max_row

                        print("start create users")

                        for row in range(1 + 1, max_rows + 1):
                            try:
                                subdivision_char_field = get_value(_col="A", _row=row, _sheet=sheet)
                                workshop_service_char_field = get_value(_col="B", _row=row, _sheet=sheet)
                                department_site_char_field = get_value(_col="C", _row=row, _sheet=sheet)
                                last_name_char_field = get_value(_col="D", _row=row, _sheet=sheet)
                                first_name_char_field = get_value(_col="E", _row=row, _sheet=sheet)
                                patronymic_char_field = get_value(_col="F", _row=row, _sheet=sheet)
                                personnel_number_slug_field = get_value(_col="G", _row=row, _sheet=sheet)
                                position_char_field = get_value(_col="H", _row=row, _sheet=sheet)
                                category_char_field = get_value(_col="I", _row=row, _sheet=sheet)
                                username = get_value(_col="J", _row=row, _sheet=sheet)
                                password_char_field = get_value(_col="K", _row=row, _sheet=sheet)
                                is_active = get_value(_col="L", _row=row, _sheet=sheet)
                                is_staff = get_value(_col="M", _row=row, _sheet=sheet)
                                is_superuser = get_value(_col="N", _row=row, _sheet=sheet)
                                is_temp_password = get_value(_col="O", _row=row, _sheet=sheet)
                                groups = get_value(_col="P", _row=row, _sheet=sheet).lower()
                                email_field = get_value(_col="Q", _row=row, _sheet=sheet)
                                secret_question_char_field = get_value(_col="R", _row=row, _sheet=sheet)
                                secret_answer_char_field = get_value(_col="S", _row=row, _sheet=sheet)

                                if len(username) <= 1:
                                    continue

                                try:
                                    user = User.objects.get(username=username)
                                    if user.is_superuser or change_user == "Не изменять уже существующего пользователя":
                                        continue
                                    new_user = False
                                except Exception as error:
                                    user = User.objects.create(
                                        username=username,
                                        password=make_password(password=password_char_field),
                                    )
                                    new_user = True

                                try:
                                    user_model = backend_models.UserModel.objects.get(
                                        user_foreign_key_field=user
                                    )
                                except Exception as error:
                                    user_model = backend_models.UserModel.objects.create(
                                        user_foreign_key_field=user
                                    )

                                if new_user:
                                    user_model.password_char_field = password_char_field
                                else:
                                    if change_user_password == "Изменять пароль уже существующего пользователя":
                                        user.password = make_password(password=password_char_field)
                                        user_model.password_char_field = password_char_field

                                user.is_staff = is_staff
                                user.is_superuser = is_superuser
                                user.email = email_field
                                user.last_name = last_name_char_field
                                user.first_name = first_name_char_field
                                user.save()

                                user_model.activity_boolean_field = is_active
                                user_model.email_field = email_field
                                user_model.secret_question_char_field = secret_question_char_field
                                user_model.secret_answer_char_field = secret_answer_char_field
                                user_model.temp_password_boolean_field = is_temp_password
                                user_model.last_name_char_field = last_name_char_field
                                user_model.first_name_char_field = first_name_char_field
                                user_model.patronymic_char_field = patronymic_char_field
                                user_model.personnel_number_slug_field = personnel_number_slug_field
                                user_model.subdivision_char_field = subdivision_char_field
                                user_model.workshop_service_char_field = workshop_service_char_field
                                user_model.department_site_char_field = department_site_char_field
                                user_model.position_char_field = position_char_field
                                user_model.category_char_field = category_char_field
                                user_model.save()

                                if clear_user_groups == "Добавлять новые группы доступа к предыдущим":
                                    for group in backend_models.GroupModel.objects.filter(
                                            user_many_to_many_field=user_model
                                    ):
                                        try:
                                            group.user_many_to_many_field.remove(user_model)
                                        except Exception as error:
                                            backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                                groups = [group.strip() for group in str(groups).lower().strip().split(',')]
                                for group in groups:
                                    if len(group) > 1:
                                        try:
                                            group_model = backend_models.GroupModel.objects.get_or_create(
                                                name_slug_field=group
                                            )[0]
                                            group_model.user_many_to_many_field.add(user_model)
                                        except Exception as error:
                                            backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                                if backend_service.DjangoClass.DefaultSettingsClass.get_actions_print_value():
                                    print(username)
                                    print(row)
                            except Exception as error:
                                backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                        response = {"response": "Пользователи успешно созданы/изменены."}
                    else:
                        response = {"error": "Ошибка чтения файла!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "EXPORT_USERS":
                try:
                    # TODO actions #####################################################################################
                    key = backend_service.UtilsClass.create_encrypted_password(
                        _random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890',
                        _length=24
                    )
                    date = backend_service.DateTimeUtils.get_current_date()
                    path = 'media/data/temp/users'
                    file_name = f"users_{key}_{date}.xlsx"
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active

                    # Delete old files
                    #######################################################
                    for root, dirs, files in os.walk(f"static/{path}", topdown=True):
                        for file in files:
                            try:
                                date_file = str(file).strip().split('.')[0].strip().split('_')[-1]
                                if date != date_file:
                                    os.remove(f'{path}/{file}')
                            except Exception as error:
                                backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    #######################################################

                    users = User.objects.filter(is_superuser=False)

                    def set_value(_col: Union[str, int], _row: Union[str, int], _value, _sheet):
                        if isinstance(_col, int):
                            _col = get_column_letter(_col)
                        if isinstance(_row, str):
                            _row = str(_row)
                        if isinstance(_value, bool):
                            if _value:
                                _value = "true"
                            else:
                                _value = "false"
                        if _value is None:
                            _value = ""
                        _sheet[str(_col).upper() + str(_row)] = str(_value)

                    titles = ['Подразделение', 'Цех/Служба', 'Отдел/Участок', 'Фамилия', 'Имя', 'Отчество',
                              'Табельный номер', 'Должность', 'Категория работника', 'Имя пользователя',
                              'Пароль аккаунта', 'Активность аккаунта', 'Доступ к панели управления',
                              'Права суперпользователя', 'Временный пароль', 'Группы доступа', 'Электронная почта',
                              'Секретный вопрос', 'Секретный ответ']
                    for title in titles:
                        set_value(_col=titles.index(title) + 1, _row=1, _value=title, _sheet=sheet)
                    _index = 1
                    for user in users:
                        try:
                            _index += 1
                            user_model = backend_models.UserModel.objects.get_or_create(user_foreign_key_field=user)[0]

                            subdivision_char_field = user_model.subdivision_char_field
                            set_value(_col="A", _row=_index, _value=subdivision_char_field, _sheet=sheet)

                            workshop_service_char_field = user_model.workshop_service_char_field
                            set_value(_col="B", _row=_index, _value=workshop_service_char_field,
                                      _sheet=sheet)

                            department_site_char_field = user_model.department_site_char_field
                            set_value(_col="C", _row=_index, _value=department_site_char_field,
                                      _sheet=sheet)

                            last_name_char_field = user_model.last_name_char_field
                            set_value(_col="D", _row=_index, _value=last_name_char_field, _sheet=sheet)

                            first_name_char_field = user_model.first_name_char_field
                            set_value(_col="E", _row=_index, _value=first_name_char_field, _sheet=sheet)

                            patronymic_char_field = user_model.patronymic_char_field
                            set_value(_col="F", _row=_index, _value=patronymic_char_field, _sheet=sheet)

                            personnel_number_slug_field = user_model.personnel_number_slug_field
                            set_value(_col="G", _row=_index, _value=personnel_number_slug_field,
                                      _sheet=sheet)

                            position_char_field = user_model.position_char_field
                            set_value(_col="H", _row=_index, _value=position_char_field, _sheet=sheet)

                            category_char_field = user_model.category_char_field
                            set_value(_col="I", _row=_index, _value=category_char_field, _sheet=sheet)

                            username = user.username
                            set_value(_col="J", _row=_index, _value=username, _sheet=sheet)

                            password_char_field = user_model.password_char_field
                            set_value(_col="K", _row=_index, _value=password_char_field, _sheet=sheet)

                            is_active = user_model.activity_boolean_field
                            set_value(_col="L", _row=_index, _value=is_active, _sheet=sheet)

                            is_staff = user.is_staff
                            set_value(_col="M", _row=_index, _value=is_staff, _sheet=sheet)

                            is_superuser = user.is_superuser
                            set_value(_col="N", _row=_index, _value=is_superuser, _sheet=sheet)

                            is_temp_password = user_model.temp_password_boolean_field
                            set_value(_col="O", _row=_index, _value=is_temp_password, _sheet=sheet)

                            group_models = backend_models.GroupModel.objects.filter(user_many_to_many_field=user_model)
                            groups = ""
                            for group_model in group_models:
                                groups += f"{str(group_model.name_slug_field).lower().strip()}, "
                            groups = groups[:-2]
                            set_value(_col="P", _row=_index, _value=groups, _sheet=sheet)

                            email_field = user_model.email_field
                            set_value(_col="Q", _row=_index, _value=email_field, _sheet=sheet)

                            secret_question_char_field = user_model.secret_question_char_field
                            set_value(_col="R", _row=_index, _value=secret_question_char_field,
                                      _sheet=sheet)

                            secret_answer_char_field = user_model.secret_answer_char_field
                            set_value(_col="S", _row=_index, _value=secret_answer_char_field, _sheet=sheet)
                        except Exception as error:
                            backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    # Set font
                    #######################################################
                    font_b = Font(name='Arial', size=8, bold=False)
                    for row in range(1, _index + 1):
                        for col in [get_column_letter(x) for x in range(1, 30 + 1)]:
                            cell = sheet[f'{col}{row}']
                            cell.font = font_b
                    #######################################################

                    # Height and width styles
                    #######################################################
                    for col in [get_column_letter(x) for x in range(1, 30 + 1)]:
                        width = 1
                        for row in range(1, _index + 1):
                            cell = sheet[f'{col}{row}']
                            value = len(str(cell.value))
                            if value > width:
                                width = value
                        sheet.column_dimensions[col].height = 1
                        sheet.column_dimensions[col].width = round((width * 0.95), 3)
                    #######################################################

                    try:
                        backend_service.ExcelClass.workbook_save(
                            workbook=workbook,
                            excel_file=f"static/{path}/{file_name}"
                        )
                    except Exception as error:
                        pass
                    response = {"response": {"excel": f"static/{path}/{file_name}"}}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "ACTIVITY_USER":
                try:
                    # TODO get_value ###################################################################################
                    username = req_inst.get_value("username")
                    active = req_inst.get_value("active")
                    # TODO actions #####################################################################################
                    user_model = backend_models.UserModel.objects.get(
                        user_foreign_key_field=User.objects.get(username=username)
                    )
                    if active:
                        user_model.activity_boolean_field = True
                        user_model.save()
                    else:
                        user_model.activity_boolean_field = False
                        user_model.save()
                    response = {"response": "Успешно изменено!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "TERMINAL_REBOOT":
                try:
                    # TODO get_value ###################################################################################
                    ips = req_inst.get_value("ips")

                    # TODO actions #####################################################################################

                    def reboot(_ip):
                        url = "htt" + f"p://{ip}/ISAPI/System/reboot"
                        h = httplib2.Http(
                            os.path.dirname(os.path.abspath('__file__')) + "/static/media/data/temp/reboot_terminal"
                        )
                        login_ = 'admin'
                        password_ = 'snrg2017'
                        h.add_credentials(login_, password_)
                        headers = {
                            'Content-type': 'text/plain;charset=UTF-8',
                            'Accept-Encoding': 'gzip, deflate',
                            'Accept-Language': 'de,en-US;q=0.7,en;q=0.3',
                            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:75.0) Gecko/20100101 '
                                          'Firefox/75.0',
                        }
                        response_, content = h.request(uri=url, method="PUT", headers=headers)
                        print(content)

                    for ip in [str(str(x).strip()) for x in ips.split(",")]:
                        if len(str(ip)) < 3:
                            continue
                        with ThreadPoolExecutor() as executor:
                            executor.submit(reboot, ip)
                    response = {"response": "Успешно перезагружено!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            else:
                return Response({"error": "This action not allowed for this method."})
        else:
            return Response({"error": "This method not allowed for this endpoint."})
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
        return render(request, "backend/404.html")


@api_view(http_method_names=["GET", "POST", "PUT", "DELETE"])
@authentication_classes([BasicAuthentication])
def api_basic_admin_user_temp(request):
    """
    basic admin django-rest-framework
    """

    try:
        # TODO Request #################################################################################################
        req_inst = backend_service.DjangoClass.TemplateClass.request(request=request)
        # TODO Methods #################################################################################################
        if req_inst.method == "GET":
            # TODO Actions #############################################################################################
            try:
                # TODO actions #########################################################################################
                user_models = backend_models.UserModel.objects.filter(
                    temp_password_boolean_field=True,
                    activity_boolean_field=True
                )
                if not request.user.is_superuser:
                    response = {"error": "Your user in not superuser."}
                else:
                    objects = []
                    for user_model in user_models:
                        if not user_model.user_foreign_key_field.is_superuser:
                            objects.append(
                                {
                                    f"""{base64.b64encode(
                                        str(user_model.user_foreign_key_field.username)[::-1].encode()
                                    ).decode()}""":
                                        base64.b64encode(
                                            str(f"12{user_model.password_char_field}345").encode()).decode()
                                }
                            )
                    # for obj in objects:
                    #     for key, value in obj.items():
                    #         print(f"{key}: {str(base64.b64decode(value).decode())[2: -3]}")
                    response = {"response": objects}
                # TODO response ########################################################################################
                backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                return Response(response)
            except Exception as error:
                backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                return Response({"error": "Произошла ошибка!"})
        else:
            return Response({"error": "Этот метод не реализован для этой точки доступа."})
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
        return render(request, "backend/404.html")


# TODO custom views ####################################################################################################
@api_view(http_method_names=["GET", "POST", "PUT", "DELETE"])
@permission_classes([IsAuthenticated])
def api_auth_salary(request):
    """
    api_salary django-rest-framework
    """

    try:
        # TODO Request #################################################################################################
        req_inst = backend_service.DjangoClass.TemplateClass.request(request=request)
        # TODO Methods #################################################################################################
        if req_inst.method == "POST":
            # TODO Actions #############################################################################################
            if req_inst.action_type == "USER_SALARY":
                try:
                    # TODO get_value ###################################################################################
                    date_time = req_inst.get_value("dateTime")
                    # TODO actions #####################################################################################
                    # Get json response from 1c
                    key = backend_service.UtilsClass.create_encrypted_password(
                        _random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890',
                        _length=10
                    )
                    hash_key_obj = hashlib.sha256()
                    hash_key_obj.update(key.encode('utf-8'))
                    key_hash = str(hash_key_obj.hexdigest().strip().upper())
                    key_hash_base64 = base64.b64encode(str(key_hash).encode()).decode()
                    iin = req_inst.user.username
                    if str(iin).lower() == '000000000000':
                        iin = 970801351179
                    iin_base64 = base64.b64encode(str(iin).encode()).decode()
                    date_base64 = base64.b64encode(f'{date_time}'.encode()).decode()
                    url = f'http://192.168.1.10/KM_1C/hs/zp/rl/{iin_base64}_{key_hash_base64}/{date_base64}'
                    h = httplib2.Http(
                        os.path.dirname(os.path.abspath('__file__')) + "/static/media/data/temp/get_salary_data"
                    )
                    _login = 'Web_adm_1c'
                    _password = '159159qqww!'
                    h.add_credentials(_login, _password)
                    response, content = h.request(url)

                    print("content: ", content)
                    print("response: ", response)

                    data = backend_service.UtilsClass.decrypt_text_with_hash(content.decode()[1:], key_hash)
                    error_word_list = ['ошибка', 'error', 'failed']
                    if data.find('send') == 0:
                        return Response({"error": f"{data.split('send')[1].strip()}"})
                    else:
                        for error_word in error_word_list:
                            if data.find(error_word.lower()) >= 0:
                                return Response({"error": f"Произошла ошибка!"})

                        # Get local test json response from 1c
                        # Временное чтение файла для отладки без доступа к 1С
                        # with open("static/media/data/json_data.json", "r", encoding="utf-8") as file:
                        #     data = json.load(file)
                        #     time.sleep(3)

                        json_data = json.loads(data)

                        try:
                            json_data["global_objects"]["3.Доходы в натуральной форме"]
                        except Exception as error:
                            json_data["global_objects"]["3.Доходы в натуральной форме"] = {
                                "Fields": {
                                    "1": "Вид",
                                    "2": "Период",
                                    "3": "Сумма"
                                },
                            }

                        # Return pretty integer and float value
                        def return_float_value(_value):
                            if isinstance(_value, int) or isinstance(_value, float):
                                if len(f'{_value:.2f}') < 10:
                                    _value = f'{_value:.2f}'[:-6] + ' ' + f'{_value:.2f}'[-6:]
                                else:
                                    _value = f'{_value:.2f}'[:-9] + ' ' + f'{_value:.2f}'[-9:-6] + ' ' + \
                                             f'{_value:.2f}'[-6:]
                            return _value

                        # Create 'Ends' and pretty integer and float value
                        def create_ends(table: str, extracols=False):
                            if extracols:
                                _days = 0
                                _hours = 0
                                _summ = 0
                                for __key in json_data['global_objects'][table].keys():
                                    if __key != 'Fields':
                                        try:
                                            _days += json_data['global_objects'][table][f'{__key}']['ВсегоДни']
                                            _hours += json_data['global_objects'][table][f'{__key}']['ВсегоЧасы']
                                            _summ_local = json_data['global_objects'][table][f'{__key}']['Сумма']
                                            json_data['global_objects'][table][f'{__key}']['Сумма'] = return_float_value(
                                                _summ_local)
                                            _summ += _summ_local
                                        except Exception as error_:
                                            backend_service.DjangoClass.LoggingClass.error(request=request, error=error_)
                                json_data['global_objects'][table]['Ends'] = {
                                    'Вид': 'Итого', 'Период': '', 'Дни': _days, 'Часы': _hours,
                                    'ВсегоДни': 0, 'ВсегоЧасы': 0, 'Сумма': return_float_value(_summ)
                                }
                            else:
                                _summ = 0
                                for __key in json_data['global_objects'][table].keys():
                                    if __key != 'Fields':
                                        try:
                                            _summ_local = json_data['global_objects'][table][f'{__key}']['Сумма']
                                            json_data['global_objects'][table][f'{__key}']['Сумма'] = return_float_value(
                                                _summ_local)
                                            _summ += _summ_local
                                        except Exception as error_:
                                            backend_service.DjangoClass.LoggingClass.error(request=request, error=error_)
                                json_data['global_objects'][table]['Ends'] = {
                                    'Вид': 'Итого', 'Период': '', 'Сумма': return_float_value(_summ)
                                }

                        create_ends(table='1.Начислено', extracols=True)
                        create_ends(table='2.Удержано', extracols=False)
                        create_ends(table='3.Доходы в натуральной форме', extracols=False)
                        create_ends(table='4.Выплачено', extracols=False)
                        create_ends(table='5.Налоговые вычеты', extracols=False)

                        # pretty integer and float value in headers
                        for _key in json_data.keys():
                            if _key != 'global_objects':
                                json_data[_key] = return_float_value(json_data[_key])

                        # create excel
                        key = backend_service.UtilsClass.create_encrypted_password(
                            _random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890',
                            _length=24
                        )
                        date = backend_service.DateTimeUtils.get_current_date()
                        path = 'media/data/temp/salary'
                        file_name = f"salary_{key}_{date}.xlsx"
                        workbook = backend_service.ExcelClass.workbook_create()
                        sheet = backend_service.ExcelClass.workbook_activate(workbook)

                        # Delete old files
                        for root, dirs, files in os.walk(f"static/{path}", topdown=True):
                            for file in files:
                                try:
                                    date_file = str(file).strip().split('.')[0].strip().split('_')[-1]
                                    if date != date_file:
                                        os.remove(f'{path}/{file}')
                                except Exception as error:
                                    pass

                        # Create 'TitleComponent'
                        backend_service.ExcelClass.set_sheet_value(
                            col=1,
                            row=1,
                            value='РАСЧЕТНЫЙ ЛИСТ',
                            sheet=sheet
                        )
                        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
                        #######################################################

                        # Create 'Headers'
                        header_arr = []
                        for key, value in json_data.items():
                            if key != 'global_objects' and key != 'Долг за организацией на начало месяца' \
                                    and key != 'Долг за организацией на конец месяца':
                                header_arr.append(f'{key}: {value}')

                        header_len_devide = len(header_arr) - 8
                        if header_len_devide % 2 != 0:
                            header_len_devide += 1

                        row_i_1 = 1 + 1
                        for header in header_arr[0:4]:
                            col_i = 1
                            backend_service.ExcelClass.set_sheet_value(
                                col=col_i,
                                row=row_i_1,
                                value=header,
                                sheet=sheet
                            )
                            sheet.merge_cells(start_row=row_i_1, start_column=col_i, end_row=row_i_1,
                                              end_column=col_i + 4)
                            row_i_1 += 1

                        row_i_2 = 1 + 1
                        for header in header_arr[4:8]:
                            col_i = 6
                            backend_service.ExcelClass.set_sheet_value(
                                col=col_i,
                                row=row_i_2,
                                value=header,
                                sheet=sheet
                            )
                            sheet.merge_cells(start_row=row_i_2, start_column=col_i, end_row=row_i_2,
                                              end_column=col_i + 2)
                            row_i_2 += 1

                        row_i_3 = row_i_1
                        for header in header_arr[8:8 + header_len_devide // 2]:
                            col_i = 1
                            backend_service.ExcelClass.set_sheet_value(
                                col=col_i,
                                row=row_i_3,
                                value=header,
                                sheet=sheet
                            )
                            sheet.merge_cells(start_row=row_i_3, start_column=col_i, end_row=row_i_3,
                                              end_column=col_i + 4)
                            row_i_3 += 1

                        row_i_4 = row_i_2
                        for header in header_arr[8 + header_len_devide // 2:]:
                            col_i = 6
                            backend_service.ExcelClass.set_sheet_value(
                                col=col_i,
                                row=row_i_4,
                                value=header,
                                sheet=sheet
                            )
                            sheet.merge_cells(start_row=row_i_4, start_column=col_i, end_row=row_i_4,
                                              end_column=col_i + 2)
                            row_i_4 += 1
                        header_low_row = row_i_4

                        #######################################################

                        # Create 'Bodyes'
                        def create_bodyes(table: str, extracols=False):
                            bodyes_arr = [[table]]
                            for __key, _value in json_data['global_objects'][table].items():
                                local_bodyes_arr = []
                                for ___key, __value in json_data['global_objects'][table][__key].items():
                                    if extracols:
                                        if ___key != '6' and ___key != '7' and \
                                                ___key != 'ВсегоДни' and ___key != 'ВсегоЧасы':
                                            local_bodyes_arr.append(__value)
                                    else:
                                        local_bodyes_arr.append(__value)
                                bodyes_arr.append(local_bodyes_arr)
                            return bodyes_arr

                        bodyes_arr_1 = create_bodyes('1.Начислено', extracols=True)
                        bodyes_arr_2 = create_bodyes('2.Удержано', extracols=False)
                        bodyes_arr_3 = create_bodyes('3.Доходы в натуральной форме', extracols=False)
                        bodyes_arr_4 = create_bodyes('4.Выплачено', extracols=False)
                        bodyes_arr_5 = create_bodyes('5.Налоговые вычеты', extracols=False)

                        bold_arr = []

                        body_low_row_1 = header_low_row
                        bold_arr.append(body_low_row_1)
                        sheet.merge_cells(start_row=body_low_row_1, start_column=1, end_row=body_low_row_1,
                                          end_column=1 + 4)
                        for body in bodyes_arr_1:
                            col_i = 1
                            for value in body:
                                if isinstance(value, int) and value == 0:
                                    value = ''
                                backend_service.ExcelClass.set_sheet_value(
                                    col=col_i,
                                    row=body_low_row_1,
                                    value=value,
                                    sheet=sheet
                                )
                                col_i += 1
                            body_low_row_1 += 1

                        body_low_row_2 = header_low_row
                        bold_arr.append(body_low_row_2)
                        sheet.merge_cells(start_row=body_low_row_2, start_column=6, end_row=body_low_row_2,
                                          end_column=6 + 2)
                        for body in bodyes_arr_2:
                            col_i = 6
                            for value in body:
                                backend_service.ExcelClass.set_sheet_value(
                                    col=col_i,
                                    row=body_low_row_2,
                                    value=value,
                                    sheet=sheet
                                )
                                col_i += 1
                            body_low_row_2 += 1

                        if body_low_row_1 >= body_low_row_2:
                            body_low_row_3 = body_low_row_1
                            body_low_row_4 = body_low_row_1
                        else:
                            body_low_row_3 = body_low_row_2
                            body_low_row_4 = body_low_row_2

                        bold_arr.append(body_low_row_3)
                        sheet.merge_cells(start_row=body_low_row_3, start_column=1, end_row=body_low_row_3,
                                          end_column=1 + 4)
                        for body in bodyes_arr_3:
                            col_i = 1
                            for value in body:
                                backend_service.ExcelClass.set_sheet_value(
                                    col=col_i,
                                    row=body_low_row_3,
                                    value=value,
                                    sheet=sheet
                                )
                                col_i += 1
                            body_low_row_3 += 1

                        bold_arr.append(body_low_row_4)
                        sheet.merge_cells(start_row=body_low_row_4, start_column=6, end_row=body_low_row_4,
                                          end_column=6 + 2)
                        for body in bodyes_arr_4:
                            col_i = 6
                            for value in body:
                                backend_service.ExcelClass.set_sheet_value(
                                    col=col_i,
                                    row=body_low_row_4,
                                    value=value,
                                    sheet=sheet
                                )
                                col_i += 1
                            body_low_row_4 += 1

                        if body_low_row_3 >= body_low_row_4:
                            body_low_row_5 = body_low_row_3
                            body_low_row_6 = body_low_row_3
                        else:
                            body_low_row_5 = body_low_row_4
                            body_low_row_6 = body_low_row_4

                        bold_arr.append(body_low_row_5)
                        sheet.merge_cells(start_row=body_low_row_5, start_column=1, end_row=body_low_row_5,
                                          end_column=1 + 4)
                        for body in bodyes_arr_5:
                            col_i = 1
                            for value in body:
                                backend_service.ExcelClass.set_sheet_value(
                                    col=col_i,
                                    row=body_low_row_5,
                                    value=value,
                                    sheet=sheet
                                )
                                col_i += 1
                            body_low_row_5 += 1

                        lowest = [
                            f'Долг за организацией на начало месяца: '
                            f'{json_data["Долг за организацией на начало месяца"]}',
                            f'Долг за организацией на конец месяца: '
                            f'{json_data["Долг за организацией на конец месяца"]}',
                        ]

                        bold_arr.append(body_low_row_6)
                        for body in ['.', 'Вид', *lowest]:
                            col_i = 6
                            backend_service.ExcelClass.set_sheet_value(
                                col=col_i,
                                row=body_low_row_6,
                                value=body,
                                sheet=sheet
                            )
                            sheet.merge_cells(start_row=body_low_row_6, start_column=col_i, end_row=body_low_row_6,
                                              end_column=col_i + 2)
                            body_low_row_6 += 1

                        if body_low_row_6 >= body_low_row_5:
                            body_low_row = body_low_row_6
                        else:
                            body_low_row = body_low_row_5

                        if body_low_row_1 >= body_low_row_2:
                            body_color_2 = body_low_row_1
                        else:
                            body_color_2 = body_low_row_2
                        if body_low_row_3 >= body_low_row_4:
                            body_color_3 = body_low_row_3
                        else:
                            body_color_3 = body_low_row_4
                        #######################################################

                        # Set fonts
                        #######################################################
                        font_headers = Font(name='Arial', size=8, bold=False)
                        for row in range(1, header_low_row):
                            for col in [backend_service.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                cell = sheet[f'{col}{row}']
                                cell.font = font_headers

                        font_bodyes = Font(name='Arial', size=7, bold=False)
                        for row in range(header_low_row, body_low_row + 1):
                            for col in [backend_service.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                cell = sheet[f'{col}{row}']
                                cell.font = font_bodyes

                        font_tables = Font(name='Arial', size=8, bold=True)
                        for row in [header_low_row, body_color_2, body_color_3]:
                            for col in [backend_service.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                cell = sheet[f'{col}{row}']
                                cell.font = font_tables

                        # Set aligments
                        # wrap_text = Alignment(wrap_text=True)
                        # shrink_to_fit = Alignment(shrink_to_fit=True)
                        aligment_center = Alignment(horizontal='center', vertical='center', wrap_text=True,
                                                    shrink_to_fit=True)
                        for col in [backend_service.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                            cell = sheet[f'{col}{1}']
                            cell.alignment = aligment_center
                        aligment_left = Alignment(horizontal='left', vertical='center', wrap_text=True,
                                                  shrink_to_fit=True)
                        for row in range(2, header_low_row):
                            for col in [backend_service.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                cell = sheet[f'{col}{row}']
                                cell.alignment = aligment_left

                        aligment_right = Alignment(horizontal='right', vertical='center', wrap_text=True,
                                                   shrink_to_fit=True)
                        for row in range(header_low_row, body_low_row + 1):
                            for col in [backend_service.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                cell = sheet[f'{col}{row}']
                                if col == 'A' or col == 'F':
                                    cell.alignment = aligment_left
                                elif col == 'E' or col == 'H':
                                    cell.alignment = aligment_right
                                else:
                                    cell.alignment = aligment_center

                        # Set borders
                        side_medium = Side(border_style="thin", color="FF808080")
                        # side_think = Side(border_style="thin", color="FF808080")
                        border_horizontal_middle = Border(
                            top=side_medium,
                            # left=side_medium,
                            # right=side_medium,
                            # bottom=side_medium
                        )
                        border_vertical_middle = Border(
                            # top=side_think,
                            left=side_medium,
                            # right=side_think,
                            # bottom=side_think
                        )
                        border_vertical_light = Border(
                            # top=side_think,
                            left=side_medium,
                            # right=side_think,
                            # bottom=side_think
                        )
                        for row in range(header_low_row, body_low_row):
                            for col in [backend_service.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                if col == 'G' and row > body_low_row_4 or col == 'H' and row > body_low_row_4:
                                    pass
                                else:
                                    cell = sheet[f'{col}{row}']
                                    cell.border = border_vertical_light
                            cell = sheet[f'{backend_service.ExcelClass.get_column_letter(1)}{row}']
                            cell.border = border_vertical_middle
                            cell = sheet[f'{backend_service.ExcelClass.get_column_letter(6)}{row}']
                            cell.border = border_vertical_middle
                            cell = sheet[f'{backend_service.ExcelClass.get_column_letter(9)}{row}']
                            cell.border = border_vertical_middle
                        side_think = Side(border_style="dotted", color="FF808080")
                        # {'mediumDashDotDot', 'thin', 'dashed', 'mediumDashed', 'dotted', 'double', 'thick',
                        # 'medium', 'dashDot','dashDotDot', 'hair', 'mediumDashDot', 'slantDashDot'}
                        border_think = Border(
                            top=side_think,
                            left=side_think,
                            right=side_think,
                            bottom=side_think
                        )
                        for row in range(1, header_low_row):
                            for col in [backend_service.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                cell = sheet[f'{col}{row}']
                                cell.border = border_think
                        side_medium = Side(border_style="thin", color="FF808080")
                        border_medium = Border(
                            top=side_medium,
                            left=side_medium,
                            right=side_medium,
                            bottom=side_medium
                        )
                        for col in [backend_service.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                            for row in [header_low_row + 1, body_color_2 + 1, body_color_3 + 1]:
                                cell = sheet[f'{col}{row - 1}']
                                cell.border = border_medium
                        for col in [backend_service.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                            cell = sheet[f'{col}{body_low_row}']
                            cell.border = border_horizontal_middle
                        #######################################################

                        # Colored styles
                        #######################################################
                        color_green = PatternFill(fgColor="E6E6FF", fill_type="solid")
                        for col in [backend_service.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                            for row in [header_low_row + 1, body_color_2 + 1, body_color_3 + 1]:
                                cell = sheet[f'{col}{row}']
                                cell.fill = color_green
                                cell = sheet[f'{col}{row - 1}']
                                cell.border = border_medium
                        color_yellow = PatternFill(fgColor="d0ffd8", fill_type="solid")
                        for col in [backend_service.ExcelClass.get_column_letter(x) for x in range(1, 5 + 1)]:
                            cell = sheet[f'{col}{body_low_row_1 - 1}']
                            cell.fill = color_yellow

                        for col in [backend_service.ExcelClass.get_column_letter(x) for x in range(6, 8 + 1)]:
                            cell = sheet[f'{col}{body_low_row_2 - 1}']
                            cell.fill = color_yellow

                        for col in [backend_service.ExcelClass.get_column_letter(x) for x in range(1, 3 + 1)]:
                            cell = sheet[f'{col}{body_low_row_3 - 1}']
                            cell.fill = color_yellow
                            cell.fill = color_yellow

                        for col in [backend_service.ExcelClass.get_column_letter(x) for x in range(6, 8 + 1)]:
                            cell = sheet[f'{col}{body_low_row_4 - 1}']
                            cell.fill = color_yellow

                        for col in [backend_service.ExcelClass.get_column_letter(x) for x in range(1, 3 + 1)]:
                            cell = sheet[f'{col}{body_low_row_5 - 1}']
                            cell.fill = color_yellow
                            cell.fill = color_yellow

                        for col in [backend_service.ExcelClass.get_column_letter(x) for x in range(6, 8 + 1)]:
                            cell = sheet[f'{col}{body_low_row_6 - 1}']
                            cell.fill = color_yellow

                        #######################################################

                        # Height and width styles
                        #######################################################
                        for col in [backend_service.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                            width = 1
                            for row in range(1, body_low_row + 1):
                                cell = sheet[f'{col}{row}']
                                value = len(str(cell.value))
                                if value > width:
                                    width = value
                            if col == 'A' or col == 'F':
                                width = width / 2
                            sheet.column_dimensions[col].height = 1
                            sheet.column_dimensions[col].width = round((width * 0.95), 3)
                        #######################################################

                        # Set global page and book settings
                        #######################################################
                        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
                        sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
                        sheet.page_margins.left = 0.05
                        sheet.page_margins.right = 0.05
                        sheet.page_margins.header = 0.1
                        sheet.page_margins.bottom = 0.2
                        sheet.page_margins.footer = 0.2
                        sheet.page_margins.top = 0.1
                        sheet.print_options.horizontalCentered = True
                        # sheet.print_options.verticalCentered = True
                        sheet.page_setup.fitToHeight = 1
                        sheet.page_setup.scale = 80
                        sheet.page_setup.fitToHeight = 1
                        sheet.page_setup.fitToWidth = 1
                        sheet.protection.password = key + '_1'
                        sheet.protection.sheet = True
                        sheet.protection.enable()
                        #######################################################
                        try:
                            backend_service.ExcelClass.workbook_save(
                                workbook=workbook,
                                excel_file=f"static/{path}/{file_name}"
                            )
                        except Exception as error:
                            pass
                        json_data["excel_path"] = f"static/{path}/{file_name}"

                        response = {"response": json_data}
                    # TODO response ###################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            else:
                return Response({"error": "This action not allowed for this method."})
        else:
            return Response({"error": "This method not allowed for this endpoint."})
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
        return render(request, "backend/404.html")


@api_view(http_method_names=["GET", "POST", "PUT", "DELETE"])
@permission_classes([IsAuthenticated])
def api_auth_vacation(request):
    """
    api_salary django-rest-framework
    """

    try:
        # TODO Request #################################################################################################
        req_inst = backend_service.DjangoClass.TemplateClass.request(request=request)
        # TODO Methods #################################################################################################
        if req_inst.method == "POST":
            # TODO Actions #############################################################################################
            if req_inst.action_type == "USER_VACATION":
                try:
                    # TODO get_value ###################################################################################
                    date_time = req_inst.get_value("dateTime")
                    # TODO actions #####################################################################################
                    # Get json response from 1c
                    key = backend_service.UtilsClass.create_encrypted_password(
                        _random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890',
                        _length=10
                    )
                    hash_key_obj = hashlib.sha256()
                    hash_key_obj.update(key.encode('utf-8'))
                    key_hash = str(hash_key_obj.hexdigest().strip().upper())
                    key_hash_base64 = base64.b64encode(str(key_hash).encode()).decode()
                    iin = req_inst.user.username
                    if str(iin).lower() == '000000000000' or str(iin).lower() == 'bogdan':
                        iin = 970801351179
                    iin_base64 = base64.b64encode(str(iin).encode()).decode()
                    date_base64 = base64.b64encode(f'{date_time}'.encode()).decode()
                    # url = f'http://192.168.1.158/Tanya_perenos/hs/pers/u_vacation/{iin_base64}_
                    # {key_hash_base64}/{date_base64}'
                    url = f'http://192.168.1.10/KM_1C/hs/pers/u_vacation/{iin_base64}_{key_hash_base64}/{date_base64}'
                    h = httplib2.Http(
                        os.path.dirname(os.path.abspath('__file__')) + "/static/media/data/temp/get_vacation_data"
                    )
                    _login = 'Web_adm_1c'
                    _password = '159159qqww!'
                    h.add_credentials(_login, _password)
                    response, content = h.request(url)

                    data = backend_service.UtilsClass.decrypt_text_with_hash(content.decode()[1:], key_hash)

                    print("data: ", data)

                    error_word_list = ['ошибка', 'error', 'failed']
                    if data.find('send') == 0:
                        return Response({"error": f"{data.split('send')[1].strip()}"})
                    json_data = json.loads(data)
                    for error_word in error_word_list:
                        if data.find(error_word.lower()) >= 0:
                            return Response({"error": f"Произошла ошибка!"})

                    headers = []
                    tables = []
                    for key, value in json_data.items():
                        try:
                            if key != "global_objects":
                                try:
                                    if key == "Работник":
                                        headers.append(["Работник", value])
                                    elif key == "Подразделение":
                                        headers.append(["Место работы", value])
                                    elif key == "Должность":
                                        headers.append(["Должность", value])
                                    elif key == "На дату":
                                        headers.append(["Дни отпуска на дату - ", value])
                                    elif key == "Фактически заработанные дни":
                                        headers.append([
                                            "Количество дней неиспользованного отпуска на выбранную дату",
                                            value
                                        ])
                                    else:
                                        headers.append([key, value])
                                except Exception as error:
                                    pass
                            else:
                                for key_, value_ in json_data["global_objects"].items():
                                    tab = []
                                    for key__, value__ in json_data["global_objects"][key_].items():
                                        try:
                                            if key__ == "По графику ":
                                                tab.append(["По графику: ", value__])
                                            elif key__ == "Период":
                                                tab.append(["Дата начала", value__[0:11]])
                                            elif key__ == "ДатаОкончания":
                                                tab.append(["Дата окончания", value__[0:11]])
                                            elif key__ == "ДнейОтпуска":
                                                tab.append(["Запланированные дни отпуска", value__])
                                            elif key__ == "Рабочий год":
                                                tab.append(["Год", value__])
                                            else:
                                                tab.append([key__, value__])
                                        except Exception as error:
                                            pass
                                    tables.append(tab)
                        except Exception as error:
                            pass
                    print("headers: ", headers)
                    print("tables: ", tables)

                    json_data = {
                        "headers": headers,
                        "tables": tables,
                    }

                    response = {"response": json_data}

                    print("response: ", response)

                    # TODO response ###################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            else:
                return Response({"error": "This action not allowed for this method."})
        else:
            return Response({"error": "This method not allowed for this endpoint."})
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
        return render(request, "backend/404.html")


@api_view(http_method_names=["GET", "POST", "PUT", "DELETE"])
@permission_classes([IsAuthenticated])
def api_auth_idea(request):
    """
    api_rational django-rest-framework
    """

    try:
        # TODO Request #################################################################################################
        req_inst = backend_service.DjangoClass.TemplateClass.request(request=request)
        # TODO Methods #################################################################################################
        if req_inst.method == "POST":
            # TODO Actions #############################################################################################
            if req_inst.action_type == "IDEA_CREATE":
                try:
                    # TODO get_value ###################################################################################
                    subdivision = req_inst.get_value("subdivision")
                    sphere = req_inst.get_value("sphere")
                    category = req_inst.get_value("category")
                    avatar = req_inst.get_value("avatar")
                    name = req_inst.get_value("name")
                    place = req_inst.get_value("place")
                    description = req_inst.get_value("description")
                    moderate = req_inst.get_value("moderate")
                    # TODO actions #####################################################################################
                    backend_models.IdeaModel.objects.create(
                        author_foreign_key_field=req_inst.user_model,
                        subdivision_char_field=subdivision,
                        sphere_char_field=sphere,
                        category_char_field=category,
                        image_field=avatar,
                        name_char_field=name,
                        place_char_field=place,
                        description_text_field=description,
                        status_moderate_char_field=moderate,
                    )
                    backend_models.NotificationModel.objects.create(
                        author_foreign_key_field=req_inst.user_model,
                        model_foreign_key_field=backend_models.GroupModel.objects.get(
                            name_slug_field="moderator_idea"
                        ),
                        name_char_field=f"Создана новая идея",
                        place_char_field="банк идей",
                        description_text_field=f"название идеи: {name}",
                    )
                    response = {"response": "Успешно отправлено на модерацию!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "IDEA_LIST":
                try:
                    # TODO get_value ###################################################################################
                    moderate = req_inst.get_value("moderate")
                    subdivision = req_inst.get_value("subdivision")
                    sphere = req_inst.get_value("sphere")
                    category = req_inst.get_value("category")
                    author = req_inst.get_value("author")
                    search = req_inst.get_value("search")
                    sort = req_inst.get_value("sort")

                    only_month = req_inst.get_value("onlyMonth")
                    # TODO actions #####################################################################################
                    # TODO onlyMonth
                    ideas = backend_models.IdeaModel.objects.all().order_by("-register_datetime_field")
                    if only_month:
                        now = (datetime.datetime.now()).strftime('%Y-%m-%d %H:%M')
                        local_objects = []
                        for idea in ideas:
                            if (idea.register_datetime_field + datetime.timedelta(days=31)).strftime('%Y-%m-%d %H:%M') \
                                    >= now:
                                local_objects.append(idea.id)
                        ideas = backend_models.IdeaModel.objects.filter(id__in=local_objects). \
                            order_by("-register_datetime_field")
                    # TODO search
                    if search:
                        ideas = ideas.filter(name_char_field__icontains=search)
                    # TODO filter
                    if subdivision:
                        ideas = ideas.filter(subdivision_char_field=subdivision)
                    if sphere:
                        ideas = ideas.filter(sphere_char_field=sphere)
                    if category:
                        ideas = ideas.filter(category_char_field=category)
                    if author:
                        if author == "self":
                            author = req_inst.user_model
                        else:
                            author = backend_models.UserModel.objects.get(
                                personnel_number_slug_field=(str(author).split(" ")[-2]).strip()
                            )
                        ideas = ideas.filter(author_foreign_key_field=author)
                    if moderate:
                        ideas = ideas.filter(status_moderate_char_field=moderate)
                    # TODO sort
                    if sort:
                        if sort == "дате публикации (свежие в начале)":
                            ideas = ideas.order_by("-register_datetime_field")
                        elif sort == "дате публикации (свежие в конце)":
                            ideas = ideas.order_by("register_datetime_field")
                        elif sort == "названию (с начала алфавита)":
                            ideas = ideas.order_by("name_char_field")
                        elif sort == "названию (с конца алфавита)":
                            ideas = ideas.order_by("-name_char_field")
                        elif sort == "рейтингу (популярные в начале)":
                            ideas_arr = []
                            for idea in ideas:
                                ratings = backend_models.IdeaRatingModel.objects.filter(
                                    idea_foreign_key_field=backend_models.IdeaModel.objects.get(id=idea.id)
                                )
                                if ratings.count() <= 0:
                                    rate = 0
                                else:
                                    rate = 0
                                    for i in ratings:
                                        rate += i.rating_integer_field
                                    rate = round(rate / ideas.count(), 2)
                                ideas_arr.append([rate, idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=True)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "рейтингу (популярные в конце)":
                            ideas_arr = []
                            for idea in ideas:
                                ratings = backend_models.IdeaRatingModel.objects.filter(
                                    idea_foreign_key_field=backend_models.IdeaModel.objects.get(id=idea.id)
                                )
                                if ratings.count() <= 0:
                                    rate = 0
                                else:
                                    rate = 0
                                    for i in ratings:
                                        rate += i.rating_integer_field
                                    rate = round(rate / ideas.count(), 2)
                                ideas_arr.append([rate, idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=False)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "отметкам рейтинга (наибольшие в начале)":
                            ideas_arr = []
                            for idea in ideas:
                                ratings = backend_models.IdeaRatingModel.objects.filter(
                                    idea_foreign_key_field=backend_models.IdeaModel.objects.get(id=idea.id)
                                )
                                ideas_arr.append([ratings.count(), idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=True)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "отметкам рейтинга (наибольшие в конце)":
                            ideas_arr = []
                            for idea in ideas:
                                ratings = backend_models.IdeaRatingModel.objects.filter(
                                    idea_foreign_key_field=backend_models.IdeaModel.objects.get(id=idea.id)
                                )
                                ideas_arr.append([ratings.count(), idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=False)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "комментариям (наибольшие в начале)":
                            ideas_arr = []
                            for idea in ideas:
                                comments = backend_models.IdeaCommentModel.objects.filter(
                                    idea_foreign_key_field=backend_models.IdeaModel.objects.get(id=idea.id)
                                )
                                ideas_arr.append([comments.count(), idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=True)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "комментариям (наибольшие в конце)":
                            ideas_arr = []
                            for idea in ideas:
                                comments = backend_models.IdeaCommentModel.objects.filter(
                                    idea_foreign_key_field=backend_models.IdeaModel.objects.get(id=idea.id)
                                )
                                ideas_arr.append([comments.count(), idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=False)
                            ideas = [x[1] for x in ideas_arr]
                    # TODO serialize
                    serializer = backend_serializers.IdeaModelSerializer(instance=ideas, many=True).data
                    response = {"response": serializer}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "IDEA_DETAIL":
                try:
                    # TODO get_value ###################################################################################
                    _id = req_inst.get_value("id")
                    # TODO actions #####################################################################################
                    if _id:
                        rational = backend_models.IdeaModel.objects.get(id=_id)
                    else:
                        rational = backend_models.IdeaModel.objects.order_by('-id')[0]
                    serializer = backend_serializers.IdeaModelSerializer(instance=rational, many=False).data
                    response = {"response": serializer}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "IDEA_CHANGE":
                try:
                    # TODO get_value ###################################################################################
                    _id = req_inst.get_value("id")
                    subdivision = req_inst.get_value("subdivision")
                    sphere = req_inst.get_value("sphere")
                    category = req_inst.get_value("category")
                    clear_image = req_inst.get_value("clearImage")
                    avatar = req_inst.get_value("avatar")
                    name = req_inst.get_value("name")
                    place = req_inst.get_value("place")
                    description = req_inst.get_value("description")
                    moderate = req_inst.get_value("moderate")
                    moderate_comment = req_inst.get_value("moderateComment")
                    # TODO actions #####################################################################################
                    idea = backend_models.IdeaModel.objects.get(id=_id)
                    if subdivision and idea.subdivision_char_field != subdivision:
                        idea.subdivision_char_field = subdivision
                    if sphere and idea.sphere_char_field != sphere:
                        idea.sphere_char_field = sphere
                    if category and idea.category_char_field != category:
                        idea.category_char_field = category
                    if clear_image:
                        idea.image_field = None
                    if avatar and idea.image_field != avatar:
                        idea.image_field = avatar
                    if name and idea.name_char_field != name:
                        idea.name_char_field = name
                    if place and idea.place_char_field != place:
                        idea.place_char_field = place
                    if description and idea.description_text_field != description:
                        idea.description_text_field = description
                    if moderate and idea.status_moderate_char_field != moderate:
                        idea.status_moderate_char_field = moderate
                    if moderate_comment and idea.comment_moderate_char_field != moderate_comment:
                        idea.comment_moderate_char_field = moderate_comment
                    idea.register_datetime_field = timezone.now()
                    idea.save()
                    if req_inst.user_model == idea.author_foreign_key_field:
                        backend_models.NotificationModel.objects.create(
                            author_foreign_key_field=req_inst.user_model,
                            model_foreign_key_field=backend_models.GroupModel.objects.get(
                                name_slug_field="moderator_idea"
                            ),
                            name_char_field="Идея отредактирована автором",
                            place_char_field="банк идей",
                            description_text_field=f"название идеи: {idea.name_char_field}",
                        )
                    response = {"response": "успешно изменено!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "IDEA_MODERATE":
                try:
                    # TODO get_value ###################################################################################
                    _id = req_inst.get_value("id")
                    moderate = req_inst.get_value("moderate")
                    moderate_comment = req_inst.get_value("moderateComment")
                    # TODO actions #####################################################################################
                    idea = backend_models.IdeaModel.objects.get(id=_id)
                    idea.status_moderate_char_field = moderate
                    idea.moderate_foreign_key_field = req_inst.user_model
                    idea.comment_moderate_char_field = moderate_comment
                    idea.register_datetime_field = timezone.now()
                    idea.save()
                    if moderate == "принято" or moderate == "на доработку":
                        message = "Действия с Вашей идеей"
                        if moderate == "принято":
                            message = "Ваша идея успешно принята и открыта в общем доступе!"
                        if moderate == "на доработку":
                            message = "Ваша идея отправлена на доработку! Проверьте список идей на доработку и " \
                                      "исправьте её!"
                        backend_models.NotificationModel.objects.create(
                            author_foreign_key_field=req_inst.user_model,
                            target_foreign_key_field=idea.author_foreign_key_field,
                            name_char_field=message,
                            place_char_field="банк идей",
                            description_text_field=f"название идеи: {idea.name_char_field}",
                        )
                    elif moderate == "скрыто":
                        message = "Действия с Вашей идеей"
                        if moderate == "скрыто":
                            message = "Автор скрыл свою идею!"
                        backend_models.NotificationModel.objects.create(
                            author_foreign_key_field=req_inst.user_model,
                            model_foreign_key_field=backend_models.GroupModel.objects.get(
                                name_slug_field="moderator_idea"
                            ),
                            name_char_field=message,
                            place_char_field="банк идей",
                            description_text_field=f"название идеи: {idea.name_char_field}",
                        )
                    response = {"response": "Модерация прошла успешно!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "IDEA_COMMENT_CREATE":
                try:
                    # TODO get_value ###################################################################################
                    _id = req_inst.get_value("id")
                    comment = req_inst.get_value("comment")
                    # TODO actions #####################################################################################
                    idea = backend_models.IdeaModel.objects.get(id=_id)
                    backend_models.IdeaCommentModel.objects.create(
                        idea_foreign_key_field=idea,
                        author_foreign_key_field=req_inst.user_model,
                        comment_text_field=comment,
                    )
                    response = {"response": "Комментарий успешно создан!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "IDEA_COMMENT_DELETE":
                try:
                    # TODO get_value ###################################################################################
                    comment_id = req_inst.get_value("commentId")
                    # TODO actions #####################################################################################
                    backend_models.IdeaCommentModel.objects.get(
                        id=comment_id,
                    ).delete()
                    response = {"response": "Комментарий успешно удалён!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "IDEA_RATING_CREATE":
                try:
                    # TODO get_value ###################################################################################
                    _id = req_inst.get_value("id")
                    rating = req_inst.get_value("rating")
                    # TODO actions #####################################################################################
                    idea = backend_models.IdeaModel.objects.get(id=_id)
                    rating_obj = backend_models.IdeaRatingModel.objects.get_or_create(
                        idea_foreign_key_field=idea, author_foreign_key_field=req_inst.user_model)[0]
                    rating_obj.rating_integer_field = rating
                    rating_obj.save()
                    ideas = backend_models.IdeaCommentModel.objects.filter(
                        idea_foreign_key_field=idea
                    ).order_by("-created_datetime_field")
                    serializer = backend_serializers.IdeaCommentModelSerializer(instance=ideas, many=True).data
                    response = {"response": serializer}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "IDEA_AUTHOR_LIST":
                try:
                    # TODO get_value ###################################################################################
                    sort = req_inst.get_value("sort")
                    only_month = req_inst.get_value("onlyMonth")
                    # TODO actions #####################################################################################
                    authors = []
                    ideas = backend_models.IdeaModel.objects.filter(
                        status_moderate_char_field="принято"
                    ).order_by("-register_datetime_field")
                    if only_month:
                        now = (datetime.datetime.now()).strftime('%Y-%m-%d %H:%M')
                        local_objects = []
                        for idea in ideas:
                            if (idea.register_datetime_field + datetime.timedelta(days=31)).strftime('%Y-%m-%d %H:%M') \
                                    >= now:
                                local_objects.append(idea.id)
                        ideas = backend_models.IdeaModel.objects.filter(
                            id__in=local_objects
                        ).order_by("-register_datetime_field")
                    for idea in ideas:
                        authors.append(idea.author_foreign_key_field)
                    authors = set(authors)
                    ideas = []
                    for author in authors:
                        ideas_arr = backend_models.IdeaModel.objects.filter(
                            author_foreign_key_field=author, status_moderate_char_field="принято"
                        )
                        idea_count = ideas_arr.count()
                        idea_rating = 0
                        idea_rating_count = 0
                        idea_comment_count = 0
                        for idea in ideas_arr:
                            ratings = backend_models.IdeaRatingModel.objects.filter(idea_foreign_key_field=idea)
                            for rate in ratings:
                                idea_rating += rate.rating_integer_field
                            idea_rating_count += ratings.count()
                            idea_comment_count = backend_models.IdeaCommentModel.objects.filter(
                                idea_foreign_key_field=idea
                            ).count()
                        if idea_rating_count == 0:
                            idea_rating_count = 1
                        serialized_user = backend_serializers.UserModelSerializer(instance=author, many=False).data
                        ideas.append({
                            "user_model": serialized_user,
                            "idea_count": idea_count, "idea_rating": round(idea_rating / idea_rating_count, 2),
                            "idea_rating_count": idea_rating_count, "idea_comment_count": idea_comment_count
                        })
                    # sort
                    if sort:
                        if sort == "количеству (наибольшие в начале)":
                            ideas_arr = []
                            for idea in ideas:
                                ideas_arr.append([idea["idea_count"], idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=True)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "количеству (наибольшие в конце)":
                            ideas_arr = []
                            for idea in ideas:
                                ideas_arr.append([idea["idea_count"], idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=False)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "рейтингу (популярные в начале)":
                            ideas_arr = []
                            for idea in ideas:
                                ideas_arr.append([idea["idea_rating"], idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=True)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "рейтингу (популярные в конце)":
                            ideas_arr = []
                            for idea in ideas:
                                ideas_arr.append([idea["idea_rating"], idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=False)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "отметкам рейтинга (наибольшие в начале)":
                            ideas_arr = []
                            for idea in ideas:
                                ideas_arr.append([idea["idea_rating_count"], idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=True)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "отметкам рейтинга (наибольшие в конце)":
                            ideas_arr = []
                            for idea in ideas:
                                ideas_arr.append([idea["idea_rating_count"], idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=False)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "комментариям (наибольшие в начале)":
                            ideas_arr = []
                            for idea in ideas:
                                ideas_arr.append([idea["idea_comment_count"], idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=True)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "комментариям (наибольшие в конце)":
                            ideas_arr = []
                            for idea in ideas:
                                ideas_arr.append([idea["idea_comment_count"], idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=False)
                            ideas = [x[1] for x in ideas_arr]
                    response = {"response": ideas}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            return Response({"error": "This action not allowed for this method."})
        else:
            return Response({"error": "This method not allowed for this endpoint."})
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
        return render(request, "backend/404.html")


# TODO test views ######################################################################################################
@api_view(http_method_names=["GET", "POST", "PUT", "DELETE"])
@permission_classes([IsAuthenticated])
def api_auth_rational(request):
    """
    api_rational django-rest-framework
    """

    try:
        # TODO Request #################################################################################################
        req_inst = backend_service.DjangoClass.TemplateClass.request(request=request)
        # TODO Methods #################################################################################################
        if req_inst.method == "POST":
            # TODO Actions #############################################################################################
            if req_inst.action_type == "RATIONAL_CREATE":
                try:
                    # TODO get_value ###################################################################################
                    subdivision = req_inst.get_value("subdivision")
                    category = req_inst.get_value("category")
                    avatar = req_inst.get_value("avatar")
                    name = req_inst.get_value("name")
                    place = req_inst.get_value("place")
                    description = req_inst.get_value("description")
                    additional_word = req_inst.get_value("additionalWord")
                    additional_pdf = req_inst.get_value("additionalPdf")
                    additional_excel = req_inst.get_value("additionalExcel")
                    user1 = req_inst.get_value("user1")
                    user2 = req_inst.get_value("user2")
                    user3 = req_inst.get_value("user3")
                    user4 = req_inst.get_value("user4")
                    user5 = req_inst.get_value("user5")
                    # TODO actions #####################################################################################
                    count = backend_models.RationalModel.objects.order_by('-id')
                    if len(count) > 0:
                        count = count[0].id + 1
                    else:
                        count = 1
                    number = f"{count}_{backend_service.DateTimeUtils.get_current_date()}"
                    user1 = f"{req_inst.user_model.last_name_char_field} " \
                            f"{req_inst.user_model.first_name_char_field} " \
                            f"{req_inst.user_model.personnel_number_slug_field} {user1}"
                    if len(user2) < 7:
                        user2 = None
                    if len(user3) < 7:
                        user3 = None
                    if len(user4) < 7:
                        user4 = None
                    if len(user5) < 7:
                        user5 = None
                    if subdivision == "автотранспортное предприятие":
                        name_slug_field = "moderator_rational_atp"
                    elif subdivision == "горно-транспортный комплекс":
                        name_slug_field = "moderator_rational_gtk"
                    elif subdivision == "обогатительный комплекс":
                        name_slug_field = "moderator_rational_ok"
                    elif subdivision == "управление предприятия":
                        name_slug_field = "moderator_rational_upravlenie"
                    elif subdivision == "энергоуправление":
                        name_slug_field = "moderator_rational_energoupravlenie"
                    else:
                        name_slug_field = "moderator_rational"
                    # TODO create ######################################################################################
                    backend_models.RationalModel.objects.create(
                        author_foreign_key_field=req_inst.user_model,
                        number_char_field=number,
                        subdivision_char_field=subdivision,
                        category_char_field=category,
                        image_field=avatar,
                        name_char_field=name,
                        place_char_field=place,
                        description_text_field=description,
                        additional_word_file_field=additional_word,
                        additional_pdf_file_field=additional_pdf,
                        additional_excel_file_field=additional_excel,
                        author_1_char_field=user1,
                        author_2_char_field=user2,
                        author_3_char_field=user3,
                        author_4_char_field=user4,
                        author_5_char_field=user5,
                        status_moderate_char_field="на модерации",
                    )
                    backend_models.NotificationModel.objects.create(
                        author_foreign_key_field=req_inst.user_model,
                        model_foreign_key_field=backend_models.GroupModel.objects.get(
                            name_slug_field=name_slug_field
                        ),
                        name_char_field="Создано новое рационализаторское предложение",
                        place_char_field="рационализаторство",
                        description_text_field=f"название рац. предложения: {name}",
                    )
                    response = {"response": "Успешно отправлено на модерацию!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "RATIONAL_LIST":
                try:
                    # TODO get_value ###################################################################################
                    moderate = req_inst.get_value("moderate")
                    subdivision = req_inst.get_value("subdivision")
                    category = req_inst.get_value("category")
                    author = req_inst.get_value("author")
                    search = req_inst.get_value("search")
                    sort = req_inst.get_value("sort")

                    only_month = req_inst.get_value("onlyMonth")
                    # TODO actions #####################################################################################
                    # TODO onlyMonth
                    objects = backend_models.RationalModel.objects.all().order_by("-register_datetime_field")
                    if only_month:
                        now = (datetime.datetime.now()).strftime('%Y-%m-%d %H:%M')
                        local_objects = []
                        for obj in objects:
                            if (obj.register_datetime_field + datetime.timedelta(days=31)).strftime('%Y-%m-%d %H:%M') \
                                    >= now:
                                local_objects.append(obj.id)
                        objects = backend_models.RationalModel.objects.filter(id__in=local_objects). \
                            order_by("-register_datetime_field")

                    # TODO search
                    if search:
                        objects = objects.filter(name_char_field__icontains=search)

                    # TODO filter
                    if subdivision:
                        objects = objects.filter(subdivision_char_field=subdivision)
                    if category:
                        objects = objects.filter(category_char_field=category)
                    if author:
                        author = backend_models.UserModel.objects.get(
                            personnel_number_slug_field=str(author).split(" ")[-2]
                        )
                        objects = objects.filter(author_foreign_key_field=author)
                    if moderate:
                        objects = objects.filter(status_moderate_char_field=moderate)

                    # TODO sort
                    if sort:
                        if sort == "дате публикации (сначала свежие)":
                            objects = objects.order_by("-register_datetime_field")
                        elif sort == "дате публикации (сначала старые)":
                            objects = objects.order_by("register_datetime_field")
                        elif sort == "названию (с начала алфавита)":
                            objects = objects.order_by("name_char_field")
                        elif sort == "названию (с конца алфавита)":
                            objects = objects.order_by("-name_char_field")
                        elif sort == "рейтингу (популярные в начале)":
                            pass
                        elif sort == "рейтингу (популярные в конце)":
                            pass
                        elif sort == "отметкам рейтинга (наибольшие в начале)":
                            pass
                        elif sort == "отметкам рейтинга (наибольшие в конце)":
                            pass
                        elif sort == "комментариям (наибольшие в начале)":
                            pass
                        elif sort == "комментариям (наибольшие в конце)":
                            pass

                    # TODO serialize
                    serializer = backend_serializers.RationalModelSerializer(instance=objects, many=True).data
                    response = {"response": serializer}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "RATIONAL_DETAIL":
                try:
                    # TODO get_value ###################################################################################
                    _id = req_inst.get_value("id")
                    # TODO actions #####################################################################################
                    if _id:
                        rational = backend_models.RationalModel.objects.get(id=_id)
                    else:
                        rational = backend_models.RationalModel.objects.order_by('-id')[0]
                    serializer = backend_serializers.RationalModelSerializer(instance=rational, many=False).data
                    response = {"response": serializer}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "RATIONAL_MODERATE":
                try:
                    # TODO get_value ###################################################################################
                    rational_id = req_inst.get_value("rationalId")
                    moderate = req_inst.get_value("moderate")
                    comment = req_inst.get_value("comment")
                    # TODO actions #####################################################################################
                    rational = backend_models.RationalModel.objects.get(id=rational_id)
                    if rational.status_moderate_char_field == "Предтехмодерация":
                        if moderate == "Принято":
                            rational.status_moderate_char_field = "Посттехмодерация"
                            rational.premoderate_foreign_key_field = req_inst.user_model
                            rational.comment_premoderate_char_field = comment
                        else:
                            rational.status_moderate_char_field = "Отклонено"
                            rational.premoderate_foreign_key_field = req_inst.user_model
                            rational.comment_premoderate_char_field = comment
                    elif rational.status_moderate_char_field == "Посттехмодерация":
                        if moderate == "Принято":
                            rational.status_moderate_char_field = "Принято"
                            rational.postmoderate_foreign_key_field = req_inst.user_model
                            rational.comment_postmoderate_char_field = comment
                        else:
                            rational.status_moderate_char_field = "Отклонено"
                            rational.postmoderate_foreign_key_field = req_inst.user_model
                            rational.comment_postmoderate_char_field = comment
                    elif rational.status_moderate_char_field == "Постнетехмодерация":
                        if moderate == "Принято":
                            rational.status_moderate_char_field = "Принято"
                            rational.postmoderate_foreign_key_field = req_inst.user_model
                            rational.comment_postmoderate_char_field = comment
                        else:
                            rational.status_moderate_char_field = "Отклонено"
                            rational.postmoderate_foreign_key_field = req_inst.user_model
                            rational.comment_postmoderate_char_field = comment
                    rational.save()
                    response = {"response": "Успешно"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            return Response({"error": "This action not allowed for this method."})
        else:
            return Response({"error": "This method not allowed for this endpoint."})
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
        return render(request, "backend/404.html")


@api_view(http_method_names=["GET", "POST", "PUT", "DELETE"])
@permission_classes([AllowAny])
def api_any_vacancy(request):
    """
    api_rational django-rest-framework
    """

    try:
        # TODO Request #################################################################################################
        req_inst = backend_service.DjangoClass.TemplateClass.request(request=request)
        # TODO Methods #################################################################################################
        if req_inst.method == "POST":
            # TODO Actions #############################################################################################
            if req_inst.action_type == "VACANCY_LIST":
                try:
                    # TODO get_value ###################################################################################
                    sphere = req_inst.get_value("sphere")
                    education = req_inst.get_value("education")
                    experience = req_inst.get_value("experience")
                    sort = req_inst.get_value("sort")
                    search = req_inst.get_value("search")
                    # TODO actions #####################################################################################
                    objects = backend_models.VacancyModel.objects.all().order_by("-id")
                    if sphere:
                        objects = objects.filter(sphere_char_field=sphere)
                    if education:
                        objects = objects.filter(education_char_field=education)
                    if experience:
                        objects = objects.filter(experience_char_field=experience)
                    if sort:
                        if sort == "Дате публикации (сначала свежие)":
                            objects = objects.order_by("-datetime_field")
                        elif sort == "Дате публикации (сначала старые)":
                            objects = objects.order_by("datetime_field")
                        elif sort == "Названию (С начала алфавита)":
                            objects = objects.order_by("qualification_char_field")
                        elif sort == "Названию (С конца алфавита)":
                            objects = objects.order_by("-qualification_char_field")
                    if search:
                        objects = objects.filter(qualification_char_field__icontains=search)
                    serializer = backend_serializers.VacancyModelSerializer(instance=objects, many=True).data
                    response = {"response": serializer}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "VACANCY_DETAIL":
                try:
                    # TODO get_value ###################################################################################
                    _id = req_inst.get_value("id")
                    # TODO actions #####################################################################################
                    if _id:
                        vacancy = backend_models.VacancyModel.objects.get(id=_id)
                    else:
                        vacancy = backend_models.VacancyModel.objects.order_by('-id')[0]
                    serializer = backend_serializers.VacancyModelSerializer(instance=vacancy, many=False).data
                    response = {"response": serializer}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            else:
                return Response({"error": "This action not allowed for this method."})
        else:
            return Response({"error": "This method not allowed for this endpoint."})
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
        return render(request, "backend/404.html")


@api_view(http_method_names=["GET", "POST", "PUT", "DELETE"])
@permission_classes([IsAuthenticated])
def api_auth_vacancy(request):
    """
    api_rational django-rest-framework
    """

    try:
        # TODO Request #################################################################################################
        req_inst = backend_service.DjangoClass.TemplateClass.request(request=request)
        # TODO Methods #################################################################################################
        if req_inst.method == "POST":
            # TODO Actions #############################################################################################
            if req_inst.action_type == "VACANCY_CREATE":
                try:
                    # TODO get_value ###################################################################################
                    qualification = req_inst.get_value("qualification")
                    rank = req_inst.get_value("rank")
                    sphere = req_inst.get_value("sphere")
                    education = req_inst.get_value("education")
                    experience = req_inst.get_value("experience")
                    schedule = req_inst.get_value("schedule")
                    image = req_inst.get_value("image")
                    description = req_inst.get_value("description")
                    # TODO actions #####################################################################################
                    if image and image != "null":
                        pass
                    else:
                        image = None
                    backend_models.VacancyModel.objects.create(
                        author_foreign_key_field=req_inst.user_model,
                        qualification_char_field=qualification,
                        rank_char_field=rank,
                        sphere_char_field=sphere,
                        education_char_field=education,
                        experience_char_field=experience,
                        schedule_char_field=schedule,
                        image_field=image,
                        description_text_field=description,
                    )
                    response = {"response": "Успешно создано!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "VACANCY_CHANGE":
                try:
                    # TODO get_value ###################################################################################
                    _id = req_inst.get_value("id")
                    qualification = req_inst.get_value("qualification")
                    rank = req_inst.get_value("rank")
                    sphere = req_inst.get_value("sphere")
                    education = req_inst.get_value("education")
                    experience = req_inst.get_value("experience")
                    schedule = req_inst.get_value("schedule")
                    image = req_inst.get_value("image")
                    clear_image = req_inst.get_value("clearImage")
                    description = req_inst.get_value("description")
                    # TODO actions #####################################################################################
                    if image and image != "null":
                        pass
                    else:
                        image = None
                    vacancy = backend_models.VacancyModel.objects.get(id=_id)
                    if req_inst.user_model and vacancy.author_foreign_key_field != req_inst.user_model:
                        vacancy.author_foreign_key_field = req_inst.user_model
                    if qualification and vacancy.qualification_char_field != qualification:
                        vacancy.qualification_char_field = qualification
                    if rank and vacancy.rank_char_field != rank:
                        vacancy.rank_char_field = rank
                    if sphere and vacancy.sphere_char_field != sphere:
                        vacancy.sphere_char_field = sphere
                    if education and vacancy.education_char_field != education:
                        vacancy.education_char_field = education
                    if experience and vacancy.experience_char_field != experience:
                        vacancy.experience_char_field = experience
                    if schedule and vacancy.schedule_char_field != schedule:
                        vacancy.schedule_char_field = schedule
                    if clear_image:
                        vacancy.image_field = None
                    if image and vacancy.image_field != image:
                        vacancy.image_field = image
                    if description and vacancy.description_text_field != description:
                        vacancy.description_text_field = description
                    vacancy.save()
                    response = {"response": "Успешно изменено!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "VACANCY_DELETE":
                try:
                    # TODO get_value ###################################################################################
                    _id = req_inst.get_value("id")
                    # TODO actions #####################################################################################
                    backend_models.VacancyModel.objects.get(id=_id).delete()
                    response = {"response": "Успешно удалено!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            else:
                return Response({"error": "This action not allowed for this method."})
        else:
            return Response({"error": "This method not allowed for this endpoint."})
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
        return render(request, "backend/404.html")


@api_view(http_method_names=["GET", "POST", "PUT", "DELETE"])
@permission_classes([AllowAny])
def api_any_resume(request):
    """
    api_rational django-rest-framework
    """

    try:
        # TODO Request #################################################################################################
        req_inst = backend_service.DjangoClass.TemplateClass.request(request=request)
        # TODO Methods #################################################################################################
        if req_inst.method == "POST":
            # TODO Actions #############################################################################################
            if req_inst.action_type == "RESUME_CREATE":
                try:
                    # TODO get_value ###################################################################################
                    qualification = req_inst.get_value("qualification")
                    last_name = req_inst.get_value("lastName")
                    first_name = req_inst.get_value("firstName")
                    patronymic = req_inst.get_value("patronymic")
                    image = req_inst.get_value("image")
                    birth_date = req_inst.get_value("birthDate")
                    education = req_inst.get_value("education")
                    experience = req_inst.get_value("experience")
                    sex = req_inst.get_value("sex")
                    contact_data = req_inst.get_value("contactData")
                    # TODO actions #####################################################################################
                    if image and image != "null":
                        pass
                    else:
                        image = None
                    backend_models.ResumeModel.objects.create(
                        qualification_char_field=qualification,
                        last_name_char_field=last_name,
                        first_name_char_field=first_name,
                        patronymic_char_field=patronymic,
                        image_field=image,
                        birth_datetime_field=birth_date,
                        education_char_field=education,
                        experience_char_field=experience,
                        sex_char_field=sex,
                        contact_data_text_field=contact_data,
                    )
                    response = {"response": "Ваше резюме успешно сохранено!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            else:
                return Response({"error": "This action not allowed for this method."})
        else:
            return Response({"error": "This method not allowed for this endpoint."})
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
        return render(request, "backend/404.html")


@api_view(http_method_names=["GET", "POST", "PUT", "DELETE"])
@permission_classes([IsAuthenticated])
def api_auth_resume(request):
    """
    api_rational django-rest-framework
    """

    try:
        # TODO Request #################################################################################################
        req_inst = backend_service.DjangoClass.TemplateClass.request(request=request)
        # TODO Methods #################################################################################################
        if req_inst.method == "POST":
            # TODO Actions #############################################################################################
            if req_inst.action_type == "RESUME_LIST":
                try:
                    # TODO get_value ###################################################################################
                    education = req_inst.get_value("education")
                    experience = req_inst.get_value("experience")
                    sex = req_inst.get_value("sex")
                    sort = req_inst.get_value("sort")
                    search_qualification = req_inst.get_value("searchQualification")
                    search_last_name = req_inst.get_value("searchLastName")
                    # TODO actions #####################################################################################
                    resume_list = backend_models.ResumeModel.objects.all().order_by("-id")
                    if education:
                        resume_list = resume_list.filter(education_char_field=education)
                    if experience:
                        resume_list = resume_list.filter(experience_char_field=experience)
                    if sex:
                        resume_list = resume_list.filter(sex_char_field=sex)
                    if sort:
                        if sort == "Дате публикации (сначала свежие)":
                            resume_list = resume_list.order_by("-create_datetime_field")
                        elif sort == "Дате публикации (сначала старые)":
                            resume_list = resume_list.order_by("create_datetime_field")
                        elif sort == "Названию вакансии (С начала алфавита)":
                            resume_list = resume_list.order_by("qualification_char_field")
                        elif sort == "Названию вакансии (С конца алфавита)":
                            resume_list = resume_list.order_by("-qualification_char_field")
                    if search_qualification:
                        resume_list = resume_list.filter(qualification_char_field__icontains=search_qualification)
                    if search_last_name:
                        resume_list = resume_list.filter(last_name_char_field__icontains=search_last_name)
                    serializer = backend_serializers.ResumeModelSerializer(instance=resume_list, many=True).data
                    response = {"response": serializer}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "RESUME_DETAIL":
                try:
                    # TODO get_value ###################################################################################
                    _id = req_inst.get_value("id")
                    # TODO actions #####################################################################################
                    if _id:
                        vacancy = backend_models.ResumeModel.objects.get(id=_id)
                    else:
                        vacancy = backend_models.ResumeModel.objects.order_by('-id')[0]
                    serializer = backend_serializers.ResumeModelSerializer(instance=vacancy, many=False).data
                    response = {"response": serializer}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            if req_inst.action_type == "RESUME_DELETE":
                try:
                    # TODO get_value ###################################################################################
                    _id = req_inst.get_value("id")
                    # TODO actions #####################################################################################
                    backend_models.ResumeModel.objects.get(id=_id).delete()
                    response = {"response": "Успешно удалено!"}
                    # TODO response ####################################################################################
                    backend_service.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            else:
                return Response({"error": "This action not allowed for this method."})
        else:
            return Response({"error": "This method not allowed for this endpoint."})
    except Exception as error:
        backend_service.DjangoClass.LoggingClass.error(request=request, error=error)
        return render(request, "backend/404.html")
