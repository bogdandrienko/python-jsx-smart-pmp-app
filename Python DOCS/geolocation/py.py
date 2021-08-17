import psycopg2 as pg
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import datetime


connection = pg.connect(
    host="192.168.1.6",
    database="navSections",
    port="5432",
    user="postgres",
    password="nF2ArtXK"
)

postgreSQL_select_Query = "SELECT * FROM public.navdata_202108 " \
                          "WHERE flags = '0' " \
                          "ORDER BY navtime DESC, device DESC LIMIT 100;"
                          # "ORDER BY navtime DESC, device DESC LIMIT 100"
                          # "ORDER BY device ASC, navtime DESC LIMIT 100"
                          # "WHERE flags = 0 " \

cursor = connection.cursor()

cursor.execute(postgreSQL_select_Query)

mobile_records = cursor.fetchall()

wb = Workbook()
sheet = wb.active

# cols = ["device", "navtime", "latitude", "longtode", "alttude", "speed", "ds", "direction", "flags"]
cols = ["устройство", "дата и время", "широта", "долгота", "высота", "скорость", "ds", "направление", "флаги ошибок"]

for col in cols:
    id_s = cols.index(col)
    sheet[f'{get_column_letter(id_s + 1)}1'] = col

for rows in mobile_records:
    for value in rows:
        id_s = rows.index(value)
        # print(f"{cols[id_s]}: {value}")
        if id_s == 1:
            sheet[f'{get_column_letter(id_s + 1)}{mobile_records.index(rows) + 2}'] = \
                datetime.datetime.fromtimestamp(int(value-21600)).strftime('%Y-%m-%d %H:%M:%S')
        else:
            sheet[f'{get_column_letter(id_s + 1)}{mobile_records.index(rows) + 2}'] = value

wb.save('data.xlsx')