import openpyxl
from openpyxl.utils import get_column_letter


def get_sheet_value(column, row, _sheet):
    return str(_sheet[str(column) + str(row)].value)


def set_sheet_value(column, row, value):
    sheet[f'{column}{row}'] = str(value)


file_1c = 'export.xlsx'
workers_id_1c = []
id_1c = 9

workbook = openpyxl.load_workbook(file_1c)
sheet = workbook.active
for num in range(1, 2500):
    workers_id_1c.append(get_sheet_value(get_column_letter(int(id_1c)), num, _sheet=sheet))
workbook.close()

file_skud = 'import.xlsx'
workers_id_skud = []
id_skud = 3

workbook = openpyxl.load_workbook(file_skud)
sheet = workbook.active
for num in range(1, 2500):
    workers_id_skud.append(get_sheet_value(get_column_letter(int(id_skud)), num, sheet=sheet))
workbook.close()

workbook = openpyxl.load_workbook(file_1c)
sheet = workbook.active

for x in workers_id_1c:
    for y in workers_id_skud:
        if x == y and x is not None:
            set_sheet_value(column=get_column_letter(int(8)), row=workers_id_1c.index(x)+1, value='+')
workbook.save('export.xlsx')
