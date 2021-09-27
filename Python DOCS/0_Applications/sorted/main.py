import openpyxl
from openpyxl.utils import get_column_letter
from threading import Thread
import tkinter


def click_button(export_file='Export.xlxs', import_file='Import.xlxs', exporting='ABCDEFINY', importing='RSTBAUCVF'):
    def get_sheet_value(column, row, sheet):
        value = str(sheet[str(column) + str(row)].value)
        if value:
            return value
        else:
            return ''

    def set_sheet_value(column, row, value_, sheet):
        if value_:
            sheet[str(column) + str(row)] = value_
        else:
            sheet[str(column) + str(row)] = ''

    def whiles():
        minerals_file = 'min.xlsx'
        minerals_iin = 'BF'
        min_minerals_value = 3
        max_minerals_value = 2200
        workers_from_minerals = []
        workbook = openpyxl.load_workbook(minerals_file)
        sheet = workbook.active
        for num in range(min_minerals_value, max_minerals_value):
            value = get_sheet_value(minerals_iin, num, sheet)
            # print(f"{minerals_iin}{num}: {value}")
            if value != 'None':
                workers_from_minerals.append(value)
        # print(workers_from_minerals)
        workbook.close()

        hospital_file = 'hospital.xlsx'
        hospital_iin = 'F'
        min_hospital_value = 2
        max_hospital_value = 7500
        workers_from_hospital = []
        workbook = openpyxl.load_workbook(hospital_file)
        sheet = workbook.active
        titles = []
        for val in 'ABCDEFGH':
            value = get_sheet_value(val, 1, sheet)
            titles.append(value)
        for num in range(min_hospital_value, max_hospital_value):
            local_workers_from_hospital = []
            for val in 'ABCDEFG':
                value = get_sheet_value(val, num, sheet)
                local_workers_from_hospital.append(value)
            # print(f"{hospital_iin}{num}: {local_workers_from_hospital}")
            workers_from_hospital.append(local_workers_from_hospital)
        # print(workers_from_hospital)
        workbook.close()

        for man in workers_from_hospital:
            try:
                truth = workers_from_minerals.index(man[5])
                workers_from_hospital[workers_from_hospital.index(man)].append('+')
                # print(f"{man[5]}: TRUE")
            except Exception as ex:
                workers_from_hospital[workers_from_hospital.index(man)].append('-')
                # print(f"{man[5]}: FALSE")
        # print(workers_from_hospital)

        hospital_new_file = 'hospital_new.xlsx'
        workbook = openpyxl.load_workbook(hospital_new_file)
        sheet = workbook.active
        for val in titles:
            set_sheet_value(get_column_letter(titles.index(val) + 1), 1, str(val), sheet)
        for ind in workers_from_hospital:
            for val in ind:
                # print(f"{workers_from_hospital.index(ind) + 1}:{ind.index(val) + 1}: {val}")
                set_sheet_value(get_column_letter(ind.index(val) + 1), workers_from_hospital.index(ind) + 2, str(val),
                                sheet)
        workbook.save(hospital_new_file)
        workbook.close()
        print("complete")

    thread_result = Thread(target=whiles)
    thread_result.start()


class Application(tkinter.Frame):
    def __init__(self, root, **kw):
        super().__init__(**kw)
        self.root = root
        self.root.title("Сравнение данных")
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        self.root.config(background="black")
        self.root.geometry('640x480')
        self.master.minsize(640, 480)
        self.master.maxsize(640, 480)
        self.id = 0
        self.iid = 0

        self.submit_button = tkinter.Button(self.root, text="Запустить", font="100", command=self.insert_data)
        self.submit_button.grid(row=0, column=0, sticky=tkinter.W)

        self.exit_button = tkinter.Button(self.root, text="Выход", font="100", command=self.quit)
        self.exit_button.grid(row=0, column=1, sticky=tkinter.W)

        self.export_label = tkinter.Label(self.root, text="Файл откуда экспортировать", font="100")
        self.export_label.grid(row=1, column=0, sticky=tkinter.W)
        self.export_entry = tkinter.Entry(self.root, font="100")
        self.export_entry.grid(row=1, column=1, sticky=tkinter.W)
        self.export_entry.insert(0, 'Export.xlsx')

        self.import_label = tkinter.Label(self.root, text="Файл куда импортировать", font="100")
        self.import_label.grid(row=2, column=0, sticky=tkinter.W)
        self.import_entry = tkinter.Entry(self.root, font="100")
        self.import_entry.grid(row=2, column=1, sticky=tkinter.W)
        self.import_entry.insert(0, 'Import.xlsx')

        self.imp_label = tkinter.Label(self.root, text="Столбцы из импорта", font="100")
        self.imp_label.grid(row=3, column=0, sticky=tkinter.W)
        self.exp_label = tkinter.Label(self.root, text="Соответствуют столбцам из экспорта", font="100")
        self.exp_label.grid(row=3, column=1, sticky=tkinter.W)

        self.importing_entry = tkinter.Entry(self.root, font="100")
        self.importing_entry.grid(row=4, column=1, sticky=tkinter.W)
        self.importing_entry.insert(0, 'RSTBAUCVF')

        self.exporting_entry = tkinter.Entry(self.root, font="100")
        self.exporting_entry.grid(row=4, column=0, sticky=tkinter.W)
        self.exporting_entry.insert(0, 'ABCDEFINY')

        self.save_label = tkinter.Label(self.root, text="Файл куда сохранять", font="100")
        self.save_label.grid(row=5, column=0, sticky=tkinter.W)
        self.save_entry = tkinter.Entry(self.root, font="100")
        self.save_entry.grid(row=5, column=1, sticky=tkinter.W)
        self.save_entry.insert(0, 'Save.xlsx')

    def insert_data(self):
        click_button(self.export_entry.get(), self.import_entry.get(),
                     self.exporting_entry.get(), self.importing_entry.get())


if __name__ == "__main__":
    app = Application(tkinter.Tk())
    thread_main = Thread(target=app.root.mainloop())
    thread_main.start()
