import time
import sys
import pyodbc
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from PySide6 import QtCore, QtWidgets
from threading import Thread


def play_analyse(temp=37.0):
    def whiles():
        def pyodbc_connect(ip="192.168.15.87", server="DESKTOP-SM7K050", port="1434", database="thirdpartydb",
                           username="sa", password="skud12345678"):
            conn_str = (
                    r'DRIVER={ODBC Driver 17 for SQL Server};SERVER=tcp:' + ip + '\\' + server + ',' + port +
                    ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password + ';'
            )
            return pyodbc.connect(conn_str)

        sql_select_query = f"SELECT * " \
                           f"FROM dbtable " \
                           f"WHERE CAST(temperature AS FLOAT) >= {temp} " \
                           f"ORDER BY date1 DESC, date2 DESC;"
        connect_db = pyodbc_connect()
        cursor = connect_db.cursor()
        cursor.fast_executemany = True
        cursor.execute(sql_select_query)

        wb = Workbook()
        sheet = wb.active

        cols = ["табельный", "доступ", "дата", "время", "данные", "точка", "номер карты", "температура", "маска"]
        for col in cols:
            id_s = cols.index(col)
            sheet[f'{get_column_letter(id_s + 1)}1'] = col

        row_index = 0
        for rows in cursor.fetchall():
            row_index += 1
            value_index = 0
            for value in rows:
                value_index += 1
                if value_index != 5:
                    sheet[f'{get_column_letter(value_index)}{row_index + 1}'] = value
                else:
                    try:
                        sheet[f'{get_column_letter(value_index)}{row_index + 1}'] = value.encode('1251').decode('utf-8')
                    except Exception as ex_1:
                        val = str(value).split(" ")
                        string = "N" + val[0][2:] + " " + "N" + val[1][2:]
                        try:
                            sheet[f'{get_column_letter(value_index)}{row_index + 1}'] = string.encode('1251').decode(
                                'utf-8')
                        except Exception as ex_2:
                            sheet[f'{get_column_letter(value_index)}{row_index + 1}'] = value
                            print(ex_2)
                        print(ex_1)

        wb.save('data.xlsx')
        widget.color_box.setText("завершено")
        time.sleep(1)
        widget.color_box.setText("ожидание")

    thread_render = Thread(target=whiles)
    thread_render.start()


class MyWidget(QtWidgets.QWidget):
    def __init__(self, title="термометрия"):
        super().__init__()
        self.play_button = QtWidgets.QPushButton("выгрузить")
        self.quit_button = QtWidgets.QPushButton("выйти")
        self.temp_box = QtWidgets.QDoubleSpinBox()
        self.color_box = QtWidgets.QLabel("ожидание")
        self.temp_box.setValue(37.0)
        self.setWindowTitle(title)

        self.ui_window = QtWidgets.QHBoxLayout(self)
        self.ui_window.addWidget(self.play_button)
        # self.ui_window.addWidget(self.temp_box)
        self.ui_window.addWidget(self.color_box)
        self.ui_window.addWidget(self.quit_button)

        self.play_button.clicked.connect(self.play_btn_func)
        self.quit_button.clicked.connect(self.quit_btn_func)

    @QtCore.Slot()
    def play_btn_func(self):
        play_analyse(temp=self.temp_box.value())
        self.set_text_func("в процессе")

    @QtCore.Slot()
    def quit_btn_func(self):
        global app
        sys.exit(app.exec_())

    @QtCore.Slot()
    def set_text_func(self, text: str):
        self.color_box.setText(text)


if __name__ == "__main__":
    app = QtWidgets.QApplication([])
    widget = MyWidget()
    widget.resize(320, 240)
    thread_main = Thread(target=widget.show())
    thread_main.start()
    sys.exit(app.exec_())
