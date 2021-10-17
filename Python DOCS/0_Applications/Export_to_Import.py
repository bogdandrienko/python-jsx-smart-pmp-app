import openpyxl
from openpyxl.utils import get_column_letter
import os
import tkinter
from threading import Thread


def click_button(export_file='Export.xlxs', import_file='Import.xlxs', exporting='ABCDEFINY', importing='RSTBAUCVF'):
    def get_sheet_value(column, row, sheet):
        return str(sheet[str(column) + str(row)].value)

    def set_sheet_value(column, row, value_, sheet):
        if value_:
            sheet[str(column) + str(row)] = value_
        else:
            sheet[str(column) + str(row)] = ''

    def whiles(_export_file=export_file, _import_file=import_file, _exporting=exporting, _importing=importing):
        app.root.config(background="red")

        relative_path = os.path.dirname(os.path.abspath('__file__'))
        _export_file = relative_path + '\\' + _export_file
        _import_file = relative_path + '\\' + _import_file

        min_export_value = 1
        max_export_value = 5000
        min_import_value = 14
        max_import_value = 5000
        workers_from_1c = []

        workbook = openpyxl.load_workbook(_export_file)
        sheet = workbook.active
        workbook.close()

        for num in range(min_export_value, max_export_value):
            worker_list = []
            for x in _exporting:
                value = get_sheet_value(x, num, sheet)
                if value == 'None' or value is None or value == '':
                    value = ''
                worker_list.append(value)
            worker_id = Worker(*worker_list)
            workers_from_1c.append(worker_id)

        workbook = openpyxl.load_workbook(_import_file)
        sheet = workbook.active

        workers_from_db = []

        for num in range(min_import_value, max_import_value):
            workers_from_db.append(get_sheet_value(get_column_letter(3), num, sheet))

        for worker_from_1C in workers_from_1c:
            for worker_from_DB in workers_from_db:
                if worker_from_1C.get_worker_id() == worker_from_DB:
                    for x in _importing:
                        try:
                            set_sheet_value(x, workers_from_db.index(worker_from_DB) + 14,
                                            worker_from_1C.get_worker_value(_importing.index(x)), sheet)
                        except:
                            pass

        workbook.save(_import_file)
        workbook.close()

    thread_result = Thread(target=whiles)
    thread_result.start()


class Application(tkinter.Frame):
    def __init__(self, root, **kw):
        super().__init__(**kw)
        self.root = root
        self.root.title("Импорт дополнительных данных")
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        self.root.config(background="black")
        self.root.geometry('640x480')
        self.master.minsize(640, 480)
        self.master.maxsize(640, 480)
        self.id = 0
        self.iid = 0

        self.submit_button = tkinter.Button(self.root, text="Запустить", font="100", command=self.insert_data)
        self.submit_button.grid(row=0, column=0, sticky=tkinter.W)

        self.exit_button = tkinter.Button(self.root, text="Выход", font="100", command=self.quit)
        self.exit_button.grid(row=0, column=1, sticky=tkinter.W)

        self.export_label = tkinter.Label(self.root, text="Файл откуда экспортировать", font="100")
        self.export_label.grid(row=1, column=0, sticky=tkinter.W)
        self.export_entry = tkinter.Entry(self.root, font="100")
        self.export_entry.grid(row=1, column=1, sticky=tkinter.W)
        self.export_entry.insert(0, 'Export.xlsx')

        self.import_label = tkinter.Label(self.root, text="Файл куда импортировать", font="100")
        self.import_label.grid(row=2, column=0, sticky=tkinter.W)
        self.import_entry = tkinter.Entry(self.root, font="100")
        self.import_entry.grid(row=2, column=1, sticky=tkinter.W)
        self.import_entry.insert(0, 'Import.xlsx')

        self.imp_label = tkinter.Label(self.root, text="Столбцы из импорта", font="100")
        self.imp_label.grid(row=3, column=0, sticky=tkinter.W)
        self.exp_label = tkinter.Label(self.root, text="Соответствуют столбцам из экспорта", font="100")
        self.exp_label.grid(row=3, column=1, sticky=tkinter.W)

        self.importing_entry = tkinter.Entry(self.root, font="100")
        self.importing_entry.grid(row=4, column=1, sticky=tkinter.W)
        self.importing_entry.insert(0, 'QRSBATUCV')

        self.exporting_entry = tkinter.Entry(self.root, font="100")
        self.exporting_entry.grid(row=4, column=0, sticky=tkinter.W)
        self.exporting_entry.insert(0, 'ABCDEFINY')

    def insert_data(self):
        click_button(self.export_entry.get(), self.import_entry.get(),
                     self.exporting_entry.get(), self.importing_entry.get())


class Worker:
    """
    Класс, который содержит в себе работника, со значениями по строке
    """

    def __init__(self, a_1='', b_1='', c_1='', d_1='', e_1='', f_1='', g_1='', h_1='', m_1=''):
        # Подразделение
        self.A_1 = a_1
        # Цех или Служба
        self.B_1 = b_1
        # Отдел или участок
        self.C_1 = c_1
        # Фамилия
        self.D_1 = d_1
        # Имя
        self.E_1 = e_1
        # Отчество
        self.F_1 = f_1
        # Табельный №
        self.G_1 = g_1
        # Категория
        self.H_1 = h_1
        # Пол
        self.M_1 = m_1

    def print_worker(self):
        print(f'{self.A_1}+{self.B_1}+{self.C_1}+{self.D_1}+{self.E_1}+{self.F_1}+{self.G_1}+{self.H_1}+{self.M_1}')

    def get_worker_value(self, index):
        value_ = list((self.A_1, self.B_1, self.C_1, self.D_1, self.E_1, self.F_1, self.G_1, self.H_1, self.M_1))
        return value_[index]

    def get_worker_id(self):
        return self.G_1


if __name__ == "__main__":
    app = Application(tkinter.Tk())
    thread_main = Thread(target=app.root.mainloop())
    thread_main.start()
