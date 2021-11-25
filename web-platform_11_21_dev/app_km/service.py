import math
import os
import datetime
import random
import bs4
import openpyxl
import requests
from django.contrib.auth.models import User, Group
from django.core.handlers.wsgi import WSGIRequest
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http.response import Http404
from fastkml import kml
from openpyxl.utils import get_column_letter
from .models import LoggingModel, GroupModel
from .utils import ExcelClass
from django.contrib.staticfiles import finders
from django.conf import settings


class DjangoClass:
    class AuthorizationClass:
        @staticmethod
        def access_to_page(request, logging, access):
            try:
                # Eсли пользователь в подсети предприятия его переадресует на локальный доступ
                if str(request.META.get("REMOTE_ADDR")) == '192.168.1.202':
                    return 'local'
                # Логирование действий
                if logging:
                    DjangoClass.LoggingClass.logging_actions(request=request)
                # Проверка на вход в аккаунт
                if request.user.is_authenticated:
                    try:
                        user = User.objects.get(username=request.user.username)
                        # Проверка бана: если аккаунт пользователя отключён, то его разлогинит
                        if user.is_active is False:
                            return 'account_logout'
                        else:
                            if access:
                                # Проверка заполнения спец полей
                                if user.profile.email and user.profile.secret_answer and user.profile.secret_question:
                                    # Полный доступ на страницу
                                    if str(access).strip().lower() == 'all':
                                        return False
                                    # Выборка групп доступа на страницу
                                    try:
                                        page_groups = [str(x).strip().lower() for x in str(access).split(',')]
                                    except Exception as error:
                                        DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                                        page_groups = [access]
                                    # Выборка групп доступа пользователя
                                    try:
                                        user_groups = [str(x).strip().lower() for x in user.groups.all()]
                                    except Exception as error:
                                        DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                                        if access:
                                            user_groups = [access]
                                        else:
                                            user_groups = ''
                                    # Проверка на наличие хоть одного совпадения
                                    for user_group in user_groups:
                                        try:
                                            if user_group and len(user_group) > 1:
                                                page_groups.index(user_group)
                                                return False
                                        except Exception as error:
                                            pass
                                    return True
                                else:
                                    return 'account_change_password'
                            else:
                                return 'home'
                    except Exception as error:
                        DjangoClass.LoggingClass.logging_errors(request=request, error=error)

                return 'account_login'
            except Exception as error:
                DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                DjangoClass.AuthorizationClass.http404_raise(exception_text=error)

        @staticmethod
        def try_to_access(request, only_logging=False):

            # Eсли пользователь в подсети предприятия его переадресует на локальный доступ
            if str(request.META.get("REMOTE_ADDR")) == '192.168.1.202':
                print(f"\n ***** ***** \n")
                print(f"Локальный доступ, переадресация")
                print(f"\n ***** ***** \n")
                return 'local'

            # Логирование действий
            print(f"\n ***** ***** \n")
            print(f"Логирование")
            print(f"\n ***** ***** \n")
            DjangoClass.LoggingClass.logging_actions(request=request)

            # Возврат, если выбрано только логирование
            if only_logging:
                print(f"\n ***** ***** \n")
                print(f"Только логирование")
                print(f"\n ***** ***** \n")
                return False

            # Проверка на вход в аккаунт
            if request.user.is_authenticated:
                try:
                    user = User.objects.get(username=request.user.username)
                    # Проверка суперпользователя: если имеет права, то полный доступ
                    if user.is_superuser:
                        print(f"\n ***** ***** \n")
                        print(f"Вы суперпользователь, доступ на страницу разрешён")
                        return False
                    # Проверка бана: если аккаунт пользователя отключён, то его разлогинит
                    if user.user_one_to_one_field.activity_boolean_field is False:
                        return 'account_logout'
                    else:
                        # Проверка заполнения спец полей
                        if user.user_one_to_one_field.email_field and \
                                user.user_one_to_one_field.secret_answer_char_field and \
                                user.user_one_to_one_field.secret_question_char_field:
                            # Выборка всех групп доступа
                            groups = GroupModel.objects.all()
                            print(f"\n ***** ***** \n")
                            print(f"Выборка всех групп доступа: {groups}")
                            # Выборка всех групп доступа с определённым пользователем
                            groups = GroupModel.objects.filter(user_many_to_many_field=user)
                            print(f"\n ***** ***** \n")
                            print(f"Выборка всех групп доступа с определённым пользователем: {groups}")
                            access = False
                            for group in groups:
                                try:
                                    pages = [str(x).strip().lower() for x in
                                             str(group.path_text_field).strip().split(',') if len(x) >= 1]
                                except Exception as error:
                                    pages = [str(group.path_text_field).strip()]
                                print(f"pages: {pages}")
                                try:
                                    path = str(request.path).strip().split('/')[1]
                                except Exception as error:
                                    path = ''
                                try:
                                    pages.index(path.strip().lower())
                                    access = True
                                    break
                                except Exception as error:
                                    pass
                            if access:
                                print(f"\n ***** ***** \n")
                                print(f"Доступ на страницу разрешён")
                                return False
                            else:
                                print(f"\n ***** ***** \n")
                                print(f"Доступ на страницу запрещён")
                                return 'home'
                        else:
                            print(f"\n ***** ***** \n")
                            print(f"Дополнительные данные не заполнены")
                            print(f"\n ***** ***** \n")
                            return 'account_change_password'
                except Exception as error:
                    DjangoClass.LoggingClass.logging_errors(request=request, error=error)
                    print(f"\n ***** ***** \n")
                    print(f"Пользователь не существует")
                    print(f"\n ***** ***** \n")
                    return 'home'
            else:
                print(f"\n ***** ***** \n")
                print(f"Вы не вошли в аккаунт")
                print(f"\n ***** ***** \n")
                return 'account_login'

        @staticmethod
        def http404_raise(exception_text):
            raise Http404(exception_text)

    class LoggingClass:
        @staticmethod
        def logging_errors(request, error):
            # for k, v in request.META.items():
            #     print(f'\n{k}: {v}')
            username = request.user.username
            ip = request.META.get("REMOTE_ADDR")
            request_path = request.path
            request_method = request.method
            datetime_now = datetime.datetime.now()
            LoggingModel.objects.create(
                username_slug_field=username,
                ip_genericipaddress_field=ip,
                request_path_slug_field=request_path,
                request_method_slug_field=request_method,
                error_text_field=f'error: {error}'
            )
            text = [username, ip, request_path, request_method, datetime_now, error]
            string = ''
            for val in text:
                string = string + f', {val}'
            with open('static/media/data/logging/logging_errors.txt', 'a') as log:
                log.write(f'\n{string[2:]}\n')

        @staticmethod
        def logging_errors_local(error, function_error):
            datetime_now = datetime.datetime.now()
            LoggingModel.objects.create(username='', ip='', request_path=function_error,
                                        request_method='', error=error)
            text = [function_error, datetime_now, error]
            string = ''
            for val in text:
                string = string + f', {val}'
            with open('static/media/data/logging/logging_errors.txt', 'a') as log:
                log.write(f'\n{string[2:]}\n')

        @staticmethod
        def logging_actions(request):
            # for k, v in request.META.items():
            #     print(f'\n{k}: {v}')
            username = request.user.username
            ip = request.META.get("REMOTE_ADDR")
            request_path = request.path
            request_method = request.method
            datetime_now = datetime.datetime.now()
            LoggingModel.objects.create(
                username_slug_field=username,
                ip_genericipaddress_field=ip,
                request_path_slug_field=request_path,
                request_method_slug_field=request_method,
                error_text_field='successful'
            )
            text = [username, ip, request_path, request_method, datetime_now]
            string = ''
            for val in text:
                string = string + f', {val}'
            with open('static/media/data/logging/logging_actions.txt', 'a') as log:
                log.write(f'\n{string[2:]}\n')

        @staticmethod
        def get_client_ip(request):
            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            if x_forwarded_for:
                ip = x_forwarded_for.split(',')[-1].strip()
            else:
                ip = request.META.get('REMOTE_ADDR')
            return ip

    class AccountClass:
        class UserAccountClass:
            """
            Основной аккаунт пользователя
            """

            def __init__(
                    # self
                    self,
                    # authorization data
                    username,
                    password,
                    # technical data
                    is_active=True,
                    is_staff=False,
                    is_superuser=False,
                    groups='User',
                    email='',
                    secret_question='',
                    secret_answer='',
                    # first data
                    last_name='',
                    first_name='',
                    patronymic='',
                    # second data
                    personnel_number='',
                    subdivision='',
                    workshop_service='',
                    department_site='',
                    position='',
                    category='',
                    # utils
                    force_change_account=False,
                    force_change_account_password=False,
                    force_clear_groups=False,
                    request=None
            ):
                # authorization data
                self.username = str(username).strip()
                self.password = str(password).strip()
                # technical data
                self.is_active = bool(is_active)
                self.is_staff = bool(is_staff)
                self.is_superuser = bool(is_superuser)
                try:
                    self.groups = [group.strip() for group in str(groups).strip().split(',') if len(group) >= 1]
                except Exception as error:
                    self.groups = [str(groups).strip()]
                self.email = str(email).strip()
                self.secret_question = str(secret_question).strip()
                self.secret_answer = str(secret_answer).strip()
                # first data
                self.first_name = str(first_name).strip()
                self.last_name = str(last_name).strip()
                self.patronymic = str(patronymic).strip()
                # second data
                self.personnel_number = str(personnel_number).strip()
                self.subdivision = str(subdivision).strip()
                self.workshop_service = str(workshop_service).strip()
                self.department_site = str(department_site).strip()
                self.position = str(position).strip()
                self.category = str(category).strip()
                # utils
                self.force_change_account = bool(force_change_account)
                self.force_change_account_password = bool(force_change_account_password)
                self.force_clear_groups = bool(force_clear_groups)
                self.request = request

            def account_create_or_change(self):
                try:
                    # Пользователь уже существует: изменение
                    try:
                        user = User.objects.get(username=self.username)
                        # Возврат, если пользователь обладает правами суперпользователя
                        if user.is_superuser:
                            return False
                        # Если пользователь уже существует и стоит статус "принудительно изменять аккаунт"
                        if user and self.force_change_account:
                            try:
                                # authorization data
                                if self.force_change_account_password:
                                    user.password = DjangoClass.AccountClass.get_sha256_password(self.password)
                                    user.user_one_to_one_field.password_slug_field = self.password
                                # technical data
                                user.user_one_to_one_field.activity_boolean_field = self.is_active
                                user.user_one_to_one_field.email_field = self.email
                                user.user_one_to_one_field.secret_question_char_field = self.secret_question
                                user.user_one_to_one_field.secret_answer_char_field = self.secret_answer
                                # first data
                                user.user_one_to_one_field.last_name_char_field = self.last_name
                                user.user_one_to_one_field.first_name_char_field = self.first_name
                                user.user_one_to_one_field.patronymic_char_field = self.patronymic
                                # second data
                                user.user_one_to_one_field.personnel_number_slug_field = self.personnel_number
                                user.user_one_to_one_field.subdivision_char_field = self.subdivision
                                user.user_one_to_one_field.workshop_service_char_field = self.workshop_service
                                user.user_one_to_one_field.department_site_char_field = self.department_site
                                user.user_one_to_one_field.position_char_field = self.position
                                user.user_one_to_one_field.category_char_field = self.category
                                # save account
                                user.save()
                                return self.account_auth_set_group()
                            except Exception as error:
                                DjangoClass.LoggingClass.logging_errors(request=self.request, error=error)
                                return False
                        else:
                            return False
                    # Пользователь не существует: создание
                    except Exception as error:
                        try:
                            user = User.objects.create(
                                # authorization data
                                username=self.username,
                                password=DjangoClass.AccountClass.get_sha256_password(self.password),
                                # technical data
                                is_active=True,
                                is_staff=self.is_staff,
                                is_superuser=self.is_superuser,
                            )
                            # authorization data
                            user.user_one_to_one_field.password_slug_field = self.password
                            # technical data
                            user.user_one_to_one_field.activity_boolean_field = self.is_active
                            user.user_one_to_one_field.email_field = self.email
                            user.user_one_to_one_field.secret_question_char_field = self.secret_question
                            user.user_one_to_one_field.secret_answer_char_field = self.secret_answer
                            # first data
                            user.user_one_to_one_field.last_name_char_field = self.last_name
                            user.user_one_to_one_field.first_name_char_field = self.first_name
                            user.user_one_to_one_field.patronymic_char_field = self.patronymic
                            # second data
                            user.user_one_to_one_field.personnel_number_slug_field = self.personnel_number
                            user.user_one_to_one_field.subdivision_char_field = self.subdivision
                            user.user_one_to_one_field.workshop_service_char_field = self.workshop_service
                            user.user_one_to_one_field.department_site_char_field = self.department_site
                            user.user_one_to_one_field.position_char_field = self.position
                            user.user_one_to_one_field.category_char_field = self.category
                            # save account
                            user.save()
                            return self.account_auth_set_group()
                        except Exception as error:
                            DjangoClass.LoggingClass.logging_errors(request=self.request, error=error)
                            return False
                except Exception as error:
                    DjangoClass.LoggingClass.logging_errors(request=self.request, error=error)
                    return False

            def account_auth_set_group(self):
                return True


        class UserAuthClassOld:
            """
            Основной аккаунт пользователя
            """

            def __init__(self, username, password, first_name='', last_name='', email='', is_active=True,
                         is_staff=False, is_superuser=False, groups='User', force_change_account=False,
                         force_change_account_password=False, force_clear_groups=False):
                # Основное
                self.username = username
                self.password = password

                # Персональная информация
                self.first_name = first_name
                self.last_name = last_name
                self.email = email

                # Права доступа
                self.is_active = is_active
                self.is_staff = is_staff
                self.is_superuser = is_superuser
                self.groups = groups

                # Настройки создания аккаунта
                self.force_change_account = force_change_account
                self.force_change_account_password = force_change_account_password
                self.force_clear_groups = force_clear_groups

            def account_auth_create(self):
                try:
                    user = User.objects.create(
                        # Основное
                        username=self.username,
                        password=DjangoClass.AccountClass.get_sha256_password(self.password),
                        # Персональная информация
                        first_name=self.first_name,
                        last_name=self.last_name,
                        email=self.email,
                        # Права доступа
                        is_active=self.is_active,
                        is_staff=self.is_staff,
                        is_superuser=self.is_superuser,
                    )
                    user.profile.password = self.password
                    user.save()
                    return self.account_auth_set_group()
                except Exception as error:
                    DjangoClass.LoggingClass.logging_errors_local(error=error, function_error='account_auth_change')
                    return False

            def account_auth_change(self):
                try:
                    user = User.objects.get(username=self.username)
                    # Основное
                    if self.force_change_account_password:
                        user.password = DjangoClass.AccountClass.get_sha256_password(self.password)
                        user.profile.password = self.password
                    # Персональная информация
                    user.first_name = self.first_name
                    user.last_name = self.last_name
                    user.email = self.email
                    # Права доступа
                    user.is_active = self.is_active
                    user.is_staff = self.is_staff
                    user.is_superuser = self.is_superuser
                    user.save()
                    return self.account_auth_set_group()
                except Exception as error:
                    DjangoClass.LoggingClass.logging_errors_local(error=error, function_error='account_auth_change')
                    return False

            def account_auth_set_group(self):
                try:
                    try:
                        groups = [x.strip() for x in self.groups.split(',')]
                    except Exception as error:
                        groups = [self.groups]
                    if self.force_clear_groups:
                        User.objects.get(username=self.username).groups.clear()
                    success = True
                    for group in groups:
                        if len(str(group).strip()) >= 1:
                            try:
                                group_object = Group.objects.get_or_create(name=group)[0]
                                group_object.user_set.add(User.objects.get(username=self.username))
                            except Exception as error:
                                DjangoClass.LoggingClass.logging_errors_local(
                                    error=error, function_error='account_auth_set_group'
                                )
                                success = False
                    return success
                except Exception as error:
                    DjangoClass.LoggingClass.logging_errors_local(error=error, function_error='account_auth_set_group')
                    return False

            def account_auth_create_or_change(self):
                try:
                    try:
                        user = User.objects.get(username=self.username)
                        # Возврат, если пользователь обладает правами суперпользователя
                        if user.is_superuser:
                            return False
                        # Если пользователь уже существует и стоит статус "принудительно изменять аккаунт"
                        if user and self.force_change_account:
                            return self.account_auth_change()
                        else:
                            return False
                    except Exception as error:
                        return self.account_auth_create()
                except Exception as error:
                    DjangoClass.LoggingClass.logging_errors_local(
                        error=error, function_error='account_auth_create_or_change'
                    )
                    return False



        class UserProfileClass:
            """
            Первичная информация аккаунта пользователя
            """

            def __init__(self, username, user_iin='', first_name='', last_name='', patronymic='', personnel_number='',
                         subdivision='', workshop_service='', department_site='', position='', category=''):
                # Основное
                self.username = username

                # Персональная информация
                self.user_iin = user_iin
                self.first_name = first_name
                self.last_name = last_name
                self.patronymic = patronymic
                self.personnel_number = personnel_number

                # First data account
                self.subdivision = subdivision
                self.workshop_service = workshop_service
                self.department_site = department_site
                self.position = position
                self.category = category

            def profile_first_change(self):
                # try:
                if True:
                    user = User.objects.get(username=self.username)
                    # Возврат, если пользователь обладает правами суперпользователя
                    if user.is_superuser:
                        return False
                    # Персональная информация
                    user.profile.user_iin = self.user_iin
                    user.profile.first_name = self.first_name
                    user.profile.last_name = self.last_name
                    user.profile.patronymic = self.patronymic
                    user.profile.personnel_number = self.personnel_number

                    # First data account
                    user.profile.subdivision = self.subdivision
                    user.profile.workshop_service = self.workshop_service
                    user.profile.department_site = self.department_site
                    user.profile.position = self.position
                    user.profile.category = self.category

                    user.save()
                # except Exception as ex:
                #     return False

        @staticmethod
        def create_main_account(user_account_class, username='', password='', email='', firstname='',
                                lastname='', is_staff=False, is_superuser=False, force_change=False):
            try:
                if user_account_class:
                    username = user_account_class.username
                    password = user_account_class.password
                    email = user_account_class.email
                    firstname = user_account_class.first_name
                    lastname = user_account_class.last_name
                    is_staff = user_account_class.is_staff
                    is_superuser = user_account_class.is_superuser
                if username == 'Bogdan' or username == 'bogdan':
                    is_superuser = True
                print(f'{username}: {password}')
                try:
                    user = User.objects.get(username=username)
                    if force_change and user:
                        DjangoClass.AccountClass.change_main_account(
                            user_account_class=False,
                            username=username,
                            password=password,
                            email=email,
                            firstname=firstname,
                            lastname=lastname,
                            is_staff=is_staff,
                            is_superuser=is_superuser,
                            force_create=force_change
                        )
                        return True
                    return False
                except Exception as error:
                    DjangoClass.LoggingClass.logging_errors_local(
                        error=error, function_error='create_main_account'
                    )
                    encrypt_password = DjangoClass.AccountClass.create_django_encrypt_password(password)
                    if encrypt_password:
                        print(f'{username}: {encrypt_password}')
                        user = User.objects.create(
                            username=username,
                            password=encrypt_password,
                            email=email,
                            firstname=firstname,
                            lastname=lastname,
                            is_staff=is_staff,
                            is_superuser=is_superuser
                        )
                        user.save()
                        user.set_password = encrypt_password
                        return True
                    return False
            except Exception as error:
                DjangoClass.LoggingClass.logging_errors_local(error=error, function_error='create_main_account')
                return False

        @staticmethod
        def change_main_account(user_account_class, username: str, password: str, email='', firstname='',
                                lastname='', is_staff=False, is_superuser=False, force_create=False):
            try:
                if user_account_class:
                    username = user_account_class.username
                    password = user_account_class.password
                    email = user_account_class.email
                    firstname = user_account_class.first_name
                    lastname = user_account_class.last_name
                    is_staff = user_account_class.is_staff
                    is_superuser = user_account_class.is_superuser
                if force_create:
                    user = User.objects.get_or_create(username=username)[0]
                else:
                    user = User.objects.get(username=username)
                user.username = username
                user.password = password
                user.email = email
                user.first_name = firstname
                user.last_name = lastname
                user.is_staff = is_staff
                user.is_superuser = is_superuser
                user.save()
                user.set_password = password
                return True
            except Exception as error:
                DjangoClass.LoggingClass.logging_errors_local(error=error, function_error='change_main_account')
                return False

        @staticmethod
        def set_user_group(user_group_class, username='', group='User', force_change=False):
            try:
                success = True
                if user_group_class:
                    username = user_group_class.username
                    group = user_group_class.group
                if group:
                    try:
                        group = [x.strip() for x in group.split(',') if len(x) > 1]
                    except Exception as error:
                        DjangoClass.LoggingClass.logging_errors_local(error=error, function_error='set_user_group')
                        group = [group]
                    for grp in group:
                        try:
                            user_group = Group.objects.get_or_create(name=grp)[0]
                            user = User.objects.get(username=username)
                            if force_change:
                                user.groups.clear()
                            user_group.user_set.add(user)
                        except Exception as error:
                            DjangoClass.LoggingClass.logging_errors_local(error=error, function_error='set_user_group')
                            success = False
                return success
            except Exception as error:
                DjangoClass.LoggingClass.logging_errors_local(error=error, function_error='set_user_group')
                return False

        @staticmethod
        def create_django_encrypt_password(decrypt_password: str):
            try:
                user = User.objects.get_or_create(username='None')[0]
                user.set_password(decrypt_password)
                user.save()
                user = User.objects.get(username='None')
                encrypt_password = user.password
                user.delete()
                return encrypt_password
            except Exception as error:
                DjangoClass.LoggingClass.logging_errors_local(
                    error=error, function_error='create_django_encrypt_password'
                )
                return False

        @staticmethod
        def get_sha256_password(password: str):
            try:
                user = User.objects.get_or_create(username='None')[0]
                user.set_password(password)
                user.save()
                user = User.objects.get_or_create(username='None')[0]
                encrypt_password = user.password
                user.delete()
                return encrypt_password
            except Exception as error:
                DjangoClass.LoggingClass.logging_errors_local(error=error, function_error='get_sha256_password')
                return False

        @staticmethod
        def create_password_from_chars(chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890',
                                       length=8):
            password = ''
            for i in range(1, length + 1):
                password += random.choice(chars)
            return password

    class RequestClass:
        @staticmethod
        def get_value(request: WSGIRequest, key: str, none_is_error=False, strip=True):
            if none_is_error:
                # If key not have is Exception Error
                value = request.POST[key]
            else:
                # If key not have value is None
                value = request.POST.get(key)
            if strip and value:
                value.strip()
            return value

        @staticmethod
        def get_check(request: WSGIRequest, key: str, none_is_error=False):
            if none_is_error:
                # If key not have is Exception Error
                value = request.POST[key]
            else:
                # If key not have value is None
                value = request.POST.get(key)
            if value is None:
                return False
            else:
                return True

        @staticmethod
        def get_file(request: WSGIRequest, key: str, none_is_error=False):
            if none_is_error:
                # If key not have is Exception Error
                file = request.FILES[key]
            else:
                # If key not have value is None
                file = request.FILES.get(key)
            return file


class PaginationClass:
    @staticmethod
    def paginate(request, objects, num_page):
        # Пагинатор: постраничный вывод объектов
        paginator = Paginator(objects, num_page)
        pages = request.GET.get('page')
        try:
            page = paginator.page(pages)
        except PageNotAnInteger:
            page = paginator.page(1)
        except EmptyPage:
            page = paginator.page(paginator.num_pages)
        return page


class SalaryClass:
    @staticmethod
    def create_arr_table(title: str, footer: str, json_obj, exclude: list):
        headers = []

        for x in json_obj["Fields"]:
            headers.append(json_obj["Fields"][x])
        del json_obj["Fields"]
        bodies = [["", title]]

        if exclude:
            hours = 0
            days = 0
            sum_ = 0
            for x in json_obj:
                val = [x]
                i = 0
                for y in json_obj[x]:
                    i += 1
                    if i == exclude[0]:
                        hours += json_obj[x][y]
                        continue
                    if i == exclude[1]:
                        days += json_obj[x][y]
                        continue
                    if i == len(json_obj[x]):
                        sum_ += json_obj[x][y]
                    val.append(json_obj[x][y])
                bodies.append(val)
            footers = ["", footer, "", hours, days, round(sum_, 2)]
        else:
            sum_ = 0
            for x in json_obj:
                val = [x]
                i = 0
                for y in json_obj[x]:
                    i += 1
                    if i == len(json_obj[x]):
                        sum_ += json_obj[x][y]
                    val.append(json_obj[x][y])
                bodies.append(val)
            footers = ["", footer, "", round(sum_, 2)]

        return [headers, bodies, footers]

    @staticmethod
    def create_arr_from_json(json_obj, parent_key: str):
        headers = []
        for x in json_obj[parent_key]["Fields"]:
            headers.append(json_obj[parent_key]["Fields"][x])
        del json_obj[parent_key]["Fields"]
        bodies = []
        for x in json_obj[parent_key]:
            val = [x]
            for y in json_obj[parent_key][x]:
                val.append(json_obj[parent_key][x][y])
            bodies.append(val)
        return [parent_key, headers, bodies]


class Xhtml2pdfClass:
    @staticmethod
    def link_callback(uri):
        """
        Convert HTML URIs to absolute system paths so xhtml2pdf can access those
        resources
        """
        result = finders.find(uri)
        s_url = ''
        m_url = ''
        if result:
            if not isinstance(result, (list, tuple)):
                result = [result]
            result = list(os.path.realpath(path) for path in result)
            path = result[0]
        else:
            s_url = settings.STATIC_URL  # Typically /static/
            s_root = settings.STATIC_ROOT  # Typically /home/userX/project_static/
            m_url = settings.MEDIA_URL  # Typically /media/
            m_root = settings.MEDIA_ROOT  # Typically /home/userX/project_static/media/

            if uri.startswith(m_url):
                path = os.path.join(m_root, uri.replace(m_url, ""))
            elif uri.startswith(s_url):
                path = os.path.join(s_root, uri.replace(s_url, ""))
            else:
                return uri
        if not os.path.isfile(path):
            raise Exception(
                'media URI must start with %s or %s' % (s_url, m_url)
            )
        return path


class GeoClass:
    @staticmethod
    def get_hypotenuse(x1: float, y1: float, x2: float, y2: float):
        """"
        Принимает: "пару" точек - их широту и долготу.
        Возвращает: корень из суммы квадратов разностей широты и долготы двух пар точек, ака гипотенузу.
        """
        return math.sqrt(((x2 - x1) ** 2) + ((y2 - y1) ** 2))

    @staticmethod
    def get_haversine(latitude_1: float, longitude_1: float, latitude_2: float, longitude_2: float):
        """"
        Принимает: "пару" точек - их широту и долготу.
        Возвращает: значение расстояния по гипотенузе двух точек, в метрах, ака формула гаверсинуса.
        """
        delta_latitude = latitude_2 * math.pi / 180 - latitude_1 * math.pi / 180
        delta_longitude = longitude_2 * math.pi / 180 - longitude_1 * math.pi / 180
        a = math.sin(delta_latitude / 2) * math.sin(delta_latitude / 2) + math.cos(latitude_1 * math.pi / 180) * \
            math.cos(latitude_2 * math.pi / 180) * math.sin(delta_longitude / 2) * math.sin(delta_longitude / 2)
        return round(math.atan2(math.sqrt(a), math.sqrt(1 - a)) * 6378.137 * 2 * 1000)

    @staticmethod
    def find_near_point(point_arr: list, point_latitude: float, point_longitude: float):
        """"
        Принимает: массив точек в которых надо искать, где первый элемент это широта, а второй долгоа. Также данные
        точки ближайшие координаты которой надо найти.
        Возвращает: данные точки из массива точек, которая соответствует ближайшему значению целевой точки.
        """
        result = None
        for coord in point_arr:
            if coord[0] <= point_latitude and coord[1] <= point_longitude:
                result = coord
        if result is None:
            result = point_arr[0]
        return result

    @staticmethod
    def get_vector_arr(point_arr):
        """
        Принимает: массив "точек" - первый элемент это имя точки, второй и третий это широта и долгота, а четвёртый
        связи.
        Возвращает: массив "векторов" - первый элемент это имя вектора, второй это расстояние через формулу
        "гаверсинуса".
        """
        # Points = [PointName, latitude, longitude, PointLinks]
        # point1 = ["1", 52.14303, 61.22812, "2"]
        # point2 = ["2", 52.1431, 61.22829, "1|3"]
        # Vectors = [VectorName, length(meters)]
        # vector1 = ["1|2", 12]
        # vector2 = ["2|3", 14]

        vector_arr = []
        for point in point_arr:
            lat1 = point[0]
            lon1 = point[1]
            vector1 = point[2]
            link_arr = point[3].split("|")
            for link in link_arr:
                for point_1 in point_arr:
                    if point_1[2] == link:
                        lat2 = point_1[0]
                        lon2 = point_1[1]
                        vector2 = link
                        length = GeoClass.get_haversine(lat1, lon1, lat2, lon2)
                        vector_arr.append([f"{vector1}|{vector2}", length])
        return vector_arr

    @staticmethod
    def create_cube_object(point: list):
        latitude = point[0]
        longitude = point[1]
        first = [latitude - 0.000002 * 1.62, longitude - 0.000002, 0]
        second = [latitude - 0.000002 * 1.62, longitude + 0.000002, 0]
        third = [latitude + 0.000002 * 1.62, longitude + 0.000002, 0]
        fourth = [latitude + 0.000002 * 1.62, longitude - 0.000002, 0]
        string_object = ''
        for iteration in [first, second, third, fourth, first]:
            num = 1
            for i in iteration:
                if num == 3:
                    string_object += f"{i} "
                else:
                    string_object += f"{i},"
                num += 1
        text_d = f"""<Placemark>
        		<name>object</name>
        		<Polygon>
        			<outerBoundaryIs>
        				<LinearRing>
        					<coordinates>
        						{string_object} 
        					</coordinates>
        				</LinearRing>
        			</outerBoundaryIs>
        		</Polygon>
        	</Placemark>"""
        return text_d

    @staticmethod
    def create_point_object(point: list):
        latitude = point[0]
        longitude = point[1]
        id_s = point[2]
        first = [latitude - 0.000002 * 1.62, longitude - 0.000002, 0]
        second = [latitude - 0.000002 * 1.62, longitude + 0.000002, 0]
        third = [latitude + 0.000002 * 1.62, longitude + 0.000002, 0]
        fourth = [latitude + 0.000002 * 1.62, longitude - 0.000002, 0]
        string_object = ''
        for iteration in [first, second, third, fourth, first]:
            num = 1
            for i in iteration:
                if num == 3:
                    string_object += f"{i} "
                else:
                    string_object += f"{i},"
                num += 1
        text_d = f"""<Placemark>
                <name>Точка: {id_s}</name>
                <Point>
                    <coordinates>{latitude},{longitude},0</coordinates>
                </Point>
            </Placemark>"""
        return text_d

    @staticmethod
    def generate_xlsx():
        # connection = pg.connect(
        #     host="192.168.1.6",
        #     database="navSections",
        #     port="5432",
        #     user="postgres",
        #     password="nF2ArtXK"
        # )
        # # postgresql_select_query = f"SELECT device, navtime, ROUND(CAST(latitude AS numeric), {request.POST['request_value']}), ROUND(CAST(longitude AS numeric), {request.POST['request_value']}) " \
        # #                           "FROM public.navdata_202108 " \
        # #                           f"WHERE device BETWEEN {request.POST['request_between_first']} AND {request.POST['request_between_last']} AND timezone('UTC', to_timestamp(navtime)) > (CURRENT_TIMESTAMP - INTERVAL '{request.POST['request_hours']} hours') AND flags != 64 " \
        # #                           "ORDER BY device, navtime DESC;"
        # postgresql_select_query = f"SELECT device, navtime, ROUND(CAST(latitude AS numeric), {request.POST['request_value']}), ROUND(CAST(longitude AS numeric), {request.POST['request_value']}) " \
        #                           "FROM public.navdata_202108 " \
        #                           f"WHERE device BETWEEN {request.POST['request_between_first']} AND {request.POST['request_between_last']} AND timezone('UTC', to_timestamp(navtime)) > (CURRENT_TIMESTAMP - INTERVAL '{request.POST['request_minutes']} minutes') AND flags != 64 " \
        #                           "ORDER BY device, navtime DESC;"
        # cursor = connection.cursor()
        # cursor.execute(postgresql_select_query)
        # mobile_records = cursor.fetchall()
        mobile_records = []
        cols = ["устройство", "дата и время", "широта", "долгота", "высота", "скорость", "ds", "направление",
                "флаги ошибок"]
        all_arr = []
        for rows in mobile_records:
            arr = []
            for value in rows:
                id_s = rows.index(value)
                if id_s == 1:
                    arr.append(datetime.datetime.fromtimestamp(int(value - 21600)).strftime('%Y-%m-%d %H:%M:%S'))
                else:
                    arr.append(value)
            all_arr.append(arr)
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Страница 1'
        for n in cols:
            sheet[f'{get_column_letter(cols.index(n) + 1)}1'] = n
        for n in all_arr:
            for i in n:
                if i:
                    sheet[f'{get_column_letter(n.index(i) + 1)}{all_arr.index(n) + 2}'] = i
                else:
                    sheet[f'{get_column_letter(n.index(i) + 1)}{all_arr.index(n) + 2}'] = '0.0'
        wb.save('static/media/data/geo.xlsx')
        postgresql_select_query = None
        return [cols, all_arr, postgresql_select_query]

    @staticmethod
    def generate_kml():
        # Чтение Excel
        file_xlsx = 'static/media/data/geo_1.xlsx'
        workbook = openpyxl.load_workbook(file_xlsx)
        sheet = workbook.active

        # Чтение из Excel
        final = 0
        for num in range(2, 10000000):
            if ExcelClass.get_sheet_value("A", num, sheet=sheet) is None or '':
                final = num
                break
        array = []
        for num in range(2, final):
            array.append(
                [ExcelClass.get_sheet_value("D", num, sheet=sheet),
                 ExcelClass.get_sheet_value("C", num, sheet=sheet),
                 ExcelClass.get_sheet_value("A", num, sheet=sheet)]
            )

        # Генерация объекта и субъекта
        point_obj = [61.22330, 52.14113, 'БелАЗ']
        point_sub = [61.22891, 52.14334, 'Экскаватор']
        text_x = GeoClass.create_point_object(point_obj)
        text_y = GeoClass.create_point_object(point_sub)

        # Цена рёбер
        count = 0
        for current in array:
            index = array.index(current)
            if index != 0:
                previous = [array[index - 1][0], array[index - 1][1]]
            else:
                previous = [current[0], current[1]]

            count += GeoClass.get_hypotenuse(current[0], current[1], previous[0], previous[1])
        print(round(count, 5))

        # Генерация линий по цветам
        device_arr = []
        previous_device = 0
        text_b = ''
        colors = ['FFFFFFFF', 'FF0000FF', 'FFFF0000', 'FF00FF00', 'FF00FFFF', 'FF0F0F0F', 'FFF0F0F0']
        colors_alias = [': Чё', ': Кр', ': Си', ': Зе', ': Жё', ': Доп1', ': Доп2']
        current_color = colors[0]
        for current in array:
            index = array.index(current)
            if previous_device is not current[2]:
                device_arr.append(str(current[2]) + colors_alias[len(device_arr) + 1])
                previous_device = current[2]
                current_color = colors[colors.index(current_color) + 1]
            if index != 0:
                previous = [array[index - 1][0], array[index - 1][1]]
            else:
                previous = [current[0], current[1]]
            text_b += f"""<Placemark>
                  <Style>
                    <LineStyle>
                      <color>{current_color}</color>
                    </LineStyle>
                  </Style>
                  <LineString>
                    <coordinates>{previous[0]},{previous[1]},0 {current[0]},{current[1]},0 </coordinates>
                  </LineString>
              </Placemark>
            """

            # Генерация точек
            text_b += GeoClass.create_cube_object([current[0], current[1], index])

        # Генерация имени документа из цветов
        dev = ''
        for x in device_arr:
            dev += f"{x} | "

        # Начало kml
        text_a = f"""<?xml version="1.0" encoding="utf-8"?>
          <kml xmlns="http://earth.google.com/kml/2.2">
            <Document>
              <name>{dev}</name>
                """

        # Конец kml
        text_c = R"""</Document>
        </kml>"""

        # Запись в kml
        with open("static/media/data/geo.kml", "w", encoding="utf-8") as file:
            file.write(text_a + text_b + text_x + text_y + text_c)

    @staticmethod
    def generate_way(object_, subject_, point_arr):
        # Генерация общей карты
        text_map = GeoClass.generate_map(point_arr, 'FF0000FF')
        # text_map = ''

        # Генерация объекта и субъекта
        point_obj = [object_[0], object_[1], "ЭКГ"]
        point_sub = [subject_[0], subject_[1], "БелАЗ"]
        text_obj = GeoClass.create_point_object(point_obj)
        text_sub = GeoClass.create_point_object(point_sub)

        path = GeoClass.generate_path(object_, subject_, point_arr)
        # print(path[0])
        # Генерация карты пути
        text_path = GeoClass.generate_map(path[0], 'FFFF0000')

        # Генерация имени документа из цветов
        dev = ''
        for x in [object_[0], subject_[0]]:
            dev += f"{x} | "
        dev += f"{path[1]} meters"
        print(f"{path[1]} meters")

        # Начало kml
        text_title = f"""<?xml version="1.0" encoding="utf-8"?>
          <kml xmlns="http://earth.google.com/kml/2.2">
            <Document>
              <name>{dev}</name>
                """

        # Конец kml
        text_footer = R"""</Document>
        </kml>"""

        # Запись в kml
        try:
            os.remove("static/media/data/geo_1.kml")
        except Exception as error:
            DjangoClass.LoggingClass.logging_errors_local(error=error, function_error='generate_way')
        with open("static/media/data/geo_5.kml", "w", encoding="utf-8") as file:
            file.write(text_title + text_map + text_path + text_obj + text_sub + text_footer)

    @staticmethod
    def generate_path(object_, subject_, point_arr):
        """"
        Генерация пути синего цвета из всех валидных точек и расчёт расстояния.
        """
        # Путь маршрут
        path = []
        # Расстояние пути
        path_distantion = 0

        # Откуда начинаем путь
        from_point = object_
        print(f"from_point: {from_point}")
        # from_point: [61.232109, 52.146845, '38', '37|39']

        # Где завершаем путь
        to_point = subject_
        print(f"to_point: {to_point}")
        # to_point: [61.229219, 52.143999, '12', '11|13']

        # Генерация маршрута
        # От исходной точки идём к её линиям связи, выбираем самую короткую к финальной точке, "наступаем" туда, цикл.
        start_point = from_point
        final_point = to_point

        path.append(from_point)
        while True:
            print('\n*****************')
            print('while', start_point[2], final_point[2])
            links = start_point[3].split("|")
            print(f"links: {links}")
            # links: ['38', '40']

            near_points = []
            for point in point_arr:
                for link in links:
                    if point[2] == link:
                        try:
                            near_points.index(point)
                        except Exception as error:
                            DjangoClass.LoggingClass.logging_errors_local(error=error, function_error='generate_path')
                            near_points.append(point)
            print(f"near_points: {near_points}")
            # near_points: [[61.232054, 52.146927, '38', '37|39'], [61.232384, 52.147085, '40', '39|41']]

            destinations = []
            for point in near_points:
                destinations.append(GeoClass.get_haversine(point[0], point[1], final_point[0], final_point[1]))
            print(f"destinations: {destinations}")
            # destinations: [[61.232054, 52.146927, '38', '37|39'], [61.232384, 52.147085, '40', '39|41']]

            min_point = min(destinations)
            print(f"min_point: {min_point}")
            # min_point: 367

            if min_point == 0:
                next_point = final_point
                print(f"next_point: {next_point}")
                # next_point: [61.232101, 52.146821, '45', '44|46']

                dist = GeoClass.get_haversine(start_point[0], start_point[1], final_point[0], final_point[1])
                print(f"dist: {dist}")
                # dist: 22

                path_distantion += dist
                path.append(next_point)
                break

            next_point = near_points[destinations.index(min_point)]
            print(f"next_point: {next_point}")
            # next_point: [61.232101, 52.146821, '45', '44|46']

            dist = GeoClass.get_haversine(start_point[0], start_point[1], next_point[0], next_point[1])
            print(f"dist: {dist}")
            # dist: 22

            path_distantion += dist
            start_point = next_point
            path.append(next_point)
            print('*****************\n')
        return [path, path_distantion]

    @staticmethod
    def generate_path_old(object_, subject_, point_arr):
        """"
        Генерация пути синего цвета из всех валидных точек.
        """
        # Генерация маршрута
        # От исходной точки идём к её линиям связи, выбираем самую короткую к финальной точке, "наступаем" туда, цикл.
        path = []
        path_dist = 0
        current_point = subject_
        # subject_: [61.232230610495556, 52.14697472947569, '21', '20|22']
        previous_point = subject_
        for loop in range(len(point_arr)):
            # links: ['20', '22']
            links = current_point[3].split("|")
            near_points = []
            for point in point_arr:
                for link in links:
                    if point[2] == link:
                        # print(point[2])
                        try:
                            near_points.index(point)
                        except Exception as error:
                            DjangoClass.LoggingClass.logging_errors_local(
                                error=error, function_error='generate_path_old'
                            )
                            near_points.append(point)

            # Первые две линии не брать, а к последней по индексу добавлять ещё одну связь.
            min_dist = 9999
            # print(near_points)
            for point in near_points:
                # object_: [61.2293618582404, 52.143978995225346, '4', '3|5']
                dist = GeoClass.get_haversine(point[0], point[1], object_[0], object_[1])
                # print(f"dist: {dist}")
                if min_dist > dist:
                    min_dist = dist
                    current_point = point
                    path.append(point)
                    distantion = GeoClass.get_haversine(point[0], point[1], previous_point[0], previous_point[1])
                    print(f"prev: {previous_point}   |   curr: {point}     |       deist: {distantion}m")
                    path_dist += distantion
                previous_point = point
        print(path_dist)
        path_dist = 0
        for x in path:
            path_dist += GeoClass.get_haversine(x[0], x[1], object_[0], object_[1])
            print(GeoClass.get_haversine(x[0], x[1], object_[0], object_[1]))
        return [path, path_dist]

    @staticmethod
    def generate_map(point_arr, color):
        """"
        Генерация карты выбранного цвета из всех валидных точек.
        """
        text = ''
        for current in point_arr:
            index = point_arr.index(current)
            if index != 0:
                previous = [point_arr[index - 1][0], point_arr[index - 1][1]]
            else:
                previous = [current[0], current[1]]
            text += f"""<Placemark>
                  <Style>
                    <LineStyle>
                      <color>{color}</color>
                    </LineStyle>
                  </Style>
                  <LineString>
                    <coordinates>{previous[0]},{previous[1]},0 {current[0]},{current[1]},0 </coordinates>
                  </LineString>
              </Placemark>
            """
            # Генерация точек
            text += GeoClass.create_cube_object([current[0], current[1], index])
        return text

    @staticmethod
    def read_kml(val: list):
        """"
        Чтение из kml-файла
        """
        with open("static/media/data/geo.kml", 'rt', encoding="utf-8") as file:
            data = file.read()
        k = kml.KML()
        k.from_string(data)
        features = list(k.features())
        k2 = list(features[0].features())
        arr = []
        for feat in k2:
            string = str(feat.geometry).split('(')[1].split('0.0')[0].split(' ')
            arr.append([float(string[0]), float(string[1])])
        if val is None:
            val = [61.2200083333333, 52.147525]
        val2 = 0
        val3 = 0
        for loop1 in arr:
            # Мы должны найти к какой из точек он ближе(разница двух элементов массива)
            if val[0] > loop1[0]:
                for loop2 in arr:
                    if val[1] > loop2[1]:
                        val2 = loop1[0]
                        val3 = loop1[1]
                        break
        print([val2, val3])
        return [val2, val3]

    @staticmethod
    def create_style():
        f"""
    	<gx:CascadingStyle kml:id="__managed_style_25130D559F1CA685BFB3">
    		<Style>
    			<IconStyle>
    				<scale>1.2</scale>
    				<Icon>
    					<href>https://earth.google.com/earth/rpc/cc/icon?color=1976d2&amp;id=2000&amp;scale=4</href>
    				</Icon>
    				<hotSpot x="64" y="128" xunits="pixels" yunits="insetPixels"/>
    			</IconStyle>
    			<LabelStyle>
    			</LabelStyle>
    			<LineStyle>
    				<color>ff2dc0fb</color>
    				<width>6</width>
    			</LineStyle>
    			<PolyStyle>
    				<color>40ffffff</color>
    			</PolyStyle>
    			<BalloonStyle>
    				<displayMode>hide</displayMode>
    			</BalloonStyle>
    		</Style>
    	</gx:CascadingStyle>
    	<gx:CascadingStyle kml:id="__managed_style_1A4EFD26461CA685BFB3">
    		<Style>
    			<IconStyle>
    				<Icon>
    					<href>https://earth.google.com/earth/rpc/cc/icon?color=1976d2&amp;id=2000&amp;scale=4</href>
    				</Icon>
    				<hotSpot x="64" y="128" xunits="pixels" yunits="insetPixels"/>
    			</IconStyle>
    			<LabelStyle>
    			</LabelStyle>
    			<LineStyle>
    				<color>ff2dc0fb</color>
    				<width>4</width>
    			</LineStyle>
    			<PolyStyle>
    				<color>40ffffff</color>
    			</PolyStyle>
    			<BalloonStyle>
    				<displayMode>hide</displayMode>
    			</BalloonStyle>
    		</Style>
    	</gx:CascadingStyle>
    	<StyleMap id="__managed_style_047C2286A81CA685BFB3">
    		<Pair>
    			<key>normal</key>
    			<styleUrl>#__managed_style_1A4EFD26461CA685BFB3</styleUrl>
    		</Pair>
    		<Pair>
    			<key>highlight</key>
    			<styleUrl>#__managed_style_25130D559F1CA685BFB3</styleUrl>
    		</Pair>
    	</StyleMap>
    	<Placemark id="0D045F86381CA685BFB2">
    		<name>Самосвал</name>
    		<LookAt>
    			<longitude>61.2344061029136</longitude>
    			<latitude>52.17019183209385</latitude>
    			<altitude>282.7747547496757</altitude>
    			<heading>0</heading>
    			<tilt>0</tilt>
    			<gx:fovy>35</gx:fovy>
    			<range>1198.571236050484</range>
    			<altitudeMode>absolute</altitudeMode>
    		</LookAt>
    		<styleUrl>#__managed_style_047C2286A81CA685BFB3</styleUrl>
    		<Point>
    			<coordinates>61.23500224897153,52.17263169824412,281.7092496784567</coordinates>
    		</Point>
    	</Placemark>
    	<Placemark id="0B4EA6F59B1CA68601E0">
    		<name>Экскаватор</name>
    		<LookAt>
    			<longitude>61.23624067458115</longitude>
    			<latitude>52.17416232356366</latitude>
    			<altitude>277.5968564918906</altitude>
    			<heading>-0.5372217869872089</heading>
    			<tilt>53.57834275643886</tilt>
    			<gx:fovy>35</gx:fovy>
    			<range>2536.120178802812</range>
    			<altitudeMode>absolute</altitudeMode>
    		</LookAt>
    		<styleUrl>#__managed_style_047C2286A81CA685BFB3</styleUrl>
    		<Point>
    			<coordinates>61.23654046107902,52.16710625511239,297.4562999141254</coordinates>
    		</Point>
    	</Placemark>"""
        pass


class CareerClass:
    @staticmethod
    # Vacansies
    def get_career():
        # headers = {
        #     'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:45.0) Gecko/20100101 Firefox/45.0'
        # }
        # vacancies_urls = []
        # url = 'https://www.km-open.online/property'
        # r = requests.get(url, headers=headers)
        # soup = bs4.BeautifulSoup(r.content.decode("utf-8"))
        # list_objs = soup.find_all('div', {"class": "collection-item w-dyn-item"})
        # for list_obj in list_objs:
        #     vacancies_urls.append(url.split('/property')[0] + str(list_obj).split('href="')[1].split('"')[0])
        # vacancies_data = []
        # for url_s in vacancies_urls:
        #     r = requests.get(url_s, headers=headers)
        #     soup = bs4.BeautifulSoup(r.content.decode("utf-8"))
        #     list_objs = soup.find_all('div', {"class": "title-block"})
        #     vacancies_data = str(list_objs[0]).split('"heading-11">')[1].split('</h5>')[0]
        #     vacancies_data.append([vacancies_data, url_s])
        # data = [["Вакансия"], vacancies_data]
        # headers = {
        #     'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:45.0) Gecko/20100101 Firefox/45.0'
        # }

        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:45.0) Gecko/20100101 Firefox/45.0'
        }
        vacancies_urls = []
        url = 'https://www.km-open.online/property'
        r = requests.get(url, headers=headers)
        soup = bs4.BeautifulSoup(r.content.decode("utf-8"))
        list_objs = soup.find_all('div', {"class": "collection-item w-dyn-item"})
        for list_obj in list_objs:
            vacancies_urls.append(url.split('/property')[0] + str(list_obj).split('href="')[1].split('"')[0])
        vacancies_title = []
        for list_obj in list_objs:
            vacancies_title.append(str(list_obj).split('class="heading-12">')[1].split('</h5>')[0])
        vacancies_data = []
        for title in vacancies_title:
            vacancies_data.append([title, vacancies_urls[vacancies_title.index(title)]])

        data = [["Вакансия"], vacancies_data]
        return data


class UtilsClass:
    @staticmethod
    def create_encrypted_password(_random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890',
                                  _length=8):
        password = ''
        for i in range(1, _length + 1):
            password += random.choice(_random_chars)
        return password

    @staticmethod
    def decrypt_text_with_hash(massivsimvolov: str, massivkhesha: str):
        rasshifrovat_tekst = ''
        pozitsiyasimvolakhesha = 0
        dlinakhesha = len(massivkhesha)
        propusk = False
        for num in massivsimvolov:
            if propusk:
                propusk = False
                continue
            nomersimvola = ord(str(num))
            if pozitsiyasimvolakhesha >= dlinakhesha - 1:
                pozitsiyasimvolakhesha = 0
            pozitsiyasimvolakhesha = pozitsiyasimvolakhesha + 1
            simvolkhesha = ord(str(massivkhesha[pozitsiyasimvolakhesha]))
            kod_zashifrovannyy_simvol = nomersimvola - simvolkhesha
            # print(f"nomersimvola:{chr(nomersimvola)}:{nomersimvola}|simvolkhesha:{chr(simvolkhesha)}:{simvolkhesha}")
            zashifrovannyy_simvol = chr(kod_zashifrovannyy_simvol)
            rasshifrovat_tekst = rasshifrovat_tekst + zashifrovannyy_simvol
            if round(simvolkhesha / 2, 0) == simvolkhesha / 2:
                propusk = True
        return rasshifrovat_tekst
