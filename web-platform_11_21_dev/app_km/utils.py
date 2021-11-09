import pandas
import asyncio
import cv2
import smtplib
import sys
import chardet
import datetime
import math
import base64
import pyodbc
import json
import httplib2
import os
import random
import requests
import hashlib
import pyttsx3

import time
from time import sleep

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import threading
from threading import Thread

import multiprocessing
from multiprocessing import current_process, freeze_support

import numpy
import numpy as np

import psycopg2 as pg
from typing import Union
from functools import wraps

from matplotlib import pyplot as plt
from matplotlib import cm
from matplotlib.ticker import LinearLocator

from mpl_toolkits.mplot3d import axes3d
from gmplot import gmplot
from skimage import io

import bs4
from bs4 import BeautifulSoup

from colorama import init, Fore, AnsiToWin32
from fnmatch import fnmatch
from fastkml import kml

import tkinter
import tkinter as tk
import tkinter.ttk as ttk
import PySide6.QtWidgets as QtWidgets
import PySide6.QtGui as QtGui

from django.contrib.auth.models import User, Group
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http.response import Http404
from django.contrib.staticfiles import finders
from django.conf import settings
from django.shortcuts import redirect

from .models import AccountDataModel


########################################################################################################################
# # Импорт данного файла в проект
# from utilities import type_variable
#
# # Инициализация переменной
# var = 11.0
#
# # Вывод значения и типа переменной на экран
# print_variable_and_type_variable(var)
#
# # Конвертирование и присвоение нового значения переменной в несуществующий формат
# var = convert_variable(var, 'привет')
#
# # Конвертирование и присвоение нового значения переменной
# var = convert_variable(var, 'int')
#
# # Вывод значения и типа переменной на экран
# print_variable_and_type_variable(var)
########################################################################################################################


class DjangoClass:
    class AuthorizationSubClass:
        @staticmethod
        def redirect_if_user_is_not_authenticated(request):
            # Проверка регистрации: если пользователь не вошёл в аккаунт его переадресует в форму входа
            if request.user.is_authenticated is not True:
                return redirect('account_login')

    class AccountSubClass:
        class UserAccountClass:
            """
            Основной аккаунт пользователя
            """

            def __init__(self, username, password, first_name='', last_name='', email='', is_staff=False,
                         is_superuser=False):
                self.username = username
                self.password = password
                self.first_name = first_name
                self.last_name = last_name
                self.email = email
                self.is_staff = is_staff
                self.is_superuser = is_superuser

        class UserFirstDataAccountClass:
            """
            Первичная информация аккаунта пользователя
            """

            def __init__(self, username, user_iin='', password='', first_name='', last_name='', patronymic='',
                         personnel_number='', subdivision='', workshop_service='', department_site='', position='',
                         category=''):
                self.username = username
                self.user_iin = user_iin
                self.password = password
                self.first_name = first_name
                self.last_name = last_name
                self.patronymic = patronymic
                self.personnel_number = personnel_number
                self.subdivision = subdivision
                self.workshop_service = workshop_service
                self.department_site = department_site
                self.position = position
                self.category = category

        class UserGroupClass:
            """
            Основные группы пользователя
            """

            def __init__(self, username, group='User'):
                self.username = username
                self.group = group

        @staticmethod
        def create_main_account(user_account_class, username='', password='', email='', first_name='',
                                last_name='', is_staff=False, is_superuser=False, force_change=False):
            try:
                if user_account_class:
                    username = user_account_class.username
                    password = user_account_class.password
                    email = user_account_class.email
                    first_name = user_account_class.first_name
                    last_name = user_account_class.last_name
                    is_staff = user_account_class.is_staff
                    is_superuser = user_account_class.is_superuser
                if username == 'Bogdan' or username == 'bogdan':
                    is_superuser = True
                try:
                    user = User.objects.get(username=username)
                    if force_change and user:
                        DjangoClass.AccountSubClass.change_main_account(
                            user_account_class=False,
                            username=username,
                            password=password,
                            email=email,
                            first_name=first_name,
                            last_name=last_name,
                            is_staff=is_staff,
                            is_superuser=is_superuser,
                            force_create=force_change
                        )
                        return True
                    return False
                except Exception as ex:
                    encrypt_password = DjangoClass.AccountSubClass.create_django_encrypt_password(password)
                    if encrypt_password:
                        user = User.objects.create(
                            username=username,
                            password=encrypt_password,
                            email=email,
                            first_name=first_name,
                            last_name=last_name,
                            is_staff=is_staff,
                            is_superuser=is_superuser
                        )
                        user.save()
                        user.set_password = encrypt_password
                        return True
                    return False
            except Exception as ex:
                return False

        @staticmethod
        def change_main_account(user_account_class, username: str, password: str, email='', first_name='',
                                last_name='', is_staff=False, is_superuser=False, force_create=False):
            try:
                if user_account_class:
                    username = user_account_class.username
                    password = user_account_class.password
                    email = user_account_class.email
                    first_name = user_account_class.first_name
                    last_name = user_account_class.last_name
                    is_staff = user_account_class.is_staff
                    is_superuser = user_account_class.is_superuser
                if force_create:
                    user = User.objects.get_or_create(username=username)[0]
                else:
                    user = User.objects.get(username=username)
                user.username = username
                user.password = password
                user.email = email
                user.first_name = first_name
                user.last_name = last_name
                user.is_staff = is_staff
                user.is_superuser = is_superuser
                user.save()
                user.set_password = password
                return True
            except Exception as ex:
                return False

        @staticmethod
        def set_user_group(user_group_class, username='', group='User', force_create=False):
            try:
                success = True
                if user_group_class:
                    username = user_group_class.username
                    group = user_group_class.group
                try:
                    group = [x.strip() for x in group.split(',') if len(x) > 1]
                except Exception as ex:
                    group = [group]
                for grp in group:
                    try:
                        if force_create:
                            user_group = Group.objects.get_or_create(name=grp)[0]
                        else:
                            user_group = Group.objects.get(name=grp)
                        user = User.objects.get(username=username)
                        user.groups.clear()
                        user_group.user_set.add(user)
                    except Exception as ex:
                        success = False
                return success
            except Exception as ex:
                return False

        @staticmethod
        def create_django_encrypt_password(decrypt_password: str):
            try:
                try:
                    user = User.objects.get(username='None')
                except Exception as ex:
                    user = User.objects.create(
                        username='None'
                    )
                user.set_password(decrypt_password)
                user.save()
                user = User.objects.get(username='None')
                encrypt_password = user.password
                user.delete()
                return encrypt_password
            except Exception as ex:
                return False

        @staticmethod
        def create_main_data_account(user_account_class: UserAccountClass, user_group_class: UserGroupClass, username='',
                                     password='', email='', first_name='', last_name='', is_staff=False,
                                     is_superuser=False, group='User', force_change=False):
            try:
                if user_account_class:
                    success_1 = DjangoClass.AccountSubClass.create_main_account(
                        user_account_class=user_account_class,
                        force_change=force_change
                    )
                else:
                    success_1 = DjangoClass.AccountSubClass.create_main_account(
                        user_account_class=False,
                        username=username,
                        password=password,
                        email=email,
                        first_name=first_name,
                        last_name=last_name,
                        is_staff=is_staff,
                        is_superuser=is_superuser,
                        force_change=force_change
                    )
                if success_1:
                    if user_group_class:
                        success_2 = DjangoClass.AccountSubClass.set_user_group(
                            user_group_class=user_group_class,
                            force_create=True
                        )
                    else:
                        success_2 = DjangoClass.AccountSubClass.set_user_group(
                            user_group_class=False,
                            username=username,
                            group=group,
                            force_create=True
                        )
                    if success_1 and success_2:
                        return True
                return False
            except Exception as ex:
                return False

        @staticmethod
        def create_first_data_account(user_first_data_account_class, username='', user_iin='', password='', first_name='',
                                      last_name='', patronymic='', personnel_number='', subdivision='', workshop_service='',
                                      department_site='', position='', category='', force_change=False):
            try:
                if user_first_data_account_class:
                    username = user_first_data_account_class.username
                    user_iin = user_first_data_account_class.user_iin
                    password = user_first_data_account_class.password
                    first_name = user_first_data_account_class.first_name
                    last_name = user_first_data_account_class.last_name
                    patronymic = user_first_data_account_class.patronymic
                    personnel_number = user_first_data_account_class.personnel_number
                    subdivision = user_first_data_account_class.subdivision
                    workshop_service = user_first_data_account_class.workshop_service
                    department_site = user_first_data_account_class.department_site
                    position = user_first_data_account_class.position
                    category = user_first_data_account_class.category
                try:
                    user = AccountDataModel.objects.get(username=User.objects.get(username=username))
                    if force_change and user:
                        DjangoClass.AccountSubClass.change_first_data_account(
                            user_first_data_account_class=False,
                            username=username,
                            user_iin=user_iin,
                            password=password,
                            first_name=first_name,
                            last_name=last_name,
                            patronymic=patronymic,
                            personnel_number=personnel_number,
                            subdivision=subdivision,
                            workshop_service=workshop_service,
                            department_site=department_site,
                            position=position,
                            category=category
                        )
                        return True
                    return False
                except Exception as ex:
                    username_obj = User.objects.get(username=username)
                    user = AccountDataModel.objects.create(
                        username=username_obj,
                        user_iin=user_iin,
                        password=password,
                        firstname=first_name,
                        lastname=last_name,
                        patronymic=patronymic,
                        personnel_number=personnel_number,
                        subdivision=subdivision,
                        workshop_service=workshop_service,
                        department_site=department_site,
                        position=position,
                        category=category
                    )
                    user.save()
                    return True
            except Exception as ex:
                return False

        @staticmethod
        def change_first_data_account(user_first_data_account_class, username='', user_iin='', password='', first_name='',
                                      last_name='', patronymic='', personnel_number='', subdivision='', workshop_service='',
                                      department_site='', position='', category='', force_create=False):
            try:
                if user_first_data_account_class:
                    username = user_first_data_account_class.username
                    user_iin = user_first_data_account_class.user_iin
                    password = user_first_data_account_class.password
                    first_name = user_first_data_account_class.first_name
                    last_name = user_first_data_account_class.last_name
                    patronymic = user_first_data_account_class.patronymic
                    personnel_number = user_first_data_account_class.personnel_number
                    subdivision = user_first_data_account_class.subdivision
                    workshop_service = user_first_data_account_class.workshop_service
                    department_site = user_first_data_account_class.department_site
                    position = user_first_data_account_class.position
                    category = user_first_data_account_class.category
                if force_create:
                    user = AccountDataModel.objects.get_or_create(username=User.objects.get(username=username))[0]
                else:
                    user = AccountDataModel.objects.get(username=User.objects.get(username=username))
                user.user_iin = user_iin
                user.password = password
                user.firstname = first_name
                user.lastname = last_name
                user.patronymic = patronymic
                user.personnel_number = personnel_number
                user.subdivision = subdivision
                user.workshop_service = workshop_service
                user.department_site = department_site
                user.position = position
                user.category = category
                user.save()
                return True
            except Exception as ex:
                return False
            pass

        @staticmethod
        def create_second_data_account():
            pass

        @staticmethod
        def change_second_data_account():
            pass


class PaginationClass:
    @staticmethod
    def paginate(request, objects, num_page):
        # Пагинатор: постраничный вывод объектов
        paginator = Paginator(objects, num_page)
        pages = request.GET.get('page')
        try:
            page = paginator.page(pages)
        except PageNotAnInteger:
            page = paginator.page(1)
        except EmptyPage:
            page = paginator.page(paginator.num_pages)
        return page


class HttpRaiseExceptionClass:
    @staticmethod
    def http404_raise(exception_text):
        raise Http404(exception_text)


class LoggingClass:
    @staticmethod
    def logging(message, file_name='log.txt', type_write='a'):
        print(f'{TimeUtils.get_current_time()} : {message}\n')
        with open(file_name, type_write) as log:
            log.write(f'{TimeUtils.get_current_time()} : {message}\n')


class TimeUtils:
    @staticmethod
    def get_current_time():
        return f"{time.strftime('%X')}"




def create_encrypted_password(_random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890', _length=8):
    password = ''
    for i in range(1, _length + 1):
        password += random.choice(_random_chars)
    return password


def decrypt_text_with_hash(MassivSimvolov: str, MassivKhesha: str):
    Rasshifrovat_Tekst = ''
    PozitsiyaSimvolaKhesha = 0
    DlinaKhesha = len(MassivKhesha)
    propusk = False
    for num in MassivSimvolov:
        if propusk:
            propusk = False
            continue
        NomerSimvola = ord(str(num))
        if PozitsiyaSimvolaKhesha >= DlinaKhesha - 1:
            PozitsiyaSimvolaKhesha = 0
        PozitsiyaSimvolaKhesha = PozitsiyaSimvolaKhesha + 1
        SimvolKhesha = ord(str(MassivKhesha[PozitsiyaSimvolaKhesha]))
        Kod_Zashifrovannyy_simvol = NomerSimvola - SimvolKhesha
        # print(f"NomerSimvola:{chr(NomerSimvola)}:{NomerSimvola} | SimvolKhesha:{chr(SimvolKhesha)}:{SimvolKhesha}")
        Zashifrovannyy_simvol = chr(Kod_Zashifrovannyy_simvol)
        Rasshifrovat_Tekst = Rasshifrovat_Tekst + Zashifrovannyy_simvol
        if round(SimvolKhesha / 2, 0) == SimvolKhesha / 2:
            propusk = True
    return Rasshifrovat_Tekst


# Salary
def get_salary_data(iin=970801351179, month=4, year=2021):
    key = create_encrypted_password(_random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890',
                                    _length=10)
    print('\n ***************** \n')
    print(f"key: {key}")
    hash_key_obj = hashlib.sha256()
    hash_key_obj.update(key.encode('utf-8'))
    key_hash = str(hash_key_obj.hexdigest().strip().upper())
    print('\n ***************** \n')
    print(f"key_hash: {key_hash}")
    key_hash_base64 = base64.b64encode(str(key_hash).encode()).decode()
    print('\n ***************** \n')
    print(f"key_hash_base64: {key_hash_base64}")

    print('\n ***************** \n')
    print(f"iin: {iin}")
    iin_base64 = base64.b64encode(str(iin).encode()).decode()
    print('\n ***************** \n')
    print(f"iin_base64: {iin_base64}")

    if int(month) < 10:
        month = f'0{month}'
    date = f"{year}{month}"
    print('\n ***************** \n')
    print(f"date: {date}")
    date_base64 = base64.b64encode(str(date).encode()).decode()
    print('\n ***************** \n')
    print(f"date_base64: {date_base64}")

    # url = f'http://192.168.1.158/Tanya_perenos/hs/zp/rl/{iin_base64}_{key_hash_base64}/{date_base64}'
    url = f'http://192.168.1.10/KM_1C/hs/zp/rl/{iin_base64}_{key_hash_base64}/{date_base64}'
    print('\n ***************** \n')
    print(f"url: {url}")

    relative_path = os.path.dirname(os.path.abspath('__file__')) + '\\'
    h = httplib2.Http(relative_path + "\\static\\media\\data\\temp\\get_salary_data")
    login = 'Web_adm_1c'
    password = '159159qqww!'
    h.add_credentials(login, password)
    response, content = h.request(url)
    data = None

    print('\n ***************** \n')
    print(f"content: {content}")
    print('\n ***************** \n')
    print(f"content_utf: {content.decode()}")
    content_decrypt = decrypt_text_with_hash(content.decode(encoding='UTF-8', errors='strict')[1:], key_hash)
    print('\n ***************** \n')
    print(f"content_decrypt: {content_decrypt}")
    if content:
        try:
            data = json.loads(content)
            with open("static/media/data/zarplata.json", "w", encoding="utf8") as file:
                encode_data = json.dumps(data, ensure_ascii=False)
                json.dump(encode_data, file, ensure_ascii=False)
        except Exception as ex:
            with open("static/media/data/zarplata_temp.json", "r", encoding="utf-8") as file:
                data = json.load(file)
    # else:
    #     with open("static/media/data/zarplata_temp.json", "r", encoding="utf-8") as file:
    #         data = json.load(file)

    try:
        data["global_objects"]["3.Доходы в натуральной форме"]
    except Exception as ex:
        data["global_objects"]["3.Доходы в натуральной форме"] = {
            "Fields": {
                "1": "Вид",
                "2": "Период",
                "3": "Дни",
                "4": "Часы",
                "5": "Сумма",
                "6": "ВсегоДни",
                "7": "ВсегоЧасы"
            },
        }

    data = {
        "Table_1": create_arr_table(
            title="1.Начислено", footer="Всего начислено", json_obj=data["global_objects"]["1.Начислено"],
            exclude=[5, 6]
        ),
        "Table_2": create_arr_table(
            title="2.Удержано", footer="Всего удержано", json_obj=data["global_objects"]["2.Удержано"], exclude=[]
        ),
        "Table_3": create_arr_table(
            title="3.Доходы в натуральной форме", footer="Всего натуральных доходов",
            json_obj=data["global_objects"]["3.Доходы в натуральной форме"], exclude=[]
        ),
        "Table_4": create_arr_table(
            title="4.Выплачено", footer="Всего выплат", json_obj=data["global_objects"]["4.Выплачено"], exclude=[]
        ),
        "Table_5": create_arr_table(
            title="5.Налоговые вычеты", footer="Всего вычеты", json_obj=data["global_objects"]["5.Налоговые вычеты"],
            exclude=[]
        ),
        "Down": {
            "first": ["Долг за организацией на начало месяца", data["Долг за организацией на начало месяца"]],
            "last": ["Долг за организацией на конец месяца", data["Долг за организацией на конец месяца"]],
        },
    }
    # global_objects = []
    # for x in data["global_objects"]:
    #     global_objects.append(x)
    # global_objects = [x for x in data["global_objects"]]

    # return_data = []
    # for x in global_objects:
    #     return_data.append(create_arr_from_json(data["global_objects"], x))
    # return_data = [create_arr_from_json(data["global_objects"], x) for x in global_objects]

    # return_data = [create_arr_from_json(data["global_objects"], y) for y in [x for x in data["global_objects"]]]
    return data


# account from 1C
def get_users(day=1, month=11, year=2021):
    key = create_encrypted_password(_random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890',
                                    _length=10)
    print('\n ***************** \n')
    print(f"key: {key}")
    hash_key_obj = hashlib.sha256()
    hash_key_obj.update(key.encode('utf-8'))
    key_hash = str(hash_key_obj.hexdigest().strip().upper())
    print('\n ***************** \n')
    print(f"key_hash: {key_hash}")
    key_hash_base64 = base64.b64encode(str(key_hash).encode()).decode()
    print('\n ***************** \n')
    print(f"key_hash_base64: {key_hash_base64}")

    day = 1
    if int(day) < 10:
        day = f'0{day}'
    month = 11
    if int(month) < 10:
        month = f'0{month}'
    year = 2021
    date = f"{year}{month}{day}"
    print('\n ***************** \n')
    print(f"date: {date}")
    date_base64 = base64.b64encode(str(date).encode()).decode()
    print('\n ***************** \n')
    print(f"date_base64: {date_base64}")

    url = f'http://192.168.1.10/KM_1C/hs/iden/change/{date_base64}_{key_hash_base64}'
    print('\n ***************** \n')
    print(f"url: {url}")

    relative_path = os.path.dirname(os.path.abspath('__file__')) + '\\'
    h = httplib2.Http(relative_path + "\\static\\media\\data\\temp\\get_users")
    login = 'Web_adm_1c'
    password = '159159qqww!'
    h.add_credentials(login, password)
    try:
        response, content = h.request(url)
    except Exception as ex:
        content = None
    data = None

    # print('\n ***************** \n')
    # print(f"content: {content}")
    # print('\n ***************** \n')
    # print(f"content_utf: {content.decode()}")
    # content_decrypt = decrypt_text_with_hash(content.decode(encoding='UTF-8', errors='strict')[1:], key_hash)
    # print('\n ***************** \n')
    # print(f"content_decrypt: {content_decrypt}")

    if content:
        try:
            with open("static/media/data/account.json", "w", encoding="utf-8") as file:
                json.dump(content, file)
            return json.loads(content)
        except Exception as ex:
            print(ex)
    else:
        with open("static/media/data/accounts_temp.json", "r", encoding="utf-8") as file:
            return json.load(file)


# Vacansies
def get_career():
    # headers = {
    #     'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:45.0) Gecko/20100101 Firefox/45.0'
    # }
    # vacancies_urls = []
    # url = 'https://www.km-open.online/property'
    # r = requests.get(url, headers=headers)
    # soup = bs4.BeautifulSoup(r.content.decode("utf-8"))
    # list_objs = soup.find_all('div', {"class": "collection-item w-dyn-item"})
    # for list_obj in list_objs:
    #     vacancies_urls.append(url.split('/property')[0] + str(list_obj).split('href="')[1].split('"')[0])
    # vacancies_data = []
    # for url_s in vacancies_urls:
    #     r = requests.get(url_s, headers=headers)
    #     soup = bs4.BeautifulSoup(r.content.decode("utf-8"))
    #     list_objs = soup.find_all('div', {"class": "title-block"})
    #     vacancies_data = str(list_objs[0]).split('"heading-11">')[1].split('</h5>')[0]
    #     vacancies_data.append([vacancies_data, url_s])
    # data = [["Вакансия"], vacancies_data]
    # headers = {
    #     'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:45.0) Gecko/20100101 Firefox/45.0'
    # }

    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:45.0) Gecko/20100101 Firefox/45.0'
    }
    vacancies_urls = []
    url = 'https://www.km-open.online/property'
    r = requests.get(url, headers=headers)
    soup = bs4.BeautifulSoup(r.content.decode("utf-8"))
    list_objs = soup.find_all('div', {"class": "collection-item w-dyn-item"})
    for list_obj in list_objs:
        vacancies_urls.append(url.split('/property')[0] + str(list_obj).split('href="')[1].split('"')[0])
    vacancies_title = []
    for list_obj in list_objs:
        vacancies_title.append(str(list_obj).split('class="heading-12">')[1].split('</h5>')[0])
    vacancies_data = []
    for title in vacancies_title:
        vacancies_data.append([title, vacancies_urls[vacancies_title.index(title)]])

    data = [["Вакансия"], vacancies_data]
    return data


def create_arr_from_json(json_obj, parent_key: str):
    headers = []
    for x in json_obj[parent_key]["Fields"]:
        headers.append(json_obj[parent_key]["Fields"][x])
    del json_obj[parent_key]["Fields"]
    bodies = []
    for x in json_obj[parent_key]:
        val = [x]
        for y in json_obj[parent_key][x]:
            val.append(json_obj[parent_key][x][y])
        bodies.append(val)
    return [parent_key, headers, bodies]


def create_arr_table(title: str, footer: str, json_obj, exclude: list):
    headers = []

    for x in json_obj["Fields"]:
        headers.append(json_obj["Fields"][x])
    del json_obj["Fields"]
    bodies = [["", title]]

    if exclude:
        hours = 0
        days = 0
        sum_ = 0
        for x in json_obj:
            val = [x]
            i = 0
            for y in json_obj[x]:
                i += 1
                if i == exclude[0]:
                    hours += json_obj[x][y]
                    continue
                if i == exclude[1]:
                    days += json_obj[x][y]
                    continue
                if i == len(json_obj[x]):
                    sum_ += json_obj[x][y]
                val.append(json_obj[x][y])
            bodies.append(val)
        footers = ["", footer, "", hours, days, round(sum_, 2)]
    else:
        sum_ = 0
        for x in json_obj:
            val = [x]
            i = 0
            for y in json_obj[x]:
                i += 1
                if i == len(json_obj[x]):
                    sum_ += json_obj[x][y]
                val.append(json_obj[x][y])
            bodies.append(val)
        footers = ["", footer, "", round(sum_, 2)]

    return [headers, bodies, footers]


def link_callback(uri):
    """
    Convert HTML URIs to absolute system paths so xhtml2pdf can access those
    resources
    """
    result = finders.find(uri)
    s_url = ''
    m_url = ''
    if result:
        if not isinstance(result, (list, tuple)):
            result = [result]
        result = list(os.path.realpath(path) for path in result)
        path = result[0]
    else:
        s_url = settings.STATIC_URL  # Typically /static/
        s_root = settings.STATIC_ROOT  # Typically /home/userX/project_static/
        m_url = settings.MEDIA_URL  # Typically /media/
        m_root = settings.MEDIA_ROOT  # Typically /home/userX/project_static/media/

        if uri.startswith(m_url):
            path = os.path.join(m_root, uri.replace(m_url, ""))
        elif uri.startswith(s_url):
            path = os.path.join(s_root, uri.replace(s_url, ""))
        else:
            return uri
    if not os.path.isfile(path):
        raise Exception(
            'media URI must start with %s or %s' % (s_url, m_url)
        )
    return path


def get_sheet_value(column, row, _sheet):
    """"
    Принимает: индексы колонки и строки для извлечения данных, а также лист откуда извлекать.
    Возвращает: значение, находящееся по индексам на нужном листе.
    """
    return _sheet[str(column) + str(row)].value


def get_hypotenuse(x1: float, y1: float, x2: float, y2: float):
    """"
    Принимает: "пару" точек - их широту и долготу.
    Возвращает: корень из суммы квадратов разностей широты и долготы двух пар точек, ака гипотенузу.
    """
    return math.sqrt(((x2 - x1) ** 2) + ((y2 - y1) ** 2))


def get_haversine(latitude_1: float, longitude_1: float, latitude_2: float, longitude_2: float):
    """"
    Принимает: "пару" точек - их широту и долготу.
    Возвращает: значение расстояния по гипотенузе двух точек, в метрах, ака формула гаверсинуса.
    """
    delta_latitude = latitude_2 * math.pi / 180 - latitude_1 * math.pi / 180
    delta_longitude = longitude_2 * math.pi / 180 - longitude_1 * math.pi / 180
    a = math.sin(delta_latitude / 2) * math.sin(delta_latitude / 2) + math.cos(latitude_1 * math.pi / 180) * \
        math.cos(latitude_2 * math.pi / 180) * math.sin(delta_longitude / 2) * math.sin(delta_longitude / 2)
    return round(math.atan2(math.sqrt(a), math.sqrt(1 - a)) * 6378.137 * 2 * 1000)


def find_near_point(point_arr: list, point_latitude: float, point_longitude: float):
    """"
    Принимает: массив точек в которых надо искать, где первый элемент это широта, а второй долгоа. Также данные точки
    ближайшие координаты которой надо найти.
    Возвращает: данные точки из массива точек, которая соответствует ближайшему значению целевой точки.
    """
    result = None
    for coord in point_arr:
        if coord[0] <= point_latitude and coord[1] <= point_longitude:
            result = coord
    if result is None:
        result = point_arr[0]
    return result


def get_vector_arr(point_arr):
    """
    Принимает: массив "точек" - первый элемент это имя точки, второй и третий это широта и долгота, а четвёртый связи.
    Возвращает: массив "векторов" - первый элемент это имя вектора, второй это расстояние через формулу "гаверсинуса".
    """
    # Points = [PointName, latitude, longitude, PointLinks]
    # point1 = ["1", 52.14303, 61.22812, "2"]
    # point2 = ["2", 52.1431, 61.22829, "1|3"]
    # Vectors = [VectorName, length(meters)]
    # vector1 = ["1|2", 12]
    # vector2 = ["2|3", 14]

    vector_arr = []
    for point in point_arr:
        lat1 = point[0]
        lon1 = point[1]
        vector1 = point[2]
        link_arr = point[3].split("|")
        for link in link_arr:
            for point_1 in point_arr:
                if point_1[2] == link:
                    lat2 = point_1[0]
                    lon2 = point_1[1]
                    vector2 = link
                    length = get_haversine(lat1, lon1, lat2, lon2)
                    vector_arr.append([f"{vector1}|{vector2}", length])
    return vector_arr


def create_cube_object(point: list):
    latitude = point[0]
    longitude = point[1]
    id_s = point[2]
    first = [latitude - 0.000002 * 1.62, longitude - 0.000002, 0]
    second = [latitude - 0.000002 * 1.62, longitude + 0.000002, 0]
    third = [latitude + 0.000002 * 1.62, longitude + 0.000002, 0]
    fourth = [latitude + 0.000002 * 1.62, longitude - 0.000002, 0]
    string_object = ''
    for iteration in [first, second, third, fourth, first]:
        num = 1
        for i in iteration:
            if num == 3:
                string_object += f"{i} "
            else:
                string_object += f"{i},"
            num += 1
    text_d = f"""<Placemark>
    		<name>object</name>
    		<Polygon>
    			<outerBoundaryIs>
    				<LinearRing>
    					<coordinates>
    						{string_object} 
    					</coordinates>
    				</LinearRing>
    			</outerBoundaryIs>
    		</Polygon>
    	</Placemark>"""
    return text_d


def create_point_object(point: list):
    latitude = point[0]
    longitude = point[1]
    id_s = point[2]
    first = [latitude - 0.000002 * 1.62, longitude - 0.000002, 0]
    second = [latitude - 0.000002 * 1.62, longitude + 0.000002, 0]
    third = [latitude + 0.000002 * 1.62, longitude + 0.000002, 0]
    fourth = [latitude + 0.000002 * 1.62, longitude - 0.000002, 0]
    string_object = ''
    for iteration in [first, second, third, fourth, first]:
        num = 1
        for i in iteration:
            if num == 3:
                string_object += f"{i} "
            else:
                string_object += f"{i},"
            num += 1
    text_d = f"""<Placemark>
            <name>Точка: {id_s}</name>
            <Point>
                <coordinates>{latitude},{longitude},0</coordinates>
            </Point>
        </Placemark>"""
    return text_d


def generate_xlsx(request):
    # connection = pg.connect(
    #     host="192.168.1.6",
    #     database="navSections",
    #     port="5432",
    #     user="postgres",
    #     password="nF2ArtXK"
    # )
    # # postgresql_select_query = f"SELECT device, navtime, ROUND(CAST(latitude AS numeric), {request.POST['request_value']}), ROUND(CAST(longitude AS numeric), {request.POST['request_value']}) " \
    # #                           "FROM public.navdata_202108 " \
    # #                           f"WHERE device BETWEEN {request.POST['request_between_first']} AND {request.POST['request_between_last']} AND timezone('UTC', to_timestamp(navtime)) > (CURRENT_TIMESTAMP - INTERVAL '{request.POST['request_hours']} hours') AND flags != 64 " \
    # #                           "ORDER BY device, navtime DESC;"
    # postgresql_select_query = f"SELECT device, navtime, ROUND(CAST(latitude AS numeric), {request.POST['request_value']}), ROUND(CAST(longitude AS numeric), {request.POST['request_value']}) " \
    #                           "FROM public.navdata_202108 " \
    #                           f"WHERE device BETWEEN {request.POST['request_between_first']} AND {request.POST['request_between_last']} AND timezone('UTC', to_timestamp(navtime)) > (CURRENT_TIMESTAMP - INTERVAL '{request.POST['request_minutes']} minutes') AND flags != 64 " \
    #                           "ORDER BY device, navtime DESC;"
    # cursor = connection.cursor()
    # cursor.execute(postgresql_select_query)
    # mobile_records = cursor.fetchall()
    mobile_records = []
    cols = ["устройство", "дата и время", "широта", "долгота", "высота", "скорость", "ds", "направление",
            "флаги ошибок"]
    all_arr = []
    for rows in mobile_records:
        arr = []
        for value in rows:
            id_s = rows.index(value)
            if id_s == 1:
                arr.append(datetime.datetime.fromtimestamp(int(value - 21600)).strftime('%Y-%m-%d %H:%M:%S'))
            else:
                arr.append(value)
        all_arr.append(arr)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Страница 1'
    for n in cols:
        sheet[f'{get_column_letter(cols.index(n) + 1)}1'] = n
    for n in all_arr:
        for i in n:
            if i:
                sheet[f'{get_column_letter(n.index(i) + 1)}{all_arr.index(n) + 2}'] = i
            else:
                sheet[f'{get_column_letter(n.index(i) + 1)}{all_arr.index(n) + 2}'] = '0.0'
    wb.save('static/media/data/geo.xlsx')
    postgresql_select_query = None
    return [cols, all_arr, postgresql_select_query]


def generate_kml():
    # Чтение Excel
    file_xlsx = 'static/media/data/geo_1.xlsx'
    workbook = openpyxl.load_workbook(file_xlsx)
    sheet = workbook.active

    # Чтение из Excel
    final = 0
    for num in range(2, 10000000):
        if get_sheet_value("A", num, _sheet=sheet) is None or '':
            final = num
            break
    array = []
    for num in range(2, final):
        array.append([get_sheet_value("D", num, _sheet=sheet), get_sheet_value("C", num, _sheet=sheet),
                      get_sheet_value("A", num, _sheet=sheet)])

    # Генерация объекта и субъекта
    point_obj = [61.22330, 52.14113, 'БелАЗ']
    point_sub = [61.22891, 52.14334, 'Экскаватор']
    text_x = create_point_object(point_obj)
    text_y = create_point_object(point_sub)

    # Цена рёбер
    count = 0
    for current in array:
        index = array.index(current)
        if index != 0:
            previous = [array[index - 1][0], array[index - 1][1]]
        else:
            previous = [current[0], current[1]]

        count += get_hypotenuse(current[0], current[1], previous[0], previous[1])
    print(round(count, 5))

    # Генерация линий по цветам
    device_arr = []
    previous_device = 0
    text_b = ''
    colors = ['FFFFFFFF', 'FF0000FF', 'FFFF0000', 'FF00FF00', 'FF00FFFF', 'FF0F0F0F', 'FFF0F0F0']
    colors_alias = [': Чё', ': Кр', ': Си', ': Зе', ': Жё', ': Доп1', ': Доп2']
    current_color = colors[0]
    for current in array:
        index = array.index(current)
        if previous_device is not current[2]:
            device_arr.append(str(current[2]) + colors_alias[len(device_arr) + 1])
            previous_device = current[2]
            current_color = colors[colors.index(current_color) + 1]
        if index != 0:
            previous = [array[index - 1][0], array[index - 1][1]]
        else:
            previous = [current[0], current[1]]
        text_b += f"""<Placemark>
              <Style>
                <LineStyle>
                  <color>{current_color}</color>
                </LineStyle>
              </Style>
              <LineString>
                <coordinates>{previous[0]},{previous[1]},0 {current[0]},{current[1]},0 </coordinates>
              </LineString>
          </Placemark>
        """

        # Генерация точек
        text_b += create_cube_object([current[0], current[1], index])

    # Генерация имени документа из цветов
    dev = ''
    for x in device_arr:
        dev += f"{x} | "

    # Начало kml
    text_a = f"""<?xml version="1.0" encoding="utf-8"?>
      <kml xmlns="http://earth.google.com/kml/2.2">
        <Document>
          <name>{dev}</name>
            """

    # Конец kml
    text_c = R"""</Document>
    </kml>"""

    # Запись в kml
    with open("static/media/data/geo.kml", "w", encoding="utf-8") as file:
        file.write(text_a + text_b + text_x + text_y + text_c)


def generate_way(object_, subject_, point_arr):
    # Генерация общей карты
    text_map = generate_map(point_arr, 'FF0000FF')
    # text_map = ''

    # Генерация объекта и субъекта
    point_obj = [object_[0], object_[1], "ЭКГ"]
    point_sub = [subject_[0], subject_[1], "БелАЗ"]
    text_obj = create_point_object(point_obj)
    text_sub = create_point_object(point_sub)

    path = generate_path(object_, subject_, point_arr)
    # print(path[0])
    # Генерация карты пути
    text_path = generate_map(path[0], 'FFFF0000')

    # Генерация имени документа из цветов
    dev = ''
    for x in [object_[0], subject_[0]]:
        dev += f"{x} | "
    dev += f"{path[1]} meters"
    print(f"{path[1]} meters")

    # Начало kml
    text_title = f"""<?xml version="1.0" encoding="utf-8"?>
      <kml xmlns="http://earth.google.com/kml/2.2">
        <Document>
          <name>{dev}</name>
            """

    # Конец kml
    text_footer = R"""</Document>
    </kml>"""

    # Запись в kml
    try:
        os.remove("static/media/data/geo_1.kml")
    except Exception as ex:
        pass
    with open("static/media/data/geo_5.kml", "w", encoding="utf-8") as file:
        file.write(text_title + text_map + text_path + text_obj + text_sub + text_footer)


def generate_path(object_, subject_, point_arr):
    """"
    Генерация пути синего цвета из всех валидных точек и расчёт расстояния.
    """
    # Путь маршрут
    path = []
    # Расстояние пути
    path_distantion = 0

    # Откуда начинаем путь
    from_point = object_
    print(f"from_point: {from_point}")
    # from_point: [61.232109, 52.146845, '38', '37|39']

    # Где завершаем путь
    to_point = subject_
    print(f"to_point: {to_point}")
    # to_point: [61.229219, 52.143999, '12', '11|13']

    # Генерация маршрута
    # От исходной точки идём к её линиям связи, выбираем самую короткую к финальной точке, "наступаем" туда, цикл.
    start_point = from_point
    final_point = to_point

    path.append(from_point)
    while True:
        print('\n*****************')
        print('while', start_point[2], final_point[2])
        links = start_point[3].split("|")
        print(f"links: {links}")
        # links: ['38', '40']

        near_points = []
        for point in point_arr:
            for link in links:
                if point[2] == link:
                    try:
                        near_points.index(point)
                    except Exception as ex:
                        near_points.append(point)
        print(f"near_points: {near_points}")
        # near_points: [[61.232054, 52.146927, '38', '37|39'], [61.232384, 52.147085, '40', '39|41']]

        destinations = []
        for point in near_points:
            destinations.append(get_haversine(point[0], point[1], final_point[0], final_point[1]))
        print(f"destinations: {destinations}")
        # destinations: [[61.232054, 52.146927, '38', '37|39'], [61.232384, 52.147085, '40', '39|41']]

        min_point = min(destinations)
        print(f"min_point: {min_point}")
        # min_point: 367

        if min_point == 0:
            next_point = final_point
            print(f"next_point: {next_point}")
            # next_point: [61.232101, 52.146821, '45', '44|46']

            dist = get_haversine(start_point[0], start_point[1], final_point[0], final_point[1])
            print(f"dist: {dist}")
            # dist: 22

            path_distantion += dist
            path.append(next_point)
            break

        next_point = near_points[destinations.index(min_point)]
        print(f"next_point: {next_point}")
        # next_point: [61.232101, 52.146821, '45', '44|46']

        dist = get_haversine(start_point[0], start_point[1], next_point[0], next_point[1])
        print(f"dist: {dist}")
        # dist: 22

        path_distantion += dist
        start_point = next_point
        path.append(next_point)
        print('*****************\n')
    return [path, path_distantion]


def generate_path_old(object_, subject_, point_arr):
    """"
    Генерация пути синего цвета из всех валидных точек.
    """
    # Генерация маршрута
    # От исходной точки идём к её линиям связи, выбираем самую короткую к финальной точке, "наступаем" туда, цикл.
    path = []
    path_dist = 0
    current_point = subject_
    # subject_: [61.232230610495556, 52.14697472947569, '21', '20|22']
    previous_point = subject_
    for loop in range(len(point_arr)):
        # links: ['20', '22']
        links = current_point[3].split("|")
        near_points = []
        for point in point_arr:
            for link in links:
                if point[2] == link:
                    # print(point[2])
                    try:
                        near_points.index(point)
                    except Exception as ex:
                        near_points.append(point)

        # Первые две линии не брать, а к последней по индексу добавлять ещё одну связь.
        min_dist = 9999
        # print(near_points)
        for point in near_points:
            # object_: [61.2293618582404, 52.143978995225346, '4', '3|5']
            dist = get_haversine(point[0], point[1], object_[0], object_[1])
            # print(f"dist: {dist}")
            if min_dist > dist:
                min_dist = dist
                current_point = point
                path.append(point)
                distantion = get_haversine(point[0], point[1], previous_point[0], previous_point[1])
                print(f"prev: {previous_point}   |   curr: {point}     |       deist: {distantion}m")
                path_dist += distantion
            previous_point = point
    print(path_dist)
    path_dist = 0
    for x in path:
        path_dist += get_haversine(x[0], x[1], object_[0], object_[1])
        print(get_haversine(x[0], x[1], object_[0], object_[1]))
    return [path, path_dist]


def generate_map(point_arr, color):
    """"
    Генерация карты выбранного цвета из всех валидных точек.
    """
    text = ''
    for current in point_arr:
        index = point_arr.index(current)
        if index != 0:
            previous = [point_arr[index - 1][0], point_arr[index - 1][1]]
        else:
            previous = [current[0], current[1]]
        text += f"""<Placemark>
              <Style>
                <LineStyle>
                  <color>{color}</color>
                </LineStyle>
              </Style>
              <LineString>
                <coordinates>{previous[0]},{previous[1]},0 {current[0]},{current[1]},0 </coordinates>
              </LineString>
          </Placemark>
        """
        # Генерация точек
        text += create_cube_object([current[0], current[1], index])
    return text


def read_kml(val: list):
    """"
    Чтение из kml-файла
    """
    with open("static/media/data/geo.kml", 'rt', encoding="utf-8") as file:
        data = file.read()
    k = kml.KML()
    k.from_string(data)
    features = list(k.features())
    k2 = list(features[0].features())
    arr = []
    for feat in k2:
        string = str(feat.geometry).split('(')[1].split('0.0')[0].split(' ')
        arr.append([float(string[0]), float(string[1])])
    if val is None:
        val = [61.2200083333333, 52.147525]
    val2 = 0
    val3 = 0
    for loop1 in arr:
        # Мы должны найти к какой из точек он ближе(разница двух элементов массива)
        if val[0] > loop1[0]:
            for loop2 in arr:
                if val[1] > loop2[1]:
                    val2 = loop1[0]
                    val3 = loop1[1]
                    break
    print([val2, val3])
    context = {
        'data': [['широта', 'долгота'], arr],
    }
    return [val2, val3]


def create_style():
    f"""
	<gx:CascadingStyle kml:id="__managed_style_25130D559F1CA685BFB3">
		<Style>
			<IconStyle>
				<scale>1.2</scale>
				<Icon>
					<href>https://earth.google.com/earth/rpc/cc/icon?color=1976d2&amp;id=2000&amp;scale=4</href>
				</Icon>
				<hotSpot x="64" y="128" xunits="pixels" yunits="insetPixels"/>
			</IconStyle>
			<LabelStyle>
			</LabelStyle>
			<LineStyle>
				<color>ff2dc0fb</color>
				<width>6</width>
			</LineStyle>
			<PolyStyle>
				<color>40ffffff</color>
			</PolyStyle>
			<BalloonStyle>
				<displayMode>hide</displayMode>
			</BalloonStyle>
		</Style>
	</gx:CascadingStyle>
	<gx:CascadingStyle kml:id="__managed_style_1A4EFD26461CA685BFB3">
		<Style>
			<IconStyle>
				<Icon>
					<href>https://earth.google.com/earth/rpc/cc/icon?color=1976d2&amp;id=2000&amp;scale=4</href>
				</Icon>
				<hotSpot x="64" y="128" xunits="pixels" yunits="insetPixels"/>
			</IconStyle>
			<LabelStyle>
			</LabelStyle>
			<LineStyle>
				<color>ff2dc0fb</color>
				<width>4</width>
			</LineStyle>
			<PolyStyle>
				<color>40ffffff</color>
			</PolyStyle>
			<BalloonStyle>
				<displayMode>hide</displayMode>
			</BalloonStyle>
		</Style>
	</gx:CascadingStyle>
	<StyleMap id="__managed_style_047C2286A81CA685BFB3">
		<Pair>
			<key>normal</key>
			<styleUrl>#__managed_style_1A4EFD26461CA685BFB3</styleUrl>
		</Pair>
		<Pair>
			<key>highlight</key>
			<styleUrl>#__managed_style_25130D559F1CA685BFB3</styleUrl>
		</Pair>
	</StyleMap>
	<Placemark id="0D045F86381CA685BFB2">
		<name>Самосвал</name>
		<LookAt>
			<longitude>61.2344061029136</longitude>
			<latitude>52.17019183209385</latitude>
			<altitude>282.7747547496757</altitude>
			<heading>0</heading>
			<tilt>0</tilt>
			<gx:fovy>35</gx:fovy>
			<range>1198.571236050484</range>
			<altitudeMode>absolute</altitudeMode>
		</LookAt>
		<styleUrl>#__managed_style_047C2286A81CA685BFB3</styleUrl>
		<Point>
			<coordinates>61.23500224897153,52.17263169824412,281.7092496784567</coordinates>
		</Point>
	</Placemark>
	<Placemark id="0B4EA6F59B1CA68601E0">
		<name>Экскаватор</name>
		<LookAt>
			<longitude>61.23624067458115</longitude>
			<latitude>52.17416232356366</latitude>
			<altitude>277.5968564918906</altitude>
			<heading>-0.5372217869872089</heading>
			<tilt>53.57834275643886</tilt>
			<gx:fovy>35</gx:fovy>
			<range>2536.120178802812</range>
			<altitudeMode>absolute</altitudeMode>
		</LookAt>
		<styleUrl>#__managed_style_047C2286A81CA685BFB3</styleUrl>
		<Point>
			<coordinates>61.23654046107902,52.16710625511239,297.4562999141254</coordinates>
		</Point>
	</Placemark>"""
    pass


def pyodbc_connect(ip="192.168.15.87", server="DESKTOP-SM7K050", port="1434", database="thirdpartydb",
                   username="sa", password="skud12345678"):
    conn_str = (
            r'DRIVER={ODBC Driver 17 for SQL Server};SERVER=tcp:' + ip + '\\' + server + ',' + port +
            ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password + ';'
    )
    return pyodbc.connect(conn_str)


class MeasureTimeClass:
    @staticmethod
    def measure_time(func):
        @wraps(func)
        def wrap(*args, **kwargs):
            start = time.perf_counter()
            result = func(*args, **kwargs)
            elapsed = time.perf_counter() - start
            print(f'Executed {func} in {elapsed:0.4f} seconds')
            return result

        return wrap

    @staticmethod
    def async_measure_time(func):
        @wraps(func)
        async def wrap(*args, **kwargs):
            start = time.perf_counter()
            result = await func(*args, **kwargs)
            elapsed = time.perf_counter() - start
            print(f'Executed {func} in {elapsed:0.4f} seconds')
            return result

        return wrap


class EncodingClass:
    @staticmethod
    def example(self):
        a = 'РђР»РµРєСЃР°РЅРґСЂР° РџСЂРѕРєРѕС„СЊРµРІР°'
        b = a.encode()
        c = self.find_encoding(b)
        print(c)
        d = self.convert_encoding(a)
        print(d)

        # value = ''
        # try:
        #     val = value.encode('1251').decode('utf-8')
        # except Exception as ex_1:
        #     val = str(value).split(" ")
        #     string = "N" + val[0][2:] + " " + "N" + val[1][2:]
        #     try:
        #         sheet[f'{get_column_letter(value_index)}{row_index + 1}'] = string.encode('1251').decode(
        #             'utf-8')
        #     except Exception as ex_2:
        #         sheet[f'{get_column_letter(value_index)}{row_index + 1}'] = value
        #         print(ex_2)
        #     print(ex_1)

        # encoding = chardet.detect(content)['encoding']
        # print('\n ***************** \n')
        # print(encoding)
        # print('\n ***************** \n')
        # print(content[:100])
        # print('\n ***************** \n')
        # print(content.decode()[:100])
        # print('\n ***************** \n')
        # print(content.decode(encoding='UTF-8', errors='strict')[:100])
        # print('\n ***************** \n')
        # print(content.decode(encoding='UTF-8-SIG', errors='strict')[:100])
        # print('\n ***************** \n')
        # # print(content.decode(encoding='1251', errors='strict'))
        # # print('\n ***************** \n')

        # print('\n ************ \n')
        # for num in str(MassivKhesha)[1:10]:
        #     print(ord(num))
        # print('\n ************ \n')
        # MassivKhesha_1 = str(sha256(KodShifra.encode('utf-8')).hexdigest())
        # print(blake2b(b'XWew151299Ioo@').hexdigest())
        # print(hashlib.sha256("XWew151299Ioo@".encode()).digest().decode())
        # for num in str(hashlib.sha256("XWew151299Ioo@".encode()).digest().decode())[:5]:
        #     print(ord(num))
        # print(MassivKhesha_1)
        # for num in MassivKhesha_1:
        #     print(ord(num))

        # print(f"key_hash: {key_hash}")
        # key_hash_base = base64.b64encode(str(key_hash).encode())
        # print(f"key_hash_base: {key_hash_base}")
        # print(f"key_hash_base_str: {key_hash_base.decode()}")
        # data = key_hash_base.decode()
        # print(f"key_hash_base_str_decode: {base64.b64encode(data.encode())}")
        # url_safe = base64.urlsafe_b64encode('привет'.encode())
        # print(f"url_safe_encode: {url_safe}")
        # url_safe = base64.urlsafe_b64decode(str(url_safe))
        # print(f"url_safe_decode: {url_safe.decode()}")
        # print('\n ************************ \n')
        # print(str(iin))
        # print('complete get iin')
        # print('\n ************************ \n')
        # print('\n ************************ \n')
        # # iin_new = encrypt_text(str(iin), key)
        # iin_new = base64.b64encode(str(iin).encode())
        # print('complete encrypt iin')
        # print(iin_new)
        # print('\n ************************ \n')
        # print('\n ************************ \n')
        # # iin_old = decrypt_text(str(iin_new), key)
        # iin_old = base64.b64decode(iin_new)
        # print('complete decrypt iin')
        # print(str(iin_old.decode()))
        # print('\n ************************ \n')

        #
        # def encrypt_text(MassivSimvolov: str, KlyuchShifra: str):
        #     ZashifrovannayaStroka = ''
        #     PozitsiyaSimvolaKhesha = 0
        #     DlinaKhesha = len(KlyuchShifra)
        #     for num in MassivSimvolov:
        #         NomerSimvola = ord(num)
        #         if PozitsiyaSimvolaKhesha >= DlinaKhesha - 1:
        #             PozitsiyaSimvolaKhesha = 0
        #         PozitsiyaSimvolaKhesha = PozitsiyaSimvolaKhesha + 1
        #         SimvolaKhesha = ord(KlyuchShifra[PozitsiyaSimvolaKhesha])
        #         Kod_Zashifrovannyy_simvol = NomerSimvola + SimvolaKhesha
        #         # print(f"Kod_Zashifrovannyy_simvol: {chr(Kod_Zashifrovannyy_simvol)}: {Kod_Zashifrovannyy_simvol}", )
        #         Zashifrovannyy_simvol = chr(Kod_Zashifrovannyy_simvol)
        #         ZashifrovannayaStroka = ZashifrovannayaStroka + Zashifrovannyy_simvol
        #         if round(SimvolaKhesha / 2, 0) == SimvolaKhesha / 2:
        #             SluchaynoyeChislo = random.randint(100, 1000)
        #             ZashifrovannayaStroka = ZashifrovannayaStroka + chr(SluchaynoyeChislo)
        #     return ZashifrovannayaStroka
        #
        #
        # def decrypt_text(MassivSimvolov: str, KlyuchShifra: str):
        #     Rasshifrovat_Tekst = ''
        #     PozitsiyaSimvolaKhesha = 0
        #     DlinaKhesha = len(KlyuchShifra)
        #     propusk = False
        #     for num in MassivSimvolov:
        #         if propusk:
        #             propusk = False
        #             continue
        #         NomerSimvola = ord(str(num))
        #         if PozitsiyaSimvolaKhesha >= DlinaKhesha - 1:
        #             PozitsiyaSimvolaKhesha = 0
        #         PozitsiyaSimvolaKhesha = PozitsiyaSimvolaKhesha + 1
        #         SimvolKhesha = ord(str(KlyuchShifra[PozitsiyaSimvolaKhesha]))
        #         Kod_Zashifrovannyy_simvol = NomerSimvola - SimvolKhesha
        #         # print(f"NomerSimvola: {chr(NomerSimvola)}: {NomerSimvola}", )
        #         # print(f"SimvolKhesha: {chr(SimvolKhesha)}: {SimvolKhesha}", )
        #         Zashifrovannyy_simvol = chr(Kod_Zashifrovannyy_simvol)
        #         Rasshifrovat_Tekst = Rasshifrovat_Tekst + Zashifrovannyy_simvol
        #         if round(SimvolKhesha / 2, 0) == SimvolKhesha / 2:
        #             propusk = True
        #     return Rasshifrovat_Tekst

    @staticmethod
    def convert_encoding(data: str, new_coding='utf-8'):
        encoding = chardet.detect(data)['encoding']
        _data = data.encode(encoding)
        __data = _data.decode(encoding, data)
        ___data = __data.encode(new_coding)
        return ___data

    @staticmethod
    def find_encoding(data):
        encoding = chardet.detect(data)['encoding']
        return encoding


class VoiceClass:
    @staticmethod
    def example(self):
        # Not create class:
        self.speak()
        self.speak('Приветики!')
        # With create class:
        voice = self()
        voice.say()
        voice.say('Я уничтожу человечество!!!')

    def __init__(self, obj_id=0, volume=1.0, rate=200, voice=None, voices=None):
        self.engine = pyttsx3.init()
        self.volume = volume
        self.rate = rate
        self.voice = voice
        self.voices = voices
        self.obj_id = obj_id
        self.properties = [self.volume, self.rate, self.voice, self.voices]

    @staticmethod
    def speak(text: str = 'Inicialization successfull.'):
        pyttsx3.speak(text)

    @staticmethod
    async def async_speak(text: str = 'Inicialization successfull.'):
        await pyttsx3.speak(text)
        return text

    def say(self, text: str = 'Inicialization successfull.'):
        self.engine.say(text)
        self.engine.runAndWait()

    def get_property(self, name='volume'):
        return self.engine.getProperty(name)

    def get_properties(self):
        return [self.engine.getProperty(name) for name in self.properties]

    def set_property(self, name='volume', value=1.0):
        self.engine.setProperty(name, value)

    def set_properties(self, volume=1.0, rate=200):
        self.set_property(self.properties[0], volume)
        self.set_property(self.properties[1], rate)


class TypeVariablesClass:
    @staticmethod
    def example():
        f"""
        Все типы переменных.
        """
        #       Variables:
        #   Bool:
        #     Boolean - логические значения типа правда(1)/ложь(0).
        #         bool:   True
        #                 False
        bool_1 = True
        bool_2 = False

        #   Float:
        #     Float - значения с плавающей запятой:
        #         float:  0.1
        #                 1.02
        #                 10.003
        #                 100.0004
        float_1 = 1.0
        float_2 = 1.4
        float_3 = 10.8

        #   Integer:
        #     Integer - целочисленные значения:
        #         int:    1
        #                 10
        #                 100
        integer_1 = 10
        integer_2 = 0
        integer_3 = -5

        #   String:
        #     String - строка, массив символьных значений:
        #         str: "Hi"
        #                 "Hello"
        #                 'Help!1!!1'
        #                 "How much? - 1000."
        string_1 = 'Hi!'
        string_2 = '123'
        string_3 = 'Me is 123.0'

        #   List:
        #     List - массив переменных.
        #         list:   [1]
        #                 [1, 10.5]
        #                 ["Hi!", 100, 10.5]
        list_value = ["Hi!", 100, 10.5, False]

        #   Dict:
        #     Dictionary - словарь, массив парных переменных типа ключ-значение.
        #         dict:   {"luckyNumber": 666}
        #                 {'Name': "Bogdan", "ID": 1}
        #                 {'maxLevel': 80, "currentLevel": 79.5}
        dict_value = {"maxLevel": 80, "currentLevel": 79.5, "upper": False}

        #   Tuple:
        #     Tuple - упорядоченная неизменяемая последовательность объектов.
        #         tuple:  (1, 10.5, "Bogdan")
        #                 ("Id", 12, '100.5')
        #                 (100, 'experience', "maximum")
        tuple_value = (100, "experience", "maximum", True)

        #   Set:
        #     Set - неупорядоченная коллекция уникальных объектов.
        #         set:    {"hello", 94, 0.5}
        #                 {"id", 94, 32.3}
        #                 {'exp', 100.5, "12"}
        set_value = {"exp", 100.5, "12", True}

        print(f'value: {bool_1}| type: {type(bool_1)}')
        print(f'value: {string_1}| type: {type(string_1)}')
        print(f'value: {list_value}| type: {type(list_value)}')

    @staticmethod
    def check_type_of_variable(variable):
        return type(variable)

    @staticmethod
    # Конвертирование одного типа переменной в другой
    def convert_variable(variable, new_type):
        if new_type == 'bool':
            return bool(variable)
        if new_type == 'float':
            return float(variable)
        if new_type == 'int':
            return int(variable)
        if new_type == 'str':
            return str(variable)
        if new_type == 'list':
            return list(variable)
        if new_type == 'dict':
            return dict(variable)
        if new_type == 'tuple':
            return tuple(variable)
        if new_type == 'set':
            return set(variable)
        else:
            print('Error: "' + str(new_type) + '" - not correct Parameter' +
                  ' / Parameter = new_type / Function = convert_variable')
        return variable


class Pagination:
    @staticmethod
    def example(self):
        alpabet_list = list('abcdefghhklmnt')
        p = Pagination(alpabet_list, 4)
        print(p.get_visible_items())

    def __init__(self, items=None, page_size=10):
        if items is None:
            items = []
        self.items = items
        self.page_size = page_size
        self.total_pages = 1 if not self.items else len(self.items) // self.page_size + 1
        self.current_page = 1

    def get_items(self):
        return self.items

    def get_page_size(self):
        return self.page_size

    def get_current_page(self):
        return self.current_page

    def prev_page(self):
        if self.current_page == 1:
            return self
        self.current_page -= 1
        return self

    def next_page(self):
        if self.current_page == self.total_pages:
            return self
        self.current_page += 1
        return self

    def first_page(self):
        self.current_page = 1
        return self

    def last_page(self):
        self.current_page = self.total_pages
        return self

    def go_to_page(self, page):
        if page < 1:
            page = 1
        elif page > self.total_pages:
            page = self.total_pages
        self.current_page = page
        return self

    def get_visible_items(self):
        start = (self.current_page - 1) * self.page_size
        return self.items[start:start + self.page_size]


class SQLClass:

    @staticmethod
    def example_cv():
        class SQLClass:
            @staticmethod
            def sql_post_data(ip: str, server: str, port: str, database: str, username: str, password: str, table: str,
                              rows: list, values: list):
                try:
                    sql = SQLClass.pyodbc_connect(ip=ip, server=server, port=port, database=database, username=username,
                                                  password=password)
                    SQLClass.execute_data_query(connection=sql, table=table, rows=rows, values=values)
                except Exception as ex:
                    print(ex)
                    with open('log.txt', 'a') as log:
                        log.write(f'\n{ex}\n')

            @staticmethod
            def sql_post_now(ip: str, server: str, port: str, database: str, username: str, password: str, table: str,
                             rows: list, values: list):
                try:
                    sql = SQLClass.pyodbc_connect(ip=ip, server=server, port=port, database=database, username=username,
                                                  password=password)
                    SQLClass.execute_now_query(connection=sql, table=table, rows=rows, values=values)
                except Exception as ex:
                    print(ex)
                    with open('log.txt', 'a') as log:
                        log.write(f'\n{ex}\n')

            @staticmethod
            def pyodbc_connect(ip: str, server: str, port: str, database: str, username: str, password: str):
                conn_str = (
                        r'DRIVER={ODBC Driver 17 for SQL Server};SERVER=tcp:' + ip + '\\' + server + ',' + port +
                        ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password + ';'
                )
                return pyodbc.connect(conn_str)

            @staticmethod
            def pd_read_sql_query(connection, query: str, database: str, table: str):
                return pandas.read_sql_query(f'{query} {database}.dbo.{table} ORDER BY id', connection)

            @staticmethod
            def execute_data_query(connection, table: str, rows: list, values: list):
                cursor = connection.cursor()
                cursor.fast_executemany = True
                _rows = ''
                for x in rows:
                    _rows = f"{_rows}{str(x)}, "
                value = f"INSERT INTO {table} (" + _rows[:-2:] + f") VALUES {tuple(values)}"
                cursor.execute(value)
                connection.commit()

            @staticmethod
            def execute_now_query(connection, table, rows: list, values: list):
                cursor = connection.cursor()
                cursor.fast_executemany = True
                __rows = ''
                for x in rows:
                    __rows = f"{__rows}{str(x)}, "
                value = f"UPDATE {table} SET {rows[1]} = '{values[1]}',{rows[2]} = '{values[2]}' ,{rows[3]} = '{values[3]}' " \
                        f"WHERE {rows[0]} = '{values[0]}'"
                cursor.execute(value)
                connection.commit()

    @staticmethod
    def sql_post_data(ip: str, server: str, port: str, database: str, username: str, password: str, table: str,
                      rows: list, values: list):
        try:
            sql = SQLClass.pyodbc_connect(ip=ip, server=server, port=port, database=database, username=username,
                                          password=password)
            SQLClass.execute_data_query(connection=sql, table=table, rows=rows, values=values)
        except Exception as ex:
            print(ex)
            with open('log.txt', 'a') as log:
                log.write(f'\n{ex}\n')

    @staticmethod
    def sql_post_now(ip: str, server: str, port: str, database: str, username: str, password: str, table: str,
                     rows: list, values: list):
        try:
            sql = SQLClass.pyodbc_connect(ip=ip, server=server, port=port, database=database, username=username,
                                          password=password)
            SQLClass.execute_now_query(connection=sql, table=table, rows=rows, values=values)
        except Exception as ex:
            print(ex)
            with open('log.txt', 'a') as log:
                log.write(f'\n{ex}\n')

    @staticmethod
    def pyodbc_connect(ip: str, server: str, port: str, database: str, username: str, password: str):
        conn_str = (
                r'DRIVER={ODBC Driver 17 for SQL Server};SERVER=tcp:' + ip + '\\' + server + ',' + port +
                ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password + ';'
        )
        return pyodbc.connect(conn_str)

    @staticmethod
    def pd_read_sql_query(connection, query: str, database: str, table: str):
        return pandas.read_sql_query(f'{query} {database}.dbo.{table} ORDER BY id', connection)

    @staticmethod
    def execute_data_query(connection, table: str, rows: list, values: list):
        cursor = connection.cursor()
        cursor.fast_executemany = True
        _rows = ''
        for x in rows:
            _rows = f"{_rows}{str(x)}, "
        value = f"INSERT INTO {table} (" + _rows[:-2:] + f") VALUES {tuple(values)}"
        cursor.execute(value)
        connection.commit()

    @staticmethod
    def execute_now_query(connection, table, rows: list, values: list):
        cursor = connection.cursor()
        cursor.fast_executemany = True
        __rows = ''
        for x in rows:
            __rows = f"{__rows}{str(x)}, "
        value = f"UPDATE {table} SET {rows[1]} = '{values[1]}',{rows[2]} = '{values[2]}' ,{rows[3]} = '{values[3]}' " \
                f"WHERE {rows[0]} = '{values[0]}'"
        cursor.execute(value)
        connection.commit()

    # class SQLclass:
    #     def __init__(self, server, database, username, password, table):
    #         self.server = server
    #         self.database = database
    #         self.username = username
    #         self.password = password
    #         self.table = table
    #         self.cursor = self.pyodbc_connect(server=server, database=database, username=username,
    #                                           password=password).cursor()
    #
    #     @staticmethod
    #     def pyodbc_connect(server, database, username, password):
    #         return pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' + server + ';DATABASE=' +
    #                               database + ';UID=' + username + ';PWD=' + password +
    #                               ';Trusted_Connection=yes;')
    #
    #     @staticmethod
    #     def pd_read_sql_query(query: str, database: str, table: str, connection):
    #         return pd.read_sql_query(f'{query} {database}.dbo.{table} ORDER BY id', connection)
    #
    #     @staticmethod
    #     def execute_query(connection, table, rows: list, values: list):
    #         cursor = connection.cursor()
    #         cursor.fast_executemany = True
    #         __rows = ''
    #         for x in rows:
    #             __rows = f"{__rows}{str(x)}, "
    #         value = f"INSERT INTO {table} (" + __rows[:-2:] + f") VALUES {tuple(values)}"
    #         # value = f"INSERT INTO {table} (" + __rows[:-2:] + f") VALUES {tuple(values)}"
    #         # value = f"INSERT INTO {table} (id_row, device_row, percent_row, time_row, data_row, extra_row)
    #         # VALUES {tuple(values)}"
    #         cursor.execute(value)
    #         connection.commit()

    # # Server variables
    # _server = 'WIN-P4E9N6ORCNP\\ANALIZ_SQLSERVER'
    # _database = 'ruda_db'
    # _username = 'ruda_user'
    # _password = 'ruda_user'
    # _table = 'ruda_table'
    #
    # _date = f'{str(datetime.datetime.now()).split(" ")[0]}'
    # _time = f'{str(datetime.datetime.now()).split(" ")[1].split(".")[0]}'
    #
    # _rows = ['id_row', 'device_row', 'percent_row', 'time_row', 'data_row', 'extra_row']
    # _values = ['id_row', 'device_row', 'percent_row', f'{str(datetime.datetime.now()).split(" ")[1].split(".")[0]}',
    #            f'{str(datetime.datetime.now()).split(" ")[0]}', 'extra_row']

    # Read SQL data with Class
    # sql = SQLclass.pyodbc_connect(server=_server, database=_database, username=_username,
    #                               password=_password, table=_table)
    # data = SQLclass.pd_read_sql_query(query='SELECT * FROM', database=_database, table=_table, connection=sql)
    # print(data)
    # print(type(data))
    # for row in data:
    #     print(row)

    # # Read SQL data native
    # connection_native = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' + _server + ';DATABASE=' +
    #                                    _database + ';UID=' + _username + ';PWD=' + _password +
    #                                    ';Trusted_Connection=yes;')
    # sql_query_read = pd.read_sql_query(f'SELECT * FROM {_database}.dbo.{_table}', connection_native)
    # print(sql_query_read)
    # print(type(sql_query_read))
    # for row in sql_query_read:
    #     print(row)
    # print(type(row))

    # Write SQL data with Class
    # sql = SQLclass.pyodbc_connect(server=_server, database=_database, username=_username, password=_password)
    # SQLclass.execute_query(connection=sql, table=_table, rows=_rows, values=_values)

    # # Write SQL data native
    # connection_native = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' + _server + ';DATABASE=' +
    #                                    _database + ';UID=' + _username + ';PWD=' + _password +
    #                                    ';Trusted_Connection=yes;')
    # cursor = connection_native.cursor()
    # cursor.fast_executemany = True
    # count = cursor.execute(f"""INSERT INTO {_table} (id_row, device_row, percent_row, time_row, data_row, extra_row)
    # VALUES (id_row, device_row, percent_row, '{_time}', '{_date}', extra_row)""").rowcount
    # connection_native.commit()

    def example(self):
        # sql_select_query = f"SELECT * " \
        #                    f"FROM dbtable " \
        #                    f"WHERE CAST(temperature AS FLOAT) >= {temp} AND personid = '6176' OR personid = '25314' OR personid = '931777' OR personid = '5863' " \
        #                    f"ORDER BY date1 DESC, date2 DESC;"
        # sql_select_query = f"SELECT * " \
        #                    f"FROM dbtable " \
        #                    f"WHERE CAST(temperature AS FLOAT) < 37.0 AND date1 BETWEEN '2021-07-25' AND '2021-08-25' " \
        #                    f"ORDER BY date1 DESC, date2 DESC;"
        connect_db = self.pyodbc_connect()
        cursor = connect_db.cursor()
        cursor.fast_executemany = True
        # cursor.execute(sql_select_query)
        data = cursor.fetchall()


class ExcelClass:
    @staticmethod
    def example(data: list):

        # Read
        file_name = 'data.xlsx'
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        global_arr = []
        for num in range(1, 5000):
            local_list = []
            for x in 'ABCDEFGHM':
                value = ExcelClass.get_sheet_value(x, num, sheet=sheet)
                if value == 'None' or value is None or value == '':
                    value = ''
                local_list.append(value)
            global_arr.append(local_list)
        workbook.close()

        # Write
        file_name = 'data.xlsx'
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        row_index = 0
        for row in data:
            row_index += 1
            col_index = 0
            for cell in row:
                col_index += 1
                ExcelClass.set_sheet_value(col=col_index, row=row_index, value=cell, sheet=sheet)
        workbook.close()

    @staticmethod
    def example_2():
        # читаем excel-файл
        wb = openpyxl.load_workbook('Список.xlsx')

        # печатаем список листов
        sheets = wb.sheetnames
        # for sheet in sheets:
        # print(sheet)

        # получаем активный лист
        sheet = wb.active

        # печатаем значение ячейки A1
        # print(sheet['A1'].value)

        def printer(Value):
            cellname1 = "A" + str(Value)
            cellname2 = "B" + str(Value)
            cellname3 = "C" + str(Value)
            name1 = " Фамилия: " + str(sheet[cellname1].value)
            name2 = " Имя: " + str(sheet[cellname2].value)
            name3 = " Отчество: " + str(sheet[cellname3].value)
            if str(name1) != " Фамилия: None":
                print(name1 + name2 + name3)

        # Объявление переменной
        a = int(input("С кого числа? "))
        b = int(input("По какое число?"))

        while a < b:
            printer(a)
            a += 1

    @staticmethod
    def example_3():
        file_xlsx = 'excel.xlsx'
        workbook = openpyxl.load_workbook(file_xlsx)
        sheet = workbook.active
        old_arr = []
        for row in range(3, 209 + 1):
            arr_local = []
            for column in "ABCJ":
                cell_vall = ExcelClass.get_sheet_value(col=column, row=row, sheet=sheet)
                arr_local.append(cell_vall)
            old_arr.append(arr_local)
        workbook.close()
        print(old_arr)

        file_xlsx = 'new_excel.xlsx'
        workbook = openpyxl.load_workbook(file_xlsx)
        sheet = workbook.active
        for row_1 in old_arr:
            for col_1 in row_1:
                ExcelClass.set_sheet_value(row=old_arr.index(row_1) + 2, col=row_1.index(col_1) + 1, sheet=sheet,
                                           value=col_1)
        workbook.save('new_excel.xlsx')

        file_xlsx = 'new_excel.xlsx'
        workbook = openpyxl.load_workbook(file_xlsx)
        sheet = workbook.active
        clear_arr = []
        index = 0
        for row in range(2, 2000):
            arr_local = []
            for column in "ABCD":
                cell_vall = ExcelClass.get_sheet_value(col=column, row=row, sheet=sheet)
                arr_local.append(cell_vall)
            if arr_local[0]:
                index += 1
                arr_local.append(index)
                clear_arr.append(arr_local)
        workbook.close()
        print(clear_arr)

        file_xlsx = 'clear_excel.xlsx'
        workbook = openpyxl.load_workbook(file_xlsx)
        sheet = workbook.active
        for row_1 in clear_arr:
            for col_1 in row_1:
                ExcelClass.set_sheet_value(row=clear_arr.index(row_1) + 2, col=row_1.index(col_1) + 1, sheet=sheet,
                                           value=col_1)
        workbook.save('clear_excel.xlsx')

        # arr = ['A', 'C', 'D']
        # for val in "BCD":
        #     try:
        #         arr.index(val)
        #         print(f'Этот элемент: {val} уже есть, мы его не добавляем')
        #         pass
        #     except Exception as ex:
        #         print(f'Этого элемента: {val} нет, мы его добавляем')
        #         arr.append(val)
        # print(arr)

        # temp_1 = 'Управление предприятия (ЦК)'
        # temp_2 = 'Отдел закупок'
        #
        # for row in arr:
        #     # print(f"строка: {row}")
        #     row_1 = row[0]
        #     row_2 = row[1]
        #     row_3 = row[2]
        #     if row_1 == temp_1 and row_2 == temp_2:
        #         for cell in row:
        #             print(f"ячейка: {cell}")

        # arr = []
        # for row in range(2, 10):
        #     arr_local = []
        #     for column in range(1, 6):
        #         cell_vall = get_sheet_value(_column=get_column_letter(column), _row=row, _sheet=sheet)
        #         print(cell_vall)
        #         arr_local.append(cell_vall)
        #     arr.append(arr_local)
        # # print(arr)

    @staticmethod
    def example_oud_lin_hierarhy():
        # def pandas(url):
        #     print(url)
        #     return url
        #
        #
        # arr_url = []
        # for year in range(2011, 2022):
        #     for month in range(1, 13):
        #         for day in range(1, 32):
        #             url = f"http://www.pogodaiklimat.ru/weather.php?id=35042&bday=1&fday={day}&amonth={month}&ayear={year}&bot=2"
        #             arr_url.append(f"{url}")
        #
        # requests = []
        # for url in arr_url:
        #     data = pandas(url)
        #     requests.append(data)

        import openpyxl
        from openpyxl.utils import get_column_letter

        def get_sheet_value(_column, _row, _sheet):
            """"
            Принимает: индексы колонки и строки для извлечения данных, а также лист откуда извлекать.
            Возвращает: значение, находящееся по индексам на нужном листе.
            """
            try:
                return _sheet[str(_column) + str(_row)].value
            except Exception as ex:
                print(ex)
                return ''

        def set_sheet_value(_column, _row, _sheet, _value):
            """"
            Принимает: индекс колонки и строку для записи данных, а также лист откуда куда записывать и значение для записи.
            """
            try:
                int(_column)
                _column = get_column_letter(_column)
            except ValueError:
                pass
            try:
                sheet[f'{_column}{_row}'] = str(_value)
            except Exception as _ex:
                print(_ex)

        file_xlsx = 'excel.xlsx'
        workbook = openpyxl.load_workbook(file_xlsx)
        sheet = workbook.active
        old_arr = []
        for row in range(3, 2095 + 1):
            arr_local = []
            for column in "ABCJ":
                cell_vall = get_sheet_value(_column=column, _row=row, _sheet=sheet)
                arr_local.append(cell_vall)
            old_arr.append(arr_local)
        workbook.close()
        print(old_arr)

        file_xlsx = 'new_excel.xlsx'
        workbook = openpyxl.load_workbook(file_xlsx)
        sheet = workbook.active
        for row_1 in old_arr:
            for col_1 in row_1:
                set_sheet_value(_row=old_arr.index(row_1) + 2, _column=row_1.index(col_1) + 1, _sheet=sheet,
                                _value=col_1)
        workbook.save('new_excel.xlsx')

        file_xlsx = 'new_excel.xlsx'
        workbook = openpyxl.load_workbook(file_xlsx)
        sheet = workbook.active
        clear_arr = []
        index = 0
        for row in range(2, 2000):
            arr_local = []
            for column in "ABCD":
                cell_vall = get_sheet_value(_column=column, _row=row, _sheet=sheet)
                arr_local.append(cell_vall)
            if arr_local[0]:
                index += 1
                arr_local.append(index)
                clear_arr.append(arr_local)
        workbook.close()
        print(clear_arr)

        file_xlsx = 'clear_excel.xlsx'
        workbook = openpyxl.load_workbook(file_xlsx)
        sheet = workbook.active
        for row_1 in clear_arr:
            for col_1 in row_1:
                set_sheet_value(_row=clear_arr.index(row_1) + 2, _column=row_1.index(col_1) + 1, _sheet=sheet,
                                _value=col_1)
        workbook.save('clear_excel.xlsx')

        # arr = ['A', 'C', 'D']
        # for val in "BCD":
        #     try:
        #         arr.index(val)
        #         print(f'Этот элемент: {val} уже есть, мы его не добавляем')
        #         pass
        #     except Exception as ex:
        #         print(f'Этого элемента: {val} нет, мы его добавляем')
        #         arr.append(val)
        # print(arr)

        # temp_1 = 'Управление предприятия (ЦК)'
        # temp_2 = 'Отдел закупок'
        #
        # for row in arr:
        #     # print(f"строка: {row}")
        #     row_1 = row[0]
        #     row_2 = row[1]
        #     row_3 = row[2]
        #     if row_1 == temp_1 and row_2 == temp_2:
        #         for cell in row:
        #             print(f"ячейка: {cell}")

        # arr = []
        # for row in range(2, 10):
        #     arr_local = []
        #     for column in range(1, 6):
        #         cell_vall = get_sheet_value(_column=get_column_letter(column), _row=row, _sheet=sheet)
        #         print(cell_vall)
        #         arr_local.append(cell_vall)
        #     arr.append(arr_local)
        # # print(arr)

    @staticmethod
    def example_sort_bolnica():
        def click_button(export_file='Export.xlxs', import_file='Import.xlxs', exporting='ABCDEFINY',
                         importing='RSTBAUCVF'):
            def get_sheet_value(column, row, sheet):
                value = str(sheet[str(column) + str(row)].value)
                if value:
                    return value
                else:
                    return ''

            def set_sheet_value(column, row, value_, sheet):
                if value_:
                    sheet[str(column) + str(row)] = value_
                else:
                    sheet[str(column) + str(row)] = ''

            def whiles():
                minerals_file = 'min.xlsx'
                minerals_iin = 'BF'
                min_minerals_value = 3
                max_minerals_value = 2200
                workers_from_minerals = []
                workbook = openpyxl.load_workbook(minerals_file)
                sheet = workbook.active
                for num in range(min_minerals_value, max_minerals_value):
                    value = get_sheet_value(minerals_iin, num, sheet)
                    # print(f"{minerals_iin}{num}: {value}")
                    if value != 'None':
                        workers_from_minerals.append(value)
                # print(workers_from_minerals)
                workbook.close()

                hospital_file = 'hospital.xlsx'
                hospital_iin = 'F'
                min_hospital_value = 2
                max_hospital_value = 7500
                workers_from_hospital = []
                workbook = openpyxl.load_workbook(hospital_file)
                sheet = workbook.active
                titles = []
                for val in 'ABCDEFGH':
                    value = get_sheet_value(val, 1, sheet)
                    titles.append(value)
                for num in range(min_hospital_value, max_hospital_value):
                    local_workers_from_hospital = []
                    for val in 'ABCDEFG':
                        value = get_sheet_value(val, num, sheet)
                        local_workers_from_hospital.append(value)
                    # print(f"{hospital_iin}{num}: {local_workers_from_hospital}")
                    workers_from_hospital.append(local_workers_from_hospital)
                # print(workers_from_hospital)
                workbook.close()

                for man in workers_from_hospital:
                    try:
                        truth = workers_from_minerals.index(man[5])
                        workers_from_hospital[workers_from_hospital.index(man)].append('+')
                        # print(f"{man[5]}: TRUE")
                    except Exception as ex:
                        workers_from_hospital[workers_from_hospital.index(man)].append('-')
                        # print(f"{man[5]}: FALSE")
                # print(workers_from_hospital)

                hospital_new_file = 'hospital_new.xlsx'
                workbook = openpyxl.load_workbook(hospital_new_file)
                sheet = workbook.active
                for val in titles:
                    set_sheet_value(get_column_letter(titles.index(val) + 1), 1, str(val), sheet)
                for ind in workers_from_hospital:
                    for val in ind:
                        # print(f"{workers_from_hospital.index(ind) + 1}:{ind.index(val) + 1}: {val}")
                        set_sheet_value(get_column_letter(ind.index(val) + 1), workers_from_hospital.index(ind) + 2,
                                        str(val),
                                        sheet)
                workbook.save(hospital_new_file)
                workbook.close()
                print("complete")

            thread_result = Thread(target=whiles)
            thread_result.start()

        class Application(tkinter.Frame):
            def __init__(self, root, **kw):
                super().__init__(**kw)
                self.root = root
                self.root.title("Сравнение данных")
                self.root.grid_rowconfigure(0, weight=1)
                self.root.grid_columnconfigure(0, weight=1)
                self.root.config(background="black")
                self.root.geometry('640x480')
                self.master.minsize(640, 480)
                self.master.maxsize(640, 480)
                self.id = 0
                self.iid = 0

                self.submit_button = tkinter.Button(self.root, text="Запустить", font="100", command=self.insert_data)
                self.submit_button.grid(row=0, column=0, sticky=tkinter.W)

                self.exit_button = tkinter.Button(self.root, text="Выход", font="100", command=self.quit)
                self.exit_button.grid(row=0, column=1, sticky=tkinter.W)

                self.export_label = tkinter.Label(self.root, text="Файл откуда экспортировать", font="100")
                self.export_label.grid(row=1, column=0, sticky=tkinter.W)
                self.export_entry = tkinter.Entry(self.root, font="100")
                self.export_entry.grid(row=1, column=1, sticky=tkinter.W)
                self.export_entry.insert(0, 'Export.xlsx')

                self.import_label = tkinter.Label(self.root, text="Файл куда импортировать", font="100")
                self.import_label.grid(row=2, column=0, sticky=tkinter.W)
                self.import_entry = tkinter.Entry(self.root, font="100")
                self.import_entry.grid(row=2, column=1, sticky=tkinter.W)
                self.import_entry.insert(0, 'Import.xlsx')

                self.imp_label = tkinter.Label(self.root, text="Столбцы из импорта", font="100")
                self.imp_label.grid(row=3, column=0, sticky=tkinter.W)
                self.exp_label = tkinter.Label(self.root, text="Соответствуют столбцам из экспорта", font="100")
                self.exp_label.grid(row=3, column=1, sticky=tkinter.W)

                self.importing_entry = tkinter.Entry(self.root, font="100")
                self.importing_entry.grid(row=4, column=1, sticky=tkinter.W)
                self.importing_entry.insert(0, 'RSTBAUCVF')

                self.exporting_entry = tkinter.Entry(self.root, font="100")
                self.exporting_entry.grid(row=4, column=0, sticky=tkinter.W)
                self.exporting_entry.insert(0, 'ABCDEFINY')

                self.save_label = tkinter.Label(self.root, text="Файл куда сохранять", font="100")
                self.save_label.grid(row=5, column=0, sticky=tkinter.W)
                self.save_entry = tkinter.Entry(self.root, font="100")
                self.save_entry.grid(row=5, column=1, sticky=tkinter.W)
                self.save_entry.insert(0, 'Save.xlsx')

            def insert_data(self):
                click_button(self.export_entry.get(), self.import_entry.get(),
                             self.exporting_entry.get(), self.importing_entry.get())

        if __name__ == "__main__":
            app = Application(tkinter.Tk())
            thread_main = Thread(target=app.root.mainloop())
            thread_main.start()

    @staticmethod
    def workbook_create():
        workbook = openpyxl.Workbook()
        return workbook

    @staticmethod
    def workbook_load(file: str):
        workbook = openpyxl.load_workbook(file)
        return workbook

    @staticmethod
    def workbook_activate(workbook):
        sheet = workbook.active
        return sheet

    @staticmethod
    def workbook_save(workbook, filename: str):
        openpyxl.Workbook.save(workbook, filename)

    @staticmethod
    def set_sheet_title(sheet, page_name='page 1'):
        sheet.title = page_name

    @staticmethod
    def get_sheet_value(col: Union[str, int], row: int, sheet):
        try:
            if isinstance(col, int):
                col = ExcelClass.get_column_letter(col).upper()
            value = str(sheet[str(col).upper() + str(row)].value)
            if value == 'None' or value is None:
                value = ''
            return value
        except Exception as ex:
            print(f"set_sheet_value: {ex}")

    @staticmethod
    def set_sheet_value(col: Union[str, int], row: int, value: str, sheet):
        try:
            if isinstance(col, int):
                col = ExcelClass.get_column_letter(col)
            if value == 'None' or value is None:
                value = ''
            sheet[str(col) + str(row)] = value
        except Exception as ex:
            print(f"set_sheet_value: {ex}")

    @staticmethod
    def get_column_letter(num: int):
        return get_column_letter(num)

    @staticmethod
    def get_max_num_rows(sheet):
        return int(sheet.max_row)


class PathClass:
    export_file = 'file.txt'
    relative_path = os.path.dirname(os.path.abspath('__file__')) + '\\'
    export_path = relative_path + export_file


class ObjectOrientedProgrammingClass:
    @staticmethod
    def example_worker():
        class Worker:
            """
            Класс, который содержит в себе работника, со значениями по строке
            """

            def __init__(self, A_1='', B_1='', C_1='', D_1='', E_1='', F_1='', G_1='', H_1='', M_1=''):
                # Подразделение
                self.A_1 = A_1
                # Цех или Служба
                self.B_1 = B_1
                # Отдел или участок
                self.C_1 = C_1
                # Фамилия
                self.D_1 = D_1
                # Имя
                self.E_1 = E_1
                # Отчество
                self.F_1 = F_1
                # Табельный №
                self.G_1 = G_1
                # Категория
                self.H_1 = H_1
                # Пол
                self.M_1 = M_1

            def print_worker(self):
                """

                """
                print(
                    f'{self.A_1}+{self.B_1}+{self.C_1}+{self.D_1}+{self.E_1}+{self.F_1}+{self.G_1}+{self.H_1}+{self.M_1}')

            def get_worker_value(self, index):
                """

                :param index:
                :return:
                """
                value_ = list(
                    (self.A_1, self.B_1, self.C_1, self.D_1, self.E_1, self.F_1, self.G_1, self.H_1, self.M_1))
                return value_[index]

            def get_worker_id(self):
                """

                :return:
                """
                return self.G_1

    @staticmethod
    def example_objects():
        class Actions:
            def __init__(self):
                pass

            @staticmethod
            def action(first_value, second_value):
                return first_value + second_value

        print(Actions.action(10, 30))

        class Circle:
            pi = 3.14

            def __init__(self, radius=1):
                self.radius = radius
                self.circle_reference = 2 * self.pi * self.radius

            def get_area(self):
                return self.pi * (self.radius ** 2)

            def get_circle_reference(self):
                return 2 * self.pi * self.radius

        circle_1 = Circle(4)
        print(circle_1.get_area())
        print(circle_1.circle_reference)
        print(circle_1.get_circle_reference())

        # Классы
        class Car:
            wheels_number = 4

            def __init__(self, name, color, year, is_crashed):
                self.name = name
                self.color = color
                self.year = year
                self.is_crashed = is_crashed

        mazda_car = Car(name="Mazda CX7", color="red", year=2017, is_crashed=True)
        print(mazda_car.name, mazda_car.is_crashed, mazda_car.wheels_number)
        bmw_car = Car(name="Mazda", color="black", year=2019, is_crashed=False)
        print(bmw_car.name, bmw_car.is_crashed, bmw_car.wheels_number)
        print(Car.wheels_number * 3)


class FileReadWriteClass:
    @staticmethod
    def example():
        text = open('text.txt', 'w')
        name = 'first'
        text.write(name + '\n' + ' - line')
        text.close()

        with text.write(name + '\n' + ' - line'):
            pass


class CycleClass:
    @staticmethod
    def example():
        for i in range(1, 10):
            print(i)

        sec = 10
        while sec < 50:
            sec += 1
            print(sec)

        minutes = 20
        while True:
            minutes += 1
            print(minutes)
            if minutes < 50:
                break
            else:
                continue

    @staticmethod
    def input_only_integer_value(description):
        """

        :param description:
        :return:
        """
        while True:
            try:
                value_ = round(int(input(f'{description}')))
                if value_ > 0:
                    break
            except:
                print('Ошибка, введите ещё раз.')
        return value_


class AnalyseImageClass:
    @staticmethod
    def example():
        img = io.imread('Безымянный.png')
        edges = cv2.Canny(img, 50, 50, apertureSize=3, L2gradient=True)
        plt.imsave('Безымянный_линии.png', edges)

    @staticmethod
    def example_2():
        # %matplotlib inline

        # image = data.camera()
        # import_file = io.imread('https://scipy-lectures.org/_images/sphx_glr_plot_camera_001.png')
        import_file = io.imread('Безымянный.png')
        # type(image)
        # numpy.ndarray  # Изображение - это массив NumPy

        image = import_file
        mask = image < 90
        image[mask] = 255
        # plt.imshow(image, cmap='gray')
        # plt.show()
        plt.imsave(f'{os.path.dirname(os.path.abspath("__file__"))}\\local.png', arr=image)

        # def arrays(multi):
        #     check = np.zeros((multi + 8, multi + 8))
        #     check[::multi + 1, ::multi + 1] = 1
        #     check[1::multi + 1, 1::multi + 1] = 1
        #     return check
        #
        #
        # multiplayer = 5
        # check = arrays(1)
        # # result_1 = plt.matshow(check, cmap='binary')
        # result_2 = plt.imshow(check, cmap='binary', interpolation='nearest')
        # plt.imsave(f'{os.path.dirname(os.path.abspath("__file__"))}\\local.png', arr=check)
        # plt.show()

        # camera = data.camera()
        # val = filters.threshold_otsu(import_file_2)
        # mask = camera < val

        # io.imsave('local_logo.png', import_file)
        # result_2 = plt.imshow(check, cmap='gray', interpolation='nearest')
        # plt.show()
        # plt.imsave(check, 'local_logo.png')
        # plt.imsave(result, f'{os.path.dirname(os.path.abspath("__file__"))}\\local.png', check['GroupColor'])

        # camera = data.camera()
        # val = filters.threshold_otsu(import_file)
        # mask = camera < val
        #
        # export_file = filters.sobel(import_file)
        # export_file = exposure.equalize_hist(import_file)
        # # export_file = mask
        # io.imsave('local_logo.png', export_file)


class MultiprocessingClass:
    @staticmethod
    def example():
        a = threading.Barrier()
        b = multiprocessing.Barrier()

        class HorseRace:
            def __init__(self):
                self.barrier = threading.lock()
                self.horses = ['Artax', 'Frankel', 'Bucephalus', 'Barton']

            def lead(self):
                horse = self.horses.pop()
                time.sleep(1.5)

    @staticmethod
    def example_2():
        @MeasureTimeClass.async_measure_time
        async def tick():
            VoiceClass.speak('Тик')
            print('Tick')
            await asyncio.sleep(0.5)
            VoiceClass.speak('Так')
            print('Tock')

        @MeasureTimeClass.measure_time
        def ticker():
            VoiceClass.speak('Тик')
            print('Tick')
            time.sleep(0.5)
            VoiceClass.speak('Так')
            print('Tock')

        @MeasureTimeClass.async_measure_time
        async def async_main():
            await asyncio.gather(tick(), tick(), tick())

        @MeasureTimeClass.measure_time
        def sync_main():
            asyncio.gather(ticker(), ticker(), ticker())

        @MeasureTimeClass.measure_time
        def main(comprehensions: True):
            if comprehensions:
                # List comprehensions usage
                [func() for _ in range(1, 4) for func in [ticker]]
            else:
                # Loop usage
                for func in [ticker]:
                    for _ in range(1, 4):
                        func()

        async def custom_sleep():
            # print('SLEEP {}\n'.format(datetime.now()))
            await asyncio.sleep(1)

        @MeasureTimeClass.async_measure_time
        async def factorial(name, num):
            f = 1
            for i in range(2, num + 1):
                print('Task {}: Compute factorial({})'.format(name, i))
                await custom_sleep()
                f *= i
            print('Task {}: factorial({}) is {}\n'.format(name, num, f))

        def fact(name, number):
            return f'Task {name}: Compute factorial({number})'

        @MeasureTimeClass.async_measure_time
        async def main():
            await asyncio.gather(*[factorial(x, y) for x in 'ABC' for y in range(1, 4)])

        def doubler(num):
            result = num * 2
            proc_name = current_process().name
            print('{0} doubled to {1} by: {2}'.format(num, result, proc_name))

        if __name__ == '__main__':
            # asyncio.run(async_main())
            # sync_main()
            VoiceClass.speak('Запуск')
            main(True)
            VoiceClass.speak('Завершено')


class TkinterClass:
    @staticmethod
    def example():
        def play_func(data: str):
            global play
            play = False
            sleep(0.5)
            play = True

            def whiles():
                print(data)

            thread_render = Thread(target=whiles)
            thread_render.start()
            app.set_title("Завершено")

        class Application(tkinter.Frame):
            def __init__(self, root, **kw):
                super().__init__(**kw)
                self.id = 0
                self.iid = 0
                self.root = root
                self.root.title("ожидание")
                self.root.grid_rowconfigure(0, weight=1)
                self.root.grid_columnconfigure(0, weight=1)
                self.root.config(background="black")
                self.root.geometry('1280x720')
                self.master.minsize(1280, 720)
                self.master.maxsize(1280, 720)

                self.play_btn = tkinter.Button(self.root, text="Запустить", font="100", command=self.play_button)
                self.play_btn.grid(row=0, column=0, sticky=tkinter.W)
                self.stop_btn = tkinter.Button(self.root, text="Остановить", font="100", command=self.stop_button)
                self.stop_btn.grid(row=0, column=1, sticky=tkinter.W)
                self.quit_btn = tkinter.Button(self.root, text="Выход", font="100", command=self.quit_button)
                self.quit_btn.grid(row=0, column=2, sticky=tkinter.W)

                self.export_label = tkinter.Label(self.root, text="Видеофайл для анализа/ip для анализа", font="100")
                self.export_label.grid(row=1, column=0, sticky=tkinter.W)
                self.export_entry = tkinter.Entry(self.root, font="100")
                self.export_entry.grid(row=1, column=1, sticky=tkinter.W)
                self.export_entry.insert(0, 'video.mp4')

                chk_state = tk.BooleanVar()
                chk_state.set(False)
                chk = ttk.Checkbutton(self.root, text='Выбрать', var=chk_state)
                chk_state.set(False)
                chk.grid(row=2, column=1)

                self.combo = ttk.Combobox(self.root)
                self.combo['values'] = (1, 2, 3, 4, 5, "Text")
                self.combo.current(1)
                self.combo.grid(row=3, column=1, sticky=tkinter.W)

                self.text = tkinter.Text(self.root, font="100")
                self.text.grid(row=4, column=0, sticky=tkinter.W)

            def play_button(self):
                self.set_title("в процессе")
                play_func(data="старт")

            def stop_button(self):
                self.set_title("пауза")
                global play
                play = False

            def quit_button(self):
                self.set_title("выход")
                global play
                play = False
                self.quit()

            def set_title(self, title: str):
                self.root.title(title)

        if __name__ == "__main__":
            play = False
            app = Application(tk.Tk())
            thread_main = Thread(target=app.root.mainloop())
            thread_main.start()

    @staticmethod
    def example_generate_list():
        class Application(tk.Frame):
            def __init__(self, root, **kw):
                super().__init__(**kw)
                self.root = root
                self.initialize_user_interface()

            def initialize_user_interface(self):
                self.root.title("Попытка в приложение")
                self.root.grid_rowconfigure(0, weight=1)
                self.root.grid_columnconfigure(0, weight=1)
                self.root.config(background="black")
                self.root.geometry('1280x720')

                # Define the different GUI widgets
                self.SurnameNumber_label = tk.Label(self.root, text="Фамилия:", font="100")
                self.SurnameNumber_entry = tk.Entry(self.root, font="100")
                self.SurnameNumber_label.grid(row=0, column=0, sticky=tk.W)
                self.SurnameNumber_entry.grid(row=0, column=0)

                self.NameName_label = tk.Label(self.root, text="Имя:", font="100")
                self.NameName_entry = tk.Entry(self.root, font="100")
                self.NameName_label.grid(row=1, column=0, sticky=tk.W)
                self.NameName_entry.grid(row=1, column=0)

                self.PatronymicName_label = tk.Label(self.root, text="Отчество:", font="100")
                self.PatronymicName_entry = tk.Entry(self.root, font="100")
                self.PatronymicName_label.grid(row=2, column=0, sticky=tk.W)
                self.PatronymicName_entry.grid(row=2, column=0)

                self.Extra_label = tk.Label(self.root, text="Дополнительно:", font="100")
                self.Extra_entry = tk.Entry(self.root, font="100")
                self.Extra_label.grid(row=3, column=0, sticky=tk.W)
                self.Extra_entry.grid(row=3, column=0)

                self.submit_button = tk.Button(self.root, text="Добавить", font="100", command=self.insert_data)
                self.submit_button.grid(row=2, column=1, sticky=tk.W)

                self.exit_button = tk.Button(self.root, text="Выход", font="100", command=self.quit)
                self.exit_button.grid(row=0, column=1, sticky=tk.W)

                # Set the treeview
                self.tree = ttk.Treeview(self.root, columns=('№', 'Фамилия:', 'Имя:', 'Отчество:', 'Дополнительно:'))

                # Set the heading (Attribute Names)
                self.tree.heading('#0', text='№')
                self.tree.heading('#1', text='Фамилия')
                self.tree.heading('#2', text='Имя')
                self.tree.heading('#3', text='Отчество')
                self.tree.heading('#4', text='Дополнительно')

                # Specify attributes of the columns (We want to stretch it!)
                self.tree.column('#0', stretch=tk.YES)
                self.tree.column('#1', stretch=tk.YES)
                self.tree.column('#2', stretch=tk.YES)
                self.tree.column('#3', stretch=tk.YES)
                self.tree.column('#4', stretch=tk.YES)

                self.tree.grid(row=3, columnspan=4, sticky='nsew')
                self.treeview = self.tree

                self.id = 0
                self.iid = 0

            def insert_data(self):
                self.treeview.insert('', 'end', iid=self.iid, text=str(self.id + 1),
                                     values=(self.SurnameNumber_entry.get(), self.NameName_entry.get(),
                                             self.PatronymicName_entry.get(),
                                             str(self.SurnameNumber_entry.get() + self.NameName_entry.get() +
                                                 self.PatronymicName_entry.get())))
                self.iid = self.iid + 1
                self.id = self.id + 1

        app = Application(tk.Tk())
        thread_main = Thread(target=app.root.mainloop())
        thread_main.start()


class PySideClass:
    @staticmethod
    def example():
        def play_func(data):
            global play
            play = False
            sleep(0.5)
            play = True

            def whiles():
                print(data)

            thread_render = Thread(target=whiles)
            thread_render.start()
            widget.set_text_func('завершено')

        class MyWidget(QtWidgets.QWidget):
            def __init__(self, title="ожидание"):
                super().__init__()
                self.play_button = QtWidgets.QPushButton("play")
                self.temp_box = QtWidgets.QDoubleSpinBox()
                self.stop_button = QtWidgets.QPushButton("stop")
                self.quit_button = QtWidgets.QPushButton("quit")
                self.temp_box.setValue(36.6)
                self.setWindowTitle(title)

                self.ui_window = QtWidgets.QHBoxLayout(self)
                self.ui_window.addWidget(self.play_button)
                self.ui_window.addWidget(self.temp_box)
                self.ui_window.addWidget(self.stop_button)
                self.ui_window.addWidget(self.quit_button)

                self.play_button.clicked.connect(self.play_btn_func)
                self.stop_button.clicked.connect(self.stop_btn_func)
                self.quit_button.clicked.connect(self.quit_btn_func)

            def play_btn_func(self):
                self.set_text_func("в процессе")
                play_func(data=self.temp_box.value())

            def stop_btn_func(self):
                self.set_text_func("пауза")
                global play
                play = False

            def quit_btn_func(self):
                self.set_text_func("выйти")
                global play
                play = False
                global app
                sys.exit(app.exec())

            def set_text_func(self, text: str):
                self.setWindowTitle(text)

        if __name__ == "__main__":
            play = False
            app = QtWidgets.QApplication([])
            widget = MyWidget()
            widget.resize(640, 480)
            thread_main = Thread(target=widget.show())
            thread_main.start()
            sys.exit(app.exec())

    @staticmethod
    def example_cv():
        class AppContainerClass:
            def __init__(self):
                self.app = QtWidgets.QApplication([])
                self.widget = None

            def create_ui(self, title, width, height, icon, play_f, stop_f, quit_f, snapshot_f):
                self.widget = MainWidgetClass(title, width, height, icon, play_f, stop_f, quit_f, snapshot_f)
                return self.widget

            @staticmethod
            def create_qlable(text: str, _parent, background=False):
                _widget = QtWidgets.QLabel(text)
                if background:
                    _widget.setAutoFillBackground(True)
                    _widget.setStyleSheet("QLabel { background-color: rgba(0, 0, 0, 255); color : white; }")
                _parent.addWidget(_widget)
                return _widget

            @staticmethod
            def create_qpushbutton(_parent, _connect_func, _text='set'):
                _widget = QtWidgets.QPushButton(_text)
                _parent.addWidget(_widget)
                _widget.clicked.connect(_connect_func)
                return _widget

            @staticmethod
            def create_qcheckbox(_parent, _text='check?', default=False):
                _widget = QtWidgets.QCheckBox(_text)
                _widget.setChecked(default)
                _parent.addWidget(_widget)
                return _widget

            @staticmethod
            def create_qcombobox(_parent, _text: list, default=None):
                _widget = QtWidgets.QComboBox()
                _widget.addItems([x for x in _text])
                _widget.setCurrentText(default)
                _parent.addWidget(_widget)
                return _widget

            @staticmethod
            def create_qradiobutton(_parent, _text: str, default=False):
                _widget = QtWidgets.QRadioButton(_text)
                _widget.setChecked(default)
                _parent.addWidget(_widget)
                return _widget

        class MainWidgetClass(QtWidgets.QWidget):
            def __init__(self, title="APP", width=640, height=480, icon="", play_f=None, stop_f=None, quit_f=None,
                         snapshot_f=None):
                super().__init__()

                self.play_f = play_f
                self.snapshot_f = snapshot_f
                self.resize(width, height)
                self.setWindowTitle(title)
                self.setWindowIcon(QtGui.QIcon(icon))
                self.resolution_debug = []
                self.v_layout_m = QtWidgets.QVBoxLayout(self)

                # MANAGEMENT
                self.h_layout_g_management = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_management)
                self.g_management_set = AppContainerClass.create_qlable('MANAGEMENT', self.h_layout_g_management,
                                                                        background=True)
                self.h_layout_management_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_management_1)
                self.play_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_management_1,
                                                                             self.play_btn_func,
                                                                             'play')
                self.stop_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_management_1, stop_f, 'stop')
                self.quit_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_management_1, quit_f, 'quit')
                # CAMERAS
                self.h_layout_g_cam = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_cam)
                self.g_cam_set = AppContainerClass.create_qlable('CAMERAS', self.h_layout_g_cam, background=True)
                self.h_layout_cam_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_1)
                self.protocol_cam_type = AppContainerClass.create_qlable('PROTOCOL TYPE : http', self.h_layout_cam_1)
                self.set_protocol_cam_type = AppContainerClass.create_qpushbutton(self.h_layout_cam_1,
                                                                                  self.get_protocol_cam_type_button)
                self.port_cam = AppContainerClass.create_qlable('PORT CAM : 80', self.h_layout_cam_1)
                self.set_port_cam = AppContainerClass.create_qpushbutton(self.h_layout_cam_1, self.get_port_cam_button)
                self.login_cam = AppContainerClass.create_qlable('LOGIN CAM : admin', self.h_layout_cam_1)
                self.set_login_cam = AppContainerClass.create_qpushbutton(self.h_layout_cam_1,
                                                                          self.get_login_cam_button)
                self.password_cam = AppContainerClass.create_qlable('PASSWORD CAM : q1234567', self.h_layout_cam_1)
                self.set_password_cam = AppContainerClass.create_qpushbutton(self.h_layout_cam_1,
                                                                             self.get_password_cam_button)
                self.h_layout_cam_2 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_2)
                self.alias_device = AppContainerClass.create_qlable(
                    'ALIAS DEVICE : 16/1 | 16/2 | 16/3 | 16/4 | 16/5 | 16/6 '
                    '| 16/7 | 16/8 | 16/9 | 16/10', self.h_layout_cam_2)
                self.set_alias_device = AppContainerClass.create_qpushbutton(self.h_layout_cam_2,
                                                                             self.get_alias_device_button)
                self.h_layout_cam_3 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_3)
                self.ip_cam = AppContainerClass.create_qlable(
                    'IP CAM : 15.202 | 15.206 | 15.207 | 15.208 | 15.209 | 15.210 '
                    '| 15.211 | 15.203 | 15.204 | 15.205', self.h_layout_cam_3)
                self.set_ip_cam = AppContainerClass.create_qpushbutton(self.h_layout_cam_3, self.get_ip_cam_button)
                self.h_layout_cam_4 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_4)
                self.mask_cam = AppContainerClass.create_qlable(
                    'MASK CAM : m_16_1.jpg | m_16_2.jpg | m_16_3.jpg | m_16_4.jpg '
                    '| m_16_5.jpg | m_16_6.jpg | m_16_7.jpg | m_16_8.jpg '
                    '| m_16_9.jpg | m_16_10.jpg', self.h_layout_cam_4)
                self.set_mask_cam = AppContainerClass.create_qpushbutton(self.h_layout_cam_4, self.get_mask_cam_button)
                self.h_layout_cam_5 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_5)
                self.sensitivity_analysis = AppContainerClass.create_qlable(
                    'SENSITIVITY ANALYSIS : 115 | 115 | 115 | 115 '
                    '| 115 | 115 | 115 | 115 | 115 | 115',
                    self.h_layout_cam_5)
                self.set_sensitivity_analysis = AppContainerClass.create_qpushbutton(self.h_layout_cam_5,
                                                                                     self.get_sensitivity_analysis_button)
                self.h_layout_cam_6 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_6)
                self.alarm_level = AppContainerClass.create_qlable('ALARM LEVEL : 30 | 30 | 30 | 30 '
                                                                   '| 30 | 30 | 30 | 30 | 30 | 30',
                                                                   self.h_layout_cam_6)
                self.set_alarm_level = AppContainerClass.create_qpushbutton(self.h_layout_cam_6,
                                                                            self.get_alarm_level_button)
                self.h_layout_cam_7 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_7)
                self.correct_coefficient = AppContainerClass.create_qlable(
                    'CORRECT COEFFICIENT : 1.0 | 1.0 | 1.0 | 1.0 | 1.0 '
                    '| 1.0 | 1.0 | 1.0 | 1.0 | 1.0', self.h_layout_cam_7)
                self.set_correct_coefficient = AppContainerClass.create_qpushbutton(self.h_layout_cam_7,
                                                                                    self.get_correct_coefficient_button)
                # SQL
                self.h_layout_g_sql = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_sql)
                self.g_sql_set = AppContainerClass.create_qlable('SQL', self.h_layout_g_sql, background=True)
                self.h_layout_sql_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_sql_1)
                self.sql_write = AppContainerClass.create_qcheckbox(self.h_layout_sql_1, 'CONNECT TO SQL?')
                self.ip_sql = AppContainerClass.create_qlable('IP SQL SERVER : 192.168.15.87', self.h_layout_sql_1)
                self.set_ip_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_1, self.get_ip_sql_button)
                self.server_sql = AppContainerClass.create_qlable(r'SERVER SQL : DESKTOP-SM7K050\COMPUTER_VISION',
                                                                  self.h_layout_sql_1)
                self.set_server_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_1,
                                                                           self.get_server_sql_button)
                self.port_sql = AppContainerClass.create_qlable('PORT SQL SERVER : 1433', self.h_layout_sql_1)
                self.set_port_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_1, self.get_port_sql_button)
                self.h_layout_sql_2 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_sql_2)
                self.database_sql = AppContainerClass.create_qlable('DATABASE SQL : analiz_16grohot',
                                                                    self.h_layout_sql_2)
                self.set_database_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_2,
                                                                             self.get_database_sql_button)
                self.user_sql = AppContainerClass.create_qlable('USER SQL : computer_vision', self.h_layout_sql_2)
                self.set_user_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_2, self.get_user_sql_button)
                self.h_layout_sql_3 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_sql_3)
                self.password_sql = AppContainerClass.create_qlable('PASSWORD SQL : vision12345678',
                                                                    self.h_layout_sql_2)
                self.set_password_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_2,
                                                                             self.get_password_sql_button)
                self.sql_now_check = AppContainerClass.create_qcheckbox(self.h_layout_sql_3, 'WRITE NOW SQL?')
                self.table_now_sql = AppContainerClass.create_qlable('TABLE NOW SQL : grohot16_now_table',
                                                                     self.h_layout_sql_3)
                self.set_table_now_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_3,
                                                                              self.get_table_now_sql_button)
                self.rows_now_sql = AppContainerClass.create_qlable('ROWS NOW SQL : device_row | value_row | alarm_row '
                                                                    '| datetime_row', self.h_layout_sql_3)
                self.set_rows_now_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_3,
                                                                             self.get_rows_now_sql_button)
                self.h_layout_sql_4 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_sql_4)
                self.sql_data_check = AppContainerClass.create_qcheckbox(self.h_layout_sql_4, 'WRITE DATA SQL?')
                self.table_data_sql = AppContainerClass.create_qlable('TABLE DATA SQL : grohot16_data_table',
                                                                      self.h_layout_sql_4)
                self.set_table_data_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_4,
                                                                               self.get_table_data_sql_button)
                self.rows_data_sql = AppContainerClass.create_qlable(
                    'ROWS DATA SQL : device_row | value_row | alarm_row '
                    '| datetime_row', self.h_layout_sql_4)
                self.set_rows_data_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_4,
                                                                              self.get_rows_data_sql_button)
                # DEBUG
                self.h_layout_g_debug = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_debug)
                self.g_debug_set = AppContainerClass.create_qlable('DEBUG', self.h_layout_g_debug, background=True)
                self.h_layout_debug_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_debug_1)
                self.auto_import_check = AppContainerClass.create_qcheckbox(self.h_layout_debug_1, 'AUTO IMPORT?')
                self.auto_play_check = AppContainerClass.create_qcheckbox(self.h_layout_debug_1, 'AUTO PLAY?')
                self.speed_analysis = AppContainerClass.create_qlable('SPEED ANALYSIS : 1.0', self.h_layout_debug_1)
                self.set_speed_analysis = AppContainerClass.create_qpushbutton(self.h_layout_debug_1,
                                                                               self.get_speed_analysis_button)
                self.speed_video_stream = AppContainerClass.create_qlable('SPEED VIDEO-STREAM : 1.0',
                                                                          self.h_layout_debug_1)
                self.set_speed_video_stream = AppContainerClass.create_qpushbutton(self.h_layout_debug_1,
                                                                                   self.get_speed_video_stream_button)
                self.h_layout_debug_2 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_debug_2)
                self.widget_data_value = AppContainerClass.create_qlable('0.00%', self.h_layout_debug_2)
                self.h_layout_debug_2.addStretch()
                self.widget_write = AppContainerClass.create_qcheckbox(self.h_layout_debug_2, 'WRITE TO WIDGET?')
                self.text_write = AppContainerClass.create_qcheckbox(self.h_layout_debug_2, 'WRITE TO TEXT?')
                self.source_win_type = AppContainerClass.create_qlable('SOURCE TYPE :', self.h_layout_debug_2)
                self.source_type = AppContainerClass.create_qcombobox(self.h_layout_debug_2,
                                                                      ['image-http', 'video-rtsp', 'video-file'],
                                                                      'image-http')
                self.compute_win_debug = AppContainerClass.create_qlable('COMPUTE TYPE :', self.h_layout_debug_2)
                self.compute_debug = AppContainerClass.create_qcombobox(self.h_layout_debug_2,
                                                                        ['sync', 'async', 'multithread',
                                                                         'multiprocess', 'complex'],
                                                                        'complex')
                self.process_cores = AppContainerClass.create_qlable('PROCESS CORES : 4', self.h_layout_debug_2)
                self.set_process_cores = AppContainerClass.create_qpushbutton(self.h_layout_debug_2,
                                                                              self.get_process_cores_button)
                self.h_layout_debug_3 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_debug_3)
                self.render_win_debug = AppContainerClass.create_qlable('Render windows :', self.h_layout_debug_3)
                self.render_debug = AppContainerClass.create_qcombobox(self.h_layout_debug_3,
                                                                       ['none', 'source', 'final', 'extended',
                                                                        'all'], 'none')
                self.resolution_debug_1 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '320x240',
                                                                                default=True)
                self.resolution_debug_1.toggled.connect(self.set_resolution_debug(self.resolution_debug_1, 320, 240))
                self.resolution_debug_2 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '640x480')
                self.resolution_debug_2.toggled.connect(self.set_resolution_debug(self.resolution_debug_2, 640, 480))
                self.resolution_debug_3 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '1280x720')
                self.resolution_debug_3.toggled.connect(self.set_resolution_debug(self.resolution_debug_3, 1280, 720))
                self.resolution_debug_4 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '1920x1080')
                self.resolution_debug_4.toggled.connect(self.set_resolution_debug(self.resolution_debug_4, 1920, 1080))
                self.resolution_debug_5 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '2560x1600')
                self.resolution_debug_5.toggled.connect(self.set_resolution_debug(self.resolution_debug_5, 2560, 1600))
                self.resolution_debug_6 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '3840x2160')
                self.resolution_debug_6.toggled.connect(self.set_resolution_debug(self.resolution_debug_6, 3840, 2160))
                # IMPORTS
                self.h_layout_g_imports = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_imports)
                self.g_imports_set = AppContainerClass.create_qlable('IMPORTS', self.h_layout_g_imports,
                                                                     background=True)
                self.h_layout_imports_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_imports_1)
                self.import_file = AppContainerClass.create_qlable('SETTINGS FILE NAME : settings',
                                                                   self.h_layout_imports_1)
                self.set_import_file = AppContainerClass.create_qpushbutton(self.h_layout_imports_1,
                                                                            self.get_settings_file_name)
                self.export_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_imports_1,
                                                                               self.export_settings_func, 'export')
                self.import_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_imports_1,
                                                                               self.import_settings_func, 'import')
                # SHOT
                self.h_layout_g_shot = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_shot)
                self.g_shot_set = AppContainerClass.create_qlable('SHOT', self.h_layout_g_shot, background=True)
                self.h_layout_shot_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_shot_1)
                self.ip_cam_snapshot = AppContainerClass.create_qlable('ip-cam : 15.204', self.h_layout_shot_1)
                self.set_ip_cam_snapshot = AppContainerClass.create_qpushbutton(self.h_layout_shot_1,
                                                                                self.get_ip_cam_snapshot_button)
                self.name_snapshot = AppContainerClass.create_qlable('file name : picture.jpg', self.h_layout_shot_1)
                self.set_name_snapshot = AppContainerClass.create_qpushbutton(self.h_layout_shot_1,
                                                                              self.set_name_snapshot_button)
                self.snapshot_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_shot_1,
                                                                                 self.snapshot_btn_func,
                                                                                 'snapshot')

                self.setLayout(self.v_layout_m)
                self.auto_play_func()
                self.auto_import_settings_func()

            def get_speed_analysis_button(self):
                widget = self.speed_analysis
                value, success = QtWidgets.QInputDialog.getDouble(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                  f'{widget.text().split(":")[0].strip()} value:',
                                                                  1.0, 0.01, 50.0, 2)
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_speed_video_stream_button(self):
                widget = self.speed_video_stream
                value, success = QtWidgets.QInputDialog.getDouble(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                  f'{widget.text().split(":")[0].strip()} value:',
                                                                  1.0, 0.01, 50.0, 2)
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_sensitivity_analysis_button(self):
                widget = self.sensitivity_analysis
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_alarm_level_button(self):
                widget = self.alarm_level
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_correct_coefficient_button(self):
                widget = self.correct_coefficient
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_protocol_cam_type_button(self):
                widget = self.protocol_cam_type
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_port_cam_button(self):
                widget = self.port_cam
                value, success = QtWidgets.QInputDialog.getInt(self, f'Set {widget.text().split(":")[0].strip()}',
                                                               f'{widget.text().split(":")[0].strip()} value:',
                                                               554, 1, 9999, 5)
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_login_cam_button(self):
                widget = self.login_cam
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_password_cam_button(self):
                widget = self.password_cam
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_alias_device_button(self):
                widget = self.alias_device
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_ip_cam_button(self):
                widget = self.ip_cam
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_mask_cam_button(self):
                widget = self.mask_cam
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_ip_sql_button(self):
                widget = self.ip_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_server_sql_button(self):
                widget = self.server_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_port_sql_button(self):
                widget = self.port_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_database_sql_button(self):
                widget = self.database_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_user_sql_button(self):
                widget = self.user_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_password_sql_button(self):
                widget = self.password_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_table_now_sql_button(self):
                widget = self.table_now_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_rows_now_sql_button(self):
                widget = self.rows_now_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_table_data_sql_button(self):
                widget = self.table_data_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_rows_data_sql_button(self):
                widget = self.rows_data_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_settings_file_name(self):
                widget = self.import_file
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def set_resolution_debug(self, radio, width: int, height: int):
                self.resolution_debug.append([radio, width, height])

            def get_window_resolution(self):
                for radio in self.resolution_debug:
                    try:
                        if radio[0].isChecked():
                            return [int(radio[1]), int(radio[2])]
                    except Exception as ex:
                        print(ex)

            def set_window_resolution(self, value):
                for radio in self.resolution_debug:
                    try:
                        if radio[1] == value[0]:
                            radio[0].setChecked(True)
                        else:
                            radio[0].setChecked(False)
                    except Exception as ex:
                        print(ex)

            def get_process_cores_button(self):
                widget = self.process_cores
                value, success = QtWidgets.QInputDialog.getInt(self, f'Set {widget.text().split(":")[0].strip()}',
                                                               f'{widget.text().split(":")[0].strip()} value:',
                                                               4, 1, 16, 5)
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_ip_cam_snapshot_button(self):
                widget = self.ip_cam_snapshot
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def set_name_snapshot_button(self):
                widget = self.name_snapshot
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_string_from_list(self, source: list):
                value = ''
                for x in source:
                    value = f'{value} | {x}'
                return value[3::]

            def set_data_func(self, value: str):
                try:
                    self.widget_data_value.setText(f"{value}")
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'set_data_func error : {ex}')

            def create_data_func(self):
                try:
                    data = {
                        'protocol_cam_type': str(self.protocol_cam_type.text().split(':')[1].strip()),
                        'port_cam': int(self.port_cam.text().split(':')[1].strip()),
                        'login_cam': str(self.login_cam.text().split(':')[1].strip()),
                        'password_cam': str(self.password_cam.text().split(':')[1].strip()),
                        'alias_device': list(
                            [x.strip() for x in self.alias_device.text().split(':')[1].strip().split('|')]),
                        'ip_cam': list([x.strip() for x in self.ip_cam.text().split(':')[1].strip().split('|')]),
                        'mask_cam': list([x.strip() for x in self.mask_cam.text().split(':')[1].strip().split('|')]),
                        'sensitivity_analysis': list([int(x.strip()) for x in
                                                      self.sensitivity_analysis.text().split(':')[1].strip().split(
                                                          '|')]),
                        'alarm_level': list(
                            [int(x.strip()) for x in self.alarm_level.text().split(':')[1].strip().split('|')]),
                        'correct_coefficient': list([float(x.strip()) for x in
                                                     self.correct_coefficient.text().split(':')[1].strip().split('|')]),

                        'sql_write': bool(self.sql_write.isChecked()),
                        'ip_sql': str(self.ip_sql.text().split(':')[1].strip()),
                        'server_sql': str(self.server_sql.text().split(':')[1].strip()),
                        'port_sql': str(self.port_sql.text().split(':')[1].strip()),
                        'database_sql': str(self.database_sql.text().split(':')[1].strip()),
                        'user_sql': str(self.user_sql.text().split(':')[1].strip()),
                        'password_sql': str(self.password_sql.text().split(':')[1].strip()),
                        'sql_now_check': bool(self.sql_now_check.isChecked()),
                        'table_now_sql': str(self.table_now_sql.text().split(':')[1].strip()),
                        'rows_now_sql': list(
                            [x.strip() for x in self.rows_now_sql.text().split(':')[1].strip().split('|')]),
                        'sql_data_check': bool(self.sql_data_check.isChecked()),
                        'table_data_sql': str(self.table_data_sql.text().split(':')[1].strip()),
                        'rows_data_sql': list(
                            [x.strip() for x in self.rows_data_sql.text().split(':')[1].strip().split('|')]),

                        'auto_import_check': bool(self.auto_import_check.isChecked()),
                        'auto_play_check': bool(self.auto_play_check.isChecked()),
                        'speed_analysis': float(self.speed_analysis.text().split(':')[1].strip()),
                        'speed_video_stream': float(self.speed_video_stream.text().split(':')[1].strip()),

                        'widget_write': bool(self.widget_write.isChecked()),
                        'text_write': bool(self.text_write.isChecked()),
                        'widget': self.set_data_func,
                        'source_type': str(self.source_type.currentText().strip()),
                        'compute_debug': str(self.compute_debug.currentText().strip()),
                        'process_cores': int(self.process_cores.text().split(':')[1].strip()),
                        'render_debug': str(self.render_debug.currentText().strip()),
                        'resolution_debug': list(self.get_window_resolution()),

                        'import_file': str(self.import_file.text().split(':')[1].strip()),

                        'ip_cam_snapshot': str(self.ip_cam_snapshot.text().split(":")[1].strip()),
                        'name_snapshot': str(self.name_snapshot.text().split(":")[1].strip()),
                    }
                    return data
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'create_data_func error : {ex}')

            def play_btn_func(self):
                try:
                    data = self.create_data_func()
                    self.play_f(data=data)
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'play_btn_func error : {ex}')

            def snapshot_btn_func(self):
                try:
                    data = self.create_data_func()
                    self.snapshot_f(data=data)
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'snapshot_btn_func error : {ex}')

            def export_settings_func(self):
                try:
                    data = self.create_data_func()
                    del data['widget']
                    FileSettings.export_settings(data)
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'export_settings_func error : {ex}')

            def import_settings_func(self):
                try:
                    data = self.create_data_func()
                    data = FileSettings.import_settings(data)
                    self.protocol_cam_type.setText(f'{self.protocol_cam_type.text().split(":")[0].strip()} : '
                                                   f'{str(data["protocol_cam_type"])}')
                    self.port_cam.setText(f'{self.port_cam.text().split(":")[0].strip()} : '
                                          f'{str(data["port_cam"])}')
                    self.login_cam.setText(f'{self.login_cam.text().split(":")[0].strip()} : '
                                           f'{str(data["login_cam"])}')
                    self.password_cam.setText(f'{self.password_cam.text().split(":")[0].strip()} : '
                                              f'{str(data["password_cam"])}')
                    self.alias_device.setText(f'{self.alias_device.text().split(":")[0].strip()} : '
                                              f'{self.get_string_from_list(data["alias_device"])}')
                    self.ip_cam.setText(f'{self.ip_cam.text().split(":")[0].strip()} : '
                                        f'{self.get_string_from_list(data["ip_cam"])}')
                    self.mask_cam.setText(f'{self.mask_cam.text().split(":")[0].strip()} : '
                                          f'{self.get_string_from_list(data["mask_cam"])}')
                    self.sensitivity_analysis.setText(f'{self.sensitivity_analysis.text().split(":")[0].strip()} : '
                                                      f'{self.get_string_from_list(data["sensitivity_analysis"])}')
                    self.alarm_level.setText(f'{self.alarm_level.text().split(":")[0].strip()} : '
                                             f'{self.get_string_from_list(data["alarm_level"])}')
                    self.correct_coefficient.setText(f'{self.correct_coefficient.text().split(":")[0].strip()} : '
                                                     f'{self.get_string_from_list(data["correct_coefficient"])}')
                    self.sql_write.setChecked(data["sql_write"])
                    self.ip_sql.setText(f'{self.ip_sql.text().split(":")[0].strip()} : {str(data["ip_sql"])}')
                    self.server_sql.setText(
                        f'{self.server_sql.text().split(":")[0].strip()} : {str(data["server_sql"])}')
                    self.port_sql.setText(f'{self.port_sql.text().split(":")[0].strip()} : {str(data["port_sql"])}')
                    self.database_sql.setText(f'{self.database_sql.text().split(":")[0].strip()} : '
                                              f'{str(data["database_sql"])}')
                    self.user_sql.setText(f'{self.user_sql.text().split(":")[0].strip()} : '
                                          f'{str(data["user_sql"])}')
                    self.password_sql.setText(f'{self.password_sql.text().split(":")[0].strip()} : '
                                              f'{str(data["password_sql"])}')
                    self.sql_now_check.setChecked(data["sql_now_check"])
                    self.table_now_sql.setText(f'{self.table_now_sql.text().split(":")[0].strip()} : '
                                               f'{str(data["table_now_sql"])}')
                    self.rows_now_sql.setText(f'{self.rows_now_sql.text().split(":")[0].strip()} : '
                                              f'{self.get_string_from_list(data["rows_now_sql"])}')
                    self.sql_data_check.setChecked(data["sql_data_check"])
                    self.table_data_sql.setText(f'{self.table_data_sql.text().split(":")[0].strip()} : '
                                                f'{str(data["table_data_sql"])}')
                    self.rows_data_sql.setText(f'{self.rows_data_sql.text().split(":")[0].strip()} : '
                                               f'{self.get_string_from_list(data["rows_data_sql"])}')
                    self.auto_import_check.setChecked(data["auto_import_check"])
                    self.auto_play_check.setChecked(data["auto_play_check"])
                    self.speed_analysis.setText(f'{self.speed_analysis.text().split(":")[0].strip()} : '
                                                f'{str(data["speed_analysis"])}')
                    self.speed_video_stream.setText(f'{self.speed_video_stream.text().split(":")[0].strip()} : '
                                                    f'{str(data["speed_video_stream"])}')
                    self.widget_write.setChecked(data["widget_write"])
                    self.text_write.setChecked(data["text_write"])
                    self.source_type.setCurrentText(data["source_type"])
                    self.compute_debug.setCurrentText(data["compute_debug"])
                    self.process_cores.setText(f'{self.process_cores.text().split(":")[0].strip()} : '
                                               f'{str(data["process_cores"])}')
                    self.render_debug.setCurrentText(data["render_debug"])
                    self.set_window_resolution(data["resolution_debug"])
                    self.import_file.setText(f'{self.import_file.text().split(":")[0].strip()} : '
                                             f'{str(data["import_file"])}')
                    self.ip_cam_snapshot.setText(f'{self.ip_cam_snapshot.text().split(":")[0].strip()} : '
                                                 f'{str(data["ip_cam_snapshot"])}')
                    self.name_snapshot.setText(f'{self.name_snapshot.text().split(":")[0].strip()} : '
                                               f'{str(data["name_snapshot"])}')
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'import_settings_func error : {ex}')

            def auto_import_settings_func(self):
                try:
                    data = self.create_data_func()
                    data = FileSettings.import_settings(data)
                    if data['auto_import_check']:
                        self.import_settings_func()
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'auto_import_settings_func error : {ex}')

            def auto_play_func(self):
                try:
                    data = self.create_data_func()
                    _data = FileSettings.import_settings(data)
                    _data['widget'] = data['widget']
                    if _data['auto_play_check']:
                        self.play_f(data=_data)
                        self.showMinimized()
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'auto_play_func error : {ex}')


class ParserClass:
    class BeautifulSoupClass:
        @staticmethod
        def example_perser_exams():
            val = 3
            num = f'{val}_'
            data = []
            with open(f'{num}data.html', 'r+', encoding='utf-8') as file:
                questions = []
                index = 1
                for lines in file:
                    c1 = f"""
                    <div id="content_test"><br><div class="workspace-header"><i>Самопроверка промежуточного рейтинга 1</i></div><p><b class="question" id="1">Сұрақ № 1 / Вопрос № 1 </b>АСУ из регулятора и объекта управления, называется:</p><p><input type="radio" name="variant" value="1" id="ans1"><label for="ans1">автоматической системой регулирования АСР</label></p><p><input type="radio" name="variant" value="2" id="ans2"><label for="ans2">дистанционным управлением</label></p><p><input type="radio" name="variant" value="3" id="ans3"><label for="ans3">автоматической системой управления АСУ</label></p><p><input type="radio" name="variant" value="4" id="ans4"><label for="ans4">автоматическим управлением</label></p><p><input type="radio" name="variant" value="5" id="ans5"><label for="ans5">комплексной автоматизацией</label></p><button class="btn btn-number" number="1" style="width:40px;margin:5px;border: 2px solid #e34761;">1</button><button class="btn btn-number" number="2" style="width:40px;margin:5px;">2</button><button class="btn btn-number" number="3" style="width:40px;margin:5px;">3</button><button class="btn btn-number" number="4" style="width:40px;margin:5px;">4</button><button class="btn btn-number" number="5" style="width:40px;margin:5px;">5</button><button class="btn btn-number" number="6" style="width:40px;margin:5px;">6</button><button class="btn btn-number" number="7" style="width:40px;margin:5px;">7</button><button class="btn btn-number" number="8" style="width:40px;margin:5px;">8</button><button class="btn btn-number" number="9" style="width:40px;margin:5px;">9</button><button class="btn btn-number" number="10" style="width:40px;margin:5px;">10</button><button class="btn btn-number" number="11" style="width:40px;margin:5px;">11</button><button class="btn btn-number" number="12" style="width:40px;margin:5px;">12</button><button class="btn btn-number" number="13" style="width:40px;margin:5px;">13</button><button class="btn btn-number" number="14" style="width:40px;margin:5px;">14</button><button class="btn btn-number" number="15" style="width:40px;margin:5px;">15</button><button class="btn btn-number" number="16" style="width:40px;margin:5px;">16</button><button class="btn btn-number" number="17" style="width:40px;margin:5px;">17</button><button class="btn btn-number" number="18" style="width:40px;margin:5px;">18</button><button class="btn btn-number" number="19" style="width:40px;margin:5px;">19</button><button class="btn btn-number" number="20" style="width:40px;margin:5px;">20</button><button class="btn btn-number" number="21" style="width:40px;margin:5px;">21</button><button class="btn btn-number" number="22" style="width:40px;margin:5px;">22</button><button class="btn btn-number" number="23" style="width:40px;margin:5px;">23</button><button class="btn btn-number" number="24" style="width:40px;margin:5px;">24</button><button class="btn btn-number" number="25" style="width:40px;margin:5px;">25</button><button class="btn btn-number" number="26" style="width:40px;margin:5px;">26</button><button class="btn btn-number" number="27" style="width:40px;margin:5px;">27</button><button class="btn btn-number" number="28" style="width:40px;margin:5px;">28</button><button class="btn btn-number" number="29" style="width:40px;margin:5px;">29</button><button class="btn btn-number" number="30" style="width:40px;margin:5px;">30</button><button class="btn btn-number" number="31" style="width:40px;margin:5px;">31</button><button class="btn btn-number" number="32" style="width:40px;margin:5px;">32</button><button class="btn btn-number" number="33" style="width:40px;margin:5px;">33</button><button class="btn btn-number" number="34" style="width:40px;margin:5px;">34</button><button class="btn btn-number" number="35" style="width:40px;margin:5px;">35</button><button class="btn btn-number" number="36" style="width:40px;margin:5px;">36</button><button class="btn btn-number" number="37" style="width:40px;margin:5px;">37</button><button class="btn btn-number" number="38" style="width:40px;margin:5px;">38</button><button class="btn btn-number" number="39" style="width:40px;margin:5px;">39</button><button class="btn btn-number" number="40" style="width:40px;margin:5px;">40</button><hr></div>
<div id="content_test"><br><div class="w
                    """
                    final_data = f"<strong>{index}</strong><br><hr><br>"
                    data = lines.split(" </b>")[1].split("<button class=")[0]
                    try:
                        new_data = data.split('<img src="')
                        if len(new_data) > 2:
                            final_data += data
                        else:
                            final_data += data.split('<img src="')[0] + """<img src="https://sdo.kineu.kz""" + \
                                          data.split('<img src="')[1]
                    except Exception as ex:
                        final_data += data
                    final_data += "<br><hr><br>"
                    reverse = False
                    for question in questions:
                        if question.split("</strong>")[1] == final_data.split("</strong>")[1]:
                            print('повторение!')
                            reverse = True
                            break
                    if reverse is False:
                        questions.append(final_data)
                        index += 1
            # print(questions)
            print(len(questions))
            with open(f'{num}new_data.html', 'w', encoding='utf-8') as file:
                title = """<!DOCTYPE html>
            <html lang="en">
            <head>
                <meta charset="UTF-8">
                <title>Title</title>
            </head>
            <body>
                """
                footer = """
                </body>
            </html>"""
                file.write(title)
                for line in questions:
                    file.write(line)
                file.write(footer)

        @staticmethod
        def example_parse_weather():
            def pandas(url):
                print(url)
                headers = {'user-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:72.0) Gecko/20100101 Firefox/72.0',
                           'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'}
                # url = 'http://www.pogodaiklimat.ru/weather.php?id=35042&bday=1&fday=1&amonth=10&ayear=2021&bot=2'
                response = requests.get(url, headers=headers)
                response.encoding = 'utf-8'
                soup = BeautifulSoup(response.text, "html.parser")
                soup.encoded = 'utf-8'
                rows = soup.find('div', class_='archive-table-wrap')
                return rows

            def set_sheet_value(column, row, value_, sheet):
                try:
                    if value_:
                        sheet[str(column) + str(row)] = value_
                    else:
                        sheet[str(column) + str(row)] = ''
                except:
                    sheet[str(column) + str(row)] = ''

            print('begin')
            arr_url = []
            for year in range(2011, 2012):
                for month in range(1, 2):
                    for day in range(1, 32):
                        url = f"http://www.pogodaiklimat.ru/weather.php?id=35042&bday={day}&fday={day}&amonth={month}&ayear={year}&bot=2"
                        # print(f"day={day} month={month} year={year}")
                        arr_url.append(url)
            # print(arr_url)

            arr_requests = []
            for url in arr_url:
                arr_requests.append(pandas(url))
            # print(arr_requests)

            all_data = []
            for request in arr_requests:
                data_list = []
                for row in request.find_all('tr'):
                    local_data = []
                    # print(row)
                    # print('\n *************** \n')
                    for col in row.find_all('td'):
                        local_data.append(col.text)
                        # print(col.text)
                        # print('\n *************** \n')
                    data_list.append(local_data)
                # print(data_list)
                all_data.append([''])
                all_data.append(data_list)
            # print(all_data)

            _import_file = 'data_1.xlsx'
            workbook = openpyxl.load_workbook(_import_file)
            sheet = workbook.active

            row_index = 0
            for data in all_data:
                # print(data)
                for row in data:
                    # print(row)
                    row_index += 1
                    col_index = 0
                    for cell in row:
                        col_index += 1
                        sheet[get_column_letter(col_index) + str(row_index)] = str(cell)

            workbook.save(_import_file)
            workbook.close()
            print('complete')

        @staticmethod
        def example_parse_wea():
            def pandas(url):
                print(url)
                headers = {'user-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:72.0) Gecko/20100101 Firefox/72.0',
                           'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'}
                # url = 'http://www.pogodaiklimat.ru/weather.php?id=35042&bday=1&fday=1&amonth=10&ayear=2021&bot=2'
                response = requests.get(url, headers=headers)
                response.encoding = 'utf-8'
                soup = BeautifulSoup(response.text, "html.parser")
                soup.encoded = 'utf-8'
                rows = soup.find('div', class_='archive-table-wrap')
                if rows is None or rows == '':
                    rows = 'ТАКОГО ДНЯ НЕ СУЩЕСТВУЕТ'
                return rows

            def set_sheet_value(column, row, value_, sheet):
                try:
                    if value_:
                        sheet[str(column) + str(row)] = value_
                    else:
                        sheet[str(column) + str(row)] = ''
                except:
                    sheet[str(column) + str(row)] = ''

            print('begin')
            arr_url = []
            for year in range(2011, 2022):
                for month in range(1, 13):
                    for day in range(1, 32):
                        url = f"http://www.pogodaiklimat.ru/weather.php?id=35042&bday={day}&fday={day}&amonth={month}&ayear={year}&bot=2"
                        print(f"day={day} month={month} year={year}")
                        arr_url.append([f"{url}", f"{day}.{month}.{year}"])
            # print(arr_url[:100])

            arr_requests = []
            for url in arr_url[:3]:
                data = pandas(url[0])
                arr_requests.append([url[1], data])
            # print(arr_requests)

            _import_file = 'data_1.xlsx'
            workbook = openpyxl.load_workbook(_import_file)
            sheet = workbook.active
            # row_index = 0
            # for req in arr_requests:
            #     row_index += 1
            #     for one_req in req:
            #         if req.index(one_req) == 0:
            #             sheet['A' + str(row_index)] = str(one_req)
            #         if req.index(one_req) == 1:
            #             sheet['B' + str(row_index)] = str(one_req)

            final_arr = []
            for one_request in arr_requests:
                data = one_request[0]
                req = one_request[1].find_all('tr')
                # req - ТУТ ЛЕЖИТ ОДИН ОТВЕТ ОТ КАЖДОЙ ДАТЫ
                # print(req)
                temp_arr = []
                sum_temp = []
                for one_req in req[1:]:
                    # one_req - ТУТ ЛЕЖИТ ОДНА СТРОКА ОТ КАЖДОЙ ТАБЛИЦЫ
                    temp_temp = []
                    try:
                        temp = str(one_req).split('class="temp')[1].split('</nobr></td>')[0].split('nobr>')[1]
                        # temp - ТУТ ЛЕЖИТ ТЕМПЕРАТУРА ДЛЯ КАЖДОЙ СТРОКИ
                        temp_temp.append(float(temp))
                        sum_temp.append(float(temp))
                        print(f"{temp}")
                        temp_arr.append([temp])
                    except Exception as ex:
                        print(ex)
                        temp = -1000
                sum_ = 0.0
                for tem in sum_temp:
                    print(f"temperature: {tem}")
                    sum_ += tem
                sum_ = sum_ / len(sum_temp)

                print(f"summ temperature: {round(sum_, 1)}")

                #     try:
                #         pressure = str(one_req).split('class="temp_33">')[1].split('</td>')[0]
                #         print(f"{pressure} \n\n\n\n\n")
                #     except:
                #         pressure = -1000
                #     temp_arr.append([temp, pressure])
                str_temp = ''
                for t in sum_temp:
                    str_temp += f", {t}"

                arr = [data, str_temp[2:], temp_arr[0][0]]
                final_arr.append(arr)

            row_index = 0
            for arr in final_arr:
                row_index += 1
                col_index = 0
                for one_req in arr:
                    col_index += 1
                    sheet[str(get_column_letter(col_index)) + str(row_index)] = one_req

                    # print(f"{one_req} \n\n\n\n\n")
                # row_index = 0
                # col_index = 0
                # for row in req.find_all('td'):
                # for row in req.find_all('td'):
                #     print(row)

                # row_index += 1
                # col_index = 0
                # for col in row.find_all('td'):
                #     col_index += 1
                #     # print(f"COL: \n {col.text} \n")
                #     if col_index == 4 and row_index != 1:
                #         value = str(col).split('</td>')[0].split('>')[1].strip()
                #     elif col_index == 5 and row_index != 1:
                #         # try:
                #         #     value = str(col.split('</')[0])
                #         # except:
                #         #     value = str(col)
                #         value = str(col).split('<')[1].split('>')[1].strip()
                #         # print(str(col))
                #         print(value)
                #     else:
                #         value = str(col).split('</td>')[0].split('>')[1].strip()
                #
                #     # print(f"COL: \n {value} \n")
                #     try:
                #         sheet[str(get_column_letter(col_index)) + str(row_index)] = value
                #     except Exception as ex:
                #         # print(f"ERROR: {ex}")
                #         pass
            workbook.save(_import_file)
            workbook.close()
            print('complete')

            stroing = '-23.7, -21.2, -18.7, -16.7, -17.7, -19.4, -23.0'

            vals = stroing.split(', ')
            print(vals)

            # url = 'http://www.pogodaiklimat.ru/weather.php?id=35042'
            # response.encoding = 'utf-8'
            # file = open('data.csv', 'w')
            # writer = csv.writer(file)
            # soup = BeautifulSoup(response.text, "html.parser")
            # soup.encoded = 'utf-8'
            # rows = soup.find('div', class_='archive-table-wrap').find_all('tr')
            # for row in rows:
            #     columns = row.find_all('td')
            #     print(columns)
            #     data_list = [columns[0].text, columns[1].text, columns[2].text, columns[3].text, columns[4].text, columns[5].text,
            #                  columns[6].text, columns[7].text, columns[8].text, columns[9].text, columns[10].text, columns[11].text,
            #                  columns[12].text, columns[13].text, columns[14].text, columns[15].text, columns[16].text]
            #     writer.writerow(data_list)
            # file.close()

        @staticmethod
        def example_parse_wea_main():
            def get_sheet_value(_column, _row, _sheet):
                """"
                Принимает: индексы колонки и строки для извлечения данных, а также лист откуда извлекать.
                Возвращает: значение, находящееся по индексам на нужном листе.
                """
                try:
                    return _sheet[str(_column) + str(_row)].value
                except Exception as ex:
                    print(ex)
                    return ''

            def set_sheet_value(_column, _row, _sheet, _value):
                """"
                Принимает: индекс колонки и строку для записи данных, а также лист откуда куда записывать и значение для записи.
                """
                try:
                    int(_column)
                    _column = get_column_letter(_column)
                except ValueError:
                    pass
                try:
                    sheet[f'{_column}{_row}'] = str(_value)
                except Exception as _ex:
                    print(_ex)

            file_xlsx = 'excel.xlsx'
            workbook = openpyxl.load_workbook(file_xlsx)
            sheet = workbook.active
            old_arr = []
            for row in range(3, 209 + 1):
                arr_local = []
                for column in "ABCJ":
                    cell_vall = get_sheet_value(_column=column, _row=row, _sheet=sheet)
                    arr_local.append(cell_vall)
                old_arr.append(arr_local)
            workbook.close()
            print(old_arr)

            file_xlsx = 'new_excel.xlsx'
            workbook = openpyxl.load_workbook(file_xlsx)
            sheet = workbook.active
            for row_1 in old_arr:
                for col_1 in row_1:
                    set_sheet_value(_row=old_arr.index(row_1) + 2, _column=row_1.index(col_1) + 1, _sheet=sheet,
                                    _value=col_1)
            workbook.save('new_excel.xlsx')

            file_xlsx = 'new_excel.xlsx'
            workbook = openpyxl.load_workbook(file_xlsx)
            sheet = workbook.active
            clear_arr = []
            index = 0
            for row in range(2, 2000):
                arr_local = []
                for column in "ABCD":
                    cell_vall = get_sheet_value(_column=column, _row=row, _sheet=sheet)
                    arr_local.append(cell_vall)
                if arr_local[0]:
                    index += 1
                    arr_local.append(index)
                    clear_arr.append(arr_local)
            workbook.close()
            print(clear_arr)

            file_xlsx = 'clear_excel.xlsx'
            workbook = openpyxl.load_workbook(file_xlsx)
            sheet = workbook.active
            for row_1 in clear_arr:
                for col_1 in row_1:
                    set_sheet_value(_row=clear_arr.index(row_1) + 2, _column=row_1.index(col_1) + 1, _sheet=sheet,
                                    _value=col_1)
            workbook.save('clear_excel.xlsx')

            # arr = ['A', 'C', 'D']
            # for val in "BCD":
            #     try:
            #         arr.index(val)
            #         print(f'Этот элемент: {val} уже есть, мы его не добавляем')
            #         pass
            #     except Exception as ex:
            #         print(f'Этого элемента: {val} нет, мы его добавляем')
            #         arr.append(val)
            # print(arr)

            # temp_1 = 'Управление предприятия (ЦК)'
            # temp_2 = 'Отдел закупок'
            #
            # for row in arr:
            #     # print(f"строка: {row}")
            #     row_1 = row[0]
            #     row_2 = row[1]
            #     row_3 = row[2]
            #     if row_1 == temp_1 and row_2 == temp_2:
            #         for cell in row:
            #             print(f"ячейка: {cell}")

            # arr = []
            # for row in range(2, 10):
            #     arr_local = []
            #     for column in range(1, 6):
            #         cell_vall = get_sheet_value(_column=get_column_letter(column), _row=row, _sheet=sheet)
            #         print(cell_vall)
            #         arr_local.append(cell_vall)
            #     arr.append(arr_local)
            # # print(arr)


class ConsoleCalculatorClass:
    @staticmethod
    def example():
        init(wrap=False)
        stream = AnsiToWin32(sys.stderr).stream

        def mathematics(first_value, argument, second_value):
            result = str('Вы ввели некорректное действие')
            if argument == '*':
                result = int(first_value * second_value)
            if argument == '/':
                if second_value == 0:
                    while second_value == 0:
                        print('Делитель равен нулю!')
                        second_value = int(input("Введите  второе число: "))
                if second_value != 0:
                    result = first_value / second_value
                if result > int(result):
                    result = float(result)
                if result == int(result):
                    result = int(result)
            if argument == '+':
                result = int(first_value + second_value)
            if argument == '-':
                result = int(first_value - second_value)
            if argument == '%':
                result = str(int(first_value * second_value / 100)) + '%'
            if argument == '**':
                result = int(first_value ** second_value)
            if argument == '^':
                result = 'Корень из первого числа:  ' + str(
                    math.sqrt(first_value)) + '     Корень из второго числа:  ' + str(
                    math.sqrt(second_value))
            return result

        while True:
            print(Fore.YELLOW + '', file=stream)
            a = int(input("Введите первое число: "))
            print(Fore.GREEN + '', file=stream)
            b = str(input("Введите действие: "))
            print(Fore.BLUE + '', file=stream)
            c = int(input("Введите второе число: "))
            print(Fore.RED + '', file=stream)
            print('Ответ: ' + str(mathematics(a, b, c)))
            print(Fore.RESET + '', file=stream)
            if str.lower(str(input("Выйти?(N/n)"))) == "n":
                break


class ConsoleTimerClass:
    @staticmethod
    def example():
        init(wrap=False)
        stream = AnsiToWin32(sys.stderr).stream

        def render(sec, minut, hour):
            print(Fore.GREEN + str(hour), Fore.WHITE + ":", Fore.BLUE + str(minut), Fore.WHITE + ":",
                  Fore.RED + str(sec), file=stream)

        def tick(seconds_, multiplayer_seconds):
            seconds_ += int(multiplayer_seconds)
            return seconds_

        seconds = float(0)
        minutes = float(0)
        hours = float(0)
        while True:
            seconds = tick(seconds, 1)
            if seconds > 59:
                seconds = float(0)
                minutes += 1
                if minutes > 59:
                    minutes = float(0)
                    hours += 1
                    if hours > 24:
                        hours = float(0)
                        minutes = float(0)
                        seconds = float(0)
            render(int(seconds), int(minutes), int(hours))
            time.sleep(1.0)


class PathWindowsClass:
    @staticmethod
    def example():
        path = r"C:\Project\Github_Projects\python-jsx-smart-pmp-app\web-platform_09_21_dev\src"
        new_path = r"C:\Project\Github_Projects\python-jsx-smart-pmp-app\web-platform_09_21_dev\new"
        pattern = '*.js'

        directories_ = []
        for root, dirs, files in os.walk(path, topdown=True):
            for name in dirs:
                directories_.append(f"{os.path.join(root, name)}")
        print(directories_)
        # for dir in directories_:
        #     print(dir)
        files_ = []
        for dir_ in directories_:
            for root, dirs, files in os.walk(dir_, topdown=True):
                for file in files:
                    # print(file)
                    if fnmatch(file, pattern):
                        files_.append(f"{dir_}\\{file}")
                    # if fnmatch(file, "*.docx"):
                    #     files_.append(f"{dir_}\\{file}")
                    # files_.append(file)
        print(files_)
        # for file in files_:
        #     print(file)
        for file in files_:
            print(file)
            try:
                file_name = file.split('\\')[len(file.split('\\')) - 1]
                print(file_name)
                os.replace(file, f"{new_path}\\{file_name}")
                # move(file, f"{new_path}\\{file_name}")
            except Exception as ex:
                print(ex)


class GeoBilazClass:
    @staticmethod
    def example_geo_1():
        connection = pg.connect(
            host="192.168.1.6",
            database="navSections",
            port="5432",
            user="postgres",
            password="nF2ArtXK"
        )

        # postgreSQL_select_Query = "SELECT * FROM public.navdata_202108 " \
        #                           "WHERE flags = '0' " \
        #                           "ORDER BY navtime DESC, device DESC LIMIT 100;"
        # "ORDER BY navtime DESC, device DESC LIMIT 100"
        # "ORDER BY device ASC, navtime DESC LIMIT 100"
        # "WHERE flags = 0 " \

        cursor = connection.cursor()

        # cursor.execute(postgreSQL_select_Query)

        mobile_records = cursor.fetchall()

        wb = Workbook()
        sheet = wb.active

        # cols = ["device", "navtime", "latitude", "longtode", "alttude", "speed", "ds", "direction", "flags"]
        cols = ["устройство", "дата и время", "широта", "долгота", "высота", "скорость", "ds", "направление",
                "флаги ошибок"]

        for col in cols:
            id_s = cols.index(col)
            sheet[f'{get_column_letter(id_s + 1)}1'] = col

        for rows in mobile_records:
            for value in rows:
                id_s = rows.index(value)
                # print(f"{cols[id_s]}: {value}")
                if id_s == 1:
                    sheet[f'{get_column_letter(id_s + 1)}{mobile_records.index(rows) + 2}'] = \
                        datetime.datetime.fromtimestamp(int(value - 21600)).strftime('%Y-%m-%d %H:%M:%S')
                else:
                    sheet[f'{get_column_letter(id_s + 1)}{mobile_records.index(rows) + 2}'] = value

        wb.save('data.xlsx')

    @staticmethod
    def example_geo_2():
        fig, ax = plt.subplots(subplot_kw={"projection": "3d"})

        # Make data.
        X = np.arange(-5, 5, 0.25)
        Y = np.arange(-5, 5, 0.25)
        X, Y = np.meshgrid(X, Y)
        R = np.sqrt(X ** 2 + Y ** 2)
        Z = np.sin(R)

        # Plot the surface.
        surf = ax.plot_surface(X, Y, Z, cmap=cm.coolwarm,
                               linewidth=0, antialiased=False)

        # Customize the z axis.
        ax.set_zlim(-1.01, 1.01)
        ax.zaxis.set_major_locator(LinearLocator(10))
        # A StrMethodFormatter is used automatically
        ax.zaxis.set_major_formatter('{x:.02f}')

        # Add a color bar which maps values to colors.
        fig.colorbar(surf, shrink=0.5, aspect=5)

        plt.show()

    @staticmethod
    def example_geo_3():
        fig = plt.figure()
        ax = fig.add_subplot(projection='3d')

        # Grab some test data.
        X, Y, Z = axes3d.get_test_data(0.1)

        # Plot a basic wireframe.
        ax.plot_wireframe(X, Y, Z, rstride=10, cstride=10)

        plt.show()

    @staticmethod
    def example_geo_4():

        # Initialize the map at a given point
        gmap = gmplot.GoogleMapPlotter(52.13047374297343, 61.33322598988672, 13)

        # Add a marker
        gmap.marker(52.13047374297343, 61.33322598988672, 'cornflowerblue')

        # Draw map into HTML file
        gmap.draw("my_map.html")


class OpenCvClass:
    @staticmethod
    def example_computer_vision():
        # UTILITES
        class LoggingClass:
            @staticmethod
            def logging(message, file_name='log.txt', type_write='a'):
                with open(file_name, type_write) as log:
                    log.write(f'{TimeUtils.get_current_time()} : {message}\n')

        class CopyDictionary:
            @staticmethod
            def get_all_sources(source: dict, values: dict):
                value = source.copy()
                for _key, _value in values.items():
                    value[_key] = _value
                return value

        class FileSettings:
            @staticmethod
            def export_settings(data: dict):
                with open(f"{data['import_file']}.json", 'w') as file:
                    json.dump(data, file)

            @staticmethod
            def import_settings(data: dict):
                try:
                    with open(f"{data['import_file']}.json", "r") as read_file:
                        data = json.load(read_file)
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'import_settings error : {ex}')
                return data

        class TimeUtils:
            @staticmethod
            def get_current_time():
                return f"{time.strftime('%X')}"

        class SendMail:
            @staticmethod
            def sender_email(subject='subj', text='text'):
                host = 'smtp.yandex.ru'
                port = '465'
                login = 'eevee.cycle'
                password = '31284bogdan'
                writer = 'eevee.cycle@yandex.ru'
                recipient = 'eevee.cycle@yandex.ru'

                message = f"""From: {recipient}\nTo: {writer}\nSubject: {subject}\n\n{text}"""

                smtpobj = smtplib.SMTP_SSL(host=host, port=port)
                smtpobj.ehlo()
                smtpobj.login(user=login, password=password)
                smtpobj.sendmail(from_addr=writer, to_addrs=recipient, msg=message)
                smtpobj.quit()

        # SQL
        class SQLClass:
            @staticmethod
            def sql_post_data(ip: str, server: str, port: str, database: str, username: str, password: str, table: str,
                              rows: list, values: list):
                try:
                    sql = SQLClass.pyodbc_connect(ip=ip, server=server, port=port, database=database, username=username,
                                                  password=password)
                    SQLClass.execute_data_query(connection=sql, table=table, rows=rows, values=values)
                except Exception as ex:
                    print(ex)
                    with open('log.txt', 'a') as log:
                        log.write(f'\n{ex}\n')

            @staticmethod
            def sql_post_now(ip: str, server: str, port: str, database: str, username: str, password: str, table: str,
                             rows: list, values: list):
                try:
                    sql = SQLClass.pyodbc_connect(ip=ip, server=server, port=port, database=database, username=username,
                                                  password=password)
                    SQLClass.execute_now_query(connection=sql, table=table, rows=rows, values=values)
                except Exception as ex:
                    print(ex)
                    with open('log.txt', 'a') as log:
                        log.write(f'\n{ex}\n')

            @staticmethod
            def pyodbc_connect(ip: str, server: str, port: str, database: str, username: str, password: str):
                conn_str = (
                        r'DRIVER={ODBC Driver 17 for SQL Server};SERVER=tcp:' + ip + '\\' + server + ',' + port +
                        ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password + ';'
                )
                return pyodbc.connect(conn_str)

            @staticmethod
            def pd_read_sql_query(connection, query: str, database: str, table: str):
                return pandas.read_sql_query(f'{query} {database}.dbo.{table} ORDER BY id', connection)

            @staticmethod
            def execute_data_query(connection, table: str, rows: list, values: list):
                cursor = connection.cursor()
                cursor.fast_executemany = True
                _rows = ''
                for x in rows:
                    _rows = f"{_rows}{str(x)}, "
                value = f"INSERT INTO {table} (" + _rows[:-2:] + f") VALUES {tuple(values)}"
                cursor.execute(value)
                connection.commit()

            @staticmethod
            def execute_now_query(connection, table, rows: list, values: list):
                cursor = connection.cursor()
                cursor.fast_executemany = True
                __rows = ''
                for x in rows:
                    __rows = f"{__rows}{str(x)}, "
                value = f"UPDATE {table} SET {rows[1]} = '{values[1]}',{rows[2]} = '{values[2]}' ,{rows[3]} = '{values[3]}' " \
                        f"WHERE {rows[0]} = '{values[0]}'"
                cursor.execute(value)
                connection.commit()

        # OPENCV
        class AnalyzeClass:
            @staticmethod
            def start_analyze(data: dict):
                sources = AnalyzeClass.get_full_sources(source_type=data['source_type'], sources=data['ip_cam'],
                                                        masks=data['mask_cam'], protocol=data['protocol_cam_type'],
                                                        login=data['login_cam'], password=data['password_cam'],
                                                        port=data['port_cam'], sensitivity=data['sensitivity_analysis'],
                                                        correct=data['correct_coefficient'],
                                                        alias_device=data['alias_device'],
                                                        alarm_level=data['alarm_level'])
                data = CopyDictionary.get_all_sources(data, {'sources': sources})
                if data['compute_debug'] == 'sync':
                    AnalyzeClass.sync_method(data)
                elif data['compute_debug'] == 'async':
                    AnalyzeClass.async_method(data)
                elif data['compute_debug'] == 'multithread':
                    AnalyzeClass.multithread_method(data)
                elif data['compute_debug'] == 'multiprocess':
                    AnalyzeClass.multiprocess_method(data)
                elif data['compute_debug'] == 'complex':
                    AnalyzeClass.complex_method(data)

            @staticmethod
            def sync_method(data: dict):
                while True:
                    if not data['pause']():
                        cv2.destroyAllWindows()
                        break
                    else:
                        for source in data['sources']:
                            AnalyzeClass.analyze(source, data)
                    time.sleep(1 / data['speed_analysis'])

            @staticmethod
            def async_method(data: dict):
                def thread_loop(source):
                    while True:
                        if not data['pause']():
                            cv2.destroyAllWindows()
                            break
                        else:
                            asyncio.run(AnalyzeClass.async_analiz(source, data))
                        time.sleep(1 / data['speed_analysis'])

                for src in data['sources']:
                    threading.Thread(target=thread_loop, args=(src,)).start()

            @staticmethod
            async def async_analiz(source, data: dict):
                image = await AnalyzeClass.async_get_source(data['source_type'], source[0], data['login_cam'],
                                                            data['password_cam'])
                mask = source[1]
                name = source[0].split("192.168.")[1].strip().split(":")[0].strip()
                AnalyzeClass.render_final(image=image, mask=mask,
                                          sensitivity_analysis=source[2],
                                          correct_coefficient=source[3],
                                          name=name,
                                          resolution_debug=data['resolution_debug'])
                cv2.waitKey(round(1000 / data['speed_video_stream']))
                values = AnalyzeClass.result_final(image=image, mask=mask,
                                                   sensitivity_analysis=source[2],
                                                   correct_coefficient=source[3])
                AnalyzeClass.write_result(ip_sql=data['ip_sql'],
                                          server_sql=data['server_sql'],
                                          port_sql=data['port_sql'],
                                          database_sql=data['database_sql'],
                                          user_sql=data['user_sql'],
                                          password_sql=data['password_sql'],
                                          sql_now_check=data['sql_now_check'],
                                          table_now_sql=data['table_now_sql'],
                                          rows_now_sql=data['rows_now_sql'],
                                          sql_data_check=data['sql_data_check'],
                                          table_data_sql=data['table_data_sql'],
                                          rows_data_sql=data['rows_data_sql'],
                                          values=values,
                                          source=name,
                                          widget=data['widget'],
                                          widget_write=data['widget_write'],
                                          text_write=data['text_write'],
                                          sql_val=data['sql_write'],
                                          alarm_level=source[5]
                                          )

            @staticmethod
            async def async_get_source(source_type: str, sources: str, login: str, password: str):
                if source_type == 'image-http':
                    h = httplib2.Http(os.path.abspath('__file__'))
                    h.add_credentials(login, password)
                    response, content = h.request(sources)
                    image = cv2.imdecode(numpy.frombuffer(content, numpy.uint8), cv2.IMREAD_COLOR)
                    return image
                elif source_type == 'video-rtsp' or 'video-file':
                    cam_stream = cv2.VideoCapture(sources)
                    _, image = cam_stream.read()
                    cam_stream.release()
                    return image

            @staticmethod
            def multithread_method(data: dict):
                def thread_loop(source):
                    while True:
                        if not data['pause']():
                            cv2.destroyAllWindows()
                            break
                        else:
                            AnalyzeClass.analyze(source, data)
                        time.sleep(1 / data['speed_analysis'])

                for src in data['sources']:
                    threading.Thread(target=thread_loop, args=(src,)).start()

            @staticmethod
            def multiprocess_method(data: dict):
                def process_loop():
                    while True:
                        if not data['pause']():
                            cv2.destroyAllWindows()
                            break
                        else:
                            with multiprocessing.Pool(data['process_cores']) as process:
                                process.map(AnalyzeClass.multi, _data)
                        time.sleep(1 / data['speed_analysis'])

                _data = []
                for loop in data['sources']:
                    __data = data.copy()
                    del __data['widget']
                    __data['source'] = loop
                    _data.append(__data)
                threading.Thread(target=process_loop).start()

            @staticmethod
            def multi(kwargs):
                source = kwargs['source']
                kwargs['widget'] = None
                AnalyzeClass.analyze(source, kwargs)

            @staticmethod
            def complex_method(data: dict):
                def thread_loop(source):
                    while True:
                        if not data['pause']():
                            cv2.destroyAllWindows()
                            break
                        else:
                            asyncio.run(AnalyzeClass.async_complex_analiz(source, data))
                        time.sleep(1 / data['speed_analysis'])

                for src in data['sources']:
                    threading.Thread(target=thread_loop, args=(src,)).start()

            @staticmethod
            async def async_complex_analiz(source, data: dict):
                try:
                    image = await AnalyzeClass.async_get_source(data['source_type'], source[0], data['login_cam'],
                                                                data['password_cam'])
                    mask = source[1]
                    name = source[4]
                    if data['render_debug'] == 'all':
                        AnalyzeClass.render_flip(image=image, flipcode=0, name=name,
                                                 resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_cvtcolor(image=image, color_type=cv2.COLOR_RGB2GRAY, name=name,
                                                     resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_shapes(name=name, resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_origin(image=image, name=name,
                                                   resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_cropping_image(image=image, name=name,
                                                           resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_bitwise_and(image=image, mask=mask, name=name,
                                                        resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_bitwise_not_white(image=image, mask=mask, name=name,
                                                              resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_cvtcolor_to_hsv(image=image, name=name,
                                                            resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_threshold(image=image, name=name,
                                                      resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_inrange(image=image, sensitivity_analysis=source[2],
                                                    name=name,
                                                    resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_canny_edges(image=image,
                                                        sensitivity_analysis=source[2],
                                                        name=name,
                                                        resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_final(image=image, mask=mask,
                                                  sensitivity_analysis=source[2],
                                                  correct_coefficient=source[3],
                                                  name=name,
                                                  resolution_debug=data['resolution_debug'])
                    elif data['render_debug'] == 'extended':
                        AnalyzeClass.render_origin(image=image, name=name,
                                                   resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_inrange(image=image, sensitivity_analysis=source[2],
                                                    name=name,
                                                    resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_final(image=image, mask=mask,
                                                  sensitivity_analysis=source[2],
                                                  correct_coefficient=source[3],
                                                  name=name,
                                                  resolution_debug=data['resolution_debug'])
                    elif data['render_debug'] == 'final':
                        AnalyzeClass.render_final(image=image, mask=mask,
                                                  sensitivity_analysis=source[2],
                                                  correct_coefficient=source[3],
                                                  name=name,
                                                  resolution_debug=data['resolution_debug'])
                    elif data['render_debug'] == 'source':
                        AnalyzeClass.render_origin(image=image, name=name,
                                                   resolution_debug=data['resolution_debug'])
                    cv2.waitKey(round(1000 / data['speed_video_stream']))
                    values = AnalyzeClass.result_final(image=image, mask=mask,
                                                       sensitivity_analysis=source[2],
                                                       correct_coefficient=source[3])
                    AnalyzeClass.write_result(ip_sql=data['ip_sql'],
                                              server_sql=data['server_sql'],
                                              port_sql=data['port_sql'],
                                              database_sql=data['database_sql'],
                                              user_sql=data['user_sql'],
                                              password_sql=data['password_sql'],
                                              sql_now_check=data['sql_now_check'],
                                              table_now_sql=data['table_now_sql'],
                                              rows_now_sql=data['rows_now_sql'],
                                              sql_data_check=data['sql_data_check'],
                                              table_data_sql=data['table_data_sql'],
                                              rows_data_sql=data['rows_data_sql'],
                                              values=values,
                                              source=name,
                                              widget=data['widget'],
                                              widget_write=data['widget_write'],
                                              text_write=data['text_write'],
                                              sql_val=data['sql_write'],
                                              alarm_level=source[5])
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'analyze func error')
                    print(ex)
                    # SendMail.sender_email(subject='error', text='analyze func error')

            @staticmethod
            def analyze(source, data: dict):
                try:
                    image = AnalyzeClass.get_source(data['source_type'], source[0], data['login_cam'],
                                                    data['password_cam'])
                    mask = source[1]
                    name = source[0].split("192.168.")[1].strip().split(":")[0].strip()
                    if data['render_debug'] == 'all':
                        AnalyzeClass.render_origin(image=image, name=name,
                                                   resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_cropping_image(image=image, name=name,
                                                           resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_bitwise_and(image=image, mask=mask, name=name,
                                                        resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_bitwise_not_white(image=image, mask=mask, name=name,
                                                              resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_cvtcolor_to_hsv(image=image, name=name,
                                                            resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_threshold(image=image, name=name,
                                                      resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_inrange(image=image, sensitivity_analysis=source[2],
                                                    name=name,
                                                    resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_canny_edges(image=image,
                                                        sensitivity_analysis=source[2],
                                                        name=name,
                                                        resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_final(image=image, mask=mask,
                                                  sensitivity_analysis=source[2],
                                                  correct_coefficient=source[3],
                                                  name=name,
                                                  resolution_debug=data['resolution_debug'])
                    elif data['render_debug'] == 'extended':
                        AnalyzeClass.render_origin(image=image, name=name,
                                                   resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_inrange(image=image, sensitivity_analysis=source[2],
                                                    name=name,
                                                    resolution_debug=data['resolution_debug'])
                        AnalyzeClass.render_final(image=image, mask=mask,
                                                  sensitivity_analysis=source[2],
                                                  correct_coefficient=source[3],
                                                  name=name,
                                                  resolution_debug=data['resolution_debug'])
                    elif data['render_debug'] == 'final':
                        AnalyzeClass.render_final(image=image, mask=mask,
                                                  sensitivity_analysis=source[2],
                                                  correct_coefficient=source[3],
                                                  name=name,
                                                  resolution_debug=data['resolution_debug'])
                    elif data['render_debug'] == 'source':
                        AnalyzeClass.render_origin(image=image, name=name,
                                                   resolution_debug=data['resolution_debug'])
                    cv2.waitKey(round(1000 / data['speed_video_stream']))
                    values = AnalyzeClass.result_final(image=image, mask=mask,
                                                       sensitivity_analysis=source[2],
                                                       correct_coefficient=source[3])
                    AnalyzeClass.write_result(
                        ip_sql=data['ip_sql'],
                        server_sql=data['server_sql'],
                        port_sql=data['port_sql'],
                        database_sql=data['database_sql'],
                        user_sql=data['user_sql'],
                        password_sql=data['password_sql'],
                        sql_now_check=data['sql_now_check'],
                        table_now_sql=data['table_now_sql'],
                        rows_now_sql=data['rows_now_sql'],
                        sql_data_check=data['sql_data_check'],
                        table_data_sql=data['table_data_sql'],
                        rows_data_sql=data['rows_data_sql'],
                        values=values,
                        source=name,
                        widget=data['widget'],
                        widget_write=data['widget_write'],
                        text_write=data['text_write'],
                        sql_val=data['sql_write'],
                        alarm_level=source[5]
                    )
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'{TimeUtils.get_current_time()} : analyze func error')
                    print(ex)

            @staticmethod
            def get_full_sources(source_type: str, sources: list, masks: list, protocol: str, login: str, password: str,
                                 port: int, sensitivity: list, correct: list, alias_device: list, alarm_level: list):
                try:
                    _sources = []
                    for x in sources:
                        index = sources.index(x)
                        if source_type == 'image-http':
                            _sources.append([f'{protocol}://192.168.{x}:{port}/ISAPI/Streaming/channels/101/'
                                             f'picture?snapShotImageType=JPEG', cv2.imread(masks[index], 0),
                                             sensitivity[index], correct[index], alias_device[index],
                                             alarm_level[index]])
                        elif source_type == 'video-rtsp':
                            _sources.append(
                                [f'{protocol}://{login}:{password}@192.168.{x}:{port}', cv2.imread(masks[index], 0),
                                 sensitivity[index], correct[index], alias_device[index], alarm_level[index]])
                        elif source_type == 'video-file':
                            _sources.append([f'{x}', cv2.imread(masks[index], 0), sensitivity[index], correct[index],
                                             alias_device[index], alarm_level[index]])
                    return _sources
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'get_full_sources error : {ex}')

            @staticmethod
            def get_source(source_type: str, sources: str, login: str, password: str):
                try:
                    if source_type == 'image-http':
                        h = httplib2.Http("/path/to/cache-directory")
                        h.add_credentials(login, password)
                        response, content = h.request(sources)
                        image = cv2.imdecode(numpy.frombuffer(content, numpy.uint8), cv2.IMREAD_COLOR)
                        return image
                    elif source_type == 'video-rtsp' or 'video-file':
                        cam_stream = cv2.VideoCapture(sources)
                        _, image = cam_stream.read()
                        cam_stream.release()
                        return image
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'get_source func error')
                    print(ex)

            @staticmethod
            def render_origin(image, name, resolution_debug):
                AnalyzeClass.render(name=f"origin : {name}", source=image, resolution_debug=resolution_debug)

            @staticmethod
            def render_cropping_image(image, name, resolution_debug):
                cropping_image = image[250:1080, 600:1720]
                AnalyzeClass.render(name=f"cropping_image : {name}", source=cropping_image,
                                    resolution_debug=resolution_debug)

            @staticmethod
            def render_bitwise_not_white(image, mask, name, resolution_debug):
                bitwise_not = cv2.bitwise_not(image, image, mask=mask)
                AnalyzeClass.render(name=f"_bitwise_not_white : {name}", source=bitwise_not,
                                    resolution_debug=resolution_debug)

            @staticmethod
            def render_bitwise_and(image, mask, name, resolution_debug):
                bitwise_and = cv2.bitwise_and(image, image, mask=mask)
                AnalyzeClass.render(name=f"bitwise_and : {name}", source=bitwise_and, resolution_debug=resolution_debug)

            @staticmethod
            def render_threshold(image, name, resolution_debug):
                _, threshold = cv2.threshold(image, 220, 255, cv2.THRESH_BINARY_INV)
                AnalyzeClass.render(name=f"threshold : {name}", source=threshold, resolution_debug=resolution_debug)

            @staticmethod
            def render_cvtcolor_to_hsv(image, name, resolution_debug):
                cvtcolor = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
                AnalyzeClass.render(name=f"cvtcolor_to_hsv : {name}", source=cvtcolor,
                                    resolution_debug=resolution_debug)

            @staticmethod
            def render_inrange(image, sensitivity_analysis, name, resolution_debug):
                cvtcolor = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
                inrange = cv2.inRange(cvtcolor, numpy.array([0, 0, 255 - sensitivity_analysis], dtype=numpy.uint8),
                                      numpy.array([255, sensitivity_analysis, 255], dtype=numpy.uint8))
                AnalyzeClass.render(name=f"inrange : {name}", source=inrange, resolution_debug=resolution_debug)

            @staticmethod
            def render_canny_edges(image, sensitivity_analysis, name, resolution_debug):
                canny = cv2.Canny(image, sensitivity_analysis, sensitivity_analysis, apertureSize=3, L2gradient=True)
                AnalyzeClass.render(name=f"canny_edges : {name}", source=canny, resolution_debug=resolution_debug)

            @staticmethod
            def render_shapes(name, resolution_debug):
                red = (0, 0, 255)
                green = (0, 255, 0)
                blue = (255, 0, 0)
                yellow = (0, 255, 255)
                numpy.set_printoptions(threshold=0)
                img = numpy.zeros(shape=(512, 512, 3), dtype=numpy.uint8)
                cv2.line(
                    img=img,
                    pt1=(0, 0),
                    pt2=(311, 511),
                    color=blue,
                    thickness=10
                )
                cv2.rectangle(
                    img=img,
                    pt1=(30, 166),
                    pt2=(130, 266),
                    color=green,
                    thickness=3
                )
                cv2.circle(
                    img=img,
                    center=(222, 222),
                    radius=50,
                    color=(255.111, 111),
                    thickness=-1
                )
                cv2.ellipse(
                    img=img,
                    center=(333, 333),
                    axes=(50, 20),
                    angle=0,
                    startAngle=0,
                    endAngle=150,
                    color=red,
                    thickness=-1
                )
                pts = numpy.array(
                    [[10, 5], [20, 30], [70, 20], [50, 10]],
                    dtype=numpy.int32
                )
                pts = pts.reshape((-1, 1, 2,))
                cv2.polylines(
                    img=img,
                    pts=[pts],
                    isClosed=True,
                    color=yellow,
                    thickness=5
                )
                cv2.putText(
                    img=img,
                    text="SOL",
                    org=(10, 400),
                    fontFace=cv2.FONT_ITALIC,
                    fontScale=3.5,
                    color=(255, 255, 255),
                    thickness=2
                )
                AnalyzeClass.render(name=f"shapes : {name}", source=img, resolution_debug=resolution_debug)

            @staticmethod
            def render_cvtcolor(image, color_type, name, resolution_debug):
                cvtcolor = cv2.cvtColor(image, color_type)
                AnalyzeClass.render(name=f"cvtcolor : {name}", source=cvtcolor, resolution_debug=resolution_debug)

            @staticmethod
            def render_flip(image, flipcode, name, resolution_debug):
                flip = cv2.flip(image, flipcode)
                AnalyzeClass.render(name=f"flip : {name}", source=flip, resolution_debug=resolution_debug)

            @staticmethod
            def render_final(image, mask, sensitivity_analysis, correct_coefficient, name, resolution_debug):
                bitwise_and = cv2.bitwise_and(image, image, mask=mask)
                cvtcolor = cv2.cvtColor(bitwise_and, cv2.COLOR_BGR2HSV)
                inrange = cv2.inRange(cvtcolor, numpy.array([0, 0, 255 - sensitivity_analysis], dtype=numpy.uint8),
                                      numpy.array([255, sensitivity_analysis, 255], dtype=numpy.uint8))
                cv2.putText(inrange,
                            f"{numpy.sum(inrange > 0) / numpy.sum(mask > 0) * 100 * correct_coefficient:0.2f}%",
                            (int(1920 / 5), int(1080 / 2)), cv2.FONT_HERSHEY_SIMPLEX, 3, (255, 255, 255), 3)
                AnalyzeClass.render(name=f"final : {name}", source=inrange, resolution_debug=resolution_debug)

            @staticmethod
            def render(name: str, source, resolution_debug: list):
                try:
                    if source is not None:
                        img = cv2.resize(source, (resolution_debug[0], resolution_debug[1]),
                                         interpolation=cv2.INTER_AREA)
                        cv2.imshow(name, img)
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(ex)

            @staticmethod
            def result_final(image, mask, sensitivity_analysis, correct_coefficient):
                bitwise_and = cv2.bitwise_and(image, image, mask=mask)
                cvtcolor = cv2.cvtColor(bitwise_and, cv2.COLOR_BGR2HSV)
                inrange = cv2.inRange(cvtcolor, numpy.array([0, 0, 255 - sensitivity_analysis], dtype=numpy.uint8),
                                      numpy.array([255, sensitivity_analysis, 255], dtype=numpy.uint8))
                try:
                    return round(numpy.sum(inrange > 0) / numpy.sum(mask > 0) * 100 * correct_coefficient, 2)
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'result_final func error : {ex}')
                    return 0.0

            @staticmethod
            def write_result(ip_sql: str, server_sql: str, port_sql: str, database_sql: str, user_sql: str,
                             password_sql: str,
                             sql_now_check: bool, table_now_sql: str, rows_now_sql: list, sql_data_check: bool,
                             table_data_sql: str, rows_data_sql: list, source: str, values: float, widget,
                             widget_write: bool,
                             text_write: bool, sql_val: bool, alarm_level: int):
                sql_datetime = datetime.datetime.now().strftime('%Y-%m-%dT%H:%M:%S')
                boolean = 0
                if values > alarm_level:
                    boolean = 1
                _values = [source, values, boolean, sql_datetime]
                if widget_write:
                    try:
                        widget(f'{_values}')
                    except Exception as ex:
                        LoggingClass.logging(ex)
                        print(ex)
                if text_write:
                    try:
                        with open('db.txt', 'a') as db:
                            db.write(f'{_values}\n')
                    except Exception as ex:
                        LoggingClass.logging(ex)
                        print(ex)
                if sql_val:
                    try:
                        if sql_now_check:
                            SQLClass.sql_post_now(ip=ip_sql, server=server_sql, port=port_sql, database=database_sql,
                                                  username=user_sql, password=password_sql, table=table_now_sql,
                                                  rows=rows_now_sql, values=_values)
                        if sql_data_check:
                            SQLClass.sql_post_data(ip=ip_sql, server=server_sql, port=port_sql, database=database_sql,
                                                   username=user_sql, password=password_sql, table=table_data_sql,
                                                   rows=rows_data_sql, values=_values)
                    except Exception as ex:
                        LoggingClass.logging(ex)
                        print(ex)

            @staticmethod
            def make_snapshot(data: dict):
                h = httplib2.Http("/path/to/cache-directory")
                h.add_credentials(name=data['login_cam'], password=data['password_cam'])
                sources = f"{data['protocol_cam_type']}://192.168.{data['ip_cam_snapshot']}:{data['port_cam']}/" \
                          f"ISAPI/Streaming/channels/101/picture?snapShotImageType=JPEG"
                response, content = h.request(sources)
                with open(data["name_snapshot"], 'wb') as file:
                    file.write(content)
                # cv2.imwrite('1.png', img, [int(cv2.IMWRITE_JPEG_QUALITY), 95])
                # cv2.imwrite('1.png', img, [int(cv2.IMWRITE_PNG_COMPRESSION), 9])

            @staticmethod
            def play_rtsp():
                cap = cv2.VideoCapture('rtsp://admin:q1234567@192.168.15.229:554/cam/realmonitor?channel=1&subtype=0')
                # cap = cv2.VideoCapture('rtsp://admin:nehrfvths123@192.168.15.140:554')
                # cap = cv2.VideoCapture('rtsp://192.168.15.229:554/cam/realmonitor?channel=1&subtype=0&unicast=true&proto=Onvif')
                # cap = cv2.VideoCapture('rtsp://192.168.15.229:554/live')
                # cap = cv2.VideoCapture('rtsp://admin:q1234567@192.168.15.229:554/cam/realmonitor?channel=2&subtype=1')
                # cap = cv2.VideoCapture('rtsp://admin:q1234567@192.168.15.229:554/cam/realmonitor?channel=33&subtype=0')
                # sources = f"http://192.168.15.227/cgi-bin/snapshot.cgi?channel=road?loginuse=admin&loginpas=q1234567"
                # sources = f"http://192.168.15.227/cgi-bin/snapshot.cgi?loginuse=admin&loginpas=q1234567"
                # sources = f"http://192.168.15.227/cgi-bin/snapshot.cgi?chn=1&u=admin&p=q1234567"
                # sources = f"http://192.168.15.227/cgi-bin/snapshot.cgi?1"
                # sources = f"rtsp://192.168.15.227:554/cam/realmonitor?channel=1&subtype=0&unicast=true&proto=Onvif"
                # sources = f"rtsp://192.168.15.227:554/live"
                # sources = f"rtsp://admin:q1234567@192.168.15.227:554/cam/realmonitor?channel=1&subtype=1"
                # sources = f"rtsp://admin:q1234567@192.168.15.227:554/cam/realmonitor?channel=1&subtype=0"
                # sources = f"rtsp://admin:nehrfvths123@192.168.15.140:554"
                while True:
                    try:
                        ret, frame = cap.read()
                        cv2.imshow("Capturing", frame)
                        if cv2.waitKey(1) & 0xFF == ord('q'):
                            break
                    except Exception as ex:
                        print(ex)
                cv2.destroyAllWindows()
                cap.release()

        # UI
        class AppContainerClass:
            def __init__(self):
                self.app = QtWidgets.QApplication([])
                self.widget = None

            def create_ui(self, title, width, height, icon, play_f, stop_f, quit_f, snapshot_f):
                self.widget = MainWidgetClass(title, width, height, icon, play_f, stop_f, quit_f, snapshot_f)
                return self.widget

            @staticmethod
            def create_qlable(text: str, _parent, background=False):
                _widget = QtWidgets.QLabel(text)
                if background:
                    _widget.setAutoFillBackground(True)
                    _widget.setStyleSheet("QLabel { background-color: rgba(0, 0, 0, 255); color : white; }")
                _parent.addWidget(_widget)
                return _widget

            @staticmethod
            def create_qpushbutton(_parent, _connect_func, _text='set'):
                _widget = QtWidgets.QPushButton(_text)
                _parent.addWidget(_widget)
                _widget.clicked.connect(_connect_func)
                return _widget

            @staticmethod
            def create_qcheckbox(_parent, _text='check?', default=False):
                _widget = QtWidgets.QCheckBox(_text)
                _widget.setChecked(default)
                _parent.addWidget(_widget)
                return _widget

            @staticmethod
            def create_qcombobox(_parent, _text: list, default=None):
                _widget = QtWidgets.QComboBox()
                _widget.addItems([x for x in _text])
                _widget.setCurrentText(default)
                _parent.addWidget(_widget)
                return _widget

            @staticmethod
            def create_qradiobutton(_parent, _text: str, default=False):
                _widget = QtWidgets.QRadioButton(_text)
                _widget.setChecked(default)
                _parent.addWidget(_widget)
                return _widget

        class MainWidgetClass(QtWidgets.QWidget):
            def __init__(self, title="APP", width=640, height=480, icon="", play_f=None, stop_f=None, quit_f=None,
                         snapshot_f=None):
                super().__init__()

                self.play_f = play_f
                self.snapshot_f = snapshot_f
                self.resize(width, height)
                self.setWindowTitle(title)
                self.setWindowIcon(QtGui.QIcon(icon))
                self.resolution_debug = []
                self.v_layout_m = QtWidgets.QVBoxLayout(self)

                # MANAGEMENT
                self.h_layout_g_management = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_management)
                self.g_management_set = AppContainerClass.create_qlable('MANAGEMENT', self.h_layout_g_management,
                                                                        background=True)
                self.h_layout_management_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_management_1)
                self.play_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_management_1,
                                                                             self.play_btn_func,
                                                                             'play')
                self.stop_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_management_1, stop_f, 'stop')
                self.quit_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_management_1, quit_f, 'quit')
                # CAMERAS
                self.h_layout_g_cam = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_cam)
                self.g_cam_set = AppContainerClass.create_qlable('CAMERAS', self.h_layout_g_cam, background=True)
                self.h_layout_cam_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_1)
                self.protocol_cam_type = AppContainerClass.create_qlable('PROTOCOL TYPE : http', self.h_layout_cam_1)
                self.set_protocol_cam_type = AppContainerClass.create_qpushbutton(self.h_layout_cam_1,
                                                                                  self.get_protocol_cam_type_button)
                self.port_cam = AppContainerClass.create_qlable('PORT CAM : 80', self.h_layout_cam_1)
                self.set_port_cam = AppContainerClass.create_qpushbutton(self.h_layout_cam_1, self.get_port_cam_button)
                self.login_cam = AppContainerClass.create_qlable('LOGIN CAM : admin', self.h_layout_cam_1)
                self.set_login_cam = AppContainerClass.create_qpushbutton(self.h_layout_cam_1,
                                                                          self.get_login_cam_button)
                self.password_cam = AppContainerClass.create_qlable('PASSWORD CAM : q1234567', self.h_layout_cam_1)
                self.set_password_cam = AppContainerClass.create_qpushbutton(self.h_layout_cam_1,
                                                                             self.get_password_cam_button)
                self.h_layout_cam_2 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_2)
                self.alias_device = AppContainerClass.create_qlable(
                    'ALIAS DEVICE : 16/1 | 16/2 | 16/3 | 16/4 | 16/5 | 16/6 '
                    '| 16/7 | 16/8 | 16/9 | 16/10', self.h_layout_cam_2)
                self.set_alias_device = AppContainerClass.create_qpushbutton(self.h_layout_cam_2,
                                                                             self.get_alias_device_button)
                self.h_layout_cam_3 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_3)
                self.ip_cam = AppContainerClass.create_qlable(
                    'IP CAM : 15.202 | 15.206 | 15.207 | 15.208 | 15.209 | 15.210 '
                    '| 15.211 | 15.203 | 15.204 | 15.205', self.h_layout_cam_3)
                self.set_ip_cam = AppContainerClass.create_qpushbutton(self.h_layout_cam_3, self.get_ip_cam_button)
                self.h_layout_cam_4 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_4)
                self.mask_cam = AppContainerClass.create_qlable(
                    'MASK CAM : m_16_1.jpg | m_16_2.jpg | m_16_3.jpg | m_16_4.jpg '
                    '| m_16_5.jpg | m_16_6.jpg | m_16_7.jpg | m_16_8.jpg '
                    '| m_16_9.jpg | m_16_10.jpg', self.h_layout_cam_4)
                self.set_mask_cam = AppContainerClass.create_qpushbutton(self.h_layout_cam_4, self.get_mask_cam_button)
                self.h_layout_cam_5 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_5)
                self.sensitivity_analysis = AppContainerClass.create_qlable(
                    'SENSITIVITY ANALYSIS : 115 | 115 | 115 | 115 '
                    '| 115 | 115 | 115 | 115 | 115 | 115',
                    self.h_layout_cam_5)
                self.set_sensitivity_analysis = AppContainerClass.create_qpushbutton(self.h_layout_cam_5,
                                                                                     self.get_sensitivity_analysis_button)
                self.h_layout_cam_6 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_6)
                self.alarm_level = AppContainerClass.create_qlable('ALARM LEVEL : 30 | 30 | 30 | 30 '
                                                                   '| 30 | 30 | 30 | 30 | 30 | 30',
                                                                   self.h_layout_cam_6)
                self.set_alarm_level = AppContainerClass.create_qpushbutton(self.h_layout_cam_6,
                                                                            self.get_alarm_level_button)
                self.h_layout_cam_7 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_cam_7)
                self.correct_coefficient = AppContainerClass.create_qlable(
                    'CORRECT COEFFICIENT : 1.0 | 1.0 | 1.0 | 1.0 | 1.0 '
                    '| 1.0 | 1.0 | 1.0 | 1.0 | 1.0', self.h_layout_cam_7)
                self.set_correct_coefficient = AppContainerClass.create_qpushbutton(self.h_layout_cam_7,
                                                                                    self.get_correct_coefficient_button)
                # SQL
                self.h_layout_g_sql = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_sql)
                self.g_sql_set = AppContainerClass.create_qlable('SQL', self.h_layout_g_sql, background=True)
                self.h_layout_sql_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_sql_1)
                self.sql_write = AppContainerClass.create_qcheckbox(self.h_layout_sql_1, 'CONNECT TO SQL?')
                self.ip_sql = AppContainerClass.create_qlable('IP SQL SERVER : 192.168.15.87', self.h_layout_sql_1)
                self.set_ip_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_1, self.get_ip_sql_button)
                self.server_sql = AppContainerClass.create_qlable(r'SERVER SQL : DESKTOP-SM7K050\COMPUTER_VISION',
                                                                  self.h_layout_sql_1)
                self.set_server_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_1,
                                                                           self.get_server_sql_button)
                self.port_sql = AppContainerClass.create_qlable('PORT SQL SERVER : 1433', self.h_layout_sql_1)
                self.set_port_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_1, self.get_port_sql_button)
                self.h_layout_sql_2 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_sql_2)
                self.database_sql = AppContainerClass.create_qlable('DATABASE SQL : analiz_16grohot',
                                                                    self.h_layout_sql_2)
                self.set_database_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_2,
                                                                             self.get_database_sql_button)
                self.user_sql = AppContainerClass.create_qlable('USER SQL : computer_vision', self.h_layout_sql_2)
                self.set_user_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_2, self.get_user_sql_button)
                self.h_layout_sql_3 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_sql_3)
                self.password_sql = AppContainerClass.create_qlable('PASSWORD SQL : vision12345678',
                                                                    self.h_layout_sql_2)
                self.set_password_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_2,
                                                                             self.get_password_sql_button)
                self.sql_now_check = AppContainerClass.create_qcheckbox(self.h_layout_sql_3, 'WRITE NOW SQL?')
                self.table_now_sql = AppContainerClass.create_qlable('TABLE NOW SQL : grohot16_now_table',
                                                                     self.h_layout_sql_3)
                self.set_table_now_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_3,
                                                                              self.get_table_now_sql_button)
                self.rows_now_sql = AppContainerClass.create_qlable('ROWS NOW SQL : device_row | value_row | alarm_row '
                                                                    '| datetime_row', self.h_layout_sql_3)
                self.set_rows_now_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_3,
                                                                             self.get_rows_now_sql_button)
                self.h_layout_sql_4 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_sql_4)
                self.sql_data_check = AppContainerClass.create_qcheckbox(self.h_layout_sql_4, 'WRITE DATA SQL?')
                self.table_data_sql = AppContainerClass.create_qlable('TABLE DATA SQL : grohot16_data_table',
                                                                      self.h_layout_sql_4)
                self.set_table_data_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_4,
                                                                               self.get_table_data_sql_button)
                self.rows_data_sql = AppContainerClass.create_qlable(
                    'ROWS DATA SQL : device_row | value_row | alarm_row '
                    '| datetime_row', self.h_layout_sql_4)
                self.set_rows_data_sql = AppContainerClass.create_qpushbutton(self.h_layout_sql_4,
                                                                              self.get_rows_data_sql_button)
                # DEBUG
                self.h_layout_g_debug = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_debug)
                self.g_debug_set = AppContainerClass.create_qlable('DEBUG', self.h_layout_g_debug, background=True)
                self.h_layout_debug_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_debug_1)
                self.auto_import_check = AppContainerClass.create_qcheckbox(self.h_layout_debug_1, 'AUTO IMPORT?')
                self.auto_play_check = AppContainerClass.create_qcheckbox(self.h_layout_debug_1, 'AUTO PLAY?')
                self.speed_analysis = AppContainerClass.create_qlable('SPEED ANALYSIS : 1.0', self.h_layout_debug_1)
                self.set_speed_analysis = AppContainerClass.create_qpushbutton(self.h_layout_debug_1,
                                                                               self.get_speed_analysis_button)
                self.speed_video_stream = AppContainerClass.create_qlable('SPEED VIDEO-STREAM : 1.0',
                                                                          self.h_layout_debug_1)
                self.set_speed_video_stream = AppContainerClass.create_qpushbutton(self.h_layout_debug_1,
                                                                                   self.get_speed_video_stream_button)
                self.h_layout_debug_2 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_debug_2)
                self.widget_data_value = AppContainerClass.create_qlable('0.00%', self.h_layout_debug_2)
                self.h_layout_debug_2.addStretch()
                self.widget_write = AppContainerClass.create_qcheckbox(self.h_layout_debug_2, 'WRITE TO WIDGET?')
                self.text_write = AppContainerClass.create_qcheckbox(self.h_layout_debug_2, 'WRITE TO TEXT?')
                self.source_win_type = AppContainerClass.create_qlable('SOURCE TYPE :', self.h_layout_debug_2)
                self.source_type = AppContainerClass.create_qcombobox(self.h_layout_debug_2,
                                                                      ['image-http', 'video-rtsp', 'video-file'],
                                                                      'image-http')
                self.compute_win_debug = AppContainerClass.create_qlable('COMPUTE TYPE :', self.h_layout_debug_2)
                self.compute_debug = AppContainerClass.create_qcombobox(self.h_layout_debug_2,
                                                                        ['sync', 'async', 'multithread',
                                                                         'multiprocess', 'complex'],
                                                                        'complex')
                self.process_cores = AppContainerClass.create_qlable('PROCESS CORES : 4', self.h_layout_debug_2)
                self.set_process_cores = AppContainerClass.create_qpushbutton(self.h_layout_debug_2,
                                                                              self.get_process_cores_button)
                self.h_layout_debug_3 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_debug_3)
                self.render_win_debug = AppContainerClass.create_qlable('Render windows :', self.h_layout_debug_3)
                self.render_debug = AppContainerClass.create_qcombobox(self.h_layout_debug_3,
                                                                       ['none', 'source', 'final', 'extended',
                                                                        'all'], 'none')
                self.resolution_debug_1 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '320x240',
                                                                                default=True)
                self.resolution_debug_1.toggled.connect(self.set_resolution_debug(self.resolution_debug_1, 320, 240))
                self.resolution_debug_2 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '640x480')
                self.resolution_debug_2.toggled.connect(self.set_resolution_debug(self.resolution_debug_2, 640, 480))
                self.resolution_debug_3 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '1280x720')
                self.resolution_debug_3.toggled.connect(self.set_resolution_debug(self.resolution_debug_3, 1280, 720))
                self.resolution_debug_4 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '1920x1080')
                self.resolution_debug_4.toggled.connect(self.set_resolution_debug(self.resolution_debug_4, 1920, 1080))
                self.resolution_debug_5 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '2560x1600')
                self.resolution_debug_5.toggled.connect(self.set_resolution_debug(self.resolution_debug_5, 2560, 1600))
                self.resolution_debug_6 = AppContainerClass.create_qradiobutton(self.h_layout_debug_3, '3840x2160')
                self.resolution_debug_6.toggled.connect(self.set_resolution_debug(self.resolution_debug_6, 3840, 2160))
                # IMPORTS
                self.h_layout_g_imports = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_imports)
                self.g_imports_set = AppContainerClass.create_qlable('IMPORTS', self.h_layout_g_imports,
                                                                     background=True)
                self.h_layout_imports_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_imports_1)
                self.import_file = AppContainerClass.create_qlable('SETTINGS FILE NAME : settings',
                                                                   self.h_layout_imports_1)
                self.set_import_file = AppContainerClass.create_qpushbutton(self.h_layout_imports_1,
                                                                            self.get_settings_file_name)
                self.export_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_imports_1,
                                                                               self.export_settings_func, 'export')
                self.import_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_imports_1,
                                                                               self.import_settings_func, 'import')
                # SHOT
                self.h_layout_g_shot = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_g_shot)
                self.g_shot_set = AppContainerClass.create_qlable('SHOT', self.h_layout_g_shot, background=True)
                self.h_layout_shot_1 = QtWidgets.QHBoxLayout()
                self.v_layout_m.addLayout(self.h_layout_shot_1)
                self.ip_cam_snapshot = AppContainerClass.create_qlable('ip-cam : 15.204', self.h_layout_shot_1)
                self.set_ip_cam_snapshot = AppContainerClass.create_qpushbutton(self.h_layout_shot_1,
                                                                                self.get_ip_cam_snapshot_button)
                self.name_snapshot = AppContainerClass.create_qlable('file name : picture.jpg', self.h_layout_shot_1)
                self.set_name_snapshot = AppContainerClass.create_qpushbutton(self.h_layout_shot_1,
                                                                              self.set_name_snapshot_button)
                self.snapshot_QPushButton = AppContainerClass.create_qpushbutton(self.h_layout_shot_1,
                                                                                 self.snapshot_btn_func,
                                                                                 'snapshot')

                self.setLayout(self.v_layout_m)
                self.auto_play_func()
                self.auto_import_settings_func()

            def get_speed_analysis_button(self):
                widget = self.speed_analysis
                value, success = QtWidgets.QInputDialog.getDouble(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                  f'{widget.text().split(":")[0].strip()} value:',
                                                                  1.0, 0.01, 50.0, 2)
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_speed_video_stream_button(self):
                widget = self.speed_video_stream
                value, success = QtWidgets.QInputDialog.getDouble(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                  f'{widget.text().split(":")[0].strip()} value:',
                                                                  1.0, 0.01, 50.0, 2)
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_sensitivity_analysis_button(self):
                widget = self.sensitivity_analysis
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_alarm_level_button(self):
                widget = self.alarm_level
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_correct_coefficient_button(self):
                widget = self.correct_coefficient
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_protocol_cam_type_button(self):
                widget = self.protocol_cam_type
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_port_cam_button(self):
                widget = self.port_cam
                value, success = QtWidgets.QInputDialog.getInt(self, f'Set {widget.text().split(":")[0].strip()}',
                                                               f'{widget.text().split(":")[0].strip()} value:',
                                                               554, 1, 9999, 5)
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_login_cam_button(self):
                widget = self.login_cam
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_password_cam_button(self):
                widget = self.password_cam
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_alias_device_button(self):
                widget = self.alias_device
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_ip_cam_button(self):
                widget = self.ip_cam
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_mask_cam_button(self):
                widget = self.mask_cam
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_ip_sql_button(self):
                widget = self.ip_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_server_sql_button(self):
                widget = self.server_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_port_sql_button(self):
                widget = self.port_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_database_sql_button(self):
                widget = self.database_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_user_sql_button(self):
                widget = self.user_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_password_sql_button(self):
                widget = self.password_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_table_now_sql_button(self):
                widget = self.table_now_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_rows_now_sql_button(self):
                widget = self.rows_now_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_table_data_sql_button(self):
                widget = self.table_data_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_rows_data_sql_button(self):
                widget = self.rows_data_sql
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_settings_file_name(self):
                widget = self.import_file
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def set_resolution_debug(self, radio, width: int, height: int):
                self.resolution_debug.append([radio, width, height])

            def get_window_resolution(self):
                for radio in self.resolution_debug:
                    try:
                        if radio[0].isChecked():
                            return [int(radio[1]), int(radio[2])]
                    except Exception as ex:
                        print(ex)

            def set_window_resolution(self, value):
                for radio in self.resolution_debug:
                    try:
                        if radio[1] == value[0]:
                            radio[0].setChecked(True)
                        else:
                            radio[0].setChecked(False)
                    except Exception as ex:
                        print(ex)

            def get_process_cores_button(self):
                widget = self.process_cores
                value, success = QtWidgets.QInputDialog.getInt(self, f'Set {widget.text().split(":")[0].strip()}',
                                                               f'{widget.text().split(":")[0].strip()} value:',
                                                               4, 1, 16, 5)
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_ip_cam_snapshot_button(self):
                widget = self.ip_cam_snapshot
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def set_name_snapshot_button(self):
                widget = self.name_snapshot
                value, success = QtWidgets.QInputDialog.getText(self, f'Set {widget.text().split(":")[0].strip()}',
                                                                f'{widget.text().split(":")[0].strip()} value:',
                                                                text=f'{widget.text().split(":")[1].strip()}')
                if success:
                    widget.setText(f'{widget.text().split(":")[0].strip()} : {str(value)}')

            def get_string_from_list(self, source: list):
                value = ''
                for x in source:
                    value = f'{value} | {x}'
                return value[3::]

            def set_data_func(self, value: str):
                try:
                    self.widget_data_value.setText(f"{value}")
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'set_data_func error : {ex}')

            def create_data_func(self):
                try:
                    data = {
                        'protocol_cam_type': str(self.protocol_cam_type.text().split(':')[1].strip()),
                        'port_cam': int(self.port_cam.text().split(':')[1].strip()),
                        'login_cam': str(self.login_cam.text().split(':')[1].strip()),
                        'password_cam': str(self.password_cam.text().split(':')[1].strip()),
                        'alias_device': list(
                            [x.strip() for x in self.alias_device.text().split(':')[1].strip().split('|')]),
                        'ip_cam': list([x.strip() for x in self.ip_cam.text().split(':')[1].strip().split('|')]),
                        'mask_cam': list([x.strip() for x in self.mask_cam.text().split(':')[1].strip().split('|')]),
                        'sensitivity_analysis': list([int(x.strip()) for x in
                                                      self.sensitivity_analysis.text().split(':')[1].strip().split(
                                                          '|')]),
                        'alarm_level': list(
                            [int(x.strip()) for x in self.alarm_level.text().split(':')[1].strip().split('|')]),
                        'correct_coefficient': list([float(x.strip()) for x in
                                                     self.correct_coefficient.text().split(':')[1].strip().split('|')]),

                        'sql_write': bool(self.sql_write.isChecked()),
                        'ip_sql': str(self.ip_sql.text().split(':')[1].strip()),
                        'server_sql': str(self.server_sql.text().split(':')[1].strip()),
                        'port_sql': str(self.port_sql.text().split(':')[1].strip()),
                        'database_sql': str(self.database_sql.text().split(':')[1].strip()),
                        'user_sql': str(self.user_sql.text().split(':')[1].strip()),
                        'password_sql': str(self.password_sql.text().split(':')[1].strip()),
                        'sql_now_check': bool(self.sql_now_check.isChecked()),
                        'table_now_sql': str(self.table_now_sql.text().split(':')[1].strip()),
                        'rows_now_sql': list(
                            [x.strip() for x in self.rows_now_sql.text().split(':')[1].strip().split('|')]),
                        'sql_data_check': bool(self.sql_data_check.isChecked()),
                        'table_data_sql': str(self.table_data_sql.text().split(':')[1].strip()),
                        'rows_data_sql': list(
                            [x.strip() for x in self.rows_data_sql.text().split(':')[1].strip().split('|')]),

                        'auto_import_check': bool(self.auto_import_check.isChecked()),
                        'auto_play_check': bool(self.auto_play_check.isChecked()),
                        'speed_analysis': float(self.speed_analysis.text().split(':')[1].strip()),
                        'speed_video_stream': float(self.speed_video_stream.text().split(':')[1].strip()),

                        'widget_write': bool(self.widget_write.isChecked()),
                        'text_write': bool(self.text_write.isChecked()),
                        'widget': self.set_data_func,
                        'source_type': str(self.source_type.currentText().strip()),
                        'compute_debug': str(self.compute_debug.currentText().strip()),
                        'process_cores': int(self.process_cores.text().split(':')[1].strip()),
                        'render_debug': str(self.render_debug.currentText().strip()),
                        'resolution_debug': list(self.get_window_resolution()),

                        'import_file': str(self.import_file.text().split(':')[1].strip()),

                        'ip_cam_snapshot': str(self.ip_cam_snapshot.text().split(":")[1].strip()),
                        'name_snapshot': str(self.name_snapshot.text().split(":")[1].strip()),
                    }
                    return data
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'create_data_func error : {ex}')

            def play_btn_func(self):
                try:
                    data = self.create_data_func()
                    self.play_f(data=data)
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'play_btn_func error : {ex}')

            def snapshot_btn_func(self):
                try:
                    data = self.create_data_func()
                    self.snapshot_f(data=data)
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'snapshot_btn_func error : {ex}')

            def export_settings_func(self):
                try:
                    data = self.create_data_func()
                    del data['widget']
                    FileSettings.export_settings(data)
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'export_settings_func error : {ex}')

            def import_settings_func(self):
                try:
                    data = self.create_data_func()
                    data = FileSettings.import_settings(data)
                    self.protocol_cam_type.setText(f'{self.protocol_cam_type.text().split(":")[0].strip()} : '
                                                   f'{str(data["protocol_cam_type"])}')
                    self.port_cam.setText(f'{self.port_cam.text().split(":")[0].strip()} : '
                                          f'{str(data["port_cam"])}')
                    self.login_cam.setText(f'{self.login_cam.text().split(":")[0].strip()} : '
                                           f'{str(data["login_cam"])}')
                    self.password_cam.setText(f'{self.password_cam.text().split(":")[0].strip()} : '
                                              f'{str(data["password_cam"])}')
                    self.alias_device.setText(f'{self.alias_device.text().split(":")[0].strip()} : '
                                              f'{self.get_string_from_list(data["alias_device"])}')
                    self.ip_cam.setText(f'{self.ip_cam.text().split(":")[0].strip()} : '
                                        f'{self.get_string_from_list(data["ip_cam"])}')
                    self.mask_cam.setText(f'{self.mask_cam.text().split(":")[0].strip()} : '
                                          f'{self.get_string_from_list(data["mask_cam"])}')
                    self.sensitivity_analysis.setText(f'{self.sensitivity_analysis.text().split(":")[0].strip()} : '
                                                      f'{self.get_string_from_list(data["sensitivity_analysis"])}')
                    self.alarm_level.setText(f'{self.alarm_level.text().split(":")[0].strip()} : '
                                             f'{self.get_string_from_list(data["alarm_level"])}')
                    self.correct_coefficient.setText(f'{self.correct_coefficient.text().split(":")[0].strip()} : '
                                                     f'{self.get_string_from_list(data["correct_coefficient"])}')
                    self.sql_write.setChecked(data["sql_write"])
                    self.ip_sql.setText(f'{self.ip_sql.text().split(":")[0].strip()} : {str(data["ip_sql"])}')
                    self.server_sql.setText(
                        f'{self.server_sql.text().split(":")[0].strip()} : {str(data["server_sql"])}')
                    self.port_sql.setText(f'{self.port_sql.text().split(":")[0].strip()} : {str(data["port_sql"])}')
                    self.database_sql.setText(f'{self.database_sql.text().split(":")[0].strip()} : '
                                              f'{str(data["database_sql"])}')
                    self.user_sql.setText(f'{self.user_sql.text().split(":")[0].strip()} : '
                                          f'{str(data["user_sql"])}')
                    self.password_sql.setText(f'{self.password_sql.text().split(":")[0].strip()} : '
                                              f'{str(data["password_sql"])}')
                    self.sql_now_check.setChecked(data["sql_now_check"])
                    self.table_now_sql.setText(f'{self.table_now_sql.text().split(":")[0].strip()} : '
                                               f'{str(data["table_now_sql"])}')
                    self.rows_now_sql.setText(f'{self.rows_now_sql.text().split(":")[0].strip()} : '
                                              f'{self.get_string_from_list(data["rows_now_sql"])}')
                    self.sql_data_check.setChecked(data["sql_data_check"])
                    self.table_data_sql.setText(f'{self.table_data_sql.text().split(":")[0].strip()} : '
                                                f'{str(data["table_data_sql"])}')
                    self.rows_data_sql.setText(f'{self.rows_data_sql.text().split(":")[0].strip()} : '
                                               f'{self.get_string_from_list(data["rows_data_sql"])}')
                    self.auto_import_check.setChecked(data["auto_import_check"])
                    self.auto_play_check.setChecked(data["auto_play_check"])
                    self.speed_analysis.setText(f'{self.speed_analysis.text().split(":")[0].strip()} : '
                                                f'{str(data["speed_analysis"])}')
                    self.speed_video_stream.setText(f'{self.speed_video_stream.text().split(":")[0].strip()} : '
                                                    f'{str(data["speed_video_stream"])}')
                    self.widget_write.setChecked(data["widget_write"])
                    self.text_write.setChecked(data["text_write"])
                    self.source_type.setCurrentText(data["source_type"])
                    self.compute_debug.setCurrentText(data["compute_debug"])
                    self.process_cores.setText(f'{self.process_cores.text().split(":")[0].strip()} : '
                                               f'{str(data["process_cores"])}')
                    self.render_debug.setCurrentText(data["render_debug"])
                    self.set_window_resolution(data["resolution_debug"])
                    self.import_file.setText(f'{self.import_file.text().split(":")[0].strip()} : '
                                             f'{str(data["import_file"])}')
                    self.ip_cam_snapshot.setText(f'{self.ip_cam_snapshot.text().split(":")[0].strip()} : '
                                                 f'{str(data["ip_cam_snapshot"])}')
                    self.name_snapshot.setText(f'{self.name_snapshot.text().split(":")[0].strip()} : '
                                               f'{str(data["name_snapshot"])}')
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'import_settings_func error : {ex}')

            def auto_import_settings_func(self):
                try:
                    data = self.create_data_func()
                    data = FileSettings.import_settings(data)
                    if data['auto_import_check']:
                        self.import_settings_func()
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'auto_import_settings_func error : {ex}')

            def auto_play_func(self):
                try:
                    data = self.create_data_func()
                    _data = FileSettings.import_settings(data)
                    _data['widget'] = data['widget']
                    if _data['auto_play_check']:
                        self.play_f(data=_data)
                        self.showMinimized()
                except Exception as ex:
                    LoggingClass.logging(ex)
                    print(f'auto_play_func error : {ex}')

        def play_func(data: dict):
            global play
            try:
                play = True

                def play_analyze():
                    AnalyzeClass.start_analyze(data=CopyDictionary.get_all_sources(data, {'pause': pause}))

                threading.Thread(target=play_analyze, args=()).start()
            except Exception as ex:
                print(ex)
                with open('log.txt', 'a') as log:
                    log.write(f'\n{ex}\n')

        def stop_func():
            global play
            play = False

        def pause():
            global play
            return play

        def quit_func():
            stop_func()
            sys.exit(app_container.app.exec())

        def snapshot_func(data: dict):
            threading.Thread(target=AnalyzeClass.make_snapshot, args=(data,)).start()

        # MAIN
        if __name__ == "__main__":
            freeze_support()
            play = True
            app_container = AppContainerClass()
            widget = app_container.create_ui(title="analysis", width=300, height=300, icon="icon.ico",
                                             play_f=play_func, stop_f=stop_func, quit_f=quit_func,
                                             snapshot_f=snapshot_func)
            ui_thread = threading.Thread(target=widget.show())
            sys.exit(app_container.app.exec())


class LoggingClass:
    @staticmethod
    def logging(message, file_name='log.txt', type_write='a'):
        with open(file_name, type_write) as log:
            log.write(f'{TimeUtils.get_current_time()} : {message}\n')


class CopyDictionary:
    @staticmethod
    def get_all_sources(source: dict, values: dict):
        value = source.copy()
        for _key, _value in values.items():
            value[_key] = _value
        return value


class FileSettings:
    @staticmethod
    def export_settings(data: dict):
        with open(f"{data['import_file']}.json", 'w') as file:
            json.dump(data, file)

    @staticmethod
    def import_settings(data: dict):
        try:
            with open(f"{data['import_file']}.json", "r") as read_file:
                data = json.load(read_file)
        except Exception as ex:
            LoggingClass.logging(ex)
            print(f'import_settings error : {ex}')
        return data


class TimeUtils:
    @staticmethod
    def get_current_time():
        return f"{time.strftime('%X')}"


class SendMail:
    @staticmethod
    def sender_email(subject='subj', text='text'):
        host = 'smtp.yandex.ru'
        port = '465'
        login = 'eevee.cycle'
        password = '31284bogdan'
        writer = 'eevee.cycle@yandex.ru'
        recipient = 'eevee.cycle@yandex.ru'

        message = f"""From: {recipient}\nTo: {writer}\nSubject: {subject}\n\n{text}"""

        smtpobj = smtplib.SMTP_SSL(host=host, port=port)
        smtpobj.ehlo()
        smtpobj.login(user=login, password=password)
        smtpobj.sendmail(from_addr=writer, to_addrs=recipient, msg=message)
        smtpobj.quit()


if __name__ == '__main__':
    pass
