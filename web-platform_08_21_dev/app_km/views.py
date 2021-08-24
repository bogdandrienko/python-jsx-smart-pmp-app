from django.shortcuts import render, redirect, get_object_or_404
from django.http.response import Http404, HttpResponseRedirect
from django.http import HttpResponse
from django.urls import reverse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.mail import BadHeaderError, send_mail
from django.conf import settings
from django.template.loader import get_template
from django.contrib import admin
from django.contrib.auth.models import Group, User
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login, authenticate, logout

from .models import RationalModel, CategoryRationalModel, LikeRationalModel, CommentRationalModel, \
    ApplicationModuleModel, ApplicationComponentModel, AccountDataModel, NotificationModel, EmailModel, ContactModel, \
    DocumentModel, MessageModel, CityModel, ArticleModel, SmsModel
from .forms import CreateUserForm, ChangeUserForm, CreateUsersForm, GeneratePasswordsForm, RationalForm, \
    NotificationForm, MessageForm, DocumentForm, ContactForm, CityForm, ArticleForm, \
    SmsForm, GeoForm
from .service import AuthorizationClass, PaginationClass, HttpRaiseExceptionClass, LoggingClass, \
    create_encrypted_password, get_users, get_salary_data, link_callback, get_career, generate_xlsx, \
    generate_kml, get_distance, find_distance

import requests
import openpyxl
from xhtml2pdf import pisa


# Admin
def admin_(request):
    return render(request, admin.site.urls)


# Home
def home(request):
    return render(request, 'components/home.html')


# Account
def login_account(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            username = request.POST['username']
            password = request.POST['password']
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('app_km:home')
            return redirect('app_km:create')
    form = AuthenticationForm()
    context = {
        'form': form,
    }
    return render(request, 'app_km/login_user.html', context)


def logout_account(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    logout(request)
    return redirect('app_km:login')


def create_user(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    if request.method == 'POST':
        form = CreateUserForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            signup_user = User.objects.get(username=username)
            try:
                user_group = Group.objects.get(name='User')
            except Exception as ex:
                print(ex)
                user_group = Group.objects.create_notification(name='User')
            user_group.user_set.add(signup_user)

        result_create_user = True
    else:
        result_create_user = False
    create_user_form = CreateUserForm()
    context = {
        'create_user_form': create_user_form,
        'result_create_user': result_create_user,
    }
    return render(request, 'app_km/create_user.html', context)


def create_user_from1c(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    data = None
    if request.method == 'POST':
        data_ = get_users()
        headers = ["Период", "Статус", "ИИН", "Фамилия", "Имя", "Отчество"]
        bodies = []
        for x in data_["global_objects"]:
            val = []
            for y in data_["global_objects"][x]:
                val.append(data_["global_objects"][x][y])
            bodies.append(val)
        data = [headers, bodies]
    context = {
        'data': data,
    }
    return render(request, 'app_km/create_user_from1c.html', context)


def change_user(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    if request.method == 'POST':
        form = ChangeUserForm(request.POST)
        if form.is_valid():
            user = User.objects.get(username=request.user.username)
            user.email = request.POST['email']
            user.password = request.POST['password1']
            user.first_name = request.POST['first_name']
            user.last_name = request.POST['last_name']
            user.save()
            user.set_password = request.POST['password1']

        result_change_user = True
    else:
        result_change_user = False
    change_user_form = ChangeUserForm()
    context = {
        'change_user_form': change_user_form,
        'result_change_user': result_change_user,
    }
    return render(request, 'app_km/change_user.html', context)


def create_users(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    if request.method == 'POST':
        excel_file = request.FILES.get('document_addition_file_1')
        if excel_file:
            def get_value(rows: str, cols: int):
                return str(sheet[rows + str(cols)].value)

            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            max_row = sheet.max_row
            for i in range(2, max_row + 1):
                username = get_value('A', i)
                password = get_value('B', i)
                email_ = get_value('C', i)
                first_name = get_value('D', i)
                last_name = get_value('E', i)
                group = get_value('F', i)
                if get_value('G', i) == 'ИСТИНА' or get_value('G', i) == 'True':
                    is_staff = True
                else:
                    is_staff = False
                if username and password:
                    try:
                        user = User.objects.get(username=username)
                        user.username = username
                        user.password = password
                        user.email = email_
                        user.first_name = first_name
                        user.last_name = last_name
                        user.is_staff = is_staff
                        if username == 'Bogdan' or username == 'bogdan':
                            user.is_superuser = True
                        else:
                            user.is_superuser = False
                        user.save()
                        user.set_password = str(password)
                    except Exception as ex:
                        print(ex)
                        User.objects.create_notification(
                            username=username,
                            password=password,
                            email=email_,
                            first_name=first_name,
                            last_name=last_name,
                            is_staff=is_staff,
                        )
                    try:
                        user_group = Group.objects.get(name=group)
                    except Exception as ex:
                        print(ex)
                        user_group = Group.objects.create_notification(name=group)
                    signup_user = User.objects.get(username=username)
                    user_group.user_set.add(signup_user)
        result_create_users = True
    else:
        result_create_users = False
    create_users_form = CreateUsersForm()
    context = {
        'create_users_form': create_users_form,
        'result_create_users': result_create_users,
    }
    return render(request, 'app_km/create_users.html', context)


def export_users(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    if request.method == 'POST':
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Страница 1'
        user_objects = User.objects.all().order_by('id')
        sheet['A1'] = 'Имя пользователя'
        sheet['B1'] = 'Зашифрованный Пароль'
        sheet['C1'] = 'Почта'
        sheet['D1'] = 'Имя'
        sheet['E1'] = 'Фамилия'
        sheet['F1'] = 'Группа'
        sheet['G1'] = 'Доступ к админ панели'
        value = 1
        for user_object in user_objects:
            value += 1
            sheet[f'A{value}'] = user_object.username
            sheet[f'B{value}'] = user_object.password
            sheet[f'C{value}'] = user_object.email
            sheet[f'D{value}'] = user_object.first_name
            sheet[f'E{value}'] = user_object.last_name
            groups = Group.objects.filter(user=user_object)
            group_list = ''
            for group in groups:
                group_list += group.name
            sheet[f'F{value}'] = group_list
            if user_object.is_staff:
                sheet[f'G{value}'] = 'ИСТИНА'
            else:
                sheet[f'G{value}'] = 'ЛОЖЬ'
        wb.save('static/media/data/accounts/export_users.xlsx')
        result_export_users = True
    else:
        result_export_users = False
    context = {
        'result_export_users': result_export_users,
    }
    return render(request, 'app_km/export_users.html', context)


def generate_password(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    if request.method == 'POST':
        not_encrypted_password = str(request.POST['password'])
        try:
            user = User.objects.get(username='None')
        except Exception as ex:
            print(ex)
            user = User.objects.create_notification(
                username='None'
            )
        user.set_password(not_encrypted_password)
        user.save()
        user = User.objects.get(username='None')
        encrypted_password = user.password
    else:
        encrypted_password = False
    context = {
        'result_generate_password': encrypted_password,
    }
    return render(request, 'app_km/generate_password.html', context)


def generate_passwords(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    if request.method == 'POST':
        passwords_quantity = int(request.POST['passwords_quantity'])
        passwords_chars = str(request.POST['passwords_chars'])
        passwords_length = int(request.POST['passwords_length'])
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Страница 1'
        sheet[f'A1'] = 'Зашифрованный Пароль'
        sheet[f'B1'] = 'Пароль'
        for n in range(2, passwords_quantity + 2):
            password = create_encrypted_password(_random_chars=passwords_chars, _length=passwords_length)
            try:
                user = User.objects.get(username='None')
            except Exception as ex:
                print(ex)
                user = User.objects.create_notification(
                    username='None',
                    email=password,
                )
            user.email = password
            user.set_password(password)
            user.save()
            user = User.objects.get(username='None')
            sheet[f'A{n}'] = user.password
            sheet[f'B{n}'] = user.email
        wb.save('static/media/data/accounts/generate_passwords.xlsx')
        result_generate_passwords = True
    else:
        result_generate_passwords = False
    generate_passwords_form = GeneratePasswordsForm()
    context = {
        'generate_passwords_form': generate_passwords_form,
        'result_generate_passwords': result_generate_passwords,
    }
    return render(request, 'app_km/generate_passwords.html', context)


def view_profile(request, username=None):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    try:
        account = AccountDataModel.objects.get(user_iin=username)
    except Exception as ex:
        print(ex)
        account = False
    context = {
        'account': account,
        'username': username
    }
    return render(request, 'app_km/view_profile.html', context)


# Application
def list_module(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    module = ApplicationModuleModel.objects.order_by('module_position')
    context = {
        'module': module,
    }
    return render(request, 'app_km/list_module.html', context)


def list_component(request, module_slug=None):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    if module_slug is not None:
        module = ApplicationModuleModel.objects.get(module_slug=module_slug)
        component = ApplicationComponentModel.objects.filter(component_Foreign=module).order_by('component_position')
    else:
        return redirect('app_km:list_module')
    context = {
        'component': component,
    }
    return render(request, 'app_km/list_component.html', context)


# Rational
def rational_list(request, category_slug=None):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    try:
        if category_slug is not None:
            category_page = get_object_or_404(CategoryRationalModel, category_slug=category_slug)
            rational = RationalModel.objects.filter(rational_category=category_page).order_by(
                '-rational_date_registered')
        else:
            rational = RationalModel.objects.order_by('-rational_date_registered')
        category = CategoryRationalModel.objects.order_by('-id')
        page = PaginationClass.paginate(request=request, objects=rational, num_page=3)
        context = {
            'page': page,
            'category': category,
        }
        return render(request, 'app_km/list.html', context)
    except Exception as ex:
        LoggingClass.logging(message=f'rational_list: {ex}')
        HttpRaiseExceptionClass.http404_raise('Страница не найдена ;(')


def rational_search(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    try:
        if request.method == 'POST':
            search = request.POST['search_text']
            rational = RationalModel.objects.filter(rational_name__icontains=search).order_by(
                '-rational_date_registered')
            context = {
                'page': rational
            }
            return render(request, 'app_km/list_search.html', context)
        else:
            return redirect('app_km:rational')
    except Exception as ex:
        LoggingClass.logging(message=f'rational_search: {ex}')
        HttpRaiseExceptionClass.http404_raise('Страница не найдена ;(')


def rational_detail(request, rational_id=1):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    try:
        rational = RationalModel.objects.get(id=rational_id)
        user = User.objects.get(id=request.user.id)
        comments = CommentRationalModel.objects.filter(comment_article=rational).order_by('-id')
        # comments = rational.comment_rational_model_set.order_by('-id')  # равнозначно предыдущему
        blog_is_liked = False
        blog_is_disliked = False
        total_like = LikeRationalModel.objects.filter(like_article=rational, like_status=True).count()
        total_dislike = LikeRationalModel.objects.filter(like_article=rational, like_status=False).count()
        try:
            LikeRationalModel.objects.get(like_article=rational, like_author=user, like_status=True)
            blog_is_liked = True
        except Exception as ex:
            print(ex)
            try:
                LikeRationalModel.objects.get(like_article=rational, like_author=user, like_status=False)
                blog_is_disliked = True
            except Exception as ex:
                print(ex)
        context = {
            'rational': rational,
            'comments': comments,
            'likes': {
                'like': blog_is_liked,
                'dislike': blog_is_disliked,
                'total_like': total_like,
                'total_dislike': total_dislike,
                'total_rating': total_like - total_dislike
            }
        }
        return render(request, 'app_km/detail.html', context)
    except Exception as ex:
        LoggingClass.logging(message=f'rational_detail: {ex}')
        HttpRaiseExceptionClass.http404_raise('Страница не найдена ;(')


def rational_create(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    try:
        if request.method == 'POST':
            form = RationalForm(request.POST, request.FILES)
            if form.is_valid():
                RationalModel.objects.create(
                    rational_structure_from=request.POST['rational_structure_from'],
                    rational_uid_registrated=request.POST['rational_uid_registered'],
                    rational_date_registrated=request.POST.get('rational_date_registered'),
                    rational_name=request.POST['rational_name'],
                    rational_place_innovation=request.POST['rational_place_innovation'],
                    rational_description=request.POST['rational_description'],
                    rational_addition_file_1=request.FILES.get('rational_addition_file_1'),
                    rational_addition_file_2=request.FILES.get('rational_addition_file_2'),
                    rational_addition_file_3=request.FILES.get('rational_addition_file_3'),
                    rational_offering_members=request.POST['rational_offering_members'],
                    rational_conclusion=request.POST['rational_conclusion'],
                    rational_change_documentations=request.POST['rational_change_documentations'],
                    rational_resolution=request.POST['rational_resolution'],
                    rational_responsible_members=request.POST['rational_responsible_members'],
                    rational_date_certification=request.POST.get('rational_date_certification'),
                    rational_category=CategoryRationalModel.objects.get(id=request.POST.get('rational_category')),
                    rational_author_name=User.objects.get(id=request.user.id),
                    # rational_date_create            = request.POST.get('rational_date_create'),
                    rational_addition_image=request.FILES.get('rational_addition_image'),
                    # rational_status                 = request.POST['rational_status'],
                )
            return redirect('app_km:rational')
        form = RationalForm(request.POST, request.FILES)
        category = CategoryRationalModel.objects.order_by('-id')
        user = User.objects.get(id=request.user.id).username
        context = {
            'form': form,
            'category': category,
            'user': user,
        }
        return render(request, 'app_km/create.html', context)
    except Exception as ex:
        LoggingClass.logging(message=f'rational_create: {ex}')
        HttpRaiseExceptionClass.http404_raise('Страница не найдена ; (')


def rational_change(request, rational_id=None):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    try:
        if request.method == 'POST':
            form = RationalForm(request.POST, request.FILES)
            if form.is_valid():
                _object = RationalModel.objects.get(id=rational_id)
                _object.rational_structure_from = request.POST['rational_structure_from']
                _object.rational_uid_registered = request.POST['rational_uid_registered']
                _object.rational_date_registered = request.POST.get('rational_date_registered')
                _object.rational_name = request.POST['rational_name']
                _object.rational_place_innovation = request.POST['rational_place_innovation']
                _object.rational_description = request.POST['rational_description']
                _object.rational_addition_file_1 = request.FILES.get('rational_addition_file_1')
                _object.rational_addition_file_2 = request.FILES.get('rational_addition_file_2')
                _object.rational_addition_file_3 = request.FILES.get('rational_addition_file_3')
                _object.rational_offering_members = request.POST['rational_offering_members']
                _object.rational_conclusion = request.POST['rational_conclusion']
                _object.rational_change_documentations = request.POST['rational_change_documentations']
                _object.rational_resolution = request.POST['rational_resolution']
                _object.rational_responsible_members = request.POST['rational_responsible_members']
                _object.rational_date_certification = request.POST.get('rational_date_certification')
                _object.rational_category = CategoryRationalModel.objects.get(id=request.POST.get('rational_category'))
                _object.rational_author_name = User.objects.get(id=request.user.id)
                # rational_date_create            = request.POST.get('rational_date_create'),
                _object.rational_addition_image = request.FILES.get('rational_addition_image')
                # rational_status                 = request.POST['rational_status'],
                _object.save()
            return redirect('app_km:rational')
        form = RationalForm(request.POST, request.FILES)
        category = CategoryRationalModel.objects.order_by('-id')
        context = {
            'form': form,
            'category': category,
            'rational_id': rational_id,
        }
        return render(request, 'app_km/change.html', context)
    except Exception as ex:
        LoggingClass.logging(message=f'rational_change: {ex}')
        HttpRaiseExceptionClass.http404_raise('Страница не найдена ;(')


def rational_leave_comment(request, rational_id):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    try:
        rational = RationalModel.objects.get(id=rational_id)
        CommentRationalModel.objects.create(comment_article=rational,
                                            comment_author=User.objects.get(id=request.user.id),
                                            comment_text=request.POST['comment_text'])
        # rational.comment_rational_model_set.create(comment_author=User.objects.get(id=request.user.id),
        #                                          comment_text=request.POST['comment_text'])  # равнозначно предыдущему
        return HttpResponseRedirect(reverse('app_km:rational_detail', args=(rational.id,)))
    except Exception as ex:
        LoggingClass.logging(message=f'rational_leave_comment: {ex}')
        HttpRaiseExceptionClass.http404_raise('Страница не найдена ;(')


def rational_change_rating(request, rational_id):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    try:
        blog_ = RationalModel.objects.get(id=rational_id)
        user = User.objects.get(id=request.user.id)
        if request.POST['status'] == '+':
            try:
                LikeRationalModel.objects.get(like_article=blog_, like_author=user, like_status=True).delete()
            except Exception as ex:
                print(ex)
                blog_.likerationalmodel_set.create(like_article=blog_, like_author=user, like_status=True)
            try:
                LikeRationalModel.objects.get(like_article=blog_, like_author=user, like_status=False).delete()
            except Exception as ex:
                print(ex)
        else:
            try:
                LikeRationalModel.objects.get(like_article=blog_, like_author=user, like_status=False).delete()
            except Exception as ex:
                print(ex)
                blog_.likerationalmodel_set.create(like_article=blog_, like_author=user, like_status=False)
            try:
                LikeRationalModel.objects.get(like_article=blog_, like_author=user, like_status=True).delete()
            except Exception as ex:
                print(ex)
        return HttpResponseRedirect(reverse('app_km:rational_detail', args=(blog_.id,)))
    except Exception as ex:
        LoggingClass.logging(message=f'rational_change_rating: {ex}')
        HttpRaiseExceptionClass.http404_raise('Страница не найдена ;(')


def rational_ratings(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    try:
        rational = RationalModel.objects.order_by('-id')
        authors = []
        for query in rational:
            authors.append(query.rational_author_name)
        user_count = {}
        for author in authors:
            user_count[author] = authors.count(author)
        user_counts = []
        for blog_s in user_count:
            rationals = RationalModel.objects.filter(rational_author_name=blog_s)
            total_rating = 0
            for rating in rationals:
                total_like = LikeRationalModel.objects.filter(like_article=rating, like_status=True).count()
                total_dislike = LikeRationalModel.objects.filter(like_article=rating, like_status=False).count()
                total_rating += total_like - total_dislike
            user_counts.append({'user': blog_s, 'count': user_count[blog_s], 'rating': total_rating})
        sorted_by_rating = True
        if request.method == 'POST':
            if request.POST['sorted'] == 'rating':
                sorted_by_rating = True
            if request.POST['sorted'] == 'count':
                sorted_by_rating = False
        if sorted_by_rating:
            page = sorted(user_counts, key=lambda k: k['rating'], reverse=True)
        else:
            page = sorted(user_counts, key=lambda k: k['count'], reverse=True)
        context = {
            'page': page,
            'sorted': sorted_by_rating
        }
        return render(request, 'app_km/ratings.html', context)
    except Exception as ex:
        LoggingClass.logging(message=f'rational_change_rating: {ex}')
        HttpRaiseExceptionClass.http404_raise('Страница не найдена ;(')


# Salary
def salary(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    try:
        data = None
        user = User.objects.get(id=request.user.id)
        if request.method == 'POST':
            # Тут мы получаем json ответ от интерфейса 1С
            data = get_salary_data(month=request.POST['transact_id'])
            # Тут мы получаем json ответ от интерфейса 1С
        context = {
            'user': user,
            'data': data,
        }
        return render(request, 'app_km/main.html', context)
    except Exception as ex:
        LoggingClass.logging(message=f'salary: {ex}')
        HttpRaiseExceptionClass.http404_raise('Страница не найдена ;(')


def view_pdf(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    data = None
    if request.method == 'POST':
        data = get_salary_data(month=request.POST['transact_id'])
    context = {
        'data': data,
    }
    return render(request, 'app_km/pdf.html', context)


def render_pdf_view(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    template_path = 'app_km/pdf.html'
    data = get_salary_data()
    # data = None
    context = {
        'data': data,
        'STATIC_ROOT': settings.STATIC_ROOT,
    }
    # Create a Django response object, and specify content_type as pdf
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="report.pdf"'
    # find the template and render it.
    template = get_template(template_path)
    html = template.render(context)
    # create a pdf
    pisa_status = pisa.CreatePDF(
        html, dest=response, encoding='utf-8', link_callback=link_callback)
    # template = render_to_string(template_path, context)
    # pdf = pisa.pisaDocument(io.BytesIO(template.encode('UTF-8')), response,
    #                         encoding='utf-8',
    #                         link_callback=link_callback)
    # pdf = pisa.pisaDocument(io.StringIO(html), response, encoding='UTF-8')
    # if error then show some funny view
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response


# Human Resources
def career(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    try:
        data = None
        if request.method == 'POST':
            data = get_career()
        context = {
            'data': data,
        }
        return render(request, 'app_km/career.html', context)
    except Exception as ex:
        LoggingClass.logging(message=f'career: {ex}')
        HttpRaiseExceptionClass.http404_raise('Страница не найдена ;(')


# Extra
def geo(request):
    # if AuthorizationClass.user_authenticated(request=request):
    #     return redirect(AuthorizationClass.user_authenticated(request=request))
    try:
        data = None
        form = GeoForm()
        if request.method == 'POST':
            print('begin')
            data = generate_xlsx(request)
            print('generate_xlsx successfully')
            # print(data)
            generate_kml()
            print('generate_kml successfully')
            # point = find_point(52.2, 61.3)

            # Точки
            point1 = [61.23500, 52.17263, '1']
            point2 = [61.23500, 52.17263, '2']
            point3 = [61.23500, 52.17263, '3']
            point4 = [61.23500, 52.17263, '4']
            point5 = [61.23500, 52.17263, '5']
            point5 = [61.23500, 52.17263, '6']
            point5 = [61.23500, 52.17263, '7']
            point5 = [61.23500, 52.17263, '8']
            # .....
            # Пути
            dist1 = [[61.23500, 52.17263, '1'], [61.23500, 52.17263, '2']]
            # 61.23500, 52.17263, 61.23500, 52.17263 'A|B'
            dist2 = [[61.23500, 52.17263, '1'], [61.23500, 52.17263, '7']]
            # 61.23500, 52.17263, 61.23500, 52.17263 'A|G'
            dist3 = [[61.23500, 52.17263, '2'], [61.23500, 52.17263, '3']]
            # 61.23500, 52.17263, 61.23500, 52.17263 'B|C'
            dist4 = [[61.23500, 52.17263, '2'], [61.23500, 52.17263, '6']]
            # 61.23500, 52.17263, 61.23500, 52.17263 'B|F'
            dist5 = [[61.23500, 52.17263, '2'], [61.23500, 52.17263, '8']]
            # 61.23500, 52.17263, 61.23500, 52.17263 'B|H'
            # .....
            # Связи
            link1 = ['1', ['2', '7']]
            link1 = ['2', ['3', '6', '8']]
            # .....

            dist1 = [round(get_distance(x1=61.23500, y1=52.17263, x2=61.23654, y2=52.16710), 5), 'BA']
            dist2 = [round(get_distance(x1=61.23550, y1=52.17263, x2=61.23654, y2=52.16710), 5), 'BC']
            dist3 = [round(get_distance(x1=61.23450, y1=52.17263, x2=61.23654, y2=52.16710), 5), 'BD']
            print([dist1, dist2, dist3])
            dist = find_distance([dist1, dist2, dist3])
            print(dist)

            print('end')
        context = {
            'data': data,
            'form': form,
        }
        return render(request, 'app_km/geo.html', context)
    except Exception as ex:
        LoggingClass.logging(message=f'geo: {ex}')
        HttpRaiseExceptionClass.http404_raise('Страница не найдена ;(')


def email(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    mails = EmailModel.objects.order_by('-id')
    # Начало пагинатора: передать массив объектов и количество объектов на одной странице, объекты будут списком
    paginator = Paginator(mails, 10)
    page = request.GET.get('page')
    try:
        page = paginator.page(page)
    except PageNotAnInteger:
        page = paginator.page(1)
    except EmptyPage:
        page = paginator.page(paginator.num_pages)
    # конец пагинатора, объекты под ключом "'page': page"

    context = {
        'page': page
    }
    return render(request, 'app_km/email.html', context)


def send_email(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    if request.method == 'POST':
        subject = request.POST.get('subject', '')
        message_s = request.POST.get('message', '')
        to_email = request.POST.get('to_email', '')
        if subject and message and to_email:
            try:
                send_mail(subject, message_s, 'eevee.cycle@yandex.ru', [to_email, ''], fail_silently=False)
                EmailModel.objects.create(
                    Email_subject=subject,
                    Email_message=message_s,
                    Email_email=to_email
                )
            except BadHeaderError:
                return redirect('home')
    return redirect('app_km:email')


def notification(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    form = NotificationForm(request.POST, request.FILES)
    pages = NotificationModel.objects.order_by('-id')
    # Начало пагинатора: передать массив объектов и количество объектов на одной странице, объекты будут списком
    paginator = Paginator(pages, 10)
    page = request.GET.get('page')
    try:
        page = paginator.page(page)
    except PageNotAnInteger:
        page = paginator.page(1)
    except EmptyPage:
        page = paginator.page(paginator.num_pages)
    # конец пагинатора, объекты под ключом "'page': page"
    context = {
        'page': page,
        'form': form,
    }
    return render(request, 'app_km/notification.html', context)


def create_notification(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    if request.method == 'POST':
        NotificationModel.objects.create(
            notification_name=request.POST['notification_name'],
            notification_slug=request.POST['notification_slug'],
            notification_description=request.POST['notification_description'],
            notification_author=User.objects.get(id=request.user.id),
        )
    return redirect('app_km:notification')


def accept(request, notify_id):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    try:
        rational = NotificationModel.objects.get(id=notify_id)
        rational.notification_status = True
        rational.save()
    except Exception as ex:
        print(ex)
    return redirect('app_km:notification')


def decline(request, notify_id):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    try:
        rational = NotificationModel.objects.get(id=notify_id)
        rational.notification_status = False
        rational.save()
    except Exception as ex:
        print(ex)
    return redirect('app_km:notification')


def documentation(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    if request.method == 'POST':
        DocumentModel.objects.create(
            document_name=request.POST['document_name'],
            document_slug=request.POST['document_slug'],
            document_description=request.POST.get('document_description'),
            document_addition_file_1=request.FILES.get('document_addition_file_1'),
            document_addition_file_2=request.FILES.get('document_addition_file_2'),
        )
        return redirect('app_km:documentation')
    docs = DocumentModel.objects.order_by('-id')
    # Начало пагинатора: передать массив объектов и количество объектов на одной странице, объекты будут списком
    paginator = Paginator(docs, 3)
    page = request.GET.get('page')
    try:
        page = paginator.page(page)
    except PageNotAnInteger:
        page = paginator.page(1)
    except EmptyPage:
        page = paginator.page(paginator.num_pages)
    # конец пагинатора, объекты под ключом "'page': page"
    form = DocumentForm(request.POST, request.FILES)
    context = {
        'page': page,
        'form': form,
    }
    return render(request, 'app_km/documentation.html', context)


def contact(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    if request.method == 'POST':
        ContactModel.objects.create(
            contact_name=request.POST['contact_name'],
            contact_slug=request.POST['contact_slug'],
            contact_description=request.POST.get('contact_description'),
            contact_image=request.FILES.get('contact_image'),
        )
        return redirect('app_km:contact')
    contacts = ContactModel.objects.order_by('-id')
    # Начало пагинатора: передать массив объектов и количество объектов на одной странице, объекты будут списком
    paginator = Paginator(contacts, 3)
    page = request.GET.get('page')
    try:
        page = paginator.page(page)
    except PageNotAnInteger:
        page = paginator.page(1)
    except EmptyPage:
        page = paginator.page(paginator.num_pages)
    # конец пагинатора, объекты под ключом "'page': page"
    form = ContactForm(request.POST, request.FILES)
    context = {
        'page': page,
        'form': form,
    }
    return render(request, 'app_km/contact.html', context)


def list_sms(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    if request.method == 'POST':
        SmsModel.objects.create(
            sms_author=User.objects.get(id=request.user.id),
            sms_description=request.POST['sms_description'],
            # sms_date=request.POST.get('contact_description'),
        )
        return redirect('chat')
    contacts = SmsModel.objects.order_by('-sms_date')
    # Начало пагинатора: передать массив объектов и количество объектов на одной странице, объекты будут списком
    paginator = Paginator(contacts, 3)
    page = request.GET.get('page')
    try:
        page = paginator.page(page)
    except PageNotAnInteger:
        page = paginator.page(1)
    except EmptyPage:
        page = paginator.page(paginator.num_pages)
    # конец пагинатора, объекты под ключом "'page': page"
    form = SmsForm(request.POST, request.FILES)
    context = {
        'page': page,
        'form': form,
    }
    return render(request, 'app_km/message.html', context)


def message(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    if request.method == 'POST':
        MessageModel.objects.create(
            message_name=request.POST['message_name'],
            message_slug=request.POST['message_slug'],
            message_description=request.POST.get('message_description'),
        )
        return redirect('app_km:message')
    form = MessageForm(request.POST, request.FILES)
    message_s = MessageModel.objects.order_by('-id')
    # Начало пагинатора: передать массив объектов и количество объектов на одной странице, объекты будут списком
    paginator = Paginator(message_s, 3)
    page = request.GET.get('page')
    try:
        page = paginator.page(page)
    except PageNotAnInteger:
        page = paginator.page(1)
    except EmptyPage:
        page = paginator.page(paginator.num_pages)
    # конец пагинатора, объекты под ключом "'page': page"
    context = {
        'page': page,
        'form': form,
    }
    return render(request, 'app_km/message.html', context)


def weather(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    appid = '82b797b6ebc625032318e16f1b42c016'
    url = 'https://api.openweathermap.org/data/2.5/weather?q={}&units=metric&appid=' + appid
    if request.method == 'POST':
        try:
            form = CityForm(request.POST)
            form.save()
        except Exception as ex:
            print(ex)
    form = CityForm()
    cities = CityModel.objects.all()
    all_cities = []
    for city in cities:
        try:
            res = requests.get(url.format(city.name)).json()
            print(res)
            city_info = {
                'city': city.name,
                'temp': res["main"]["temp"],
                'icon': res["weather"][0]["icon"]
            }
            all_cities.append(city_info)
        except Exception as ex:
            print(ex)
    # Начало пагинатора: передать модель и количество объектов на одной странице, объекты будут списком
    paginator = Paginator(all_cities, 3)
    page = request.GET.get('page')
    try:
        page = paginator.page(page)
    except PageNotAnInteger:
        page = paginator.page(1)
    except EmptyPage:
        page = paginator.page(paginator.num_pages)
    # конец пагинатора, объекты под ключом "'page': page"
    context = {
        'form': form,
        'page': page,
    }
    return render(request, 'app_km/weather_list.html', context)


def news_list(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    latest_article_list = ArticleModel.objects.order_by('-article_pub_date')[:10]
    return render(request, 'app_km/news_list.html', {'latest_article_list': latest_article_list})


def news_create(request):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    if request.method == 'POST':
        ArticleModel.objects.create(
            article_title=request.POST['article_title'],
            article_text=request.POST['article_text']
        )
        latest_article_list = ArticleModel.objects.order_by('-article_pub_date')[:10]
        return render(request, 'app_km/news_list.html', {'latest_article_list': latest_article_list})
    post_form = ArticleForm(request.POST)
    return render(request, 'app_km/news_create.html', {'post_form': post_form})


def news_detail(request, article_id):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    try:
        a = ArticleModel.objects.get(id=article_id)
    except Exception as ex:
        print(ex)
        raise Http404('Статья не найдена')
    latest_comments_list = a.comment_set.order_by('-id')[:10]
    return render(request, 'app_km/news_detail.html', {'article': a, 'latest_comments_list': latest_comments_list})


def leave_comment(request, article_id):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    try:
        a = ArticleModel.objects.get(id=article_id)
    except Exception as ex:
        print(ex)
        raise Http404('Статья не найдена')
    a.comment_set.create(author_name=request.POST['name'], comment_text=request.POST['text'])
    return HttpResponseRedirect(reverse('news_detail', args=(a.id,)))


def increase_rating(request, article_id):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    a = ArticleModel.objects.get(id=article_id)
    a.increase()
    a.save()
    return HttpResponseRedirect(reverse('news_detail', args=(a.id,)))


def decrease_rating(request, article_id):
    if AuthorizationClass.user_authenticated(request=request):
        return redirect(AuthorizationClass.user_authenticated(request=request))
    a = ArticleModel.objects.get(id=article_id)
    a.decrease()
    a.save()
    return HttpResponseRedirect(reverse('news_detail', args=(a.id,)))


# React
def react(request):
    return render(request, 'app_km/react.html')


# Bootstrap
def example(request):
    return render(request, 'app_km/bootstrap/example.html')


def album(request):
    return render(request, 'app_km/bootstrap/album.html')


def blog(request):
    return render(request, 'app_km/bootstrap/blog.html')


def carousel(request):
    return render(request, 'app_km/bootstrap/carousel.html')


def checkout(request):
    return render(request, 'app_km/bootstrap/checkout.html')


def cover(request):
    return render(request, 'app_km/bootstrap/cover.html')


def dashboard(request):
    return render(request, 'app_km/bootstrap/dashboard.html')


def pricing(request):
    return render(request, 'app_km/bootstrap/pricing.html')


def product(request):
    return render(request, 'app_km/bootstrap/product.html')


def sign_in(request):
    return render(request, 'app_km/bootstrap/sign-in.html')


def sticky_footer(request):
    return render(request, 'app_km/bootstrap/sticky-footer.html')


def sticky_footer_navbar(request):
    return render(request, 'app_km/bootstrap/sticky-footer-navbar.html')


def starter_template(request):
    return render(request, 'app_km/bootstrap/starter-template.html')


def grid(request):
    return render(request, 'app_km/bootstrap/grid.html')


def cheatsheet(request):
    return render(request, 'app_km/bootstrap/cheatsheet.html')


def nav_bars(request):
    return render(request, 'app_km/bootstrap/nav_bars.html')


def off_canvas(request):
    return render(request, 'app_km/bootstrap/off_canvas.html')


def masonry(request):
    return render(request, 'app_km/bootstrap/masonry.html')


def navbar_static(request):
    return render(request, 'app_km/bootstrap/navbar-static.html')


def navbar_fixed(request):
    return render(request, 'app_km/bootstrap/navbar-fixed.html')


def navbar_bottom(request):
    return render(request, 'app_km/bootstrap/navbar-bottom.html')
