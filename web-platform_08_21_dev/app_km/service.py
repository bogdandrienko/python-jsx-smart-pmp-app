import datetime

import openpyxl
from django.contrib.auth.models import User
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http.response import Http404
from django.contrib.staticfiles import finders
from django.conf import settings
import json
import httplib2
import os
import time
import random
import requests
import bs4
import psycopg2 as pg
from openpyxl.utils import get_column_letter


class AuthorizationClass:
    @staticmethod
    def user_authenticated(request):
        # Проверка регистрации: если пользователь не вошёл в аккаунт его переадресует в форму входа
        if request.user.is_authenticated is not True:
            return 'app_km:login'
        return None


class PaginationClass:
    @staticmethod
    def paginate(request, objects, num_page):
        # Пагинатор: постраничный вывод объектов
        paginator = Paginator(objects, num_page)
        pages = request.GET.get('page')
        try:
            page = paginator.page(pages)
        except PageNotAnInteger:
            page = paginator.page(1)
        except EmptyPage:
            page = paginator.page(paginator.num_pages)
        return page


class HttpRaiseExceptionClass:
    @staticmethod
    def http404_raise(exception_text):
        raise Http404(exception_text)


class LoggingClass:
    @staticmethod
    def logging(message, file_name='log.txt', type_write='a'):
        print(f'{TimeUtils.get_current_time()} : {message}\n')
        with open(file_name, type_write) as log:
            log.write(f'{TimeUtils.get_current_time()} : {message}\n')


class TimeUtils:
    @staticmethod
    def get_current_time():
        return f"{time.strftime('%X')}"


def create_account(_login, _not_encrypted_password, _email, _name, _surname, _is_staff):
    try:
        user = User.objects.create_notification(
            username=_login,
            password=_not_encrypted_password,
            email=_email,
            first_name=_name,
            last_name=_surname,
            is_staff=_is_staff,
        )
        user.save()
        user.set_password = _not_encrypted_password
    except Exception as ex:
        print(ex)


def create_encrypted_password(_random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890', _length=8):
    password = ''
    for i in range(1, _length + 1):
        password += random.choice(_random_chars)
    return password


# Salary
def get_salary_data(month=4):
    url = f'http://192.168.1.158/Tanya_perenos/hs/zp/rl/970801351179/20210{month}'
    relative_path = os.path.dirname(os.path.abspath('__file__')) + '\\'
    h = httplib2.Http(relative_path + "\\static\\media\\data\\temp")
    login = 'zpadmin'
    password = '159159qo'
    h.add_credentials(login, password)
    response, content = h.request(url)
    if content:
        try:
            with open("static/media/data/zarplata.json", "w", encoding="utf-8") as file:
                json.dump(content, file)
            data = json.loads(content)
        except Exception as ex:
            print(ex)
            with open("static/media/data/zarplata_temp.json", "r", encoding="utf-8") as file:
                data = json.load(file)
    else:
        with open("static/media/data/zarplata_temp.json", "r", encoding="utf-8") as file:
            data = json.load(file)
    try:
        data["global_objects"]["3.Доходы в натуральной форме"]
    except Exception as ex:
        print(ex)
        data["global_objects"]["3.Доходы в натуральной форме"] = {
            "Fields": {
                "1": "Вид",
                "2": "Период",
                "3": "Дни",
                "4": "Часы",
                "5": "Сумма",
                "6": "ВсегоДни",
                "7": "ВсегоЧасы"
            },
        }
    data = {
        "Table_1": create_arr_table(
            title="1.Начислено", footer="Всего начислено", json_obj=data["global_objects"]["1.Начислено"],
            exclude=[5, 6]
        ),
        "Table_2": create_arr_table(
            title="2.Удержано", footer="Всего удержано", json_obj=data["global_objects"]["2.Удержано"], exclude=[]
        ),
        "Table_3": create_arr_table(
            title="3.Доходы в натуральной форме", footer="Всего натуральных доходов",
            json_obj=data["global_objects"]["3.Доходы в натуральной форме"], exclude=[]
        ),
        "Table_4": create_arr_table(
            title="4.Выплачено", footer="Всего выплат", json_obj=data["global_objects"]["4.Выплачено"], exclude=[]
        ),
        "Table_5": create_arr_table(
            title="5.Налоговые вычеты", footer="Всего вычеты", json_obj=data["global_objects"]["5.Налоговые вычеты"],
            exclude=[]
        ),
        "Down": {
            "first": ["Долг за организацией на начало месяца", data["Долг за организацией на начало месяца"]],
            "last": ["Долг за организацией на конец месяца", data["Долг за организацией на конец месяца"]],
        },
    }
    # global_objects = []
    # for x in data["global_objects"]:
    #     global_objects.append(x)
    # global_objects = [x for x in data["global_objects"]]

    # return_data = []
    # for x in global_objects:
    #     return_data.append(create_arr_from_json(data["global_objects"], x))
    # return_data = [create_arr_from_json(data["global_objects"], x) for x in global_objects]

    # return_data = [create_arr_from_json(data["global_objects"], y) for y in [x for x in data["global_objects"]]]
    return data


def get_users(day=1):
    url = f'http://192.168.1.158/Tanya_perenos/hs/iden/change/2021030{day}'
    relative_path = os.path.dirname(os.path.abspath('__file__')) + '\\'
    h = httplib2.Http(relative_path + "\\static\\media\\data\\temp")
    login = 'zpadmin'
    password = '159159qqq'
    h.add_credentials(login, password)
    response, content = h.request(url)
    if content:
        try:
            with open("static/media/data/accounts.json", "w", encoding="utf-8") as file:
                json.dump(content, file)
            return json.loads(content)
        except Exception as ex:
            print(ex)
    with open("static/media/data/accounts_temp.json", "r", encoding="utf-8") as file:
        data = json.load(file)
    return data


def get_career():
    # headers = {
    #     'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:45.0) Gecko/20100101 Firefox/45.0'
    # }
    # vacancies_urls = []
    # url = 'https://www.km-open.online/property'
    # r = requests.get(url, headers=headers)
    # soup = bs4.BeautifulSoup(r.content.decode("utf-8"))
    # list_objs = soup.find_all('div', {"class": "collection-item w-dyn-item"})
    # for list_obj in list_objs:
    #     vacancies_urls.append(url.split('/property')[0] + str(list_obj).split('href="')[1].split('"')[0])
    # vacancies_data = []
    # for url_s in vacancies_urls:
    #     r = requests.get(url_s, headers=headers)
    #     soup = bs4.BeautifulSoup(r.content.decode("utf-8"))
    #     list_objs = soup.find_all('div', {"class": "title-block"})
    #     vacancies_data = str(list_objs[0]).split('"heading-11">')[1].split('</h5>')[0]
    #     vacancies_data.append([vacancies_data, url_s])
    # data = [["Вакансия"], vacancies_data]
    # headers = {
    #     'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:45.0) Gecko/20100101 Firefox/45.0'
    # }

    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:45.0) Gecko/20100101 Firefox/45.0'
    }
    vacancies_urls = []
    url = 'https://www.km-open.online/property'
    r = requests.get(url, headers=headers)
    soup = bs4.BeautifulSoup(r.content.decode("utf-8"))
    list_objs = soup.find_all('div', {"class": "collection-item w-dyn-item"})
    for list_obj in list_objs:
        vacancies_urls.append(url.split('/property')[0] + str(list_obj).split('href="')[1].split('"')[0])
    vacancies_title = []
    for list_obj in list_objs:
        vacancies_title.append(str(list_obj).split('class="heading-12">')[1].split('</h5>')[0])
    vacancies_data = []
    for title in vacancies_title:
        vacancies_data.append([title, vacancies_urls[vacancies_title.index(title)]])

    data = [["Вакансия"], vacancies_data]
    return data


def create_arr_from_json(json_obj, parent_key: str):
    headers = []
    for x in json_obj[parent_key]["Fields"]:
        headers.append(json_obj[parent_key]["Fields"][x])
    del json_obj[parent_key]["Fields"]
    bodies = []
    for x in json_obj[parent_key]:
        val = [x]
        for y in json_obj[parent_key][x]:
            val.append(json_obj[parent_key][x][y])
        bodies.append(val)
    return [parent_key, headers, bodies]


def create_arr_table(title: str, footer: str, json_obj, exclude: list):
    headers = []

    for x in json_obj["Fields"]:
        headers.append(json_obj["Fields"][x])
    del json_obj["Fields"]
    bodies = [["", title]]

    if exclude:
        hours = 0
        days = 0
        sum_ = 0
        for x in json_obj:
            val = [x]
            i = 0
            for y in json_obj[x]:
                i += 1
                if i == exclude[0]:
                    hours += json_obj[x][y]
                    continue
                if i == exclude[1]:
                    days += json_obj[x][y]
                    continue
                if i == len(json_obj[x]):
                    sum_ += json_obj[x][y]
                val.append(json_obj[x][y])
            bodies.append(val)
        footers = ["", footer, "", hours, days, round(sum_, 2)]
    else:
        sum_ = 0
        for x in json_obj:
            val = [x]
            i = 0
            for y in json_obj[x]:
                i += 1
                if i == len(json_obj[x]):
                    sum_ += json_obj[x][y]
                val.append(json_obj[x][y])
            bodies.append(val)
        footers = ["", footer, "", round(sum_, 2)]

    return [headers, bodies, footers]


def link_callback(uri):
    """
    Convert HTML URIs to absolute system paths so xhtml2pdf can access those
    resources
    """
    result = finders.find(uri)
    s_url = ''
    m_url = ''
    if result:
        if not isinstance(result, (list, tuple)):
            result = [result]
        result = list(os.path.realpath(path) for path in result)
        path = result[0]
    else:
        s_url = settings.STATIC_URL  # Typically /static/
        s_root = settings.STATIC_ROOT  # Typically /home/userX/project_static/
        m_url = settings.MEDIA_URL  # Typically /media/
        m_root = settings.MEDIA_ROOT  # Typically /home/userX/project_static/media/

        if uri.startswith(m_url):
            path = os.path.join(m_root, uri.replace(m_url, ""))
        elif uri.startswith(s_url):
            path = os.path.join(s_root, uri.replace(s_url, ""))
        else:
            return uri
    if not os.path.isfile(path):
        raise Exception(
            'media URI must start with %s or %s' % (s_url, m_url)
        )
    return path


def get_sheet_value(column, row, _sheet):
    return _sheet[str(column) + str(row)].value


def generate_xlsx(request):
    connection = pg.connect(
        host="192.168.1.6",
        database="navSections",
        port="5432",
        user="postgres",
        password="nF2ArtXK"
    )
    # postgresql_select_query = f"SELECT device, navtime, ROUND(CAST(latitude AS numeric), {request.POST['request_value']}), ROUND(CAST(longitude AS numeric), {request.POST['request_value']}) " \
    #                           "FROM public.navdata_202108 " \
    #                           f"WHERE device BETWEEN {request.POST['request_between_first']} AND {request.POST['request_between_last']} AND timezone('UTC', to_timestamp(navtime)) > (CURRENT_TIMESTAMP - INTERVAL '{request.POST['request_hours']} hours') AND flags != 64 " \
    #                           "ORDER BY device, navtime DESC;"
    postgresql_select_query = f"SELECT device, navtime, ROUND(CAST(latitude AS numeric), {request.POST['request_value']}), ROUND(CAST(longitude AS numeric), {request.POST['request_value']}) " \
                              "FROM public.navdata_202108 " \
                              f"WHERE device BETWEEN {request.POST['request_between_first']} AND {request.POST['request_between_last']} AND timezone('UTC', to_timestamp(navtime)) > (CURRENT_TIMESTAMP - INTERVAL '{request.POST['request_hours']} hours') AND flags != 64 " \
                              "ORDER BY device, navtime DESC;"
    cursor = connection.cursor()
    cursor.execute(postgresql_select_query)
    mobile_records = cursor.fetchall()
    cols = ["устройство", "дата и время", "широта", "долгота", "высота", "скорость", "ds", "направление",
            "флаги ошибок"]
    all_arr = []
    for rows in mobile_records:
        arr = []
        for value in rows:
            id_s = rows.index(value)
            if id_s == 1:
                arr.append(datetime.datetime.fromtimestamp(int(value - 21600)).strftime('%Y-%m-%d %H:%M:%S'))
            else:
                arr.append(value)
        all_arr.append(arr)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Страница 1'
    for n in cols:
        sheet[f'{get_column_letter(cols.index(n) + 1)}1'] = n
    for n in all_arr:
        for i in n:
            if i:
                sheet[f'{get_column_letter(n.index(i) + 1)}{all_arr.index(n) + 2}'] = i
            else:
                sheet[f'{get_column_letter(n.index(i) + 1)}{all_arr.index(n) + 2}'] = '0.0'
    wb.save('static/media/data/geo.xlsx')
    return [cols, all_arr]


def generate_kml():
    file_xlsx = 'static/media/data/geo.xlsx'
    workbook = openpyxl.load_workbook(file_xlsx)
    sheet = workbook.active
    array = []
    # colors = ['FFFFFFFF', 'FF0000FF', 'FFFF0000', 'FF00FF00', 'FF00FFFF', 'FF000000']
    colors = ['FFFFFFFF', 'FF0000FF', 'FFFF0000', 'FF00FF00', 'FF00FFFF', 'FF0F0F0F', 'FFF0F0F0']
    colors_alias = [': Чё', ': Кр', ': Си', ': Зе', ': Жё', ': Бе', ': Доп1', ': Доп2']
    final = 0
    for num in range(2, 10000000):
        if get_sheet_value("A", num, _sheet=sheet) is None or '':
            final = num
            break
    for num in range(2, final):
        array.append([get_sheet_value("D", num, _sheet=sheet), get_sheet_value("C", num, _sheet=sheet),
                      get_sheet_value("A", num, _sheet=sheet)])
    text_b = ''
    current_color = colors[0]
    previous_device = 0
    device_arr = []
    for current in array:
        index = array.index(current)
        if previous_device is not current[2]:
            device_arr.append(str(current[2]) + colors_alias[len(device_arr)+1])
            previous_device = current[2]
            current_color = colors[colors.index(current_color)+1]
        if index != 0:
            previous = [array[index - 1][0], array[index - 1][1]]
        else:
            previous = [current[0], current[1]]
        text_b += f"""<Placemark>
              <Style>
                <LineStyle>
                  <color>{current_color}</color>
                </LineStyle>
              </Style>
              <LineString>
                <coordinates>{previous[0]},{previous[1]},0 {current[0]},{current[1]},0 </coordinates>
              </LineString>
          </Placemark>
        """
    dev = ''
    for x in device_arr:
        dev += f"{x} | "
    text_a = f"""<?xml version="1.0" encoding="utf-8"?>
      <kml xmlns="http://earth.google.com/kml/2.2">
        <Document>
          <name>{dev}</name>
            """
    text_c = R"""</Document>
    </kml>"""
    text = text_a + text_b + text_c
    with open("static/media/data/geo.kml", "w", encoding="utf-8") as file:
        file.write(text)


def find_point(latitude: float, longitude: float):
    # with open("static/media/data/geo.kml", 'rt', encoding="utf-8") as file:
    #     data = file.read()
    # k = kml.KML()
    # k.from_string(data)
    # features = list(k.features())
    # k2 = list(features[0].features())
    # arr = []
    # for feat in k2:
    #     string = str(feat.geometry).split('(')[1].split('0.0')[0].split(' ')
    #     arr.append([float(string[0]), float(string[1])])
    # if val is None:
    #     val = [61.2200083333333, 52.147525]
    # val2 = 0
    # val3 = 0
    # for loop1 in arr:
    #     # Мы должны найти к какой из точек он ближе(разница двух элементов массива)
    #     if val[0] > loop1[0]:
    #         for loop2 in arr:
    #             if val[1] > loop2[1]:
    #                 val2 = loop1[0]
    #                 val3 = loop1[1]
    #                 break
    # print([val2, val3])
    # context = {
    #     'data': [['широта', 'долгота'], arr],
    # }
    return [latitude, longitude]
