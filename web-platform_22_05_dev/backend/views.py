# TODO download modules ################################################################################################

import base64
import datetime
import hashlib
import json
import os
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import random
from typing import Union
import httplib2
import openpyxl
from django.core.paginator import Paginator
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill
from openpyxl.utils import get_column_letter

# TODO django modules ##################################################################################################

from django.contrib.auth import authenticate
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User, Group, update_last_login
from django.core.mail import send_mail
from django.shortcuts import render
from django.utils import timezone

# TODO drf modules #####################################################################################################

from rest_framework import viewsets, permissions
from rest_framework.authentication import BasicAuthentication
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken

# TODO custom modules ##################################################################################################

from backend import models as backend_models, serializers as backend_serializers, utils as backend_utils

# TODO custom variables ################################################################################################

HTTP_METHOD_NAMES = ["GET", "POST", "PUT", "DELETE", "OPTIONS"]


# TODO viewsets ########################################################################################################

class ExamplesModelViewSet(viewsets.ModelViewSet):
    queryset = backend_models.ExamplesModel.objects.all()
    serializer_class = backend_serializers.ExamplesModelSerializer
    permission_classes = [permissions.AllowAny]


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = backend_serializers.UserSerializer
    permission_classes = [permissions.AllowAny]


class GroupViewSet(viewsets.ModelViewSet):
    queryset = Group.objects.all()
    serializer_class = backend_serializers.GroupSerializer
    permission_classes = [permissions.AllowAny]


# TODO routes ##########################################################################################################

@api_view(http_method_names=HTTP_METHOD_NAMES)
@authentication_classes([BasicAuthentication])
def api_auth_routes(request):
    """
    routes django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "GET":

            # TODO action ##############################################################################################

            if req_inst.action_type == "":
                try:
                    response = {"response": [
                        {
                            'Endpoints': '$BASEPATH$/router/',
                            'methods': 'GET, HEAD, OPTIONS',
                            'descriptions': 'default router for rest_framework ViewSet',
                            'helps': '''$BASEPATH$/router/''',
                            'code': '''from rest_framework import routers\n
                            router = routers.DefaultRouter()\n
                            router.register(prefix=r'users', viewset=views.UserViewSet, basename='users')'''
                        },
                        {
                            'name': '''JWT token''',
                            'endpoint': '''$BASEPATH$/tokens/''',
                            'method': '''...''',
                            'body': '''...''',
                            'descriptions': '''All endpoints for "JWT token" with django-rest_framework api''',
                            'helps': '''$BASEPATH$/tokens/''',
                            'code': '''...'''
                        },
                        {
                            'name': '''get token''',
                            'endpoint': '''$BASEPATH$/token/''',
                            'method': '''...''',
                            'body': '''...''',
                            'descriptions': '''All actions for "note" with django-rest_framework api''',
                            'helps': '''...''',
                            'code': '''...'''
                        },
                        {
                            'name': '''refresh token''',
                            'endpoint': '''$BASEPATH$/token/refresh/''',
                            'method': '''...''',
                            'body': '''...''',
                            'descriptions': '''refresh token with django-rest_framework api''',
                            'helps': '''-''',
                            'code': '''...'''
                        },
                        {
                            'name': '''verify token''',
                            'endpoint': '''$BASEPATH$/token/verify/''',
                            'method': '''...''',
                            'body': '''...''',
                            'descriptions': '''verify token with django-rest_framework api''',
                            'helps': '''-''',
                            'code': '''...'''
                        },
                        {
                            'name': '''note_api''',
                            'endpoint': '''$BASEPATH$/note_api/ && $BASEPATH$/note_api/<id>/''',
                            'method': '''...''',
                            'body': '''...''',
                            'descriptions': '''All actions for "note" with django-rest_framework api''',
                            'helps': '''$BASEPATH$/note_api/-1/''',
                            'code': '''...'''
                        },
                    ]}

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            else:
                return Response({"error": req_inst.return_not_allowed_action_type()})
        else:
            return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


# TODO main ############################################################################################################

@permission_classes([AllowAny])
def index(request):
    """
    index React Single Page Applications
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "GET":

            # TODO action ##############################################################################################

            if req_inst.action_type == "":
                try:
                    response = {"response": "Данные успешно получены!"}

                    # TODO response ####################################################################################

                    return render(request, 'production/index.html', response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return render(request, "backend/404.html")
            else:
                return Response({"error": req_inst.return_not_allowed_action_type()})
        else:
            return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([IsAuthenticated])
def api_user_ratings(request):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "GET":

            # TODO action ##############################################################################################

            if req_inst.action_type == "ratingsListStore":
                try:

                    # TODO get_value ###################################################################################

                    sort = req_inst.get_value(key="sort", default='')
                    only_month = req_inst.get_value(key="onlyMonth", default='')

                    # TODO action ######################################################################################

                    authors = []
                    ideas = backend_models.IdeaModel.objects.filter(
                        moderate_status="принято"
                    ).order_by("-updated")
                    if only_month:
                        now = (datetime.datetime.now()).strftime('%Y-%m-%d %H:%M')
                        local_objects = []
                        for idea in ideas:
                            if (idea.updated + datetime.timedelta(days=31)).strftime('%Y-%m-%d %H:%M') \
                                    >= now:
                                local_objects.append(idea.id)
                        ideas = backend_models.IdeaModel.objects.filter(
                            id__in=local_objects
                        ).order_by("-updated")
                    for idea in ideas:
                        authors.append(idea.author)
                    authors = set(authors)
                    ideas = []
                    for author in authors:
                        ideas_arr = backend_models.IdeaModel.objects.filter(
                            author=author, moderate_status="принято"
                        )
                        idea_count = ideas_arr.count()
                        idea_rating = 0
                        idea_rating_count = 0
                        idea_comment_count = 0
                        for idea in ideas_arr:
                            ratings = backend_models.IdeaRatingModel.objects.filter(idea=idea)
                            for rate in ratings:
                                idea_rating += rate.rating
                            idea_rating_count += ratings.count()
                            idea_comment_count = backend_models.IdeaCommentModel.objects.filter(
                                idea=idea
                            ).count()
                        if idea_rating_count == 0:
                            idea_rating = 0
                        else:
                            idea_rating = round(idea_rating / idea_rating_count, 2)
                        serialized_user = backend_serializers.UserModelSerializer(instance=author, many=False).data
                        ideas.append({
                            "user_model": serialized_user,
                            "idea_count": idea_count, "idea_rating": idea_rating,
                            "idea_rating_count": idea_rating_count, "idea_comment_count": idea_comment_count
                        })
                    # sort
                    if sort:
                        if sort == "количеству (наибольшие в начале)":
                            ideas_arr = []
                            for idea in ideas:
                                ideas_arr.append([idea["idea_count"], idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=True)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "количеству (наибольшие в конце)":
                            ideas_arr = []
                            for idea in ideas:
                                ideas_arr.append([idea["idea_count"], idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=False)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "рейтингу (популярные в начале)":
                            ideas_arr = []
                            for idea in ideas:
                                ideas_arr.append([idea["idea_rating"], idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=True)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "рейтингу (популярные в конце)":
                            ideas_arr = []
                            for idea in ideas:
                                ideas_arr.append([idea["idea_rating"], idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=False)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "отметкам рейтинга (наибольшие в начале)":
                            ideas_arr = []
                            for idea in ideas:
                                ideas_arr.append([idea["idea_rating_count"], idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=True)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "отметкам рейтинга (наибольшие в конце)":
                            ideas_arr = []
                            for idea in ideas:
                                ideas_arr.append([idea["idea_rating_count"], idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=False)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "комментариям (наибольшие в начале)":
                            ideas_arr = []
                            for idea in ideas:
                                ideas_arr.append([idea["idea_comment_count"], idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=True)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "комментариям (наибольшие в конце)":
                            ideas_arr = []
                            for idea in ideas:
                                ideas_arr.append([idea["idea_comment_count"], idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=False)
                            ideas = [x[1] for x in ideas_arr]
                    response = {
                        "response": {
                            "list": ideas,
                            "x-total-count": len(ideas)
                        }
                    }

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)

                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        else:
            return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


# TODO profile #########################################################################################################


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([AllowAny])
def api_captcha(request):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "GET":

            # TODO action ##############################################################################################

            if req_inst.action_type == "captchaCheckStore":
                try:

                    # TODO action ######################################################################################

                    response = {
                        "response": "Вы не робот!"
                    }
                    # TODO response ####################################################################################

                    time.sleep(round(random.uniform(1, 3), 2))

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([AllowAny])
def api_user_login(request):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "GET":

            # TODO action ##############################################################################################

            if req_inst.action_type == "userLoginStore":
                try:

                    # TODO get_value ###################################################################################

                    username = req_inst.get_value(key="username", default='')
                    password = req_inst.get_value(key="password", default='')

                    # TODO action ######################################################################################

                    access_count = 0
                    for log in backend_models.LoggingModel.objects.filter(
                            username=username,
                            ip=req_inst.ip,
                            path=req_inst.path,
                            method=f"{req_inst.method} | {req_inst.action_type}",
                            error="-"
                    ):
                        if (log.created +
                            datetime.timedelta(hours=6, minutes=59)).strftime('%Y-%m-%d %H:%M') >= \
                                (datetime.datetime.now()).strftime('%Y-%m-%d %H:%M'):
                            access_count += 1
                    if access_count < 20:
                        backend_models.LoggingModel.objects.create(
                            username=username,
                            ip=req_inst.ip,
                            path=req_inst.path,
                            method=f"{req_inst.method} | {req_inst.action_type}",
                            error="-"
                        )
                        is_authenticated = authenticate(username=username, password=password)
                        if is_authenticated is not None:
                            user = User.objects.get(username=username)
                            user_model = backend_models.UserModel.objects.get(user=user)
                            if user_model.is_active_account is False:
                                return Response({"error": "Внимание, Ваш аккаунт заблокирован!"})
                            update_last_login(sender=None, user=user)
                            refresh = RefreshToken.for_user(user=user)
                            response = {"response": {
                                "token": str(refresh.access_token),
                                "full name": f"{user_model.last_name} {user_model.first_name}"
                            }}

                            token = f"{username}_{password}_{backend_utils.DateTimeUtils.get_current_date()}"
                            token = make_password(token)
                            token_obj = backend_models.TokenModel.objects.get_or_create(user=user)[0]
                            token_obj.token = token
                            token_obj.updated = timezone.now()
                            token_obj.save()
                            response = {"response": {
                                "token": str(token),
                                "full name": f"{user_model.last_name} {user_model.first_name}"
                            }}

                        else:
                            response = {
                                "error": "Внимание, данные не совпадают! Проверьте правильность введения пароля!"
                            }
                    else:
                        response = {"error": "Внимание, попыток входа можно совершать не более 20 в час!"}

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        else:
            return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


from django.http import HttpResponse
from rest_framework.renderers import JSONRenderer
from rest_framework import status

class JSONResponse(HttpResponse):
    def __init__(self, data, **kwargs):
        content = JSONRenderer().render(data)
        kwargs['content_type'] = 'application/json'
        super(JSONResponse, self).__init__(content, **kwargs)


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([AllowAny])
# @permission_classes([IsAuthenticated])
# @authentication_classes([AllowAny])
def api_user_detail(request):
    """
    django-rest-framework
    """

    try:

        # try:
        #     token = str(request.META.get("HTTP_AUTHORIZATION", "1 0")).split(' ')[1]
        #     print("token: ", token)
        #     token_obj = backend_models.TokenModel.objects.get(token=token)
        #     user = token_obj.user
        #     print("user: ", user)
        #     response = {"response": backend_serializers.UserSerializer(user, many=False).data}
        #     return Response(response)
        # except Exception as error:
        #     return JSONResponse(data="error", status=status.HTTP_401_UNAUTHORIZED)

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "GET":

            # TODO action ##############################################################################################

            if req_inst.action_type == "userDetailStore":
                try:

                    # TODO action ######################################################################################

                    response = {"response": backend_serializers.UserSerializer(req_inst.user, many=False).data}

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)

                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        else:
            return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([IsAuthenticated])
def api_user_password_change(request):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "PUT":

            # TODO action ##############################################################################################

            if req_inst.action_type == "userPasswordUpdateStore":
                try:

                    # TODO get_value ###################################################################################

                    secret_question = req_inst.get_value(key="secretQuestion", default='')
                    secret_answer = req_inst.get_value(key="secretAnswer", default='')
                    email = req_inst.get_value(key="email", default=None)
                    password = req_inst.get_value(key="password", default='')
                    password2 = req_inst.get_value(key="password2", default='')

                    # TODO action ######################################################################################

                    user = User.objects.get(id=req_inst.user.id)
                    author = backend_models.UserModel.objects.get(user=user)
                    if len(password) < 1:
                        response = {"error": "Пароль пустой!"}
                    elif password != password2:
                        response = {"error": "Пароли не совпадают!"}
                    elif password == author.password:
                        response = {"error": "Пароль такой же как и предыдущий!"}
                    else:
                        user.set_password(password)
                        user.save()
                        author.password = password
                        author.is_temp_password = False
                        if secret_question and secret_question != author.secret_question:
                            author.secret_question = secret_question
                        if secret_answer and secret_answer != author.secret_answer:
                            author.secret_answer = secret_answer
                        if email and email != author.email:
                            author.email = email
                        author.save()
                        response = {"response": "Изменение успешно проведено."}

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)

                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        else:
            return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([AllowAny])
def api_user_recover(request):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "POST":

            # TODO action ##############################################################################################

            if req_inst.action_type == "userRecoverPasswordStore":
                try:

                    # TODO get_value ###################################################################################

                    username = req_inst.get_value(key="username", default='')
                    secret_answer = req_inst.get_value(key="secretAnswer", default='')
                    recover_password = req_inst.get_value(key="recoverPassword", default='')

                    # TODO action ######################################################################################

                    response = {"error": "ошибка действия!"}

                    if not secret_answer and not recover_password:
                        author = backend_models.UserModel.objects.get(
                            user=User.objects.get(username=username)
                        )
                        if author.is_temp_password:
                            response = {"error": {"Пользователь ещё ни разу не менял пароль!"}}
                        else:
                            response = {"response": {
                                "username": str(author.user.username),
                                "secretQuestion": str(author.secret_question),
                                "email": str(author.email),
                                "stage": "Second",
                            }}
                    elif secret_answer:
                        author = backend_models.UserModel.objects.get(
                            user=User.objects.get(username=username)
                        )
                        if str(secret_answer).strip().lower() == \
                                str(author.secret_answer).strip().lower():
                            response = {"response": {
                                "username": author.user.username,
                                "stage": "Third",
                            }}
                        else:
                            response = {"error": "Секретный ответ не совпадает!"}
                    elif recover_password:
                        author = backend_models.UserModel.objects.get(
                            user=User.objects.get(username=username)
                        )
                        decrypt_text = backend_utils.EncryptingClass.decrypt_text(
                            recover_password, author.password
                        )
                        text = f"{datetime.datetime.now().strftime('%Y-%m-%dT%H:%M')}_" \
                               f"{author.password[-1]}" \
                               f" {str(author.user)}{author.password} "
                        time1 = int(
                            decrypt_text.split('_')[0].split('T')[1].split(":")[0] +
                            decrypt_text.split('_')[0].split('T')[1].split(":")[1]
                        )
                        time2 = int(
                            text.split('_')[0].split('T')[1].split(":")[0] +
                            text.split('_')[0].split('T')[1].split(":")[1]
                        )
                        if time1 - time2 > -60:
                            if str(decrypt_text.split('_')[1]).strip() == str(text.split('_')[1]).strip():
                                response = {"response": {
                                    "username": author.user.username,
                                    "stage": "Third",
                                }}
                            else:
                                response = {"error": "Код не верный!"}
                        else:
                            response = {"error": "Код не верный или просрочен!"}

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        else:
            return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([AllowAny])
def api_user_recover_email(request):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "POST":

            # TODO action ##############################################################################################

            if req_inst.action_type == "userRecoverPasswordSendEmailStore":
                try:

                    # TODO get_value ###################################################################################

                    username = req_inst.get_value(key="username", default='')

                    # TODO action ######################################################################################

                    author = backend_models.UserModel.objects.get(
                        user=User.objects.get(username=username)
                    )
                    access_count = 0
                    for log in backend_models.LoggingModel.objects.filter(
                            username=username,
                            ip=req_inst.ip,
                            path=req_inst.path,
                            method=f"{req_inst.method} | {req_inst.action_type}",
                            error="-",
                    ):
                        if (log.created +
                            datetime.timedelta(hours=6, minutes=3)).strftime('%Y-%m-%d %H:%M') \
                                >= (datetime.datetime.now()).strftime('%Y-%m-%d %H:%M'):
                            access_count += 1
                    if access_count < 1:
                        backend_models.LoggingModel.objects.create(
                            username=username,
                            ip=req_inst.ip,
                            path=req_inst.path,
                            method=f"{req_inst.method} | {req_inst.action_type}",
                            error="-",
                        )
                        text = f"{datetime.datetime.now().strftime('%Y-%m-%dT%H:%M')}_" \
                               f"{author.password[-1]} {str(author.user)}" \
                               f"{author.password}"
                        encrypt_text = backend_utils.EncryptingClass.encrypt_text(
                            text,
                            '31284'
                        )
                        subject = 'Восстановление пароля от веб платформы'
                        message_s = f'{author.first_name} {author.last_name}, ' \
                                    f'перейдите по ссылке: https://web.km.kz/recover_password => войти => ' \
                                    f'восстановить доступ к аккаунту => введите Ваш ИИН и затем в окне восстановления' \
                                    f' через почту введите этот код восстановления (без кавычек): "{encrypt_text}". ' \
                                    f'Внимание! Этот код действует в течении часа с момента отправки!'
                        if subject and message_s and author.email:
                            send_mail(
                                subject,
                                message_s,
                                "kostanayminerals@web.km.kz",
                                [author.email, ''],
                                fail_silently=False
                            )
                        response = {
                            "response":
                                {"username": str(author.user.username),
                                 "secret_question": str(author.secret_question),
                                 "email": str(author.email),
                                 "success": False}
                        }
                    else:
                        response = {"error": "Внимание, отправлять письмо можно не чаще раза в 3 минуты!"}

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        else:
            return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([AllowAny])
def api_user_recover_password(request):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "POST":

            # TODO action ##############################################################################################

            if req_inst.action_type == "userRecoverPasswordChangePasswordStore":
                try:

                    # TODO get_value ###################################################################################

                    username = req_inst.get_value(key="username", default='')
                    secret_question = req_inst.get_value(key="secretQuestion", default='')
                    secret_answer = req_inst.get_value(key="secretAnswer", default='')
                    email = req_inst.get_value(key="email", default=None)
                    password = req_inst.get_value(key="password", default='')
                    password2 = req_inst.get_value(key="password2", default='')

                    # TODO action ######################################################################################

                    user = User.objects.get(username=username)
                    author = backend_models.UserModel.objects.get(user=user)
                    if len(password) < 1:
                        response = {"error": "Пароль пустой!"}
                    elif password != password2:
                        response = {"error": "Пароли не совпадают!"}
                    elif password == author.password:
                        response = {"error": "Пароль такой же как и предыдущий!"}
                    else:
                        user.set_password(password)
                        user.save()
                        author.password = password
                        author.is_temp_password = False
                        author.secret_question = secret_question
                        author.secret_answer = secret_answer
                        author.email = email
                        author.save()
                        response = {"response": "Изменение успешно проведено."}

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        else:
            return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([IsAuthenticated])
def api_notification(request):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "GET":

            # TODO action ##############################################################################################

            if req_inst.action_type == "notificationReadListStore":
                try:

                    # TODO get_value ###################################################################################

                    page = req_inst.get_value(key="page", default=1)
                    limit = req_inst.get_value(key="limit", default=10)

                    # TODO action ######################################################################################

                    objects = backend_models.NotificationModel.objects.filter(
                        target_group_model=None,
                        target_user_model=req_inst.user_model,
                        is_visible=True,
                    )
                    serializer = backend_serializers.NotificationModelSerializer(objects, many=True).data
                    group_models = backend_models.GroupModel.objects.filter(users=req_inst.user_model)
                    objects = backend_models.NotificationModel.objects.filter(
                        target_group_model__in=group_models,
                        target_user_model=None,
                        is_visible=True,
                    ).order_by("-updated")
                    serializer1 = backend_serializers.NotificationModelSerializer(objects, many=True).data

                    serializers = serializer + serializer1

                    num_pages = len(serializers)
                    if limit > 0:
                        p = Paginator(serializers, limit)
                        serializers = p.page(page).object_list
                    response = {
                        "response": {
                            "list": serializers,
                            "x-total-count": num_pages
                        }
                    }

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)

                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        if req_inst.method == "POST":

            # TODO action ##############################################################################################

            if req_inst.action_type == "notificationCreateStore":
                try:

                    # TODO get_value ###################################################################################

                    name = req_inst.get_value(key="name", default="")
                    place = req_inst.get_value(key="place", default="")
                    description = req_inst.get_value(key="description", default="")

                    # TODO action ######################################################################################

                    if place == "банк идей":
                        model = backend_models.GroupModel.objects.get(name="moderator_idea")
                    else:
                        model = backend_models.GroupModel.objects.get(name="superuser")

                    backend_models.NotificationModel.objects.create(
                        author=req_inst.user_model,
                        target_group_model=model,
                        name=name,
                        place=place,
                        description=description,
                    )
                    response = {"response": "Успешно отправлено!"}

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)

                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        else:
            return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([IsAuthenticated])
def api_notification_id(request, notification_id):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "PUT":

            # TODO action ##############################################################################################

            if req_inst.action_type == "notificationUpdateStore":
                try:

                    # TODO action ######################################################################################

                    backend_models.NotificationModel.objects.get(id=notification_id).delete()
                    response = {
                        "response": "Успешно скрыто!",
                    }

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)

                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        else:
            return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([IsAuthenticated])
def api_user(request):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "GET":

            # TODO action ##############################################################################################

            if req_inst.action_type == "userReadListStore":
                try:

                    # TODO get_value ###################################################################################

                    authors = backend_models.UserModel.objects.filter(
                        is_active_account=True
                    ).order_by("last_name")

                    # TODO action ######################################################################################

                    users = []
                    for author in authors:
                        if author.user.is_superuser:
                            continue
                        users.append(f"{author.last_name} {author.first_name} "
                                     f"{author.personnel_number} ")
                    response = {
                        "response": {
                            "list": users,
                            "x-total-count": len(users)
                        }
                    }

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)

                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        else:
            return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


# TODO progress ########################################################################################################


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([IsAuthenticated])
def api_idea(request):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "GET":

            # TODO action ##############################################################################################

            if req_inst.action_type == "ideaReadListStore":
                try:

                    # TODO get_value ###################################################################################

                    page = req_inst.get_value(key="page", default=1)
                    limit = req_inst.get_value(key="limit", default=10)

                    moderate = req_inst.get_value(key="moderate", default="")
                    subdivision = req_inst.get_value(key="subdivision", default="")
                    sphere = req_inst.get_value(key="sphere", default="")
                    category = req_inst.get_value(key="category", default="")
                    author = req_inst.get_value(key="author", default="")
                    search = req_inst.get_value(key="search", default="")
                    sort = req_inst.get_value(key="sort", default="")

                    only_month = req_inst.get_value(key="onlyMonth", default=False)

                    # TODO action ######################################################################################

                    # TODO onlyMonth
                    ideas = backend_models.IdeaModel.objects.all().order_by("-updated")
                    if only_month:
                        now = (datetime.datetime.now()).strftime('%Y-%m-%d %H:%M')
                        local_objects = []
                        for idea in ideas:
                            if (idea.updated + datetime.timedelta(days=31)).strftime('%Y-%m-%d %H:%M') \
                                    >= now:
                                local_objects.append(idea.id)
                        ideas = backend_models.IdeaModel.objects.filter(id__in=local_objects). \
                            order_by("-updated")
                    # TODO search
                    if search:
                        ideas = ideas.filter(name__icontains=search)
                    # TODO filter
                    if subdivision:
                        ideas = ideas.filter(subdivision=subdivision)
                    if sphere:
                        ideas = ideas.filter(sphere=sphere)
                    if category:
                        ideas = ideas.filter(category=category)
                    if author:
                        if author == "self":
                            author = req_inst.user_model
                        else:
                            author = backend_models.UserModel.objects.get(
                                personnel_number=(str(author).split(" ")[-2]).strip()
                            )
                        ideas = ideas.filter(author=author)
                    if moderate:
                        ideas = ideas.filter(moderate_status=moderate)
                    # TODO sort
                    if sort:
                        if sort == "дате публикации (свежие в начале)":
                            ideas = ideas.order_by("-updated")
                        elif sort == "дате публикации (свежие в конце)":
                            ideas = ideas.order_by("updated")
                        elif sort == "названию (с начала алфавита)":
                            ideas = ideas.order_by("name")
                        elif sort == "названию (с конца алфавита)":
                            ideas = ideas.order_by("-name")
                        elif sort == "рейтингу (популярные в начале)":
                            ideas_arr = []
                            for idea in ideas:
                                ratings = backend_models.IdeaRatingModel.objects.filter(
                                    idea=idea
                                )
                                if ratings.count() <= 0:
                                    rate = 0
                                else:
                                    rate = 0
                                    for i in ratings:
                                        rate += i.rating
                                    rate = round(rate / ratings.count(), 3)
                                ideas_arr.append([rate, idea])

                            ideas_arr.sort(key=lambda x: x[0], reverse=True)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "рейтингу (популярные в конце)":
                            ideas_arr = []
                            for idea in ideas:
                                ratings = backend_models.IdeaRatingModel.objects.filter(
                                    idea=idea
                                )
                                if ratings.count() <= 0:
                                    rate = 0
                                else:
                                    rate = 0
                                    for i in ratings:
                                        rate += i.rating
                                    rate = round(rate / ratings.count(), 3)
                                ideas_arr.append([rate, idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=False)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "отметкам рейтинга (наибольшие в начале)":
                            ideas_arr = []
                            for idea in ideas:
                                ratings = backend_models.IdeaRatingModel.objects.filter(
                                    idea=backend_models.IdeaModel.objects.get(id=idea.id)
                                )
                                ideas_arr.append([ratings.count(), idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=True)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "отметкам рейтинга (наибольшие в конце)":
                            ideas_arr = []
                            for idea in ideas:
                                ratings = backend_models.IdeaRatingModel.objects.filter(
                                    idea=backend_models.IdeaModel.objects.get(id=idea.id)
                                )
                                ideas_arr.append([ratings.count(), idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=False)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "комментариям (наибольшие в начале)":
                            ideas_arr = []
                            for idea in ideas:
                                comments = backend_models.IdeaCommentModel.objects.filter(
                                    idea=backend_models.IdeaModel.objects.get(id=idea.id)
                                )
                                ideas_arr.append([comments.count(), idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=True)
                            ideas = [x[1] for x in ideas_arr]
                        elif sort == "комментариям (наибольшие в конце)":
                            ideas_arr = []
                            for idea in ideas:
                                comments = backend_models.IdeaCommentModel.objects.filter(
                                    idea=backend_models.IdeaModel.objects.get(id=idea.id)
                                )
                                ideas_arr.append([comments.count(), idea])
                            ideas_arr.sort(key=lambda x: x[0], reverse=False)
                            ideas = [x[1] for x in ideas_arr]

                    num_pages = len(ideas)
                    if limit > 0:
                        p = Paginator(ideas, limit)
                        ideas = p.page(page).object_list
                    response = {
                        "response": {
                            "list": backend_serializers.IdeaModelSerializer(instance=ideas, many=True).data,
                            "x-total-count": num_pages
                        }
                    }

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)

                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        if req_inst.method == "POST":

            # TODO action ##############################################################################################

            if req_inst.action_type == "ideaCreateStore":
                try:

                    # TODO get_value ###################################################################################

                    token = str(request.META.get("HTTP_AUTHORIZATION", "1 0")).split(' ')[1]
                    print("token: ", token)
                    token_obj = backend_models.TokenModel.objects.get(token=token)
                    user = token_obj.user
                    print("user: ", user)
                    user_model = backend_models.UserModel.objects.get(user=user)
                    print("user_model: ", user_model)

                    subdivision = req_inst.get_value(key="subdivision", default="")
                    sphere = req_inst.get_value(key="sphere", default="")
                    category = req_inst.get_value(key="category", default="")
                    avatar = req_inst.get_value(key="avatar", default=None)
                    name = req_inst.get_value(key="name", default="")
                    place = req_inst.get_value(key="place", default="")
                    description = req_inst.get_value(key="description", default="")
                    moderate = req_inst.get_value(key="moderate", default="на модерации")

                    # TODO action ######################################################################################

                    backend_models.IdeaModel.objects.create(
                        # author=req_inst.user_model,
                        author=user_model,
                        subdivision=subdivision,
                        sphere=sphere,
                        category=category,
                        image=avatar,
                        name=name,
                        place=place,
                        description=description,
                        moderate_status=moderate
                    )

                    backend_models.NotificationModel.objects.create(
                        # author=req_inst.user_model,
                        author=user_model,
                        target_group_model=backend_models.GroupModel.objects.get(
                            name="moderator_idea"
                        ),
                        name=f"Создана новая идея",
                        place="банк идей",
                        description=f"название идеи: {name}",
                    )

                    response = {
                        "response": "Идея успешно отправлена на модерацию! Ожидайте уведомлений о статусе."
                    }

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)

                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([IsAuthenticated])
def api_idea_id(request, idea_id):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "GET":

            # TODO action ##############################################################################################

            if req_inst.action_type == "ideaReadStore":
                try:

                    # TODO action ######################################################################################

                    obj = backend_models.IdeaModel.objects.get(id=idea_id)
                    response = {
                        "response": backend_serializers.IdeaModelSerializer(instance=obj, many=False).data,
                    }

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)

                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        if req_inst.method == "PUT":

            # TODO action ##############################################################################################

            if req_inst.action_type == "ideaUpdateStore":
                try:

                    # TODO get_value ###################################################################################

                    subdivision = req_inst.get_value(key="subdivision", default="")
                    sphere = req_inst.get_value(key="sphere", default="")
                    category = req_inst.get_value(key="category", default="")
                    clear_image = req_inst.get_value("clearImage", False)
                    avatar = req_inst.get_value(key="avatar", default=None)
                    name = req_inst.get_value(key="name", default="")
                    place = req_inst.get_value(key="place", default="")
                    description = req_inst.get_value(key="description", default="")
                    moderate = req_inst.get_value(key="moderate", default="")

                    moderate_comment = req_inst.get_value(key="moderateComment", default="")

                    # TODO action ######################################################################################

                    idea = backend_models.IdeaModel.objects.get(id=idea_id)
                    if subdivision and idea.subdivision != subdivision:
                        idea.subdivision = subdivision
                    if sphere and idea.sphere != sphere:
                        idea.sphere = sphere
                    if category and idea.category != category:
                        idea.category = category
                    if clear_image:
                        idea.image = None
                    if avatar and idea.image != avatar:
                        idea.image = avatar
                    if name and idea.name != name:
                        idea.name = name
                    if place and idea.place != place:
                        idea.place = place
                    if description and idea.description != description:
                        idea.description = description
                    if moderate and idea.moderate_status != moderate:
                        idea.moderate_status = moderate
                        idea.moderate_author = req_inst.user_model
                    if moderate_comment and idea.moderate_comment != moderate_comment:
                        idea.moderate_comment = moderate_comment

                    idea.updated = timezone.now()
                    idea.save()
                    if req_inst.user_model == idea.author:
                        message = "Идея отредактирована автором"
                        if moderate == "скрыто":
                            message = "Автор скрыл свою идею!"
                        backend_models.NotificationModel.objects.create(
                            author=req_inst.user_model,
                            target_group_model=backend_models.GroupModel.objects.get(
                                name="moderator_idea"
                            ),
                            name=message,
                            place="банк идей",
                            description=f"название идеи: {idea.name}",
                        )
                    else:
                        if moderate == "принято" or moderate == "на доработку":
                            message = "Действия с Вашей идеей"
                            if moderate == "принято":
                                message = "Ваша идея успешно принята и открыта в общем доступе!"
                            if moderate == "на доработку":
                                message = "Ваша идея отправлена на доработку! Проверьте список идей на доработку и " \
                                          "исправьте её!"
                            backend_models.NotificationModel.objects.create(
                                author=req_inst.user_model,
                                target_user_model=idea.author,
                                name=message,
                                place="банк идей",
                                description=f"название идеи: {idea.name}",
                            )

                    response = {"response": "успешно изменено!"}

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": "Произошла ошибка!"})
        return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([IsAuthenticated])
def api_idea_comment(request, idea_id):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "GET":

            # TODO action ##############################################################################################

            if req_inst.action_type == "ideaCommentReadListStore":
                try:

                    # TODO get_value ###################################################################################

                    page = req_inst.get_value(key="page", default=1)
                    limit = req_inst.get_value(key="limit", default=10)

                    # TODO action ######################################################################################

                    idea = backend_models.IdeaModel.objects.get(id=idea_id)
                    comments = backend_models.IdeaCommentModel.objects.filter(idea=idea)
                    total_count = len(comments)
                    current_page = 1
                    if limit > 0:
                        p = Paginator(comments, limit)
                        comments = p.page(page).object_list
                        current_page = p.num_pages
                    response = {
                        "response": {
                            "list": backend_serializers.IdeaCommentModelSerializer(instance=comments, many=True).data,
                            "x-total-count": total_count,
                            "current-page": current_page
                        }
                    }

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)

                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        if req_inst.method == "POST":

            # TODO action ##############################################################################################

            if req_inst.action_type == "ideaCommentCreateStore":
                try:
                    # TODO get_value ###################################################################################

                    comment = req_inst.get_value(key="comment", default="")

                    # TODO action ######################################################################################

                    backend_models.IdeaCommentModel.objects.create(
                        author=req_inst.user_model,
                        idea=backend_models.IdeaModel.objects.get(id=idea_id),
                        comment=comment,
                    )

                    response = {
                        "response": "Комментарий успешно отправлен!"
                    }

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)

                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([IsAuthenticated])
def api_idea_comment_id(request, comment_id):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "DELETE":

            # TODO action ##############################################################################################

            if req_inst.action_type == "ideaCommentDeleteStore":
                try:

                    # TODO action ######################################################################################

                    backend_models.IdeaCommentModel.objects.get(id=comment_id).delete()
                    response = {
                        "response": "Успешно удалено!",
                    }

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)

                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([IsAuthenticated])
def api_idea_rating(request, idea_id):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "GET":

            # TODO action ##############################################################################################

            if req_inst.action_type == "ideaRatingReadListStore":
                try:

                    page = req_inst.get_value(key="page", default=1)
                    limit = req_inst.get_value(key="limit", default=10)

                    # TODO get_value ###################################################################################

                    objects = backend_models.IdeaRatingModel.objects.filter(
                        idea=backend_models.IdeaModel.objects.get(id=idea_id)
                    )
                    total_count = len(objects)
                    current_page = 1

                    rate = 0
                    if total_count > 0:
                        for i in objects:
                            rate += i.rating
                        rate = round(rate / total_count, 2)

                    if limit > 0:
                        p = Paginator(objects, limit)
                        objects = p.page(page).object_list
                        current_page = p.num_pages

                    try:
                        self_rating = backend_models.IdeaRatingModel.objects.get(
                            author=req_inst.user_model,
                            idea=backend_models.IdeaModel.objects.get(id=idea_id)
                        )
                        self_rate = self_rating.rating

                    except Exception as error:
                        self_rate = 0
                    response = {
                        "response": {
                            "list": backend_serializers.IdeaRatingModelSerializer(instance=objects, many=True).data,
                            "x-total-count": total_count,
                            "total_rate": rate,
                            "self_rate": self_rate,
                            "current-page": current_page
                        }
                    }

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)

                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        if req_inst.method == "PUT":

            # TODO action ##############################################################################################

            if req_inst.action_type == "ideaRatingUpdateStore":
                try:
                    # TODO get_value ###################################################################################

                    rating = req_inst.get_value(key="rating", default=0)

                    # TODO action ######################################################################################

                    obj = backend_models.IdeaRatingModel.objects.get_or_create(
                        author=req_inst.user_model,
                        idea=backend_models.IdeaModel.objects.get(id=idea_id),
                    )[0]

                    if rating != obj.rating:
                        obj.rating = rating
                        obj.save()

                    response = {
                        "response": "Рейтинг успешно применён!"
                    }

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)

                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


# TODO buh #############################################################################################################


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([IsAuthenticated])
def api_salary(request):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "GET":

            # TODO action ##############################################################################################

            if req_inst.action_type == "salaryReadStore":
                try:

                    # TODO get_value ###################################################################################

                    date_time = req_inst.get_value(key="dateTime", default="202201")

                    # TODO action ######################################################################################

                    # Get json response from 1c
                    key = backend_utils.UtilsClass.create_encrypted_password(
                        _random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890',
                        _length=10
                    )
                    hash_key_obj = hashlib.sha256()
                    hash_key_obj.update(key.encode('utf-8'))
                    key_hash = str(hash_key_obj.hexdigest().strip().upper())
                    key_hash_base64 = base64.b64encode(str(key_hash).encode()).decode()
                    iin = req_inst.user.username
                    if str(iin).lower() == '000000000000':
                        iin = 970801351179
                    iin_base64 = base64.b64encode(str(iin).encode()).decode()
                    date_base64 = base64.b64encode(f'{date_time}'.encode()).decode()
                    url = f'http://192.168.1.10/KM_1C/hs/zp/rl/{iin_base64}_{key_hash_base64}/{date_base64}'
                    h = httplib2.Http(
                        os.path.dirname(os.path.abspath('__file__')) + "/static/media/data/temp/get_salary_data"
                    )
                    _login = 'Web_adm_1c'
                    _password = '159159qqww!'
                    h.add_credentials(_login, _password)
                    response, content = h.request(url)

                    data = backend_utils.UtilsClass.decrypt_text_with_hash(content.decode()[1:], key_hash)
                    error_word_list = ['ошибка', 'error', 'failed']
                    if data.find('send') == 0:
                        return Response({"error": f"{data.split('send')[1].strip()}"})
                    else:
                        for error_word in error_word_list:
                            if data.find(error_word.lower()) >= 0:
                                return Response({"error": f"Произошла ошибка!"})

                        # Get local test json response from 1c
                        # Временное чтение файла для отладки без доступа к 1С
                        # with open("static/media/data/json_data.json", "r", encoding="utf-8") as file:
                        #     data = json.load(file)
                        #     time.sleep(3)
                        json_data = json.loads(data)

                        try:
                            json_data["global_objects"]["1.Начислено"]
                        except Exception as error:
                            json_data["global_objects"]["1.Начислено"] = {
                                "Fields": {
                                    "1": "Вид",
                                    "2": "Период",
                                    "3": "Дни",
                                    "4": "Период",
                                    "5": "Часы",
                                    "6": "ВсегоДни",
                                    "7": "ВсегоЧасы",
                                    "8": "Сумма"
                                },
                            }
                        try:
                            json_data["global_objects"]["2.Удержано"]
                        except Exception as error:
                            json_data["global_objects"]["2.Удержано"] = {
                                "Fields": {
                                    "1": "Вид",
                                    "2": "Период",
                                    "3": "Сумма"
                                },
                            }
                        try:
                            json_data["global_objects"]["3.Доходы в натуральной форме"]
                        except Exception as error:
                            json_data["global_objects"]["3.Доходы в натуральной форме"] = {
                                "Fields": {
                                    "1": "Вид",
                                    "2": "Период",
                                    "3": "Сумма"
                                },
                            }
                        try:
                            json_data["global_objects"]["4.Выплачено"]
                        except Exception as error:
                            json_data["global_objects"]["4.Выплачено"] = {
                                "Fields": {
                                    "1": "Вид",
                                    "2": "Период",
                                    "3": "Сумма"
                                },
                            }
                        try:
                            json_data["global_objects"]["5.Налоговые вычеты"]
                        except Exception as error:
                            json_data["global_objects"]["5.Налоговые вычеты"] = {
                                "Fields": {
                                    "1": "Вид",
                                    "2": "Период",
                                    "3": "Сумма"
                                },
                            }

                        # Return pretty integer and float value
                        def return_float_value(_value):
                            if isinstance(_value, int) or isinstance(_value, float):
                                if len(f'{_value:.2f}') < 10:
                                    _value = f'{_value:.2f}'[:-6] + ' ' + f'{_value:.2f}'[-6:]
                                else:
                                    _value = f'{_value:.2f}'[:-9] + ' ' + f'{_value:.2f}'[-9:-6] + ' ' + \
                                             f'{_value:.2f}'[-6:]
                            return _value

                        # Create 'Ends' and pretty integer and float value
                        def create_ends(table: str, extracols=False):
                            if extracols:
                                _days = 0
                                _hours = 0
                                _summ = 0
                                for __key in json_data['global_objects'][table].keys():
                                    if __key != 'Fields':
                                        try:
                                            _days += json_data['global_objects'][table][f'{__key}']['ВсегоДни']
                                            _hours += json_data['global_objects'][table][f'{__key}']['ВсегоЧасы']
                                            _summ_local = json_data['global_objects'][table][f'{__key}']['Сумма']
                                            json_data['global_objects'][table][f'{__key}'][
                                                'Сумма'] = return_float_value(
                                                _summ_local)
                                            _summ += _summ_local
                                        except Exception as error_:
                                            backend_utils.DjangoClass.LoggingClass.error(request=request,
                                                                                         error=error_)
                                json_data['global_objects'][table]['Ends'] = {
                                    'Вид': 'Итого', 'Период': '', 'Дни': _days, 'Часы': _hours,
                                    'ВсегоДни': 0, 'ВсегоЧасы': 0, 'Сумма': return_float_value(_summ)
                                }
                            else:
                                _summ = 0
                                for __key in json_data['global_objects'][table].keys():
                                    if __key != 'Fields':
                                        try:
                                            _summ_local = json_data['global_objects'][table][f'{__key}']['Сумма']
                                            json_data['global_objects'][table][f'{__key}'][
                                                'Сумма'] = return_float_value(
                                                _summ_local)
                                            _summ += _summ_local
                                        except Exception as error_:
                                            backend_utils.DjangoClass.LoggingClass.error(request=request,
                                                                                         error=error_)
                                json_data['global_objects'][table]['Ends'] = {
                                    'Вид': 'Итого', 'Период': '', 'Сумма': return_float_value(_summ)
                                }

                        create_ends(table='1.Начислено', extracols=True)
                        create_ends(table='2.Удержано', extracols=False)
                        create_ends(table='3.Доходы в натуральной форме', extracols=False)
                        create_ends(table='4.Выплачено', extracols=False)
                        create_ends(table='5.Налоговые вычеты', extracols=False)

                        # pretty integer and float value in headers
                        for _key in json_data.keys():
                            if _key != 'global_objects':
                                json_data[_key] = return_float_value(json_data[_key])

                        # create excel
                        key = backend_utils.UtilsClass.create_encrypted_password(
                            _random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890',
                            _length=24
                        )
                        date = backend_utils.DateTimeUtils.get_current_date()
                        path = 'media/data/temp/salary'
                        file_name = f"salary_{key}_{date}.xlsx"
                        workbook = backend_utils.ExcelClass.workbook_create()
                        sheet = backend_utils.ExcelClass.workbook_activate(workbook)

                        # Delete old files
                        for root, dirs, files in os.walk(f"static/{path}", topdown=True):
                            for file in files:
                                try:
                                    date_file = str(file).strip().split('.')[0].strip().split('_')[-1]
                                    if date != date_file:
                                        os.remove(f'{path}/{file}')
                                except Exception as error:
                                    pass

                        # Create 'TitleComponent'
                        backend_utils.ExcelClass.set_sheet_value(
                            col=1,
                            row=1,
                            value='РАСЧЕТНЫЙ ЛИСТ',
                            sheet=sheet
                        )
                        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
                        #######################################################

                        # Create 'Headers'
                        header_arr = []
                        for key, value in json_data.items():
                            if key != 'global_objects' and key != 'Долг за организацией на начало месяца' \
                                    and key != 'Долг за организацией на конец месяца':
                                header_arr.append(f'{key}: {value}')

                        header_len_devide = len(header_arr) - 8
                        if header_len_devide % 2 != 0:
                            header_len_devide += 1

                        row_i_1 = 1 + 1
                        for header in header_arr[0:4]:
                            col_i = 1
                            backend_utils.ExcelClass.set_sheet_value(
                                col=col_i,
                                row=row_i_1,
                                value=header,
                                sheet=sheet
                            )
                            sheet.merge_cells(start_row=row_i_1, start_column=col_i, end_row=row_i_1,
                                              end_column=col_i + 4)
                            row_i_1 += 1

                        row_i_2 = 1 + 1
                        for header in header_arr[4:8]:
                            col_i = 6
                            backend_utils.ExcelClass.set_sheet_value(
                                col=col_i,
                                row=row_i_2,
                                value=header,
                                sheet=sheet
                            )
                            sheet.merge_cells(start_row=row_i_2, start_column=col_i, end_row=row_i_2,
                                              end_column=col_i + 2)
                            row_i_2 += 1

                        row_i_3 = row_i_1
                        for header in header_arr[8:8 + header_len_devide // 2]:
                            col_i = 1
                            backend_utils.ExcelClass.set_sheet_value(
                                col=col_i,
                                row=row_i_3,
                                value=header,
                                sheet=sheet
                            )
                            sheet.merge_cells(start_row=row_i_3, start_column=col_i, end_row=row_i_3,
                                              end_column=col_i + 4)
                            row_i_3 += 1

                        row_i_4 = row_i_2
                        for header in header_arr[8 + header_len_devide // 2:]:
                            col_i = 6
                            backend_utils.ExcelClass.set_sheet_value(
                                col=col_i,
                                row=row_i_4,
                                value=header,
                                sheet=sheet
                            )
                            sheet.merge_cells(start_row=row_i_4, start_column=col_i, end_row=row_i_4,
                                              end_column=col_i + 2)
                            row_i_4 += 1
                        header_low_row = row_i_4

                        #######################################################

                        # Create 'Bodyes'
                        def create_bodyes(table: str, extracols=False):
                            bodyes_arr = [[table]]
                            for __key, _value in json_data['global_objects'][table].items():
                                local_bodyes_arr = []
                                for ___key, __value in json_data['global_objects'][table][__key].items():
                                    if extracols:
                                        if ___key != '6' and ___key != '7' and \
                                                ___key != 'ВсегоДни' and ___key != 'ВсегоЧасы':
                                            local_bodyes_arr.append(__value)
                                    else:
                                        local_bodyes_arr.append(__value)
                                bodyes_arr.append(local_bodyes_arr)
                            return bodyes_arr

                        bodyes_arr_1 = create_bodyes('1.Начислено', extracols=True)
                        bodyes_arr_2 = create_bodyes('2.Удержано', extracols=False)
                        bodyes_arr_3 = create_bodyes('3.Доходы в натуральной форме', extracols=False)
                        bodyes_arr_4 = create_bodyes('4.Выплачено', extracols=False)
                        bodyes_arr_5 = create_bodyes('5.Налоговые вычеты', extracols=False)

                        bold_arr = []

                        body_low_row_1 = header_low_row
                        bold_arr.append(body_low_row_1)
                        sheet.merge_cells(start_row=body_low_row_1, start_column=1, end_row=body_low_row_1,
                                          end_column=1 + 4)
                        for body in bodyes_arr_1:
                            col_i = 1
                            for value in body:
                                if isinstance(value, int) and value == 0:
                                    value = ''
                                backend_utils.ExcelClass.set_sheet_value(
                                    col=col_i,
                                    row=body_low_row_1,
                                    value=value,
                                    sheet=sheet
                                )
                                col_i += 1
                            body_low_row_1 += 1

                        body_low_row_2 = header_low_row
                        bold_arr.append(body_low_row_2)
                        sheet.merge_cells(start_row=body_low_row_2, start_column=6, end_row=body_low_row_2,
                                          end_column=6 + 2)
                        for body in bodyes_arr_2:
                            col_i = 6
                            for value in body:
                                backend_utils.ExcelClass.set_sheet_value(
                                    col=col_i,
                                    row=body_low_row_2,
                                    value=value,
                                    sheet=sheet
                                )
                                col_i += 1
                            body_low_row_2 += 1

                        if body_low_row_1 >= body_low_row_2:
                            body_low_row_3 = body_low_row_1
                            body_low_row_4 = body_low_row_1
                        else:
                            body_low_row_3 = body_low_row_2
                            body_low_row_4 = body_low_row_2

                        bold_arr.append(body_low_row_3)
                        sheet.merge_cells(start_row=body_low_row_3, start_column=1, end_row=body_low_row_3,
                                          end_column=1 + 4)
                        for body in bodyes_arr_3:
                            col_i = 1
                            for value in body:
                                backend_utils.ExcelClass.set_sheet_value(
                                    col=col_i,
                                    row=body_low_row_3,
                                    value=value,
                                    sheet=sheet
                                )
                                col_i += 1
                            body_low_row_3 += 1

                        bold_arr.append(body_low_row_4)
                        sheet.merge_cells(start_row=body_low_row_4, start_column=6, end_row=body_low_row_4,
                                          end_column=6 + 2)
                        for body in bodyes_arr_4:
                            col_i = 6
                            for value in body:
                                backend_utils.ExcelClass.set_sheet_value(
                                    col=col_i,
                                    row=body_low_row_4,
                                    value=value,
                                    sheet=sheet
                                )
                                col_i += 1
                            body_low_row_4 += 1

                        if body_low_row_3 >= body_low_row_4:
                            body_low_row_5 = body_low_row_3
                            body_low_row_6 = body_low_row_3
                        else:
                            body_low_row_5 = body_low_row_4
                            body_low_row_6 = body_low_row_4

                        bold_arr.append(body_low_row_5)
                        sheet.merge_cells(start_row=body_low_row_5, start_column=1, end_row=body_low_row_5,
                                          end_column=1 + 4)
                        for body in bodyes_arr_5:
                            col_i = 1
                            for value in body:
                                backend_utils.ExcelClass.set_sheet_value(
                                    col=col_i,
                                    row=body_low_row_5,
                                    value=value,
                                    sheet=sheet
                                )
                                col_i += 1
                            body_low_row_5 += 1

                        lowest = [
                            f'Долг за организацией на начало месяца: '
                            f'{json_data["Долг за организацией на начало месяца"]}',
                            f'Долг за организацией на конец месяца: '
                            f'{json_data["Долг за организацией на конец месяца"]}',
                        ]

                        bold_arr.append(body_low_row_6)
                        for body in ['.', 'Вид', *lowest]:
                            col_i = 6
                            backend_utils.ExcelClass.set_sheet_value(
                                col=col_i,
                                row=body_low_row_6,
                                value=body,
                                sheet=sheet
                            )
                            sheet.merge_cells(start_row=body_low_row_6, start_column=col_i, end_row=body_low_row_6,
                                              end_column=col_i + 2)
                            body_low_row_6 += 1

                        if body_low_row_6 >= body_low_row_5:
                            body_low_row = body_low_row_6
                        else:
                            body_low_row = body_low_row_5

                        if body_low_row_1 >= body_low_row_2:
                            body_color_2 = body_low_row_1
                        else:
                            body_color_2 = body_low_row_2
                        if body_low_row_3 >= body_low_row_4:
                            body_color_3 = body_low_row_3
                        else:
                            body_color_3 = body_low_row_4
                        #######################################################

                        # Set fonts
                        #######################################################
                        font_headers = Font(name='Arial', size=8, bold=False)
                        for row in range(1, header_low_row):
                            for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                cell = sheet[f'{col}{row}']
                                cell.font = font_headers

                        font_bodyes = Font(name='Arial', size=7, bold=False)
                        for row in range(header_low_row, body_low_row + 1):
                            for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                cell = sheet[f'{col}{row}']
                                cell.font = font_bodyes

                        font_tables = Font(name='Arial', size=8, bold=True)
                        for row in [header_low_row, body_color_2, body_color_3]:
                            for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                cell = sheet[f'{col}{row}']
                                cell.font = font_tables

                        # Set aligments
                        # wrap_text = Alignment(wrap_text=True)
                        # shrink_to_fit = Alignment(shrink_to_fit=True)
                        aligment_center = Alignment(horizontal='center', vertical='center', wrap_text=True,
                                                    shrink_to_fit=True)
                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                            cell = sheet[f'{col}{1}']
                            cell.alignment = aligment_center
                        aligment_left = Alignment(horizontal='left', vertical='center', wrap_text=True,
                                                  shrink_to_fit=True)
                        for row in range(2, header_low_row):
                            for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                cell = sheet[f'{col}{row}']
                                cell.alignment = aligment_left

                        aligment_right = Alignment(horizontal='right', vertical='center', wrap_text=True,
                                                   shrink_to_fit=True)
                        for row in range(header_low_row, body_low_row + 1):
                            for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                cell = sheet[f'{col}{row}']
                                if col == 'A' or col == 'F':
                                    cell.alignment = aligment_left
                                elif col == 'E' or col == 'H':
                                    cell.alignment = aligment_right
                                else:
                                    cell.alignment = aligment_center

                        # Set borders
                        side_medium = Side(border_style="thin", color="FF808080")
                        # side_think = Side(border_style="thin", color="FF808080")
                        border_horizontal_middle = Border(
                            top=side_medium,
                            # left=side_medium,
                            # right=side_medium,
                            # bottom=side_medium
                        )
                        border_vertical_middle = Border(
                            # top=side_think,
                            left=side_medium,
                            # right=side_think,
                            # bottom=side_think
                        )
                        border_vertical_light = Border(
                            # top=side_think,
                            left=side_medium,
                            # right=side_think,
                            # bottom=side_think
                        )
                        for row in range(header_low_row, body_low_row):
                            for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                if col == 'G' and row > body_low_row_4 or col == 'H' and row > body_low_row_4:
                                    pass
                                else:
                                    cell = sheet[f'{col}{row}']
                                    cell.border = border_vertical_light
                            cell = sheet[f'{backend_utils.ExcelClass.get_column_letter(1)}{row}']
                            cell.border = border_vertical_middle
                            cell = sheet[f'{backend_utils.ExcelClass.get_column_letter(6)}{row}']
                            cell.border = border_vertical_middle
                            cell = sheet[f'{backend_utils.ExcelClass.get_column_letter(9)}{row}']
                            cell.border = border_vertical_middle
                        side_think = Side(border_style="dotted", color="FF808080")
                        # {'mediumDashDotDot', 'thin', 'dashed', 'mediumDashed', 'dotted', 'double', 'thick',
                        # 'medium', 'dashDot','dashDotDot', 'hair', 'mediumDashDot', 'slantDashDot'}
                        border_think = Border(
                            top=side_think,
                            left=side_think,
                            right=side_think,
                            bottom=side_think
                        )
                        for row in range(1, header_low_row):
                            for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                                cell = sheet[f'{col}{row}']
                                cell.border = border_think
                        side_medium = Side(border_style="thin", color="FF808080")
                        border_medium = Border(
                            top=side_medium,
                            left=side_medium,
                            right=side_medium,
                            bottom=side_medium
                        )
                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                            for row in [header_low_row + 1, body_color_2 + 1, body_color_3 + 1]:
                                cell = sheet[f'{col}{row - 1}']
                                cell.border = border_medium
                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                            cell = sheet[f'{col}{body_low_row}']
                            cell.border = border_horizontal_middle
                        #######################################################

                        # Colored styles
                        #######################################################
                        color_green = PatternFill(fgColor="E6E6FF", fill_type="solid")
                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                            for row in [header_low_row + 1, body_color_2 + 1, body_color_3 + 1]:
                                cell = sheet[f'{col}{row}']
                                cell.fill = color_green
                                cell = sheet[f'{col}{row - 1}']
                                cell.border = border_medium
                        color_yellow = PatternFill(fgColor="d0ffd8", fill_type="solid")
                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 5 + 1)]:
                            cell = sheet[f'{col}{body_low_row_1 - 1}']
                            cell.fill = color_yellow

                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(6, 8 + 1)]:
                            cell = sheet[f'{col}{body_low_row_2 - 1}']
                            cell.fill = color_yellow

                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 3 + 1)]:
                            cell = sheet[f'{col}{body_low_row_3 - 1}']
                            cell.fill = color_yellow
                            cell.fill = color_yellow

                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(6, 8 + 1)]:
                            cell = sheet[f'{col}{body_low_row_4 - 1}']
                            cell.fill = color_yellow

                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 3 + 1)]:
                            cell = sheet[f'{col}{body_low_row_5 - 1}']
                            cell.fill = color_yellow
                            cell.fill = color_yellow

                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(6, 8 + 1)]:
                            cell = sheet[f'{col}{body_low_row_6 - 1}']
                            cell.fill = color_yellow

                        #######################################################

                        # Height and width styles
                        #######################################################
                        for col in [backend_utils.ExcelClass.get_column_letter(x) for x in range(1, 8 + 1)]:
                            width = 1
                            for row in range(1, body_low_row + 1):
                                cell = sheet[f'{col}{row}']
                                value = len(str(cell.value))
                                if value > width:
                                    width = value
                            if col == 'A' or col == 'F':
                                width = width / 2
                            sheet.column_dimensions[col].height = 1
                            sheet.column_dimensions[col].width = round((width * 0.95), 3)
                        #######################################################

                        # Set global page and book settings
                        #######################################################
                        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
                        sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
                        sheet.page_margins.left = 0.05
                        sheet.page_margins.right = 0.05
                        sheet.page_margins.header = 0.1
                        sheet.page_margins.bottom = 0.2
                        sheet.page_margins.footer = 0.2
                        sheet.page_margins.top = 0.1
                        sheet.print_options.horizontalCentered = True
                        # sheet.print_options.verticalCentered = True
                        sheet.page_setup.fitToHeight = 1
                        sheet.page_setup.scale = 80
                        sheet.page_setup.fitToHeight = 1
                        sheet.page_setup.fitToWidth = 1
                        sheet.protection.password = key + '_1'
                        sheet.protection.sheet = True
                        sheet.protection.enable()
                        #######################################################
                        try:
                            backend_utils.ExcelClass.workbook_save(
                                workbook=workbook,
                                excel_file=f"static/{path}/{file_name}"
                            )
                        except Exception as error:
                            pass

                        headers = []
                        for x in json_data.keys():
                            if x != "global_objects":
                                headers.append([x, json_data[x]])
                        tables = [
                            ["1.Начислено", json_data["global_objects"]["1.Начислено"]],
                            ["2.Удержано", json_data["global_objects"]["2.Удержано"]],
                            [
                                "3.Доходы в натуральной форме",
                                json_data["global_objects"]["3.Доходы в натуральной форме"]
                            ],
                            ["4.Выплачено", json_data["global_objects"]["4.Выплачено"]],
                            ["5.Налоговые вычеты", json_data["global_objects"]["5.Налоговые вычеты"]]
                        ]
                        data = {"excelPath": f"static/{path}/{file_name}", "headers": headers, "tables": tables}

                        response = {"response": data}

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        else:
            return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


# TODO sup #############################################################################################################


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([IsAuthenticated])
def api_vacation(request):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "GET":

            # TODO action ##############################################################################################

            if req_inst.action_type == "vacationReadStore":
                try:

                    # TODO get_value ###################################################################################

                    date_time = req_inst.get_value(key="dateTime", default="20220101")

                    # TODO action ######################################################################################

                    key = backend_utils.UtilsClass.create_encrypted_password(
                        _random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890',
                        _length=10
                    )
                    hash_key_obj = hashlib.sha256()
                    hash_key_obj.update(key.encode('utf-8'))
                    key_hash = str(hash_key_obj.hexdigest().strip().upper())
                    key_hash_base64 = base64.b64encode(str(key_hash).encode()).decode()
                    iin = req_inst.user.username
                    if str(iin).lower() == '000000000000' or str(iin).lower() == 'bogdan':
                        iin = 970801351179
                    iin_base64 = base64.b64encode(str(iin).encode()).decode()
                    date_base64 = base64.b64encode(f'{date_time}'.encode()).decode()
                    url = f'http://192.168.1.10/KM_1C/hs/pers/u_vacation/{iin_base64}_{key_hash_base64}/{date_base64}'
                    h = httplib2.Http(
                        os.path.dirname(os.path.abspath('__file__')) + "/static/media/data/temp/get_vacation_data"
                    )
                    _login = 'Web_adm_1c'
                    _password = '159159qqww!'
                    h.add_credentials(_login, _password)
                    response, content = h.request(url)

                    data = backend_utils.UtilsClass.decrypt_text_with_hash(content.decode()[1:], key_hash)
                    error_word_list = ['ошибка', 'error', 'failed']
                    if data.find('send') == 0:
                        return Response({"error": f"{data.split('send')[1].strip()}"})
                    json_data = json.loads(data)
                    for error_word in error_word_list:
                        if data.find(error_word.lower()) >= 0:
                            return Response({"error": f"Произошла ошибка!"})

                    headers = []
                    tables = []
                    for key, value in json_data.items():
                        try:
                            if key != "global_objects":
                                try:
                                    if key == "Работник":
                                        headers.append(["Работник", value])
                                    elif key == "Подразделение":
                                        headers.append(["Место работы", value])
                                    elif key == "Должность":
                                        headers.append(["Должность", value])
                                    elif key == "На дату":
                                        headers.append(["Дни отпуска на дату - ", value])
                                    elif key == "Фактически заработанные дни":
                                        headers.append([
                                            "Количество дней неиспользованного отпуска на выбранную дату",
                                            value
                                        ])
                                    else:
                                        headers.append([key, value])
                                except Exception as error:
                                    pass
                            else:
                                for key_, value_ in json_data["global_objects"].items():
                                    tab = []
                                    for key__, value__ in json_data["global_objects"][key_].items():
                                        try:
                                            if key__ == "По графику ":
                                                tab.append(["По графику: ", value__])
                                            elif key__ == "Период":
                                                tab.append(["Дата начала", value__[0:11]])
                                            elif key__ == "ДатаОкончания":
                                                tab.append(["Дата окончания", value__[0:11]])
                                            elif key__ == "ДнейОтпуска":
                                                tab.append(["Запланированные дни отпуска", value__])
                                            elif key__ == "Рабочий год":
                                                tab.append(["Год", value__])
                                            else:
                                                tab.append([key__, value__])
                                        except Exception as error:
                                            pass
                                    tables.append(tab)
                        except Exception as error:
                            pass
                    json_data = {
                        "headers": headers,
                        "tables": tables,
                    }
                    response = {"response": json_data}

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        else:
            return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


# TODO moderator #######################################################################################################

@api_view(http_method_names=HTTP_METHOD_NAMES)
@authentication_classes([BasicAuthentication])
def api_basic_admin_user_temp(request):
    """
    basic admin django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "GET":

            # TODO action ##############################################################################################

            try:

                # TODO action ##########################################################################################

                if req_inst.action_type == "":

                    authors = backend_models.UserModel.objects.filter(
                        is_temp_password=True,
                        is_active_account=True
                    )
                    if not request.user.is_superuser:
                        response = {"error": "Your user in not superuser."}
                    else:
                        objects = []
                        for author in authors:
                            if not author.user.is_superuser:
                                objects.append(
                                    {
                                        f"""{base64.b64encode(
                                            str(author.user.username)[::-1].encode()
                                        ).decode()}""":
                                            base64.b64encode(
                                                str(f"12{author.password}345").encode()).decode()
                                    }
                                )
                        # for obj in objects:
                        #     for key, value in obj.items():
                        #         print(f"{key}: {str(base64.b64decode(value).decode())[2: -3]}")
                        response = {"response": objects}

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)
                    return Response(response)
                else:
                    return Response({"error": req_inst.return_not_allowed_action_type()})
            except Exception as error:
                backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                return Response({"error": "Произошла ошибка!"})
        else:
            return Response({"error": "Этот метод не реализован для этой точки доступа."})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([IsAuthenticated])
def api_admin_export_users(request):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "GET":

            # TODO action ##############################################################################################

            if req_inst.action_type == "":
                try:

                    # TODO action ######################################################################################

                    key = backend_utils.UtilsClass.create_encrypted_password(
                        _random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890',
                        _length=24
                    )
                    date = backend_utils.DateTimeUtils.get_current_date()
                    path = 'media/data/temp/users'
                    file_name = f"users_{key}_{date}.xlsx"
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active

                    # Delete old files
                    #######################################################
                    for root, dirs, files in os.walk(f"static/{path}", topdown=True):
                        for file in files:
                            try:
                                date_file = str(file).strip().split('.')[0].strip().split('_')[-1]
                                if date != date_file:
                                    os.remove(f'{path}/{file}')
                            except Exception as error:
                                backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    #######################################################

                    users = User.objects.filter(is_superuser=False)

                    def set_value(_col: Union[str, int], _row: Union[str, int], _value, _sheet):
                        if isinstance(_col, int):
                            _col = get_column_letter(_col)
                        if isinstance(_row, str):
                            _row = str(_row)
                        if isinstance(_value, bool):
                            if _value:
                                _value = "true"
                            else:
                                _value = "false"
                        if _value is None:
                            _value = ""
                        _sheet[str(_col).upper() + str(_row)] = str(_value)

                    titles = ['Подразделение', 'Цех/Служба', 'Отдел/Участок', 'Фамилия', 'Имя', 'Отчество',
                              'Табельный номер', 'Должность', 'Категория работника', 'Имя пользователя',
                              'Пароль аккаунта', 'Активность аккаунта', 'Доступ к панели управления',
                              'Права суперпользователя', 'Временный пароль', 'Группы доступа', 'Электронная почта',
                              'Секретный вопрос', 'Секретный ответ']
                    for title in titles:
                        set_value(_col=titles.index(title) + 1, _row=1, _value=title, _sheet=sheet)
                    _index = 1
                    for user in users:
                        try:
                            _index += 1
                            author = backend_models.UserModel.objects.get_or_create(user=user)[0]

                            subdivision = author.subdivision
                            set_value(_col="A", _row=_index, _value=subdivision, _sheet=sheet)

                            workshop_service = author.workshop_service
                            set_value(_col="B", _row=_index, _value=workshop_service,
                                      _sheet=sheet)

                            department_site = author.department_site
                            set_value(_col="C", _row=_index, _value=department_site,
                                      _sheet=sheet)

                            last_name = author.last_name
                            set_value(_col="D", _row=_index, _value=last_name, _sheet=sheet)

                            first_name = author.first_name
                            set_value(_col="E", _row=_index, _value=first_name, _sheet=sheet)

                            patronymic = author.patronymic
                            set_value(_col="F", _row=_index, _value=patronymic, _sheet=sheet)

                            personnel_number = author.personnel_number
                            set_value(_col="G", _row=_index, _value=personnel_number,
                                      _sheet=sheet)

                            position = author.position
                            set_value(_col="H", _row=_index, _value=position, _sheet=sheet)

                            category = author.category
                            set_value(_col="I", _row=_index, _value=category, _sheet=sheet)

                            username = user.username
                            set_value(_col="J", _row=_index, _value=username, _sheet=sheet)

                            password = author.password
                            set_value(_col="K", _row=_index, _value=password, _sheet=sheet)

                            is_active = author.is_active_account
                            set_value(_col="L", _row=_index, _value=is_active, _sheet=sheet)

                            is_staff = user.is_staff
                            set_value(_col="M", _row=_index, _value=is_staff, _sheet=sheet)

                            is_superuser = user.is_superuser
                            set_value(_col="N", _row=_index, _value=is_superuser, _sheet=sheet)

                            is_temp_password = author.is_temp_password
                            set_value(_col="O", _row=_index, _value=is_temp_password, _sheet=sheet)

                            group_models = backend_models.GroupModel.objects.filter(users=author)
                            groups = ""
                            for group_model in group_models:
                                groups += f"{str(group_model.name).lower().strip()}, "
                            groups = groups[:-2]
                            set_value(_col="P", _row=_index, _value=groups, _sheet=sheet)

                            email = author.email
                            set_value(_col="Q", _row=_index, _value=email, _sheet=sheet)

                            secret_question = author.secret_question
                            set_value(_col="R", _row=_index, _value=secret_question,
                                      _sheet=sheet)

                            secret_answer = author.secret_answer
                            set_value(_col="S", _row=_index, _value=secret_answer, _sheet=sheet)
                        except Exception as error:
                            backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    # Set font
                    #######################################################
                    font_b = Font(name='Arial', size=8, bold=False)
                    for row in range(1, _index + 1):
                        for col in [get_column_letter(x) for x in range(1, 30 + 1)]:
                            cell = sheet[f'{col}{row}']
                            cell.font = font_b
                    #######################################################

                    # Height and width styles
                    #######################################################
                    for col in [get_column_letter(x) for x in range(1, 30 + 1)]:
                        width = 1
                        for row in range(1, _index + 1):
                            cell = sheet[f'{col}{row}']
                            value = len(str(cell.value))
                            if value > width:
                                width = value
                        sheet.column_dimensions[col].height = 1
                        sheet.column_dimensions[col].width = round((width * 0.95), 3)
                    #######################################################

                    try:
                        backend_utils.ExcelClass.workbook_save(
                            workbook=workbook,
                            excel_file=f"static/{path}/{file_name}"
                        )
                    except Exception as error:
                        pass
                    response = {"response": {"excel": f"static/{path}/{file_name}"}}

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)

                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        else:
            return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([IsAuthenticated])
def api_admin_create_users(request):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "POST":

            # TODO action ##############################################################################################

            if req_inst.action_type == "adminCreateUsersStore":
                try:

                    # TODO get_value ###################################################################################

                    change_user = req_inst.get_value(key="changeUser", default="")
                    change_user_password = req_inst.get_value(key="changeUserPassword", default="")
                    clear_user_groups = req_inst.get_value(key="clearUserGroups", default="")
                    additional_excel = req_inst.get_value(key="additionalExcel", default=None)

                    # TODO action ######################################################################################

                    if additional_excel:
                        def get_value(_col: Union[str, int], _row: Union[str, int], _sheet):
                            if isinstance(_col, int):
                                _col = get_column_letter(_col)
                            if isinstance(_row, str):
                                _row = str(_row)
                            _value = str(_sheet[str(_col).upper() + str(_row)].value).strip()
                            if _value.lower() == "none":
                                return ""
                            elif _value.lower() == "true":
                                return True
                            elif _value.lower() == "false":
                                return False
                            else:
                                return _value

                        workbook = openpyxl.load_workbook(additional_excel)
                        sheet = workbook.active
                        max_rows = sheet.max_row

                        for row in range(1 + 1, max_rows + 1):
                            try:
                                subdivision = get_value(_col="A", _row=row, _sheet=sheet)
                                workshop_service = get_value(_col="B", _row=row, _sheet=sheet)
                                department_site = get_value(_col="C", _row=row, _sheet=sheet)
                                last_name = get_value(_col="D", _row=row, _sheet=sheet)
                                first_name = get_value(_col="E", _row=row, _sheet=sheet)
                                patronymic = get_value(_col="F", _row=row, _sheet=sheet)
                                personnel_number = get_value(_col="G", _row=row, _sheet=sheet)
                                position = get_value(_col="H", _row=row, _sheet=sheet)
                                category = get_value(_col="I", _row=row, _sheet=sheet)
                                username = get_value(_col="J", _row=row, _sheet=sheet)
                                password = get_value(_col="K", _row=row, _sheet=sheet)
                                is_active = get_value(_col="L", _row=row, _sheet=sheet)
                                is_staff = get_value(_col="M", _row=row, _sheet=sheet)
                                is_superuser = get_value(_col="N", _row=row, _sheet=sheet)
                                is_temp_password = get_value(_col="O", _row=row, _sheet=sheet)
                                groups = get_value(_col="P", _row=row, _sheet=sheet).lower()
                                email = get_value(_col="Q", _row=row, _sheet=sheet)
                                secret_question = get_value(_col="R", _row=row, _sheet=sheet)
                                secret_answer = get_value(_col="S", _row=row, _sheet=sheet)

                                if len(username) <= 1:
                                    continue

                                try:
                                    user = User.objects.get(username=username)
                                    if user.is_superuser or change_user == "Не изменять уже существующего пользователя":
                                        continue
                                    new_user = False
                                except Exception as error:
                                    user = User.objects.create(
                                        username=username,
                                        password=make_password(password=password),
                                    )
                                    new_user = True

                                try:
                                    author = backend_models.UserModel.objects.get(
                                        user=user
                                    )
                                except Exception as error:
                                    author = backend_models.UserModel.objects.create(
                                        user=user
                                    )

                                if new_user:
                                    author.password = password
                                else:
                                    if change_user_password == "Изменять пароль уже существующего пользователя":
                                        user.password = make_password(password=password)
                                        author.password = password

                                user.is_staff = is_staff
                                user.is_superuser = is_superuser
                                user.email = email
                                user.last_name = last_name
                                user.first_name = first_name
                                user.save()

                                author.is_active_account = is_active
                                author.email = email
                                author.secret_question = secret_question
                                author.secret_answer = secret_answer
                                author.is_temp_password = is_temp_password
                                author.last_name = last_name
                                author.first_name = first_name
                                author.patronymic = patronymic
                                author.personnel_number = personnel_number
                                author.subdivision = subdivision
                                author.workshop_service = workshop_service
                                author.department_site = department_site
                                author.position = position
                                author.category = category
                                author.save()

                                if clear_user_groups == "Добавлять новые группы доступа к предыдущим":
                                    for group in backend_models.GroupModel.objects.filter(
                                            users=author
                                    ):
                                        try:
                                            group.users.remove(author)
                                        except Exception as error:
                                            backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                                groups = [group.strip() for group in str(groups).lower().strip().split(',')]
                                for group in groups:
                                    if len(group) > 1:
                                        try:
                                            group_model = backend_models.GroupModel.objects.get_or_create(
                                                name=group
                                            )[0]
                                            group_model.users.add(author)
                                        except Exception as error:
                                            backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                                if backend_utils.DjangoClass.DefaultSettingsClass.get_actions_print_value():
                                    print(row, username)
                            except Exception as error:
                                backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                        response = {"response": "Пользователи успешно созданы/изменены."}
                    else:
                        response = {"error": "Ошибка чтения файла!"}

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)

                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        else:
            return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([IsAuthenticated])
def api_admin_terminal_reboot(request):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "POST":

            # TODO action ##############################################################################################

            if req_inst.action_type == "":
                try:

                    # TODO get_value ###################################################################################

                    ips = req_inst.get_value(key="ips", default="")

                    # TODO action ######################################################################################

                    def reboot(_ip):
                        url = "htt" + f"p://{ip}/ISAPI/System/reboot"
                        h = httplib2.Http(
                            os.path.dirname(os.path.abspath('__file__')) + "/static/media/data/temp/reboot_terminal"
                        )
                        login_ = 'admin'
                        password_ = 'snrg2017'
                        h.add_credentials(login_, password_)
                        headers = {
                            'Content-type': 'text/plain;charset=UTF-8',
                            'Accept-Encoding': 'gzip, deflate',
                            'Accept-Language': 'de,en-US;q=0.7,en;q=0.3',
                            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:75.0) Gecko/20100101 '
                                          'Firefox/75.0',
                        }
                        response_, content = h.request(uri=url, method="PUT", headers=headers)
                        return [
                            _ip,
                            content.decode().split("<moderate_statusString>")[1].split("</moderate_statusString>")[0]
                        ]

                    with ThreadPoolExecutor() as executor:
                        futures = []
                        for ip in [str(str(x).strip()) for x in ips.split(",")]:
                            if len(str(ip)) < 3:
                                continue
                            futures.append(executor.submit(reboot, ip))
                        responses = []
                        for future in as_completed(futures):
                            responses.append(future.result())

                    response = {"response": responses}

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)

                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        else:
            return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


@api_view(http_method_names=HTTP_METHOD_NAMES)
# @permission_classes([IsAuthenticated])
def api_admin_recover_password(request):
    """
    django-rest-framework
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "GET":

            # TODO action ##############################################################################################

            if req_inst.action_type == "":
                try:

                    # TODO get_value ###################################################################################

                    username = req_inst.get_value(key="username", default="")

                    # TODO action ######################################################################################

                    user = User.objects.get(username=username)

                    response = {"response": backend_serializers.UserSerializer(user, many=False).data}

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)

                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        if req_inst.method == "POST":

            # TODO action ##############################################################################################

            if req_inst.action_type == "":
                try:

                    # TODO get_value ###################################################################################

                    username = req_inst.get_value(key="username", default="")
                    password = req_inst.get_value(key="password", default="")
                    password2 = req_inst.get_value(key="password2", default="")

                    # TODO action ######################################################################################

                    user = User.objects.get(username=username)
                    author = backend_models.UserModel.objects.get(user=user)
                    if len(password) < 1:
                        response = {"error": "Пароль пустой!"}
                    elif password != password2:
                        response = {"error": "Пароли не совпадают!"}
                    elif password == author.password:
                        response = {"error": "Пароль такой же как и предыдущий!"}
                    else:
                        user.set_password(password)
                        user.save()
                        author.password = password
                        author.is_temp_password = True
                        author.secret_question = ""
                        author.secret_answer = ""
                        author.email = None
                        author.save()
                        response = {"response": "Изменение успешно проведено."}

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)

                    return Response(response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return Response({"error": req_inst.return_action_type_error(error)})
            return Response({"error": req_inst.return_not_allowed_action_type()})
        else:
            return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)


# TODO develop #########################################################################################################

@permission_classes([AllowAny])
def test(request):
    """
    index React Single Page Applications
    """

    try:

        # TODO request #################################################################################################

        req_inst = backend_utils.DjangoClass.TemplateClass.request(request=request)

        # TODO method ##################################################################################################

        if req_inst.method == "GET":

            # TODO action ##############################################################################################

            if req_inst.action_type == "":
                try:
                    response = {"response": "Данные успешно получены!"}

                    # TODO response ####################################################################################

                    backend_utils.DjangoClass.TemplateClass.response(request=request, response=response)
                    return render(request, 'test/index.html', response)
                except Exception as error:
                    backend_utils.DjangoClass.LoggingClass.error(request=request, error=error)
                    return render(request, "backend/404.html")
            else:
                return Response({"error": req_inst.return_not_allowed_action_type()})
        else:
            return Response({"error": req_inst.return_not_allowed_method()})
    except Exception as error:
        return backend_utils.DjangoClass.DRFClass.RequestClass.return_global_error(request=request, error=error)
