import openpyxl
import os

relative_path =  os.path.dirname(os.path.abspath('__file__'))

export_file = relative_path+r'\Export.xlsx'
import_file = relative_path+r'\Import.xlsx'

export_file = relative_path+'\\'+str(input(f'Введите название файла для экспорта: '))
import_file = relative_path+'\\'+str(input(f'Введите название файла для импорта: '))

def input_integer_value(description):
    while True:
        try:
            value = round(int(input(f'{description}')))
            if value > 0:
                break
        except:
            print('Ошибка, введите ещё раз.')
    return value

def get_sheet_value(column, row):
    return str(sheet[str(column)+str(row)].value)

def set_sheet_value(column, row, value):
    if value:
        sheet[str(column)+str(row)] = value
    else:
        sheet[str(column)+str(row)] = ''

# min_export_value = input_integer_value('Введите номер строки, с которой начинается экспорт: ')
# max_export_value = input_integer_value('Введите номер строки, на которой заканчивается экспорт: ')
# min_import_value = input_integer_value('Введите номер строки, с которой начинается импорт: ')
# max_import_value = input_integer_value('Введите номер строки, на которой заканчивается импорт: ')
min_export_value = 1
max_export_value = 5000
min_import_value = 14
max_import_value = 5000

# print(get_sheet_value('A', 2))
# print(openpyxl.utils.get_column_letter(7))
# print(openpyxl.utils.column_index_from_string('W'))
# print(f'подразделение :{Workers[6].A_1}')
# print(f'участок :{Workers[6].B_1}')

Workers_from_1C = []


class Worker():
    """
    Класс, который содержит в себе работника, со значениями по строке
    """
    def __init__(self, A_1 = '', B_1 = '', C_1 = '', D_1 = '', E_1 = '', F_1 = '', G_1 = '', H_1 = '', M_1 = ''):
        # Подразделение
        self.A_1 = A_1
        # Цех или Служба
        self.B_1 = B_1
        # Отдел или участок
        self.C_1 = C_1
        # Фамилия
        self.D_1 = D_1
        # Имя
        self.E_1 = E_1
        # Отчество
        self.F_1 = F_1
        # Табельный №
        self.G_1 = G_1
        # Категория
        self.H_1 = H_1
        # Пол
        self.M_1 = M_1

    def print_worker(self):
        print(f'{self.A_1}+{self.B_1}+{self.C_1}+{self.D_1}+{self.E_1}+{self.F_1}+{self.G_1}+{self.H_1}+{self.M_1}')

    def get_worker_value(self, index):
        value = list((self.A_1, self.B_1, self.C_1, self.D_1, self.E_1, self.F_1, self.G_1, self.H_1, self.M_1))
        return value[index]

    def get_worker_id(self):
        return self.G_1

workbook = openpyxl.load_workbook(export_file)
sheet = workbook.active
workbook.close()

for num in range(min_export_value, max_export_value):
    Worker_list = []
    for x in 'ABCDEFGHM':
        value = get_sheet_value(x, num)
        if value == 'None' or value == None or value == '':
            value = ''
        Worker_list.append(value)
    Worker_id = Worker(*Worker_list)
    Workers_from_1C.append(Worker_id)

workbook = openpyxl.load_workbook(import_file)
sheet = workbook.active

Workers_from_DB = []

for num in range(min_import_value, max_import_value):
    Workers_from_DB.append(get_sheet_value('C', num))

for worker_from_1C in Workers_from_1C:
    for worker_from_DB in Workers_from_DB:
        if worker_from_1C.get_worker_id() == worker_from_DB: 
            for x in 'RSTBAUCVF':
                try:
                    set_sheet_value(x, Workers_from_DB.index(worker_from_DB)+14, worker_from_1C.get_worker_value('RSTBAUCVF'.index(x)))
                except:
                    pass

workbook.save(import_file)
workbook.close()
