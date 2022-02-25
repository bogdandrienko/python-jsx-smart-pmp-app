import openpyxl
from openpyxl.utils import get_column_letter

# wb = openpyxl.load_workbook('pzdht.xlsx')
# ws = wb.active
# max_row = ws.max_row
#
# id_inx = []
# global_list = []
# for row in range(1, 610):
#     local_list = []
#     if ws[f'A{row}'].value not in id_inx:
#         id_inx.append(ws[f'A{row}'].value)
#         for char in 'ABCDEFGH':
#             value = ws[f'{char}{row}'].value
#             if char == 'E':
#                 try:
#                     value = value.encode('1251').decode('utf-8')
#                 except Exception as error:
#                     try:
#                         value = str(value).split(" ")
#                         try:
#                             name = value[0].encode('1251').decode('utf-8')
#                         except Exception as error:
#                             name = "И" + \
#                                    value[0][2:].encode('1251').decode('utf-8')
#                         try:
#                             surname = value[1].encode(
#                                 '1251').decode('utf-8')
#                         except Exception as error:
#                             surname = "И" + \
#                                       value[1][2:].encode('1251').decode('utf-8')
#                         string = name + " " + surname
#                         value = string
#                     except Exception as error:
#                         pass
#             local_list.append(value)
#     global_list.append(local_list)
#
# print(len(id_inx))
#
# wb = openpyxl.Workbook()
# ws = wb.active
# row_index = 1
# for row in global_list:
#     col_index = 1
#     for col in row:
#         try:
#             ws[f"{get_column_letter(col_index)}{row_index}"] = col
#         except Exception as error:
#             print(error)
#         col_index += 1
#     row_index += 1
# wb.save("pzdht_1.xlsx")

