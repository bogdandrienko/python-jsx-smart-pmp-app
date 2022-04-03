# TODO download modules ################################################################################################
import base64
import datetime
import hashlib
import json
import os
import random
import threading
import time
import httplib2
import openpyxl
from django.contrib.auth.hashers import make_password
from openpyxl.utils import get_column_letter
from typing import Union
# TODO django modules ##################################################################################################
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.models import User
from django.core.handlers.wsgi import WSGIRequest
# TODO custom modules ##################################################################################################
from backend import models as backend_models


# TODO base service ####################################################################################################
class DjangoClass:
    class LoggingClass:
        @staticmethod
        def error(request, error):
            if DjangoClass.DefaultSettingsClass.get_actions_print_value():
                req_inst = DjangoClass.DRFClass.RequestClass(request=request)
                print(f"\n\nusername: {req_inst.user}"
                      f"\nip: {req_inst.ip}"
                      f"\nrequest_path: {req_inst.path}"
                      f"\nrequest_method: {req_inst.method}"
                      f"\nrequest_action_type: {req_inst.action_type}"
                      f"\nerror: {error}")

            if DjangoClass.DefaultSettingsClass.get_error_logging_value():
                req_inst = DjangoClass.DRFClass.RequestClass(request=request)
                backend_models.LoggingModel.objects.create(
                    username_slug_field=req_inst.user,
                    ip_genericipaddress_field=req_inst.ip,
                    request_path_slug_field=req_inst.path,
                    request_method_slug_field=f"ERROR | {req_inst.method} | {req_inst.action_type}",
                    error_text_field=f'error: {error}'
                )
                text = [req_inst.user, req_inst.ip, req_inst.path, req_inst.method, req_inst.action_type, error,
                        datetime.datetime.now()]
                string = ''
                for val in text:
                    string = string + f', {val}'
                with open('static/media/admin/logging/logging_errors.txt', 'a') as log:
                    log.write(f'\n{string[2:]}\n')

        @staticmethod
        def error_local(error, function_error):
            if DjangoClass.DefaultSettingsClass.get_actions_print_value():
                print(f"\n\nrequest_path: error_local"
                      f"\nrequest_method: {function_error}"
                      f"\nerror: {error}")

            if DjangoClass.DefaultSettingsClass.get_actions_logging_value():
                backend_models.LoggingModel.objects.create(
                    request_path_slug_field="error_local",
                    request_method_slug_field=function_error,
                    error_text_field=f'error: {error}'
                )
                text = ["error_local", function_error, error, datetime.datetime.now()]
                string = ''
                for val in text:
                    string = string + f', {val}'
                with open('static/media/admin/logging/logging_errors.txt', 'a') as log:
                    log.write(f'\n{string[2:]}\n')

        @staticmethod
        def action(request):
            if DjangoClass.DefaultSettingsClass.get_actions_print_value():
                request_instance = DjangoClass.DRFClass.RequestClass(request=request)
                print(f"\n\n{DateTimeUtils.get_current_datetime()}"
                      f"\nrequest_path: {request_instance.path}"
                      f"\nrequest_action_type: {request_instance.action_type}"
                      f"\nrequest_user: {request_instance.user}"
                      f"\nrequest.data: {request_instance.data}")

            if DjangoClass.DefaultSettingsClass.get_actions_logging_value():
                req_inst = DjangoClass.DRFClass.RequestClass(request=request)
                backend_models.LoggingModel.objects.create(
                    username_slug_field=req_inst.user,
                    ip_genericipaddress_field=req_inst.ip,
                    request_path_slug_field=req_inst.path,
                    request_method_slug_field=f"ACTION | {req_inst.method} | {req_inst.action_type}",
                    error_text_field=f'-'
                )
                text = [req_inst.user, req_inst.ip, req_inst.path, req_inst.method, req_inst.action_type,
                        datetime.datetime.now()]
                string = ''
                for val in text:
                    string = string + f', {val}'
                with open('static/media/admin/logging/logging_actions.txt', 'a') as log:
                    log.write(f'\n{string[2:]}\n')

        @staticmethod
        def response(request, response):
            if DjangoClass.DefaultSettingsClass.get_response_print_value():
                request_instance = DjangoClass.DRFClass.RequestClass(request=request)
                print(f"\n\n{DateTimeUtils.get_current_datetime()}"
                      f"\nrequest_path: {request_instance.path}"
                      f"\nrequest_action_type: {request_instance.action_type}"
                      f"\nrequest_user: {request_instance.user}"
                      f"\nrequest.data: {request_instance.data}"
                      f"\nrequest.response: {response}")

            if DjangoClass.DefaultSettingsClass.get_response_logging_value():
                req_inst = DjangoClass.DRFClass.RequestClass(request=request)
                backend_models.LoggingModel.objects.create(
                    username_slug_field=req_inst.user,
                    ip_genericipaddress_field=req_inst.ip,
                    request_path_slug_field=req_inst.path,
                    request_method_slug_field=f"RESPONSE | {req_inst.method} | {req_inst.action_type}",
                    error_text_field=f'response {response}'
                )
                text = [req_inst.user, req_inst.ip, req_inst.path, req_inst.method, req_inst.action_type,
                        datetime.datetime.now()]
                string = ''
                for val in text:
                    string = string + f', {val}'
                with open('static/media/admin/logging/logging_responses.txt', 'a') as log:
                    log.write(f'\n{string[2:]}\n')

    class TemplateClass:
        @staticmethod
        def request(request):
            threading.Thread(target=DjangoClass.DefaultSettingsClass.check_settings, args=(request,)).start()
            threading.Thread(target=DjangoClass.LoggingClass.action, args=(request,)).start()
            request_instance = DjangoClass.DRFClass.RequestClass(request=request)

            return request_instance

        @staticmethod
        def response(request, response):
            threading.Thread(target=DjangoClass.LoggingClass.response, args=(request, response,)).start()

    class DRFClass:
        class RequestClass:
            def __init__(self, request):
                # print("request.scheme: ", request.scheme)
                # print("request.body: ", request.body)
                # print("request.path: ", request.path)
                # print("request.path_info: ", request.path_info)
                # print("request.method: ", request.method)
                # print("request.encoding: ", request.encoding)
                # print("request.content_type: ", request.content_type)
                # print("request.GET: ", request.GET)
                # print("request.POST: ", request.POST)
                # print("request.COOKIES: ", request.COOKIES)
                # print("request.FILES: ", request.FILES)
                # print("request.META: ", request.META)
                # print("request.META: ", request.META)
                # for key, value in request.META.items():
                #     print(f"{key}: {value}")
                # print("request.META.HTTP_HOST : ", request.META.get("HTTP_HOST"))
                # print("request.META.REMOTE_ADDR: ", request.META.get("REMOTE_ADDR"))
                # print("request.META.HTTP_REFERER: ", request.META.get("HTTP_REFERER"))
                try:
                    self.request = request
                except Exception as error:
                    self.request = None
                try:
                    self.path = str(request.path)
                except Exception as error:
                    self.path = ""
                try:
                    self.ip = str(request.META.get("REMOTE_ADDR"))
                except Exception as error:
                    self.ip = ""
                try:
                    self.method = str(request.method).upper()
                except Exception as error:
                    self.method = "GET"
                try:
                    self.action_type = str(request.data["Action-type"]).upper()
                except Exception as error:
                    self.action_type = ""
                try:
                    self.user = User.objects.get(username=str(request.user.username))
                except Exception as error:
                    self.user = None
                try:
                    self.user_model = \
                        backend_models.UserModel.objects.get(user_foreign_key_field=self.user)
                except Exception as error:
                    self.user_model = None
                try:
                    self.data = request.data
                except Exception as error:
                    self.data = None
                try:
                    self.body = request.data["body"]
                except Exception as error:
                    self.body = None

            def get_value(self, key: str, except_error=False, strip=False):
                try:
                    if self.data[key] == "null":
                        return None
                    elif self.data[key] == "true":
                        return True
                    elif self.data[key] == "false":
                        return False
                    else:
                        if strip:
                            return str(self.data[key]).strip()
                        else:
                            return self.data[key]
                except Exception as error:
                    pass

        class OldRequestClass:
            @staticmethod
            def get_value(request: WSGIRequest, key: str, none_is_error=False, strip=True):
                if none_is_error:
                    # If key not have is Exception Error
                    value = request.POST[key]
                else:
                    # If key not have value is None
                    value = request.POST.get(key)
                if strip and value:
                    value.strip()
                return value

            @staticmethod
            def get_check(request: WSGIRequest, key: str, none_is_error=False):
                if none_is_error:
                    # If key not have is Exception Error
                    value = request.POST[key]
                else:
                    # If key not have value is None
                    value = request.POST.get(key)
                if value is None:
                    return False
                else:
                    return True

            @staticmethod
            def get_file(request: WSGIRequest, key: str, none_is_error=False):
                if none_is_error:
                    # If key not have is Exception Error
                    file = request.FILES[key]
                else:
                    # If key not have value is None
                    file = request.FILES.get(key)
                return file

            @staticmethod
            def request_utils(request):
                try:
                    request_method = str(request.method).upper()
                except Exception as error:
                    request_method = None

                try:
                    request_action_type = str(request.data["Action-type"]).upper()
                except Exception as error:
                    request_action_type = None

                try:
                    request_user = User.objects.get(username=str(request.user.username))
                except Exception as error:
                    request_user = None

                try:
                    request_body = request.data["body"]
                except Exception as error:
                    request_body = None

                return [request_method, request_action_type, request_user, request_body]

    class DefaultSettingsClass:

        @staticmethod
        def check_settings(request):
            try:
                obj = backend_models.SettingsModel.objects.filter(type_char_field="logging_action")
                if obj.count() < 1:
                    backend_models.SettingsModel.objects.create(
                        type_char_field="logging_action",
                        char_field="Логирование действий: вкл/выкл(boolean_field)",
                        boolean_field=True
                    )
            except Exception as error:
                print(error)

            try:
                objects = backend_models.SettingsModel.objects.filter(type_char_field="print_action")
                if objects.count() < 1:
                    backend_models.SettingsModel.objects.create(
                        type_char_field="print_action",
                        char_field="Вывод в консоль действий, вкл/выкл(boolean_field)",
                        boolean_field=False
                    )
            except Exception as error:
                print(error)

            try:
                objects = backend_models.SettingsModel.objects.filter(type_char_field="logging_error")
                if objects.count() < 1:
                    backend_models.SettingsModel.objects.create(
                        type_char_field="logging_error",
                        char_field="Логирование ошибок, вкл/выкл(boolean_field)",
                        boolean_field=True
                    )
            except Exception as error:
                print(error)

            try:
                objects = backend_models.SettingsModel.objects.filter(type_char_field="print_error")
                if objects.count() < 1:
                    backend_models.SettingsModel.objects.create(
                        type_char_field="print_error",
                        char_field="Вывод в консоль ошибок, вкл/выкл(boolean_field)",
                        boolean_field=False
                    )
            except Exception as error:
                print(error)

            try:
                objects = backend_models.SettingsModel.objects.filter(type_char_field="logging_response")
                if objects.count() < 1:
                    backend_models.SettingsModel.objects.create(
                        type_char_field="logging_response",
                        char_field="Логирование ответов, вкл/выкл(boolean_field)",
                        boolean_field=False
                    )
            except Exception as error:
                print(error)

            try:
                objects = backend_models.SettingsModel.objects.filter(type_char_field="print_response")
                if objects.count() < 1:
                    backend_models.SettingsModel.objects.create(
                        type_char_field="print_response",
                        char_field="Вывод в консоль ответов, вкл/выкл(boolean_field)",
                        boolean_field=False
                    )
            except Exception as error:
                print(error)

            try:
                objects = backend_models.SettingsModel.objects.filter(type_char_field="scheduler_personal")
                if objects.count() < 1:
                    backend_models.SettingsModel.objects.create(
                        type_char_field="scheduler_personal",
                        char_field="Планировщик обновления персонала из 1С",
                        text_field="http://192.168.1.10/KM_1C/hs/iden/change/ | Web_adm_1c | 159159qqww!",
                        integer_field=60,
                        boolean_field=True
                    )
                else:
                    obj = objects[0]
                    if obj.boolean_field:
                        try:
                            objects = backend_models.LoggingModel.objects.filter(
                                request_path_slug_field="/scheduler_personal/"
                            )
                            update = True
                            if objects.count() > 0:
                                if (objects[0].created_datetime_field +
                                    datetime.timedelta(hours=6, minutes=obj.integer_field)).strftime(
                                    '%Y-%m-%d %H:%M') >= \
                                        (datetime.datetime.now()).strftime('%Y-%m-%d %H:%M'):
                                    update = False
                            if update:
                                req_inst = DjangoClass.DRFClass.RequestClass(request=request)
                                backend_models.LoggingModel.objects.create(
                                    username_slug_field=req_inst.user,
                                    ip_genericipaddress_field=req_inst.ip,
                                    request_path_slug_field="/scheduler_personal/",
                                    request_method_slug_field=req_inst.method + " | SCHEDULER",
                                    error_text_field=f'-'
                                )
                                threading.Thread(
                                    target=DjangoClass.DefaultSettingsClass.SchedulerClass.scheduler_personal, args=()
                                ).start()
                        except Exception as error:
                            DjangoClass.LoggingClass.error(request=request, error=error)
            except Exception as error:
                print(error)

            try:
                objects = backend_models.SettingsModel.objects.filter(type_char_field="scheduler_superuser")
                if objects.count() < 1:
                    backend_models.SettingsModel.objects.create(
                        type_char_field="scheduler_superuser",
                        char_field="Планировщик создания стандартных суперпользователей",
                        text_field="000000000000, 31284bogdan | Web_adm_1c, 159159qqww!",
                        boolean_field=True
                    )
                else:
                    obj = objects[0]
                    if obj.boolean_field:
                        users_strings = str(obj.text_field).strip().split("|")
                        superusers = []
                        for user_strings in users_strings:
                            superusers.append(
                                [str(x).strip() for x in str(user_strings).strip().split(",") if len(x) > 1])
                        threading.Thread(
                            target=DjangoClass.DefaultSettingsClass.SchedulerClass.scheduler_default_superusers,
                            args=(superusers,)
                        ).start()
                        obj.boolean_field = False
                        obj.save()
            except Exception as error:
                print(error)

            try:
                objects = backend_models.SettingsModel.objects.filter(type_char_field="scheduler_group")
                if objects.count() < 1:
                    backend_models.SettingsModel.objects.create(
                        type_char_field="scheduler_group",
                        char_field="Планировщик создания стандартных групп",
                        text_field="user, moderator, superuser, " +

                                   "moderator_idea, moderator_vacancy, " +

                                   "moderator_rational, moderator_rational_atp, moderator_rational_gtk, "
                                   "moderator_rational_ok, moderator_rational_upravlenie, "
                                   "moderator_rational_energoupravlenie, " +
                                   "",
                        boolean_field=True
                    )
                else:
                    obj = objects[0]
                    if obj.boolean_field:
                        groups = [str(x).strip() for x in str(obj.text_field).strip().split(",") if len(x) > 1]
                        threading.Thread(
                            target=DjangoClass.DefaultSettingsClass.SchedulerClass.scheduler_default_groups,
                            args=(groups,)
                        ).start()
                        obj.boolean_field = False
                        obj.save()
            except Exception as error:
                print(error)

        @staticmethod
        def get_actions_logging_value():
            try:
                obj = backend_models.SettingsModel.objects.filter(type_char_field="logging_action")[0]
                return obj.boolean_field
            except Exception as error:
                return False

        @staticmethod
        def get_actions_print_value():
            try:
                obj = backend_models.SettingsModel.objects.filter(type_char_field="print_action")[0]
                return obj.boolean_field
            except Exception as error:
                return False

        @staticmethod
        def get_error_logging_value():
            try:
                obj = backend_models.SettingsModel.objects.filter(type_char_field="logging_error")[0]
                return obj.boolean_field
            except Exception as error:
                return False

        @staticmethod
        def get_error_print_value():
            try:
                obj = backend_models.SettingsModel.objects.filter(type_char_field="print_error")[0]
                return obj.boolean_field
            except Exception as error:
                return False

        @staticmethod
        def get_response_logging_value():
            try:
                obj = backend_models.SettingsModel.objects.filter(type_char_field="logging_response")[0]
                return obj.boolean_field
            except Exception as error:
                return False

        @staticmethod
        def get_response_print_value():
            try:
                obj = backend_models.SettingsModel.objects.filter(type_char_field="print_response")[0]
                return obj.boolean_field
            except Exception as error:
                return False

        @staticmethod
        def get_salary_value():
            try:
                obj = backend_models.SettingsModel.objects.filter(type_char_field="scheduler_personal")[0]
                objects = []
                for x in str(obj.text_field).strip().split("|"):
                    objects.append(str(x).strip())
                return objects
            except Exception as error:
                return False

        class SchedulerClass:
            @staticmethod
            def scheduler_personal():
                try:
                    key = UtilsClass.create_encrypted_password(
                        _random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890',
                        _length=10
                    )
                    hash_key_obj = hashlib.sha256()
                    hash_key_obj.update(key.encode('utf-8'))
                    key_hash = str(hash_key_obj.hexdigest().strip().upper())
                    key_hash_base64 = base64.b64encode(str(key_hash).encode()).decode()
                    date = datetime.datetime.now().strftime("%Y%m%d")
                    date_base64 = base64.b64encode(str(date).encode()).decode()
                    get_salary_value = DjangoClass.DefaultSettingsClass.get_salary_value()
                    url = f'{get_salary_value[0]}{date_base64}_{key_hash_base64}'
                    h = httplib2.Http(
                        os.path.dirname(os.path.abspath('__file__')) + "/static/media/data/temp/get_users")
                    _login = f'{get_salary_value[1]}'
                    password = f'{get_salary_value[2]}'
                    h.add_credentials(_login, password)
                    response_, content = h.request(url)
                    json_data = json.loads(
                        UtilsClass.decrypt_text_with_hash(content.decode()[1:], key_hash)
                    )

                    class Worker:
                        def __init__(self, date_time_: str, status_: str, username_: str, last_name_: str,
                                     first_name_: str,
                                     patronymic_: str, personnel_number_: str, subdivision_: str,
                                     workshop_service_: str,
                                     department_site_: str, position_: str, category_: str):
                            self.date_time_ = date_time_
                            self.status_ = status_
                            self.username_ = username_
                            self.last_name_ = last_name_
                            self.first_name_ = first_name_
                            self.patronymic_ = patronymic_
                            self.personnel_number_ = personnel_number_
                            self.subdivision_ = subdivision_
                            self.workshop_service_ = workshop_service_
                            self.department_site_ = department_site_
                            self.position_ = position_
                            self.category_ = category_

                        @staticmethod
                        def get_value(dict_: dict, user_, key_: str):
                            try:
                                value = dict_["global_objects"][user_][key_]
                                return value
                            except Exception as error__:
                                DjangoClass.LoggingClass.error_local(
                                    error=error__,
                                    function_error="DefaultSettingsClass.SchedulerClass.scheduler_personal."
                                                   "Worker.get_value"
                                )
                                return ''
                    index = 0
                    for user in json_data["global_objects"]:
                        worker = Worker(
                            date_time_=Worker.get_value(dict_=json_data, user_=user, key_="Период"),
                            status_=Worker.get_value(dict_=json_data, user_=user, key_="Статус"),
                            username_=Worker.get_value(dict_=json_data, user_=user, key_="ИИН"),
                            last_name_=Worker.get_value(dict_=json_data, user_=user, key_="Фамилия"),
                            first_name_=Worker.get_value(dict_=json_data, user_=user, key_="Имя"),
                            patronymic_=Worker.get_value(dict_=json_data, user_=user, key_="Отчество"),
                            personnel_number_=Worker.get_value(dict_=json_data, user_=user, key_="ТабельныйНомер"),
                            subdivision_=Worker.get_value(dict_=json_data, user_=user, key_="Подразделение"),
                            workshop_service_=Worker.get_value(dict_=json_data, user_=user, key_="Цех_Служба"),
                            department_site_=Worker.get_value(dict_=json_data, user_=user, key_="Отдел_Участок"),
                            position_=Worker.get_value(dict_=json_data, user_=user, key_="Должность"),
                            category_=Worker.get_value(dict_=json_data, user_=user, key_="Категория")
                        )

                        if len(worker.username_) <= 1:
                            continue

                        password = ""
                        try:
                            user = User.objects.get(username=worker.username_)
                            if user.is_superuser:
                                continue
                            new_user = False
                        except Exception as error_:
                            for i in range(1, 6 + 1):
                                password += random.choice(
                                    "abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890"
                                )
                            password = "temp_" + password
                            user = User.objects.create(
                                username=worker.username_,
                                password=make_password(password=password),
                            )
                            new_user = True

                        try:
                            user_model = backend_models.UserModel.objects.get(
                                user_foreign_key_field=user
                            )
                        except Exception as error:
                            user_model = backend_models.UserModel.objects.create(
                                user_foreign_key_field=user
                            )

                        if new_user:
                            user_model.password_char_field = password

                        if user.last_name != worker.last_name_:
                            user.last_name = worker.last_name_
                        if user.first_name != worker.first_name_:
                            user.first_name = worker.first_name_
                        user.save()

                        if worker.status_ == 'created':
                            user_model.activity_boolean_field = True
                        elif worker.status_ == 'changed':
                            user_model.activity_boolean_field = True
                        elif worker.status_ == 'disabled':
                            user_model.activity_boolean_field = False
                        else:
                            user_model.activity_boolean_field = False
                        if user_model.last_name_char_field != worker.last_name_:
                            user_model.last_name_char_field = worker.last_name_
                        if user_model.first_name_char_field != worker.first_name_:
                            user_model.first_name_char_field = worker.first_name_
                        if user_model.patronymic_char_field != worker.patronymic_:
                            user_model.patronymic_char_field = worker.patronymic_
                        if user_model.personnel_number_slug_field != worker.personnel_number_:
                            user_model.personnel_number_slug_field = worker.personnel_number_
                        if user_model.subdivision_char_field != worker.subdivision_:
                            user_model.subdivision_char_field = worker.subdivision_
                        if user_model.workshop_service_char_field != worker.workshop_service_:
                            user_model.workshop_service_char_field = worker.workshop_service_
                        if user_model.department_site_char_field != worker.department_site_:
                            user_model.department_site_char_field = worker.department_site_
                        if user_model.position_char_field != worker.position_:
                            user_model.position_char_field = worker.position_
                        if user_model.category_char_field != worker.category_:
                            user_model.category_char_field = worker.category_
                        user_model.save()
                        try:
                            group_model = backend_models.GroupModel.objects.get_or_create(name_slug_field="user")[0]
                            group_model.user_many_to_many_field.add(user_model)
                        except Exception as error_:
                            pass
                        index += 1
                        if DjangoClass.DefaultSettingsClass.get_actions_print_value():
                            print(worker.username_)
                            print(index)
                except Exception as error:
                    DjangoClass.LoggingClass.error_local(
                        error=error,
                        function_error="DefaultSettingsClass.SchedulerClass.scheduler_personal"
                    )

            @staticmethod
            def scheduler_default_superusers(superusers=None):
                if superusers is None:
                    superusers = [["000000000000", "31284bogdan"], ]
                for superuser in superusers:
                    try:
                        username_ = superuser[0]
                        password_ = superuser[1]
                        try:
                            user = User.objects.get(username=username_)
                        except Exception as error__:
                            user = User.objects.create(username=username_)
                        user.set_password(password_)
                        user.is_staff = True
                        user.is_superuser = True
                        user.last_name = "Andrienko"
                        user.first_name = "Bogdan"
                        user.save()
                        user_model = backend_models.UserModel.objects.get_or_create(user_foreign_key_field=user)[0]
                        user_model.password_char_field = password_
                        user_model.last_name_char_field = "Andrienko"
                        user_model.first_name_char_field = "Bogdan"
                        user_model.patronymic_char_field = ""
                        user_model.position_char_field = "Administrator"
                        user_model.save()
                    except Exception as error__:
                        DjangoClass.LoggingClass.error_local(
                            error=error__,
                            function_error="DefaultSettingsClass.SchedulerClass.scheduler_default_superusers"
                        )

            @staticmethod
            def scheduler_default_groups(groups=None):
                if groups is None:
                    groups = ["user", "moderator", "superuser"]
                for grp in groups:
                    try:
                        action_model = backend_models.ActionModel.objects.get_or_create(action_slug_field=grp)[0]
                        group_model = backend_models.GroupModel.objects.get_or_create(name_slug_field=grp)[0]
                        group_model.action_many_to_many_field.add(action_model)
                        group_model.save()
                    except Exception as error:
                        DjangoClass.LoggingClass.error_local(
                            error=error,
                            function_error="DefaultSettingsClass.SchedulerClass.scheduler_default_groups"
                        )

    class PaginationClass:
        @staticmethod
        def paginate(request, objects, num_page):
            paginator = Paginator(objects, num_page)
            pages = request.GET.get('page')
            try:
                page = paginator.page(pages)
            except PageNotAnInteger:
                page = paginator.page(1)
            except EmptyPage:
                page = paginator.page(paginator.num_pages)
            return page

    @staticmethod
    def check_access(user_model, slug=""):
        if backend_models.GroupModel.objects.filter(
                user_many_to_many_field=user_model,
                name_slug_field=slug
        ).count() > 0:
            return True
        else:
            return False


# TODO custom service ##################################################################################################
class DateTimeUtils:
    @staticmethod
    def get_current_datetime():
        return f"{time.strftime('%Y-%m-%d %H:%M:%S')}"

    @staticmethod
    def get_current_date():
        return f"{time.strftime('%Y-%m-%d')}"

    @staticmethod
    def get_current_time():
        return f"{time.strftime('%H:%M:%S')}"

    @staticmethod
    def get_difference_datetime(days=0.0, hours=0.0, minutes=0.0, seconds=0.0):
        value = datetime.datetime.now() + datetime.timedelta(days=days, hours=hours, minutes=minutes, seconds=seconds)
        return value.replace(tzinfo=datetime.timezone.utc)

    class Example:
        @staticmethod
        def example_get_datetime():
            print(DateTimeUtils.get_current_datetime())

        @staticmethod
        def example_get_difference_datetime():
            print(DateTimeUtils.get_difference_datetime(hours=-2))


class DirPathFolderPathClass:
    @staticmethod
    def create_folder_in_this_dir(folder_name='new_folder', current_path=os.path.dirname(os.path.abspath('__file__'))):
        full_path = current_path + f'/{folder_name}'
        try:
            os.makedirs(full_path)
        except Exception as error:
            DjangoClass.LoggingClass.error_local(
                error=error,
                function_error="DirPathFolderPathClass.create_folder_in_this_dir"
            )
            # print(f'directory already yet | {error}')
            pass
        finally:
            return full_path

    @staticmethod
    def get_all_files_in_path(path=os.path.dirname(os.path.abspath('__file__'))):
        files_list = []
        for root, dirs, files in os.walk(path, topdown=True):
            for name in files:
                files_list.append(f"{os.path.join(root, name)}")
        return files_list

    @staticmethod
    def get_all_dirs_in_path(path=os.path.dirname(os.path.abspath('__file__'))):
        directories_list = []
        for root, dirs, files in os.walk(path, topdown=True):
            for name in dirs:
                directories_list.append(f"{os.path.join(root, name)}")
        return directories_list

    class Example:
        @staticmethod
        def example_create_folder_in_this_folder():
            path = DirPathFolderPathClass.create_folder_in_this_dir(folder_name='new_folder')
            print(path)

        @staticmethod
        def example_create_folder_in_folder_in_this_folder():
            path = DirPathFolderPathClass.create_folder_in_this_dir(folder_name='new_folder\\new')
            print(path)

        @staticmethod
        def example_create_folder_in_external_folder():
            path = DirPathFolderPathClass.create_folder_in_this_dir(folder_name='new_folder\\new', current_path='C:\\')
            print(path)

        @staticmethod
        def example_get_all_files_in_path():
            files_list = DirPathFolderPathClass.get_all_files_in_path(
                path=r'C:\Project\Github_Projects\python-jsx-smart-pmp-app\web-platform_12_21_dev\app_admin'
            )
            for file in files_list:
                print(file)

        @staticmethod
        def example_get_all_folders_in_path():
            path_list = DirPathFolderPathClass.get_all_dirs_in_path(
                path=r'C:\Project\Github_Projects\python-jsx-smart-pmp-app\web-platform_12_21_dev\app_admin'
            )
            for path in path_list:
                print(path)


class EncryptingClass:
    @staticmethod
    def encrypt_text(text: str, hash_chars: str):
        chars = 'abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890' + ' _-:' + hash_chars
        chars = set(chars)
        forward_ord_list = []
        for char in chars:
            forward_ord_list.append(ord(char))
        forward_ord_list.sort(reverse=False)
        reverse_ord_list = forward_ord_list.copy()
        reverse_ord_list.sort(reverse=True)
        forward_dictinary = {reverse_ord_list[forward_ord_list.index(x)]: x for x in forward_ord_list}
        return ''.join([chr(forward_dictinary[ord(x)]) for x in text])

    @staticmethod
    def decrypt_text(text: str, hash_chars: str):
        chars = 'abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890' + ' _-:' + hash_chars
        chars = set(chars)
        forward_ord_list = []
        for char in chars:
            forward_ord_list.append(ord(char))
        forward_ord_list.sort(reverse=False)
        reverse_ord_list = forward_ord_list.copy()
        reverse_ord_list.sort(reverse=True)
        reverse_dictinary = {forward_ord_list[reverse_ord_list.index(x)]: x for x in reverse_ord_list}
        return ''.join([chr(reverse_dictinary[ord(x)]) for x in text])

    class Example:
        @staticmethod
        def example_encrypt_text():
            value = EncryptingClass.encrypt_text(text='12345', hash_chars='321')
            print(value)

        @staticmethod
        def example_decrypt_text():
            value = EncryptingClass.decrypt_text(text='wvuts', hash_chars='321')
            print(value)


class UtilsClass:
    @staticmethod
    def create_encrypted_password(_random_chars='abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890',
                                  _length=8):
        password = ''
        for i in range(1, _length + 1):
            password += random.choice(_random_chars)
        return password

    @staticmethod
    def decrypt_text_with_hash(massivsimvolov: str, massivkhesha: str):
        rasshifrovat_tekst = ''
        pozitsiyasimvolakhesha = 0
        dlinakhesha = len(massivkhesha)
        propusk = False
        for num in massivsimvolov:
            if propusk:
                propusk = False
                continue
            nomersimvola = ord(str(num))
            if pozitsiyasimvolakhesha >= dlinakhesha - 1:
                pozitsiyasimvolakhesha = 0
            pozitsiyasimvolakhesha = pozitsiyasimvolakhesha + 1
            simvolkhesha = ord(str(massivkhesha[pozitsiyasimvolakhesha]))
            kod_zashifrovannyy_simvol = nomersimvola - simvolkhesha
            # print(f"nomersimvola:{chr(nomersimvola)}:{nomersimvola}|simvolkhesha:{chr(simvolkhesha)}:{simvolkhesha}")
            zashifrovannyy_simvol = chr(kod_zashifrovannyy_simvol)
            rasshifrovat_tekst = rasshifrovat_tekst + zashifrovannyy_simvol
            if round(simvolkhesha / 2, 0) == simvolkhesha / 2:
                propusk = True
        return rasshifrovat_tekst


class ExcelClass:
    @staticmethod
    def workbook_create():
        workbook = openpyxl.Workbook()
        return workbook

    @staticmethod
    def workbook_load(excel_file: str):
        workbook = openpyxl.load_workbook(excel_file)
        return workbook

    @staticmethod
    def workbook_activate(workbook):
        sheet = workbook.active
        return sheet

    @staticmethod
    def workbook_save(workbook, excel_file: str):
        try:
            os.remove(excel_file)
        except Exception as error:
            DjangoClass.LoggingClass.error_local(
                error=error,
                function_error="ExcelClass.workbook_save"
            )
            pass
        try:
            workbook.save(excel_file)
        except Exception as error:
            DjangoClass.LoggingClass.error_local(
                error=error,
                function_error="ExcelClass.workbook_save"
            )
            print(f'\n ! Please, close the excel_file! \n: {excel_file} | {error}')

    @staticmethod
    def workbook_close(workbook):
        openpyxl.Workbook.close(workbook)

    @staticmethod
    def set_sheet_title(sheet, page_name='page 1'):
        sheet.title = page_name

    @staticmethod
    def get_sheet_value(col: Union[str, int], row: int, sheet):
        if isinstance(col, int):
            col = ExcelClass.get_column_letter(col)
        value = str(sheet[str(col).upper() + str(row)].value)
        if value == 'None' or value is None:
            value = ''
        else:
            value = str(value)
        return value

    @staticmethod
    def set_sheet_value(col: Union[str, int], row: int, value: str, sheet):
        if isinstance(col, int):
            col = ExcelClass.get_column_letter(col)
        if value == 'None' or value is None:
            value = ''
        sheet[str(col) + str(row)] = str(value)

    @staticmethod
    def get_column_letter(num: int):
        return get_column_letter(num)

    @staticmethod
    def get_max_num_rows(sheet):
        return int(sheet.max_row)

    class Example:
        @staticmethod
        def example_read_from_excel_file_col_int():
            excel_file = 'export.xlsx'
            workbook = ExcelClass.workbook_load(excel_file=excel_file)
            sheet = ExcelClass.workbook_activate(workbook=workbook)
            max_num_rows = ExcelClass.get_max_num_rows(sheet=sheet)
            max_num_cols = 10
            global_list = []
            for row in range(1, max_num_rows + 1):
                local_list = []
                for col in range(1, max_num_cols + 1):
                    value = ExcelClass.get_sheet_value(col=col, row=row, sheet=sheet)
                    local_list.append(value)
                global_list.append(local_list)
            for row in global_list:
                print(row)
            ExcelClass.workbook_close(workbook=workbook)

        @staticmethod
        def example_read_from_excel_file_col_char():
            excel_file = 'export.xlsx'
            workbook = ExcelClass.workbook_load(excel_file=excel_file)
            sheet = ExcelClass.workbook_activate(workbook=workbook)
            max_num_rows = ExcelClass.get_max_num_rows(sheet=sheet)
            char_cols = 'ACDF'
            global_list = []
            for row in range(1, max_num_rows + 1):
                local_list = []
                for col in char_cols:
                    value = ExcelClass.get_sheet_value(col=col, row=row, sheet=sheet)
                    local_list.append(value)
                global_list.append(local_list)
            for row in global_list:
                print(row)
            ExcelClass.workbook_close(workbook=workbook)

        @staticmethod
        def example_write_to_excel_file():
            global_list = [
                ['title_1', 'title_2', 'title_3'],
                ['body_1_1', 'body_1_2', 'body_1_3'],
                ['body_2_1', 'body_2_2', 'body_2_3'],
                ['body_3_1', 'body_3_3', 'body_3_3'],
            ]

            excel_file = 'import.xlsx'
            workbook = ExcelClass.workbook_create()
            sheet = ExcelClass.workbook_activate(workbook=workbook)
            for row in global_list:
                for value in row:
                    ExcelClass.set_sheet_value(
                        col=row.index(value) + 1,
                        row=global_list.index(row) + 1,
                        value=value,
                        sheet=sheet
                    )
            ExcelClass.workbook_save(workbook=workbook, excel_file=excel_file)
